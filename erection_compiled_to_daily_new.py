#!/usr/bin/env python3
"""
Parse 'Erection Compiled' sheets from one or many Excel files, compute productivity,
expand to daily rows, and write a consolidated workbook:

Sheets written:
  - ProdDailyExpandedSingles : per-day rows including single-occurrence gangs
  - RawData           : per-erection (unexpanded) rows (the 6 fields)
  - Data Issues       : row-level problems (requested columns + 'Issues')
  - Issues            : file-level problems (missing sheet/headers, no valid rows, etc.)
  - Diagnostics       : which sheet used, detected header row, and normalized column names
  - README_Assumptions: assumptions and cleaning notes

Usage examples (Windows CMD):
  python erection_compiled_to_daily.py ^
    --input "C:\\path\\to\\DPR_Files" ^
    --output "C:\\path\\ErectionCompiled_Output.xlsx"

  python erection_compiled_to_daily.py ^
    --files "C:\\path\\TA 408.xlsx" "C:\\path\\TA413-PBNTL-20.08.2025.xlsx" ^
           "C:\\path\\TA325 ANTL KEC DPR_18-08-2025.xlsx" "C:\\path\\DPR TA-416,14-09-25.xlsx" ^
    --output "C:\\path\\ErectionCompiled_Output.xlsx"
"""

import argparse
import logging
import re
import shutil
import tempfile
import warnings
import zipfile
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Tuple
import xml.etree.ElementTree as ET

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from dashboard.project_identity import (
    build_project_display,
    build_project_scope_key,
    parse_sheet_line_entries,
    normalize_line_name,
    parse_project_identity_from_filename,
)

logger = logging.getLogger("erection_compiled")


class ExcelCOMUnavailable(RuntimeError):
    """Raised when Excel COM automation cannot be used."""


# ---------- Config ----------
EXPECTED_HEADERS = [
    "location no",
    "type of tower",
    "starting date",
    "completion date",
    "gang name",
    "tower weight",
    "status",
]

HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "location no": ("location number", "loc no"),
    "tower weight": ("tower weight (mt)", "tower weight(mt)", "total tower weight", "weight mt"),
}

# Accepts: "Erection Compiled", "Erection-Compiled", "Erection Compiled v2", etc.
TARGET_SHEET_REGEX = re.compile(r"^\s*erection\s*.*\s*compiled\s*$", flags=re.I)

# Business rules (centralized here)
START_CUTOFF = pd.Timestamp("2021-01-01")
TODAY = pd.Timestamp.today().normalize()
TOWER_MIN_MT = 0.0
TOWER_MAX_MT = 500.0
DEFAULT_TOWER_WEIGHT_ASSUMED_MT = 45.0

# Column order for per-day expanded output
PER_DAY_COLUMNS = [
    "Work Date",
    "Start Date",
    "Complete Date",
    "Gang name",
    "Tower Weight",
    "Tower Type",
    "Productivity",
    "Project Code",
    "Line Name",
    "Project Name",
    "Project Display",
    "Project Scope Key",
    "Location No.",
    "Status",
]

# ---------- Helpers ----------
def nrm_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\n", " ").replace("\r", " ")
    s = s.replace("_", " ")
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.lower()


def canonical_header_key(s: str) -> str:
    """Return a compact key (alnum only) for header comparison."""
    normalized = nrm_header(s)
    return re.sub(r"\s+", "", normalized)


EXPECTED_HEADER_KEYS: Dict[str, str] = {}
for exp in EXPECTED_HEADERS:
    aliases = (exp, *HEADER_ALIASES.get(exp, ()))
    for alias in aliases:
        EXPECTED_HEADER_KEYS[canonical_header_key(alias)] = exp


def normalize_tower_type_label(value: object) -> str:
    text = "" if value is None else str(value).strip().upper()
    if not text or text in {"NAN", "NA", "NONE"}:
        return ""
    text = text.replace("\u00a0", " ")
    compact = re.sub(r"\s+", "", text)
    base_match = re.search(r"(DA|DB|DC|DD)", compact)
    if base_match:
        compact = compact[base_match.start():]
    match = re.match(r"^(DA|DB|DC|DD)(?:\+?(\d+))?$", compact)
    if match:
        base = match.group(1)
        ext = match.group(2)
        ext_value = str(int(ext)) if ext is not None else "0"
        return f"{base}+{ext_value}"
    return compact


def find_header_row(
    df_raw: pd.DataFrame,
    search_rows: int = 30,
    *,
    min_score: float = 3.0,
) -> Tuple[Optional[int], Optional[list]]:
    best = None
    best_score = -1
    nrows = min(search_rows, df_raw.shape[0])

    for r in range(nrows):
        row_vals = [nrm_header(x) for x in list(df_raw.iloc[r, :].values)]
        row_keys = [re.sub(r"\s+", "", val) for val in row_vals]

        score = 0
        mapping = {}
        used_expected = set()
        for i, (val, key) in enumerate(zip(row_vals, row_keys)):
            if not key:
                continue
            exp = EXPECTED_HEADER_KEYS.get(key)
            if exp and exp not in used_expected:
                mapping[i] = exp
                score += 1
                used_expected.add(exp)

        non_empty = sum(1 for v in row_vals if v)
        score += max(0, non_empty - 3) * 0.02

        if score > best_score:
            cols = [mapping.get(i, row_vals[i]) for i in range(len(row_vals))]
            best = (r, cols)
            best_score = score

    if best and best_score >= min_score:
        return best
    return None, None


def _drop_duplicate_columns_keep_first(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    """
    Drop duplicate column names while keeping the most informative column per name.

    Historically this kept the first duplicate only, but template column mappings can
    intentionally remap a later column to the same canonical header (for example
    ``tower weight``). In those cases the first match may be empty while a later one
    contains the actual values. We therefore keep the duplicate with the highest count
    of non-empty cells (ties resolved by first occurrence).
    """

    columns = list(df.columns)
    groups: dict[object, list[int]] = {}
    for idx, name in enumerate(columns):
        groups.setdefault(name, []).append(idx)

    duplicate_names: List[str] = []
    keep_indexes: set[int] = set()

    for name, idxs in groups.items():
        if len(idxs) == 1:
            keep_indexes.add(idxs[0])
            continue

        duplicate_names.append(str(name))
        best_idx = idxs[0]
        best_score = -1
        for idx in idxs:
            series = df.iloc[:, idx]
            non_empty = series.notna()
            if non_empty.any():
                as_text = series[non_empty].astype(str).str.strip().str.lower()
                non_empty = non_empty & ~as_text.isin({"", "nan", "none", "null"})
            score = int(non_empty.sum())
            if score > best_score:
                best_score = score
                best_idx = idx
        keep_indexes.add(best_idx)

    deduped = df.iloc[:, [idx for idx in range(len(columns)) if idx in keep_indexes]].copy()
    return deduped, duplicate_names


def find_target_sheet(sheet_names: List[str]) -> Optional[str]:
    for s in sheet_names:
        if s.strip().lower() == "erection compiled":
            return s
    for s in sheet_names:
        if TARGET_SHEET_REGEX.search(s):
            return s
    return None


def find_project_details_sheet(sheet_names: List[str]) -> Optional[str]:
    for s in sheet_names:
        if s.strip().lower() == "project details":
            return s
    return None


PROJECT_DETAILS_COLUMN_ALIASES: Dict[str, Tuple[str, ...]] = {
    "project_code": ("Project Code", "project_code", "code"),
    "project_name": ("Project Name", "project_name", "name"),
    "client_name": ("Client Name", "client", "client_name"),
    "noa_start": ("NOA Start Date", "noa start", "start date"),
    "loa_end": ("LOA End Date", "loa end", "end date"),
    "project_mgr": ("Project Manager", "project manger", "pm"),
    "regional_mgr": ("Regional Manager", "regional_manager"),
    "planning_eng": ("Planning Engineer", "planning_engineer"),
    "pch": ("PCH",),
    "section_inch": ("Section Incharge", "section_incharge"),
    "supervisor": ("Supervisor", "supervisor"),
}

PROJECT_DETAILS_HEADER_KEYS = {
    canonical_header_key(alias)
    for opts in PROJECT_DETAILS_COLUMN_ALIASES.values()
    for alias in opts
}


def find_project_details_header(
    df_raw: pd.DataFrame,
    *,
    search_rows: int = 10,
    search_cols: int = 10,
    min_matches: int = 3,
) -> Tuple[Optional[int], int]:
    """
    Search the top-left grid (rows/columns limited) for a row that looks like the header.
    Returns (row_index, left_trim_columns). If no candidate found, row_index is None.
    """
    if df_raw is None or df_raw.empty:
        return None, 0

    max_rows = min(search_rows, df_raw.shape[0])
    max_cols = min(search_cols, df_raw.shape[1])

    best_row = None
    best_score = -1
    best_first_idx = 0

    for r in range(max_rows):
        row_slice = list(df_raw.iloc[r, :max_cols].values)
        row_keys = [canonical_header_key(val) for val in row_slice]

        hits = [idx for idx, key in enumerate(row_keys) if key and key in PROJECT_DETAILS_HEADER_KEYS]
        score = len(hits)
        if score == 0:
            continue

        first_idx = hits[0]
        if score > best_score or (score == best_score and first_idx < best_first_idx):
            best_row = r
            best_score = score
            best_first_idx = first_idx

    if best_row is None or best_score < min_matches:
        return None, 0
    return best_row, best_first_idx


def to_number_mt(x):
    """Parse weight values like '5.5 MT', '7 t', '3,200' â†’ float (MT)."""
    if pd.isna(x):
        return np.nan
    s = str(x).strip().lower().replace(",", "")
    s = re.sub(r"(mt|tons?|t)\b", "", s)
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return float(s)
    except Exception:
        return np.nan


def _coerce_excel_serial(value: object) -> pd.Timestamp | None:
    """Return timestamp if *value* looks like an Excel serial date."""
    try:
        numeric = float(str(value).strip())
    except Exception:
        return None
    if not np.isfinite(numeric):
        return None
    if 20000 <= numeric <= 80000:
        return pd.to_datetime("1899-12-30") + pd.to_timedelta(numeric, unit="D")
    return None


def to_date(x):
    """Parse text dates (DD/MM/YYYY, etc.) and Excel serials."""
    excel_ts = _coerce_excel_serial(x)
    if excel_ts is not None:
        return excel_ts
    return pd.to_datetime(x, errors="coerce", dayfirst=True)


def to_date_monthfirst(x):
    """Parse text dates with month-first preference while preserving Excel serials."""
    excel_ts = _coerce_excel_serial(x)
    if excel_ts is not None:
        return excel_ts
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        return pd.to_datetime(x, errors="coerce", dayfirst=False)


_AMBIGUOUS_NUMERIC_DATE_RE = re.compile(r"^\s*(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})(?:\s+.*)?$")


def _is_ambiguous_numeric_date(value: object) -> bool:
    """True when *value* is textual and both day/month positions can be <= 12."""
    if pd.isna(value):
        return False
    if isinstance(value, (pd.Timestamp, np.datetime64)):
        return False
    text = str(value).strip()
    if not text:
        return False
    match = _AMBIGUOUS_NUMERIC_DATE_RE.match(text)
    if not match:
        return False
    first = int(match.group(1))
    second = int(match.group(2))
    return 1 <= first <= 12 and 1 <= second <= 12


def _repair_ambiguous_non_positive_dates(work: pd.DataFrame) -> int:
    """
    For rows parsed as Start>Date, retry ambiguous numeric dates in month-first
    mode and accept only repairs that produce a positive duration.
    """
    required = {"starting date", "completion date", "Start Date", "Complete Date"}
    if not required.issubset(work.columns):
        return 0
    if work.empty:
        return 0

    work["Start Date"] = pd.to_datetime(work["Start Date"], errors="coerce")
    work["Complete Date"] = pd.to_datetime(work["Complete Date"], errors="coerce")

    missing_dt_mask = work["Start Date"].isna() | work["Complete Date"].isna()
    days = (work["Complete Date"] - work["Start Date"]).dt.days + 1
    non_positive_mask = (~missing_dt_mask) & (days <= 0)
    if not non_positive_mask.any():
        return 0

    repaired = 0
    for idx in work.index[non_positive_mask]:
        raw_start = work.at[idx, "starting date"]
        raw_end = work.at[idx, "completion date"]
        if not (_is_ambiguous_numeric_date(raw_start) or _is_ambiguous_numeric_date(raw_end)):
            continue

        parsed_start = work.at[idx, "Start Date"]
        parsed_end = work.at[idx, "Complete Date"]
        start_monthfirst = to_date_monthfirst(raw_start)
        end_monthfirst = to_date_monthfirst(raw_end)

        for cand_start, cand_end in (
            (start_monthfirst, parsed_end),
            (parsed_start, end_monthfirst),
            (start_monthfirst, end_monthfirst),
        ):
            if pd.isna(cand_start) or pd.isna(cand_end):
                continue
            if (cand_end - cand_start).days + 1 <= 0:
                continue
            work.at[idx, "Start Date"] = cand_start
            work.at[idx, "Complete Date"] = cand_end
            repaired += 1
            break

    return repaired


def parse_project_from_filename(name: str) -> str:
    identity = parse_project_identity_from_filename(name)
    project_code = str(identity.get("project_code", "")).strip()
    if project_code:
        compact = re.sub(r"[^A-Z0-9]+", "", project_code.upper())
        if compact:
            return compact
    return Path(name).stem


def _normalize_space_only(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _normalize_project_code_key(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _resolve_dpr_config_path(input_folder: Optional[Path]) -> Optional[Path]:
    if input_folder is not None:
        candidate = input_folder.parent / "DPR_Config.xlsx"
        if candidate.exists():
            return candidate
    fallback = Path(__file__).resolve().parent / "Raw Data" / "DPR_Config.xlsx"
    if fallback.exists():
        return fallback
    return None


def load_erection_sheet_config(input_folder: Optional[Path]) -> Dict[str, List[Dict[str, str]]]:
    config_path = _resolve_dpr_config_path(input_folder)
    if config_path is None:
        return {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception as exc:
        logger.warning("Erection config: failed to read DPR config '%s': %s", config_path, exc)
        return {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            logger.warning("Erection config: 'Sheet Names Check' not found in '%s'", config_path)
            return {}

        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}

        normalized_headers = [_normalize_space_only(v) for v in header_row]
        try:
            project_idx = normalized_headers.index("project code")
            erection_idx = normalized_headers.index("erection sheet names")
        except ValueError:
            logger.warning(
                "Erection config: DPR config missing 'Project Code' or 'Erection Sheet Names' columns."
            )
            return {}

        line_idx = normalized_headers.index("erection line names") if "erection line names" in normalized_headers else None

        mapping: Dict[str, List[Dict[str, str]]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project = row[project_idx] if project_idx < len(row) else None
            if project in (None, ""):
                continue

            project_key = _normalize_project_code_key(project)
            raw_names = row[erection_idx] if erection_idx < len(row) else None
            if raw_names in (None, ""):
                mapping[project_key] = []
                continue
            raw_line_names = row[line_idx] if line_idx is not None and line_idx < len(row) else None

            sheet_count = len([chunk for chunk in re.split(r"[;,]", str(raw_names)) if str(chunk).strip()])
            line_chunks = [str(chunk).strip() for chunk in re.split(r"[;,]", str(raw_line_names))] if raw_line_names not in (None, "") else []
            if line_chunks and len(line_chunks) != sheet_count:
                logger.warning(
                    "Erection config: project '%s' has mismatched 'Erection Line Names'; falling back to sheet-name inference.",
                    str(project).strip(),
                )
                raw_line_names = ""

            entries = parse_sheet_line_entries(raw_names, raw_line_names, "erection")
            deduped_entries: List[Dict[str, str]] = []
            seen_sheet_keys = set()
            for entry in entries:
                key = _normalize_space_only(entry.get("sheet_name"))
                if not key or key in seen_sheet_keys:
                    continue
                seen_sheet_keys.add(key)
                deduped_entries.append(entry)
            mapping[project_key] = deduped_entries
        return mapping
    finally:
        wb.close()


def _resolve_named_template_sheet(wb, expected_name: str) -> Optional[str]:
    expected_key = _normalize_space_only(expected_name)
    for name in wb.sheetnames:
        if _normalize_space_only(name) == expected_key:
            return name
    return None


def _resolve_project_template_sheets(wb, project_name: object, discipline: str) -> List[str]:
    project_text = str(project_name or "").strip()
    if not project_text:
        return []

    resolved: List[str] = []
    seen: set[str] = set()

    for expected in (
        f"{project_text} {discipline}",
        f"{project_text} {discipline} Template Check",
    ):
        hit = _resolve_named_template_sheet(wb, expected)
        if hit and hit not in seen:
            seen.add(hit)
            resolved.append(hit)

    project_key = _normalize_space_only(project_text)
    discipline_key = _normalize_space_only(discipline)
    for name in wb.sheetnames:
        key = _normalize_space_only(name)
        if not key:
            continue
        if key.startswith(project_key) and key.endswith(discipline_key) and name not in seen:
            seen.add(name)
            resolved.append(name)

    return resolved


def _extract_template_column_map(ws) -> Dict[int, str]:
    to_map_row = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if _normalize_space_only(cell) == "to map":
                to_map_row = row_idx
                break
        if to_map_row is not None:
            break

    if to_map_row is None:
        return {}

    labels_row = to_map_row + 1
    col_map: Dict[int, str] = {}
    for col_idx, val in enumerate(next(ws.iter_rows(min_row=labels_row, max_row=labels_row, values_only=True), ()), start=0):
        normalized = nrm_header(val)
        if not normalized or normalized in {"nan", "none"}:
            continue
        col_map[col_idx] = normalized
    return col_map


def _extract_voltage_token(value: object) -> str:
    text = _normalize_space_only(value)
    if not text:
        return ""
    match = re.search(r"(\d{2,4})\s*kv\b", text)
    if match:
        return match.group(1)
    compact = re.sub(r"\s+", "", text)
    match = re.search(r"(\d{2,4})kv\b", compact)
    return match.group(1) if match else ""


def _select_template_mapping_for_request(
    template_entries: List[Dict[str, object]] | None,
    configured_sheet_name: object = "",
    line_name: object = "",
) -> Optional[Dict[str, object]]:
    if not template_entries:
        return None
    if len(template_entries) == 1:
        return template_entries[0]

    request_token = _extract_voltage_token(configured_sheet_name) or _extract_voltage_token(line_name)
    if request_token:
        token_matches = [
            entry for entry in template_entries if str(entry.get("voltage_token", "")).strip() == request_token
        ]
        if token_matches:
            token_matches.sort(key=lambda item: len(item.get("column_map", {}) or {}), reverse=True)
            return token_matches[0]

    ranked = sorted(
        template_entries,
        key=lambda item: len(item.get("column_map", {}) or {}),
        reverse=True,
    )
    return ranked[0] if ranked else None


def load_erection_template_mapping_config(
    input_folder: Optional[Path],
) -> Tuple[Dict[str, List[Dict[str, object]]], Dict[str, str]]:
    config_path = _resolve_dpr_config_path(input_folder)
    if config_path is None:
        return {}, {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception as exc:
        logger.warning("Erection template config: failed to read DPR config '%s': %s", config_path, exc)
        return {}, {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}, {}

        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}, {}

        normalized_headers = [_normalize_space_only(v) for v in header_row]
        project_idx = normalized_headers.index("project code") if "project code" in normalized_headers else None
        check_idx = None
        for candidate in ("erection template check", "erection"):
            if candidate in normalized_headers:
                check_idx = normalized_headers.index(candidate)
                break
        if project_idx is None or check_idx is None:
            return {}, {}

        mapping: Dict[str, List[Dict[str, object]]] = {}
        errors: Dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project = row[project_idx] if project_idx < len(row) else None
            if project in (None, ""):
                continue

            check_val = row[check_idx] if check_idx < len(row) else None
            if _normalize_space_only(check_val) != "yes":
                continue

            project_key = _normalize_project_code_key(project)
            template_sheets = _resolve_project_template_sheets(wb, project, "Erection")
            if not template_sheets:
                errors[project_key] = (
                    f"Erection Template Check is Yes but no mapping tab matching project '{str(project).strip()}' was found."
                )
                continue

            template_entries: List[Dict[str, object]] = []
            for sheet_name in template_sheets:
                col_map = _extract_template_column_map(wb[sheet_name])
                if not col_map:
                    continue
                template_entries.append(
                    {
                        "template_sheet": sheet_name,
                        "column_map": col_map,
                        "voltage_token": _extract_voltage_token(sheet_name),
                    }
                )

            if not template_entries:
                errors[project_key] = (
                    f"Erection template tab(s) for project '{str(project).strip()}' have no usable 'To Map' mapping row."
                )
                continue

            mapping[project_key] = template_entries
        return mapping, errors
    finally:
        wb.close()


def load_erection_section_config(input_folder: Optional[Path]) -> Dict[str, List[Dict[str, str]]]:
    config_path = _resolve_dpr_config_path(input_folder)
    if config_path is None:
        return {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception as exc:
        logger.warning("Erection section config: failed to read DPR config '%s': %s", config_path, exc)
        return {}

    required_headers = {
        "source sheet": "source_sheet",
        "section start text": "section_start_text",
        "section end text": "section_end_text",
        "line name": "line_name",
        "line column": "line_column",
        "line filter column": "line_filter_column",
        "line filter value": "line_filter_value",
        "template sheet": "template_sheet",
    }
    try:
        mapping: Dict[str, List[Dict[str, str]]] = {}
        suffix = " erection sections"
        for sheet_name in wb.sheetnames:
            sheet_key = _normalize_space_only(sheet_name)
            if not sheet_key.endswith(suffix):
                continue
            project_name = str(sheet_name[: -len(" Erection Sections")]).strip()
            project_key = _normalize_project_code_key(project_name)
            if not project_key:
                continue

            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue
            normalized_headers = [_normalize_space_only(v) for v in header_row]
            header_indexes: Dict[str, int] = {}
            for header, field in required_headers.items():
                if header in normalized_headers:
                    header_indexes[field] = normalized_headers.index(header)

            if "source_sheet" not in header_indexes:
                logger.warning("Erection section config: '%s' missing Source Sheet column", sheet_name)
                continue

            entries: List[Dict[str, str]] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                entry: Dict[str, str] = {}
                for field, idx in header_indexes.items():
                    raw = row[idx] if idx < len(row) else None
                    entry[field] = "" if raw in (None, "") else str(raw).strip()
                if not entry.get("source_sheet"):
                    continue
                entries.append(entry)
            if entries:
                mapping[project_key] = entries
        return mapping
    finally:
        wb.close()


def _select_template_mapping_by_sheet(
    template_entries: List[Dict[str, object]] | None,
    template_sheet_name: object,
) -> Optional[Dict[str, object]]:
    template_key = _normalize_space_only(template_sheet_name)
    if not template_entries or not template_key:
        return None
    for entry in template_entries:
        if _normalize_space_only(entry.get("template_sheet")) == template_key:
            return entry
    return None


def _row_contains_marker(row: pd.Series, marker: str) -> bool:
    marker_key = _normalize_space_only(marker)
    if not marker_key:
        return False
    for value in row.tolist():
        if marker_key in _normalize_space_only(value):
            return True
    return False


def _slice_section_frame(
    df_raw: pd.DataFrame,
    start_text: object = "",
    end_text: object = "",
) -> Tuple[pd.DataFrame, Optional[int], Optional[int]]:
    start_marker = str(start_text or "").strip()
    end_marker = str(end_text or "").strip()
    start_idx = 0
    end_idx = len(df_raw.index)

    if start_marker:
        for idx, (_, row) in enumerate(df_raw.iterrows()):
            if _row_contains_marker(row, start_marker):
                start_idx = idx
                break

    if end_marker:
        for idx in range(start_idx + 1, len(df_raw.index)):
            if _row_contains_marker(df_raw.iloc[idx], end_marker):
                end_idx = idx
                break

    section = df_raw.iloc[start_idx:end_idx].copy().reset_index(drop=True)
    return section, start_idx, (end_idx if end_idx < len(df_raw.index) else None)


def _find_df_column_by_normalized(df: pd.DataFrame, candidates: Iterable[object]) -> Optional[object]:
    candidate_keys = [_normalize_space_only(candidate) for candidate in candidates if _normalize_space_only(candidate)]
    if not candidate_keys:
        return None
    by_key: Dict[str, object] = {}
    for column in df.columns:
        key = _normalize_space_only(column)
        if key and key not in by_key:
            by_key[key] = column
    for key in candidate_keys:
        if key in by_key:
            return by_key[key]
    return None


def _line_weight_key(value: object) -> str:
    return _normalize_space_only(value)


def build_exact_sheet_selector(sheet_name: str):
    expected_key = _normalize_space_only(sheet_name)

    def _selector(sheet_names: List[str]) -> Optional[str]:
        by_key: Dict[str, str] = {}
        for existing in sheet_names:
            key = _normalize_space_only(existing)
            if key and key not in by_key:
                by_key[key] = existing
        return by_key.get(expected_key)

    return _selector


def normalize_gang_name(name: str) -> str:
    """
    - Strip special chars (keep letters, digits, spaces)
    - Each word Title Case
    - If ends with digits stuck to a letter (e.g., 'Xyz4'), insert a space â†’ 'Xyz 4'
    Examples:
      'sobha devi' -> 'Sobha Devi'
      'sobha-devi' -> 'Sobha Devi'
      'xyz4' -> 'Xyz 4'
      'xyz-4' -> 'Xyz 4'
    """
    if name is None:
        return "undefined"
    s = str(name).strip().lower()
    # keep letters, digits, spaces â†’ replace other runs with a space
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return "undefined"
    # Title case
    s = s.title()
    # Insert a space before trailing digits if jammed to a letter
    s = re.sub(r"([A-Za-z])(\d+)$", r"\1 \2", s)
    cleaned = s.strip()
    return cleaned if cleaned else "undefined"

# --- NEW: tolerant loader for a single source file's "Project Details" ---
def load_project_details_from_source(dfp: pd.DataFrame, source_file: Path) -> pd.DataFrame:
    # only proceed if the sheet produced data
    if dfp is None or dfp.empty:
        return pd.DataFrame()

    header_row, trim_left = find_project_details_header(dfp, search_rows=10, search_cols=10)
    if header_row is None:
        return pd.DataFrame()

    trim_left = max(0, trim_left or 0)
    header_values = list(dfp.iloc[header_row, trim_left:])
    dfp = dfp.iloc[header_row + 1 :, trim_left:].copy()
    if dfp.empty or not header_values:
        return pd.DataFrame()
    dfp.columns = header_values

    # tolerant column picks (allow minor header variations)
    def pick(df, *opts):
        cols = {str(c).strip().lower(): c for c in df.columns}
        for o in opts:
            k = o.strip().lower()
            if k in cols:
                return cols[k]
        # contains fallback
        for k, c in cols.items():
            if any(o.lower() in k for o in opts):
                return c
        raise KeyError(f"Missing one of {opts} in {list(df.columns)}")

    def pick_field(df, field_key: str):
        return pick(df, *PROJECT_DETAILS_COLUMN_ALIASES[field_key])

    try:
        c_code = pick_field(dfp, "project_code")
        c_name = pick_field(dfp, "project_name")
        c_client = pick_field(dfp, "client_name")
        c_noa = pick_field(dfp, "noa_start")
        c_loa = pick_field(dfp, "loa_end")
        c_pm = pick_field(dfp, "project_mgr")
        c_rm = pick_field(dfp, "regional_mgr")
        c_pe = pick_field(dfp, "planning_eng")
        c_pch = pick_field(dfp, "pch")
        c_si = pick_field(dfp, "section_inch")
        c_sup = pick_field(dfp, "supervisor")
    except KeyError:
        # If the sheet is weirdly formatted, skip gracefully
        return pd.DataFrame()

        # Build a narrow frame with the columns we care about
    meta = pd.DataFrame({
        "project_code": dfp[c_code],
        "project_name": dfp[c_name],
        "client_name":  dfp[c_client],
        "noa_start":    pd.to_datetime(dfp[c_noa], errors="coerce"),
        "loa_end":      pd.to_datetime(dfp[c_loa], errors="coerce"),
        "project_mgr":  dfp[c_pm],
        "regional_mgr": dfp[c_rm],
        "planning_eng": dfp[c_pe],
        "pch":          dfp[c_pch],
        "section_inch": dfp[c_si],
        "supervisor":   dfp[c_sup],
    }).astype(object)

    # Forward-fill project metadata so blank rows (extra names) inherit the project
    meta[["project_code","project_name","client_name","noa_start","loa_end",
          "project_mgr","regional_mgr","planning_eng","pch"]] = \
        meta[["project_code","project_name","client_name","noa_start","loa_end",
              "project_mgr","regional_mgr","planning_eng","pch"]].ffill()

    # Helper: join unique non-empty values, preserving order
    def uniq_join(series):
        vals = [str(x).strip() for x in series if pd.notna(x) and str(x).strip() != ""]
        seen = set(); out = []
        for v in vals:
            if v not in seen:
                seen.add(v); out.append(v)
        return ", ".join(out)

    # Aggregate to one row per project_code (collect multiples for the two fields)
    out = (meta
           .groupby("project_code", dropna=False)
           .agg({
               "project_name": "first",
               "client_name":  "first",
               "noa_start":    "first",
               "loa_end":      "first",
               "project_mgr":  "first",
               "regional_mgr": "first",
               "planning_eng": "first",
               "pch":          "first",
               "section_inch": uniq_join,
               "supervisor":   uniq_join,
           })
           .reset_index()
    )

    # Uppercase/trim code; drop empties
    out["project_code"] = out["project_code"].astype(str).str.strip().str.upper()
    out = out[out["project_code"].ne("")]

    # File name for traceability + pass-through literal "Project Name" label if present
    out["_source_file"] = source_file.name
    if "Project Name" in dfp.columns:
        # keep the human label from sheet; itâ€™s already ffilled via meta
        out["Project Name"] = meta.groupby("project_code")["project_name"].first().astype(str).str.strip()

    return out


SheetSelector = Callable[[List[str]], Optional[str]]


def export_sheet_via_excel_to_df(
    source: Path,
    selector: SheetSelector,
    read_csv_kwargs: Optional[dict] = None,
) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    """Open Excel via COM, export the selected sheet to CSV, and load it into pandas."""
    read_csv_kwargs = read_csv_kwargs or {}
    try:
        import win32com.client  # type: ignore
    except ImportError as exc:  # pragma: no cover - depends on environment
        raise ExcelCOMUnavailable("Excel COM automation (win32com) is not available") from exc

    temp_dir = Path(tempfile.mkdtemp(prefix="excel_csv_"))
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    wb = None
    csv_path: Optional[Path] = None
    try:
        wb = excel.Workbooks.Open(str(source.resolve()), UpdateLinks=False, ReadOnly=True)
        sheet_names = [ws.Name for ws in wb.Worksheets]
        target = selector(sheet_names)
        if not target:
            return None, None

        target_ws = next((ws for ws in wb.Worksheets if ws.Name == target), None)
        if target_ws is None:
            raise RuntimeError(f"Sheet '{target}' could not be accessed via Excel COM")

        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", target)
        csv_path = temp_dir / f"{safe_name}.csv"

        target_ws.Copy()  # make it the active workbook
        wb_export = excel.ActiveWorkbook
        try:
            wb_export.SaveAs(str(csv_path), FileFormat=6)  # xlCSV
        finally:
            wb_export.Close(SaveChanges=False)

        df = pd.read_csv(csv_path, **read_csv_kwargs)
        logger.info("Excel COM CSV export succeeded for sheet '%s' in '%s'", target, source.name)
        return df, target
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()
        if csv_path and csv_path.exists():
            csv_path.unlink(missing_ok=True)
        shutil.rmtree(temp_dir, ignore_errors=True)


def scrub_defined_names_from_workbook(source: Path) -> Path:
    """Create a temp copy of the workbook with <definedNames> removed."""
    temp_dir = Path(tempfile.mkdtemp(prefix="excel_scrub_"))
    scrubbed_path = temp_dir / f"{source.stem}_scrubbed.xlsx"
    removed_any = False

    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(scrubbed_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/workbook.xml":
                try:
                    root = ET.fromstring(data)
                    ns_uri = ""
                    if root.tag.startswith("{"):
                        ns_uri = root.tag.split("}")[0][1:]
                    nsmap = {"ns": ns_uri} if ns_uri else {}
                    defined_nodes = root.findall("ns:definedNames", nsmap) if nsmap else root.findall("definedNames")
                    if defined_nodes:
                        for node in defined_nodes:
                            root.remove(node)
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                        removed_any = True
                except ET.ParseError:
                    text = data.decode("utf-8", errors="ignore")
                    if "<definedNames" in text and "</definedNames>" in text:
                        start = text.find("<definedNames")
                        end = text.find("</definedNames>")
                        if start != -1 and end != -1:
                            end = text.find(">", end)
                            if end != -1:
                                text = text[:start] + text[end + 1 :]
                                removed_any = True
                    data = text.encode("utf-8")
            zout.writestr(item, data)

    if removed_any:
        logger.info("Scrubbed defined names for '%s'", source.name)
    else:
        logger.info("Workbook '%s' had no defined names to scrub", source.name)
    return scrubbed_path


def load_sheet_from_scrubbed_copy(
    source: Path,
    selector: SheetSelector,
    read_excel_kwargs: Optional[dict] = None,
) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    """Scrub defined names then reload the desired sheet via openpyxl."""
    read_excel_kwargs = read_excel_kwargs or {}
    scrubbed = scrub_defined_names_from_workbook(source)
    try:
        with pd.ExcelFile(scrubbed, engine="openpyxl") as xl:
            sheet_name = selector(list(xl.sheet_names))
            if not sheet_name:
                return None, None
            df = pd.read_excel(xl, sheet_name=sheet_name, **read_excel_kwargs)
            logger.info("XML scrub load succeeded for sheet '%s' in '%s'", sheet_name, source.name)
            return df, sheet_name
    finally:
        shutil.rmtree(scrubbed.parent, ignore_errors=True)


def load_sheet_with_csv_fallback(
    source: Path,
    selector: SheetSelector,
    *,
    read_excel_kwargs: Optional[dict] = None,
    read_csv_kwargs: Optional[dict] = None,
) -> Tuple[Optional[pd.DataFrame], Optional[str], Optional[str]]:
    """
    Try reading a sheet with openpyxl first; on failure, export to CSV via Excel COM.

    Returns (df, sheet_name, fallback_note). df/sheet_name are None if the selector
    doesn't find a matching sheet.
    """
    read_excel_kwargs = read_excel_kwargs or {}
    read_csv_kwargs = read_csv_kwargs or {}

    try:
        with pd.ExcelFile(source, engine="openpyxl") as xl:
            sheet_name = selector(list(xl.sheet_names))
            if not sheet_name:
                return None, None, None
            df = pd.read_excel(xl, sheet_name=sheet_name, **read_excel_kwargs)
            logger.debug("Loaded sheet '%s' from '%s' via openpyxl", sheet_name, source.name)
            return df, sheet_name, None
    except Exception as primary_error:
        logger.warning("openpyxl failed to load sheet from '%s': %s", source.name, primary_error)
        fallback_messages = []

        try:
            df_scrub, sheet_name = load_sheet_from_scrubbed_copy(
                source, selector, read_excel_kwargs=read_excel_kwargs
            )
        except Exception as scrub_error:
            logger.warning("XML scrub fallback failed for '%s': %s", source.name, scrub_error)
            fallback_messages.append(f"XML scrub fallback failed ({scrub_error})")
        else:
            if sheet_name is None:
                return None, None, None
            note = f"XML scrub fallback used for sheet '{sheet_name}'"
            logger.warning("%s in '%s'", note, source.name)
            return df_scrub, sheet_name, note

        try:
            df_csv, sheet_name = export_sheet_via_excel_to_df(
                source, selector, read_csv_kwargs=read_csv_kwargs
            )
        except ExcelCOMUnavailable as com_exc:
            logger.warning(
                "Excel COM CSV fallback unavailable for '%s': %s", source.name, com_exc
            )
            fallback_messages.append(f"Excel COM CSV fallback unavailable ({com_exc})")
        except Exception as fallback_error:
            logger.warning(
                "Excel COM CSV fallback failed for '%s': %s", source.name, fallback_error
            )
            fallback_messages.append(f"Excel COM CSV fallback failed ({fallback_error})")
        else:
            if sheet_name is None:
                return None, None, None
            note = f"Excel COM CSV fallback used for sheet '{sheet_name}' after XML scrub failure"
            logger.warning("%s in '%s'", note, source.name)
            return df_csv, sheet_name, note

        raise RuntimeError(
            f"openpyxl load failed ({primary_error}); " + "; ".join(fallback_messages)
        ) from primary_error


# ---------- Core (per file) ----------
def process_file(
    path: Path,
    configured_sheet_names: Optional[List[Dict[str, str]]] = None,
    template_column_map: Optional[Dict[int, str]] = None,
    template_sheet_name: Optional[str] = None,
    template_mappings: Optional[List[Dict[str, object]]] = None,
    section_config: Optional[List[Dict[str, str]]] = None,
):
    """
    Process a single workbook; return:
      per_day_with_singles : per-day rows including single-occurrence gangs
      per_erection         : per-erection (unexpanded) rows
      diagnostics          : list[dict] (one row per processed sheet)
      issues               : list[dict] (file-level)
      data_issues_df       : DataFrame with requested columns + 'Issues' (row-level)
    """
    issues: List[dict] = []
    per_day_frames: List[pd.DataFrame] = []
    per_erection_frames: List[pd.DataFrame] = []
    data_issue_frames: List[pd.DataFrame] = []
    diagnostics: List[dict] = []
    project_identity = parse_project_identity_from_filename(path.name)
    project_code_display = str(project_identity.get("project_code", "")).strip()
    file_line_name = normalize_line_name(project_identity.get("line_name", ""))
    base_project_name = build_project_display(
        project_code_display,
        file_line_name,
        parse_project_from_filename(path.name),
    )

    sheet_requests: List[Dict[str, object]] = []
    if section_config:
        for section_entry in section_config:
            configured_name = str(section_entry.get("source_sheet", "")).strip()
            if not configured_name:
                continue
            fixed_line_name = normalize_line_name(section_entry.get("line_name", ""))
            sheet_requests.append(
                {
                    "requested_name": configured_name,
                    "selector": build_exact_sheet_selector(configured_name),
                    "line_name": fixed_line_name,
                    "line_name_source": "config" if fixed_line_name else "column",
                    "section_start_text": str(section_entry.get("section_start_text", "")).strip(),
                    "section_end_text": str(section_entry.get("section_end_text", "")).strip(),
                    "line_column": str(section_entry.get("line_column", "")).strip(),
                    "line_filter_column": str(section_entry.get("line_filter_column", "")).strip(),
                    "line_filter_value": str(section_entry.get("line_filter_value", "")).strip(),
                    "template_sheet": str(section_entry.get("template_sheet", "")).strip(),
                    "drop_blank_location_rows": True,
                }
            )
    elif configured_sheet_names:
        seen_keys = set()
        for configured_entry in configured_sheet_names:
            configured_name = str(configured_entry.get("sheet_name", "")).strip()
            key = _normalize_space_only(configured_name)
            if not key or key in seen_keys:
                continue
            seen_keys.add(key)
            sheet_requests.append(
                {
                    "requested_name": configured_name,
                    "selector": build_exact_sheet_selector(configured_name),
                    "line_name": normalize_line_name(configured_entry.get("line_name", "")),
                    "line_name_source": str(configured_entry.get("line_name_source", "")).strip(),
                }
            )
    if not sheet_requests:
        sheet_requests = [
            {
                "requested_name": None,
                "selector": find_target_sheet,
                "line_name": "",
                "line_name_source": "",
            }
        ]

    line_weight_defaults: Dict[str, float] = {}

    def expand_per_day(source: pd.DataFrame) -> pd.DataFrame:
        if source.empty:
            return pd.DataFrame(columns=PER_DAY_COLUMNS)

        rows = []
        for _, r in source.iterrows():
            for d in pd.date_range(r["Start Date"], r["Complete Date"], freq="D"):
                rows.append(
                    {
                        "Work Date": d.normalize(),
                        "Start Date": r["Start Date"].normalize(),
                        "Complete Date": r["Complete Date"].normalize(),
                        "Gang name": r["Gang name"],
                        "Tower Weight": r["Tower Weight"],
                        "Tower Type": r["Tower Type"],
                        "Productivity": r["Productivity"],
                        "Project Code": r["Project Code"],
                        "Line Name": r["Line Name"],
                        "Project Name": r["Project Name"],
                        "Project Display": r["Project Display"],
                        "Project Scope Key": r["Project Scope Key"],
                        "Location No.": r["Location No."],
                        "Status": r["Status"],
                    }
                )

        result = pd.DataFrame(rows)
        if result.empty:
            return result.reindex(columns=PER_DAY_COLUMNS)

        result = result.sort_values(
            ["Project Name", "Work Date", "Gang name", "Start Date"],
            ignore_index=True,
        )
        return result.reindex(columns=PER_DAY_COLUMNS)

    for sheet_request in sheet_requests:
        requested_name = sheet_request.get("requested_name")
        selector = sheet_request["selector"]
        sheet_line_name = normalize_line_name(sheet_request.get("line_name", "")) or file_line_name
        line_name_source = str(sheet_request.get("line_name_source", "")).strip()
        selected_template_column_map = template_column_map
        selected_template_sheet_name = template_sheet_name
        requested_template_sheet = str(sheet_request.get("template_sheet", "")).strip()
        if template_mappings:
            selected_template = (
                _select_template_mapping_by_sheet(template_mappings, requested_template_sheet)
                if requested_template_sheet
                else None
            )
            if selected_template is None:
                selected_template = _select_template_mapping_for_request(
                    template_mappings,
                    configured_sheet_name=requested_template_sheet or requested_name,
                    line_name=sheet_line_name,
                )
            if selected_template:
                selected_template_column_map = selected_template.get("column_map")
                selected_template_sheet_name = str(selected_template.get("template_sheet", "")).strip() or None
        project_name = build_project_display(project_code_display, sheet_line_name, base_project_name)
        project_scope_key = build_project_scope_key(project_code_display, sheet_line_name, project_name)
        try:
            df_raw, target, fallback_note = load_sheet_with_csv_fallback(
                path,
                selector,
                read_excel_kwargs={"header": None},
                read_csv_kwargs={"header": None},
            )
        except Exception as e:
            issues.append(
                {
                    "file": path.name,
                    "sheet": requested_name or "",
                    "issue": f"'Erection' load error: {e}",
                }
            )
            continue

        if df_raw is None or target is None:
            if requested_name:
                issues.append(
                    {
                        "file": path.name,
                        "sheet": requested_name,
                        "issue": f"Configured erection sheet not found: '{requested_name}'",
                    }
                )
            else:
                issues.append({"file": path.name, "issue": "Sheet 'Erection Compiled' not found"})
            continue

        if fallback_note:
            logger.warning("Fallback note for '%s': %s", path.name, fallback_note)
            issues.append({"file": path.name, "sheet": target, "issue": fallback_note})

        section_start_text = str(sheet_request.get("section_start_text", "")).strip()
        section_end_text = str(sheet_request.get("section_end_text", "")).strip()
        section_start_row = None
        section_end_row = None
        if section_start_text or section_end_text:
            df_raw, section_start_row, section_end_row = _slice_section_frame(
                df_raw,
                start_text=section_start_text,
                end_text=section_end_text,
            )
            if df_raw.empty:
                issues.append(
                    {
                        "file": path.name,
                        "sheet": target,
                        "issue": (
                            "Configured section is empty"
                            f" (start='{section_start_text}', end='{section_end_text}')"
                        ),
                    }
                )
                continue

        hdr_row, cols = find_header_row(
            df_raw,
            search_rows=30,
            min_score=1.0 if selected_template_column_map else 3.0,
        )
        if hdr_row is None:
            issues.append(
                {
                    "file": path.name,
                    "sheet": target,
                    "issue": "Could not detect header row automatically",
                }
            )
            continue

        df = df_raw.iloc[hdr_row + 1 :].copy()
        df.columns = cols

        applied_mapping: List[str] = []
        if selected_template_column_map:
            remapped_columns = list(df.columns)
            for idx, mapped in sorted(selected_template_column_map.items()):
                if idx >= len(remapped_columns):
                    continue
                current = nrm_header(remapped_columns[idx])
                mapped_target = nrm_header(mapped)
                if not mapped_target:
                    continue
                remapped_columns[idx] = mapped_target
                if current != mapped_target:
                    applied_mapping.append(f"C{idx + 1}:{current}->{mapped_target}")
            df.columns = remapped_columns
        df, duplicate_names = _drop_duplicate_columns_keep_first(df)
        if duplicate_names:
            issues.append(
                {
                    "file": path.name,
                    "sheet": target,
                    "issue": (
                        "Duplicate columns detected after header mapping; kept first occurrence for: "
                        + ", ".join(sorted(set(duplicate_names)))
                    ),
                }
            )

        diag = {
            "file": path.name,
            "project": project_name,
            "sheet": target,
            "line_name": sheet_line_name,
            "line_name_source": line_name_source,
            "detected_header_row": hdr_row,
            "columns_detected": ", ".join(cols[:20]),
        }
        if section_start_text or section_end_text:
            diag["section_start_text"] = section_start_text
            diag["section_end_text"] = section_end_text
            diag["section_start_row"] = "" if section_start_row is None else int(section_start_row) + 1
            diag["section_end_row"] = "" if section_end_row is None else int(section_end_row)
        if requested_name:
            diag["configured_sheet"] = requested_name
        if selected_template_column_map:
            diag["template_mapping_sheet"] = selected_template_sheet_name or ""
            diag["template_mapping_applied"] = bool(applied_mapping)
            diag["template_mapping_changes"] = "; ".join(applied_mapping)

        line_filter_column = str(sheet_request.get("line_filter_column", "")).strip()
        line_filter_value = str(sheet_request.get("line_filter_value", "")).strip()
        if line_filter_column and line_filter_value:
            filter_col = _find_df_column_by_normalized(df, [line_filter_column])
            if filter_col is None:
                issues.append(
                    {
                        "file": path.name,
                        "sheet": target,
                        "issue": f"Configured line filter column not found: '{line_filter_column}'",
                        "columns": list(df.columns),
                    }
                )
                diagnostics.append(diag)
                continue
            filter_key = _normalize_space_only(line_filter_value)
            filter_series = df[filter_col].astype(object).map(_normalize_space_only)
            df = df.loc[filter_series.eq(filter_key)].copy()
            diag["line_filter_column"] = str(filter_col)
            diag["line_filter_value"] = line_filter_value
            diag["line_filter_rows"] = int(len(df.index))

        if sheet_request.get("drop_blank_location_rows"):
            location_col = _find_df_column_by_normalized(df, ["location no"])
            if location_col is not None:
                location_text = df[location_col].astype(object).map(lambda x: str(x).strip() if pd.notna(x) else "")
                before_rows = len(df.index)
                df = df.loc[
                    location_text.ne("")
                    & ~location_text.str.casefold().isin({"nan", "none", "nat"})
                ].copy()
                diag["blank_location_rows_dropped"] = int(before_rows - len(df.index))

        # Only the fields we need for computation. Tower Weight may be absent.
        needed = ["starting date", "completion date", "location no"]
        missing = [c for c in needed if c not in df.columns]
        if missing:
            issues.append(
                {
                    "file": path.name,
                    "sheet": target,
                    "issue": f"Missing required columns after header-detect: {missing}",
                    "columns": list(df.columns),
                }
            )
            diagnostics.append(diag)
            continue

        if "gang name" not in df.columns:
            issues.append(
                {
                    "file": path.name,
                    "sheet": target,
                    "issue": "'gang name' header missing; using 'undefined' for affected rows",
                }
            )

        line_column = str(sheet_request.get("line_column", "")).strip()
        row_line_source = pd.Series(sheet_line_name or file_line_name, index=df.index, dtype="object")
        if line_column:
            line_col = _find_df_column_by_normalized(df, [line_column, "line name"])
            if line_col is not None:
                line_values = df[line_col].astype(object).map(normalize_line_name)
                fallback_line = sheet_line_name or file_line_name
                row_line_source = line_values.where(line_values.astype(str).str.strip().ne(""), fallback_line)
                diag["line_column"] = str(line_col)
            else:
                issues.append(
                    {
                        "file": path.name,
                        "sheet": target,
                        "issue": f"Configured line column not found: '{line_column}'",
                        "columns": list(df.columns),
                    }
                )

        work = df[needed].copy()
        if "tower weight" in df.columns:
            work["tower weight"] = df["tower weight"]
            diag["tower_weight_assumed_mt"] = ""
            diag["tower_weight_assumption_rows"] = 0
        elif sheet_request.get("drop_blank_location_rows"):
            work["tower weight"] = pd.NA
            diag["tower_weight_assumed_mt"] = ""
            diag["tower_weight_assumption_rows"] = int(len(work.index))
            issues.append(
                {
                    "file": path.name,
                    "sheet": target,
                    "issue": (
                        "'tower weight' header missing in configured section; "
                        "will use same-file line averages where available"
                    ),
                }
            )
        else:
            work["tower weight"] = DEFAULT_TOWER_WEIGHT_ASSUMED_MT
            diag["tower_weight_assumed_mt"] = float(DEFAULT_TOWER_WEIGHT_ASSUMED_MT)
            diag["tower_weight_assumption_rows"] = int(len(work.index))
            issues.append(
                {
                    "file": path.name,
                    "sheet": target,
                    "issue": (
                        f"'tower weight' header missing; assumed "
                        f"{DEFAULT_TOWER_WEIGHT_ASSUMED_MT:.0f} MT for all rows in this sheet"
                    ),
                }
            )

        tower_type_col = None
        for candidate in ("type of tower", "tower type", "type"):
            if candidate in df.columns:
                tower_type_col = candidate
                break
        if tower_type_col:
            work["Tower Type"] = df[tower_type_col].apply(normalize_tower_type_label)
        else:
            work["Tower Type"] = ""

        tightening_col = None
        for candidate in ("tower tightening", "tower tightening date", "tower_tightening", "tightening date"):
            if candidate in df.columns:
                tightening_col = candidate
                break
        if tightening_col:
            work["tower tightening"] = df[tightening_col]
        else:
            work["tower tightening"] = pd.NA

        # Parse + clean (do not drop yet; capture issues first)
        work["Start Date"] = work["starting date"].apply(to_date)
        work["Complete Date"] = work["completion date"].apply(to_date)
        work["Tower Tightening"] = work["tower tightening"].apply(to_date)
        repaired_dates = _repair_ambiguous_non_positive_dates(work)
        work["Start Date"] = pd.to_datetime(work["Start Date"], errors="coerce")
        work["Complete Date"] = pd.to_datetime(work["Complete Date"], errors="coerce")
        work["Tower Tightening"] = pd.to_datetime(work["Tower Tightening"], errors="coerce").dt.normalize()
        diag["date_repairs_applied"] = int(repaired_dates)
        if repaired_dates:
            issues.append(
                {
                    "file": path.name,
                    "sheet": target,
                    "issue": f"Auto-corrected {repaired_dates} ambiguous Start/Complete date row(s)",
                }
            )
        gang_source = df["gang name"] if "gang name" in df.columns else pd.Series("undefined", index=df.index)
        work["Gang name"] = gang_source.apply(normalize_gang_name)
        work.loc[work["Gang name"].astype(str).str.strip().eq(""), "Gang name"] = "undefined"
        work["Tower Weight"] = work["tower weight"].apply(to_number_mt)
        work["Line Name"] = row_line_source.reindex(work.index).fillna("").astype(object)
        assumed_weight_messages: list[str] = []
        existing_weight_mask = work["Tower Weight"].notna()
        valid_existing_weight_mask = existing_weight_mask & work["Tower Weight"].between(
            TOWER_MIN_MT, TOWER_MAX_MT, inclusive="both"
        )
        missing_weight_mask = ~existing_weight_mask

        if missing_weight_mask.any() and line_weight_defaults:
            filled_by_line: Dict[str, int] = {}
            for idx in work.index[missing_weight_mask]:
                line_key = _line_weight_key(work.at[idx, "Line Name"])
                default_weight = line_weight_defaults.get(line_key)
                if default_weight is None:
                    continue
                work.at[idx, "Tower Weight"] = default_weight
                line_label = str(work.at[idx, "Line Name"]).strip() or "blank line"
                filled_by_line[line_label] = filled_by_line.get(line_label, 0) + 1
            if filled_by_line:
                assumed_weight_messages.append(
                    "Assumed Tower Weight from same-file historical line averages: "
                    + "; ".join(
                        f"{line}={line_weight_defaults.get(_line_weight_key(line), 0):.3f} MT for {count} row(s)"
                        for line, count in sorted(filled_by_line.items())
                    )
                )
                diag["tower_weight_line_average_rows"] = int(sum(filled_by_line.values()))
                missing_weight_mask = work["Tower Weight"].isna()

        # If a sheet has mixed tower weight quality, impute blanks with the average of
        # valid entered values. If no usable values exist, keep the historical 45 MT fallback.
        if missing_weight_mask.any():
            if valid_existing_weight_mask.any():
                assumed_weight = float(work.loc[valid_existing_weight_mask, "Tower Weight"].mean())
                work.loc[missing_weight_mask, "Tower Weight"] = assumed_weight
                diag["tower_weight_assumed_mt"] = float(assumed_weight)
                diag["tower_weight_assumption_rows"] = int(missing_weight_mask.sum())
                assumed_weight_messages.append(
                    f"Assumed Tower Weight {assumed_weight:.3f} MT for {int(missing_weight_mask.sum())} row(s) "
                    "with blank/missing tower weight using average of available towers"
                )
            else:
                work.loc[missing_weight_mask, "Tower Weight"] = DEFAULT_TOWER_WEIGHT_ASSUMED_MT
                diag["tower_weight_assumed_mt"] = float(DEFAULT_TOWER_WEIGHT_ASSUMED_MT)
                diag["tower_weight_assumption_rows"] = int(missing_weight_mask.sum())
                assumed_weight_messages.append(
                    f"No usable tower weights found; assumed {DEFAULT_TOWER_WEIGHT_ASSUMED_MT:.0f} MT "
                    f"for {int(missing_weight_mask.sum())} row(s)"
                )

        for message in assumed_weight_messages:
            issues.append(
                {
                    "file": path.name,
                    "sheet": target,
                    "issue": message,
                }
            )
        if valid_existing_weight_mask.any():
            weighted_lines = work.loc[valid_existing_weight_mask, ["Line Name", "Tower Weight"]].copy()
            for line_name, line_frame in weighted_lines.groupby("Line Name", dropna=False):
                line_key = _line_weight_key(line_name)
                if not line_key:
                    continue
                line_weight_defaults[line_key] = float(line_frame["Tower Weight"].mean())

        work["Project Code"] = project_code_display or project_name
        work["Project Name"] = work["Line Name"].map(
            lambda line: build_project_display(project_code_display, line, base_project_name)
        )
        work["Project Display"] = work["Project Name"]
        work["Project Scope Key"] = work["Line Name"].map(
            lambda line: build_project_scope_key(project_code_display, line, project_name)
        )
        work["Source File"] = path.name
        work["Source Sheet"] = target
        work["Location No."] = work["location no"].astype(object).map(
            lambda x: str(x).strip() if pd.notna(x) else pd.NA
        )
        work["Tower Tightening Raw"] = work["tower tightening"].astype(object).map(
            lambda x: str(x).strip() if pd.notna(x) else pd.NA
        )

        status_series = df["status"] if "status" in df.columns else pd.Series(pd.NA, index=df.index)
        work["Status"] = status_series.astype(object).map(lambda x: str(x).strip() if pd.notna(x) else pd.NA)

        # Precompute validity flags
        missing_dt_mask = work["Start Date"].isna() | work["Complete Date"].isna()
        days = (work["Complete Date"] - work["Start Date"]).dt.days + 1
        non_positive_days_mask = (~missing_dt_mask) & (days <= 0)
        old_start_mask = (~work["Start Date"].isna()) & (work["Start Date"] < START_CUTOFF)
        future_completion_mask = (~work["Complete Date"].isna()) & (work["Complete Date"] >= TODAY)
        tw_out_of_range_mask = (~work["Tower Weight"].isna()) & (
            (work["Tower Weight"] < TOWER_MIN_MT) | (work["Tower Weight"] > TOWER_MAX_MT)
        )

        # Productivity (only where days valid)
        prod = pd.Series(np.nan, index=work.index, dtype="float")
        valid_for_prod = (~missing_dt_mask) & (days > 0)
        prod[valid_for_prod] = work.loc[valid_for_prod, "Tower Weight"] / days[valid_for_prod]
        work["Productivity"] = prod

        data_issues_rows: List[pd.DataFrame] = []

        def push_data_issue(mask, reason: str):
            if mask.any():
                sub = work.loc[
                    mask,
                    [
                        "Start Date",
                        "Complete Date",
                        "Gang name",
                        "Tower Weight",
                        "Tower Type",
                        "Productivity",
                        "Project Code",
                        "Line Name",
                        "Project Name",
                        "Project Display",
                        "Project Scope Key",
                        "Source File",
                        "Source Sheet",
                        "Location No.",
                        "Tower Tightening Raw",
                        "Tower Tightening",
                        "Status",
                    ],
                ].copy()
                sub["Issues"] = reason
                data_issues_rows.append(sub)

        push_data_issue(missing_dt_mask, "Missing start/end date (not expanded)")
        push_data_issue(non_positive_days_mask, "Non-positive duration (Start > End or 0) (not expanded)")
        push_data_issue(old_start_mask, f"Start before {START_CUTOFF.date()} (not expanded)")
        push_data_issue(future_completion_mask, f"Completion on/after {TODAY.date()} (not expanded)")
        push_data_issue(
            tw_out_of_range_mask,
            f"Tower Weight out of range (<{TOWER_MIN_MT} or >{TOWER_MAX_MT}) (not expanded)",
        )

        invalid_mask = (
            missing_dt_mask | non_positive_days_mask | old_start_mask | future_completion_mask | tw_out_of_range_mask
        )
        work_valid = work.loc[~invalid_mask].copy()

        per_erection = work[
            [
                "Start Date",
                "Complete Date",
                "Gang name",
                "Tower Weight",
                "Tower Type",
                "Productivity",
                "Project Code",
                "Line Name",
                "Project Name",
                "Project Display",
                "Project Scope Key",
                "Source File",
                "Source Sheet",
                "Location No.",
                "Tower Tightening Raw",
                "Tower Tightening",
                "Status",
            ]
        ].copy()
        per_day_with_singles = expand_per_day(work_valid)
        data_issues_df = pd.concat(data_issues_rows, ignore_index=True) if data_issues_rows else pd.DataFrame()

        if not per_day_with_singles.empty:
            per_day_frames.append(per_day_with_singles)
        if not per_erection.empty:
            per_erection_frames.append(per_erection)
        if not data_issues_df.empty:
            data_issue_frames.append(data_issues_df)
        diagnostics.append(diag)

    per_day_consol = pd.concat(per_day_frames, ignore_index=True) if per_day_frames else pd.DataFrame()
    per_erection_consol = pd.concat(per_erection_frames, ignore_index=True) if per_erection_frames else pd.DataFrame()
    data_issues_consol = pd.concat(data_issue_frames, ignore_index=True) if data_issue_frames else pd.DataFrame()
    return per_day_consol, per_erection_consol, diagnostics, issues, data_issues_consol


# ---------- Styling ----------
def style_sheet(ws, tab_color="99CCFF"):
    """Minimalist styling: colored tab + bold header with fill + borders + freeze header + friendly widths."""
    try:
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    except Exception:
        return  # styling optional; skip if openpyxl is limited

    # Tab color
    ws.sheet_properties.tabColor = tab_color

    if ws.max_row < 1 or ws.max_column < 1:
        return

    # Freeze header
    ws.freeze_panes = "A2"

    # Header styling
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for c in ws[1]:
        c.fill = header_fill
        c.font = bold_font
        c.border = border
        c.alignment = Alignment(vertical="center")

    # Light borders on data cells + basic width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # Column widths (simple heuristic)
    for col_cells in ws.columns:
        max_len = 10
        for cell in col_cells[: min(200, ws.max_row)]:  # don't scan entire huge sheets
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 45)


# ---------- Main ----------
# def main():
def main(argv=None):
    if not logging.getLogger().handlers:
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s [%(levelname)s] %(message)s",
        )
    logger.info("Starting erection compiled pipeline")

    ap = argparse.ArgumentParser(
        description="Parse 'Erection Compiled', compute productivity, expand to daily rows."
    )
    grp = ap.add_mutually_exclusive_group(required=True)
    grp.add_argument("--input", help="Folder containing .xlsx/.xlsm files")
    grp.add_argument("--files", nargs="+", help="Explicit list of .xlsx/.xlsm files")
    ap.add_argument("--output", required=True, help="Output Excel path")
    # args = ap.parse_args()
    args = ap.parse_args(argv)

    # Resolve input files
    paths: List[Path] = []
    input_folder: Optional[Path] = None
    if args.input:
        input_folder = Path(args.input)
        if not input_folder.exists():
            raise SystemExit(f"Input folder not found: {input_folder}")
        for fp in input_folder.iterdir():
            if fp.is_file() and fp.suffix.lower() in (".xlsx", ".xlsm"):
                if fp.name.startswith("~$"):
                    continue
                name_lower = fp.name.lower()
                if any(k in name_lower for k in ("consolidated", "output", "compiled")) and "erection" not in name_lower:
                    continue
                paths.append(fp)
    else:
        paths = [Path(p) for p in args.files]

    all_per_day_with_singles, all_per_erection = [], []
    all_issues, all_diag = [], []
    all_data_issues = []
    all_proj_details = []
    erection_sheet_config = load_erection_sheet_config(input_folder or (paths[0].parent if paths else None))
    erection_template_config, erection_template_errors = load_erection_template_mapping_config(
        input_folder or (paths[0].parent if paths else None)
    )
    erection_section_config = load_erection_section_config(input_folder or (paths[0].parent if paths else None))
    skipped_no_erection = 0

    for p in paths:
        if not p.exists():
            all_issues.append({"file": p.name, "issue": "missing"})
            continue

        project_code = parse_project_from_filename(p.name)
        project_key = _normalize_project_code_key(project_code)
        configured_erection_sheets = erection_sheet_config.get(project_key)
        configured_erection_sections = erection_section_config.get(project_key)
        template_column_map = None
        template_sheet_name = None
        template_mappings = None
        template_error = erection_template_errors.get(project_key)
        if project_key in erection_template_config:
            template_mappings = erection_template_config[project_key]
            selected_project_template = _select_template_mapping_for_request(template_mappings)
            if selected_project_template:
                template_column_map = selected_project_template.get("column_map")
                template_sheet_name = str(selected_project_template.get("template_sheet", "")).strip() or None
        if template_error:
            all_issues.append(
                {
                    "file": p.name,
                    "issue": f"Template mapping configuration error: {template_error}",
                }
            )
            continue
        if configured_erection_sheets is not None and not configured_erection_sheets:
            skipped_no_erection += 1
            all_issues.append(
                {
                    "file": p.name,
                    "issue": "Skipped: no erection sheet configured for this project in DPR_Config.",
                }
            )
            continue

        try:
            per_day_with_singles, per_erection, diagnostics, issues, data_issues_df = process_file(
                p,
                configured_sheet_names=configured_erection_sheets,
                template_column_map=template_column_map,
                template_sheet_name=template_sheet_name,
                template_mappings=template_mappings,
                section_config=configured_erection_sections,
            )
        except Exception as e:
            # Guardrail: never let a single bad file crash the whole pipeline
            all_issues.append({"file": p.name, "issue": f"Unhandled processing error: {e}"})
            continue

        # --- NEW: attempt to read "Project Details" from this source ---
        try:
            raw_pd, _pd_sheet, pd_note = load_sheet_with_csv_fallback(
                p,
                find_project_details_sheet,
                read_excel_kwargs={"header": None},
                read_csv_kwargs={"header": None},
            )
            if raw_pd is not None and not raw_pd.empty:
                dfp = load_project_details_from_source(raw_pd, p)
                if not dfp.empty:
                    fn_project_name = parse_project_from_filename(p.name)
                    dfp["Project Name"] = fn_project_name
                    all_proj_details.append(dfp)
            if pd_note:
                logger.warning("Fallback note for '%s': %s", p.name, pd_note)
                all_issues.append({"file": p.name, "issue": pd_note})
        except Exception as e:
            all_issues.append({"file": p.name, "issue": f"Project Details read error: {e}"})

        if not per_day_with_singles.empty:
            all_per_day_with_singles.append(per_day_with_singles.assign(_source_file=p.name))
        if not per_erection.empty:
            all_per_erection.append(per_erection.assign(_source_file=p.name))
        if not data_issues_df.empty:
            all_data_issues.append(data_issues_df.assign(_source_file=p.name))

        if diagnostics:
            all_diag.extend(diagnostics)
        all_issues.extend(issues)

    if skipped_no_erection:
        logger.info(
            "Skipped %d workbook(s) because DPR_Config has no erection sheet configured.",
            skipped_no_erection,
        )


    # Consolidate across all inputs
    per_day_with_singles_consol = pd.concat(all_per_day_with_singles, ignore_index=True) if all_per_day_with_singles else pd.DataFrame()
    per_erection_consol = pd.concat(all_per_erection, ignore_index=True) if all_per_erection else pd.DataFrame()
    data_issues_consol = pd.concat(all_data_issues, ignore_index=True) if all_data_issues else pd.DataFrame()
    issues_df = pd.DataFrame(all_issues) if all_issues else pd.DataFrame()
    diag_df = pd.DataFrame(all_diag) if all_diag else pd.DataFrame()
    projdetails_df = pd.DataFrame()
    projdetails_out = pd.DataFrame()
        # --- NEW: consolidate Project Details across inputs ---
        # --- consolidate Project Details across inputs (NEW) ---
    if all_proj_details:
        projdetails_df = pd.concat(all_proj_details, ignore_index=True)

        # Deduplicate by project_code; latest file wins
        projdetails_df = (
            projdetails_df.sort_values("_source_file")
                        .drop_duplicates(subset=["project_code"], keep="last")
        )

        # Order & friendly headers (align with other sheets' style)
        projdetails_out = projdetails_df.rename(columns={
            "project_code": "Project Code",
            "client_name": "Client Name",
            "noa_start": "NOA Start Date",
            "loa_end": "LOA End Date",
            "project_mgr": "Project Manager",
            "regional_mgr": "Regional Manager",
            "planning_eng": "Planning Engineer",
            "pch": "PCH",
            "section_inch": "Section Incharge",
            "supervisor": "Supervisor",
        })[
            [
                "Project Code",
                "Project Name",
                "project_name",           # <-- ensure this is present
                "Client Name",
                "NOA Start Date",
                "LOA End Date",
                "Project Manager",
                "Regional Manager",
                "Planning Engineer",
                "PCH",
                "Section Incharge",
                "Supervisor",
            ]
        ]
    else:
        projdetails_out = pd.DataFrame()



    # README / Assumptions
    readme_lines = [
        "Assumptions & Cleaning Rules:",
        f"- Start date cutoff: rows with Start Date before {START_CUTOFF.date()} are not expanded and logged in 'Data Issues'.",
        f"- Completion date must be before {TODAY.date()} (future completions go to 'Data Issues').",
        f"- Tower Weight range: only [{TOWER_MIN_MT:.0f}, {TOWER_MAX_MT:.0f}] MT is considered valid for expansion; out-of-range rows go to 'Data Issues'.",
        "- Missing or non-positive durations (Start/End missing or Start > End or 0) are logged to 'Data Issues' and not expanded.",
        "- Ambiguous numeric dates (e.g., 1/12/2025) that initially parse as Start > End are retried with month-first before being marked invalid.",
        "- Gang name normalization: remove special characters (digits kept), Title Case words, and insert a space before trailing digits (e.g., 'xyz4' â†’ 'Xyz 4').",
        f"- If 'Tower Weight' header is missing in a sheet, Tower Weight is assumed as {DEFAULT_TOWER_WEIGHT_ASSUMED_MT:.0f} MT for all rows in that sheet (noted in 'Issues').",
        "- If some Tower Weight values are present and some are blank in a sheet, blank rows are filled with the average of available in-range tower weights for that sheet (noted in 'Issues').",
        f"- If a sheet has no usable in-range Tower Weight values at all, blank rows fall back to {DEFAULT_TOWER_WEIGHT_ASSUMED_MT:.0f} MT (noted in 'Issues').",
        "- DPR_Config support: when multiple erection sheet names are configured for a project, each listed sheet is processed independently, assigned its own line identity, and then concatenated.",
        "- DPR_Config template mapping: for projects with discipline-specific template check marked Yes, column-wise mapping from the discipline template tab is applied before required-field validation.",
        "- DPR_Config section mapping: optional '<Project> Erection Sections' tabs can split a physical sheet by marker text, filter rows, derive line names from a column, and use section-specific template tabs.",
        "- Sectioned rows with missing Tower Weight first use same-file historical averages for the same Line Name before falling back to the default assumption.",
        "- Sheets:",
        "    â€¢ ProdDailyExpanded  : per-day expanded rows used by the dashboard",
        "    â€¢ ProdDailyExpandedSingles : per-day expanded rows including single-occurrence gangs",
        "    â€¢ RawData        : per-erection rows (unexpanded), for traceability",
        "    â€¢ Data Issues    : row-level data problems (reason in 'Issues' column)",
        "    â€¢ Issues         : file-level problems (missing sheet/headers, read/open errors, etc.)",
        "    â€¢ Diagnostics    : which sheet used, detected header row, and normalized header text",
        "- Dashboard note: a 15-day productivity loss cap is used in downstream analytics.",
    ]
    readme_df = pd.DataFrame({"Notes": readme_lines})

    # Write output workbook (+ minimal styling)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        if not per_day_with_singles_consol.empty:
            per_day_with_singles_consol.drop(columns=["_source_file"], errors="ignore").to_excel(w, sheet_name="ProdDailyExpandedSingles", index=False)
        if not per_erection_consol.empty:
            per_erection_consol.drop(columns=["_source_file"], errors="ignore").to_excel(w, sheet_name="RawData", index=False)
        if not data_issues_consol.empty:
            data_issues_consol.drop(columns=["_source_file"], errors="ignore").to_excel(w, sheet_name="Data Issues", index=False)
        if not issues_df.empty:
            issues_df.to_excel(w, sheet_name="Issues", index=False)
        if not diag_df.empty:
            diag_df.to_excel(w, sheet_name="Diagnostics", index=False)
        # --- NEW: write consolidated Project Details ---
        if not projdetails_df.empty:
            projdetails_df.to_excel(w, sheet_name="ProjectDetails", index=False)
        readme_df.to_excel(w, sheet_name="README_Assumptions", index=False)

        # Apply styling
        wb = w.book
        for sheet_name, color in [
            ("ProdDailyExpandedSingles", "9CC3E6"),  # blue variant including singles
            ("RawData", "C6E0B4"),         # green
            ("Data Issues", "F8CBAD"),     # light red
            ("Issues", "D9D2E9"),          # purple
            ("Diagnostics", "FFE699"),     # yellow
            ("README_Assumptions", "99CCFF"),
            ("ProjectDetails", "99E6E6"),  # --- NEW ---
        ]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                style_sheet(ws, tab_color=color)

    print(f"Done. Wrote {out_path}")


def run_pipeline(input_path=None, output_path=None, files=None, extra_args=None):
    """Convenience wrapper to run the CLI pipeline programmatically."""
    cli_args = []
    if input_path:
        cli_args.extend(["--input", str(input_path)])
    elif files:
        cli_args.append("--files")
        cli_args.extend(str(p) for p in files)
    else:
        raise ValueError("Either input_path or files must be provided.")

    if output_path:
        cli_args.extend(["--output", str(output_path)])
    else:
        raise ValueError("An output_path is required to write the compiled workbook.")

    if extra_args:
        cli_args.extend(list(extra_args))

    main(cli_args)


if __name__ == "__main__":
    main()
