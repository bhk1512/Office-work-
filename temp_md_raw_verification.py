from __future__ import annotations

from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from erection_compiled_to_daily_new import load_sheet_with_csv_fallback


ROOT = Path.cwd()
RAW_DIR = ROOT / "Raw Data" / "DPRs"
OUT = ROOT / "Productivity Summaries" / "Foundation_Erection_MD_Raw_Verification_2026-08-13.xlsx"
OBS_DATE = pd.Timestamp("2026-08-12")


def compact(value: object) -> str:
    text = "" if value is None else str(value).strip().lower()
    if text in {"", "nan", "none", "null", "nat"}:
        return ""
    text = re.sub(r"^n['`’]\s*", "", text, flags=re.I)
    return re.sub(r"[^a-z0-9]", "", text)


def loc_alias(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = re.sub(r"^N['`’]\s*", "", text, flags=re.I)
    text = re.sub(r"\s+[A-Za-z](?:\s*,\s*[A-Za-z])+$", "", text)
    text = re.sub(r"\s+[A-Za-z]$", "", text)
    return compact(text)


def base_project_key(value: object) -> str:
    text = compact(value)
    match = re.search(r"(ta|tb)(\d{3,4})", text)
    return f"{match.group(1)}{int(match.group(2))}" if match else text


def sheet_selector_exact(target: object):
    def select(names: list[str]) -> str | None:
        for name in names:
            if str(name).strip().lower() == str(target).strip().lower():
                return name
        return None

    return select


def norm_date(value: object) -> pd.Timestamp | None:
    if value is None:
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.notna(parsed):
        return pd.Timestamp(parsed).normalize()
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.notna(numeric) and 20000 <= float(numeric) <= 80000:
        converted = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")
        if pd.notna(converted):
            return pd.Timestamp(converted).normalize()
    return None


def row_has_loc_and_date(df: pd.DataFrame, loc: object, dt: object) -> tuple[bool, str, str]:
    if df is None or df.empty or pd.isna(dt):
        return False, "", ""
    loc_norm = compact(loc)
    target = pd.Timestamp(dt).normalize()
    for row_idx in range(min(len(df.index), 1000)):
        values = df.iloc[row_idx].tolist()
        loc_cols: list[int] = []
        date_cols: list[int] = []
        for col_idx, value in enumerate(values):
            if compact(value) == loc_norm:
                loc_cols.append(col_idx + 1)
            parsed = norm_date(value)
            if parsed is not None and parsed == target:
                date_cols.append(col_idx + 1)
        if loc_cols and date_cols:
            return True, str(row_idx + 1), f"loc_cols={loc_cols}; date_cols={date_cols}"
    return False, "", ""


def main() -> int:
    foundation = pd.read_parquet(ROOT / "Parquets" / "Foundation" / "FoundationCompletions.parquet")
    erection = pd.read_parquet(ROOT / "Parquets" / "Erection" / "RawData.parquet")
    workbook_raw = pd.read_excel(
        ROOT / "Productivity Summaries" / "Foundation_Erection_Gap_Raw_and_Analysis_Q1_Q2_2026.xlsx",
        sheet_name="Raw Project Wise",
    )

    foundation = foundation[foundation["source_type"].astype(str).str.lower().eq("detail")].copy()
    foundation["foundation_date"] = pd.to_datetime(foundation["event_date"], errors="coerce").dt.normalize()
    foundation = foundation[foundation["foundation_date"].notna()].copy()
    erection = erection.copy()
    erection["start_date"] = pd.to_datetime(erection["Start Date"], errors="coerce").dt.normalize()
    erection["complete_date"] = pd.to_datetime(erection["Complete Date"], errors="coerce").dt.normalize()

    foundation["project_key"] = foundation["project_code"].map(base_project_key)
    foundation["line_key"] = foundation["line_name"].map(compact)
    foundation["loc_key"] = foundation["location_no"].map(compact)
    foundation["loc_alias_key"] = foundation["location_no"].map(loc_alias)
    erection["project_key"] = erection["Project Code"].map(base_project_key)
    erection["line_key"] = erection["Line Name"].map(compact)
    erection["loc_key"] = erection["Location No."].map(compact)
    erection["loc_alias_key"] = erection["Location No."].map(loc_alias)

    foundation_events = (
        foundation.sort_values(["project_key", "foundation_date", "line_key", "location_no"])
        .groupby(["project_key", "line_key", "loc_key"], as_index=False)
        .agg(
            project_code=("project_code", "first"),
            project_display=("project_display", "first"),
            location_no=("location_no", "first"),
            loc_alias_key=("loc_alias_key", "first"),
            foundation_date=("foundation_date", "min"),
            foundation_source_file=("source_file", "first"),
            foundation_source_sheet=("source_sheet", "first"),
        )
    )
    erection_events = (
        erection[erection["start_date"].notna()]
        .sort_values(["project_key", "start_date", "complete_date", "line_key", "Location No."])
        .groupby(["project_key", "line_key", "loc_key"], as_index=False)
        .agg(
            erection_location_no=("Location No.", "first"),
            erection_start_date=("start_date", "min"),
            erection_complete_date=("complete_date", "first"),
            erection_source_file=("Source File", "first"),
            erection_source_sheet=("Source Sheet", "first"),
        )
    )

    merged = foundation_events.merge(
        erection_events.rename(
            columns={
                "loc_key": "ere_loc_key",
                "erection_location_no": "erection_location_no_exact",
                "erection_start_date": "erection_start_exact",
                "erection_complete_date": "erection_complete_exact",
                "erection_source_file": "erection_source_file_exact",
                "erection_source_sheet": "erection_source_sheet_exact",
            }
        ),
        left_on=["project_key", "line_key", "loc_key"],
        right_on=["project_key", "line_key", "ere_loc_key"],
        how="left",
    )
    alias_events = erection_events.rename(
        columns={
            "loc_key": "ere_alias_key",
            "erection_location_no": "erection_location_no_alias",
            "erection_start_date": "erection_start_alias",
            "erection_complete_date": "erection_complete_alias",
            "erection_source_file": "erection_source_file_alias",
            "erection_source_sheet": "erection_source_sheet_alias",
        }
    )
    merged = merged.merge(
        alias_events,
        left_on=["project_key", "line_key", "loc_alias_key"],
        right_on=["project_key", "line_key", "ere_alias_key"],
        how="left",
    )
    merged["erection_location_no"] = merged["erection_location_no_exact"].where(
        merged["erection_location_no_exact"].fillna("").astype(str).str.strip().astype(bool),
        merged["erection_location_no_alias"],
    )
    for target, exact, alias in (
        ("erection_start", "erection_start_exact", "erection_start_alias"),
        ("erection_complete", "erection_complete_exact", "erection_complete_alias"),
        ("erection_source_file", "erection_source_file_exact", "erection_source_file_alias"),
        ("erection_source_sheet", "erection_source_sheet_exact", "erection_source_sheet_alias"),
    ):
        exact_series = merged[exact]
        mask = exact_series.notna() if pd.api.types.is_datetime64_any_dtype(exact_series) else exact_series.fillna("").astype(str).str.strip().astype(bool)
        merged[target] = exact_series.where(mask, merged[alias])
    merged["match_basis"] = "unmatched"
    merged.loc[merged["erection_start_exact"].notna(), "match_basis"] = "exact"
    merged.loc[merged["erection_start_exact"].isna() & merged["erection_start_alias"].notna(), "match_basis"] = "alias"
    merged["delay_to_start_days"] = (merged["erection_start"] - merged["foundation_date"]).dt.days
    merged["delay_to_completion_days"] = (merged["erection_complete"] - merged["foundation_date"]).dt.days
    merged["foundation_month"] = merged["foundation_date"].dt.to_period("M").astype(str)
    merged["exposure_days"] = (OBS_DATE - merged["foundation_date"]).dt.days
    for horizon in (30, 60):
        merged[f"eligible_{horizon}d"] = merged["exposure_days"].ge(horizon)
        merged[f"start_within_{horizon}d"] = merged["delay_to_start_days"].between(0, horizon, inclusive="both")
        merged[f"completion_within_{horizon}d"] = merged["delay_to_completion_days"].between(0, horizon, inclusive="both")

    def horizon_row(df: pd.DataFrame, label: str, horizon: int, basis: str) -> dict[str, object]:
        eligible = df[df[f"eligible_{horizon}d"]]
        denom = len(eligible)
        col = f"{basis}_within_{horizon}d"
        within = int(eligible[col].fillna(False).sum())
        return {
            "Cohort": label,
            "Basis": "Foundation to erection START" if basis == "start" else "Foundation to erection COMPLETION",
            "Horizon Days": horizon,
            "Foundation Locations": int(len(df)),
            "Eligible Locations": int(denom),
            f"Erected <= {horizon}D": within,
            f"% <= {horizon}D": round(within / denom * 100, 1) if denom else pd.NA,
            "Full Cohort Exposure": "Yes" if denom == len(df) and len(df) else "No",
        }

    horizon_rows: list[dict[str, object]] = []
    for basis in ("start", "completion"):
        for horizon in (30, 60):
            for year in (2025, 2026):
                for month in (4, 5, 6):
                    period = pd.Period(f"{year}-{month:02d}", freq="M")
                    horizon_rows.append(
                        horizon_row(
                            merged[merged["foundation_date"].dt.to_period("M").eq(period)],
                            str(period),
                            horizon,
                            basis,
                        )
                    )
    horizon_compare = pd.DataFrame(horizon_rows)

    aggregate_rows: list[dict[str, object]] = []
    for basis in ("start", "completion"):
        for year in (2025, 2026):
            periods = set(pd.period_range(f"{year}-04", f"{year}-06", freq="M"))
            aggregate_rows.append(
                horizon_row(
                    merged[merged["foundation_date"].dt.to_period("M").isin(periods)],
                    f"Apr-Jun {year}",
                    30,
                    basis,
                )
            )
    aggregate = pd.DataFrame(aggregate_rows)

    workbook_key = workbook_raw.copy()
    workbook_key["foundation_date"] = pd.to_datetime(workbook_key["Foundation Date"], errors="coerce").dt.normalize()
    workbook_key["project_key"] = workbook_key["Project"].map(base_project_key)
    workbook_key["loc_key"] = workbook_key["Location"].map(compact)
    workbook_key["delay_workbook"] = pd.to_numeric(workbook_key["Delay Days"], errors="coerce")
    compare = merged.merge(
        workbook_key[["project_key", "loc_key", "foundation_date", "delay_workbook"]],
        on=["project_key", "loc_key", "foundation_date"],
        how="left",
    )
    compare["workbook_delay_matches_recompute"] = (
        (compare["delay_workbook"].isna() & compare["delay_to_start_days"].isna())
        | (pd.to_numeric(compare["delay_workbook"], errors="coerce") == pd.to_numeric(compare["delay_to_start_days"], errors="coerce"))
    )
    reconcile = pd.DataFrame(
        [
            {"Check": "Foundation detail events", "Rows": len(foundation_events), "Pass Rows": len(foundation_events), "Pass %": 100.0},
            {
                "Check": "Matched to erection start",
                "Rows": len(merged),
                "Pass Rows": int(merged["erection_start"].notna().sum()),
                "Pass %": round(merged["erection_start"].notna().mean() * 100, 1),
            },
            {
                "Check": "Workbook delay equals recomputed start delay",
                "Rows": len(compare),
                "Pass Rows": int(compare["workbook_delay_matches_recompute"].sum()),
                "Pass %": round(compare["workbook_delay_matches_recompute"].mean() * 100, 2),
            },
        ]
    )

    sheet_cache: dict[tuple[object, object], tuple[pd.DataFrame, object, object]] = {}

    def load_raw_sheet(file_name: object, sheet_name: object) -> tuple[pd.DataFrame, object, object]:
        key = (file_name, sheet_name)
        if key in sheet_cache:
            return sheet_cache[key]
        df, actual, note = load_sheet_with_csv_fallback(
            RAW_DIR / str(file_name),
            sheet_selector_exact(sheet_name),
            read_excel_kwargs={"header": None},
            read_csv_kwargs={"header": None},
        )
        if df is None:
            df = pd.DataFrame()
        sheet_cache[key] = (df, actual, note)
        return sheet_cache[key]

    sample_pool = merged[merged["foundation_month"].isin(["2026-04", "2026-05", "2026-06"])].copy()
    hits = sample_pool[sample_pool["start_within_30d"]].sample(n=min(12, int(sample_pool["start_within_30d"].sum())), random_state=13)
    misses = sample_pool[~sample_pool["start_within_30d"]].sample(n=min(12, int((~sample_pool["start_within_30d"]).sum())), random_state=17)
    focus_projects = {"TA 413", "TA 414", "TA 419", "TA 512", "TA 602", "TB 507", "TB 608"}
    project_reps = (
        sample_pool[sample_pool["project_code"].isin(focus_projects)]
        .sort_values(["project_code", "foundation_date"])
        .groupby("project_code", as_index=False)
        .head(1)
    )
    tb408 = merged[merged["project_code"].eq("TB 408")]
    sample = (
        pd.concat([hits, misses, project_reps, tb408], ignore_index=True)
        .drop_duplicates(subset=["project_key", "line_key", "loc_key", "foundation_date"])
        .head(40)
    )

    trace_rows: list[dict[str, object]] = []
    for _, row in sample.iterrows():
        f_found = f_row = f_note = ""
        e_start_found = e_start_row = e_start_note = ""
        e_complete_found = e_complete_row = e_complete_note = ""
        f_load_note = e_load_note = ""
        try:
            fdf, _, load_note = load_raw_sheet(row["foundation_source_file"], row["foundation_source_sheet"])
            f_load_note = load_note or ""
            ok, row_number, note = row_has_loc_and_date(fdf, row["location_no"], row["foundation_date"])
            f_found, f_row, f_note = ("Yes" if ok else "No"), row_number, note
        except Exception as exc:
            f_found, f_note = "ERROR", str(exc)[:180]

        if pd.notna(row["erection_start"]) and str(row.get("erection_source_file", "")).strip():
            try:
                edf, _, load_note = load_raw_sheet(row["erection_source_file"], row["erection_source_sheet"])
                e_load_note = load_note or ""
                loc = row["erection_location_no"] if pd.notna(row["erection_location_no"]) else row["location_no"]
                ok, row_number, note = row_has_loc_and_date(edf, loc, row["erection_start"])
                e_start_found, e_start_row, e_start_note = ("Yes" if ok else "No"), row_number, note
                ok, row_number, note = row_has_loc_and_date(edf, loc, row["erection_complete"])
                e_complete_found, e_complete_row, e_complete_note = ("Yes" if ok else "No"), row_number, note
            except Exception as exc:
                e_start_found, e_start_note = "ERROR", str(exc)[:180]
        else:
            e_start_found = "Unmatched"
            e_complete_found = "Unmatched"

        trace_rows.append(
            {
                "Project": row["project_code"],
                "Location": row["location_no"],
                "Foundation Date": row["foundation_date"],
                "Erection Start": row["erection_start"],
                "Erection Complete": row["erection_complete"],
                "Delay to Start": row["delay_to_start_days"],
                "Delay to Completion": row["delay_to_completion_days"],
                "Start <=30D": bool(row["start_within_30d"]),
                "Completion <=30D": bool(row["completion_within_30d"]),
                "Match Basis": row["match_basis"],
                "Foundation Source File": row["foundation_source_file"],
                "Foundation Sheet": row["foundation_source_sheet"],
                "Raw Foundation Loc+Date Found": f_found,
                "Raw Foundation Row": f_row,
                "Raw Foundation Evidence": f_note,
                "Foundation Load Note": f_load_note,
                "Erection Source File": row["erection_source_file"],
                "Erection Sheet": row["erection_source_sheet"],
                "Raw Erection Loc+Start Found": e_start_found,
                "Raw Erection Start Row": e_start_row,
                "Raw Erection Start Evidence": e_start_note,
                "Raw Erection Loc+Complete Found": e_complete_found,
                "Raw Erection Complete Row": e_complete_row,
                "Raw Erection Complete Evidence": e_complete_note,
                "Erection Load Note": e_load_note,
            }
        )
    trace = pd.DataFrame(trace_rows)
    matched_start = trace[~trace["Raw Erection Loc+Start Found"].eq("Unmatched")]
    matched_complete = trace[~trace["Raw Erection Loc+Complete Found"].eq("Unmatched")]
    trace_summary = pd.DataFrame(
        [
            {
                "Trace Check": "Sample foundation loc+date found in raw DPR sheet",
                "Rows": len(trace),
                "Pass Rows": int(trace["Raw Foundation Loc+Date Found"].eq("Yes").sum()),
                "Pass %": round(trace["Raw Foundation Loc+Date Found"].eq("Yes").mean() * 100, 1),
            },
            {
                "Trace Check": "Sample matched erection loc+start found in raw DPR sheet",
                "Rows": len(matched_start),
                "Pass Rows": int(matched_start["Raw Erection Loc+Start Found"].eq("Yes").sum()),
                "Pass %": round(matched_start["Raw Erection Loc+Start Found"].eq("Yes").mean() * 100, 1) if len(matched_start) else pd.NA,
            },
            {
                "Trace Check": "Sample matched erection loc+complete found in raw DPR sheet",
                "Rows": len(matched_complete),
                "Pass Rows": int(matched_complete["Raw Erection Loc+Complete Found"].eq("Yes").sum()),
                "Pass %": round(matched_complete["Raw Erection Loc+Complete Found"].eq("Yes").mean() * 100, 1) if len(matched_complete) else pd.NA,
            },
        ]
    )

    notes = pd.DataFrame(
        [
            {
                "Point": "Definition used in current analysis",
                "Detail": "Foundation completion date to tower erection START date. A row is <=30D if erection start date - foundation completion date is between 0 and 30 days.",
            },
            {
                "Point": "Alternative checked here",
                "Detail": "Foundation completion date to tower erection COMPLETION date is also computed. It is stricter and gives lower <=30D percentages.",
            },
            {
                "Point": "Recommendation",
                "Detail": "For MD, label the chosen metric explicitly as start-based pickup. If leadership expects complete erection within 30 days, use the completion-based column instead.",
            },
            {
                "Point": "Raw-file check",
                "Detail": "Trace Audit opens the referenced raw DPR sheets and searches for the same location and date values used in the analysis.",
            },
        ]
    )

    fact_cols = [
        "project_code",
        "location_no",
        "foundation_date",
        "erection_start",
        "erection_complete",
        "delay_to_start_days",
        "delay_to_completion_days",
        "foundation_month",
        "exposure_days",
        "start_within_30d",
        "completion_within_30d",
        "match_basis",
        "foundation_source_file",
        "foundation_source_sheet",
        "erection_source_file",
        "erection_source_sheet",
    ]
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        notes.to_excel(writer, sheet_name="Definition", index=False)
        reconcile.to_excel(writer, sheet_name="Recompute Checks", index=False)
        aggregate.to_excel(writer, sheet_name="30D Start vs Completion", index=False)
        horizon_compare.to_excel(writer, sheet_name="Apr-Jun Horizon Detail", index=False)
        trace_summary.to_excel(writer, sheet_name="Raw Trace Summary", index=False)
        trace.to_excel(writer, sheet_name="Raw Trace Sample", index=False)
        merged[fact_cols].to_excel(writer, sheet_name="Recomputed Raw Fact", index=False)

    workbook = load_workbook(OUT)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    for ws in workbook.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 22
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(OUT)

    print("OUTPUT", OUT)
    print("RECONCILE")
    print(reconcile.to_string(index=False))
    print("30D AGG")
    print(aggregate.to_string(index=False))
    print("TRACE SUMMARY")
    print(trace_summary.to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
