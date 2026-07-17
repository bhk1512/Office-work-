from __future__ import annotations

import copy
import hashlib
import json
import math
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook

import erection_compiled_to_daily_new as erection
from dashboard import stringing_ingest
from dashboard.project_identity import normalize_line_name, parse_project_identity_from_filename
from pipeline_runner import _apply_template_if_improves, _sanitize_stringing_columns


REPO = Path(r"C:\Users\kaushikb\Documents\Work\Git\Office-work-")
DOWNLOADS = Path(r"C:\Users\kaushikb\Downloads")
SOURCE = DOWNLOADS / "Copy of TnDIndia_Locations.xlsx"
OUTPUT = DOWNLOADS / "Copy of TnDIndia_Locations_filled_2026-07-17.xlsx"
PROJECT_DIR = DOWNLOADS / "TnDIndia_Locations_Project_Confirmation_2026-07-17"
RAW_ROOT = REPO / "Raw Data" / "DPRs"
CONFIG = REPO / "Raw Data" / "DPR_Config.xlsx"

TARGET_HEADERS = [
    "ProjectID",
    "ProjectCode",
    "SectionID",
    "Section",
    "LocationID",
    "LocationNo",
    "PipeLength",
    "Geolocation",
    "Tower Type",
    "Weight",
    "Stringing Method",
]
IMMUTABLE_HEADERS = TARGET_HEADERS[:8]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            block = handle.read(1024 * 1024)
            if not block:
                break
            digest.update(block)
    return digest.hexdigest()


def project_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def canonical_project_label(key: str) -> str:
    match = re.fullmatch(r"(ta|tb)(\d+)", key.lower())
    if not match:
        return key.upper()
    return f"{match.group(1).upper()} {match.group(2)}"


def text_value(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).replace("\u00a0", " ").strip()
    if text.lower() in {"", "nan", "none", "null", "nat", "#n/a", "#value!"}:
        return ""
    return text


def location_key(value: object) -> str:
    text = text_value(value).upper()
    if not text:
        return ""
    text = re.sub(r"^\s*AP[\s\-_./]*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^\s*T[\s\-_]+(?=\d)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", "", text)
    text = text.strip("-_.")
    # Canonical tower labels: AP25, 25 and 25/0 all represent the same AP.
    match = re.fullmatch(r"0*(\d+)([A-Z]+)?(?:/0*(\d+))?", text)
    if match:
        main = str(int(match.group(1)))
        letters = match.group(2) or ""
        sub = str(int(match.group(3) or "0"))
        return f"{main}{letters}/{sub}"
    if re.fullmatch(r"GANTRY(?:/0)?", text):
        return "GANTRY"
    return text


def exact_sheet_selector(expected: str):
    expected_key = re.sub(r"\s+", " ", expected.strip()).casefold()

    def select(names: list[str]) -> str | None:
        for name in names:
            if re.sub(r"\s+", " ", str(name).strip()).casefold() == expected_key:
                return name
        return None

    return select


def latest_dpr_files() -> dict[str, list[Path]]:
    grouped: dict[str, list[Path]] = defaultdict(list)
    for path in RAW_ROOT.iterdir():
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            continue
        identity = parse_project_identity_from_filename(path.name)
        key = project_key(identity.get("project_code", ""))
        if key:
            grouped[key].append(path)

    selected: dict[str, list[Path]] = {}
    for key, paths in grouped.items():
        if key == "tb507":
            main = [path for path in paths if "[MAIN]" in path.name.upper()]
            other = [path for path in paths if "[MAIN]" not in path.name.upper()]
            chosen: list[Path] = []
            if main:
                chosen.append(max(main, key=lambda path: (path.stat().st_mtime, path.name)))
            if other:
                chosen.append(max(other, key=lambda path: (path.stat().st_mtime, path.name)))
            selected[key] = chosen
        else:
            selected[key] = [max(paths, key=lambda path: (path.stat().st_mtime, path.name))]
    return selected


def header_map(ws) -> dict[str, int]:
    values = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return {str(value).strip(): index + 1 for index, value in enumerate(values) if value is not None}


def load_target_rows() -> tuple[list[dict[str, Any]], dict[str, int], list[str]]:
    wb = load_workbook(SOURCE, read_only=True, data_only=False, keep_links=False)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        index = {str(value).strip(): idx for idx, value in enumerate(headers) if value is not None}
        missing = [header for header in TARGET_HEADERS if header not in index]
        if missing:
            raise RuntimeError(f"Target workbook is missing headers: {missing}")
        rows: list[dict[str, Any]] = []
        for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = {header: values[index[header]] for header in TARGET_HEADERS}
            row["_excel_row"] = excel_row
            row["_project_key"] = project_key(row["ProjectCode"])
            row["_location_key"] = location_key(row["LocationNo"])
            rows.append(row)
        return rows, index, ["" if value is None else str(value) for value in headers]
    finally:
        wb.close()


def erection_scope(key: str, path: Path, configured: dict[str, str]) -> str:
    sheet = text_value(configured.get("sheet_name", "")).lower()
    line = text_value(configured.get("line_name", "")).lower()
    if key == "tb501":
        return "220kv" if "220" in (line or sheet) else "132kv"
    if key == "tb507":
        return "main" if "[MAIN]" in path.name.upper() else "other"
    if key == "tb605":
        label = line or sheet
        if "jammu" in label:
            return "jammu"
        if re.search(r"\bsk\b", label) or "-sk" in label:
            return "sk"
        if "punjab" in label:
            return "punjab"
    if key == "ta513":
        if "s-p" in sheet:
            return "sp"
        if "s-f" in sheet:
            return "sf"
    if key == "ta504":
        return "main"
    return "default"


def target_erection_scope(key: str, section: object) -> str | None:
    section_text = text_value(section).lower()
    if key == "tb501":
        return "220kv"
    if key == "tb507":
        return "main"
    if key == "tb605":
        if "ks line" in section_text:
            return "sk"
        if "js line" in section_text and "ap5-ap46" in section_text.replace(" ", ""):
            return "jammu"
        if "js line" in section_text:
            return "punjab"
        return None
    if key == "ta513":
        if "patran" in section_text:
            return "sp"
        if "fatehabad" in section_text:
            return "sf"
        return None
    if key == "ta504":
        return "main" if "main" in section_text else None
    return "default"


def find_column(df: pd.DataFrame, candidates: Iterable[str]) -> str | None:
    normalized = {erection.nrm_header(column): column for column in df.columns}
    for candidate in candidates:
        hit = normalized.get(erection.nrm_header(candidate))
        if hit is not None:
            return hit
    return None


def parse_weight(value: object) -> float | None:
    text = text_value(value)
    if not text:
        return None
    try:
        number = float(str(text).replace(",", ""))
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number) or number <= 0:
        return None
    return number


def extract_erection_records(
    selected_files: dict[str, list[Path]], target_projects: set[str]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    sheet_config = erection.load_erection_sheet_config(RAW_ROOT)
    template_config, template_errors = erection.load_erection_template_mapping_config(RAW_ROOT)
    records: list[dict[str, Any]] = []
    diagnostics: list[dict[str, Any]] = []

    for key in sorted(target_projects):
        files = selected_files.get(key, [])
        configured_entries = sheet_config.get(key)
        if not files or not configured_entries:
            diagnostics.append({"kind": "erection", "project": key, "status": "NO_DPR_OR_CONFIG"})
            continue
        if key in template_errors:
            diagnostics.append(
                {"kind": "erection", "project": key, "status": "TEMPLATE_WARNING", "note": template_errors[key]}
            )
        for path in files:
            if key == "tb507" and "[MAIN]" not in path.name.upper():
                # The location master is the MAIN line; the other TB507 DPR has overlapping numbering.
                continue
            for configured in configured_entries:
                if key == "tb501" and "220" not in (
                    text_value(configured.get("line_name")) + text_value(configured.get("sheet_name"))
                ):
                    continue
                requested = text_value(configured.get("sheet_name"))
                if not requested:
                    continue
                template = erection._select_template_mapping_for_request(
                    template_config.get(key),
                    configured_sheet_name=requested,
                    line_name=configured.get("line_name", ""),
                )
                template_map = template.get("column_map") if template else None
                try:
                    df_raw, resolved, fallback = erection.load_sheet_with_csv_fallback(
                        path,
                        exact_sheet_selector(requested),
                        read_excel_kwargs={"header": None},
                        read_csv_kwargs={"header": None},
                    )
                except Exception as exc:
                    diagnostics.append(
                        {
                            "kind": "erection",
                            "project": key,
                            "file": path.name,
                            "sheet": requested,
                            "status": "READ_FAIL",
                            "note": f"{type(exc).__name__}: {exc}",
                        }
                    )
                    continue
                if df_raw is None or resolved is None or df_raw.empty:
                    diagnostics.append(
                        {
                            "kind": "erection",
                            "project": key,
                            "file": path.name,
                            "sheet": requested,
                            "status": "SHEET_NOT_FOUND_OR_EMPTY",
                        }
                    )
                    continue
                header_row, columns = erection.find_header_row(
                    df_raw, search_rows=30, min_score=1.0 if template_map else 2.0
                )
                if header_row is None or columns is None:
                    diagnostics.append(
                        {
                            "kind": "erection",
                            "project": key,
                            "file": path.name,
                            "sheet": resolved,
                            "status": "HEADER_NOT_FOUND",
                        }
                    )
                    continue
                df = df_raw.iloc[header_row + 1 :].copy()
                df.columns = columns
                if template_map:
                    remapped = list(df.columns)
                    for column_index, mapped_name in sorted(template_map.items()):
                        if column_index < len(remapped) and erection.nrm_header(mapped_name):
                            remapped[column_index] = erection.nrm_header(mapped_name)
                    df.columns = remapped
                df, _duplicates = erection._drop_duplicate_columns_keep_first(df)
                location_col = find_column(df, ["location no", "location number", "loc no"])
                type_col = find_column(df, ["type of tower", "tower type", "type"])
                weight_col = find_column(
                    df, ["tower weight", "total tower weight", "tower weight mt", "weight mt", "mt", "weight"]
                )
                if location_col is None:
                    diagnostics.append(
                        {
                            "kind": "erection",
                            "project": key,
                            "file": path.name,
                            "sheet": resolved,
                            "status": "NO_LOCATION_COLUMN",
                        }
                    )
                    continue
                scope = erection_scope(key, path, configured)
                extracted = 0
                for _, source_row in df.iterrows():
                    loc = location_key(source_row.get(location_col))
                    if not loc:
                        continue
                    tower_type = text_value(source_row.get(type_col)) if type_col else ""
                    weight = parse_weight(source_row.get(weight_col)) if weight_col else None
                    if not tower_type and weight is None:
                        continue
                    records.append(
                        {
                            "project": key,
                            "scope": scope,
                            "location": loc,
                            "tower_type": tower_type,
                            "weight": weight,
                            "file": path.name,
                            "sheet": resolved,
                        }
                    )
                    extracted += 1
                diagnostics.append(
                    {
                        "kind": "erection",
                        "project": key,
                        "file": path.name,
                        "sheet": resolved,
                        "scope": scope,
                        "status": "OK",
                        "rows": extracted,
                        "type_column": type_col or "",
                        "weight_column": weight_col or "",
                        "fallback": fallback or "",
                    }
                )
                print(f"[erection] {key} | {path.name} | {resolved} | extracted={extracted}", flush=True)
    return records, diagnostics


def collapse_erection_records(
    records: list[dict[str, Any]],
) -> tuple[dict[tuple[str, str, str], dict[str, Any]], list[dict[str, Any]]]:
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        grouped[(record["project"], record["scope"], record["location"])].append(record)
    collapsed: dict[tuple[str, str, str], dict[str, Any]] = {}
    conflicts: list[dict[str, Any]] = []
    for key, group in grouped.items():
        type_values: dict[str, str] = {}
        weight_values: dict[float, float] = {}
        for record in group:
            tower_type = text_value(record.get("tower_type"))
            if tower_type:
                type_values.setdefault(re.sub(r"\s+", "", tower_type).upper(), tower_type)
            weight = record.get("weight")
            if weight is not None:
                weight_values.setdefault(round(float(weight), 6), float(weight))
        tower_type = next(iter(type_values.values())) if len(type_values) == 1 else ""
        weight = next(iter(weight_values.values())) if len(weight_values) == 1 else None
        if len(type_values) > 1 or len(weight_values) > 1:
            conflicts.append(
                {
                    "project": key[0],
                    "scope": key[1],
                    "location": key[2],
                    "tower_types": sorted(type_values.values()),
                    "weights": sorted(weight_values.values()),
                }
            )
        collapsed[key] = {"tower_type": tower_type, "weight": weight}
    return collapsed, conflicts


def normalize_method(value: object) -> str:
    text = re.sub(r"[\s_-]+", " ", text_value(value).strip().lower()).strip()
    if text == "tse":
        return "TSE"
    if text in {"manual", "m"}:
        return "Manual"
    if text.replace(" ", "") == "hotline":
        return "Hotline"
    return ""


def stringing_scope(key: str, path: Path, configured: dict[str, str]) -> str:
    sheet = text_value(configured.get("sheet_name", "")).lower()
    line = text_value(configured.get("line_name", "")).lower()
    if key == "tb501":
        return "220kv" if "220" in (line or sheet) else "132kv"
    if key == "tb507":
        return "main" if "[MAIN]" in path.name.upper() else "other"
    if key == "tb605":
        return "jammu"
    if key == "ta504":
        return "main"
    return "default"


def target_stringing_scope(key: str, section: object) -> str | None:
    section_text = text_value(section).lower()
    if key == "tb501":
        return "220kv"
    if key == "tb507":
        return "main"
    if key == "tb605":
        compact = section_text.replace(" ", "")
        return "jammu" if "jsline" in compact and "ap5-ap46" in compact else None
    if key == "ta504":
        return "main" if "main" in section_text else None
    return "default"


def extract_stringing_spans(
    selected_files: dict[str, list[Path]], target_projects: set[str]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    config = stringing_ingest.load_stringing_sheet_config(RAW_ROOT, repo_root=REPO)
    checked_catalog, checked_errors = stringing_ingest.load_stringing_template_mapping_catalog(
        RAW_ROOT, repo_root=REPO
    )
    all_catalog, _ = stringing_ingest.load_stringing_template_mapping_catalog(
        RAW_ROOT, repo_root=REPO, include_unchecked=True
    )
    spans: list[dict[str, Any]] = []
    diagnostics: list[dict[str, Any]] = []

    for key in sorted(target_projects):
        files = selected_files.get(key, [])
        configured_entries = config.get(key)
        if not files or not configured_entries:
            diagnostics.append({"kind": "stringing", "project": key, "status": "NO_DPR_OR_CONFIG"})
            continue
        for path in files:
            if key == "tb507" and "[MAIN]" not in path.name.upper():
                continue
            for configured in configured_entries:
                requested = text_value(configured.get("sheet_name"))
                line = normalize_line_name(configured.get("line_name", ""))
                if key == "tb501" and "220" not in (line + requested):
                    continue
                initial_pair = stringing_ingest.select_template_map_for_sheet(
                    all_catalog.get(key), configured_sheet_name=requested, line_name=line
                )
                min_columns = max(initial_pair[0].keys()) + 1 if initial_pair else None
                try:
                    result = stringing_ingest.load_stringing_sheet_frame(
                        path,
                        configured_sheet_name=requested,
                        min_columns=min_columns,
                        section_start_text=text_value(configured.get("section_start_text")),
                        section_end_text=text_value(configured.get("section_end_text")),
                    )
                except Exception as exc:
                    diagnostics.append(
                        {
                            "kind": "stringing",
                            "project": key,
                            "file": path.name,
                            "sheet": requested,
                            "status": "READ_FAIL",
                            "note": f"{type(exc).__name__}: {exc}",
                        }
                    )
                    continue
                if result.frame is None or result.frame.empty or result.resolved_sheet is None:
                    diagnostics.append(
                        {
                            "kind": "stringing",
                            "project": key,
                            "file": path.name,
                            "sheet": requested,
                            "status": "EMPTY",
                        }
                    )
                    continue
                frame = _sanitize_stringing_columns(result.frame)
                selected_pair = stringing_ingest.select_template_map_for_sheet(
                    checked_catalog.get(key),
                    configured_sheet_name=requested,
                    resolved_sheet_name=result.resolved_sheet,
                    line_name=line,
                )
                fallback_pair = stringing_ingest.select_template_map_for_sheet(
                    all_catalog.get(key),
                    configured_sheet_name=requested,
                    resolved_sheet_name=result.resolved_sheet,
                    line_name=line,
                )
                pair = selected_pair or fallback_pair
                template_map = pair[0] if pair else None
                _source, normalized, _report, _classification, _changes, _applied = _apply_template_if_improves(
                    frame, template_map
                )
                scope = stringing_scope(key, path, configured)
                extracted = 0
                for _, source_row in normalized.iterrows():
                    method = normalize_method(source_row.get("method"))
                    from_location = location_key(source_row.get("from_ap"))
                    to_location = location_key(source_row.get("to_ap"))
                    if not method or not from_location or not to_location or from_location == to_location:
                        continue
                    spans.append(
                        {
                            "project": key,
                            "scope": scope,
                            "from": from_location,
                            "to": to_location,
                            "method": method,
                            "file": path.name,
                            "sheet": result.resolved_sheet,
                        }
                    )
                    extracted += 1
                diagnostics.append(
                    {
                        "kind": "stringing",
                        "project": key,
                        "file": path.name,
                        "sheet": result.resolved_sheet,
                        "scope": scope,
                        "status": "OK",
                        "rows": extracted,
                        "fallback": result.fallback_note,
                        "template_warning": checked_errors.get(key, ""),
                    }
                )
                print(f"[stringing] {key} | {path.name} | {result.resolved_sheet} | explicit_spans={extracted}", flush=True)
    return spans, diagnostics


def map_stringing_methods(
    target_rows: list[dict[str, Any]], spans: list[dict[str, Any]]
) -> tuple[dict[int, str], list[dict[str, Any]], Counter]:
    sequences: dict[tuple[str, str], list[int]] = defaultdict(list)
    for target_index, row in enumerate(target_rows):
        scope = target_stringing_scope(row["_project_key"], row["Section"])
        if scope is not None:
            sequences[(row["_project_key"], scope)].append(target_index)

    spans_by_scope: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for span in spans:
        spans_by_scope[(span["project"], span["scope"])].append(span)

    proposals: dict[int, set[str]] = defaultdict(set)
    unresolved: list[dict[str, Any]] = []
    stats: Counter = Counter()
    for scope_key, scope_spans in spans_by_scope.items():
        sequence = sequences.get(scope_key, [])
        if not sequence:
            stats["no_target_sequence"] += len(scope_spans)
            unresolved.extend({**span, "reason": "NO_TARGET_SEQUENCE"} for span in scope_spans)
            continue
        positions: dict[str, list[int]] = defaultdict(list)
        for position, target_index in enumerate(sequence):
            locations = target_rows[target_index]["_location_key"]
            if locations:
                positions[locations].append(position)
        for span in scope_spans:
            from_positions = positions.get(span["from"], [])
            to_positions = positions.get(span["to"], [])
            if len(from_positions) != 1 or len(to_positions) != 1:
                reason = "MISSING_ENDPOINT" if not from_positions or not to_positions else "DUPLICATE_ENDPOINT"
                stats[reason.lower()] += 1
                unresolved.append({**span, "reason": reason})
                continue
            start = min(from_positions[0], to_positions[0])
            end = max(from_positions[0], to_positions[0])
            if start == end:
                stats["zero_length"] += 1
                unresolved.append({**span, "reason": "ZERO_LENGTH"})
                continue
            for position in range(start, end):
                proposals[sequence[position]].add(span["method"])
            stats["mapped_spans"] += 1

    mapped: dict[int, str] = {}
    for target_index, methods in proposals.items():
        if len(methods) == 1:
            mapped[target_index] = next(iter(methods))
        else:
            stats["conflicting_target_rows"] += 1
    stats["mapped_target_rows"] = len(mapped)
    return mapped, unresolved, stats


def apply_mappings(
    target_rows: list[dict[str, Any]],
    erection_lookup: dict[tuple[str, str, str], dict[str, Any]],
    stringing_map: dict[int, str],
) -> Counter:
    stats: Counter = Counter()
    for target_index, row in enumerate(target_rows):
        scope = target_erection_scope(row["_project_key"], row["Section"])
        attributes = (
            erection_lookup.get((row["_project_key"], scope, row["_location_key"]))
            if scope is not None
            else None
        )
        if attributes:
            if attributes.get("tower_type"):
                row["Tower Type"] = attributes["tower_type"]
                stats["tower_type_filled"] += 1
            if attributes.get("weight") is not None:
                row["Weight"] = attributes["weight"]
                stats["weight_filled"] += 1
        method = stringing_map.get(target_index, "")
        if method:
            row["Stringing Method"] = method
            stats["method_filled"] += 1
    return stats


def copy_cell(source_cell, target_cell) -> None:
    target_cell.value = source_cell.value
    if source_cell.has_style:
        # Copy the public style components so openpyxl registers each component
        # in the destination workbook's own style tables. Copying _style across
        # workbooks leaves foreign alignment/border IDs and produces invalid xlsx.
        target_cell.font = copy.copy(source_cell.font)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format
    if source_cell.hyperlink:
        target_cell._hyperlink = copy.copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy.copy(source_cell.comment)


def write_outputs(target_rows: list[dict[str, Any]]) -> dict[str, int]:
    wb = load_workbook(SOURCE, data_only=False, keep_links=True)
    ws = wb[wb.sheetnames[0]]
    headers = header_map(ws)
    for row in target_rows:
        excel_row = int(row["_excel_row"])
        ws.cell(excel_row, headers["Tower Type"], row.get("Tower Type") or None)
        ws.cell(excel_row, headers["Weight"], row.get("Weight") if row.get("Weight") is not None else None)
        ws.cell(excel_row, headers["Stringing Method"], row.get("Stringing Method") or None)
    wb.save(OUTPUT)

    PROJECT_DIR.mkdir(parents=True, exist_ok=False)
    project_rows: dict[str, list[int]] = defaultdict(list)
    for row in target_rows:
        project_rows[row["_project_key"]].append(int(row["_excel_row"]))

    row_counts: dict[str, int] = {}
    max_column = ws.max_column
    for key in sorted(project_rows):
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = ws.title
        output_wb.properties = copy.copy(wb.properties)
        output_ws.freeze_panes = ws.freeze_panes
        output_ws.sheet_format = copy.copy(ws.sheet_format)
        output_ws.sheet_properties = copy.copy(ws.sheet_properties)
        output_ws.page_margins = copy.copy(ws.page_margins)
        output_ws.page_setup = copy.copy(ws.page_setup)
        output_ws.print_options = copy.copy(ws.print_options)
        output_ws.sheet_view.showGridLines = ws.sheet_view.showGridLines
        for column_letter, dimension in ws.column_dimensions.items():
            output_ws.column_dimensions[column_letter] = copy.copy(dimension)

        output_row = 1
        for column in range(1, max_column + 1):
            copy_cell(ws.cell(1, column), output_ws.cell(output_row, column))
        if ws.row_dimensions[1].height is not None:
            output_ws.row_dimensions[1].height = ws.row_dimensions[1].height

        for source_row in project_rows[key]:
            output_row += 1
            for column in range(1, max_column + 1):
                copy_cell(ws.cell(source_row, column), output_ws.cell(output_row, column))
            if ws.row_dimensions[source_row].height is not None:
                output_ws.row_dimensions[output_row].height = ws.row_dimensions[source_row].height

        if ws.auto_filter.ref:
            start_cell = ws.auto_filter.ref.split(":", 1)[0]
            end_column = re.match(r"[A-Z]+", ws.auto_filter.ref.split(":", 1)[-1]).group(0)
            output_ws.auto_filter.ref = f"{start_cell}:{end_column}{output_row}"
        label = canonical_project_label(key)
        output_path = PROJECT_DIR / f"{label}_Locations.xlsx"
        output_wb.save(output_path)
        output_wb.close()
        row_counts[key] = len(project_rows[key])
        print(f"[export] {output_path.name} | rows={row_counts[key]}", flush=True)
    wb.close()
    return row_counts


def export_projects_from_filled_master() -> dict[str, int]:
    """Create the 27 project workbooks from an already validated filled master."""
    wb = load_workbook(OUTPUT, data_only=False, keep_links=True)
    ws = wb[wb.sheetnames[0]]
    headers = header_map(ws)
    project_rows: dict[str, list[int]] = defaultdict(list)
    for source_row in range(2, ws.max_row + 1):
        key = project_key(ws.cell(source_row, headers["ProjectCode"]).value)
        if key:
            project_rows[key].append(source_row)
    PROJECT_DIR.mkdir(parents=True, exist_ok=False)
    row_counts: dict[str, int] = {}
    max_column = ws.max_column
    for key in sorted(project_rows):
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = ws.title
        output_wb.properties = copy.copy(wb.properties)
        output_ws.freeze_panes = ws.freeze_panes
        output_ws.sheet_format = copy.copy(ws.sheet_format)
        output_ws.sheet_properties = copy.copy(ws.sheet_properties)
        output_ws.page_margins = copy.copy(ws.page_margins)
        output_ws.page_setup = copy.copy(ws.page_setup)
        output_ws.print_options = copy.copy(ws.print_options)
        output_ws.sheet_view.showGridLines = ws.sheet_view.showGridLines
        for column_letter, dimension in ws.column_dimensions.items():
            output_ws.column_dimensions[column_letter] = copy.copy(dimension)
        for column in range(1, max_column + 1):
            copy_cell(ws.cell(1, column), output_ws.cell(1, column))
        if ws.row_dimensions[1].height is not None:
            output_ws.row_dimensions[1].height = ws.row_dimensions[1].height
        output_row = 1
        for source_row in project_rows[key]:
            output_row += 1
            for column in range(1, max_column + 1):
                copy_cell(ws.cell(source_row, column), output_ws.cell(output_row, column))
            if ws.row_dimensions[source_row].height is not None:
                output_ws.row_dimensions[output_row].height = ws.row_dimensions[source_row].height
        if ws.auto_filter.ref:
            start_cell = ws.auto_filter.ref.split(":", 1)[0]
            end_match = re.match(r"[A-Z]+", ws.auto_filter.ref.split(":", 1)[-1])
            end_column = end_match.group(0) if end_match else "M"
            output_ws.auto_filter.ref = f"{start_cell}:{end_column}{output_row}"
        output_path = PROJECT_DIR / f"{canonical_project_label(key)}_Locations.xlsx"
        output_wb.save(output_path)
        output_wb.close()
        row_counts[key] = len(project_rows[key])
        print(f"[export] {output_path.name} | rows={row_counts[key]}", flush=True)
    wb.close()
    return row_counts


def normalized_cell(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def validate_outputs(
    target_rows: list[dict[str, Any]],
    row_counts: dict[str, int],
    input_hashes: dict[str, str],
) -> dict[str, Any]:
    source_wb = load_workbook(SOURCE, read_only=True, data_only=False, keep_links=False)
    output_wb = load_workbook(OUTPUT, read_only=True, data_only=False, keep_links=False)
    try:
        source_ws = source_wb[source_wb.sheetnames[0]]
        output_ws = output_wb[output_wb.sheetnames[0]]
        source_headers = header_map(source_ws)
        output_headers = header_map(output_ws)
        if source_ws.max_row != 9510 or output_ws.max_row != 9510:
            raise AssertionError(f"Expected 9510 rows including header, got source={source_ws.max_row}, output={output_ws.max_row}")
        for row_number in range(2, source_ws.max_row + 1):
            for header in IMMUTABLE_HEADERS:
                source_value = normalized_cell(source_ws.cell(row_number, source_headers[header]).value)
                output_value = normalized_cell(output_ws.cell(row_number, output_headers[header]).value)
                if source_value != output_value:
                    raise AssertionError(f"Immutable value changed at row {row_number}, column {header}")
        accepted_methods = {None, "", "TSE", "Manual", "Hotline"}
        for row_number in range(2, output_ws.max_row + 1):
            method = output_ws.cell(row_number, output_headers["Stringing Method"]).value
            if method not in accepted_methods:
                raise AssertionError(f"Unexpected method at row {row_number}: {method!r}")

        def find_row(project: str, location: str) -> int:
            wanted_project = project_key(project)
            wanted_location = location_key(location)
            for row_number in range(2, output_ws.max_row + 1):
                if (
                    project_key(output_ws.cell(row_number, output_headers["ProjectCode"]).value) == wanted_project
                    and location_key(output_ws.cell(row_number, output_headers["LocationNo"]).value) == wanted_location
                ):
                    return row_number
            raise AssertionError(f"Representative row not found: {project} {location}")

        representative = {}
        for project, location in [("TA416", "2/8"), ("TA414", "118/0"), ("TA414", "121/0"), ("TA414", "121/7")]:
            row_number = find_row(project, location)
            representative[f"{project}:{location}"] = {
                "tower_type": output_ws.cell(row_number, output_headers["Tower Type"]).value,
                "weight": output_ws.cell(row_number, output_headers["Weight"]).value,
                "method": output_ws.cell(row_number, output_headers["Stringing Method"]).value,
            }
        if representative["TA414:118/0"]["method"] != "Manual":
            raise AssertionError(f"TA414 118/0 method mismatch: {representative['TA414:118/0']}")
        if representative["TA414:121/0"]["method"] != "Hotline":
            raise AssertionError(f"TA414 121/0 method mismatch: {representative['TA414:121/0']}")
        if representative["TA414:121/7"]["method"] not in {None, ""}:
            raise AssertionError(f"TA414 121/7 mixed method was not blank: {representative['TA414:121/7']}")

        missing_projects = {"ta609", "ta514", "tb610", "tb611"}
        for row in target_rows:
            if row["_project_key"] in missing_projects:
                if text_value(row.get("Tower Type")) or row.get("Weight") is not None or text_value(row.get("Stringing Method")):
                    raise AssertionError(f"Missing-DPR project unexpectedly populated: {row['_project_key']}")
    finally:
        source_wb.close()
        output_wb.close()

    if len(row_counts) != 27 or sum(row_counts.values()) != 9509:
        raise AssertionError(f"Project exports invalid: files={len(row_counts)}, rows={sum(row_counts.values())}")
    actual_files = list(PROJECT_DIR.glob("*.xlsx"))
    if len(actual_files) != 27:
        raise AssertionError(f"Expected 27 project workbooks, found {len(actual_files)}")
    exported_rows = 0
    for project_file in actual_files:
        project_wb = load_workbook(project_file, read_only=True, data_only=False, keep_links=False)
        try:
            project_ws = project_wb[project_wb.sheetnames[0]]
            project_headers = header_map(project_ws)
            for required in TARGET_HEADERS:
                if required not in project_headers:
                    raise AssertionError(f"{project_file.name} is missing {required}")
            expected_key = project_key(project_file.stem.replace("_Locations", ""))
            for row_number in range(2, project_ws.max_row + 1):
                actual_key = project_key(project_ws.cell(row_number, project_headers["ProjectCode"]).value)
                if actual_key != expected_key:
                    raise AssertionError(
                        f"{project_file.name} contains project {actual_key!r} at row {row_number}"
                    )
                exported_rows += 1
        finally:
            project_wb.close()
    if exported_rows != 9509:
        raise AssertionError(f"Project workbook row total is {exported_rows}, expected 9509")

    after_hashes = {str(path): sha256(path) for path in [SOURCE, CONFIG, *sorted(RAW_ROOT.glob("*.xls*"))]}
    changed_sources = [path for path, before in input_hashes.items() if after_hashes.get(path) != before]
    if changed_sources:
        raise AssertionError(f"Source files changed during run: {changed_sources}")

    return {
        "master_rows": 9509,
        "project_files": len(row_counts),
        "project_rows": sum(row_counts.values()),
        "source_files_unchanged": True,
        "representative": representative,
    }


def main() -> None:
    if "--export-only" in sys.argv:
        if not OUTPUT.exists():
            raise FileNotFoundError(OUTPUT)
        if PROJECT_DIR.exists():
            raise FileExistsError(PROJECT_DIR)
        input_paths = [SOURCE, CONFIG, *sorted(RAW_ROOT.glob("*.xls*"))]
        input_hashes = {str(path): sha256(path) for path in input_paths}
        target_rows, _target_index, _headers = load_target_rows()
        row_counts = export_projects_from_filled_master()
        validation = validate_outputs(target_rows, row_counts, input_hashes)
        output_wb = load_workbook(OUTPUT, read_only=True, data_only=False, keep_links=False)
        try:
            output_ws = output_wb[output_wb.sheetnames[0]]
            columns = header_map(output_ws)
            fill_counts = {
                "tower_type_filled": sum(
                    bool(text_value(output_ws.cell(row, columns["Tower Type"]).value))
                    for row in range(2, output_ws.max_row + 1)
                ),
                "weight_filled": sum(
                    output_ws.cell(row, columns["Weight"]).value is not None
                    for row in range(2, output_ws.max_row + 1)
                ),
                "method_filled": sum(
                    bool(text_value(output_ws.cell(row, columns["Stringing Method"]).value))
                    for row in range(2, output_ws.max_row + 1)
                ),
            }
        finally:
            output_wb.close()
        print(
            "FINAL_SUMMARY="
            + json.dumps(
                {
                    "output": str(OUTPUT),
                    "project_directory": str(PROJECT_DIR),
                    "fill_stats": fill_counts,
                    "validation": validation,
                },
                indent=2,
                default=str,
            ),
            flush=True,
        )
        return
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)
    if OUTPUT.exists() or PROJECT_DIR.exists():
        raise FileExistsError("One or more output paths already exist; refusing to merge with a prior run.")

    input_paths = [SOURCE, CONFIG, *sorted(RAW_ROOT.glob("*.xls*"))]
    input_hashes = {str(path): sha256(path) for path in input_paths}
    target_rows, _target_index, _headers = load_target_rows()
    if len(target_rows) != 9509:
        raise AssertionError(f"Expected 9509 target rows, found {len(target_rows)}")
    target_projects = {row["_project_key"] for row in target_rows}
    if len(target_projects) != 27:
        raise AssertionError(f"Expected 27 target projects, found {len(target_projects)}")

    selected_files = latest_dpr_files()
    print("[snapshot] " + json.dumps({key: [path.name for path in value] for key, value in sorted(selected_files.items())}), flush=True)

    erection_records, erection_diagnostics = extract_erection_records(selected_files, target_projects)
    erection_lookup, erection_conflicts = collapse_erection_records(erection_records)
    stringing_spans, stringing_diagnostics = extract_stringing_spans(selected_files, target_projects)
    stringing_map, stringing_unresolved, stringing_stats = map_stringing_methods(target_rows, stringing_spans)
    fill_stats = apply_mappings(target_rows, erection_lookup, stringing_map)

    print(f"[mapping] fill_stats={dict(fill_stats)}", flush=True)
    print(f"[mapping] erection_conflicts={len(erection_conflicts)}", flush=True)
    print(f"[mapping] stringing_stats={dict(stringing_stats)} unresolved={len(stringing_unresolved)}", flush=True)

    row_counts = write_outputs(target_rows)
    validation = validate_outputs(target_rows, row_counts, input_hashes)

    summary = {
        "output": str(OUTPUT),
        "project_directory": str(PROJECT_DIR),
        "fill_stats": dict(fill_stats),
        "erection_source_records": len(erection_records),
        "erection_conflicts": len(erection_conflicts),
        "stringing_explicit_spans": len(stringing_spans),
        "stringing_stats": dict(stringing_stats),
        "stringing_unresolved_spans": len(stringing_unresolved),
        "erection_diagnostics": Counter(row.get("status", "") for row in erection_diagnostics),
        "stringing_diagnostics": Counter(row.get("status", "") for row in stringing_diagnostics),
        "validation": validation,
    }
    # Counter is not directly JSON serializable.
    summary["erection_diagnostics"] = dict(summary["erection_diagnostics"])
    summary["stringing_diagnostics"] = dict(summary["stringing_diagnostics"])
    print("FINAL_SUMMARY=" + json.dumps(summary, indent=2, default=str), flush=True)


if __name__ == "__main__":
    main()
