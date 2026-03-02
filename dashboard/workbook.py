"""Excel workbook export helpers."""
from __future__ import annotations

import logging
import re
from io import BytesIO
from pathlib import Path
from math import ceil
from typing import Sequence, TYPE_CHECKING

import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import AppConfig
from .metrics import (
    calc_idle_and_loss,
    compute_gang_baseline_maps,
    compute_idle_intervals_per_gang,
    compute_project_baseline_maps_for,
)
from .pch_normalizer import CANONICAL_PCH_PRIMARY, normalize_pch

if TYPE_CHECKING:  # pragma: no cover - type checking only
    from .state import AppDataStore

LOGGER = logging.getLogger(__name__)

AVG_PRODUCTIVITY_COLUMN = "Avg Productivity (MTD)"
AVG_PRODUCTIVITY_OVERALL_COLUMN = "Avg Productivity (Overall)"
TOTAL_MT_COLUMN = "Total MT (MTD)"
TOTAL_COUNT_COLUMN = "Total No. of Erections (MTD)"
TOWER_BUCKETS = ("DA", "DB", "DC", "DD")
EXCEPTION_BUCKET = "Exception"
AVG_TOWER_WEIGHT_OVERALL_COLUMN = "Avg Tower Weight Overall"
TOWER_BUCKET_COUNT_COLUMNS = [*TOWER_BUCKETS, EXCEPTION_BUCKET]
TOWER_BUCKET_AVG_COLUMNS = [f"{bucket} Avg Tower Weight" for bucket in TOWER_BUCKETS]
TOWER_BUCKET_AVG_COLUMNS.append(f"{EXCEPTION_BUCKET} Avg Tower Weight")
TOWER_WEIGHT_COLUMNS = [AVG_TOWER_WEIGHT_OVERALL_COLUMN]
IDLE_DAYS_COLUMN = "Idle Counted (days)"
IDLE_INTERVALS_COLUMN = "Number of Idle Intervals"
YTD_METRIC_LABELS = {
    AVG_PRODUCTIVITY_COLUMN: "Avg Productivity (YTD)",
    TOTAL_MT_COLUMN: "Total MT (YTD)",
    TOTAL_COUNT_COLUMN: "Total No. of Erections (YTD)",
}
SUMMARY_METRIC_LABELS = {
    AVG_PRODUCTIVITY_COLUMN: "Avg Productivity",
    TOTAL_MT_COLUMN: "Total MT",
    TOTAL_COUNT_COLUMN: "Total No. of Erections",
}
MONTHLY_AVG_LABEL = "Avg Productivity"
MONTHLY_COUNT_LABEL = "No. of Erections"
MONTHLY_MT_LABEL = "MT Erected"
_DEFAULT_SHEET_NAME = "Erection Summary"
_DEFAULT_KV_SHEET_NAME = "KV Productivity"
_PCH_SORT_ORDER = {name: idx for idx, name in enumerate(CANONICAL_PCH_PRIMARY)}

# TODO: Align the week bucket helper below with the dashboard's official week mapping
# once that logic is exposed outside the callbacks module.


def _generate_week_labels(month_start: pd.Timestamp, month_end: pd.Timestamp) -> list[str]:
    """Generate sequential week labels covering the provided date window."""
    days = max(1, int((month_end - month_start).days) + 1)
    week_count = max(1, ceil(days / 7))
    return [f"Week {idx}" for idx in range(1, week_count + 1)]


def _generate_quarter_week_labels(month_start: pd.Timestamp, month_end: pd.Timestamp) -> list[str]:
    """Generate week labels for a quarter split into 4 weeks per month."""
    if month_start > month_end:
        return []
    start_period = month_start.to_period("M")
    end_period = month_end.to_period("M")
    months = (end_period.year - start_period.year) * 12 + (end_period.month - start_period.month) + 1
    total_weeks = max(1, months * 4)
    return [f"Week {idx}" for idx in range(1, total_weeks + 1)]


_DPR_FILE_SUFFIXES = (".xlsx", ".xlsm", ".xlsb", ".xls")
_DPR_HEADER_REQUIRED_MARKERS = {
    "sl_no",
    "location_no",
    "tower_type",
    "start_date",
    "tower_weight",
    "completion_date",
    "gang_name",
    "manpower",
}
_DPR_HEADER_OPTIONAL_MARKERS = {"status"}
_DPR_HEADER_MARKERS = _DPR_HEADER_REQUIRED_MARKERS | _DPR_HEADER_OPTIONAL_MARKERS
_DPR_COLUMN_ALIASES = {
    "sl. no.": "sl_no",
    "sl no": "sl_no",
    "sl": "sl_no",
    "location no.": "location_no",
    "location number": "location_no",
    "loc no": "location_no",
    "location": "location_no",
    "type of tower": "tower_type",
    "tower type": "tower_type",
    "starting date": "start_date",
    "start date": "start_date",
    "tower weight": "tower_weight",
    "tower wt": "tower_weight",
    "tower weight (mt)": "tower_weight",
    "tower weight in mt": "tower_weight",
    "status": "status",
    "completion date": "completion_date",
    "complete date": "completion_date",
    "gang name": "gang_name",
    "gang": "gang_name",
    "manpower": "manpower",
    "man power": "manpower",
    "hindarnce days": "hindrance_days",
    "hindrance days": "hindrance_days",
    "remarks": "remarks",
    "tackweldng status": "tackwelding_status",
    "tackwelding start date": "tackwelding_start_date",
    "tackwelding completion date": "tackwelding_completion_date",
    "sub contractor name": "sub_contractor_name",
}
_COMPILED_COLUMN_ALIASES = {
    "work date": "work_date",
    "work_date": "work_date",
    "start date": "start_date",
    "starting date": "start_date",
    "complete date": "completion_date",
    "completion date": "completion_date",
    "completion": "completion_date",
    "gang name": "gang_name",
    "gang": "gang_name",
    "tower weight": "tower_weight",
    "tower weight (mt)": "tower_weight",
    "tower weight(mt)": "tower_weight",
    "tower type": "tower_type",
    "productivity": "productivity",
    "daily prod": "productivity",
    "project name": "project_name",
    "project": "project_name",
    "project code": "project_code",
    "project_code": "project_code",
    "location no.": "location_no",
    "location number": "location_no",
    "location no": "location_no",
    "location": "location_no",
    "status": "status",
}


def _normalize_dpr_label(value: object) -> str:
    text = "" if value is None else str(value)
    normalized = re.sub(r"\s+", " ", text).strip().lower()
    return normalized


def _normalize_tower_bucket(value: object) -> str:
    text = "" if value is None else str(value).lower()
    compact = re.sub(r"[^a-z]+", "", text)
    if not compact:
        return EXCEPTION_BUCKET
    for bucket in TOWER_BUCKETS:
        if bucket.lower() in compact:
            return bucket
    return EXCEPTION_BUCKET


def _tokenize_project_code(value: str) -> list[str]:
    text = re.sub(r"[^a-zA-Z0-9]+", " ", str(value or "").strip()).lower()
    if not text:
        return []
    base_tokens = [token for token in text.split() if token]
    split_tokens: list[str] = []
    for token in base_tokens:
        split_tokens.extend(re.findall(r"[a-z]+|\d+", token))
    ordered: list[str] = []
    for token in [*base_tokens, *split_tokens]:
        if token and token not in ordered:
            ordered.append(token)
    return ordered


def _pick_dpr_sheet(sheet_names: Sequence[str]) -> str | None:
    for name in sheet_names:
        lowered = _normalize_dpr_label(name)
        if "erection" in lowered and "compil" in lowered:
            return name
    for name in sheet_names:
        lowered = _normalize_dpr_label(name)
        if "erection" in lowered:
            return name
    return sheet_names[0] if sheet_names else None


def _find_dpr_header_row(df: pd.DataFrame) -> int | None:
    for idx, row in df.iterrows():
        normalized = {_normalize_dpr_label(value) for value in row if isinstance(value, str)}
        canonical = {_DPR_COLUMN_ALIASES.get(value, value) for value in normalized}
        if _DPR_HEADER_REQUIRED_MARKERS.issubset(canonical):
            return idx
    return None


def _standardize_dpr_columns(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.copy()
    used_names: list[str] = []
    renamed: list[str] = []
    for column in cleaned.columns:
        normalized = _normalize_dpr_label(column)
        canonical = _DPR_COLUMN_ALIASES.get(normalized, normalized)
        candidate = canonical or column
        if candidate in used_names:
            suffix = 2
            while f"{candidate}_{suffix}" in used_names:
                suffix += 1
            candidate = f"{candidate}_{suffix}"
        used_names.append(candidate)
        renamed.append(candidate)
    cleaned.columns = renamed
    return cleaned


def _standardize_compiled_columns(df: pd.DataFrame) -> pd.DataFrame:
    renamed: dict[str, str] = {}
    for column in df.columns:
        normalized = _normalize_dpr_label(column)
        canonical = _COMPILED_COLUMN_ALIASES.get(normalized)
        if canonical:
            renamed[column] = canonical
    if not renamed:
        return df
    return df.rename(columns=renamed)


def _load_dpr_table(path: Path) -> pd.DataFrame:
    try:
        workbook = pd.ExcelFile(path)
    except Exception as exc:
        LOGGER.warning("Unable to open DPR workbook '%s': %s", path, exc)
        return pd.DataFrame()

    sheet_name = _pick_dpr_sheet(workbook.sheet_names)
    if sheet_name is None:
        LOGGER.warning("DPR workbook '%s' has no sheet resembling 'Erection Compiled'; skipping.", path)
        return pd.DataFrame()

    raw = workbook.parse(sheet_name, header=None)
    header_idx = _find_dpr_header_row(raw)
    if header_idx is None:
        LOGGER.warning("DPR workbook '%s' did not expose the expected header row; skipping.", path)
        return pd.DataFrame()

    header_values = raw.iloc[header_idx].tolist()
    data = raw.iloc[header_idx + 1 :].copy()
    data.columns = header_values
    data = data.dropna(how="all")
    if data.empty:
        return pd.DataFrame()
    data = _standardize_dpr_columns(data)
    data["source_file"] = path.name
    return data


def _load_compiled_completion_rows(path: Path | None, project_code: str) -> pd.DataFrame:
    if path is None:
        return pd.DataFrame()

    resolved = Path(path).expanduser()
    if not resolved.exists():
        LOGGER.info("Compiled erection dataset '%s' not found; skipping.", resolved)
        return pd.DataFrame()

    try:
        suffix = resolved.suffix.lower()
        if suffix in {".parquet", ".pq"}:
            daily_df = pd.read_parquet(resolved)
        elif suffix in {".xlsx", ".xlsm", ".xlsb", ".xls"}:
            daily_df = pd.read_excel(resolved)
        elif suffix in {".csv", ".txt"}:
            daily_df = pd.read_csv(resolved)
        else:
            LOGGER.warning("Unsupported compiled dataset format '%s'; provide parquet or Excel.", resolved)
            return pd.DataFrame()
    except Exception as exc:
        LOGGER.warning("Unable to load compiled erection dataset '%s': %s", resolved, exc)
        return pd.DataFrame()

    if daily_df.empty:
        return pd.DataFrame()
    daily_df = _standardize_compiled_columns(daily_df)

    def _ensure_string_column(frame: pd.DataFrame, column: str) -> None:
        series = frame.get(column)
        if isinstance(series, pd.Series):
            frame[column] = series.fillna("").astype(str).str.strip()
        else:
            frame[column] = ""

    _ensure_string_column(daily_df, "project_name")
    if "project_code" in daily_df.columns:
        _ensure_string_column(daily_df, "project_code")
    else:
        daily_df["project_code"] = daily_df["project_name"]

    target_key = _compact_project_key(project_code)
    if not target_key:
        return pd.DataFrame()

    project_code_keys = daily_df["project_code"].map(_compact_project_key)
    project_name_keys = daily_df["project_name"].map(_compact_project_key)
    scoped = daily_df[project_code_keys.eq(target_key) | project_name_keys.eq(target_key)].copy()
    if scoped.empty:
        return pd.DataFrame()

    scoped["work_date"] = pd.to_datetime(scoped.get("work_date"), errors="coerce").dt.normalize()
    scoped["start_date"] = pd.to_datetime(scoped.get("start_date"), errors="coerce").dt.normalize()
    scoped["completion_date"] = pd.to_datetime(scoped.get("completion_date"), errors="coerce").dt.normalize()
    scoped["tower_weight"] = _coerce_numeric_series(scoped, "tower_weight")
    scoped["productivity"] = _coerce_numeric_series(scoped, "productivity")
    for column in ("location_no", "tower_type", "gang_name", "status"):
        _ensure_string_column(scoped, column)

    completed = scoped.dropna(subset=["completion_date"]).copy()
    if completed.empty:
        return completed

    order_cols = [col for col in ("completion_date", "location_no", "work_date") if col in completed.columns]
    if order_cols:
        completed = completed.sort_values(order_cols)
    dedup_columns = [col for col in ("project_code", "location_no", "completion_date") if col in completed.columns]
    if dedup_columns:
        completed = completed.drop_duplicates(subset=dedup_columns, keep="last")

    durations = (completed["completion_date"] - completed["start_date"]).dt.days + 1
    durations = durations.where(durations > 0)
    derived = completed["tower_weight"] / durations
    productivity_series = completed["productivity"]
    completed["derived_productivity"] = derived.fillna(productivity_series)
    completed["manpower"] = pd.Series(pd.NA, index=completed.index, dtype="Float64")
    completed["source_file"] = resolved.name

    columns = [
        "project_code",
        "project_name",
        "completion_date",
        "start_date",
        "location_no",
        "tower_type",
        "tower_weight",
        "status",
        "gang_name",
        "manpower",
        "derived_productivity",
        "source_file",
    ]
    present_columns = [column for column in columns if column in completed.columns]
    return completed[present_columns].copy()


def _coerce_numeric_series(frame: pd.DataFrame, column: str, dtype: str = "float64") -> pd.Series:
    """Return a numeric series for *column* even if it is missing."""

    series = frame.get(column)
    if isinstance(series, pd.Series):
        return pd.to_numeric(series, errors="coerce")
    return pd.Series(index=frame.index, dtype=dtype)


def _resolve_dpr_paths(
    project_code: str,
    explicit_paths: Sequence[str | Path] | None,
    dpr_folder: str | Path | None,
) -> list[Path]:
    resolved: list[Path] = []
    if explicit_paths:
        for candidate in explicit_paths:
            candidate_path = Path(candidate).expanduser()
            if candidate_path.exists():
                resolved.append(candidate_path.resolve())
            else:
                LOGGER.warning("Explicit DPR path '%s' does not exist; skipping.", candidate)
    if resolved:
        return resolved

    search_root = Path(dpr_folder) if dpr_folder else Path("Raw Data") / "DPRs"
    if not search_root.exists():
        LOGGER.warning("DPR folder '%s' not found; provide explicit DPR paths.", search_root)
        return []

    tokens = _tokenize_project_code(project_code)
    compact_code = re.sub(r"[^a-z0-9]+", "", _normalize_dpr_label(project_code))
    if not tokens:
        tokens = [_normalize_dpr_label(project_code)]
    matches: list[Path] = []
    for candidate in search_root.rglob("*"):
        if not candidate.is_file():
            continue
        if candidate.suffix.lower() not in _DPR_FILE_SUFFIXES:
            continue
        lowered_name = candidate.name.lower()
        compact_name = re.sub(r"[^a-z0-9]+", "", lowered_name)
        token_match = all(token in lowered_name for token in tokens if token)
        compact_match = bool(compact_code and compact_code in compact_name)
        if token_match or compact_match:
            matches.append(candidate.resolve())
    matches.sort()
    return matches


def _sanitize_sheet_name(value: str) -> str:
    """Return a value safe for use as an Excel sheet name."""

    sanitized = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(value))
    return sanitized[:31]


def make_trace_workbook_bytes(
    scoped_data: pd.DataFrame,
    months_ts: Sequence[pd.Timestamp] | None,
    projects: Sequence[str] | None,
    gangs: Sequence[str] | None,
    bench: float,
    *,
    gang_for_sheet: str | None = None,
    config: AppConfig | None = None,
    project_info: pd.DataFrame | None = None,
    erections_completed: pd.DataFrame | None = None,
    erections_context: dict[str, object] | None = None,
) -> bytes:
    """Build the Excel export with summary, idle intervals, and daily detail."""

    active_config = config or AppConfig()
    LOGGER.info("Building trace workbook (rows=%d)", len(scoped_data))

    overall_baseline_map, monthly_baseline_map = compute_gang_baseline_maps(scoped_data)

    summary_rows: list[dict[str, object]] = []
    for gang_name, gang_df in scoped_data.groupby("gang_name"):
        overall_baseline = overall_baseline_map.get(gang_name)
        monthly_baseline = monthly_baseline_map.get(gang_name)
        idle, baseline, loss_mt, delivered, potential = calc_idle_and_loss(
            gang_df,
            loss_max_gap_days=active_config.loss_max_gap_days,
            baseline_mt_per_day=overall_baseline,
            baseline_by_month=monthly_baseline,
        )
        summary_rows.append(
            {
                "gang_name": gang_name,
                "delivered_mt": delivered,
                "lost_mt": loss_mt,
                "potential_mt": potential,
                "baseline_mt_per_day": baseline,
                "idle_days_capped": idle,
                "first_date": gang_df["date"].min(),
                "last_date": gang_df["date"].max(),
                "active_days": gang_df["date"].nunique(),
            }
        )
    per_gang_summary = (
        pd.DataFrame(summary_rows).sort_values("potential_mt", ascending=False)
        if summary_rows
        else pd.DataFrame(
            columns=[
                "gang_name",
                "delivered_mt",
                "lost_mt",
                "potential_mt",
                "baseline_mt_per_day",
                "idle_days_capped",
                "first_date",
                "last_date",
                "active_days",
            ]
        )
    )

    idle_df = compute_idle_intervals_per_gang(
        scoped_data,
        loss_max_gap_days=active_config.loss_max_gap_days,
        baseline_month_lookup=monthly_baseline_map,
        baseline_fallback_map=overall_baseline_map,
    )
    daily_df = (
        scoped_data.sort_values(["gang_name", "date"])
        [["date", "gang_name", "project_name", "daily_prod_mt"]]
        .copy()
    )
    project_month = (
        scoped_data.groupby(["project_name", "month"])["daily_prod_mt"].mean().reset_index()
    )

    context_row = {
        "projects": ", ".join(projects or []) or "(all)",
        "gangs": ", ".join(gangs or []) or "(all)",
        "months": ", ".join(
            [timestamp.strftime("%Y-%m") for timestamp in (months_ts or [])]
        )
        or "(all / overall)",
        "benchmark": bench,
        "loss_cap_days": active_config.loss_max_gap_days,
    }
    if erections_context:
        def _format_context_date(value: object) -> str:
            if isinstance(value, pd.Timestamp):
                return value.strftime("%Y-%m-%d")
            if value:
                return str(value)
            return ""
        context_row["erections_range_start"] = _format_context_date(
            erections_context.get("range_start")
        )
        context_row["erections_range_end"] = _format_context_date(
            erections_context.get("range_end")
        )
        search_value = erections_context.get("search_text")
        context_row["erections_search"] = (search_value or "").strip()
    context_df = pd.DataFrame([context_row])

    assumptions = pd.DataFrame(
        {
            "Notes": [
                f"Loss cap per gap: {active_config.loss_max_gap_days} days.",
                "Efficiency = delivered / (delivered + lost). Lost = baseline * capped idle days.",
                "Idle interval = gaps between observed work dates; dates inferred from current filtered scope.",
                "All numbers reflect current dashboard filters (project, period, and gang if applied).",
            ]
        }
    )

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        per_gang_summary.to_excel(writer, "PerGangSummary", index=False)
        idle_df.to_excel(writer, "IdleIntervals", index=False)
        daily_df.to_excel(writer, "DailyProductivity", index=False)
        project_month.to_excel(writer, "ProjectsMonthly", index=False)
        context_df.to_excel(writer, "SelectionContext", index=False)
        assumptions.to_excel(writer, "Assumptions", index=False)

        if erections_completed is not None:
            completions_sheet = erections_completed.copy()
            if "completion_date" in completions_sheet.columns:
                completions_sheet["completion_date"] = completions_sheet["completion_date"].apply(
                    lambda value: value.date() if isinstance(value, pd.Timestamp) else value
                )
            if "start_date" in completions_sheet.columns:
                completions_sheet["start_date"] = completions_sheet["start_date"].apply(
                    lambda value: value.date() if isinstance(value, pd.Timestamp) else (None if pd.isna(value) else value)
                )
            if "supervisor_name" in completions_sheet.columns:
                completions_sheet["supervisor_name"] = completions_sheet["supervisor_name"].fillna("")
            else:
                completions_sheet["supervisor_name"] = ""
            if "section_incharge_name" in completions_sheet.columns:
                completions_sheet["section_incharge_name"] = completions_sheet["section_incharge_name"].fillna("")
            else:
                completions_sheet["section_incharge_name"] = ""
            completions_sheet = completions_sheet.rename(
                columns={
                    "completion_date": "Completion Date",
                    "project_name": "Project",
                    "location_no": "Location",
                    "tower_weight_mt": "Tower Weight (MT)",
                    "daily_prod_mt": "Productivity (MT/day)",
                    "gang_name": "Gang",
                    "start_date": "Start Date",
                    "supervisor_name": "Supervisor",
                    "section_incharge_name": "Section Incharge",
                    "revenue_value": "Revenue",
                }
            )
            completions_sheet.to_excel(writer, "ErectionsCompleted", index=False)

        if gang_for_sheet:
            selected = scoped_data[scoped_data["gang_name"] == gang_for_sheet]
            if not selected.empty:
                single_idle = compute_idle_intervals_per_gang(
                    selected,
                    loss_max_gap_days=active_config.loss_max_gap_days,
                    baseline_month_lookup=monthly_baseline_map,
                    baseline_fallback_map=overall_baseline_map,
                )
                single_idle.to_excel(
                    writer,
                    _sanitize_sheet_name(f"Idle_{gang_for_sheet}"),
                    index=False,
                )
                (
                    selected.sort_values("date")[["date", "project_name", "daily_prod_mt"]]
                    .assign(date=lambda frame: frame["date"].dt.strftime("%Y-%m-%d"))
                    .to_excel(writer, _sanitize_sheet_name(f"Daily_{gang_for_sheet}"), index=False)
                )

                idle, baseline, loss_mt, delivered, potential = calc_idle_and_loss(
                    selected,
                    loss_max_gap_days=active_config.loss_max_gap_days,
                    baseline_mt_per_day=overall_baseline_map.get(gang_for_sheet),
                    baseline_by_month=monthly_baseline_map.get(gang_for_sheet),
                )
                efficiency = (delivered / potential * 100) if potential > 0 else 0.0
                pd.DataFrame(
                    [
                        {
                            "gang_name": gang_for_sheet,
                            "delivered_mt": delivered,
                            "lost_mt": loss_mt,
                            "potential_mt": potential,
                            "efficiency_%": efficiency,
                            "baseline_mt_per_day": baseline,
                            "idle_days_capped": idle,
                        }
                    ]
                ).to_excel(
                    writer,
                    _sanitize_sheet_name(f"Summary_{gang_for_sheet}"),
                    index=False,
                )
        
        
        # Optional: include ProjectDetails if exactly one project is selected
        if projects and len(projects) == 1 and project_info is not None and not project_info.empty:
            pname = str(projects[0]).strip()
            # project_name -> project_code from scoped_data (or fall back to global if needed)
            name_to_code = (
                scoped_data.dropna(subset=["project_name", "project_code"])
                           .drop_duplicates(subset=["project_name"])
                           .set_index("project_name")["project_code"]
                           .to_dict()
            )
            pcode = name_to_code.get(pname)
            if pcode:
                row = project_info[project_info["project_code"] == pcode]
                if not row.empty:
                    r = row.iloc[0]
                    pd.DataFrame([{
                        "Project Code": pcode,
                        "Project Name": pname,
                        "Client Name": r.get("client_name"),
                        "NOA Start Date": r.get("noa_start"),
                        "LOA End Date": r.get("loa_end"),
                        "Project Manager": r.get("project_mgr"),
                        "Regional Manager": r.get("regional_mgr"),
                        "Planning Engineer": r.get("planning_eng"),
                        "PCH": r.get("pch"),
                        "Section Incharge": r.get("section_inch"),
                        "Supervisor": r.get("supervisor"),
                    }]).to_excel(writer, "ProjectDetails_Selected", index=False)


    buffer.seek(0)
    return buffer.getvalue()



def _project_lookup_series(project_info: pd.DataFrame | None, column: str) -> pd.Series:
    if project_info is None or project_info.empty:
        return pd.Series(dtype="object")
    if "key_name" not in project_info.columns or column not in project_info.columns:
        return pd.Series(dtype="object")
    lookup = (
        project_info[["key_name", column]]
        .dropna(subset=["key_name"])
        .drop_duplicates("key_name")
        .set_index("key_name")[column]
    )
    return lookup


def _compact_project_key(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def _prepare_project_voltage_lookup(project_kv: pd.DataFrame | None) -> dict[str, str]:
    if project_kv is None or project_kv.empty:
        return {}

    working = project_kv.copy()

    def _pick_col(options: Sequence[str]) -> str:
        lowered = {str(col).strip().lower(): col for col in working.columns}
        for option in options:
            normalized = option.strip().lower()
            if normalized in lowered:
                return lowered[normalized]
        for option in options:
            normalized = option.strip().lower()
            for candidate, original in lowered.items():
                if normalized and normalized in candidate:
                    return original
        raise KeyError

    try:
        project_col = _pick_col(["project", "project_code", "project name"])
        voltage_col = _pick_col(["voltage", "kv", "kv level"])
    except KeyError:
        LOGGER.warning("Projects_KV workbook missing required columns; voltage lookup disabled.")
        return {}

    lookup: dict[str, str] = {}
    for _, row in working.iterrows():
        project_value = row.get(project_col)
        voltage_value = row.get(voltage_col)
        key = _compact_project_key(project_value)
        if not key:
            continue
        if voltage_value is None or (isinstance(voltage_value, float) and pd.isna(voltage_value)):
            continue
        voltage_text = str(voltage_value).strip()
        if not voltage_text:
            continue
        lookup[key] = voltage_text
    return lookup


def _annotate_scope_with_voltage(scope: pd.DataFrame, voltage_lookup: dict[str, str]) -> pd.DataFrame:
    if scope is None or scope.empty:
        return pd.DataFrame()

    working = scope.copy()
    tower_series = working.get("tower_type")
    if tower_series is None:
        working["tower_type"] = "Unknown"
    else:
        normalized = tower_series.fillna(" ").astype(str).str.strip().str.upper()
        working["tower_type"] = normalized.where(normalized.astype(bool), "Unknown")
    family = working["tower_type"].str.extract(r"^(DA|DB|DC|DD)", expand=False)
    working["tower_family"] = family.fillna("Unknown")

    code_series = working.get("project_code")
    code_keys = code_series.fillna("" ).map(_compact_project_key) if code_series is not None else pd.Series("", index=working.index)
    name_series = working.get("project_name")
    fallback_keys = name_series.fillna("" ).map(_compact_project_key) if name_series is not None else pd.Series("", index=working.index)
    key_series = code_keys.where(code_keys.astype(bool), fallback_keys)
    working["voltage_label"] = key_series.map(voltage_lookup)
    working["voltage_label"] = working["voltage_label"].fillna("Unmapped")

    return working


def _build_project_meta_lookup(project_info: pd.DataFrame | None) -> dict[str, dict[str, str]]:
    if project_info is None or project_info.empty:
        return {}
    info = project_info.copy()
    info["project_code_value"] = (
        info.get("project_code", info.get("Project Code", ""))
        .fillna("")
        .astype(str)
        .str.strip()
    )
    info["project_display_value"] = (
        info.get("Project Name", info.get("project_name", info["project_code_value"]))
        .fillna("")
        .astype(str)
        .str.strip()
    )
    pch_col = None
    for candidate in ("pch", "PCH", "PCH Name", "pch_name"):
        if candidate in info.columns:
            pch_col = candidate
            break
    if pch_col is None:
        info["pch_value"] = ""
    else:
        info["pch_value"] = info[pch_col]

    lookup: dict[str, dict[str, str]] = {}
    for _, row in info.iterrows():
        entry = {
            "project_code": str(row.get("project_code_value", "")).strip(),
            "project_display": str(row.get("project_display_value", "")).strip(),
            "pch": normalize_pch(row.get("pch_value", "")),
        }
        keys = {
            _compact_project_key(row.get("project_code_value")),
            _compact_project_key(row.get("project_display_value")),
            _compact_project_key(row.get("key_name", "")),
        }
        for key in keys:
            if key and key not in lookup:
                lookup[key] = entry
    return lookup


def _prepare_dpr_records(table: pd.DataFrame, project_code: str) -> pd.DataFrame:
    if table.empty:
        return pd.DataFrame()

    working = table.copy()
    required_columns = {"location_no", "tower_type", "tower_weight", "completion_date", "gang_name", "manpower"}
    if not required_columns.intersection(set(working.columns)):
        return pd.DataFrame()

    working["completion_date"] = pd.to_datetime(working.get("completion_date"), errors="coerce").dt.normalize()
    working["start_date"] = pd.to_datetime(working.get("start_date"), errors="coerce").dt.normalize()
    working["tower_weight"] = _coerce_numeric_series(working, "tower_weight")
    working["manpower"] = _coerce_numeric_series(working, "manpower")

    for column in ("location_no", "tower_type", "status", "gang_name"):
        if column in working.columns:
            working[column] = working[column].fillna("").astype(str).str.strip()
        else:
            working[column] = ""

    working = working.dropna(subset=["completion_date"])
    working = working[working["location_no"].astype(bool)]
    if working.empty:
        return pd.DataFrame()

    durations = (working["completion_date"] - working["start_date"]).dt.days + 1
    durations = durations.where(durations > 0)
    working["derived_productivity"] = working["tower_weight"] / durations

    working = working.assign(
        project_code=str(project_code).strip(),
    )
    columns = [
        "project_code",
        "source_file",
        "completion_date",
        "start_date",
        "location_no",
        "tower_type",
        "tower_weight",
        "status",
        "gang_name",
        "manpower",
        "derived_productivity",
    ]
    present_columns = [column for column in columns if column in working.columns]
    return working[present_columns].copy()


def _format_week_range(week_index: int, month_start: pd.Timestamp, month_end: pd.Timestamp) -> str:
    if week_index < 1:
        return ""
    start = month_start + pd.Timedelta(days=(week_index - 1) * 7)
    end = min(start + pd.Timedelta(days=6), month_end)
    start = max(start, month_start)
    return f"{start.strftime('%d-%b-%Y')} - {end.strftime('%d-%b-%Y')}"


def _assign_week_bucket(dates: pd.Series, month_start: pd.Timestamp, week_labels: Sequence[str]) -> pd.Series:
    offsets = (dates - month_start).dt.days
    buckets = ((offsets // 7) + 1).clip(lower=1, upper=len(week_labels))
    return buckets.fillna(0).astype(int)


def _assign_quarter_week_bucket(
    dates: pd.Series,
    quarter_start: pd.Timestamp,
    *,
    months_count: int = 3,
) -> pd.Series:
    """Assign dates to quarter week buckets (4 weeks per month, last week has remaining days)."""
    date_series = pd.to_datetime(dates, errors="coerce")
    if date_series.empty:
        return pd.Series(index=dates.index, dtype="int64")
    start_key = quarter_start.year * 12 + quarter_start.month
    month_key = date_series.dt.year * 12 + date_series.dt.month
    month_offset = month_key - start_key
    day = date_series.dt.day
    week_in_month = ((day - 1) // 7) + 1
    week_in_month = week_in_month.clip(lower=1, upper=4)
    week_index = month_offset * 4 + week_in_month
    valid = date_series.notna() & (month_offset >= 0) & (month_offset < months_count)
    return week_index.where(valid, 0).fillna(0).astype(int)


def _prepare_month_scope(
    daily_df: pd.DataFrame | None,
    project_info: pd.DataFrame | None,
    month_start: pd.Timestamp,
    month_end: pd.Timestamp,
    week_labels: Sequence[str],
    *,
    filter_by_date: bool = True,
) -> pd.DataFrame:
    if daily_df is None or daily_df.empty:
        return pd.DataFrame()

    working = daily_df.copy()
    working["date"] = pd.to_datetime(working["date"], errors="coerce").dt.normalize()
    if filter_by_date:
        working = working.dropna(subset=["date"])
    if filter_by_date:
        scope = working[(working["date"] >= month_start) & (working["date"] <= month_end)].copy()
    else:
        scope = working.copy()
    if scope.empty:
        return scope

    scope["daily_prod_mt"] = pd.to_numeric(scope.get("daily_prod_mt"), errors="coerce").fillna(0.0)

    scope["project_name"] = scope.get("project_name", "").fillna("").astype(str).str.strip()
    meta_lookup = _build_project_meta_lookup(project_info)
    scope["project_code"] = scope.get("project_code", "").fillna("").astype(str).str.strip()
    scope["project_display"] = scope["project_code"].where(scope["project_code"].astype(bool), scope["project_name"])
    scope["project_display"] = scope["project_display"].fillna("").astype(str).str.strip()
    scope["project_display"] = scope["project_display"].where(scope["project_display"].astype(bool), scope["project_name"])
    scope["project_display"] = scope["project_display"].where(
        scope["project_display"].astype(bool), "(Unlabeled Project)"
    )
    scope["project_name"] = scope["project_name"].where(scope["project_name"].astype(bool), scope["project_display"])

    def _lookup_meta(value: object) -> dict[str, str] | None:
        if not meta_lookup:
            return None
        key = _compact_project_key(value)
        return meta_lookup.get(key)

    if meta_lookup:
        meta_from_code = scope["project_code"].map(_lookup_meta)
        meta_from_name = scope["project_name"].map(_lookup_meta)
        meta_series = meta_from_code.where(meta_from_code.notna(), meta_from_name)

        scope["project_code"] = scope["project_code"].where(
            scope["project_code"].astype(bool),
            meta_series.map(lambda rec: rec.get("project_code", "") if isinstance(rec, dict) else ""),
        )
        scope["project_display"] = scope["project_display"].where(
            scope["project_display"].astype(bool),
            meta_series.map(lambda rec: rec.get("project_display", "") if isinstance(rec, dict) else ""),
        )
        scope["project_display"] = scope["project_display"].fillna("").astype(str).str.strip()
        scope["project_display"] = scope["project_display"].where(scope["project_display"].astype(bool), scope["project_name"])
        scope["project_display"] = scope["project_display"].where(
            scope["project_display"].astype(bool), "(Unlabeled Project)"
        )
        scope["pch_display"] = meta_series.map(lambda rec: rec.get("pch", "") if isinstance(rec, dict) else "")
    else:
        scope["pch_display"] = ""

    scope["pch_display"] = scope["pch_display"].where(scope["pch_display"].astype(bool), scope.get("pch", ""))
    scope["pch_display"] = scope["pch_display"].fillna("").astype(str)
    scope["pch_display"] = scope["pch_display"].map(normalize_pch)
    scope["pch_display"] = scope["pch_display"].where(scope["pch_display"].astype(bool), "Unassigned")

    scope["location_no"] = scope.get("location_no", "").fillna("").astype(str).str.strip()
    scope["completion_date"] = pd.to_datetime(scope.get("completion_date"), errors="coerce").dt.normalize()
    scope["week_index"] = _assign_week_bucket(scope["date"], month_start, week_labels)

    return scope


def _prepare_completion_scope(
    scope: pd.DataFrame,
    month_start: pd.Timestamp,
    month_end: pd.Timestamp,
    week_labels: Sequence[str],
) -> pd.DataFrame:
    if scope.empty:
        return pd.DataFrame()
    completion_mask = (
        scope["completion_date"].notna()
        & (scope["completion_date"] >= month_start)
        & (scope["completion_date"] <= month_end)
    )
    completed = scope[completion_mask].copy()
    if completed.empty:
        return completed
    dedup_cols = [col for col in ("project_name_key", "location_no", "completion_date") if col in completed.columns]
    if dedup_cols:
        completed = completed.drop_duplicates(subset=dedup_cols)
    completed["completion_week_index"] = _assign_week_bucket(completed["completion_date"], month_start, week_labels)
    return completed


def _build_excluded_data_audit_tables(
    scope: pd.DataFrame,
    month_start: pd.Timestamp,
    month_end: pd.Timestamp,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    summary_columns = [
        "PCH",
        "Project",
        "Productivity Missing Date",
        "Productivity Date Outside Range",
        "Completion Missing Date",
        "Completion Date Outside Range",
        "Completion Duplicate Rows",
        "Total Excluded Rows",
        "Total Rows",
    ]
    detail_columns = [
        "PCH",
        "Project",
        "Project Code",
        "Project Name",
        "Work Date",
        "Completion Date",
        "Location No",
        "Gang Name",
        "Daily Productivity (MT)",
        "Tower Weight (MT)",
        "Status",
        "Productivity Exclusion",
        "Completion Exclusion",
        "Duplicate Completion Row",
    ]
    if scope is None or scope.empty:
        return (
            pd.DataFrame(columns=summary_columns),
            pd.DataFrame(columns=detail_columns),
        )

    working = scope.copy()
    working["date"] = pd.to_datetime(working.get("date"), errors="coerce").dt.normalize()
    working["completion_date"] = pd.to_datetime(
        working.get("completion_date"), errors="coerce"
    ).dt.normalize()

    def _safe_text_series(frame: pd.DataFrame, column: str, fallback: str | None = None) -> pd.Series:
        if isinstance(frame.get(column), pd.Series):
            series = frame[column]
        elif fallback and isinstance(frame.get(fallback), pd.Series):
            series = frame[fallback]
        else:
            return pd.Series("", index=frame.index, dtype="object")
        return series.fillna("").astype(str).str.strip()

    project_display = _safe_text_series(working, "project_display", "project_name")
    project_code = _safe_text_series(working, "project_code")
    project_display = project_display.where(project_display.astype(bool), project_code)
    project_display = project_display.where(project_display.astype(bool), "(Unlabeled Project)")
    working["project_display"] = project_display
    working["project_code"] = project_code
    working["project_name"] = _safe_text_series(working, "project_name")

    pch_display = _safe_text_series(working, "pch_display", "pch")
    pch_display = pch_display.map(normalize_pch)
    pch_display = pch_display.where(pch_display.astype(bool), "Unassigned")
    working["pch_display"] = pch_display

    date_valid = working["date"].notna()
    date_in_range = working["date"].between(month_start, month_end, inclusive="both")
    completion_valid = working["completion_date"].notna()
    completion_in_range = working["completion_date"].between(month_start, month_end, inclusive="both")

    productivity_reason = pd.Series("", index=working.index, dtype="object")
    productivity_reason.loc[~date_valid] = "Missing/invalid date"
    productivity_reason.loc[date_valid & ~date_in_range] = "Date outside selected range"

    completion_reason = pd.Series("", index=working.index, dtype="object")
    completion_reason.loc[~completion_valid] = "Missing/invalid completion date"
    completion_reason.loc[completion_valid & ~completion_in_range] = "Completion date outside selected range"

    duplicate_completion = pd.Series(False, index=working.index)
    dedup_cols = [col for col in ("project_name_key", "location_no", "completion_date") if col in working.columns]
    if dedup_cols and completion_in_range.any():
        completion_scope = working[completion_in_range]
        dupes = completion_scope.duplicated(subset=dedup_cols, keep="last")
        duplicate_completion.loc[completion_scope.index] = dupes

    has_issue = (productivity_reason != "") | (completion_reason != "") | duplicate_completion
    working["_prod_missing"] = (~date_valid).astype(int)
    working["_prod_out"] = (date_valid & ~date_in_range).astype(int)
    working["_comp_missing"] = (~completion_valid).astype(int)
    working["_comp_out"] = (completion_valid & ~completion_in_range).astype(int)
    working["_comp_dup"] = duplicate_completion.astype(int)
    working["_any_issue"] = has_issue.astype(int)
    working["_row_count"] = 1

    group_cols = ["pch_display", "project_display"]
    total_rows = working.groupby(group_cols, dropna=False)["_row_count"].sum().rename("Total Rows")
    issue_summary = (
        working[has_issue]
        .groupby(group_cols, dropna=False)
        .agg(
            **{
                "Productivity Missing Date": ("_prod_missing", "sum"),
                "Productivity Date Outside Range": ("_prod_out", "sum"),
                "Completion Missing Date": ("_comp_missing", "sum"),
                "Completion Date Outside Range": ("_comp_out", "sum"),
                "Completion Duplicate Rows": ("_comp_dup", "sum"),
                "Total Excluded Rows": ("_any_issue", "sum"),
            }
        )
    )
    if issue_summary.empty:
        summary = pd.DataFrame(columns=summary_columns)
    else:
        summary = (
            issue_summary.merge(total_rows, left_index=True, right_index=True, how="left")
            .reset_index()
            .rename(columns={"pch_display": "PCH", "project_display": "Project"})
        )
        summary = summary.reindex(columns=summary_columns)

    detail = working[has_issue].copy()
    detail["Productivity Exclusion"] = productivity_reason.loc[detail.index]
    detail["Completion Exclusion"] = completion_reason.loc[detail.index]
    detail["Duplicate Completion Row"] = duplicate_completion.loc[detail.index].map(
        lambda value: "Yes" if value else "No"
    )
    for date_column in ("date", "completion_date"):
        if date_column in detail.columns:
            detail[date_column] = pd.to_datetime(detail[date_column], errors="coerce").dt.strftime("%d-%m-%Y")
            detail[date_column] = detail[date_column].fillna("")
    for numeric_column in ("daily_prod_mt", "tower_weight"):
        if numeric_column in detail.columns:
            detail[numeric_column] = pd.to_numeric(detail[numeric_column], errors="coerce").round(3)

    detail = detail.rename(
        columns={
            "pch_display": "PCH",
            "project_display": "Project",
            "project_code": "Project Code",
            "project_name": "Project Name",
            "date": "Work Date",
            "completion_date": "Completion Date",
            "location_no": "Location No",
            "gang_name": "Gang Name",
            "daily_prod_mt": "Daily Productivity (MT)",
            "tower_weight": "Tower Weight (MT)",
            "status": "Status",
        }
    )
    detail = detail.reindex(columns=[column for column in detail_columns if column in detail.columns])
    if detail.empty:
        detail = pd.DataFrame(columns=detail_columns)

    return summary, detail


def _safe_average(values: list[object]) -> float:
    filtered = [float(v) for v in values if v is not None and not pd.isna(v)]
    return round(sum(filtered) / len(filtered), 2) if filtered else 0.0


def _summarize_tower_weight_metrics(completions: pd.DataFrame) -> dict[str, object]:
    empty = {column: 0.0 for column in TOWER_WEIGHT_COLUMNS}
    if completions is None or completions.empty:
        return empty
    working = completions.copy()
    working["_tower_weight"] = _coerce_numeric_series(working, "tower_weight")
    working["_tower_bucket"] = working.get("tower_type", "").map(_normalize_tower_bucket)

    bucket_counts = working.groupby("_tower_bucket").size().to_dict()
    bucket_avgs = (
        working.groupby("_tower_bucket")["_tower_weight"]
        .mean()
        .to_dict()
    )
    overall_avg = working["_tower_weight"].mean() if not working.empty else 0.0

    summary = {column: 0.0 for column in TOWER_WEIGHT_COLUMNS}
    for bucket in TOWER_BUCKETS:
        summary[bucket] = int(bucket_counts.get(bucket, 0))
        summary[f"{bucket} Avg Tower Weight"] = round(float(bucket_avgs.get(bucket, 0.0) or 0.0), 2)
    summary[EXCEPTION_BUCKET] = int(bucket_counts.get(EXCEPTION_BUCKET, 0))
    summary[f"{EXCEPTION_BUCKET} Avg Tower Weight"] = round(
        float(bucket_avgs.get(EXCEPTION_BUCKET, 0.0) or 0.0), 2
    )
    summary[AVG_TOWER_WEIGHT_OVERALL_COLUMN] = round(float(overall_avg or 0.0), 2)
    return summary


def _build_tower_weight_metrics(
    completions: pd.DataFrame,
    *,
    group_columns: Sequence[str],
) -> pd.DataFrame:
    columns = [*group_columns, *TOWER_WEIGHT_COLUMNS]
    if completions is None or completions.empty:
        return pd.DataFrame(columns=columns)

    working = completions.copy()
    working["_tower_weight"] = _coerce_numeric_series(working, "tower_weight")
    working["_tower_bucket"] = working.get("tower_type", "").map(_normalize_tower_bucket)

    if group_columns:
        group_keys = list(group_columns)
    else:
        working["_all_scope"] = "All"
        group_keys = ["_all_scope"]

    counts = (
        working.groupby(group_keys + ["_tower_bucket"], dropna=False)
        .size()
        .reset_index(name="count")
    )
    counts_pivot = counts.pivot_table(
        index=group_keys, columns="_tower_bucket", values="count", fill_value=0
    )

    avg_weights = (
        working.groupby(group_keys + ["_tower_bucket"], dropna=False)["_tower_weight"]
        .mean()
        .reset_index(name="avg_weight")
    )
    avg_pivot = avg_weights.pivot_table(
        index=group_keys, columns="_tower_bucket", values="avg_weight", fill_value=0.0
    )

    overall_avg = (
        working.groupby(group_keys, dropna=False)["_tower_weight"]
        .mean()
        .reset_index(name=AVG_TOWER_WEIGHT_OVERALL_COLUMN)
    )

    metrics = pd.DataFrame(index=counts_pivot.index.union(avg_pivot.index))
    def _ensure_series(frame: pd.DataFrame, column: str, default_value: float) -> pd.Series:
        series = frame.get(column)
        if isinstance(series, pd.Series):
            return series
        return pd.Series(default_value, index=metrics.index)

    for bucket in TOWER_BUCKETS:
        count_series = _ensure_series(counts_pivot, bucket, 0)
        metrics[bucket] = count_series.fillna(0).astype(int)
        avg_series = _ensure_series(avg_pivot, bucket, 0.0)
        metrics[f"{bucket} Avg Tower Weight"] = avg_series.fillna(0.0).round(2)
    exception_counts = _ensure_series(counts_pivot, EXCEPTION_BUCKET, 0)
    metrics[EXCEPTION_BUCKET] = exception_counts.fillna(0).astype(int)
    exception_avg = _ensure_series(avg_pivot, EXCEPTION_BUCKET, 0.0)
    metrics[f"{EXCEPTION_BUCKET} Avg Tower Weight"] = exception_avg.fillna(0.0).round(2)
    metrics = metrics.reset_index()
    metrics = metrics.merge(overall_avg, on=group_keys, how="left")
    metrics[AVG_TOWER_WEIGHT_OVERALL_COLUMN] = (
        metrics[AVG_TOWER_WEIGHT_OVERALL_COLUMN].fillna(0.0).round(2)
    )

    if not group_columns:
        metrics = metrics.drop(columns=["_all_scope"], errors="ignore")

    return metrics.reindex(columns=columns)


def _compute_weekly_productivity_maps(scope: pd.DataFrame, week_labels: Sequence[str]) -> dict[str, dict[str, float]]:
    maps = {label: {} for label in week_labels}
    if scope.empty:
        return maps
    for week_idx, label in enumerate(week_labels, start=1):
        week_scope = scope[scope.get("week_index") == week_idx]
        if week_scope.empty:
            continue
        project_map, _ = compute_project_baseline_maps_for(week_scope, "daily_prod_mt")
        if project_map:
            maps[label] = project_map
    return maps


def _build_overall_productivity_table(
    scope: pd.DataFrame,
    completions: pd.DataFrame,
    week_labels: Sequence[str],
    week_maps: dict[str, dict[str, float]],
) -> pd.DataFrame:
    columns = [*week_labels, AVG_PRODUCTIVITY_COLUMN, TOTAL_MT_COLUMN, TOTAL_COUNT_COLUMN, *TOWER_WEIGHT_COLUMNS]
    if scope is None or scope.empty:
        return pd.DataFrame(columns=columns)

    row: dict[str, object] = {}
    for label in week_labels:
        label_map = week_maps.get(label, {})
        row[label] = _safe_average(list(label_map.values()))

    metric_series = scope.get("daily_prod_mt")
    if isinstance(metric_series, pd.Series):
        productivity_series = pd.to_numeric(metric_series, errors="coerce")
        valid_values = productivity_series.dropna()
        row[AVG_PRODUCTIVITY_COLUMN] = round(float(valid_values.mean()), 2) if not valid_values.empty else 0.0
    else:
        row[AVG_PRODUCTIVITY_COLUMN] = 0.0

    if isinstance(completions, pd.DataFrame) and not completions.empty and "_tower_weight" in completions.columns:
        total_mt = completions["_tower_weight"].sum()
        total_count = len(completions)
    else:
        total_mt = 0.0
        total_count = 0
    row[TOTAL_MT_COLUMN] = round(float(total_mt), 2)
    row[TOTAL_COUNT_COLUMN] = int(total_count)
    row.update(_summarize_tower_weight_metrics(completions))

    return pd.DataFrame([row], columns=columns).fillna(0.0)


def _build_productivity_tables(
    scope: pd.DataFrame,
    completions: pd.DataFrame,
    week_labels: Sequence[str],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    pch_columns = [
        "PCH",
        *week_labels,
        AVG_PRODUCTIVITY_COLUMN,
        TOTAL_MT_COLUMN,
        TOTAL_COUNT_COLUMN,
        *TOWER_WEIGHT_COLUMNS,
    ]
    project_columns = [
        "PCH",
        "Project",
        *week_labels,
        AVG_PRODUCTIVITY_COLUMN,
        TOTAL_MT_COLUMN,
        TOTAL_COUNT_COLUMN,
        *TOWER_WEIGHT_COLUMNS,
    ]

    if scope.empty:
        empty_pch = pd.DataFrame(columns=pch_columns)
        empty_project = pd.DataFrame(columns=project_columns)
        empty_overall = pd.DataFrame(
            columns=[*week_labels, AVG_PRODUCTIVITY_COLUMN, TOTAL_MT_COLUMN, TOTAL_COUNT_COLUMN, *TOWER_WEIGHT_COLUMNS]
        )
        return empty_overall, empty_pch, empty_project

    project_meta = (
        scope[["pch_display", "project_name", "project_display"]]
        .dropna(subset=["project_name"])
        .drop_duplicates(subset=["pch_display", "project_name"])
        .reset_index(drop=True)
    )
    week_maps = _compute_weekly_productivity_maps(scope, week_labels)

    project_avg_map = (
        scope.groupby("project_name")["daily_prod_mt"].mean().to_dict()
        if not scope.empty
        else {}
    )
    pch_avg_map = (
        scope.groupby("pch_display")["daily_prod_mt"].mean().to_dict()
        if not scope.empty
        else {}
    )

    tower_weights = pd.to_numeric(completions.get("tower_weight"), errors="coerce") if isinstance(completions, pd.DataFrame) else pd.Series(dtype=float)
    if isinstance(completions, pd.DataFrame):
        completions = completions.assign(_tower_weight=tower_weights.fillna(0.0))
    else:
        completions = pd.DataFrame(columns=["pch_display", "project_name", "_tower_weight"])

    overall_summary = _build_overall_productivity_table(scope, completions, week_labels, week_maps)

    tower_metrics_pch = _build_tower_weight_metrics(completions, group_columns=["pch_display"])
    tower_metrics_project = _build_tower_weight_metrics(
        completions, group_columns=["pch_display", "project_name"]
    )
    tower_metrics_pch_map = (
        tower_metrics_pch.set_index("pch_display").to_dict(orient="index")
        if not tower_metrics_pch.empty
        else {}
    )
    tower_metrics_project_map = (
        tower_metrics_project.set_index(["pch_display", "project_name"]).to_dict(orient="index")
        if not tower_metrics_project.empty
        else {}
    )

    mt_totals_by_project = (
        completions.groupby(["pch_display", "project_name"])["_tower_weight"].sum().to_dict()
        if not completions.empty
        else {}
    )
    mt_totals_by_pch = (
        completions.groupby("pch_display")["_tower_weight"].sum().to_dict()
        if not completions.empty
        else {}
    )

    counts_by_project: dict[tuple[str, str], int] = {}
    counts_by_pch: dict[str, int] = {}
    if isinstance(completions, pd.DataFrame) and not completions.empty:
        counts_by_project = completions.groupby(["pch_display", "project_name"]).size().to_dict()
        counts_by_pch = completions.groupby("pch_display").size().to_dict()

    pch_rows: list[dict[str, object]] = []
    if not project_meta.empty:
        for pch_name in sorted(project_meta["pch_display"].unique(), key=lambda name: _pch_sort_components(name)):
            projects_in_pch = project_meta[project_meta["pch_display"] == pch_name]["project_name"].tolist()
            row = {"PCH": pch_name}
            for label in week_labels:
                project_values = [week_maps.get(label, {}).get(name) for name in projects_in_pch]
                row[label] = _safe_average(project_values)
            avg_value = pch_avg_map.get(pch_name)
            row[AVG_PRODUCTIVITY_COLUMN] = round(float(avg_value), 2) if avg_value is not None and not pd.isna(avg_value) else 0.0
            row[TOTAL_MT_COLUMN] = round(float(mt_totals_by_pch.get(pch_name, 0.0)), 2)
            row[TOTAL_COUNT_COLUMN] = int(counts_by_pch.get(pch_name, 0))
            row.update(tower_metrics_pch_map.get(pch_name, {column: 0.0 for column in TOWER_WEIGHT_COLUMNS}))
            pch_rows.append(row)

    if pch_rows:
        pch_summary = pd.DataFrame(pch_rows).reindex(columns=pch_columns).fillna(0.0)
        pch_summary = _sort_pch_frame(pch_summary, column="PCH")
        pch_summary = _append_totals_row(pch_summary)
    else:
        pch_summary = pd.DataFrame(columns=pch_columns)

    project_rows: list[dict[str, object]] = []
    for _, meta_row in project_meta.sort_values(["pch_display", "project_display"], key=lambda col: col.astype(str)).iterrows():
        pch_value = meta_row["pch_display"]
        project_name = meta_row["project_name"]
        display_name = meta_row["project_display"]
        row = {"PCH": pch_value, "Project": display_name}
        for label in week_labels:
            value = week_maps.get(label, {}).get(project_name)
            row[label] = round(float(value), 2) if value is not None and not pd.isna(value) else 0.0
        avg_value = project_avg_map.get(project_name)
        row[AVG_PRODUCTIVITY_COLUMN] = round(float(avg_value), 2) if avg_value is not None and not pd.isna(avg_value) else 0.0
        row[TOTAL_MT_COLUMN] = round(float(mt_totals_by_project.get((pch_value, project_name), 0.0)), 2)
        row[TOTAL_COUNT_COLUMN] = int(counts_by_project.get((pch_value, project_name), 0))
        row.update(
            tower_metrics_project_map.get(
                (pch_value, project_name),
                {column: 0.0 for column in TOWER_WEIGHT_COLUMNS},
            )
        )
        project_rows.append(row)

    if project_rows:
        project_summary = pd.DataFrame(project_rows).reindex(columns=project_columns).fillna(0.0)
        order_components = project_summary["PCH"].map(_pch_sort_components)
        project_summary = project_summary.assign(
            _pch_order_bucket=order_components.map(lambda pair: pair[0]),
            _pch_order_value=order_components.map(lambda pair: pair[1]),
            _project_order=project_summary["Project"].astype(str).str.lower(),
        )
        project_summary = (
            project_summary.sort_values(by=["_pch_order_bucket", "_pch_order_value", "_project_order"])
            .drop(columns=["_pch_order_bucket", "_pch_order_value", "_project_order"])
            .reset_index(drop=True)
        )
        project_summary = _append_totals_row(project_summary)
    else:
        project_summary = pd.DataFrame(columns=project_columns)

    return overall_summary, pch_summary, project_summary


def _monthly_avg_column(month_label: str) -> str:
    return f"{month_label}__avg"


def _monthly_count_column(month_label: str) -> str:
    return f"{month_label}__count"


def _monthly_mt_column(month_label: str) -> str:
    return f"{month_label}__mt"


def _monthly_column_layout(
    month_labels: Sequence[str],
    *,
    group_labels: Sequence[str],
) -> tuple[list[str], list[str], list[str]]:
    ordered_columns = [*group_labels]
    top_headers: list[str] = []
    bottom_headers: list[str] = []

    for label in group_labels:
        top_headers.append("")
        bottom_headers.append(label)

    for month_label in month_labels:
        ordered_columns.extend(
            [
                _monthly_avg_column(month_label),
                _monthly_count_column(month_label),
                _monthly_mt_column(month_label),
            ]
        )
        top_headers.extend([month_label, "", ""])
        bottom_headers.extend([MONTHLY_AVG_LABEL, MONTHLY_COUNT_LABEL, MONTHLY_MT_LABEL])

    ordered_columns.extend(
        [
            _monthly_avg_column("Overall"),
            _monthly_count_column("Overall"),
            _monthly_mt_column("Overall"),
        ]
    )
    top_headers.extend(["Overall", "", ""])
    bottom_headers.extend([MONTHLY_AVG_LABEL, MONTHLY_COUNT_LABEL, MONTHLY_MT_LABEL])
    return ordered_columns, top_headers, bottom_headers


def _build_monthly_group_productivity_table(
    scope: pd.DataFrame,
    completions: pd.DataFrame,
    *,
    month_keys: Sequence[pd.Timestamp],
    month_labels: Sequence[str],
    group_columns: Sequence[str],
    group_labels: Sequence[str],
) -> pd.DataFrame:
    ordered_columns, _, _ = _monthly_column_layout(month_labels, group_labels=group_labels)
    if scope is None or scope.empty:
        return pd.DataFrame(columns=ordered_columns)

    working_scope = scope.copy()
    working_scope["date"] = pd.to_datetime(working_scope.get("date"), errors="coerce").dt.normalize()
    working_scope["daily_prod_mt"] = pd.to_numeric(working_scope.get("daily_prod_mt"), errors="coerce")
    working_scope = working_scope.dropna(subset=["date"])
    if working_scope.empty:
        return pd.DataFrame(columns=ordered_columns)
    working_scope["month_key"] = working_scope["date"].dt.to_period("M").dt.to_timestamp()

    groups = list(group_columns)
    rename_map = {source: label for source, label in zip(group_columns, group_labels)}

    def _month_avg_key(label: str) -> str:
        return _monthly_avg_column(label)

    def _month_count_key(label: str) -> str:
        return _monthly_count_column(label)

    def _month_mt_key(label: str) -> str:
        return _monthly_mt_column(label)

    overall_avg_key = _month_avg_key("Overall")
    overall_count_key = _month_count_key("Overall")
    overall_mt_key = _month_mt_key("Overall")
    numeric_columns = [overall_avg_key, overall_count_key, overall_mt_key]

    if not groups:
        row: dict[str, object] = {}
        for month_key, month_label in zip(month_keys, month_labels):
            month_scope = working_scope[working_scope["month_key"] == month_key]
            avg_value = month_scope["daily_prod_mt"].dropna().mean() if not month_scope.empty else 0.0
            row[_month_avg_key(month_label)] = round(float(avg_value) if not pd.isna(avg_value) else 0.0, 2)
            row[_month_count_key(month_label)] = 0
            row[_month_mt_key(month_label)] = 0.0
            numeric_columns.extend(
                [_month_avg_key(month_label), _month_count_key(month_label), _month_mt_key(month_label)]
            )

        overall_avg_value = working_scope["daily_prod_mt"].dropna().mean()
        row[overall_avg_key] = round(float(overall_avg_value) if not pd.isna(overall_avg_value) else 0.0, 2)
        row[overall_count_key] = 0
        row[overall_mt_key] = 0.0

        if isinstance(completions, pd.DataFrame) and not completions.empty:
            completion_scope = completions.copy()
            completion_scope["completion_date"] = pd.to_datetime(
                completion_scope.get("completion_date"), errors="coerce"
            ).dt.normalize()
            completion_scope = completion_scope.dropna(subset=["completion_date"])
            if not completion_scope.empty:
                completion_scope["month_key"] = completion_scope["completion_date"].dt.to_period("M").dt.to_timestamp()
                completion_scope["_tower_weight"] = _coerce_numeric_series(
                    completion_scope, "tower_weight"
                ).fillna(0.0)
                for month_key, month_label in zip(month_keys, month_labels):
                    month_comp = completion_scope[completion_scope["month_key"] == month_key]
                    row[_month_count_key(month_label)] = int(len(month_comp))
                    row[_month_mt_key(month_label)] = round(float(month_comp["_tower_weight"].sum()), 2)
                row[overall_count_key] = int(len(completion_scope))
                row[overall_mt_key] = round(float(completion_scope["_tower_weight"].sum()), 2)

        result = pd.DataFrame([row], columns=ordered_columns).fillna(0.0)
        for month_label in month_labels:
            count_key = _month_count_key(month_label)
            if count_key in result.columns:
                result[count_key] = pd.to_numeric(result[count_key], errors="coerce").fillna(0).astype(int)
            mt_key = _month_mt_key(month_label)
            if mt_key in result.columns:
                result[mt_key] = pd.to_numeric(result[mt_key], errors="coerce").fillna(0.0).round(2)
        result[overall_count_key] = pd.to_numeric(result[overall_count_key], errors="coerce").fillna(0).astype(int)
        result[overall_mt_key] = pd.to_numeric(result[overall_mt_key], errors="coerce").fillna(0.0).round(2)
        return result

    for column in groups:
        if column not in working_scope.columns:
            working_scope[column] = ""
        working_scope[column] = working_scope[column].fillna("").astype(str).str.strip()

    summary = working_scope[groups].drop_duplicates().reset_index(drop=True)
    overall_avg = (
        working_scope.groupby(groups, dropna=False)["daily_prod_mt"]
        .mean()
        .reset_index(name=overall_avg_key)
    )
    summary = summary.merge(overall_avg, on=groups, how="left")

    for month_key, month_label in zip(month_keys, month_labels):
        avg_key = _month_avg_key(month_label)
        month_avg = (
            working_scope[working_scope["month_key"] == month_key]
            .groupby(groups, dropna=False)["daily_prod_mt"]
            .mean()
            .reset_index(name=avg_key)
        )
        summary = summary.merge(month_avg, on=groups, how="left")
        numeric_columns.append(avg_key)

    if isinstance(completions, pd.DataFrame) and not completions.empty:
        completion_scope = completions.copy()
        completion_scope["completion_date"] = pd.to_datetime(
            completion_scope.get("completion_date"), errors="coerce"
        ).dt.normalize()
        completion_scope = completion_scope.dropna(subset=["completion_date"])
        if not completion_scope.empty:
            for column in groups:
                if column not in completion_scope.columns:
                    completion_scope[column] = ""
                completion_scope[column] = completion_scope[column].fillna("").astype(str).str.strip()
            completion_scope["month_key"] = completion_scope["completion_date"].dt.to_period("M").dt.to_timestamp()

            counts_overall = (
                completion_scope.groupby(groups, dropna=False)
                .size()
                .reset_index(name=overall_count_key)
            )
            summary = summary.merge(counts_overall, on=groups, how="left")
            weights_overall = (
                completion_scope.assign(
                    _tower_weight=_coerce_numeric_series(completion_scope, "tower_weight").fillna(0.0)
                )
                .groupby(groups, dropna=False)["_tower_weight"]
                .sum()
                .reset_index(name=overall_mt_key)
            )
            summary = summary.merge(weights_overall, on=groups, how="left")
            for month_key, month_label in zip(month_keys, month_labels):
                count_key = _month_count_key(month_label)
                mt_key = _month_mt_key(month_label)
                month_counts = (
                    completion_scope[completion_scope["month_key"] == month_key]
                    .groupby(groups, dropna=False)
                    .size()
                    .reset_index(name=count_key)
                )
                summary = summary.merge(month_counts, on=groups, how="left")
                month_weights = (
                    completion_scope[completion_scope["month_key"] == month_key]
                    .assign(
                        _tower_weight=lambda frame: _coerce_numeric_series(frame, "tower_weight").fillna(0.0)
                    )
                    .groupby(groups, dropna=False)["_tower_weight"]
                    .sum()
                    .reset_index(name=mt_key)
                )
                summary = summary.merge(month_weights, on=groups, how="left")
                numeric_columns.append(count_key)
                numeric_columns.append(mt_key)
        else:
            summary[overall_count_key] = 0
            summary[overall_mt_key] = 0.0
            for month_label in month_labels:
                summary[_month_count_key(month_label)] = 0
                summary[_month_mt_key(month_label)] = 0.0
    else:
        summary[overall_count_key] = 0
        summary[overall_mt_key] = 0.0
        for month_label in month_labels:
            summary[_month_count_key(month_label)] = 0
            summary[_month_mt_key(month_label)] = 0.0

    summary = summary.rename(columns=rename_map)
    for month_label in month_labels:
        count_key = _month_count_key(month_label)
        if count_key not in summary.columns:
            summary[count_key] = 0
        mt_key = _month_mt_key(month_label)
        if mt_key not in summary.columns:
            summary[mt_key] = 0.0
        numeric_columns.append(count_key)
        numeric_columns.append(mt_key)

    for key in set(numeric_columns):
        if key not in summary.columns:
            summary[key] = 0.0
        if key.endswith("__count"):
            summary[key] = pd.to_numeric(summary[key], errors="coerce").fillna(0).astype(int)
        else:
            summary[key] = pd.to_numeric(summary[key], errors="coerce").fillna(0.0).round(2)

    summary = summary.reindex(columns=ordered_columns)

    if group_labels == ["PCH"]:
        summary = _sort_pch_frame(summary, column="PCH")
    elif group_labels == ["PCH", "Project"] and not summary.empty:
        order_components = summary["PCH"].map(_pch_sort_components)
        summary = summary.assign(
            _pch_order_bucket=order_components.map(lambda pair: pair[0]),
            _pch_order_value=order_components.map(lambda pair: pair[1]),
            _project_order=summary["Project"].astype(str).str.lower(),
        )
        summary = (
            summary.sort_values(by=["_pch_order_bucket", "_pch_order_value", "_project_order"])
            .drop(columns=["_pch_order_bucket", "_pch_order_value", "_project_order"])
            .reset_index(drop=True)
        )
    return summary


def _build_monthly_productivity_tables(
    scope: pd.DataFrame,
    completions: pd.DataFrame,
    *,
    months: Sequence[pd.Timestamp],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list[str]]:
    normalized_months = [
        pd.Timestamp(month).normalize().to_period("M").to_timestamp()
        for month in months
    ]
    month_keys = sorted(dict.fromkeys(normalized_months))
    month_labels = [month.strftime("%b-%y") for month in month_keys]
    overall_summary = _build_monthly_group_productivity_table(
        scope,
        completions,
        month_keys=month_keys,
        month_labels=month_labels,
        group_columns=[],
        group_labels=[],
    )
    pch_summary = _build_monthly_group_productivity_table(
        scope,
        completions,
        month_keys=month_keys,
        month_labels=month_labels,
        group_columns=["pch_display"],
        group_labels=["PCH"],
    )
    project_summary = _build_monthly_group_productivity_table(
        scope,
        completions,
        month_keys=month_keys,
        month_labels=month_labels,
        group_columns=["pch_display", "project_display"],
        group_labels=["PCH", "Project"],
    )
    return overall_summary, pch_summary, project_summary, month_labels


def _build_project_gang_rankings(
    scope: pd.DataFrame,
    completions: pd.DataFrame,
    *,
    top_n: int = 3,
    min_erections: int = 4,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    columns = [
        "PCH",
        "Project",
        "Gang Name",
        AVG_PRODUCTIVITY_COLUMN,
        AVG_PRODUCTIVITY_OVERALL_COLUMN,
        TOTAL_MT_COLUMN,
        TOTAL_COUNT_COLUMN,
        *TOWER_WEIGHT_COLUMNS,
        IDLE_DAYS_COLUMN,
        IDLE_INTERVALS_COLUMN,
    ]
    if scope is None or scope.empty or "gang_name" not in scope.columns:
        empty = pd.DataFrame(columns=columns)
        return empty, empty

    working = scope.copy()
    working["gang_name"] = working["gang_name"].fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty:
        empty = pd.DataFrame(columns=columns)
        return empty, empty

    group_cols = ["pch_display", "project_name", "project_display", "gang_name"]
    avg_prod = (
        working.groupby(group_cols, dropna=False)["daily_prod_mt"]
        .mean()
        .reset_index()
        .rename(columns={"daily_prod_mt": AVG_PRODUCTIVITY_COLUMN})
    )
    avg_prod[AVG_PRODUCTIVITY_COLUMN] = pd.to_numeric(
        avg_prod[AVG_PRODUCTIVITY_COLUMN], errors="coerce"
    ).fillna(0.0).round(2)

    totals = pd.DataFrame(columns=[*group_cols, TOTAL_MT_COLUMN, TOTAL_COUNT_COLUMN])
    if isinstance(completions, pd.DataFrame) and not completions.empty:
        comp = completions.copy()
        comp["gang_name"] = comp.get("gang_name", "").fillna("").astype(str).str.strip()
        comp = comp[comp["gang_name"].astype(bool)]
        if not comp.empty:
            tower_weights = pd.to_numeric(comp.get("tower_weight"), errors="coerce").fillna(0.0)
            totals = (
                comp.assign(_tower_weight=tower_weights)
                .groupby(group_cols, dropna=False)
                .agg(
                    **{
                        TOTAL_MT_COLUMN: ("_tower_weight", "sum"),
                        TOTAL_COUNT_COLUMN: ("_tower_weight", "size"),
                    }
                )
                .reset_index()
            )
            totals[TOTAL_MT_COLUMN] = totals[TOTAL_MT_COLUMN].round(2)
            totals[TOTAL_COUNT_COLUMN] = totals[TOTAL_COUNT_COLUMN].astype(int)

    merged = avg_prod.merge(totals, on=group_cols, how="left")
    merged[TOTAL_MT_COLUMN] = merged[TOTAL_MT_COLUMN].fillna(0.0).round(2)
    merged[TOTAL_COUNT_COLUMN] = merged[TOTAL_COUNT_COLUMN].fillna(0).astype(int)
    merged = merged[merged[TOTAL_COUNT_COLUMN] >= min_erections].copy()
    if merged.empty:
        empty = pd.DataFrame(columns=columns)
        return empty, empty
    tower_metrics = _build_tower_weight_metrics(
        completions, group_columns=["pch_display", "project_name", "gang_name"]
    )
    if not tower_metrics.empty:
        merged = merged.merge(
            tower_metrics,
            on=["pch_display", "project_name", "gang_name"],
            how="left",
        )
    for column in TOWER_WEIGHT_COLUMNS:
        if column not in merged.columns:
            merged[column] = 0.0
        else:
            merged[column] = merged[column].fillna(0.0)
            if column in TOWER_BUCKET_COUNT_COLUMNS:
                merged[column] = merged[column].astype(int)
            else:
                merged[column] = pd.to_numeric(merged[column], errors="coerce").fillna(0.0).round(2)
    for column in (IDLE_DAYS_COLUMN, IDLE_INTERVALS_COLUMN):
        if column not in merged.columns:
            merged[column] = 0
        else:
            merged[column] = pd.to_numeric(merged[column], errors="coerce").fillna(0).astype(int)
    if AVG_PRODUCTIVITY_OVERALL_COLUMN not in merged.columns:
        merged[AVG_PRODUCTIVITY_OVERALL_COLUMN] = 0.0
    else:
        merged[AVG_PRODUCTIVITY_OVERALL_COLUMN] = (
            pd.to_numeric(merged[AVG_PRODUCTIVITY_OVERALL_COLUMN], errors="coerce").fillna(0.0).round(2)
        )

    def _collect_ranked(top: bool) -> pd.DataFrame:
        rows: list[pd.DataFrame] = []
        for _, group in merged.groupby(["pch_display", "project_name", "project_display"], dropna=False):
            ordered = group.sort_values(
                by=[AVG_PRODUCTIVITY_COLUMN, "gang_name"],
                ascending=[not top, True],
                kind="mergesort",
            )
            rows.append(ordered.head(top_n))
        if not rows:
            return pd.DataFrame(columns=columns)
        ranked = pd.concat(rows, ignore_index=True)
        ranked = ranked.rename(
            columns={
                "pch_display": "PCH",
                "project_display": "Project",
                "gang_name": "Gang Name",
            }
        )
        ranked = ranked[columns].copy()
        order_components = ranked["PCH"].map(_pch_sort_components)
        ranked = ranked.assign(
            _pch_order_bucket=order_components.map(lambda pair: pair[0]),
            _pch_order_value=order_components.map(lambda pair: pair[1]),
            _project_order=ranked["Project"].astype(str).str.lower(),
        )
        ranked = (
            ranked.sort_values(by=["_pch_order_bucket", "_pch_order_value", "_project_order"])
            .drop(columns=["_pch_order_bucket", "_pch_order_value", "_project_order"])
            .reset_index(drop=True)
        )
        return ranked

    top_ranked = _collect_ranked(top=True)
    bottom_ranked = _collect_ranked(top=False)
    return top_ranked, bottom_ranked


def _build_gang_level_productivity_table(
    scope: pd.DataFrame,
    completions: pd.DataFrame,
) -> pd.DataFrame:
    columns = [
        "PCH",
        "Project",
        "Gang Name",
        AVG_PRODUCTIVITY_COLUMN,
        TOTAL_MT_COLUMN,
        TOTAL_COUNT_COLUMN,
        *TOWER_WEIGHT_COLUMNS,
    ]
    if scope is None or scope.empty or "gang_name" not in scope.columns:
        return pd.DataFrame(columns=columns)

    working = scope.copy()
    working["gang_name"] = working["gang_name"].fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty:
        return pd.DataFrame(columns=columns)

    if completions is None or completions.empty:
        return pd.DataFrame(columns=columns)

    comp = completions.copy()
    comp["gang_name"] = comp.get("gang_name", "").fillna("").astype(str).str.strip()
    comp = comp[comp["gang_name"].astype(bool)]
    if comp.empty:
        return pd.DataFrame(columns=columns)

    group_cols = ["pch_display", "project_name", "project_display", "gang_name"]
    avg_prod = (
        working.groupby(group_cols, dropna=False)["daily_prod_mt"]
        .mean()
        .reset_index()
        .rename(columns={"daily_prod_mt": AVG_PRODUCTIVITY_COLUMN})
    )
    avg_prod[AVG_PRODUCTIVITY_COLUMN] = pd.to_numeric(
        avg_prod[AVG_PRODUCTIVITY_COLUMN], errors="coerce"
    ).fillna(0.0).round(2)

    tower_weights = pd.to_numeric(comp.get("tower_weight"), errors="coerce").fillna(0.0)
    totals = (
        comp.assign(_tower_weight=tower_weights)
        .groupby(group_cols, dropna=False)
        .agg(
            **{
                TOTAL_MT_COLUMN: ("_tower_weight", "sum"),
                TOTAL_COUNT_COLUMN: ("_tower_weight", "size"),
            }
        )
        .reset_index()
    )
    totals[TOTAL_MT_COLUMN] = totals[TOTAL_MT_COLUMN].round(2)
    totals[TOTAL_COUNT_COLUMN] = totals[TOTAL_COUNT_COLUMN].astype(int)

    merged = avg_prod.merge(totals, on=group_cols, how="inner")
    if merged.empty:
        return pd.DataFrame(columns=columns)

    tower_metrics = _build_tower_weight_metrics(comp, group_columns=group_cols)
    if not tower_metrics.empty:
        merged = merged.merge(tower_metrics, on=group_cols, how="left")

    for column in TOWER_WEIGHT_COLUMNS:
        if column not in merged.columns:
            merged[column] = 0.0
        else:
            merged[column] = merged[column].fillna(0.0)
            if column in TOWER_BUCKET_COUNT_COLUMNS:
                merged[column] = merged[column].astype(int)
            else:
                merged[column] = pd.to_numeric(merged[column], errors="coerce").fillna(0.0).round(2)

    merged = merged.rename(
        columns={
            "pch_display": "PCH",
            "project_display": "Project",
            "gang_name": "Gang Name",
        }
    )
    merged = merged[columns].copy()

    order_components = merged["PCH"].map(_pch_sort_components)
    merged = merged.assign(
        _pch_order_bucket=order_components.map(lambda pair: pair[0]),
        _pch_order_value=order_components.map(lambda pair: pair[1]),
        _project_order=merged["Project"].astype(str).str.lower(),
        _gang_order=merged["Gang Name"].astype(str).str.lower(),
    )
    merged = (
        merged.sort_values(by=["_pch_order_bucket", "_pch_order_value", "_project_order", "_gang_order"])
        .drop(columns=["_pch_order_bucket", "_pch_order_value", "_project_order", "_gang_order"])
        .reset_index(drop=True)
    )
    return merged


def _build_overall_gang_productivity_map(daily_df: pd.DataFrame | None) -> dict[str, float]:
    if daily_df is None or daily_df.empty or "gang_name" not in daily_df.columns:
        return {}
    working = daily_df.copy()
    working["gang_name"] = working.get("gang_name", "").fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty or "daily_prod_mt" not in working.columns:
        return {}
    productivity = pd.to_numeric(working["daily_prod_mt"], errors="coerce")
    grouped = productivity.groupby(working["gang_name"]).mean().dropna()
    return {gang: round(float(value), 2) for gang, value in grouped.items()}


def _build_gang_productivity_map(scope: pd.DataFrame, metric_column: str = "daily_prod_mt") -> dict[str, float]:
    if scope is None or scope.empty or "gang_name" not in scope.columns:
        return {}
    working = scope.copy()
    working["gang_name"] = working.get("gang_name", "").fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty or metric_column not in working.columns:
        return {}
    metric = pd.to_numeric(working[metric_column], errors="coerce")
    grouped = metric.groupby(working["gang_name"]).mean().dropna()
    return {gang: round(float(value), 2) for gang, value in grouped.items()}


def _build_gang_total_map(scope: pd.DataFrame | None, metric_column: str = "daily_prod_mt") -> dict[str, float]:
    if scope is None or scope.empty or "gang_name" not in scope.columns:
        return {}
    working = scope.copy()
    working["gang_name"] = working.get("gang_name", "").fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty or metric_column not in working.columns:
        return {}
    metric = pd.to_numeric(working[metric_column], errors="coerce").fillna(0.0)
    grouped = metric.groupby(working["gang_name"]).sum()
    return {gang: round(float(value), 2) for gang, value in grouped.items()}


def _extract_completion_rows(daily_df: pd.DataFrame | None) -> pd.DataFrame:
    if daily_df is None or daily_df.empty:
        return pd.DataFrame()
    working = daily_df.copy()
    working["date"] = pd.to_datetime(working.get("date"), errors="coerce").dt.normalize()
    working["completion_date"] = pd.to_datetime(
        working.get("completion_date"), errors="coerce"
    ).dt.normalize()
    working = working.dropna(subset=["date", "completion_date"])
    if working.empty:
        return pd.DataFrame()
    completions = working[working["date"].eq(working["completion_date"])].copy()
    if completions.empty:
        return pd.DataFrame()
    dedup_cols = [
        col
        for col in ("project_name_key", "project_name", "location_no", "completion_date")
        if col in completions.columns
    ]
    if dedup_cols:
        completions = completions.drop_duplicates(subset=dedup_cols, keep="last")
    return completions


def _build_completion_counts_by_gang(completions: pd.DataFrame | None) -> dict[str, int]:
    if completions is None or completions.empty or "gang_name" not in completions.columns:
        return {}
    working = completions.copy()
    working["gang_name"] = working.get("gang_name", "").fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty:
        return {}
    return working.groupby("gang_name").size().astype(int).to_dict()


def _idle_productivity_buckets() -> list[tuple[str, callable]]:
    return [
        ("0-1", lambda value: 0 <= value < 1),
        ("1-2", lambda value: 1 <= value < 2),
        ("2-3", lambda value: 2 <= value < 3),
        ("3-4", lambda value: 3 <= value < 4),
        ("4-5", lambda value: 4 <= value < 5),
        ("5-6", lambda value: 5 <= value < 6),
        ("6-7", lambda value: 6 <= value < 7),
        ("7-8", lambda value: 7 <= value < 8),
        ("8-9", lambda value: 8 <= value < 9),
        ("9-10", lambda value: 9 <= value < 10),
        ("10-11", lambda value: 10 <= value < 11),
        ("11-12", lambda value: 11 <= value <= 12),
        (">12", lambda value: value > 12),
    ]


def _assign_idle_bucket(value: float, buckets: Sequence[tuple[str, callable]]) -> str:
    for label, predicate in buckets:
        if predicate(value):
            return label
    return ""


def _build_gang_project_code_map(completions: pd.DataFrame | None) -> dict[str, str]:
    if completions is None or completions.empty:
        return {}
    working = completions.copy()
    working["gang_name"] = working.get("gang_name", "").fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty:
        return {}

    working["project_code"] = working.get("project_code", "").fillna("").astype(str).str.strip()
    working["project_name"] = working.get("project_name", "").fillna("").astype(str).str.strip()
    working["project_label"] = working["project_code"].where(
        working["project_code"].astype(bool),
        working["project_name"],
    )

    def _join_project_codes(series: pd.Series) -> str:
        seen: set[str] = set()
        ordered: list[str] = []
        for value in series.fillna("").astype(str):
            label = value.strip()
            if not label or label in seen:
                continue
            seen.add(label)
            ordered.append(label)
        return ", ".join(ordered)

    grouped = (
        working.groupby("gang_name")["project_label"]
        .apply(_join_project_codes)
        .to_dict()
    )
    return {str(k): str(v) for k, v in grouped.items() if v}


def _build_idle_days_bucket_tables(
    *,
    productivity_map: dict[str, float],
    idle_days_map: dict[str, int],
    erection_counts: dict[str, int],
    idle_interval_counts: dict[str, int] | None = None,
    daily_weight_map: dict[str, float] | None = None,
    tower_weight_map: dict[str, float] | None = None,
    thresholds: Sequence[int] = (1, 2, 3),
) -> dict[int, pd.DataFrame]:
    buckets = _idle_productivity_buckets()
    metric_label = "Metric"
    columns = [metric_label, *[label for label, _ in buckets]]
    empty = pd.DataFrame(columns=columns)
    if not productivity_map:
        return {threshold: empty.copy() for threshold in thresholds}

    effective_interval_counts = idle_interval_counts or {}
    effective_daily_weight_map = daily_weight_map or {}
    effective_tower_weight_map = tower_weight_map or {}

    row_order = [
        "Days per idle window (severity)",
        "Avg Idle Days (Magnitude)",
        "Idle windows per gang (Frequency)",
        "Gang Count",
        "Weight Erected (MT) - Daily Prod",
        "Weight Erected (MT) - Tower Weight",
    ]

    tables: dict[int, pd.DataFrame] = {}
    for threshold in thresholds:
        rows: dict[str, dict[str, object]] = {label: {} for label in row_order}
        for bucket_label, predicate in buckets:
            gangs = [
                gang
                for gang, productivity in productivity_map.items()
                if productivity is not None
                and not pd.isna(productivity)
                and predicate(productivity)
                and int(erection_counts.get(gang, 0)) >= threshold
            ]
            gang_count = len(gangs)
            idle_days_total = sum(float(idle_days_map.get(gang, 0)) for gang in gangs)
            idle_windows_total = sum(int(effective_interval_counts.get(gang, 0)) for gang in gangs)
            severity_sum = sum(
                (
                    float(idle_days_map.get(gang, 0)) / int(effective_interval_counts.get(gang, 0))
                    if int(effective_interval_counts.get(gang, 0)) > 0
                    else 0.0
                )
                for gang in gangs
            )
            avg_severity = severity_sum / gang_count if gang_count else 0.0
            avg_idle_days = idle_days_total / gang_count if gang_count else 0.0
            avg_idle_windows = idle_windows_total / gang_count if gang_count else 0.0
            weight_daily = sum(float(effective_daily_weight_map.get(gang, 0.0)) for gang in gangs)
            weight_tower = sum(float(effective_tower_weight_map.get(gang, 0.0)) for gang in gangs)

            rows["Days per idle window (severity)"][bucket_label] = round(avg_severity, 2)
            rows["Avg Idle Days (Magnitude)"][bucket_label] = round(avg_idle_days, 2)
            rows["Idle windows per gang (Frequency)"][bucket_label] = round(avg_idle_windows, 2)
            rows["Gang Count"][bucket_label] = int(gang_count)
            rows["Weight Erected (MT) - Daily Prod"][bucket_label] = round(weight_daily, 2)
            rows["Weight Erected (MT) - Tower Weight"][bucket_label] = round(weight_tower, 2)

        table_rows: list[dict[str, object]] = []
        for label in row_order:
            row = {metric_label: label}
            row.update(rows[label])
            table_rows.append(row)
        tables[threshold] = pd.DataFrame(table_rows, columns=columns)

    return tables


def _build_idle_days_bucket_audit(
    *,
    productivity_map: dict[str, float],
    idle_days_map: dict[str, int],
    erection_counts: dict[str, int],
    idle_interval_counts: dict[str, int] | None,
    daily_weight_map: dict[str, float] | None,
    tower_weight_map: dict[str, float] | None,
    completions: pd.DataFrame | None,
    threshold: int = 3,
) -> pd.DataFrame:
    columns = [
        "Gang Name",
        "Productivity (MT/day)",
        "Bucket",
        "Idle Days",
        "Idle Intervals",
        "Idle Days per Interval",
        "Towers Erected",
        "Daily Prod Sum (MT)",
        "Tower Weight Sum (MT)",
        "Project Codes",
    ]
    if not productivity_map:
        return pd.DataFrame(columns=columns)

    buckets = _idle_productivity_buckets()
    idle_intervals = idle_interval_counts or {}
    daily_weights = daily_weight_map or {}
    tower_weights = tower_weight_map or {}
    project_codes = _build_gang_project_code_map(completions)

    rows: list[dict[str, object]] = []
    for gang, productivity in productivity_map.items():
        if productivity is None or pd.isna(productivity):
            continue
        if int(erection_counts.get(gang, 0)) < threshold:
            continue
        bucket_label = _assign_idle_bucket(float(productivity), buckets)
        if not bucket_label:
            continue
        idle_days = float(idle_days_map.get(gang, 0))
        interval_count = int(idle_intervals.get(gang, 0))
        idle_per_interval = idle_days / interval_count if interval_count > 0 else 0.0
        erections = int(erection_counts.get(gang, 0))
        rows.append(
            {
                "Gang Name": gang,
                "Productivity (MT/day)": round(float(productivity), 2),
                "Bucket": bucket_label,
                "Idle Days": round(idle_days, 2),
                "Idle Intervals": int(interval_count),
                "Idle Days per Interval": round(idle_per_interval, 2),
                "Towers Erected": erections,
                "Daily Prod Sum (MT)": round(float(daily_weights.get(gang, 0.0)), 2),
                "Tower Weight Sum (MT)": round(float(tower_weights.get(gang, 0.0)), 2),
                "Project Codes": project_codes.get(gang, ""),
            }
        )

    if not rows:
        return pd.DataFrame(columns=columns)

    audit_df = pd.DataFrame(rows, columns=columns)
    bucket_order = [label for label, _ in buckets]
    audit_df["Bucket"] = pd.Categorical(audit_df["Bucket"], categories=bucket_order, ordered=True)
    audit_df = audit_df.sort_values(["Bucket", "Gang Name"]).reset_index(drop=True)
    audit_df["Bucket"] = audit_df["Bucket"].astype(str)
    return audit_df


def _build_monthly_gang_production_buckets(
    daily_df: pd.DataFrame | None,
    *,
    bucket_min: int = 80,
    bucket_max: int = 260,
    bucket_step: int = 20,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    count_columns = ["Bucket", "Gang-Months", "Sum of Tower Weight (MT)"]
    detail_columns = [
        "Gang Name",
        "Month",
        "Monthly Production (MT)",
        "Monthly Tower Weight (MT)",
        "Project Codes",
        "Towers Erected",
        "Bucket",
    ]
    if daily_df is None or daily_df.empty:
        return pd.DataFrame(columns=count_columns), pd.DataFrame(columns=detail_columns)

    working = daily_df.copy()
    working["date"] = pd.to_datetime(working.get("date"), errors="coerce").dt.normalize()
    working = working.dropna(subset=["date"])
    if working.empty:
        return pd.DataFrame(columns=count_columns), pd.DataFrame(columns=detail_columns)

    working["gang_name"] = working.get("gang_name", "").fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty:
        return pd.DataFrame(columns=count_columns), pd.DataFrame(columns=detail_columns)

    working["daily_prod_mt"] = pd.to_numeric(working.get("daily_prod_mt"), errors="coerce").fillna(0.0)
    working["month"] = working["date"].dt.to_period("M").dt.to_timestamp()
    monthly = (
        working.groupby(["gang_name", "month"], dropna=False)["daily_prod_mt"]
        .sum()
        .reset_index()
        .rename(
            columns={
                "gang_name": "Gang Name",
                "month": "Month",
                "daily_prod_mt": "Monthly Production (MT)",
            }
        )
    )
    if monthly.empty:
        return pd.DataFrame(columns=count_columns), pd.DataFrame(columns=detail_columns)

    monthly["Monthly Production (MT)"] = pd.to_numeric(
        monthly.get("Monthly Production (MT)"), errors="coerce"
    ).fillna(0.0)

    completion_rows = _extract_completion_rows(daily_df)
    if not completion_rows.empty:
        completion_rows["gang_name"] = (
            completion_rows.get("gang_name", "").fillna("").astype(str).str.strip()
        )
        completion_rows = completion_rows[completion_rows["gang_name"].astype(bool)]
        completion_rows["tower_weight"] = pd.to_numeric(
            completion_rows.get("tower_weight"), errors="coerce"
        ).fillna(0.0)
        completion_rows["project_code"] = (
            completion_rows.get("project_code", "").fillna("").astype(str).str.strip()
        )
        completion_rows["project_name"] = (
            completion_rows.get("project_name", "").fillna("").astype(str).str.strip()
        )
        completion_rows["project_label"] = completion_rows["project_code"].where(
            completion_rows["project_code"].astype(bool),
            completion_rows["project_name"],
        )
        completion_rows["month_key"] = (
            pd.to_datetime(completion_rows.get("completion_date"), errors="coerce")
            .dt.to_period("M")
            .dt.to_timestamp()
        )

        def _join_project_codes(series: pd.Series) -> str:
            seen: set[str] = set()
            ordered: list[str] = []
            for value in series.fillna("").astype(str):
                label = value.strip()
                if not label or label in seen:
                    continue
                seen.add(label)
                ordered.append(label)
            return ", ".join(ordered)

        tower_month = (
            completion_rows.groupby(["gang_name", "month_key"], sort=False)
            .agg(
                tower_weight=("tower_weight", "sum"),
                towers_erected=("tower_weight", "size"),
                project_codes=("project_label", _join_project_codes),
            )
            .reset_index()
        )
    else:
        tower_month = pd.DataFrame(
            columns=["gang_name", "month_key", "tower_weight", "towers_erected", "project_codes"]
        )

    monthly = monthly.rename(columns={"Month": "month_key"})
    monthly = monthly.merge(
        tower_month,
        left_on=["Gang Name", "month_key"],
        right_on=["gang_name", "month_key"],
        how="left",
    )
    monthly["Monthly Tower Weight (MT)"] = pd.to_numeric(
        monthly.get("tower_weight"), errors="coerce"
    ).fillna(0.0)
    monthly["Project Codes"] = monthly.get("project_codes", "").fillna("").astype(str)
    monthly["Towers Erected"] = pd.to_numeric(
        monthly.get("towers_erected"), errors="coerce"
    ).fillna(0).astype(int)
    monthly["Month"] = pd.to_datetime(monthly["month_key"], errors="coerce").dt.strftime("%Y-%m")
    monthly["Monthly Production (MT)"] = monthly["Monthly Production (MT)"].round(2)
    monthly["Monthly Tower Weight (MT)"] = monthly["Monthly Tower Weight (MT)"].round(2)
    monthly = monthly.drop(
        columns=[
            col
            for col in (
                "gang_name",
                "tower_weight",
                "month_key",
                "project_codes",
                "towers_erected",
            )
            if col in monthly.columns
        ]
    )

    bucket_edges = list(range(int(bucket_min), int(bucket_max) + int(bucket_step), int(bucket_step)))
    if not bucket_edges:
        return pd.DataFrame(columns=count_columns), pd.DataFrame(columns=detail_columns)
    labels = [f"<{bucket_edges[0]}"]
    labels.extend(
        f"{bucket_edges[idx]}-{bucket_edges[idx + 1]}"
        for idx in range(len(bucket_edges) - 1)
    )
    labels.append(f">={bucket_edges[-1]}")
    bins = [-float("inf"), *bucket_edges, float("inf")]

    bucketed = pd.cut(
        monthly["Monthly Production (MT)"],
        bins=bins,
        labels=labels,
        right=False,
        include_lowest=True,
    )
    bucketed = pd.Categorical(bucketed, categories=labels, ordered=True)
    monthly["Bucket"] = bucketed

    summary = (
        monthly.groupby("Bucket", observed=False)
        .agg(
            **{
                "Gang-Months": ("Bucket", "size"),
                "Sum of Tower Weight (MT)": ("Monthly Tower Weight (MT)", "sum"),
            }
        )
        .reindex(labels)
        .fillna(0.0)
    )
    summary["Gang-Months"] = summary["Gang-Months"].astype(int)
    summary["Sum of Tower Weight (MT)"] = summary["Sum of Tower Weight (MT)"].round(2)
    counts_df = summary.reset_index().rename(columns={"Bucket": "Bucket"})
    monthly["Bucket"] = monthly["Bucket"].astype(str)
    details_df = monthly.reindex(columns=detail_columns)
    return counts_df, details_df


def _pch_sort_components(value: object) -> tuple[int, str]:
    label = str(value or "").strip()
    if label in _PCH_SORT_ORDER:
        return (0, f"{_PCH_SORT_ORDER[label]:03d}")
    lowered = label.lower()
    if not lowered or lowered == "unassigned":
        return (2, lowered)
    return (1, lowered)


def _sort_pch_frame(df: pd.DataFrame, column: str = "PCH") -> pd.DataFrame:
    if df.empty or column not in df.columns:
        return df
    order_components = df[column].map(_pch_sort_components)
    df = df.assign(
        _pch_order_bucket=order_components.map(lambda pair: pair[0]),
        _pch_order_value=order_components.map(lambda pair: pair[1]),
    )
    df = df.sort_values(by=["_pch_order_bucket", "_pch_order_value"]).drop(columns=["_pch_order_bucket", "_pch_order_value"])
    return df.reset_index(drop=True)


def _append_totals_row(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    total_mt = df[TOTAL_MT_COLUMN].sum() if TOTAL_MT_COLUMN in df.columns else 0.0
    total_count = df[TOTAL_COUNT_COLUMN].sum() if TOTAL_COUNT_COLUMN in df.columns else 0
    totals = {column: "" for column in df.columns}
    totals[AVG_PRODUCTIVITY_COLUMN] = "Total"
    totals[TOTAL_MT_COLUMN] = round(float(total_mt), 2)
    totals[TOTAL_COUNT_COLUMN] = int(total_count)
    return pd.concat([df, pd.DataFrame([totals])], ignore_index=True)


def _count_unique_nonempty(series: pd.Series | None) -> int:
    if series is None or not isinstance(series, pd.Series):
        return 0
    cleaned = series.fillna("").astype(str).str.strip()
    return int(cleaned[cleaned.astype(bool)].nunique())


def _resolve_project_series(df: pd.DataFrame | None) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype="object")
    for column in ("project_code", "project_display", "project_name", "project"):
        if column not in df.columns:
            continue
        series = df[column]
        if not isinstance(series, pd.Series):
            continue
        cleaned = series.fillna("").astype(str).str.strip()
        if cleaned[cleaned.astype(bool)].any():
            return series
    return pd.Series(dtype="object")


def _build_scope_label_from_data(
    daily_df: pd.DataFrame | None,
    completions: pd.DataFrame | None,
    *,
    date_start: pd.Timestamp | None = None,
    date_end: pd.Timestamp | None = None,
) -> str:
    date_start_ts = pd.Timestamp(date_start) if date_start is not None else None
    date_end_ts = pd.Timestamp(date_end) if date_end is not None else None
    if daily_df is not None and not daily_df.empty and (date_start_ts is None or date_end_ts is None):
        date_series = pd.to_datetime(daily_df.get("date"), errors="coerce").dropna()
        if not date_series.empty:
            if date_start_ts is None:
                date_start_ts = date_series.min()
            if date_end_ts is None:
                date_end_ts = date_series.max()

    start_label = date_start_ts.strftime("%Y-%m-%d") if date_start_ts is not None else ""
    end_label = date_end_ts.strftime("%Y-%m-%d") if date_end_ts is not None else ""
    project_count = _count_unique_nonempty(_resolve_project_series(daily_df))
    gang_count = _count_unique_nonempty(
        daily_df["gang_name"] if daily_df is not None and "gang_name" in daily_df.columns else None
    )
    erection_count = int(len(completions)) if completions is not None else 0
    return (
        f"Scope: {start_label} to {end_label} | "
        f"Projects: {project_count} | Erections: {erection_count} | Gangs: {gang_count}"
    )


def _build_kv_export_issues(
    daily_df: pd.DataFrame | None,
    month_start: pd.Timestamp,
    month_end: pd.Timestamp,
    voltage_lookup: dict[str, str],
) -> pd.DataFrame:
    if daily_df is None or daily_df.empty:
        return pd.DataFrame([{"Issue": "No data loaded"}])

    working = daily_df.copy()

    date_series = working.get("date")
    if isinstance(date_series, pd.Series):
        working["date"] = pd.to_datetime(date_series, errors="coerce").dt.normalize()
    else:
        working["date"] = pd.NaT

    completion_series = working.get("completion_date")
    if isinstance(completion_series, pd.Series):
        working["completion_date"] = pd.to_datetime(completion_series, errors="coerce").dt.normalize()
    else:
        working["completion_date"] = pd.NaT

    working = _annotate_scope_with_voltage(working, voltage_lookup)
    if working.empty:
        return pd.DataFrame([{"Issue": "No data loaded"}])

    preferred_columns = [
        "date",
        "completion_date",
        "project_code",
        "project_name",
        "project_display",
        "pch_display",
        "location_no",
        "gang_name",
        "tower_type",
        "tower_family",
        "tower_weight",
        "daily_prod_mt",
        "voltage_label",
    ]
    base_columns = [col for col in preferred_columns if col in working.columns]

    date_in_scope = working["date"].notna() & (working["date"] >= month_start) & (working["date"] <= month_end)
    completion_in_scope = working["completion_date"].notna() & (
        (working["completion_date"] >= month_start) & (working["completion_date"] <= month_end)
    )
    excluded_mask = ~(date_in_scope | completion_in_scope)
    if not excluded_mask.any():
        return pd.DataFrame([{"Issue": "No issues detected"}])

    def _build_issue_row(row: pd.Series) -> str:
        parts: list[str] = []
        if pd.isna(row.get("date")):
            parts.append("Invalid or missing date")
        else:
            date_value = row.get("date")
            if isinstance(date_value, pd.Timestamp) and (date_value < month_start or date_value > month_end):
                parts.append("Date outside selected range")
        if pd.isna(row.get("completion_date")):
            parts.append("Missing completion date")
        else:
            comp_value = row.get("completion_date")
            if isinstance(comp_value, pd.Timestamp) and (comp_value < month_start or comp_value > month_end):
                parts.append("Completion date outside selected range")
        return "; ".join(parts) if parts else "Excluded from calculations"

    subset = working.loc[excluded_mask, base_columns].copy()
    subset["Issue"] = subset.apply(_build_issue_row, axis=1)
    subset["Excluded From"] = "Avg Productivity + Totals"
    ordered = ["Issue", "Excluded From", *base_columns]
    return subset.reindex(columns=ordered)


def _build_group_productivity_table(
    scope: pd.DataFrame,
    completions: pd.DataFrame,
    week_labels: Sequence[str],
    *,
    group_columns: Sequence[str],
    column_labels: dict[str, str],
    include_weekly: bool = True,
    metric_labels: dict[str, str] | None = None,
) -> pd.DataFrame:
    metric_labels = metric_labels or {}
    labeled_columns = [column_labels.get(col, col) for col in group_columns]
    week_columns = list(week_labels) if include_weekly else []
    avg_label = metric_labels.get(AVG_PRODUCTIVITY_COLUMN, AVG_PRODUCTIVITY_COLUMN)
    total_mt_label = metric_labels.get(TOTAL_MT_COLUMN, TOTAL_MT_COLUMN)
    total_count_label = metric_labels.get(TOTAL_COUNT_COLUMN, TOTAL_COUNT_COLUMN)
    ordered_columns = [*labeled_columns, *week_columns, avg_label, total_mt_label, total_count_label]

    if scope is None or scope.empty:
        return pd.DataFrame(columns=ordered_columns)

    grouping = list(group_columns)
    summary = (
        scope.groupby(grouping, dropna=False)["daily_prod_mt"]
        .mean()
        .reset_index()
        .rename(columns={"daily_prod_mt": AVG_PRODUCTIVITY_COLUMN})
    )

    if include_weekly:
        for idx, label in enumerate(week_labels, start=1):
            week_scope = scope[scope.get("week_index") == idx]
            if week_scope.empty:
                summary[label] = 0.0
                continue
            week_avg = (
                week_scope.groupby(grouping, dropna=False)["daily_prod_mt"]
                .mean()
                .reset_index()
                .rename(columns={"daily_prod_mt": label})
            )
            summary = summary.merge(week_avg, on=grouping, how="left")
            summary[label] = summary[label].fillna(0.0)

    weights = None
    counts = None
    if isinstance(completions, pd.DataFrame) and not completions.empty:
        tower_series = pd.to_numeric(completions.get("_tower_weight"), errors="coerce")
        weights = (
            completions.assign(_tower_weight=tower_series.fillna(0.0))
            .groupby(grouping, dropna=False)["_tower_weight"]
            .sum()
            .reset_index()
            .rename(columns={"_tower_weight": TOTAL_MT_COLUMN})
        )
        counts = (
            completions.groupby(grouping, dropna=False)
            .size()
            .reset_index(name=TOTAL_COUNT_COLUMN)
        )

    if weights is not None:
        summary = summary.merge(weights, on=grouping, how="left")
    else:
        summary[TOTAL_MT_COLUMN] = 0.0
    if counts is not None:
        summary = summary.merge(counts, on=grouping, how="left")
    else:
        summary[TOTAL_COUNT_COLUMN] = 0

    summary[TOTAL_MT_COLUMN] = summary[TOTAL_MT_COLUMN].fillna(0.0)
    summary[TOTAL_COUNT_COLUMN] = summary[TOTAL_COUNT_COLUMN].fillna(0).astype(int)

    for column in [AVG_PRODUCTIVITY_COLUMN, *week_columns, TOTAL_MT_COLUMN]:
        summary[column] = summary[column].fillna(0.0).round(2)

    summary = summary.rename(columns=column_labels)
    if metric_labels:
        summary = summary.rename(columns=metric_labels)
    summary = summary.sort_values(labeled_columns).reset_index(drop=True)
    return summary.reindex(columns=ordered_columns)


def _build_gang_weekly_productivity_table(scope: pd.DataFrame, week_labels: Sequence[str]) -> pd.DataFrame:
    columns = ["Gang Name", *week_labels, "Avg Productivity (MT/day)", "Erections Completed"]
    if scope is None or scope.empty or "gang_name" not in scope.columns:
        return pd.DataFrame(columns=columns)

    working = scope.copy()
    working["gang_name"] = working["gang_name"].fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty:
        return pd.DataFrame(columns=columns)

    working["derived_productivity"] = _coerce_numeric_series(working, "derived_productivity")
    working["week_index"] = _coerce_numeric_series(working, "week_index").fillna(0).astype(int)

    weekly = (
        working.groupby(["gang_name", "week_index"])["derived_productivity"]
        .mean()
        .reset_index()
    )
    if not weekly.empty:
        weekly["week_label"] = weekly["week_index"].map(
            lambda idx: week_labels[idx - 1] if 1 <= idx <= len(week_labels) else None
        )
        weekly = weekly.dropna(subset=["week_label"])
        pivot = (
            weekly.pivot_table(index="gang_name", columns="week_label", values="derived_productivity", aggfunc="mean")
            .reset_index()
            if not weekly.empty
            else pd.DataFrame(columns=["gang_name"])
        )
    else:
        pivot = pd.DataFrame(columns=["gang_name"])

    def _count_unique_locations(series: pd.Series) -> int:
        normalized = series.fillna("").astype(str).str.strip()
        normalized = normalized.replace("", pd.NA).dropna()
        return int(normalized.nunique())

    overall = (
        working.groupby("gang_name")
        .agg(
            overall_avg=("derived_productivity", lambda values: values.dropna().mean()),
            completions=("location_no", _count_unique_locations),
        )
        .reset_index()
    )
    summary = overall.merge(pivot, on="gang_name", how="left")
    summary = summary.rename(
        columns={
            "gang_name": "Gang Name",
            "overall_avg": "Avg Productivity (MT/day)",
            "completions": "Erections Completed",
        }
    )

    for label in week_labels:
        if label not in summary.columns:
            summary[label] = 0.0
    summary["Avg Productivity (MT/day)"] = summary["Avg Productivity (MT/day)"].fillna(0.0).round(2)
    for label in week_labels:
        summary[label] = summary[label].fillna(0.0).round(2)
    summary["Erections Completed"] = summary["Erections Completed"].fillna(0).astype(int)

    ordered_columns = ["Gang Name", *week_labels, "Avg Productivity (MT/day)", "Erections Completed"]
    summary = summary.reindex(columns=ordered_columns)
    if summary.empty:
        return summary
    return summary.sort_values("Gang Name").reset_index(drop=True)


def _suppress_repeated_pch(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "PCH" not in df.columns:
        return df
    rows: list[pd.Series] = []
    current = None
    for _, row in df.iterrows():
        row_copy = row.copy()
        if current == row_copy["PCH"]:
            row_copy["PCH"] = ""
        else:
            current = row_copy["PCH"]
        rows.append(row_copy)
    return pd.DataFrame(rows, columns=df.columns)


def export_erection_productivity_summary(
    output_path: str | Path,
    *,
    data_store: AppDataStore | None = None,
    daily_df: pd.DataFrame | None = None,
    project_info: pd.DataFrame | None = None,
    as_of_date: pd.Timestamp | str | None = None,
    range_start: pd.Timestamp | str | None = None,
    range_end: pd.Timestamp | str | None = None,
    gang_min_erections: int = 4,
    sheet_name: str = _DEFAULT_SHEET_NAME,
    blank_rows_between_tables: int = 3,
) -> Path:
    """
    Uses existing dashboard data to build and save the PCH and project-wise erection
    productivity summary Excel. Returns the path to the saved file.
    """

    source_daily = daily_df
    source_project_info = project_info
    candidate_date = pd.Timestamp(as_of_date) if as_of_date is not None else None
    range_start_ts = pd.Timestamp(range_start) if range_start is not None else None
    range_end_ts = pd.Timestamp(range_end) if range_end is not None else None

    if data_store is not None:
        if source_daily is None:
            source_daily = data_store.get_daily()
        if source_project_info is None:
            source_project_info = data_store.get_project_info()
        if candidate_date is None or pd.isna(candidate_date):
            candidate_date = data_store.metadata.last_data_date

    if source_daily is None or source_daily.empty:
        raise ValueError("No erection daily dataframe is available to export.")

    if range_start_ts is not None or range_end_ts is not None:
        if range_start_ts is None or range_end_ts is None:
            raise ValueError("Both range_start and range_end are required when selecting a month range.")
        range_start_ts = range_start_ts.normalize().to_period("M").to_timestamp()
        range_end_ts = range_end_ts.normalize().to_period("M").to_timestamp() + pd.offsets.MonthEnd(1)
        if range_start_ts > range_end_ts:
            raise ValueError("range_start must be before or equal to range_end.")
        month_start = range_start_ts
        month_end = range_end_ts
        active_date = month_end
    else:
        if candidate_date is None or pd.isna(candidate_date):
            date_series = pd.to_datetime(source_daily.get("date"), errors="coerce").dropna()
            if date_series.empty:
                raise ValueError("Unable to determine the current month for the erection summary export.")
            candidate_date = date_series.max()
        active_date = pd.Timestamp(candidate_date).normalize()
        month_start = active_date.to_period("M").to_timestamp()
        month_end = (month_start + pd.offsets.MonthEnd(1)).normalize()

    quarter_mode = False
    if range_start_ts is not None and range_end_ts is not None:
        start_period = range_start_ts.to_period("M")
        end_period = range_end_ts.to_period("M")
        months_count = (end_period.year - start_period.year) * 12 + (end_period.month - start_period.month) + 1
        quarter_mode = months_count == 3
    else:
        months_count = 1

    if quarter_mode:
        week_labels = _generate_quarter_week_labels(month_start, month_end)
        week_index = _assign_quarter_week_bucket(pd.Series([active_date]), month_start, months_count=months_count)
        current_week_idx = int(week_index.iloc[0]) if not week_index.empty else 1
        current_week_idx = max(1, min(len(week_labels), current_week_idx))
    else:
        week_labels = _generate_week_labels(month_start, month_end)
        offset_days = max(0, (active_date - month_start).days)
        week_count = len(week_labels) or 1
        current_week_idx = min(week_count, max(1, (offset_days // 7) + 1))
    current_week_label = week_labels[current_week_idx - 1]

    project_info_frame = source_project_info.copy() if isinstance(source_project_info, pd.DataFrame) else pd.DataFrame()

    scope = _prepare_month_scope(
        source_daily, project_info_frame, month_start, month_end, week_labels, filter_by_date=True
    )
    completion_base = _prepare_month_scope(
        source_daily, project_info_frame, month_start, month_end, week_labels, filter_by_date=False
    )
    completions = _prepare_completion_scope(completion_base, month_start, month_end, week_labels)
    if quarter_mode:
        scope["week_index"] = _assign_quarter_week_bucket(scope["date"], month_start, months_count=months_count)
        if not completions.empty:
            completions["completion_week_index"] = _assign_quarter_week_bucket(
                completions["completion_date"], month_start, months_count=months_count
            )

    overall_summary, pch_summary, project_summary = _build_productivity_tables(
        scope,
        completions,
        week_labels=week_labels,
    )
    gang_level_summary = _build_gang_level_productivity_table(scope, completions)
    top_gangs, bottom_gangs = _build_project_gang_rankings(
        scope,
        completions,
        min_erections=int(gang_min_erections) if gang_min_erections is not None else 4,
    )
    overall_gang_productivity = _build_overall_gang_productivity_map(source_daily)
    month_gang_productivity = _build_gang_productivity_map(scope)
    active_config = data_store._config if data_store is not None else AppConfig()
    idle_intervals = compute_idle_intervals_per_gang(
        scope,
        loss_max_gap_days=active_config.loss_max_gap_days,
    )
    overall_idle_source = source_daily.copy()
    overall_idle_source["date"] = pd.to_datetime(
        overall_idle_source.get("date"), errors="coerce"
    ).dt.normalize()
    overall_idle_source = overall_idle_source.dropna(subset=["date"])
    overall_idle_source["gang_name"] = (
        overall_idle_source.get("gang_name", "").fillna("").astype(str).str.strip()
    )
    overall_idle_source = overall_idle_source[overall_idle_source["gang_name"].astype(bool)]
    overall_idle_source["daily_prod_mt"] = pd.to_numeric(
        overall_idle_source.get("daily_prod_mt"), errors="coerce"
    ).fillna(0.0)
    overall_idle_intervals = compute_idle_intervals_per_gang(
        overall_idle_source,
        loss_max_gap_days=active_config.loss_max_gap_days,
    )
    idle_days_by_gang: dict[str, int] = {}
    idle_interval_counts_by_gang: dict[str, int] = {}
    if not idle_intervals.empty and "gang_name" in idle_intervals.columns:
        idle_days_by_gang = (
            idle_intervals.groupby("gang_name")["idle_days_capped"].sum().fillna(0).astype(int).to_dict()
        )
        idle_interval_counts_by_gang = (
            idle_intervals.groupby("gang_name").size().fillna(0).astype(int).to_dict()
        )
    overall_idle_days_by_gang: dict[str, int] = {}
    overall_idle_interval_counts_by_gang: dict[str, int] = {}
    if not overall_idle_intervals.empty and "gang_name" in overall_idle_intervals.columns:
        overall_idle_days_by_gang = (
            overall_idle_intervals.groupby("gang_name")["idle_days_capped"].sum().fillna(0).astype(int).to_dict()
        )
        overall_idle_interval_counts_by_gang = (
            overall_idle_intervals.groupby("gang_name").size().fillna(0).astype(int).to_dict()
        )

    if AVG_PRODUCTIVITY_OVERALL_COLUMN not in gang_level_summary.columns:
        gang_level_summary[AVG_PRODUCTIVITY_OVERALL_COLUMN] = pd.Series(
            index=gang_level_summary.index, dtype="float64"
        )
    if IDLE_DAYS_COLUMN not in gang_level_summary.columns:
        gang_level_summary[IDLE_DAYS_COLUMN] = pd.Series(index=gang_level_summary.index, dtype="int64")
    if IDLE_INTERVALS_COLUMN not in gang_level_summary.columns:
        gang_level_summary[IDLE_INTERVALS_COLUMN] = pd.Series(index=gang_level_summary.index, dtype="int64")
    if "Gang Name" in gang_level_summary.columns and not gang_level_summary.empty:
        gang_names = gang_level_summary["Gang Name"].astype(str).str.strip()
        gang_level_summary[AVG_PRODUCTIVITY_OVERALL_COLUMN] = (
            gang_names.map(overall_gang_productivity).fillna(0.0).round(2)
        )
        gang_level_summary[IDLE_DAYS_COLUMN] = gang_names.map(idle_days_by_gang).fillna(0).astype(int)
        gang_level_summary[IDLE_INTERVALS_COLUMN] = (
            gang_names.map(idle_interval_counts_by_gang).fillna(0).astype(int)
        )

    gang_level_columns = [
        "PCH",
        "Project",
        "Gang Name",
        AVG_PRODUCTIVITY_COLUMN,
        AVG_PRODUCTIVITY_OVERALL_COLUMN,
        TOTAL_MT_COLUMN,
        TOTAL_COUNT_COLUMN,
        AVG_TOWER_WEIGHT_OVERALL_COLUMN,
        IDLE_DAYS_COLUMN,
        IDLE_INTERVALS_COLUMN,
    ]
    gang_level_summary = gang_level_summary.reindex(
        columns=[column for column in gang_level_columns if column in gang_level_summary.columns]
    )

    def _apply_gang_metrics(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            df[AVG_PRODUCTIVITY_OVERALL_COLUMN] = pd.Series(dtype="float64")
            df[IDLE_DAYS_COLUMN] = pd.Series(dtype="int64")
            df[IDLE_INTERVALS_COLUMN] = pd.Series(dtype="int64")
            return df
        if "Gang Name" not in df.columns:
            df[AVG_PRODUCTIVITY_OVERALL_COLUMN] = 0.0
            df[IDLE_DAYS_COLUMN] = 0
            df[IDLE_INTERVALS_COLUMN] = 0
            return df
        names = df["Gang Name"].astype(str).str.strip()
        df[AVG_PRODUCTIVITY_OVERALL_COLUMN] = (
            names.map(overall_gang_productivity).fillna(0.0).round(2)
        )
        df[IDLE_DAYS_COLUMN] = names.map(idle_days_by_gang).fillna(0).astype(int)
        df[IDLE_INTERVALS_COLUMN] = names.map(idle_interval_counts_by_gang).fillna(0).astype(int)
        return df

    ranking_columns = [
        "PCH",
        "Project",
        "Gang Name",
        AVG_PRODUCTIVITY_COLUMN,
        AVG_PRODUCTIVITY_OVERALL_COLUMN,
        TOTAL_MT_COLUMN,
        TOTAL_COUNT_COLUMN,
        AVG_TOWER_WEIGHT_OVERALL_COLUMN,
        IDLE_DAYS_COLUMN,
        IDLE_INTERVALS_COLUMN,
    ]
    top_gangs = _apply_gang_metrics(top_gangs)
    bottom_gangs = _apply_gang_metrics(bottom_gangs)
    top_gangs = top_gangs.reindex(columns=ranking_columns)
    bottom_gangs = bottom_gangs.reindex(columns=ranking_columns)
    for column in (IDLE_DAYS_COLUMN, IDLE_INTERVALS_COLUMN):
        top_gangs[column] = pd.to_numeric(top_gangs[column], errors="coerce").fillna(0).astype(int)
        bottom_gangs[column] = pd.to_numeric(bottom_gangs[column], errors="coerce").fillna(0).astype(int)
    if AVG_PRODUCTIVITY_OVERALL_COLUMN in top_gangs.columns:
        top_gangs[AVG_PRODUCTIVITY_OVERALL_COLUMN] = (
            pd.to_numeric(top_gangs[AVG_PRODUCTIVITY_OVERALL_COLUMN], errors="coerce").fillna(0.0).round(2)
        )
    if AVG_PRODUCTIVITY_OVERALL_COLUMN in bottom_gangs.columns:
        bottom_gangs[AVG_PRODUCTIVITY_OVERALL_COLUMN] = (
            pd.to_numeric(bottom_gangs[AVG_PRODUCTIVITY_OVERALL_COLUMN], errors="coerce").fillna(0.0).round(2)
        )
    if AVG_TOWER_WEIGHT_OVERALL_COLUMN in top_gangs.columns:
        top_gangs[AVG_TOWER_WEIGHT_OVERALL_COLUMN] = (
            pd.to_numeric(top_gangs[AVG_TOWER_WEIGHT_OVERALL_COLUMN], errors="coerce").fillna(0.0).round(2)
        )
    if AVG_TOWER_WEIGHT_OVERALL_COLUMN in bottom_gangs.columns:
        bottom_gangs[AVG_TOWER_WEIGHT_OVERALL_COLUMN] = (
            pd.to_numeric(bottom_gangs[AVG_TOWER_WEIGHT_OVERALL_COLUMN], errors="coerce").fillna(0.0).round(2)
        )

    mtd_erection_counts = _build_completion_counts_by_gang(completions)
    overall_completions = _extract_completion_rows(source_daily)
    overall_erection_counts = _build_completion_counts_by_gang(overall_completions)
    mtd_total_mt_by_gang = _build_gang_total_map(scope)
    overall_total_mt_by_gang = _build_gang_total_map(overall_idle_source)
    mtd_tower_weight_by_gang = _build_gang_total_map(completions, metric_column="tower_weight")
    overall_tower_weight_by_gang = _build_gang_total_map(overall_completions, metric_column="tower_weight")
    mtd_idle_bucket_tables = _build_idle_days_bucket_tables(
        productivity_map=month_gang_productivity,
        idle_days_map=idle_days_by_gang,
        erection_counts=mtd_erection_counts,
        idle_interval_counts=idle_interval_counts_by_gang,
        daily_weight_map=mtd_total_mt_by_gang,
        tower_weight_map=mtd_tower_weight_by_gang,
    )
    overall_idle_bucket_tables = _build_idle_days_bucket_tables(
        productivity_map=overall_gang_productivity,
        idle_days_map=overall_idle_days_by_gang,
        erection_counts=overall_erection_counts,
        idle_interval_counts=overall_idle_interval_counts_by_gang,
        daily_weight_map=overall_total_mt_by_gang,
        tower_weight_map=overall_tower_weight_by_gang,
    )
    mtd_idle_bucket_audit = _build_idle_days_bucket_audit(
        productivity_map=month_gang_productivity,
        idle_days_map=idle_days_by_gang,
        erection_counts=mtd_erection_counts,
        idle_interval_counts=idle_interval_counts_by_gang,
        daily_weight_map=mtd_total_mt_by_gang,
        tower_weight_map=mtd_tower_weight_by_gang,
        completions=completions,
        threshold=3,
    )
    overall_idle_bucket_audit = _build_idle_days_bucket_audit(
        productivity_map=overall_gang_productivity,
        idle_days_map=overall_idle_days_by_gang,
        erection_counts=overall_erection_counts,
        idle_interval_counts=overall_idle_interval_counts_by_gang,
        daily_weight_map=overall_total_mt_by_gang,
        tower_weight_map=overall_tower_weight_by_gang,
        completions=overall_completions,
        threshold=3,
    )
    monthly_bucket_counts, monthly_bucket_details = _build_monthly_gang_production_buckets(source_daily)
    excluded_summary, excluded_detail = _build_excluded_data_audit_tables(
        completion_base,
        month_start,
        month_end,
    )

    base_sheet_name = str(sheet_name or _DEFAULT_SHEET_NAME).strip() or _DEFAULT_SHEET_NAME
    weekly_sheet_label = _sanitize_sheet_name(f"{base_sheet_name} Weekly")
    monthly_summary_sheet_label = _sanitize_sheet_name(f"{base_sheet_name} Monthly")
    if not weekly_sheet_label:
        weekly_sheet_label = _sanitize_sheet_name(f"{_DEFAULT_SHEET_NAME} Weekly")
    if not monthly_summary_sheet_label:
        monthly_summary_sheet_label = _sanitize_sheet_name(f"{_DEFAULT_SHEET_NAME} Monthly")
    if monthly_summary_sheet_label == weekly_sheet_label:
        monthly_summary_sheet_label = _sanitize_sheet_name(f"{base_sheet_name} Monthwise")
    review_sheet_label = _sanitize_sheet_name("Review_Format")
    gangs_sheet_label = _sanitize_sheet_name("Project Gang Rankings")
    gang_level_sheet_label = _sanitize_sheet_name("Gang Level Productivity")
    idle_sheet_label = _sanitize_sheet_name("Idle Intervals")
    idle_bucket_sheet_label = _sanitize_sheet_name("Idle Days Buckets")
    monthly_bucket_sheet_label = _sanitize_sheet_name("Monthly Gang Buckets")
    excluded_sheet_label = _sanitize_sheet_name("Excluded Data Audit")
    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    month_cursor = month_start.to_period("M").to_timestamp()
    month_end_cursor = month_end.to_period("M").to_timestamp()
    months: list[pd.Timestamp] = []
    while month_cursor <= month_end_cursor:
        months.append(month_cursor)
        month_cursor = (month_cursor + pd.offsets.MonthBegin(1)).normalize()
    month_labels = [month.strftime("%b'%y") for month in months]
    (
        monthly_overall_summary,
        monthly_pch_summary,
        monthly_project_summary,
        monthly_summary_labels,
    ) = _build_monthly_productivity_tables(
        scope,
        completions,
        months=months,
    )
    scope_label = _build_scope_label_from_data(
        scope,
        completions,
        date_start=month_start,
        date_end=month_end,
    )
    overall_scope_label = _build_scope_label_from_data(
        source_daily,
        overall_completions,
    )

    def _build_review_table_two(scoped: pd.DataFrame, *, label_col: str = "Erection productivity") -> pd.DataFrame:
        weight_row: dict[str, object] = {label_col: "Weight/day (Avg)"}
        gang_row: dict[str, object] = {label_col: "No of Gangs engaged (Avg)"}
        if scoped.empty:
            for label in month_labels:
                weight_row[label] = 0.0
                gang_row[label] = 0
            return pd.DataFrame([weight_row, gang_row], columns=[label_col, *month_labels])

        for month_start_ts, label in zip(months, month_labels):
            month_end_ts = (month_start_ts + pd.offsets.MonthEnd(1)).normalize()
            month_scope = scoped[(scoped["date"] >= month_start_ts) & (scoped["date"] <= month_end_ts)]
            if month_scope.empty:
                weight_row[label] = 0.0
                gang_row[label] = 0
                continue

            weight_series = pd.to_numeric(month_scope.get("daily_prod_mt"), errors="coerce").dropna()
            weight_avg = weight_series.mean() if not weight_series.empty else 0.0

            gangs_scope = month_scope.copy()
            if "gang_name" in gangs_scope.columns:
                gangs_scope["gang_name"] = gangs_scope["gang_name"].fillna("").astype(str).str.strip()
                gangs_scope = gangs_scope[gangs_scope["gang_name"].astype(bool)]
            else:
                gangs_scope = gangs_scope.iloc[0:0]
            gangs_unique = gangs_scope["gang_name"].nunique() if not gangs_scope.empty else 0

            weight_row[label] = round(float(weight_avg), 2)
            gang_row[label] = int(gangs_unique)

        return pd.DataFrame([weight_row, gang_row], columns=[label_col, *month_labels])

    review_columns = [
        "PCH",
        "Project",
        *week_labels,
        AVG_PRODUCTIVITY_COLUMN,
        TOTAL_MT_COLUMN,
        TOTAL_COUNT_COLUMN,
    ]
    review_table_raw = project_summary.reindex(columns=review_columns)
    if not review_table_raw.empty and AVG_PRODUCTIVITY_COLUMN in review_table_raw.columns:
        review_table_raw = review_table_raw[review_table_raw[AVG_PRODUCTIVITY_COLUMN] != "Total"]

    review_table_one_groups: list[pd.DataFrame] = []
    if review_table_raw.empty or "PCH" not in review_table_raw.columns:
        review_table_one_groups.append(_append_totals_row(review_table_raw))
    else:
        unique_pch = sorted(review_table_raw["PCH"].fillna("").astype(str).unique(), key=_pch_sort_components)
        for pch_name in unique_pch:
            if not pch_name:
                continue
            pch_table = review_table_raw[review_table_raw["PCH"] == pch_name].copy()
            if pch_table.empty:
                continue
            header_row: dict[str, object] = {"PCH": f"PCH: {pch_name}"}
            for column in review_columns:
                if column != "PCH":
                    header_row[column] = ""
            review_table_one_groups.append(pd.DataFrame([header_row], columns=review_columns))
            review_table_one_groups.append(_append_totals_row(pch_table))

    review_table_two_groups: list[pd.DataFrame] = []
    if scope.empty or "pch_display" not in scope.columns:
        review_table_two_groups.append(_build_review_table_two(scope))
    else:
        unique_pch = sorted(scope["pch_display"].fillna("").astype(str).unique(), key=_pch_sort_components)
        for pch_name in unique_pch:
            if not pch_name:
                continue
            pch_scope = scope[scope["pch_display"] == pch_name]
            header_row = {"Erection productivity": f"PCH: {pch_name}"}
            for label in month_labels:
                header_row[label] = ""
            review_table_two_groups.append(
                pd.DataFrame([header_row], columns=["Erection productivity", *month_labels])
            )
            review_table_two_groups.append(_build_review_table_two(pch_scope))

    idle_export_columns = {
        "gang_name": "Gang",
        "interval_start": "Interval Start",
        "interval_end": "Interval End",
        "raw_gap_days": "Raw Gap (days)",
        "idle_days_capped": IDLE_DAYS_COLUMN,
    }
    if idle_intervals.empty:
        idle_export = pd.DataFrame(columns=list(idle_export_columns.values()))
    else:
        idle_export = idle_intervals.rename(columns=idle_export_columns).copy()
        idle_export = idle_export.reindex(columns=list(idle_export_columns.values()))
        for date_column in ("Interval Start", "Interval End"):
            if date_column in idle_export.columns:
                idle_export[date_column] = pd.to_datetime(
                    idle_export[date_column], errors="coerce"
                ).dt.strftime("%d-%m-%Y")
        if "Raw Gap (days)" in idle_export.columns:
            idle_export["Raw Gap (days)"] = pd.to_numeric(
                idle_export["Raw Gap (days)"], errors="coerce"
            ).fillna(0).astype(int)
        if IDLE_DAYS_COLUMN in idle_export.columns:
            idle_export[IDLE_DAYS_COLUMN] = pd.to_numeric(
                idle_export[IDLE_DAYS_COLUMN], errors="coerce"
            ).fillna(0).astype(int)

    def _unique_sheet_name(base: str, used: set[str]) -> str:
        sanitized = _sanitize_sheet_name(base) or "Sheet"
        candidate = sanitized
        suffix = 2
        while candidate in used:
            trimmed = sanitized[: max(0, 31 - len(str(suffix)) - 1)]
            candidate = f"{trimmed}-{suffix}"
            suffix += 1
        used.add(candidate)
        return candidate

    def _write_scope_row(
        writer: pd.ExcelWriter,
        sheet: str,
        label: str,
        *,
        startrow: int,
        startcol: int = 0,
    ) -> int:
        pd.DataFrame([[label]]).to_excel(
            writer, sheet_name=sheet, index=False, header=False, startrow=startrow, startcol=startcol
        )
        return startrow + 1

    def _write_scoped_table(
        writer: pd.ExcelWriter,
        sheet: str,
        table: pd.DataFrame,
        label: str,
        *,
        startrow: int,
        startcol: int = 0,
    ) -> int:
        next_row = _write_scope_row(writer, sheet, label, startrow=startrow, startcol=startcol)
        table.to_excel(writer, sheet_name=sheet, index=False, startrow=next_row, startcol=startcol)
        return next_row + len(table) + 1

    def _write_monthly_scoped_table(
        writer: pd.ExcelWriter,
        sheet: str,
        table: pd.DataFrame,
        label: str,
        *,
        month_headers: Sequence[str],
        group_headers: Sequence[str],
        startrow: int,
        startcol: int = 0,
    ) -> tuple[int, dict[str, int]]:
        next_row = _write_scope_row(writer, sheet, label, startrow=startrow, startcol=startcol)
        ordered_columns, top_headers, bottom_headers = _monthly_column_layout(
            month_headers,
            group_labels=group_headers,
        )
        aligned = table.reindex(columns=ordered_columns)
        pd.DataFrame([top_headers, bottom_headers], columns=ordered_columns).to_excel(
            writer,
            sheet_name=sheet,
            index=False,
            header=False,
            startrow=next_row,
            startcol=startcol,
        )
        data_row = next_row + 2
        aligned.to_excel(
            writer,
            sheet_name=sheet,
            index=False,
            header=False,
            startrow=data_row,
            startcol=startcol,
        )
        next_table_row = data_row + len(aligned) + 1
        block = {
            "scope_row": startrow,
            "header_top_row": next_row,
            "header_bottom_row": next_row + 1,
            "data_start_row": data_row,
            "data_end_row": data_row + max(len(aligned) - 1, 0),
            "group_count": len(group_headers),
            "month_count": len(month_headers),
            "column_count": len(ordered_columns),
            "start_col": startcol,
        }
        return next_table_row, block

    def _style_monthly_summary_sheet(
        writer: pd.ExcelWriter,
        sheet: str,
        blocks: Sequence[dict[str, int]],
    ) -> None:
        worksheet = writer.sheets.get(sheet)
        if worksheet is None or not blocks:
            return

        thin = Side(style="thin", color="D9E2EC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        scope_font = Font(bold=True, color="334155")
        header_font = Font(bold=True, color="1F2937")
        data_font = Font(color="111827")
        scope_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
        top_header_fill = PatternFill(fill_type="solid", fgColor="EAF1F8")
        bottom_header_fill = PatternFill(fill_type="solid", fgColor="F5F8FC")
        zebra_fill = PatternFill(fill_type="solid", fgColor="FAFCFF")
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        max_columns = max(block["column_count"] + block["start_col"] for block in blocks)
        for col_idx in range(1, max_columns + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1:
                worksheet.column_dimensions[col_letter].width = 30
            elif col_idx == 2:
                worksheet.column_dimensions[col_letter].width = 20
            else:
                worksheet.column_dimensions[col_letter].width = 14

        for block in blocks:
            col_offset = block["start_col"]
            first_col = col_offset + 1
            last_col = col_offset + block["column_count"]
            scope_row = block["scope_row"] + 1
            top_row = block["header_top_row"] + 1
            bottom_row = block["header_bottom_row"] + 1
            data_start = block["data_start_row"] + 1
            data_end = block["data_end_row"] + 1
            group_count = block["group_count"]

            scope_cell = worksheet.cell(row=scope_row, column=first_col)
            scope_cell.font = scope_font
            scope_cell.fill = scope_fill
            scope_cell.alignment = left

            for col in range(first_col, last_col + 1):
                top_cell = worksheet.cell(row=top_row, column=col)
                top_cell.font = header_font
                top_cell.fill = top_header_fill
                top_cell.alignment = center
                top_cell.border = border

                bottom_cell = worksheet.cell(row=bottom_row, column=col)
                bottom_cell.font = header_font
                bottom_cell.fill = bottom_header_fill
                bottom_cell.alignment = center
                bottom_cell.border = border

            metric_span = 3
            month_block_count = block["month_count"] + 1  # includes Overall
            month_start_col = first_col + group_count
            for idx in range(month_block_count):
                span_start = month_start_col + (idx * metric_span)
                span_end = span_start + metric_span - 1
                worksheet.merge_cells(
                    start_row=top_row,
                    start_column=span_start,
                    end_row=top_row,
                    end_column=span_end,
                )

            if group_count > 0:
                for offset in range(group_count):
                    col = first_col + offset
                    label_value = worksheet.cell(row=bottom_row, column=col).value
                    worksheet.merge_cells(
                        start_row=top_row,
                        start_column=col,
                        end_row=bottom_row,
                        end_column=col,
                    )
                    merged_cell = worksheet.cell(row=top_row, column=col)
                    merged_cell.value = label_value
                    merged_cell.font = header_font
                    merged_cell.fill = top_header_fill
                    merged_cell.alignment = center
                    merged_cell.border = border

            if data_end < data_start:
                continue
            for row in range(data_start, data_end + 1):
                apply_zebra = (row - data_start) % 2 == 1
                for col in range(first_col, last_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = border
                    is_group_col = col < (first_col + group_count)
                    cell.alignment = left if is_group_col else center
                    if apply_zebra:
                        cell.fill = zebra_fill
                    if not is_group_col:
                        metric_idx = (col - (first_col + group_count)) % metric_span
                        if metric_idx == 1:
                            cell.number_format = "#,##0"
                        else:
                            cell.number_format = "#,##0.00"

            worksheet.row_dimensions[top_row].height = 22
            worksheet.row_dimensions[bottom_row].height = 22

    def _style_clean_sheet(writer: pd.ExcelWriter, sheet: str) -> None:
        worksheet = writer.sheets.get(sheet)
        if worksheet is None:
            return

        thin = Side(style="thin", color="D9E2EC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        label_font = Font(bold=True, color="334155")
        header_font = Font(bold=True, color="1F2937")
        data_font = Font(color="111827")
        label_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
        header_fill = PatternFill(fill_type="solid", fgColor="F3F7FB")
        zebra_fill = PatternFill(fill_type="solid", fgColor="FAFCFF")
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        max_row = worksheet.max_row or 0
        max_col = worksheet.max_column or 0
        if max_row == 0 or max_col == 0:
            return

        column_widths: dict[int, int] = {}
        data_row_band = 0

        for row_idx in range(1, max_row + 1):
            row_values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
            non_empty_pairs = [
                (idx + 1, value)
                for idx, value in enumerate(row_values)
                if value is not None and str(value).strip() != ""
            ]
            if not non_empty_pairs:
                continue

            non_empty_values = [value for _, value in non_empty_pairs]
            string_count = sum(1 for value in non_empty_values if isinstance(value, str))
            numeric_count = sum(
                1 for value in non_empty_values if isinstance(value, (int, float)) and not isinstance(value, bool)
            )
            is_label_row = len(non_empty_pairs) == 1 and string_count == 1
            is_header_row = (
                (not is_label_row)
                and len(non_empty_pairs) >= 2
                and string_count >= max(2, int(len(non_empty_pairs) * 0.6))
                and numeric_count <= max(2, int(len(non_empty_pairs) * 0.4))
            )
            is_data_row = not is_label_row and not is_header_row
            if is_data_row:
                data_row_band += 1

            for col_idx, value in non_empty_pairs:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue

                cell_text = str(value)
                width_hint = min(48, max(8, len(cell_text) + 2))
                prev_width = column_widths.get(col_idx, 0)
                if width_hint > prev_width:
                    column_widths[col_idx] = width_hint

                cell.border = border
                if is_label_row:
                    cell.font = label_font
                    cell.fill = label_fill
                    cell.alignment = left
                    continue
                if is_header_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
                    continue

                cell.font = data_font
                if data_row_band % 2 == 0:
                    cell.fill = zebra_fill
                cell.alignment = left if col_idx <= 2 else center
                if isinstance(value, bool):
                    continue
                if cell.number_format in {"General", "0", ""}:
                    if isinstance(value, int):
                        cell.number_format = "#,##0"
                    elif isinstance(value, float):
                        cell.number_format = "#,##0.00"

        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = width

    LOGGER.info(
        "Exporting erection productivity summary for %s (current week: %s) to %s",
        active_date.strftime("%Y-%m"),
        current_week_label,
        target_path,
    )

    gap_rows = max(2, int(blank_rows_between_tables))

    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        main_row = 0
        main_row = _write_scoped_table(
            writer, weekly_sheet_label, overall_summary, scope_label, startrow=main_row
        )
        main_row += gap_rows
        main_row = _write_scoped_table(
            writer, weekly_sheet_label, pch_summary, scope_label, startrow=main_row
        )
        main_row += gap_rows
        _write_scoped_table(
            writer, weekly_sheet_label, project_summary, scope_label, startrow=main_row
        )

        monthly_summary_row = 0
        monthly_blocks: list[dict[str, int]] = []
        monthly_summary_row, monthly_block = _write_monthly_scoped_table(
            writer,
            monthly_summary_sheet_label,
            monthly_overall_summary,
            scope_label,
            month_headers=monthly_summary_labels,
            group_headers=[],
            startrow=monthly_summary_row,
        )
        monthly_blocks.append(monthly_block)
        monthly_summary_row += gap_rows
        monthly_summary_row, monthly_block = _write_monthly_scoped_table(
            writer,
            monthly_summary_sheet_label,
            monthly_pch_summary,
            scope_label,
            month_headers=monthly_summary_labels,
            group_headers=["PCH"],
            startrow=monthly_summary_row,
        )
        monthly_blocks.append(monthly_block)
        monthly_summary_row += gap_rows
        _, monthly_block = _write_monthly_scoped_table(
            writer,
            monthly_summary_sheet_label,
            monthly_project_summary,
            scope_label,
            month_headers=monthly_summary_labels,
            group_headers=["PCH", "Project"],
            startrow=monthly_summary_row,
        )
        monthly_blocks.append(monthly_block)
        _style_monthly_summary_sheet(writer, monthly_summary_sheet_label, monthly_blocks)

        current_row = 0
        current_row = _write_scope_row(writer, review_sheet_label, scope_label, startrow=current_row)
        current_row += 1
        group_gap_rows = 1
        for table in review_table_one_groups:
            if table.empty:
                continue
            table.to_excel(
                writer,
                sheet_name=review_sheet_label,
                index=False,
                startrow=current_row,
            )
            current_row += len(table) + 1 + group_gap_rows

        review_gap_rows = 2
        current_row += review_gap_rows
        for table in review_table_two_groups:
            if table.empty:
                continue
            table.to_excel(
                writer,
                sheet_name=review_sheet_label,
                index=False,
                startrow=current_row,
            )
            current_row += len(table) + 1 + group_gap_rows
        gang_row = _write_scope_row(writer, gang_level_sheet_label, scope_label, startrow=0)
        gang_level_summary.to_excel(writer, sheet_name=gang_level_sheet_label, index=False, startrow=gang_row)

        idle_row = _write_scope_row(writer, idle_sheet_label, scope_label, startrow=0)
        idle_export.to_excel(writer, sheet_name=idle_sheet_label, index=False, startrow=idle_row)

        rankings_row = _write_scope_row(writer, gangs_sheet_label, scope_label, startrow=0)
        top_title_row = {column: "" for column in ranking_columns}
        top_title_row["PCH"] = "Top 3 Gangs"
        bottom_title_row = {column: "" for column in ranking_columns}
        bottom_title_row["PCH"] = "Bottom 3 Gangs"
        top_title = pd.DataFrame([top_title_row], columns=ranking_columns)
        bottom_title = pd.DataFrame([bottom_title_row], columns=ranking_columns)
        top_title_row_start = rankings_row
        top_title.to_excel(
            writer, sheet_name=gangs_sheet_label, index=False, header=False, startrow=top_title_row_start
        )
        top_gangs_start = top_title_row_start + 1
        top_gangs.to_excel(writer, sheet_name=gangs_sheet_label, index=False, startrow=top_gangs_start)
        top_table_height = len(top_gangs) + 1
        bottom_title_row = top_gangs_start + top_table_height + gap_rows
        bottom_title.to_excel(
            writer, sheet_name=gangs_sheet_label, index=False, header=False, startrow=bottom_title_row
        )
        bottom_gangs.to_excel(writer, sheet_name=gangs_sheet_label, index=False, startrow=bottom_title_row + 1)

        bucket_row = 0
        bucket_gap_rows = 2
        bucket_gap_cols = 3
        thresholds = (1, 2, 3)
        sample_table = overall_idle_bucket_tables.get(thresholds[0])
        if sample_table is None or not len(sample_table.columns):
            sample_table = mtd_idle_bucket_tables.get(thresholds[0])
        table_columns = list(sample_table.columns) if sample_table is not None else []
        empty_bucket_table = pd.DataFrame(columns=table_columns)
        table_width = len(table_columns)
        audit_width = max(len(overall_idle_bucket_audit.columns), len(mtd_idle_bucket_audit.columns))
        table_span = max(table_width, audit_width)
        overall_startcol = 0
        mtd_startcol = overall_startcol + table_span + bucket_gap_cols

        overall_title = pd.DataFrame([["Overall Idle Days Buckets"]])
        overall_title.to_excel(
            writer,
            sheet_name=idle_bucket_sheet_label,
            index=False,
            header=False,
            startrow=bucket_row,
            startcol=overall_startcol,
        )
        mtd_title = pd.DataFrame([["MTD Idle Days Buckets"]])
        mtd_title.to_excel(
            writer,
            sheet_name=idle_bucket_sheet_label,
            index=False,
            header=False,
            startrow=bucket_row,
            startcol=mtd_startcol,
        )
        bucket_row += len(overall_title) + 1

        for threshold in thresholds:
            label = f"Erections >={threshold}"
            pd.DataFrame([[label]]).to_excel(
                writer,
                sheet_name=idle_bucket_sheet_label,
                index=False,
                header=False,
                startrow=bucket_row,
                startcol=overall_startcol,
            )
            pd.DataFrame([[label]]).to_excel(
                writer,
                sheet_name=idle_bucket_sheet_label,
                index=False,
                header=False,
                startrow=bucket_row,
                startcol=mtd_startcol,
            )
            bucket_row += 1
            _write_scope_row(
                writer,
                idle_bucket_sheet_label,
                overall_scope_label,
                startrow=bucket_row,
                startcol=overall_startcol,
            )
            _write_scope_row(
                writer,
                idle_bucket_sheet_label,
                scope_label,
                startrow=bucket_row,
                startcol=mtd_startcol,
            )
            bucket_row += 1

            overall_table = overall_idle_bucket_tables.get(threshold, empty_bucket_table)
            if overall_table is None:
                overall_table = empty_bucket_table
            mtd_table = mtd_idle_bucket_tables.get(threshold, empty_bucket_table)
            if mtd_table is None:
                mtd_table = empty_bucket_table

            overall_table.to_excel(
                writer,
                sheet_name=idle_bucket_sheet_label,
                index=False,
                startrow=bucket_row,
                startcol=overall_startcol,
            )
            mtd_table.to_excel(
                writer,
                sheet_name=idle_bucket_sheet_label,
                index=False,
                startrow=bucket_row,
                startcol=mtd_startcol,
            )
            table_height = max(len(overall_table), len(mtd_table)) + 1
            bucket_row += table_height + bucket_gap_rows

        audit_title = pd.DataFrame([["Idle Days Audit (Erections >=3)"]])
        audit_title.to_excel(
            writer,
            sheet_name=idle_bucket_sheet_label,
            index=False,
            header=False,
            startrow=bucket_row,
            startcol=overall_startcol,
        )
        audit_title.to_excel(
            writer,
            sheet_name=idle_bucket_sheet_label,
            index=False,
            header=False,
            startrow=bucket_row,
            startcol=mtd_startcol,
        )
        bucket_row += len(audit_title) + 1
        _write_scope_row(
            writer,
            idle_bucket_sheet_label,
            overall_scope_label,
            startrow=bucket_row,
            startcol=overall_startcol,
        )
        _write_scope_row(
            writer,
            idle_bucket_sheet_label,
            scope_label,
            startrow=bucket_row,
            startcol=mtd_startcol,
        )
        bucket_row += 1
        overall_idle_bucket_audit.to_excel(
            writer,
            sheet_name=idle_bucket_sheet_label,
            index=False,
            startrow=bucket_row,
            startcol=overall_startcol,
        )
        mtd_idle_bucket_audit.to_excel(
            writer,
            sheet_name=idle_bucket_sheet_label,
            index=False,
            startrow=bucket_row,
            startcol=mtd_startcol,
        )

        monthly_row = 0
        monthly_gap_rows = 2
        monthly_title = pd.DataFrame([["Monthly Gang Production Buckets (All Data)"]])
        monthly_title.to_excel(
            writer, sheet_name=monthly_bucket_sheet_label, index=False, header=False, startrow=monthly_row
        )
        scope_row = monthly_row + len(monthly_title)
        _write_scope_row(
            writer, monthly_bucket_sheet_label, overall_scope_label, startrow=scope_row
        )
        monthly_row += len(monthly_title) + 2
        monthly_bucket_counts.to_excel(
            writer, sheet_name=monthly_bucket_sheet_label, index=False, startrow=monthly_row
        )
        monthly_row += len(monthly_bucket_counts) + 1 + monthly_gap_rows
        monthly_bucket_details.to_excel(
            writer, sheet_name=monthly_bucket_sheet_label, index=False, startrow=monthly_row
        )

        audit_row = 0
        audit_row = _write_scope_row(writer, excluded_sheet_label, scope_label, startrow=audit_row)
        audit_row += 1
        pd.DataFrame([["Project-wise Excluded Rows"]]).to_excel(
            writer,
            sheet_name=excluded_sheet_label,
            index=False,
            header=False,
            startrow=audit_row,
        )
        audit_row += 2
        excluded_summary.to_excel(
            writer, sheet_name=excluded_sheet_label, index=False, startrow=audit_row
        )
        audit_row += len(excluded_summary) + 2
        pd.DataFrame([["Excluded Row Details"]]).to_excel(
            writer,
            sheet_name=excluded_sheet_label,
            index=False,
            header=False,
            startrow=audit_row,
        )
        audit_row += 2
        excluded_detail.to_excel(
            writer, sheet_name=excluded_sheet_label, index=False, startrow=audit_row
        )

        used_sheet_names: set[str] = {
            weekly_sheet_label,
            monthly_summary_sheet_label,
            review_sheet_label,
            gangs_sheet_label,
            gang_level_sheet_label,
            idle_sheet_label,
            idle_bucket_sheet_label,
            monthly_bucket_sheet_label,
            excluded_sheet_label,
        }
        if not review_table_raw.empty and "PCH" in review_table_raw.columns and "pch_display" in scope.columns:
            for pch_name in sorted(review_table_raw["PCH"].fillna("").astype(str).unique(), key=_pch_sort_components):
                if not pch_name:
                    continue
                pch_sheet_name = _unique_sheet_name(f"PCH {pch_name}", used_sheet_names)
                pch_table_one = review_table_raw[review_table_raw["PCH"] == pch_name].copy()
                if pch_table_one.empty:
                    continue
                pch_table_one = _append_totals_row(pch_table_one)
                pch_scope = scope[scope["pch_display"] == pch_name]
                pch_table_two_overall = _build_review_table_two(pch_scope)

                current_row = 0
                current_row = _write_scope_row(
                    writer, pch_sheet_name, scope_label, startrow=current_row
                )
                current_row += 1
                pch_table_one.to_excel(writer, sheet_name=pch_sheet_name, index=False, startrow=current_row)
                current_row += len(pch_table_one) + 1 + group_gap_rows

                overall_header = {"Erection productivity": "PCH Overall"}
                for label in month_labels:
                    overall_header[label] = ""
                overall_header_df = pd.DataFrame(
                    [overall_header], columns=["Erection productivity", *month_labels]
                )
                overall_header_df.to_excel(writer, sheet_name=pch_sheet_name, index=False, startrow=current_row)
                current_row += len(overall_header_df) + 1
                pch_table_two_overall.to_excel(writer, sheet_name=pch_sheet_name, index=False, startrow=current_row)
                current_row += len(pch_table_two_overall) + 1 + group_gap_rows

                if "project_display" in pch_scope.columns:
                    project_names = (
                        pch_table_one["Project"]
                        .dropna()
                        .astype(str)
                        .map(str.strip)
                        .unique()
                    )
                    for project_name in project_names:
                        if not project_name or project_name == "Total":
                            continue
                        project_scope = pch_scope[pch_scope["project_display"] == project_name]
                        header_row = {"Erection productivity": f"Project: {project_name}"}
                        for label in month_labels:
                            header_row[label] = ""
                        header_df = pd.DataFrame(
                            [header_row], columns=["Erection productivity", *month_labels]
                        )
                        header_df.to_excel(writer, sheet_name=pch_sheet_name, index=False, startrow=current_row)
                        current_row += len(header_df) + 1
                        project_table_two = _build_review_table_two(project_scope)
                        project_table_two.to_excel(writer, sheet_name=pch_sheet_name, index=False, startrow=current_row)
                        current_row += len(project_table_two) + 1 + group_gap_rows

        for styled_sheet in sorted(used_sheet_names):
            if styled_sheet == monthly_summary_sheet_label:
                continue
            _style_clean_sheet(writer, styled_sheet)

    return target_path


def export_voltage_tower_productivity_summary(
    output_path: str | Path,
    *,
    data_store: AppDataStore | None = None,
    daily_df: pd.DataFrame | None = None,
    project_info: pd.DataFrame | None = None,
    project_voltage_df: pd.DataFrame | None = None,
    project_voltage_path: Path | None = None,
    as_of_date: pd.Timestamp | str | None = None,
    range_start: pd.Timestamp | str | None = None,
    range_end: pd.Timestamp | str | None = None,
    scope_mode: str = "auto",
    sheet_name: str = _DEFAULT_KV_SHEET_NAME,
    issues_sheet_name: str = "Issues",
    blank_rows_between_tables: int = 3,
) -> Path:
    """Export productivity grouped by KV level and tower type."""

    source_daily = daily_df
    source_project_info = project_info
    candidate_date = pd.Timestamp(as_of_date) if as_of_date is not None else None
    voltage_frame = project_voltage_df

    if project_voltage_df is None and project_voltage_path is not None:
        try:
            voltage_frame = pd.read_excel(project_voltage_path)
        except FileNotFoundError:
            LOGGER.warning("Projects_KV workbook not found at %s; voltage lookup disabled.", project_voltage_path)
            voltage_frame = pd.DataFrame()
        except Exception as exc:
            LOGGER.warning("Unable to read Projects_KV workbook '%s': %s", project_voltage_path, exc)
            voltage_frame = pd.DataFrame()

    if data_store is not None:
        if source_daily is None:
            source_daily = data_store.get_daily()
        if source_project_info is None:
            source_project_info = data_store.get_project_info()

    if source_daily is None or source_daily.empty:
        raise ValueError("No erection daily dataframe is available to export.")

    date_series = pd.to_datetime(source_daily.get("date"), errors="coerce").dropna()
    if date_series.empty:
        raise ValueError("Unable to determine a date range for the KV productivity export.")

    range_start_ts = pd.Timestamp(range_start) if range_start is not None else None
    range_end_ts = pd.Timestamp(range_end) if range_end is not None else None

    scope_mode_norm = (scope_mode or "auto").strip().lower()
    if scope_mode_norm not in {"auto", "full", "month", "range"}:
        raise ValueError(f"Unknown scope_mode '{scope_mode}'. Expected one of: auto, full, month, range.")

    if scope_mode_norm == "full":
        if range_start_ts is not None or range_end_ts is not None or candidate_date is not None:
            raise ValueError("scope_mode='full' cannot be combined with range_start/range_end/as_of_date.")
        range_mode = False
        full_range = True
        month_start = date_series.min().normalize()
        month_end = date_series.max().normalize()
        active_date = month_end
    elif scope_mode_norm == "range":
        range_mode = True
        if range_start_ts is None or range_end_ts is None:
            raise ValueError("Both range_start and range_end are required for a date range export.")
        start_month = range_start_ts.to_period("M").to_timestamp()
        end_month = range_end_ts.to_period("M").to_timestamp()
        if end_month < start_month:
            raise ValueError("range_end must be the same month or after range_start.")
        month_start = start_month.normalize()
        month_end = (end_month + pd.offsets.MonthEnd(1)).normalize()
        active_date = month_end
        full_range = False
    elif scope_mode_norm == "month":
        if range_start_ts is not None or range_end_ts is not None:
            raise ValueError("scope_mode='month' cannot be combined with range_start/range_end.")
        if candidate_date is None or pd.isna(candidate_date):
            candidate_date = date_series.max()
        range_mode = False
        full_range = False
        active_date = pd.Timestamp(candidate_date).normalize()
        month_start = active_date.to_period("M").to_timestamp()
        month_end = (month_start + pd.offsets.MonthEnd(1)).normalize()
    else:
        range_mode = range_start_ts is not None or range_end_ts is not None
        if range_mode:
            if range_start_ts is None or range_end_ts is None:
                raise ValueError("Both range_start and range_end are required for a date range export.")
            start_month = range_start_ts.to_period("M").to_timestamp()
            end_month = range_end_ts.to_period("M").to_timestamp()
            if end_month < start_month:
                raise ValueError("range_end must be the same month or after range_start.")
            month_start = start_month.normalize()
            month_end = (end_month + pd.offsets.MonthEnd(1)).normalize()
            active_date = month_end
            full_range = False
        else:
            full_range = candidate_date is None or pd.isna(candidate_date)
            if full_range:
                month_start = date_series.min().normalize()
                month_end = date_series.max().normalize()
                active_date = month_end
            else:
                active_date = pd.Timestamp(candidate_date).normalize()
                month_start = active_date.to_period("M").to_timestamp()
                month_end = (month_start + pd.offsets.MonthEnd(1)).normalize()

    week_labels = _generate_week_labels(month_start, month_end)

    project_info_frame = source_project_info.copy() if isinstance(source_project_info, pd.DataFrame) else pd.DataFrame()
    voltage_lookup = _prepare_project_voltage_lookup(voltage_frame)

    scope = _prepare_month_scope(
        source_daily, project_info_frame, month_start, month_end, week_labels, filter_by_date=True
    )
    if scope.empty:
        raise ValueError("No erection rows found for the requested period.")
    scope = _annotate_scope_with_voltage(scope, voltage_lookup)

    completion_base = _prepare_month_scope(
        source_daily, project_info_frame, month_start, month_end, week_labels, filter_by_date=False
    )
    completion_base = _annotate_scope_with_voltage(completion_base, voltage_lookup)
    completions = _prepare_completion_scope(completion_base, month_start, month_end, week_labels)
    if isinstance(completions, pd.DataFrame) and not completions.empty:
        tower_weights = pd.to_numeric(completions.get("tower_weight"), errors="coerce")
        completions = completions.assign(_tower_weight=tower_weights.fillna(0.0))
    else:
        completions = pd.DataFrame(columns=list(scope.columns) + ["_tower_weight"])

    voltage_table = _build_group_productivity_table(
        scope,
        completions,
        week_labels,
        group_columns=["voltage_label"],
        column_labels={"voltage_label": "Voltage"},
        include_weekly=False,
        metric_labels=YTD_METRIC_LABELS,
    )
    family_table = _build_group_productivity_table(
        scope,
        completions,
        week_labels,
        group_columns=["voltage_label", "tower_family"],
        column_labels={"voltage_label": "Voltage", "tower_family": "Tower Family"},
        include_weekly=False,
        metric_labels=SUMMARY_METRIC_LABELS,
    )
    tower_table = _build_group_productivity_table(
        scope,
        completions,
        week_labels,
        group_columns=["voltage_label", "tower_type"],
        column_labels={"voltage_label": "Voltage", "tower_type": "Tower Type"},
        include_weekly=False,
        metric_labels=YTD_METRIC_LABELS,
    )

    sheet_label = _sanitize_sheet_name(sheet_name or _DEFAULT_KV_SHEET_NAME) or _DEFAULT_KV_SHEET_NAME
    issues_label = _sanitize_sheet_name(issues_sheet_name or "Issues") or "Issues"
    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    if range_mode:
        period_label = f"{month_start:%Y-%m-%d} to {month_end:%Y-%m-%d}"
    elif full_range:
        period_label = f"{month_start:%Y-%m-%d} to {month_end:%Y-%m-%d}"
    else:
        period_label = active_date.strftime("%Y-%m")
    LOGGER.info("Exporting KV productivity summary for %s to %s", period_label, target_path)

    gap_rows = max(2, int(blank_rows_between_tables))
    tables = [voltage_table, family_table, tower_table]
    issues_table = _build_kv_export_issues(source_daily, month_start, month_end, voltage_lookup)
    start_row = 0
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        for table in tables:
            table.to_excel(writer, sheet_name=sheet_label, index=False, startrow=start_row)
            start_row += len(table) + 1 + gap_rows
        if issues_table is not None:
            issues_table.to_excel(writer, sheet_name=issues_label, index=False)

    return target_path


def export_weekly_dpr_analysis(
    project_code: str,
    output_path: str | Path,
    *,
    dpr_paths: Sequence[str | Path] | None = None,
    dpr_folder: str | Path | None = None,
    as_of_date: pd.Timestamp | str | None = None,
    daily_path: str | Path | None = None,
    sheet_summary: str = "WeeklySummary",
    sheet_details: str = "WeeklyDetails",
    sheet_context: str = "SelectionContext",
    sheet_gang_weekly: str = "GangWeekly",
) -> Path:
    """
    Build a weekly completion detail workbook for a project's DPR, including manpower,
    tower type, and tower weight per location along with gang-wise productivity.
    """

    project_label = str(project_code or "").strip()
    if not project_label:
        raise ValueError("A project code is required for the DPR weekly analysis.")

    source_labels: list[str] = []
    combined = pd.DataFrame()

    compiled_scope = _load_compiled_completion_rows(daily_path, project_label)
    if not compiled_scope.empty:
        combined = compiled_scope
        dataset_label = Path(daily_path).name if daily_path else "compiled"
        source_labels = [f"{dataset_label} (compiled)"]
    else:
        candidate_paths = _resolve_dpr_paths(project_label, dpr_paths, dpr_folder)
        if not candidate_paths:
            raise FileNotFoundError(
                f"No DPR workbooks were found for '{project_label}'. Provide --dpr-path or place the file under Raw Data/DPRs."
            )

        records: list[pd.DataFrame] = []
        for path in candidate_paths:
            table = _load_dpr_table(path)
            cleaned = _prepare_dpr_records(table, project_label)
            if cleaned.empty:
                LOGGER.warning("DPR workbook '%s' did not yield usable completion rows; skipping.", path)
                continue
            records.append(cleaned)

        if not records:
            raise ValueError(
                f"No completion rows were extracted from the compiled dataset or DPR files for '{project_label}'."
            )

        combined = pd.concat(records, ignore_index=True)
        source_labels = [path.name for path in candidate_paths]
    completion_series = pd.to_datetime(combined["completion_date"], errors="coerce")
    latest_completion = completion_series.max()
    candidate_date = pd.Timestamp(as_of_date) if as_of_date is not None else latest_completion
    if pd.isna(candidate_date):
        raise ValueError("Unable to determine the reference month; no completion dates were detected.")

    active_date = pd.Timestamp(candidate_date).normalize()
    month_start = active_date.to_period("M").to_timestamp()
    month_end = (month_start + pd.offsets.MonthEnd(1)).normalize()

    scope = combined[
        (completion_series >= month_start)
        & (completion_series <= month_end)
        & completion_series.notna()
    ].copy()
    if scope.empty:
        raise ValueError(
            f"No completed locations for '{project_label}' fall within {month_start:%B %Y}. "
            "Try adjusting --as-of-date or provide DPR files for the desired month."
        )

    week_labels = _generate_week_labels(month_start, month_end)
    scope["week_index"] = _assign_week_bucket(scope["completion_date"], month_start, week_labels)

    summary_rows: list[dict[str, object]] = []
    for idx, label in enumerate(week_labels, start=1):
        week_scope = scope[scope["week_index"] == idx]
        tower_weight_sum = week_scope["tower_weight"].sum() if not week_scope.empty else 0.0
        tower_weight_avg = week_scope["tower_weight"].mean() if not week_scope.empty else 0.0
        manpower_avg = week_scope["manpower"].mean() if not week_scope.empty else 0.0
        productivity_avg = week_scope["derived_productivity"].mean() if not week_scope.empty else 0.0
        if pd.isna(tower_weight_sum):
            tower_weight_sum = 0.0
        if pd.isna(tower_weight_avg):
            tower_weight_avg = 0.0
        if pd.isna(manpower_avg):
            manpower_avg = 0.0
        if pd.isna(productivity_avg):
            productivity_avg = 0.0
        unique_gangs = sorted({name for name in week_scope["gang_name"] if isinstance(name, str) and name})
        summary_rows.append(
            {
                "Week": label,
                "Week Range": _format_week_range(idx, month_start, month_end),
                "Erections Completed": int(len(week_scope)),
                "Total Tower Weight (MT)": round(float(tower_weight_sum), 2),
                "Avg Tower Weight (MT)": round(float(tower_weight_avg), 2),
                "Avg Manpower": round(float(manpower_avg), 1),
                "Avg Productivity (MT/day)": round(float(productivity_avg), 2),
                "Active Gangs": ", ".join(unique_gangs),
            }
        )
    summary_df = pd.DataFrame(summary_rows)
    gang_weekly_table = _build_gang_weekly_productivity_table(scope, week_labels)

    detail = scope.copy()
    detail["Week"] = detail["week_index"].map(lambda idx: week_labels[idx - 1] if idx > 0 and idx <= len(week_labels) else "")
    detail = detail.sort_values(["week_index", "completion_date", "location_no"]).reset_index(drop=True)
    detail_table = detail.rename(
        columns={
            "completion_date": "Completion Date",
            "start_date": "Starting Date",
            "location_no": "Location No.",
            "tower_type": "Tower Type",
            "tower_weight": "Tower Weight (MT)",
            "gang_name": "Gang Name",
            "manpower": "Manpower",
            "status": "Status",
            "derived_productivity": "Productivity (MT/day)",
        }
    )
    # Ensure consistent column order and friendly data types
    detail_table["Completion Date"] = pd.to_datetime(detail_table["Completion Date"], errors="coerce").dt.date
    start_series = (
        pd.to_datetime(detail_table["Starting Date"], errors="coerce")
        if "Starting Date" in detail_table.columns
        else pd.Series(pd.NaT, index=detail_table.index, dtype="datetime64[ns]")
    )
    detail_table["Starting Date"] = start_series.dt.date
    if "Tower Weight (MT)" in detail_table.columns:
        detail_table["Tower Weight (MT)"] = pd.to_numeric(detail_table["Tower Weight (MT)"], errors="coerce").round(3)
    else:
        detail_table["Tower Weight (MT)"] = pd.Series(index=detail_table.index, dtype="float64")
    if "Manpower" in detail_table.columns:
        manpower_series = pd.to_numeric(detail_table["Manpower"], errors="coerce")
        detail_table["Manpower"] = manpower_series.round(0).astype("Int64")
    else:
        detail_table["Manpower"] = pd.Series(pd.NA, index=detail_table.index, dtype="Int64")
    if "Productivity (MT/day)" in detail_table.columns:
        detail_table["Productivity (MT/day)"] = (
            pd.to_numeric(detail_table["Productivity (MT/day)"], errors="coerce").round(2)
        )
    else:
        detail_table["Productivity (MT/day)"] = pd.Series(index=detail_table.index, dtype="float64")
    ordered_columns = [
        "Week",
        "Completion Date",
        "Location No.",
        "Tower Type",
        "Tower Weight (MT)",
        "Manpower",
        "Gang Name",
        "Status",
        "Starting Date",
        "Productivity (MT/day)",
        "source_file",
    ]
    present_detail_cols = [column for column in ordered_columns if column in detail_table.columns]
    detail_table = detail_table[present_detail_cols].rename(columns={"source_file": "Source"})

    context = pd.DataFrame(
        [
            {
                "Project Code": project_label,
                "Analysis Month": month_start.strftime("%B %Y"),
                "Month Start": month_start.date(),
                "Month End": month_end.date(),
                "As Of Date": active_date.date(),
                "Source DPR Files": ", ".join(source_labels) if source_labels else "",
                "Total Completions (Month)": len(scope),
                "Unique Locations": scope["location_no"].nunique(),
                "Unique Gangs": scope["gang_name"].nunique(),
            }
        ]
    )

    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name=_sanitize_sheet_name(sheet_summary), index=False)
        detail_table.to_excel(writer, sheet_name=_sanitize_sheet_name(sheet_details), index=False)
        gang_weekly_table.to_excel(writer, sheet_name=_sanitize_sheet_name(sheet_gang_weekly), index=False)
        context.to_excel(writer, sheet_name=_sanitize_sheet_name(sheet_context), index=False)

    LOGGER.info(
        "Exported DPR weekly analysis for %s (%s completions) to %s",
        project_label,
        len(scope),
        target_path,
    )
    return target_path


def _summarize_top_gangs(month_df: pd.DataFrame, *, limit: int = 3) -> str:
    """
    Return a comma-separated label for the top gangs and their average productivity.
    """

    if "derived_productivity" not in month_df.columns:
        return ""
    working = month_df.dropna(subset=["derived_productivity"]).copy()
    if working.empty:
        return ""
    working["gang_name"] = working.get("gang_name", "").fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty:
        return ""
    grouped = (
        working.groupby("gang_name")["derived_productivity"]
        .mean()
        .sort_values(ascending=False)
    )
    labels: list[str] = []
    for gang, avg in grouped.head(limit).items():
        labels.append(f"{gang} ({avg:.2f})")
    return ", ".join(labels)


def _build_monthly_productivity_summary(completions: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Month",
        "Erections Completed",
        "Total Tower Weight (MT)",
        "Avg Tower Weight (MT)",
        "Avg Productivity (MT/day)",
        "Active Gangs",
        "Top 3 Gangs (Avg Productivity)",
    ]
    if completions.empty:
        return pd.DataFrame(columns=columns)

    working = completions.copy()
    working["completion_date"] = pd.to_datetime(working.get("completion_date"), errors="coerce").dt.normalize()
    working = working.dropna(subset=["completion_date"])
    if working.empty:
        return pd.DataFrame(columns=columns)

    working["tower_weight"] = pd.to_numeric(working.get("tower_weight"), errors="coerce")
    working["derived_productivity"] = pd.to_numeric(working.get("derived_productivity"), errors="coerce")
    working["gang_name"] = working.get("gang_name", "").fillna("").astype(str).str.strip()
    working["month_key"] = working["completion_date"].dt.to_period("M").dt.to_timestamp()

    rows: list[dict[str, object]] = []
    for month_key, month_df in working.groupby("month_key"):
        tower_weights = month_df["tower_weight"].dropna()
        productivity = month_df["derived_productivity"].dropna()

        gangs = sorted({name for name in month_df["gang_name"] if name})
        rows.append(
            {
                "_month_sort": month_key,
                "Month": month_key.strftime("%b %Y"),
                "Erections Completed": int(len(month_df)),
                "Total Tower Weight (MT)": round(float(tower_weights.sum()), 2) if not tower_weights.empty else 0.0,
                "Avg Tower Weight (MT)": round(float(tower_weights.mean()), 2) if not tower_weights.empty else 0.0,
                "Avg Productivity (MT/day)": round(float(productivity.mean()), 2) if not productivity.empty else 0.0,
                "Active Gangs": ", ".join(gangs),
                "Top 3 Gangs (Avg Productivity)": _summarize_top_gangs(month_df),
            }
        )
    summary = pd.DataFrame(rows)
    if summary.empty:
        return pd.DataFrame(columns=columns)
    ordered = summary.sort_values("_month_sort").drop(columns="_month_sort")
    return ordered.reindex(columns=columns)


def export_monthly_productivity_summary(
    project_code: str,
    output_path: str | Path,
    *,
    compiled_path: str | Path | None = None,
    sheet_summary: str = "MonthlySummary",
    sheet_details: str = "MonthlyDetails",
    sheet_context: str = "SelectionContext",
) -> Path:
    """
    Generate a month-wise productivity workbook using the compiled erection dataset.
    """

    project_label = str(project_code or "").strip()
    if not project_label:
        raise ValueError("A project code is required for the monthly productivity export.")

    dataset_path = Path(compiled_path) if compiled_path else Path("Parquets") / "Erection" / "ErectionCompiled_Output.xlsx"
    completions = _load_compiled_completion_rows(dataset_path, project_label)
    if completions.empty:
        raise ValueError(
            f"No completion rows were found in '{dataset_path}' for project '{project_label}'. "
            "Ensure the compiled workbook includes the requested project."
        )

    summary_df = _build_monthly_productivity_summary(completions)
    if summary_df.empty:
        raise ValueError(
            f"The compiled dataset '{dataset_path}' did not expose usable completion months for '{project_label}'."
        )

    detail_table = completions.copy()
    detail_table["Completion Month"] = pd.to_datetime(detail_table.get("completion_date"), errors="coerce").dt.strftime("%b %Y")
    detail_table = detail_table.sort_values(["completion_date", "location_no"]).reset_index(drop=True)
    detail_table = detail_table.rename(
        columns={
            "completion_date": "Completion Date",
            "start_date": "Starting Date",
            "location_no": "Location No.",
            "tower_type": "Tower Type",
            "tower_weight": "Tower Weight (MT)",
            "gang_name": "Gang Name",
            "status": "Status",
            "derived_productivity": "Productivity (MT/day)",
            "project_name": "Project Name",
            "source_file": "Source",
        }
    )
    for column in ("Tower Weight (MT)", "Productivity (MT/day)"):
        if column in detail_table.columns:
            detail_table[column] = pd.to_numeric(detail_table[column], errors="coerce").round(2)
    if "Starting Date" in detail_table.columns:
        detail_table["Starting Date"] = (
            pd.to_datetime(detail_table["Starting Date"], errors="coerce").dt.date
        )
    if "Completion Date" in detail_table.columns:
        detail_table["Completion Date"] = (
            pd.to_datetime(detail_table["Completion Date"], errors="coerce").dt.date
        )
    ordered_detail_cols = [
        "Completion Month",
        "Completion Date",
        "Location No.",
        "Project Name",
        "Tower Type",
        "Tower Weight (MT)",
        "Gang Name",
        "Status",
        "Starting Date",
        "Productivity (MT/day)",
        "Source",
    ]
    detail_table = detail_table[[col for col in ordered_detail_cols if col in detail_table.columns]]

    first_completion = pd.to_datetime(completions["completion_date"], errors="coerce").min()
    last_completion = pd.to_datetime(completions["completion_date"], errors="coerce").max()
    project_names = sorted({name for name in completions.get("project_name", []) if isinstance(name, str) and name.strip()})

    context = pd.DataFrame(
        [
            {
                "Project Code": project_label,
                "Project Names": ", ".join(project_names),
                "Dataset Path": str(Path(dataset_path).resolve()),
                "Range Start": first_completion.date() if pd.notna(first_completion) else "",
                "Range End": last_completion.date() if pd.notna(last_completion) else "",
                "Months Covered": len(summary_df),
                "Total Completions": len(completions),
            }
        ]
    )

    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name=_sanitize_sheet_name(sheet_summary), index=False)
        detail_table.to_excel(writer, sheet_name=_sanitize_sheet_name(sheet_details), index=False)
        context.to_excel(writer, sheet_name=_sanitize_sheet_name(sheet_context), index=False)

    LOGGER.info(
        "Exported monthly productivity summary for %s (%d months, %d completions) to %s",
        project_label,
        len(summary_df),
        len(completions),
        target_path,
    )
    return target_path
