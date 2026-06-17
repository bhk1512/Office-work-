"""Config-driven DPR foundation ingestion."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional
import re

MIN_COMPLETION_YEAR = 2015
MAX_FUTURE_DAYS_FROM_REPORT = 14

import pandas as pd
from openpyxl import load_workbook

from erection_compiled_to_daily_new import load_sheet_with_csv_fallback
from . import stringing_ingest as ingest
from .completed_projects import is_completed_project
from .project_identity import (
    build_project_display,
    build_project_scope_key,
    normalize_line_name,
    parse_project_identity_from_filename,
    parse_sheet_line_entries,
)


FOUNDATION_RAW_COLUMNS = [
    "project_code",
    "project_display",
    "project_scope_key",
    "line_name",
    "line_name_source",
    "report_date",
    "start_date",
    "event_date",
    "source_file",
    "source_sheet",
    "configured_sheet",
    "source_type",
    "quality_flag",
    "location_no",
    "gang_name",
    "status_text",
    "cumulative_foundation",
    "fallback_note",
]

FOUNDATION_COMPLETIONS_COLUMNS = [
    "project_code",
    "project_display",
    "project_scope_key",
    "line_name",
    "line_name_source",
    "report_date",
    "start_date",
    "event_date",
    "period_week_start",
    "period_month",
    "source_file",
    "source_sheet",
    "configured_sheet",
    "source_type",
    "quality_flag",
    "location_no",
    "gang_name",
    "event_value",
    "cumulative_foundation",
]

COVERAGE_COLUMNS = [
    "project_code",
    "project_display",
    "status",
    "reason_code",
    "reason",
    "source_used",
    "snapshot_limited",
    "workbook",
    "configured_sheet",
    "resolved_sheet",
    "detail_rows",
    "detail_completions",
    "snapshot_rows",
    "first_event_date",
    "last_event_date",
    "first_report_date",
    "last_report_date",
    "available_sheets",
]

DIAGNOSTICS_COLUMNS = [
    "Workbook",
    "Project",
    "Sheet",
    "ConfiguredSheet",
    "LineName",
    "LineNameSource",
    "Rows",
    "DetailRows",
    "Completions",
    "ParserMode",
    "FallbackNote",
    "Status",
    "Reason",
    "TemplateSheet",
    "TemplateApplied",
    "TemplateChanges",
]

ISSUES_COLUMNS = [
    "Workbook",
    "Project",
    "Sheet",
    "ConfiguredSheet",
    "LineName",
    "LineNameSource",
    "Issue",
    "Reason",
]


@dataclass(frozen=True)
class FoundationParseResult:
    raw_rows: list[dict[str, object]]
    completion_rows: list[dict[str, object]]
    parse_status: str
    parse_reason: str
    parser_mode: str
    rows_examined: int
    template_sheet: str = ""
    template_applied: bool = False
    template_changes: tuple[str, ...] = ()


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip().lower()
    if not text:
        return ""
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _normalize_text(value)).strip("_")


def _split_tokens(value: object) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [part.strip() for part in re.split(r"[;,|]+", text) if str(part).strip()]


def _as_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).replace("\u00a0", " ").strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def _coerce_numeric(value: object) -> float | None:
    if value is None:
        return None
    try:
        numeric = pd.to_numeric([value], errors="coerce")[0]
    except Exception:
        return None
    if pd.isna(numeric):
        return None
    return float(numeric)


def _parse_report_timestamp(value: object) -> pd.Timestamp | pd.NaT:
    text = str(value or "").strip()
    if not text:
        return pd.NaT
    ts = pd.to_datetime(text, errors="coerce")
    if pd.notna(ts):
        return pd.Timestamp(ts).normalize()
    return pd.NaT


def _has_date_signal(text: str) -> bool:
    value = str(text or "").strip()
    if not value:
        return False
    if re.fullmatch(r"\d{1,4}", value):
        return False
    if re.search(r"\d{1,4}[-/.]\d{1,2}[-/.]\d{1,4}", value):
        return True
    if re.search(r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\b", value, flags=re.IGNORECASE):
        return True
    return False


def _is_plausible_completion_date(
    value: pd.Timestamp,
    *,
    report_ts: pd.Timestamp | pd.NaT,
) -> bool:
    ts = pd.Timestamp(value).normalize()
    if ts.year < MIN_COMPLETION_YEAR:
        return False
    if ts.year > 2100:
        return False
    if pd.notna(report_ts):
        max_allowed = pd.Timestamp(report_ts).normalize() + pd.Timedelta(days=MAX_FUTURE_DAYS_FROM_REPORT)
        if ts > max_allowed:
            return False
    return True


def _coerce_date(value: object, *, report_ts: pd.Timestamp | pd.NaT = pd.NaT) -> pd.Timestamp | pd.NaT:
    if value is None:
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        normalized = value.normalize()
        return normalized if _is_plausible_completion_date(normalized, report_ts=report_ts) else pd.NaT
    if isinstance(value, datetime):
        normalized = pd.Timestamp(value).normalize()
        return normalized if _is_plausible_completion_date(normalized, report_ts=report_ts) else pd.NaT
    if isinstance(value, date):
        normalized = pd.Timestamp(value).normalize()
        return normalized if _is_plausible_completion_date(normalized, report_ts=report_ts) else pd.NaT

    numeric = _coerce_numeric(value)
    if numeric is not None and 20000 <= numeric <= 80000:
        converted = pd.to_datetime(numeric, errors="coerce", unit="D", origin="1899-12-30")
        if pd.notna(converted):
            normalized = pd.Timestamp(converted).normalize()
            return normalized if _is_plausible_completion_date(normalized, report_ts=report_ts) else pd.NaT

    text = _as_text(value)
    if not text:
        return pd.NaT
    if not _has_date_signal(text):
        return pd.NaT

    iso_like = bool(re.fullmatch(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", text))
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=not iso_like)
    if pd.notna(parsed):
        normalized = pd.Timestamp(parsed).normalize()
        return normalized if _is_plausible_completion_date(normalized, report_ts=report_ts) else pd.NaT
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=iso_like)
    if pd.notna(parsed):
        normalized = pd.Timestamp(parsed).normalize()
        return normalized if _is_plausible_completion_date(normalized, report_ts=report_ts) else pd.NaT
    return pd.NaT


def _extract_report_date_from_filename(name: str) -> str:
    match = re.search(r"(20\d{2}-\d{2}-\d{2})", str(name))
    if match:
        return match.group(1)
    return ""


def _looks_like_location(value: object) -> bool:
    text = _as_text(value)
    if not text:
        return False
    if len(text) > 48:
        return False
    lowered = text.lower()
    if any(token in lowered for token in ("activity", "status", "remarks", "progress", "cumulative", "balance", "plan")):
        return False
    if re.match(r"^\d+[a-z]?(?:[/\-]\d+[a-z]?)?$", lowered):
        return True
    if re.match(r"^[a-z]{1,5}[-/]\d+[a-z]?(?:[/\-]\d+[a-z]?)?$", lowered):
        return True
    if re.match(r"^[a-z]{1,5}\d+(?:[/\-]\d+[a-z]?)?$", lowered):
        return True
    if re.search(r"\d", lowered) and "/" in lowered:
        return True
    if ("loc" in lowered or "tower" in lowered) and re.search(r"\d", lowered):
        return True
    return False


def _looks_completed(value: object) -> bool:
    text = _normalize_text(value)
    if not text:
        return False
    if "not complete" in text or "incomplete" in text:
        return False
    return any(token in text for token in ("complete", "completed", "done", "finished"))


def _make_unique_headers(labels: list[str]) -> list[str]:
    unique: list[str] = []
    seen: dict[str, int] = {}
    for idx, label in enumerate(labels, start=1):
        base = label.strip() if label else f"unnamed_col_{idx}"
        key = ingest.normalize_space_only(base) or base.lower()
        count = seen.get(key, 0) + 1
        seen[key] = count
        unique.append(base if count == 1 else f"{base}__{count}")
    return unique


def _resolve_named_template_sheet(wb, expected_name: str) -> str | None:
    expected_key = ingest.normalize_space_only(expected_name)
    for name in wb.sheetnames:
        if ingest.normalize_space_only(name) == expected_key:
            return name
    return None


def _resolve_project_template_sheets(wb, project_name: object, discipline: str) -> list[str]:
    project_text = str(project_name or "").strip()
    if not project_text:
        return []

    resolved: list[str] = []
    seen: set[str] = set()
    for expected in (
        f"{project_text} {discipline}",
        f"{project_text} {discipline} Template Check",
    ):
        hit = _resolve_named_template_sheet(wb, expected)
        if hit and hit not in seen:
            seen.add(hit)
            resolved.append(hit)

    project_key = ingest.normalize_space_only(project_text)
    discipline_key = ingest.normalize_space_only(discipline)
    for name in wb.sheetnames:
        key = ingest.normalize_space_only(name)
        if key and key.startswith(project_key) and key.endswith(discipline_key) and name not in seen:
            seen.add(name)
            resolved.append(name)
    return resolved


def _extract_numeric_tokens(value: object) -> set[str]:
    text = str(value or "")
    return {token for token in re.findall(r"\d{2,4}", text)}


def _extract_template_column_map(ws) -> dict[int, str]:
    to_map_row = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if ingest.normalize_space_only(cell) == "to map":
                to_map_row = row_idx
                break
        if to_map_row is not None:
            break
    if to_map_row is None:
        return {}

    labels_row = to_map_row + 1
    row_values = next(ws.iter_rows(min_row=labels_row, max_row=labels_row, values_only=True), ())
    mapping: dict[int, str] = {}
    for col_idx, value in enumerate(row_values):
        label = str(value).strip() if value is not None else ""
        if label:
            mapping[col_idx] = label
    return mapping


def _apply_template_column_mapping(
    df: pd.DataFrame,
    template_map: dict[int, str],
) -> tuple[pd.DataFrame, list[str]]:
    if df is None or df.empty or not template_map:
        return df, []
    remapped = df.copy()
    columns = list(remapped.columns)
    changes: list[str] = []
    for idx, mapped_name in sorted(template_map.items()):
        if idx >= len(columns):
            continue
        current = str(columns[idx]).strip()
        target = str(mapped_name).strip()
        if not target:
            continue
        columns[idx] = target
        if ingest.normalize_space_only(current) != ingest.normalize_space_only(target):
            changes.append(f"C{idx + 1}:{current}->{target}")
    remapped.columns = columns
    return remapped, changes


def _looks_like_gang_label(label: str) -> bool:
    normalized = _normalize_text(label)
    label_key = re.sub(r"[^a-z0-9]+", " ", normalized).strip()
    if not label_key:
        return False
    gang_tokens = (
        "gang name",
        "gang",
        "contractor",
        "sub contractor",
        "subcontractor",
        "agency",
        "vendor",
    )
    return any(token in label_key for token in gang_tokens)


def _label_key(label: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _normalize_text(label)).strip()


def _looks_like_start_label(label: object) -> bool:
    label_key = _label_key(label)
    if not label_key:
        return False
    if _looks_like_completion_label(label):
        return False
    return (
        label_key in {"start", "starting", "dos"}
        or "start date" in label_key
        or "starting date" in label_key
        or "date of start" in label_key
        or "foundation start" in label_key
        or "fdn start" in label_key
    )


def _looks_like_completion_label(label: object) -> bool:
    label_key = _label_key(label)
    if not label_key:
        return False
    has_start_token = any(token in label_key for token in ("start", "starting", "begin", "began"))
    return (
        "completion date" in label_key
        or "complete date" in label_key
        or "date of completion" in label_key
        or "ending date" in label_key
        or label_key in {"completion", "completed", "complete", "doc"}
        or ("completed" in label_key and "date" in label_key)
        or ("complete" in label_key and "date" in label_key)
        or (
            ("date" in label_key)
            and ("foundation" in label_key or "fdn" in label_key)
            and not has_start_token
        )
    )


def _detect_header(df_raw: pd.DataFrame) -> tuple[int | None, str | None, str | None, str | None, str | None, str | None]:
    best_row: int | None = None
    best_score = -1
    best_location = None
    best_start = None
    best_completion = None
    best_status = None
    best_gang = None

    max_rows = min(len(df_raw.index), 40)
    max_cols = min(len(df_raw.columns), 50)
    for row_idx in range(max_rows):
        row_labels = [_normalize_text(df_raw.iat[row_idx, col_idx]) for col_idx in range(max_cols)]
        location_col = None
        start_col = None
        completion_col = None
        status_col = None
        gang_col = None
        score = 0
        for col_idx, label in enumerate(row_labels):
            if not label:
                continue
            if location_col is None and (
                "location" in label
                or "loc no" in label
                or "tower no" in label
                or label == "loc"
            ):
                location_col = col_idx
                score += 3
            if start_col is None and _looks_like_start_label(label):
                start_col = col_idx
                score += 2
            if completion_col is None:
                if _looks_like_completion_label(label):
                    completion_col = col_idx
                    score += 6
                elif "completed on" in label:
                    completion_col = col_idx
                    score += 5
                elif "date" in label and not _looks_like_start_label(label):
                    completion_col = col_idx
                    score += 1
            if status_col is None and ("status" in label or "remark" in label):
                status_col = col_idx
                score += 1
            if gang_col is None and _looks_like_gang_label(label):
                gang_col = col_idx
                score += 1
        if location_col is not None and completion_col is not None and score > best_score:
            best_score = score
            best_row = row_idx
            best_location = f"c{location_col}"
            best_start = f"c{start_col}" if start_col is not None else None
            best_completion = f"c{completion_col}"
            best_status = f"c{status_col}" if status_col is not None else None
            best_gang = f"c{gang_col}" if gang_col is not None else None

    return best_row, best_location, best_start, best_completion, best_status, best_gang


def _extract_location_date_rowwise(
    values: list[object],
    *,
    report_ts: pd.Timestamp | pd.NaT,
) -> tuple[str, pd.Timestamp | pd.NaT, str]:
    location = ""
    event_date = pd.NaT
    status_bits: list[str] = []
    for value in values:
        text = _as_text(value)
        if text:
            status_bits.append(text)
        if not location and _looks_like_location(value):
            location = text
        if pd.isna(event_date):
            parsed = _coerce_date(value, report_ts=report_ts)
            if pd.notna(parsed):
                event_date = parsed
    return location, event_date, " | ".join(status_bits[:8])


def _build_exact_sheet_selector(configured_sheet: str):
    wanted_space = ingest.normalize_space_only(configured_sheet)
    wanted_compact = ingest.normalize_sheet_key(configured_sheet)

    def _selector(names: list[str]) -> str | None:
        if not names:
            return None
        for name in names:
            if ingest.normalize_space_only(name) == wanted_space:
                return name
        for name in names:
            if ingest.normalize_sheet_key(name) == wanted_compact:
                return name
        return None

    return _selector


def _pick_matching_workbooks_for_sheet(
    project_workbooks: list[Path],
    workbook_sheet_cache: dict[str, list[str]],
    configured_sheet: str,
    configured_file_identifier: str,
) -> list[tuple[Path, str]]:
    matches: list[tuple[Path, str]] = []
    wanted_space = ingest.normalize_space_only(configured_sheet)
    wanted_compact = ingest.normalize_sheet_key(configured_sheet)
    file_id_key = ingest.normalize_space_only(configured_file_identifier)

    for workbook in sorted(project_workbooks, key=lambda path: path.name):
        if file_id_key and file_id_key not in ingest.normalize_space_only(workbook.name):
            continue
        names = workbook_sheet_cache.get(str(workbook.resolve()), [])
        resolved = None
        for name in names:
            if ingest.normalize_space_only(name) == wanted_space:
                resolved = name
                break
        if resolved is None:
            for name in names:
                if ingest.normalize_sheet_key(name) == wanted_compact:
                    resolved = name
                    break
        if resolved:
            matches.append((workbook, resolved))
    return matches


def _resolve_template_column_indices(
    template_map: dict[int, str],
) -> tuple[int | None, int | None, list[int], int | None, int | None]:
    location_idx = None
    start_idx = None
    completion_indices: list[int] = []
    status_idx = None
    gang_idx = None

    for idx, raw_label in sorted(template_map.items()):
        label = _normalize_text(raw_label)
        label_key = _label_key(label)
        if not label:
            continue
        if location_idx is None and (
            "location" in label
            or "loc no" in label
            or "loc no" in label_key
            or "tower no" in label
            or label == "loc"
            or label_key == "loc"
        ):
            location_idx = int(idx)
        if start_idx is None and _looks_like_start_label(label):
            start_idx = int(idx)
        if _looks_like_completion_label(label):
            completion_indices.append(int(idx))
        if status_idx is None and ("status" in label or "remark" in label):
            status_idx = int(idx)
        if gang_idx is None and _looks_like_gang_label(label):
            gang_idx = int(idx)
    completion_indices = sorted({idx for idx in completion_indices})
    return location_idx, start_idx, completion_indices, status_idx, gang_idx


def _parse_foundation_sheet_with_template_map(
    df_raw: pd.DataFrame,
    *,
    project_code: str,
    project_display: str,
    project_scope_key: str,
    line_name: str,
    line_name_source: str,
    report_date: str,
    source_file: str,
    source_sheet: str,
    configured_sheet: str,
    fallback_note: str,
    template_map: dict[int, str],
    template_sheet: str,
) -> FoundationParseResult:
    remapped_df, template_changes = _apply_template_column_mapping(df_raw, template_map)
    location_idx, start_idx, completion_indices, status_idx, gang_idx = _resolve_template_column_indices(template_map)
    if location_idx is None or not completion_indices:
        return FoundationParseResult(
            raw_rows=[],
            completion_rows=[],
            parse_status="TEMPLATE_MAP_INCOMPLETE",
            parse_reason=(
                "Template map missing required location/completion-date column labels. "
                "Expected labels containing Location and Completion Date."
            ),
            parser_mode="template",
            rows_examined=int(len(df_raw.index)),
            template_sheet=template_sheet,
            template_applied=True,
            template_changes=tuple(template_changes),
        )

    base_record = {
        "project_code": project_code,
        "project_display": project_display,
        "project_scope_key": project_scope_key,
        "line_name": line_name,
        "line_name_source": line_name_source,
        "report_date": report_date,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "configured_sheet": configured_sheet,
        "source_type": "detail",
        "fallback_note": fallback_note or "",
    }

    report_ts = _parse_report_timestamp(report_date)
    raw_rows: list[dict[str, object]] = []
    completion_rows: list[dict[str, object]] = []
    max_col_count = len(remapped_df.columns)
    if location_idx >= max_col_count or any(idx >= max_col_count for idx in completion_indices):
        return FoundationParseResult(
            raw_rows=[],
            completion_rows=[],
            parse_status="TEMPLATE_INDEX_OUT_OF_RANGE",
            parse_reason=(
                f"Template map column index out of range for sheet width={max_col_count}. "
                f"location_idx={location_idx}, completion_indices={completion_indices}."
            ),
            parser_mode="template",
            rows_examined=int(len(df_raw.index)),
            template_sheet=template_sheet,
            template_applied=True,
            template_changes=tuple(template_changes),
        )

    for row_idx in range(len(remapped_df.index)):
        location_value = remapped_df.iat[row_idx, location_idx]
        location = _as_text(location_value)
        if not _looks_like_location(location):
            continue

        status_text = ""
        if status_idx is not None and status_idx < max_col_count:
            status_text = _as_text(remapped_df.iat[row_idx, status_idx])
        gang_name = ""
        if gang_idx is not None and gang_idx < max_col_count:
            gang_name = _as_text(remapped_df.iat[row_idx, gang_idx])
        start_date = pd.NaT
        if start_idx is not None and start_idx < max_col_count:
            start_date = _coerce_date(remapped_df.iat[row_idx, start_idx], report_ts=report_ts)
            if pd.notna(start_date):
                start_date = pd.Timestamp(start_date).normalize()
        event_date = pd.NaT
        for completion_idx in completion_indices:
            event_date = _coerce_date(remapped_df.iat[row_idx, completion_idx], report_ts=report_ts)
            if pd.notna(event_date):
                break
        if pd.notna(event_date):
            event_date = pd.Timestamp(event_date).normalize()

        raw_rows.append(
            {
                **base_record,
                "start_date": start_date if pd.notna(start_date) else pd.NaT,
                "event_date": event_date if pd.notna(event_date) else pd.NaT,
                "quality_flag": "detail_date" if pd.notna(event_date) else (
                    "detail_marker_no_date" if _looks_completed(status_text) else "detail_unparsed"
                ),
                "location_no": location,
                "gang_name": gang_name,
                "status_text": status_text,
                "cumulative_foundation": pd.NA,
            }
        )
        if pd.notna(event_date):
            completion_rows.append(
                {
                    **base_record,
                    "start_date": start_date if pd.notna(start_date) else pd.NaT,
                    "event_date": event_date,
                    "quality_flag": "detail_date",
                    "location_no": location,
                    "gang_name": gang_name,
                    "event_value": 1.0,
                    "cumulative_foundation": pd.NA,
                }
            )

    if completion_rows:
        return FoundationParseResult(
            raw_rows=raw_rows,
            completion_rows=completion_rows,
            parse_status="OK_TEMPLATE",
            parse_reason="",
            parser_mode="template",
            rows_examined=int(len(df_raw.index)),
            template_sheet=template_sheet,
            template_applied=True,
            template_changes=tuple(template_changes),
        )
    if raw_rows:
        return FoundationParseResult(
            raw_rows=raw_rows,
            completion_rows=[],
            parse_status="NO_COMPLETION_DATES_TEMPLATE",
            parse_reason="Template columns resolved but completion dates were not parseable.",
            parser_mode="template",
            rows_examined=int(len(df_raw.index)),
            template_sheet=template_sheet,
            template_applied=True,
            template_changes=tuple(template_changes),
        )
    return FoundationParseResult(
        raw_rows=[],
        completion_rows=[],
        parse_status="NO_FOUNDATION_ROWS_TEMPLATE",
        parse_reason="Template columns resolved but no foundation-like location rows were found.",
        parser_mode="template",
        rows_examined=int(len(df_raw.index)),
        template_sheet=template_sheet,
        template_applied=True,
        template_changes=tuple(template_changes),
    )


def _parse_foundation_sheet_dataframe(
    df_raw: pd.DataFrame,
    *,
    project_code: str,
    project_display: str,
    project_scope_key: str,
    line_name: str,
    line_name_source: str,
    report_date: str,
    source_file: str,
    source_sheet: str,
    configured_sheet: str,
    fallback_note: str,
    template_map: dict[int, str] | None = None,
    template_sheet: str = "",
) -> FoundationParseResult:
    if df_raw is None or df_raw.empty:
        return FoundationParseResult(
            raw_rows=[],
            completion_rows=[],
            parse_status="EMPTY_SHEET",
            parse_reason="Configured foundation sheet is empty.",
            parser_mode="none",
            rows_examined=0,
        )

    template_map = dict(template_map or {})
    if template_map:
        template_result = _parse_foundation_sheet_with_template_map(
            df_raw,
            project_code=project_code,
            project_display=project_display,
            project_scope_key=project_scope_key,
            line_name=line_name,
            line_name_source=line_name_source,
            report_date=report_date,
            source_file=source_file,
            source_sheet=source_sheet,
            configured_sheet=configured_sheet,
            fallback_note=fallback_note,
            template_map=template_map,
            template_sheet=template_sheet,
        )
        if template_result.parse_status not in {"TEMPLATE_MAP_INCOMPLETE", "TEMPLATE_INDEX_OUT_OF_RANGE"}:
            return template_result

    base_record = {
        "project_code": project_code,
        "project_display": project_display,
        "project_scope_key": project_scope_key,
        "line_name": line_name,
        "line_name_source": line_name_source,
        "report_date": report_date,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "configured_sheet": configured_sheet,
        "source_type": "detail",
        "fallback_note": fallback_note or "",
    }

    raw_rows: list[dict[str, object]] = []
    completion_rows: list[dict[str, object]] = []
    report_ts = _parse_report_timestamp(report_date)

    header_row, location_key, start_key, completion_key, status_key, gang_key = _detect_header(df_raw)
    rows_examined = int(len(df_raw.index))

    if header_row is not None and location_key and completion_key:
        labels = [_as_text(df_raw.iat[header_row, col_idx]) for col_idx in range(len(df_raw.columns))]
        labels = _make_unique_headers(labels)
        data = df_raw.iloc[header_row + 1 :].copy()
        data.columns = labels[: len(data.columns)]

        location_idx = int(location_key[1:])
        start_idx = int(start_key[1:]) if start_key else None
        completion_idx = int(completion_key[1:])
        status_idx = int(status_key[1:]) if status_key else None
        gang_idx = int(gang_key[1:]) if gang_key else None
        location_col = data.columns[location_idx] if location_idx < len(data.columns) else None
        start_col = data.columns[start_idx] if start_idx is not None and start_idx < len(data.columns) else None
        completion_col = data.columns[completion_idx] if completion_idx < len(data.columns) else None
        status_col = data.columns[status_idx] if status_idx is not None and status_idx < len(data.columns) else None
        gang_col = data.columns[gang_idx] if gang_idx is not None and gang_idx < len(data.columns) else None
        completion_label = _normalize_text(completion_col or "")

        location_candidates = 0
        parsed_date_candidates = 0
        for _, row in data.iterrows():
            location_value = row.get(location_col) if location_col else None
            location = _as_text(location_value)
            if not _looks_like_location(location):
                continue
            location_candidates += 1
            status_text = _as_text(row.get(status_col)) if status_col else ""
            gang_name = _as_text(row.get(gang_col)) if gang_col else ""
            start_date = _coerce_date(
                row.get(start_col),
                report_ts=report_ts,
            ) if start_col else pd.NaT
            if pd.notna(start_date):
                start_date = pd.Timestamp(start_date).normalize()
            event_date = _coerce_date(
                row.get(completion_col),
                report_ts=report_ts,
            ) if completion_col else pd.NaT
            if pd.notna(event_date):
                event_date = pd.Timestamp(event_date).normalize()
                parsed_date_candidates += 1
            cumulative_val = None
            for key in data.columns:
                norm = _normalize_text(key)
                if "cumulative" in norm and ("foundation" in norm or "progress" in norm):
                    cumulative_val = _coerce_numeric(row.get(key))
                    if cumulative_val is not None:
                        break

            raw_entry = {
                **base_record,
                "start_date": start_date if pd.notna(start_date) else pd.NaT,
                "event_date": event_date if pd.notna(event_date) else pd.NaT,
                "quality_flag": "detail_date" if pd.notna(event_date) else (
                    "detail_marker_no_date" if _looks_completed(status_text) else "detail_unparsed"
                ),
                "location_no": location,
                "gang_name": gang_name,
                "status_text": status_text,
                "cumulative_foundation": cumulative_val,
            }
            raw_rows.append(raw_entry)
            if pd.notna(event_date):
                completion_rows.append(
                    {
                        **base_record,
                        "start_date": start_date if pd.notna(start_date) else pd.NaT,
                        "event_date": event_date,
                        "quality_flag": "detail_date",
                        "location_no": location,
                        "gang_name": gang_name,
                        "event_value": 1.0,
                        "cumulative_foundation": pd.NA,
                    }
                )

        parse_ratio = (
            float(parsed_date_candidates) / float(location_candidates)
            if location_candidates > 0
            else 0.0
        )
        ambiguous_completion_label = bool(completion_label) and ("date" not in completion_label and "dt" not in completion_label)
        if location_candidates >= 8 and (parse_ratio < 0.25 or ambiguous_completion_label):
            return FoundationParseResult(
                raw_rows=raw_rows,
                completion_rows=[],
                parse_status="AMBIGUOUS_DATE_COLUMN",
                parse_reason=(
                    f"Detected location rows={location_candidates}, parseable completion dates={parsed_date_candidates} "
                    f"(ratio={parse_ratio:.2f}) from header '{completion_col}'."
                ),
                parser_mode="header",
                rows_examined=rows_examined,
            )

        if completion_rows:
            return FoundationParseResult(
                raw_rows=raw_rows,
                completion_rows=completion_rows,
                parse_status="OK",
                parse_reason="",
                parser_mode="header",
                rows_examined=rows_examined,
            )
        if raw_rows:
            return FoundationParseResult(
                raw_rows=raw_rows,
                completion_rows=[],
                parse_status="NO_COMPLETION_DATES",
                parse_reason="Location rows found but completion dates were not parseable.",
                parser_mode="header",
                rows_examined=rows_examined,
            )

    max_cols = min(len(df_raw.columns), 120)
    for row_idx in range(len(df_raw.index)):
        values = [df_raw.iat[row_idx, col_idx] for col_idx in range(max_cols)]
        location, event_date, status_text = _extract_location_date_rowwise(values, report_ts=report_ts)
        if not location:
            continue
        quality = "detail_date" if pd.notna(event_date) else (
            "detail_marker_no_date" if _looks_completed(status_text) else "detail_unparsed"
        )
        raw_entry = {
            **base_record,
            "start_date": pd.NaT,
            "event_date": event_date if pd.notna(event_date) else pd.NaT,
            "quality_flag": quality,
            "location_no": location,
            "gang_name": "",
            "status_text": status_text,
            "cumulative_foundation": pd.NA,
        }
        raw_rows.append(raw_entry)
        if pd.notna(event_date):
            completion_rows.append(
                {
                    **base_record,
                    "start_date": pd.NaT,
                    "event_date": pd.Timestamp(event_date).normalize(),
                    "quality_flag": "detail_date",
                    "location_no": location,
                    "gang_name": "",
                    "event_value": 1.0,
                    "cumulative_foundation": pd.NA,
                }
            )

    if completion_rows:
        return FoundationParseResult(
            raw_rows=raw_rows,
            completion_rows=completion_rows,
            parse_status="OK_ROWWISE",
            parse_reason="",
            parser_mode="rowwise",
            rows_examined=rows_examined,
        )
    if raw_rows:
        return FoundationParseResult(
            raw_rows=raw_rows,
            completion_rows=[],
            parse_status="NO_COMPLETION_DATES",
            parse_reason="Foundation-like rows found but no parseable completion dates.",
            parser_mode="rowwise",
            rows_examined=rows_examined,
        )

    return FoundationParseResult(
        raw_rows=[],
        completion_rows=[],
        parse_status="NO_FOUNDATION_ROWS",
        parse_reason="No parseable foundation rows found in configured sheet.",
        parser_mode="none",
        rows_examined=rows_examined,
    )


def _status_candidates(input_dir: Optional[Path], files: Optional[list[Path]]) -> list[Path]:
    if files:
        return [
            path
            for path in files
            if path.suffix.lower() in (".xlsx", ".xlsm", ".xls")
            and path.exists()
            and not path.name.startswith("~$")
        ]
    if input_dir and input_dir.exists():
        return sorted(
            [
                path
                for path in input_dir.rglob("*.xls*")
                if path.is_file() and not path.name.startswith("~$")
            ]
        )
    return []


def load_foundation_sheet_config(raw_root: Path, *, repo_root: Path | None = None) -> dict[str, list[dict[str, str]]]:
    config_path = ingest.resolve_dpr_config_path(raw_root, repo_root=repo_root)
    if config_path is None:
        return {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception:
        return {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}
        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}
        headers = [ingest.normalize_space_only(value) for value in header_row]
        if "project code" not in headers or "foundation sheet names" not in headers:
            return {}
        project_idx = headers.index("project code")
        foundation_idx = headers.index("foundation sheet names")
        line_idx = headers.index("foundation line names") if "foundation line names" in headers else None
        file_id_idx = headers.index("status file identifier") if "status file identifier" in headers else None

        mapping: dict[str, list[dict[str, str]]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project_val = row[project_idx] if project_idx < len(row) else None
            if project_val in (None, ""):
                continue
            project_key = ingest.normalize_project_code_key(project_val)
            raw_foundation = row[foundation_idx] if foundation_idx < len(row) else None
            if raw_foundation in (None, ""):
                mapping[project_key] = []
                continue
            raw_line_names = row[line_idx] if line_idx is not None and line_idx < len(row) else None
            raw_file_ids = row[file_id_idx] if file_id_idx is not None and file_id_idx < len(row) else None
            entries = parse_sheet_line_entries(
                raw_foundation,
                raw_line_names,
                "foundation",
                infer_from_sheet_name=False,
            )
            if raw_file_ids:
                file_id_parts = _split_tokens(raw_file_ids)
                for idx, entry in enumerate(entries):
                    entry["file_identifier"] = file_id_parts[idx] if idx < len(file_id_parts) else ""
            deduped_entries: list[dict[str, str]] = []
            seen_entry_keys: set[tuple[str, str, str]] = set()
            for entry in entries:
                sheet_key = ingest.normalize_space_only(entry.get("sheet_name"))
                if not sheet_key:
                    continue
                line_key = ingest.normalize_space_only(entry.get("line_name"))
                file_key = ingest.normalize_space_only(entry.get("file_identifier"))
                entry_key = (sheet_key, line_key, file_key)
                if entry_key in seen_entry_keys:
                    continue
                seen_entry_keys.add(entry_key)
                entry.setdefault("file_identifier", "")
                deduped_entries.append(entry)
            mapping[project_key] = deduped_entries
        return mapping
    finally:
        wb.close()


def load_foundation_template_mapping_catalog(
    raw_root: Path,
    *,
    repo_root: Path | None = None,
    include_unchecked: bool = False,
) -> tuple[dict[str, list[dict[str, object]]], dict[str, str]]:
    config_path = ingest.resolve_dpr_config_path(raw_root, repo_root=repo_root)
    if config_path is None:
        return {}, {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception:
        return {}, {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}, {}
        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}, {}
        headers = [ingest.normalize_space_only(value) for value in header_row]
        project_idx = headers.index("project code") if "project code" in headers else None
        check_idx = None
        for candidate in ("foundation template check", "foundation"):
            if candidate in headers:
                check_idx = headers.index(candidate)
                break
        if project_idx is None or (check_idx is None and not include_unchecked):
            return {}, {}

        catalog: dict[str, list[dict[str, object]]] = {}
        errors: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project_val = row[project_idx] if project_idx < len(row) else None
            if project_val in (None, ""):
                continue
            check_val = row[check_idx] if check_idx is not None and check_idx < len(row) else None
            check_enabled = ingest.normalize_space_only(check_val) == "yes"
            if not check_enabled and not include_unchecked:
                continue

            project_key = ingest.normalize_project_code_key(project_val)
            template_sheets = _resolve_project_template_sheets(wb, project_val, "Foundation")
            if not template_sheets:
                if check_enabled:
                    errors[project_key] = (
                        f"Foundation Template Check is Yes but no mapping tab matching project "
                        f"'{str(project_val).strip()}' was found."
                    )
                continue

            options: list[dict[str, object]] = []
            for sheet_name in template_sheets:
                ws_template = wb[sheet_name]
                col_map = _extract_template_column_map(ws_template)
                if not col_map:
                    continue
                options.append(
                    {
                        "column_map": col_map,
                        "template_sheet": sheet_name,
                        "numeric_tokens": _extract_numeric_tokens(sheet_name),
                    }
                )
            if not options:
                if check_enabled:
                    errors[project_key] = (
                        f"Foundation template tab(s) for project '{str(project_val).strip()}' have no usable "
                        f"'To Map' mapping row."
                    )
                continue
            catalog[project_key] = options
        return catalog, errors
    finally:
        wb.close()


def select_foundation_template_for_sheet(
    template_options: list[dict[str, object]] | None,
    *,
    configured_sheet_name: str = "",
    resolved_sheet_name: str = "",
    line_name: str = "",
) -> dict[str, object] | None:
    if not template_options:
        return None
    hints = [configured_sheet_name or "", resolved_sheet_name or "", line_name or ""]
    hint_keys = [ingest.normalize_space_only(value) for value in hints if ingest.normalize_space_only(value)]
    hint_numbers: set[str] = set()
    for value in hints:
        hint_numbers.update(_extract_numeric_tokens(value))

    best: dict[str, object] | None = None
    best_score = float("-inf")
    for idx, option in enumerate(template_options):
        sheet_name = str(option.get("template_sheet", "")).strip()
        sheet_key = ingest.normalize_space_only(sheet_name)
        sheet_numbers = set(option.get("numeric_tokens", set()) or set())
        col_map = option.get("column_map", {}) or {}
        score = 0.0
        for hint in hint_keys:
            if sheet_key == hint:
                score += 1000.0
            elif hint and (hint in sheet_key or sheet_key in hint):
                score += 200.0
        if hint_numbers:
            score += float(len(sheet_numbers & hint_numbers)) * 120.0
        score += float(len(col_map))
        score -= idx * 1e-4
        if score > best_score:
            best_score = score
            best = option
    return best


def _load_status_raw_for_fallback(output_path: Path, repo_root: Path | None = None) -> pd.DataFrame:
    candidates: list[Path] = []
    foundation_root = Path(output_path).parent
    if foundation_root.name.lower() == "foundation":
        status_root = foundation_root.parent / "ProgressStatus"
    else:
        status_root = foundation_root / "ProgressStatus"
    candidates.append(status_root / "RawData.parquet")
    candidates.append(status_root / "ProgressStatus_Output.xlsx")
    if repo_root is not None:
        candidates.append(Path(repo_root) / "Parquets" / "ProgressStatus" / "RawData.parquet")
        candidates.append(Path(repo_root) / "Parquets" / "ProgressStatus" / "ProgressStatus_Output.xlsx")

    for candidate in candidates:
        try:
            if not candidate.exists():
                continue
            if candidate.suffix.lower() == ".parquet":
                frame = pd.read_parquet(candidate)
                if isinstance(frame, pd.DataFrame) and not frame.empty:
                    return frame
            elif candidate.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
                with pd.ExcelFile(candidate) as xl:
                    if "RawData" in xl.sheet_names:
                        frame = xl.parse("RawData")
                        if isinstance(frame, pd.DataFrame) and not frame.empty:
                            return frame
        except Exception:
            continue
    return pd.DataFrame()


def _build_snapshot_rows_from_status(
    status_raw: pd.DataFrame,
    *,
    project_key: str,
    project_code: str,
    project_display: str,
    source_file_hint: str,
    configured_sheet_text: str,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    if status_raw is None or status_raw.empty:
        return [], []

    work = status_raw.copy()
    for required in ("project_code", "activity_norm", "report_date"):
        if required not in work.columns:
            return [], []

    work["project_key"] = work["project_code"].map(ingest.normalize_project_code_key)
    work["activity_key"] = work["activity_norm"].map(_normalize_key)
    work = work[(work["project_key"] == project_key) & (work["activity_key"] == "foundation")]
    if work.empty:
        return [], []

    work["report_date"] = pd.to_datetime(work.get("report_date"), errors="coerce").dt.normalize()
    work["cumulative_progress"] = pd.to_numeric(work.get("cumulative_progress"), errors="coerce")
    if "quantity_primary" in work.columns:
        quantity_primary = pd.to_numeric(work.get("quantity_primary"), errors="coerce")
        work["cumulative_progress"] = work["cumulative_progress"].fillna(quantity_primary)
    work = work.dropna(subset=["report_date", "cumulative_progress"])
    if work.empty:
        return [], []

    # Status snapshots can contain duplicate foundation rows for the same report date/sheet.
    # Drop exact duplicates before rollup to avoid inflated cumulative values.
    dedupe_cols = [
        "report_date",
        "source_sheet",
        "configured_sheet",
        "source_file",
        "cumulative_progress",
    ]
    available_dedupe_cols = [col for col in dedupe_cols if col in work.columns]
    if available_dedupe_cols:
        work = work.drop_duplicates(subset=available_dedupe_cols, keep="last")

    grouped = (
        work.groupby("report_date", as_index=False)
        .agg(
            cumulative_foundation=("cumulative_progress", "sum"),
            source_sheet=("source_sheet", lambda values: "; ".join(sorted({str(v).strip() for v in values if str(v).strip()}))),
            configured_sheet=("configured_sheet", lambda values: "; ".join(sorted({str(v).strip() for v in values if str(v).strip()}))),
            source_file=("source_file", lambda values: "; ".join(sorted({str(v).strip() for v in values if str(v).strip()}))),
        )
        .sort_values("report_date")
    )

    raw_rows: list[dict[str, object]] = []
    completion_rows: list[dict[str, object]] = []
    for _, row in grouped.iterrows():
        report_date = pd.Timestamp(row["report_date"]).normalize()
        source_sheet = str(row.get("source_sheet", "")).strip()
        configured_sheet = str(row.get("configured_sheet", "")).strip() or configured_sheet_text
        source_file = str(row.get("source_file", "")).strip() or source_file_hint
        base = {
            "project_code": project_code,
            "project_display": project_display,
            "project_scope_key": build_project_scope_key(project_code, "", project_display),
            "line_name": "",
            "line_name_source": "",
            "report_date": report_date.strftime("%Y-%m-%d"),
            "source_file": source_file,
            "source_sheet": source_sheet,
            "configured_sheet": configured_sheet,
            "source_type": "snapshot_fallback",
            "location_no": "",
            "gang_name": "",
            "status_text": "Derived from ProgressStatus foundation cumulative snapshot.",
            "fallback_note": "snapshot_from_progress_status",
        }
        raw_rows.append(
            {
                **base,
                "start_date": pd.NaT,
                "event_date": report_date,
                "quality_flag": "snapshot_cumulative",
                "cumulative_foundation": float(row["cumulative_foundation"]),
            }
        )
        completion_rows.append(
            {
                **base,
                "start_date": pd.NaT,
                "event_date": report_date,
                "quality_flag": "snapshot_cumulative",
                "event_value": pd.NA,
                "cumulative_foundation": float(row["cumulative_foundation"]),
            }
        )
    return raw_rows, completion_rows


def _format_date_bounds(values: pd.Series) -> tuple[str, str]:
    parsed = pd.to_datetime(values, errors="coerce").dropna()
    if parsed.empty:
        return "", ""
    return parsed.min().strftime("%Y-%m-%d"), parsed.max().strftime("%Y-%m-%d")


def compile_foundation_to_workbook(
    input_dir: Optional[Path],
    files: Optional[list[Path]],
    output_path: Path,
    *,
    repo_root: Path | None = None,
    completed_project_keys: set[str] | None = None,
    status_raw: pd.DataFrame | None = None,
) -> Path | None:
    candidates = _status_candidates(input_dir, files)
    if completed_project_keys:
        kept: list[Path] = []
        skipped_files = 0
        skipped_projects: set[str] = set()
        for workbook in candidates:
            identity = parse_project_identity_from_filename(workbook.name)
            project_code = str(identity.get("project_code", "")).strip() or workbook.stem
            if project_code and is_completed_project(project_code, completed_project_keys):
                skipped_files += 1
                skipped_projects.add(project_code.upper())
                continue
            kept.append(workbook)
        if skipped_files:
            print(
                f"[pipeline] Foundation: skipped_completed_files={skipped_files}, "
                f"skipped_completed_projects={len(skipped_projects)}"
            )
        candidates = kept

    if not candidates:
        print("[pipeline] Foundation: no candidate files found; skipping.")
        return None

    if input_dir is not None:
        raw_root = input_dir
    elif files:
        raw_root = files[0].parent
    else:
        raw_root = Path(".")

    foundation_sheet_config = load_foundation_sheet_config(raw_root, repo_root=repo_root)
    has_foundation_config = bool(foundation_sheet_config)
    foundation_template_catalog, foundation_template_errors = load_foundation_template_mapping_catalog(
        raw_root,
        repo_root=repo_root,
    )
    foundation_template_all_catalog, _ = load_foundation_template_mapping_catalog(
        raw_root,
        repo_root=repo_root,
        include_unchecked=True,
    )

    workbook_sheet_cache: dict[str, list[str]] = {}
    workbooks_by_project: dict[str, list[Path]] = {}
    for workbook in candidates:
        identity = parse_project_identity_from_filename(workbook.name)
        project_code = str(identity.get("project_code", "")).strip() or workbook.stem
        project_key = ingest.normalize_project_code_key(project_code)
        workbooks_by_project.setdefault(project_key, []).append(workbook)
        sheet_names, _ = ingest.list_excel_sheet_names(workbook)
        workbook_sheet_cache[str(workbook.resolve())] = list(sheet_names or [])

    if status_raw is None:
        status_raw = _load_status_raw_for_fallback(Path(output_path), repo_root=repo_root)

    raw_rows_all: list[dict[str, object]] = []
    completion_rows_all: list[dict[str, object]] = []
    diagnostics_rows: list[dict[str, object]] = []
    issue_rows: list[dict[str, object]] = []
    coverage_rows: list[dict[str, object]] = []

    skipped_not_in_config = 0
    skipped_blank_config = 0
    fallback_used = 0
    template_error_logged: set[str] = set()

    for project_key, project_workbooks in sorted(workbooks_by_project.items()):
        identity = parse_project_identity_from_filename(project_workbooks[0].name)
        project_code = str(identity.get("project_code", "")).strip() or project_workbooks[0].stem
        base_project_display = project_code
        configured_entries_opt = foundation_sheet_config.get(project_key)

        if has_foundation_config and configured_entries_opt is None:
            skipped_not_in_config += len(project_workbooks)
            coverage_rows.append(
                {
                    "project_code": project_code,
                    "project_display": base_project_display,
                    "status": "SKIPPED_NOT_IN_CONFIG",
                    "reason_code": "SKIPPED_NOT_IN_CONFIG",
                    "reason": "Project not listed in DPR_Config foundation mapping.",
                    "source_used": "missing",
                    "snapshot_limited": "No",
                    "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                    "configured_sheet": "",
                    "resolved_sheet": "",
                    "detail_rows": 0,
                    "detail_completions": 0,
                    "snapshot_rows": 0,
                    "first_event_date": "",
                    "last_event_date": "",
                    "first_report_date": "",
                    "last_report_date": "",
                    "available_sheets": "",
                }
            )
            continue

        configured_entries = configured_entries_opt if configured_entries_opt is not None else []
        explicit_blank_skip = bool(
            has_foundation_config
            and configured_entries_opt is not None
            and not configured_entries
        )
        if explicit_blank_skip:
            skipped_blank_config += len(project_workbooks)
            coverage_rows.append(
                {
                    "project_code": project_code,
                    "project_display": base_project_display,
                    "status": "SKIPPED_BLANK_CONFIG",
                    "reason_code": "SKIPPED_BLANK_CONFIG",
                    "reason": "Foundation sheet mapping left blank in DPR_Config; project explicitly skipped.",
                    "source_used": "missing",
                    "snapshot_limited": "No",
                    "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                    "configured_sheet": "",
                    "resolved_sheet": "",
                    "detail_rows": 0,
                    "detail_completions": 0,
                    "snapshot_rows": 0,
                    "first_event_date": "",
                    "last_event_date": "",
                    "first_report_date": "",
                    "last_report_date": "",
                    "available_sheets": "; ".join(
                        sorted(
                            {
                                sheet
                                for workbook in project_workbooks
                                for sheet in workbook_sheet_cache.get(str(workbook.resolve()), [])
                            }
                        )
                    ),
                }
            )
            continue

        project_raw_rows: list[dict[str, object]] = []
        project_completion_rows: list[dict[str, object]] = []
        project_issue_no_target = False
        configured_sheet_tokens: list[str] = []
        project_template_options = foundation_template_catalog.get(project_key)
        project_template_all_options = foundation_template_all_catalog.get(project_key)
        project_template_error = foundation_template_errors.get(project_key)
        if project_template_error and project_key not in template_error_logged:
            template_error_logged.add(project_key)
            issue_rows.append(
                {
                    "Workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                    "Project": project_code,
                    "Sheet": "",
                    "ConfiguredSheet": "; ".join(
                        [
                            str(entry.get("sheet_name", "")).strip()
                            for entry in (configured_entries or [])
                            if str(entry.get("sheet_name", "")).strip()
                        ]
                    ),
                    "LineName": "",
                    "LineNameSource": "",
                    "Issue": "TEMPLATE_CONFIG_ERROR",
                    "Reason": project_template_error,
                }
            )

        for request in configured_entries:
            configured_sheet = str(request.get("sheet_name", "")).strip()
            configured_line_name = normalize_line_name(request.get("line_name", ""))
            configured_line_source = str(request.get("line_name_source", "")).strip()
            configured_file_identifier = str(request.get("file_identifier", "")).strip()
            configured_sheet_tokens.append(configured_sheet)

            selected_workbooks = _pick_matching_workbooks_for_sheet(
                project_workbooks,
                workbook_sheet_cache,
                configured_sheet,
                configured_file_identifier,
            )
            if not selected_workbooks:
                project_issue_no_target = True
                issue_rows.append(
                    {
                        "Workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                        "Project": project_code,
                        "Sheet": "",
                        "ConfiguredSheet": configured_sheet,
                        "LineName": configured_line_name,
                        "LineNameSource": configured_line_source,
                        "Issue": "NO_TARGET_SHEET",
                        "Reason": "Configured foundation sheet not found in project workbooks.",
                    }
                )
                continue

            selector = _build_exact_sheet_selector(configured_sheet)
            for selected_workbook, resolved_sheet in selected_workbooks:
                selected_identity = parse_project_identity_from_filename(selected_workbook.name)
                line_name = configured_line_name or normalize_line_name(selected_identity.get("line_name", ""))
                line_source = configured_line_source or ("config" if configured_line_name else "filename")
                project_display = build_project_display(project_code, line_name, base_project_display) or base_project_display
                project_scope_key = build_project_scope_key(project_code, line_name, project_display)
                report_date = _extract_report_date_from_filename(selected_workbook.name)
                fallback_note = ""
                selected_template = select_foundation_template_for_sheet(
                    project_template_options,
                    configured_sheet_name=configured_sheet,
                    resolved_sheet_name=resolved_sheet,
                    line_name=line_name,
                )
                template_fallback_used = False
                if selected_template is None:
                    selected_template = select_foundation_template_for_sheet(
                        project_template_all_options,
                        configured_sheet_name=configured_sheet,
                        resolved_sheet_name=resolved_sheet,
                        line_name=line_name,
                    )
                    template_fallback_used = selected_template is not None
                template_map = dict(selected_template.get("column_map", {}) if selected_template else {})
                template_sheet_name = str(selected_template.get("template_sheet", "")).strip() if selected_template else ""

                try:
                    df_raw, resolved_sheet_loaded, fallback_note = load_sheet_with_csv_fallback(
                        selected_workbook,
                        selector,
                        read_excel_kwargs={"header": None},
                        read_csv_kwargs={"header": None},
                    )
                except Exception as exc:
                    issue_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Sheet": resolved_sheet,
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "Issue": "READ_FAIL",
                            "Reason": str(exc),
                        }
                    )
                    diagnostics_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Sheet": resolved_sheet,
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "Rows": 0,
                            "DetailRows": 0,
                            "Completions": 0,
                            "ParserMode": "none",
                            "FallbackNote": "",
                            "Status": "READ_FAIL",
                            "Reason": str(exc),
                            "TemplateSheet": template_sheet_name,
                            "TemplateApplied": bool(template_map),
                            "TemplateChanges": "",
                        }
                    )
                    continue

                resolved_sheet_final = resolved_sheet_loaded or resolved_sheet
                effective_fallback_note = fallback_note or ""
                if template_fallback_used:
                    extra = f"template fallback used: {template_sheet_name}"
                    effective_fallback_note = f"{effective_fallback_note}; {extra}".strip("; ").strip()
                parse_result = _parse_foundation_sheet_dataframe(
                    df_raw if isinstance(df_raw, pd.DataFrame) else pd.DataFrame(),
                    project_code=project_code,
                    project_display=project_display,
                    project_scope_key=project_scope_key,
                    line_name=line_name,
                    line_name_source=line_source,
                    report_date=report_date,
                    source_file=selected_workbook.name,
                    source_sheet=resolved_sheet_final,
                    configured_sheet=configured_sheet,
                    fallback_note=effective_fallback_note,
                    template_map=template_map,
                    template_sheet=template_sheet_name,
                )
                project_raw_rows.extend(parse_result.raw_rows)
                project_completion_rows.extend(parse_result.completion_rows)
                diagnostics_rows.append(
                    {
                        "Workbook": selected_workbook.name,
                        "Project": project_code,
                        "Sheet": resolved_sheet_final,
                        "ConfiguredSheet": configured_sheet,
                        "LineName": line_name,
                        "LineNameSource": line_source,
                        "Rows": parse_result.rows_examined,
                        "DetailRows": len(parse_result.raw_rows),
                        "Completions": len(parse_result.completion_rows),
                        "ParserMode": parse_result.parser_mode,
                        "FallbackNote": effective_fallback_note,
                        "Status": parse_result.parse_status,
                        "Reason": parse_result.parse_reason,
                        "TemplateSheet": parse_result.template_sheet or template_sheet_name,
                        "TemplateApplied": bool(parse_result.template_applied or template_map),
                        "TemplateChanges": "; ".join(parse_result.template_changes or ()),
                    }
                )
                if parse_result.parse_status not in {"OK", "OK_ROWWISE", "OK_TEMPLATE"}:
                    issue_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Sheet": resolved_sheet_final,
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "Issue": parse_result.parse_status,
                            "Reason": parse_result.parse_reason,
                        }
                    )

        detail_rows = [row for row in project_raw_rows if str(row.get("source_type", "")).strip().lower() == "detail"]
        detail_completions = [row for row in project_completion_rows if str(row.get("source_type", "")).strip().lower() == "detail"]
        snapshot_rows, snapshot_completions = _build_snapshot_rows_from_status(
            status_raw if isinstance(status_raw, pd.DataFrame) else pd.DataFrame(),
            project_key=project_key,
            project_code=project_code,
            project_display=base_project_display,
            source_file_hint="; ".join(sorted({w.name for w in project_workbooks})),
            configured_sheet_text="; ".join(configured_sheet_tokens),
        )
        snapshot_latest = 0.0
        if snapshot_completions:
            snapshot_values = [
                _coerce_numeric(row.get("cumulative_foundation"))
                for row in snapshot_completions
            ]
            snapshot_values = [value for value in snapshot_values if value is not None]
            if snapshot_values:
                snapshot_latest = float(max(snapshot_values))
        if snapshot_completions and snapshot_latest <= 0:
            issue_rows.append(
                {
                    "Workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                    "Project": project_code,
                    "Sheet": "",
                    "ConfiguredSheet": "; ".join(configured_sheet_tokens),
                    "LineName": "",
                    "LineNameSource": "",
                    "Issue": "MAPPING_CONFIRMATION_REQUIRED",
                    "Reason": (
                        "Status snapshot fallback returned non-positive cumulative foundation value. "
                        "Please confirm status/foundation mapping for this project."
                    ),
                }
            )
            snapshot_rows = []
            snapshot_completions = []
            snapshot_latest = 0.0

        use_snapshot = False
        status_value = "OK_DETAIL" if detail_completions else "MISSING"
        reason = "Detail foundation completion dates parsed from configured sheet(s)." if detail_completions else "No parseable foundation detail rows found."
        source_used = "detail" if detail_completions else "missing"
        reason_code = status_value
        snapshot_limited = "No"

        detail_completion_count = int(len(detail_completions))
        if detail_completion_count == 0 and snapshot_completions:
            use_snapshot = True
            status_value = "SNAPSHOT_FALLBACK"
            reason_code = "SNAPSHOT_FALLBACK"
            reason = "No parseable detail completion dates; used ProgressStatus cumulative foundation snapshots."
        elif detail_completion_count > 0 and snapshot_completions and snapshot_latest >= 10:
            coverage_ratio = detail_completion_count / snapshot_latest if snapshot_latest else 0.0
            if coverage_ratio < 0.70:
                status_value = "DETAIL_STATUS_MISMATCH"
                reason_code = "DETAIL_STATUS_MISMATCH"
                reason = (
                    "Detail completion dates were parsed, but counts are materially below status cumulative "
                    "foundation progress; keeping detail source and flagging mapping/status mismatch."
                )
                issue_rows.append(
                    {
                        "Workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                        "Project": project_code,
                        "Sheet": "",
                        "ConfiguredSheet": "; ".join(configured_sheet_tokens),
                        "LineName": "",
                        "LineNameSource": "",
                        "Issue": "MAPPING_CONFIRMATION_REQUIRED",
                        "Reason": (
                            f"Detail completion rows ({detail_completion_count}) are significantly below "
                            f"status snapshot cumulative ({snapshot_latest:.0f}). Please confirm foundation "
                            f"column mapping/status snapshot semantics."
                        ),
                    }
                )

        if use_snapshot:
            fallback_used += 1
            source_used = "snapshot_fallback"
            snapshot_limited = "Yes"
            project_completion_rows = [
                row for row in project_completion_rows
                if str(row.get("source_type", "")).strip().lower() != "detail"
            ]
            project_raw_rows.extend(snapshot_rows)
            project_completion_rows.extend(snapshot_completions)
            detail_completions = []
        elif detail_completion_count > 0 and not snapshot_completions and len(detail_rows) >= 8:
            coverage_ratio_no_snapshot = detail_completion_count / float(len(detail_rows))
            if coverage_ratio_no_snapshot < 0.25:
                status_value = "MAPPING_CONFIRMATION_REQUIRED"
                reason_code = "MAPPING_CONFIRMATION_REQUIRED"
                reason = (
                    "Detail rows are present but very few parseable completion dates were found. "
                    "Foundation column mapping likely needs confirmation."
                )
                source_used = "missing"
                snapshot_limited = "No"
                project_completion_rows = [
                    row for row in project_completion_rows
                    if str(row.get("source_type", "")).strip().lower() != "detail"
                ]
                detail_completions = []
                issue_rows.append(
                    {
                        "Workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                        "Project": project_code,
                        "Sheet": "",
                        "ConfiguredSheet": "; ".join(configured_sheet_tokens),
                        "LineName": "",
                        "LineNameSource": "",
                        "Issue": "MAPPING_CONFIRMATION_REQUIRED",
                        "Reason": (
                            f"Detail rows={len(detail_rows)}, parseable completion rows={detail_completion_count}. "
                            "Please confirm foundation completion-date column mapping."
                        ),
                    }
                )
        elif project_issue_no_target and detail_completion_count == 0 and not snapshot_completions:
            if project_code.strip().upper() == "TB 408":
                status_value = "BLOCKED_NO_SOURCE"
                reason_code = "BLOCKED_NO_SOURCE"
                reason = "Configured foundation sheet unavailable and no usable foundation/status fallback in raw DPR."
            else:
                status_value = "NO_TARGET_SHEET"
                reason_code = "NO_TARGET_SHEET"
                reason = "Configured foundation sheet not found in project workbook(s); no fallback snapshot available."
            source_used = "missing"
            snapshot_limited = "No"
        elif detail_completion_count == 0 and not snapshot_completions:
            status_value = "MISSING"
            reason_code = "MISSING"
            reason = "No parseable foundation detail/snapshot source found."
            source_used = "missing"
            snapshot_limited = "No"

        raw_rows_all.extend(project_raw_rows)
        completion_rows_all.extend(project_completion_rows)

        events = pd.DataFrame(project_completion_rows)
        first_event, last_event = ("", "")
        first_report, last_report = ("", "")
        if not events.empty:
            first_event, last_event = _format_date_bounds(events.get("event_date", pd.Series(dtype="object")))
            first_report, last_report = _format_date_bounds(events.get("report_date", pd.Series(dtype="object")))

        coverage_rows.append(
            {
                "project_code": project_code,
                "project_display": base_project_display,
                "status": status_value,
                "reason_code": reason_code,
                "reason": reason,
                "source_used": source_used,
                "snapshot_limited": snapshot_limited,
                "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                "configured_sheet": "; ".join(configured_sheet_tokens),
                "resolved_sheet": "; ".join(
                    sorted(
                        {
                            str(row.get("source_sheet", "")).strip()
                            for row in project_raw_rows
                            if str(row.get("source_sheet", "")).strip()
                        }
                    )
                ),
                "detail_rows": int(len(detail_rows)),
                "detail_completions": int(len(detail_completions)),
                "snapshot_rows": int(len(snapshot_completions)),
                "first_event_date": first_event,
                "last_event_date": last_event,
                "first_report_date": first_report,
                "last_report_date": last_report,
                "available_sheets": "; ".join(
                    sorted(
                        {
                            sheet
                            for workbook in project_workbooks
                            for sheet in workbook_sheet_cache.get(str(workbook.resolve()), [])
                        }
                    )
                ),
            }
        )

    if skipped_blank_config:
        print(f"[pipeline] Foundation: skipped_blank_config={skipped_blank_config} workbook(s) with blank foundation sheet config.")
    if skipped_not_in_config:
        print(f"[pipeline] Foundation: skipped_not_in_config={skipped_not_in_config} workbook(s) not listed in DPR_Config.")
    if fallback_used:
        print(f"[pipeline] Foundation: fallback_used={fallback_used} project(s) used snapshot fallback.")

    raw_df = pd.DataFrame(raw_rows_all, columns=FOUNDATION_RAW_COLUMNS)
    completions_df = pd.DataFrame(completion_rows_all, columns=FOUNDATION_COMPLETIONS_COLUMNS)
    if not raw_df.empty:
        raw_df["start_date"] = pd.to_datetime(raw_df["start_date"], errors="coerce").dt.normalize()
        raw_df["event_date"] = pd.to_datetime(raw_df["event_date"], errors="coerce").dt.normalize()
    if not completions_df.empty:
        completions_df["start_date"] = pd.to_datetime(completions_df["start_date"], errors="coerce").dt.normalize()
        completions_df["event_date"] = pd.to_datetime(completions_df["event_date"], errors="coerce").dt.normalize()
        completions_df["report_date"] = pd.to_datetime(completions_df["report_date"], errors="coerce").dt.normalize()
        completions_df["location_no"] = completions_df["location_no"].fillna("").astype(str).str.strip()
        completions_df["line_name"] = completions_df["line_name"].fillna("").astype(str).str.strip()
        completions_df["project_key"] = completions_df["project_code"].map(ingest.normalize_project_code_key)
        completions_df["line_key"] = completions_df["line_name"].map(_normalize_key)
        completions_df["location_key"] = completions_df["location_no"].map(_normalize_key)

        detail_mask = completions_df["source_type"].astype(str).str.strip().str.lower() == "detail"
        detail = completions_df[detail_mask].copy()
        if not detail.empty:
            detail = detail.dropna(subset=["event_date"])
            detail = detail.drop_duplicates(
                subset=["project_key", "line_key", "location_key", "event_date"],
                keep="last",
            )

        snapshot = completions_df[~detail_mask].copy()
        if not snapshot.empty:
            snapshot = snapshot.dropna(subset=["event_date"])
            snapshot["cumulative_foundation"] = pd.to_numeric(snapshot["cumulative_foundation"], errors="coerce")
            snapshot = (
                snapshot.sort_values("event_date")
                .groupby(["project_key", "event_date"], as_index=False)
                .agg(
                    project_code=("project_code", "first"),
                    project_display=("project_display", "first"),
                    project_scope_key=("project_scope_key", "first"),
                    line_name=("line_name", "first"),
                    line_name_source=("line_name_source", "first"),
                    report_date=("report_date", "first"),
                    start_date=("start_date", "first"),
                    source_file=("source_file", "first"),
                    source_sheet=("source_sheet", "first"),
                    configured_sheet=("configured_sheet", "first"),
                    source_type=("source_type", "first"),
                    quality_flag=("quality_flag", "first"),
                    location_no=("location_no", "first"),
                    gang_name=("gang_name", "first"),
                    event_value=("event_value", "first"),
                    cumulative_foundation=("cumulative_foundation", "max"),
                )
            )

        completions_df = pd.concat([detail, snapshot], ignore_index=True, sort=False)
        completions_df["period_week_start"] = completions_df["event_date"] - pd.to_timedelta(
            (completions_df["event_date"].dt.weekday + 1) % 7, unit="D"
        )
        completions_df["period_month"] = completions_df["event_date"].dt.to_period("M").dt.to_timestamp()
        completions_df = completions_df.reindex(columns=FOUNDATION_COMPLETIONS_COLUMNS)
        completions_df = completions_df.sort_values(["project_code", "event_date", "line_name", "location_no"])
    else:
        completions_df = pd.DataFrame(columns=FOUNDATION_COMPLETIONS_COLUMNS)

    diagnostics_df = pd.DataFrame(diagnostics_rows, columns=DIAGNOSTICS_COLUMNS)
    issues_df = pd.DataFrame(issue_rows, columns=ISSUES_COLUMNS)
    coverage_df = pd.DataFrame(coverage_rows, columns=COVERAGE_COLUMNS)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output_path.with_suffix(f"{output_path.suffix}.tmp")

    try:
        with pd.ExcelWriter(temp_output, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name="FoundationRaw", index=False)
            completions_df.to_excel(writer, sheet_name="FoundationCompletions", index=False)
            coverage_df.to_excel(writer, sheet_name="Coverage", index=False)
            diagnostics_df.to_excel(writer, sheet_name="Diagnostics", index=False)
            issues_df.to_excel(writer, sheet_name="Issues", index=False)
        temp_output.replace(output_path)
    finally:
        if temp_output.exists():
            try:
                temp_output.unlink()
            except Exception:
                pass

    print(
        f"[pipeline] Foundation: wrote workbook {output_path} "
        f"(raw_rows={len(raw_df.index)}, completion_rows={len(completions_df.index)}, coverage_rows={len(coverage_df.index)})."
    )
    return output_path
