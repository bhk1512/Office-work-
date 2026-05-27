"""Shared stringing ingestion helpers used by pipeline and dashboard loaders."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
import re
import xml.etree.ElementTree as ET
import zipfile

import pandas as pd
from openpyxl import load_workbook

from erection_compiled_to_daily_new import load_sheet_with_csv_fallback
from .project_identity import parse_sheet_line_entries
from .stringing import find_stringing_header_row


def normalize_project_code_key(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def normalize_space_only(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def normalize_sheet_key(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def resolve_dpr_config_path(raw_root: Path, *, repo_root: Path | None = None) -> Path | None:
    candidate = raw_root.parent / "DPR_Config.xlsx"
    if candidate.exists():
        return candidate
    if repo_root is not None:
        fallback = repo_root / "Raw Data" / "DPR_Config.xlsx"
        if fallback.exists():
            return fallback
    return None


def load_stringing_sheet_config(raw_root: Path, *, repo_root: Path | None = None) -> dict[str, list[dict[str, str]]]:
    config_path = resolve_dpr_config_path(raw_root, repo_root=repo_root)
    if config_path is None:
        return {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception:
        return {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}
        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}
        headers = [normalize_space_only(v) for v in header_row]
        if "project code" not in headers or "stringing sheet names" not in headers:
            return {}
        project_idx = headers.index("project code")
        stringing_idx = headers.index("stringing sheet names")
        line_idx = headers.index("stringing line names") if "stringing line names" in headers else None

        mapping: dict[str, list[dict[str, str]]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project_val = row[project_idx] if project_idx < len(row) else None
            if project_val in (None, ""):
                continue
            project_key = normalize_project_code_key(project_val)
            raw_stringing = row[stringing_idx] if stringing_idx < len(row) else None
            if raw_stringing in (None, ""):
                mapping[project_key] = []
                continue

            raw_line_names = row[line_idx] if line_idx is not None and line_idx < len(row) else None
            entries = parse_sheet_line_entries(
                raw_stringing,
                raw_line_names,
                "stringing",
                infer_from_sheet_name=False,
            )
            deduped_entries: list[dict[str, str]] = []
            seen_sheet_keys: set[str] = set()
            for entry in entries:
                key = normalize_space_only(entry.get("sheet_name"))
                if not key or key in seen_sheet_keys:
                    continue
                seen_sheet_keys.add(key)
                deduped_entries.append(entry)
            mapping[project_key] = deduped_entries
        return mapping
    finally:
        wb.close()


def _resolve_named_template_sheet(wb, expected_name: str) -> Optional[str]:
    expected_key = normalize_space_only(expected_name)
    for name in wb.sheetnames:
        if normalize_space_only(name) == expected_key:
            return name
    return None


def _resolve_project_template_sheets(wb, project_name: object, discipline: str) -> List[str]:
    project_text = str(project_name or "").strip()
    if not project_text:
        return []

    resolved: List[str] = []
    seen: set[str] = set()
    for expected in (
        f"{project_text} {discipline}",
        f"{project_text} {discipline} Template Check",
    ):
        hit = _resolve_named_template_sheet(wb, expected)
        if hit and hit not in seen:
            seen.add(hit)
            resolved.append(hit)

    project_key = normalize_space_only(project_text)
    discipline_key = normalize_space_only(discipline)
    for name in wb.sheetnames:
        key = normalize_space_only(name)
        if key and key.startswith(project_key) and key.endswith(discipline_key) and name not in seen:
            seen.add(name)
            resolved.append(name)
    return resolved


def _numeric_tokens(value: object) -> set[str]:
    text = str(value or "")
    return {token for token in re.findall(r"\d{2,4}", text)}


def _extract_template_column_map(ws) -> Dict[int, str]:
    to_map_row = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if normalize_space_only(cell) == "to map":
                to_map_row = row_idx
                break
        if to_map_row is not None:
            break
    if to_map_row is None:
        return {}

    labels_row = to_map_row + 1
    row_values = next(ws.iter_rows(min_row=labels_row, max_row=labels_row, values_only=True), ())
    mapping: Dict[int, str] = {}
    for col_idx, value in enumerate(row_values):
        label = str(value).strip() if value is not None else ""
        if label:
            mapping[col_idx] = label
    return mapping


def load_stringing_template_mapping_catalog(
    raw_root: Path,
    *,
    repo_root: Path | None = None,
    include_unchecked: bool = False,
) -> Tuple[Dict[str, List[Tuple[Dict[int, str], str]]], Dict[str, str]]:
    """Return all usable template maps per project, preserving sheet specificity."""

    config_path = resolve_dpr_config_path(raw_root, repo_root=repo_root)
    if config_path is None:
        return {}, {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception:
        return {}, {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}, {}
        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}, {}

        headers = [normalize_space_only(value) for value in header_row]
        project_idx = headers.index("project code") if "project code" in headers else None
        check_idx = None
        for candidate in ("stringing template check", "stringing"):
            if candidate in headers:
                check_idx = headers.index(candidate)
                break
        if project_idx is None or (check_idx is None and not include_unchecked):
            return {}, {}

        catalog: Dict[str, List[Tuple[Dict[int, str], str]]] = {}
        errors: Dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project_val = row[project_idx] if project_idx < len(row) else None
            if project_val in (None, ""):
                continue
            check_val = row[check_idx] if check_idx is not None and check_idx < len(row) else None
            check_enabled = normalize_space_only(check_val) == "yes"
            if not check_enabled and not include_unchecked:
                continue

            project_key = normalize_project_code_key(project_val)
            template_sheets = _resolve_project_template_sheets(wb, project_val, "Stringing")
            if not template_sheets:
                if check_enabled:
                    errors[project_key] = (
                        f"Stringing Template Check is Yes but no mapping tab matching project '{str(project_val).strip()}' was found."
                    )
                continue

            options: List[Tuple[Dict[int, str], str]] = []
            for sheet_name in template_sheets:
                col_map = _extract_template_column_map(wb[sheet_name])
                if col_map:
                    options.append((col_map, sheet_name))
            if not options:
                if check_enabled:
                    errors[project_key] = (
                        f"Stringing template tab(s) for project '{str(project_val).strip()}' have no usable 'To Map' mapping row."
                    )
                continue
            catalog[project_key] = options
        return catalog, errors
    finally:
        wb.close()


def select_template_map_for_sheet(
    template_options: List[Tuple[Dict[int, str], str]] | None,
    *,
    configured_sheet_name: str = "",
    resolved_sheet_name: str = "",
    line_name: str = "",
) -> tuple[dict[int, str], str] | None:
    """Choose the best template map for a specific stringing sheet request."""

    if not template_options:
        return None

    hints = [configured_sheet_name or "", resolved_sheet_name or "", line_name or ""]
    hint_keys = [normalize_space_only(value) for value in hints if normalize_space_only(value)]
    hint_numbers: set[str] = set()
    for value in hints:
        hint_numbers.update(_numeric_tokens(value))

    best: tuple[dict[int, str], str] | None = None
    best_score = float("-inf")
    for idx, (col_map, sheet_name) in enumerate(template_options):
        sheet_key = normalize_space_only(sheet_name)
        sheet_numbers = _numeric_tokens(sheet_name)
        score = 0.0

        for hint in hint_keys:
            if sheet_key == hint:
                score += 1000.0
            elif hint and (hint in sheet_key or sheet_key in hint):
                score += 200.0
        if hint_numbers:
            score += float(len(sheet_numbers & hint_numbers)) * 120.0

        # Preserve previous fallback behavior (prefer richer maps) when hints are inconclusive.
        score += float(len(col_map))
        score -= idx * 1e-4

        if score > best_score:
            best_score = score
            best = (col_map, sheet_name)
    return best


def load_stringing_template_mapping_config(
    raw_root: Path,
    *,
    repo_root: Path | None = None,
    include_unchecked: bool = False,
) -> Tuple[Dict[str, Tuple[Dict[int, str], str]], Dict[str, str]]:
    catalog, errors = load_stringing_template_mapping_catalog(
        raw_root,
        repo_root=repo_root,
        include_unchecked=include_unchecked,
    )
    mappings: Dict[str, Tuple[Dict[int, str], str]] = {}
    for project_key, options in catalog.items():
        selected = select_template_map_for_sheet(options)
        if selected is not None:
            mappings[project_key] = selected
    return mappings, errors


def resolve_template_fallback_for_project(
    raw_root: Path,
    project_value: object,
    *,
    repo_root: Path | None = None,
) -> tuple[dict[int, str], str] | None:
    project_key = normalize_project_code_key(project_value)
    if not project_key:
        return None
    mappings, _ = load_stringing_template_mapping_config(
        raw_root,
        repo_root=repo_root,
        include_unchecked=True,
    )
    return mappings.get(project_key)


def apply_template_column_mapping(df: pd.DataFrame, template_map: Dict[int, str]) -> Tuple[pd.DataFrame, List[str]]:
    if df is None or df.empty or not template_map:
        return df, []
    remapped = df.copy()
    columns = list(remapped.columns)
    changes: List[str] = []
    for idx, mapped_name in sorted(template_map.items()):
        if idx >= len(columns):
            continue
        current = str(columns[idx]).strip()
        target = str(mapped_name).strip()
        if not target:
            continue
        columns[idx] = target
        if normalize_space_only(current) != normalize_space_only(target):
            changes.append(f"C{idx + 1}:{current}->{target}")
    remapped.columns = columns
    return remapped, changes


def resolve_project_sheet_name(sheet_names: Iterable[str], project_candidates: list[str]) -> str | None:
    by_space_key: dict[str, str] = {}
    by_sheet_key: dict[str, str] = {}
    for name in sheet_names:
        sheet_name = str(name)
        key = normalize_space_only(sheet_name)
        if key and key not in by_space_key:
            by_space_key[key] = sheet_name
        compact_key = normalize_sheet_key(sheet_name)
        if compact_key and compact_key not in by_sheet_key:
            by_sheet_key[compact_key] = sheet_name
    for candidate in project_candidates:
        hit = by_space_key.get(normalize_space_only(candidate))
        if hit:
            return hit
        hit = by_sheet_key.get(normalize_sheet_key(candidate))
        if hit:
            return hit
    return None


def _normalize_sheet_label(label: Optional[str]) -> str:
    if label is None:
        return ""
    lowered = str(label).lower().strip()
    return re.sub(r"[^a-z0-9]+", "", lowered)


def find_stringing_sheet_name_from_list(
    names: list[str],
    preferred: Optional[str],
    project_candidates: Optional[List[str]] = None,
) -> Optional[str]:
    if not names:
        return None
    if project_candidates:
        by_space_key: Dict[str, str] = {}
        by_sheet_key: Dict[str, str] = {}
        for name in names:
            key = normalize_space_only(name)
            if key and key not in by_space_key:
                by_space_key[key] = name
            compact_key = normalize_sheet_key(name)
            if compact_key and compact_key not in by_sheet_key:
                by_sheet_key[compact_key] = name
        for candidate in project_candidates:
            hit = by_space_key.get(normalize_space_only(candidate))
            if hit:
                return hit
            hit = by_sheet_key.get(normalize_sheet_key(candidate))
            if hit:
                return hit
        candidate_keys = {normalize_sheet_key(candidate) for candidate in project_candidates}
        if candidate_keys <= {"stringing", ""}:
            stringing_sheets = [
                name
                for name in names
                if "stringing" in normalize_space_only(name)
            ]
            compiled_sheets = [
                name
                for name in stringing_sheets
                if "compiled" in normalize_space_only(name)
            ]
            if len(stringing_sheets) == 1 and len(compiled_sheets) == 1:
                return compiled_sheets[0]
        return None
    if preferred:
        target = preferred.strip().lower()
        target_normalized = _normalize_sheet_label(preferred)
        for name in names:
            if name.strip().lower() == target:
                return name
        if target_normalized:
            for name in names:
                if _normalize_sheet_label(name) == target_normalized:
                    return name
    for name in names:
        low = name.strip().lower()
        if "stringing" in low and "compiled" in low:
            return name
    return None


def list_excel_sheet_names(xlsx_path: Path | str) -> tuple[list[str], str | None]:
    try:
        with pd.ExcelFile(xlsx_path) as xl:
            return list(xl.sheet_names), None
    except Exception as exc:
        try:
            with zipfile.ZipFile(xlsx_path) as zf:
                workbook_xml = zf.read("xl/workbook.xml")
            root = ET.fromstring(workbook_xml)
            ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            names = [node.attrib.get("name", "") for node in root.findall(".//x:sheets/x:sheet", ns)]
            names = [name for name in names if str(name).strip()]
            if names:
                return names, None
        except Exception:
            pass
        return [], f"{type(exc).__name__}: {exc}"


def _clean_header_label(label: object) -> str:
    if label is None:
        return ""
    try:
        if pd.isna(label):
            return ""
    except Exception:
        pass
    text = str(label).strip()
    return "" if text.lower() == "nan" else text


def _make_unique_headers(labels: List[str]) -> List[str]:
    unique: List[str] = []
    seen: Dict[str, int] = {}
    for idx, label in enumerate(labels, start=1):
        base = label if label else f"unnamed_col_{idx}"
        key = normalize_space_only(base) or base.lower()
        count = seen.get(key, 0) + 1
        seen[key] = count
        unique.append(base if count == 1 else f"{base}__{count}")
    return unique


def materialize_stringing_data(
    df_raw: pd.DataFrame,
    *,
    min_columns: int | None = None,
) -> tuple[pd.DataFrame, int | None, list[str]]:
    if df_raw is None or df_raw.empty:
        return pd.DataFrame(), None, []
    header_row, header_labels = find_stringing_header_row(df_raw)
    keep_min = int(min_columns or 0)
    if header_row is None or header_labels is None:
        labels = list(df_raw.iloc[0, :].values)
        data = df_raw.iloc[1:].copy()
        if keep_min > 0 and data.shape[1] < keep_min:
            missing = keep_min - data.shape[1]
            for _ in range(missing):
                data[data.shape[1]] = pd.NA
                labels.append("")
        clean_labels = _make_unique_headers([_clean_header_label(label) for label in labels])
        data.columns = clean_labels
        return data.reset_index(drop=True), 0, clean_labels

    labels = [_clean_header_label(label) for label in list(header_labels)]
    data = df_raw.iloc[header_row + 1 :].copy()
    labels_series = pd.Series(labels)
    last_non_empty = labels_series.replace("", pd.NA).last_valid_index()
    if last_non_empty is not None:
        target_cols = last_non_empty + 1
        if keep_min > 0:
            target_cols = max(target_cols, keep_min)
        if data.shape[1] < target_cols:
            missing = target_cols - data.shape[1]
            for _ in range(missing):
                data[data.shape[1]] = pd.NA
        if labels_series.shape[0] < target_cols:
            labels_series = labels_series.reindex(range(target_cols), fill_value="")
        data = data.iloc[:, :target_cols]
        labels_series = labels_series.iloc[:target_cols]
    clean_labels = _make_unique_headers([_clean_header_label(label) for label in labels_series.values])
    data.columns = clean_labels
    return data.reset_index(drop=True), int(header_row), clean_labels


@dataclass(frozen=True)
class StringingSheetLoadResult:
    frame: pd.DataFrame | None
    resolved_sheet: str | None
    fallback_note: str
    header_row: int | None
    header_labels: list[str]


def load_stringing_sheet_frame(
    source: Path,
    *,
    configured_sheet_name: str = "",
    preferred_sheet_name: str = "",
    min_columns: int | None = None,
) -> StringingSheetLoadResult:
    selector = (
        (lambda names: find_stringing_sheet_name_from_list(list(names), None, [configured_sheet_name]))
        if configured_sheet_name
        else (lambda names: find_stringing_sheet_name_from_list(list(names), preferred_sheet_name, None))
    )

    try:
        with pd.ExcelFile(source) as xl:
            found = selector(list(xl.sheet_names))
            if not found:
                raise ValueError("NO_TARGET_SHEET")
            df_raw = xl.parse(sheet_name=found, header=None)
            frame, header_row, header_labels = materialize_stringing_data(df_raw, min_columns=min_columns)
            fallback_note = ""
            if configured_sheet_name and normalize_sheet_key(found) != normalize_sheet_key(configured_sheet_name):
                fallback_note = f"Configured sheet '{configured_sheet_name}' not found; used '{found}'."
            return StringingSheetLoadResult(
                frame=frame,
                resolved_sheet=found,
                fallback_note=fallback_note,
                header_row=header_row,
                header_labels=header_labels,
            )
    except Exception:
        df_raw, found, fallback_note = load_sheet_with_csv_fallback(
            source,
            selector,
            read_excel_kwargs={"header": None},
            read_csv_kwargs={"header": None},
        )
        if found is None or df_raw is None or df_raw.empty:
            raise ValueError("NO_TARGET_SHEET")
        frame, header_row, header_labels = materialize_stringing_data(df_raw, min_columns=min_columns)
        note = fallback_note or ""
        if configured_sheet_name and normalize_sheet_key(found) != normalize_sheet_key(configured_sheet_name):
            sheet_note = f"Configured sheet '{configured_sheet_name}' not found; used '{found}'."
            note = f"{note}; {sheet_note}" if note else sheet_note
        return StringingSheetLoadResult(
            frame=frame,
            resolved_sheet=found,
            fallback_note=note,
            header_row=header_row,
            header_labels=header_labels,
        )
