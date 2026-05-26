"""Config-driven DPR progress/status ingestion."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Optional
import re
import warnings

import pandas as pd
from openpyxl import load_workbook

from erection_compiled_to_daily_new import load_sheet_with_csv_fallback
from . import stringing_ingest as ingest
from .stringing import add_length_units, normalize_stringing_columns
from .project_identity import (
    build_project_display,
    build_project_scope_key,
    normalize_line_name,
    parse_project_identity_from_filename,
    parse_sheet_line_entries,
)
from .completed_projects import is_completed_project


DEFAULT_ACTIVITY_ALLOWLIST = (
    "route alignment",
    "detail survey",
    "detailed survey",
    "check survey",
    "soil investigation",
    "excavation",
    "foundation",
    "earthing",
    "tower erection",
    "erection",
    "tower tightening",
    "tack welding",
    "tackwelding",
    "nh crossing",
    "power line crossing",
    "stringing",
    "paying out",
    "final sag",
    "opgw",
)

DEFAULT_REQUIRED_HEADER_TOKENS = ("activity", "progress", "cumulative")

RAWDATA_COLUMNS = [
    "project_code",
    "project_display",
    "project_scope_key",
    "line_name",
    "line_name_source",
    "section_label",
    "source_file",
    "source_sheet",
    "configured_sheet",
    "template_sheet",
    "header_row_number",
    "source_row_number",
    "activity_raw",
    "activity_norm",
    "quantity_loa",
    "quantity_estimated_or_total",
    "quantity_primary",
    "cumulative_last_month",
    "plan_for_month",
    "progress_for_month",
    "today_progress",
    "cumulative_progress",
    "balance_progress",
    "gangs_working",
    "remarks",
]

DIAGNOSTICS_COLUMNS = [
    "Workbook",
    "Project",
    "Sheet",
    "ConfiguredSheet",
    "LineName",
    "LineNameSource",
    "TemplateSheet",
    "TemplateApplied",
    "TemplateChanges",
    "FallbackNote",
    "SectionsDetected",
    "HeadersDetected",
    "Rows",
    "Status",
    "Reason",
]

ISSUES_COLUMNS = [
    "Workbook",
    "Project",
    "Sheet",
    "ConfiguredSheet",
    "LineName",
    "LineNameSource",
    "Issue",
    "Reason",
]

COVERAGE_COLUMNS = [
    "project_code",
    "project_display",
    "status",
    "reason_code",
    "reason",
    "workbook",
    "configured_sheet",
    "resolved_sheet",
    "rows",
    "available_sheets",
]


def _normalize_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"[.,;:()\\[\\]{}]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _split_tokens(value: object) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    tokens = [part.strip() for part in re.split(r"[;,|]+", text)]
    return [token for token in tokens if token]


def _coerce_numeric(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        cleaned = cleaned.replace(",", "")
        value = cleaned
    try:
        num = pd.to_numeric([value], errors="coerce")[0]
    except Exception:
        return None
    if pd.isna(num):
        return None
    return float(num)


def _as_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def _make_unique_headers(labels: list[str]) -> list[str]:
    unique: list[str] = []
    seen: dict[str, int] = {}
    for idx, label in enumerate(labels, start=1):
        base = label.strip() if label else f"unnamed_col_{idx}"
        key = ingest.normalize_space_only(base) or base.lower()
        count = seen.get(key, 0) + 1
        seen[key] = count
        unique.append(base if count == 1 else f"{base}__{count}")
    return unique


def _extract_numeric_tokens(value: object) -> set[str]:
    text = str(value or "")
    return {token for token in re.findall(r"\d{2,4}", text)}


def _normalize_activity(raw: object) -> str:
    text = _normalize_text(raw)
    if not text:
        return ""
    if "route alignment" in text:
        return "route_alignment"
    if "detailed survey" in text:
        return "detailed_survey"
    if "check survey" in text:
        return "check_survey"
    if "soil investigation" in text:
        return "soil_investigation"
    if "excavation" in text:
        return "excavation"
    if "foundation" in text:
        return "foundation"
    if "earthing" in text:
        return "earthing"
    if "tower erection" in text or text == "erection":
        return "tower_erection"
    if "tack welding" in text or "tackwelding" in text:
        return "tack_welding"
    if "paying out" in text:
        return "paying_out"
    if "final sag" in text:
        return "final_sag"
    if "opgw" in text:
        return "opgw_stringing"
    if "stringing" in text:
        return "stringing"
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _header_tokens_joined(labels: list[str]) -> str:
    return " ".join(_normalize_text(label) for label in labels if _normalize_text(label))


def _combine_header_labels(df: pd.DataFrame, start_row: int, header_rows: int, width: int) -> list[str]:
    labels: list[str] = []
    for col_idx in range(width):
        parts: list[str] = []
        for offset in range(header_rows):
            row_idx = start_row + offset
            if row_idx >= len(df.index):
                continue
            value = _as_text(df.iat[row_idx, col_idx] if col_idx < len(df.columns) else "")
            if value and value not in parts:
                parts.append(value)
        labels.append(" ".join(parts).strip())
    return _make_unique_headers(labels)


def _row_text(df: pd.DataFrame, row_idx: int, max_cols: int = 30) -> str:
    if row_idx < 0 or row_idx >= len(df.index):
        return ""
    values: list[str] = []
    width = min(len(df.columns), max_cols)
    for col_idx in range(width):
        values.append(_as_text(df.iat[row_idx, col_idx]))
    return " | ".join(values)


def _row_is_blank(values: list[object]) -> bool:
    for value in values:
        if _as_text(value):
            return False
    return True


def _parse_header_rows_guardrail(value: object) -> list[int]:
    tokens = _split_tokens(value)
    if not tokens:
        token_text = str(value or "").strip()
        if token_text:
            tokens = [token_text]
    parsed: list[int] = []
    for token in tokens:
        try:
            intval = int(str(token).strip())
        except Exception:
            continue
        if intval > 0:
            parsed.append(intval)
    deduped = sorted(set(parsed))
    return deduped if deduped else [1, 2, 3]


def _find_anchor_row(
    df: pd.DataFrame,
    start_row: int,
    end_row: int,
    *,
    anchor_text: str = "",
    anchor_regex: str = "",
) -> int:
    if anchor_regex:
        try:
            pattern = re.compile(anchor_regex, flags=re.IGNORECASE)
        except re.error:
            pattern = None
        if pattern is not None:
            for row_idx in range(start_row, end_row + 1):
                text = _row_text(df, row_idx, max_cols=25)
                if pattern.search(text):
                    return row_idx

    anchor = _normalize_text(anchor_text)
    if anchor:
        for row_idx in range(start_row, end_row + 1):
            text = _normalize_text(_row_text(df, row_idx, max_cols=25))
            if anchor in text:
                return row_idx
    return start_row


def _detect_header(
    df: pd.DataFrame,
    *,
    start_row: int,
    end_row: int,
    required_tokens: list[str],
    header_rows_options: list[int],
    max_scan_rows: int = 45,
) -> tuple[int | None, int | None, list[str]]:
    if df.empty:
        return None, None, []
    width = min(len(df.columns), 60)
    scan_end = min(end_row, start_row + max_scan_rows)
    for row_idx in range(start_row, scan_end + 1):
        for span in header_rows_options:
            if row_idx + span - 1 > end_row:
                continue
            labels = _combine_header_labels(df, row_idx, span, width)
            joined = _header_tokens_joined(labels)
            if not joined:
                continue
            if all(token in joined for token in required_tokens):
                return row_idx, span, labels
    return None, None, []


def _extract_section_ranges(
    df: pd.DataFrame,
    *,
    start_row: int,
    end_row: int,
    section_contains: str = "",
    section_regex: str = "",
    section_label_regex: str = "",
) -> list[tuple[str, int, int]]:
    contains = _normalize_text(section_contains)
    regex_obj = None
    label_regex_obj = None
    if section_regex:
        try:
            regex_obj = re.compile(section_regex, flags=re.IGNORECASE)
        except re.error:
            regex_obj = None
    if section_label_regex:
        try:
            label_regex_obj = re.compile(section_label_regex, flags=re.IGNORECASE)
        except re.error:
            label_regex_obj = None

    if not contains and regex_obj is None:
        return [("", start_row, end_row)]

    hits: list[tuple[str, int]] = []
    for row_idx in range(start_row, end_row + 1):
        text = _row_text(df, row_idx, max_cols=30)
        text_norm = _normalize_text(text)
        matched = False
        if contains and contains in text_norm:
            matched = True
        if not matched and regex_obj is not None and regex_obj.search(text):
            matched = True
        if not matched:
            continue
        label = text.strip()
        if label_regex_obj is not None:
            label_match = label_regex_obj.search(text)
            if label_match:
                label = label_match.group(1).strip() if label_match.groups() else label_match.group(0).strip()
        label = re.sub(r"\s+", " ", label).strip()
        hits.append((label, row_idx))

    if not hits:
        return [("", start_row, end_row)]

    ranges: list[tuple[str, int, int]] = []
    for idx, (label, section_row) in enumerate(hits):
        section_end = hits[idx + 1][1] - 1 if idx + 1 < len(hits) else end_row
        section_start = max(start_row, section_row)
        if section_start <= section_end:
            ranges.append((label, section_start, section_end))
    return ranges if ranges else [("", start_row, end_row)]


def _apply_template_column_mapping(df: pd.DataFrame, template_map: dict[int, str]) -> tuple[pd.DataFrame, list[str]]:
    if df is None or df.empty or not template_map:
        return df, []
    remapped = df.copy()
    columns = list(remapped.columns)
    changes: list[str] = []
    for idx, mapped_name in sorted(template_map.items()):
        if idx >= len(columns):
            continue
        current = str(columns[idx]).strip()
        target = str(mapped_name).strip()
        if not target:
            continue
        columns[idx] = target
        if ingest.normalize_space_only(current) != ingest.normalize_space_only(target):
            changes.append(f"C{idx + 1}:{current}->{target}")
    remapped.columns = columns
    return remapped, changes


def _resolve_activity_column(df: pd.DataFrame) -> str:
    for column in df.columns:
        key = _normalize_text(column)
        if "activity" in key:
            return str(column)
    return str(df.columns[0]) if len(df.columns) else "activity_raw"


def _first_non_null(*values: object) -> float | None:
    for value in values:
        num = _coerce_numeric(value)
        if num is not None:
            return num
    return None


def _build_status_rows(
    df: pd.DataFrame,
    *,
    project_code: str,
    project_display: str,
    project_scope_key: str,
    line_name: str,
    line_name_source: str,
    section_label: str,
    source_file: str,
    source_sheet: str,
    configured_sheet: str,
    template_sheet: str,
    header_row_number: int,
    activity_allowlist: list[str],
    activity_exclude: list[str],
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=RAWDATA_COLUMNS)

    rows: list[dict[str, object]] = []
    activity_col = "activity_raw" if "activity_raw" in df.columns else _resolve_activity_column(df)
    allowlist_norm = [_normalize_text(token) for token in activity_allowlist if _normalize_text(token)]
    exclude_norm = [_normalize_text(token) for token in activity_exclude if _normalize_text(token)]

    for _, row in df.iterrows():
        raw_activity = _as_text(row.get(activity_col, ""))
        activity_norm_text = _normalize_text(raw_activity)
        if not raw_activity:
            continue
        if allowlist_norm and not any(token in activity_norm_text for token in allowlist_norm):
            continue
        if exclude_norm and any(token in activity_norm_text for token in exclude_norm):
            continue

        quantity_loa = _coerce_numeric(row.get("quantity_loa"))
        quantity_estimated = _coerce_numeric(row.get("quantity_estimated_or_total"))
        cumulative_last_month = _first_non_null(
            row.get("cumulative_last_month"),
            row.get("cumulative_progress_last_month"),
        )
        plan_for_month = _coerce_numeric(row.get("plan_for_month"))
        progress_for_month = _coerce_numeric(row.get("progress_for_month"))
        today_progress = _coerce_numeric(row.get("today_progress"))
        cumulative_progress = _coerce_numeric(row.get("cumulative_progress"))
        balance_progress = _coerce_numeric(row.get("balance_progress"))

        rows.append(
            {
                "project_code": project_code,
                "project_display": project_display,
                "project_scope_key": project_scope_key,
                "line_name": line_name,
                "line_name_source": line_name_source,
                "section_label": section_label,
                "source_file": source_file,
                "source_sheet": source_sheet,
                "configured_sheet": configured_sheet,
                "template_sheet": template_sheet,
                "header_row_number": int(header_row_number),
                "source_row_number": int(row.get("__source_row_number", 0) or 0),
                "activity_raw": raw_activity,
                "activity_norm": _normalize_activity(raw_activity),
                "quantity_loa": quantity_loa,
                "quantity_estimated_or_total": quantity_estimated,
                "quantity_primary": quantity_estimated if quantity_estimated is not None else quantity_loa,
                "cumulative_last_month": cumulative_last_month,
                "plan_for_month": plan_for_month,
                "progress_for_month": progress_for_month,
                "today_progress": today_progress,
                "cumulative_progress": cumulative_progress,
                "balance_progress": balance_progress,
                "gangs_working": _as_text(row.get("gangs_working")),
                "remarks": _as_text(row.get("remarks")),
            }
        )
    return pd.DataFrame(rows, columns=RAWDATA_COLUMNS)


def _sheet_tokens(value: object) -> set[str]:
    text = str(value or "").strip().lower()
    if not text:
        return set()
    return {token for token in re.findall(r"[a-z0-9]+", text) if token}


def _sheet_match_score(configured_sheet: str, workbook_sheet: str) -> tuple[int, int, int]:
    configured = str(configured_sheet or "").strip()
    workbook = str(workbook_sheet or "").strip()
    if not configured or not workbook:
        return (0, 0, 0)
    if ingest.normalize_space_only(configured) == ingest.normalize_space_only(workbook):
        return (1, 0, 0)
    if ingest.normalize_sheet_key(configured) == ingest.normalize_sheet_key(workbook):
        return (0, 1, 0)
    c_tokens = _sheet_tokens(configured)
    w_tokens = _sheet_tokens(workbook)
    overlap = len(c_tokens.intersection(w_tokens)) if c_tokens and w_tokens else 0
    return (0, 0, overlap)


def _pick_best_workbook_for_sheet(
    workbooks: list[Path],
    workbook_sheets: dict[str, list[str]],
    configured_sheet: str,
) -> tuple[Path | None, str]:
    scored: list[tuple[tuple[int, int, int, float], Path, str]] = []
    for workbook in workbooks:
        names = workbook_sheets.get(str(workbook.resolve()), [])
        best_score = (0, 0, 0)
        best_name = ""
        for name in names:
            score = _sheet_match_score(configured_sheet, name)
            if score > best_score:
                best_score = score
                best_name = name
        mtime = workbook.stat().st_mtime if workbook.exists() else 0.0
        scored.append(((best_score[0], best_score[1], best_score[2], mtime), workbook, best_name))

    if not scored:
        return None, ""
    scored.sort(key=lambda item: item[0], reverse=True)
    top = scored[0]
    if top[0][:3] == (0, 0, 0):
        return None, ""
    return top[1], top[2]


def _resolve_named_template_sheet(wb, expected_name: str) -> Optional[str]:
    expected_key = ingest.normalize_space_only(expected_name)
    for name in wb.sheetnames:
        if ingest.normalize_space_only(name) == expected_key:
            return name
    return None


def _resolve_project_template_sheets(wb, project_name: object, discipline: str) -> list[str]:
    project_text = str(project_name or "").strip()
    if not project_text:
        return []

    resolved: list[str] = []
    seen: set[str] = set()
    for expected in (
        f"{project_text} {discipline}",
        f"{project_text} {discipline} Template Check",
    ):
        hit = _resolve_named_template_sheet(wb, expected)
        if hit and hit not in seen:
            seen.add(hit)
            resolved.append(hit)

    project_key = ingest.normalize_space_only(project_text)
    discipline_key = ingest.normalize_space_only(discipline)
    for name in wb.sheetnames:
        key = ingest.normalize_space_only(name)
        if key and key.startswith(project_key) and key.endswith(discipline_key) and name not in seen:
            seen.add(name)
            resolved.append(name)
    return resolved


def _extract_template_column_map(ws) -> dict[int, str]:
    to_map_row = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if ingest.normalize_space_only(cell) == "to map":
                to_map_row = row_idx
                break
        if to_map_row is not None:
            break
    if to_map_row is None:
        return {}

    labels_row = to_map_row + 1
    row_values = next(ws.iter_rows(min_row=labels_row, max_row=labels_row, values_only=True), ())
    mapping: dict[int, str] = {}
    for col_idx, value in enumerate(row_values):
        label = str(value).strip() if value is not None else ""
        if label:
            mapping[col_idx] = label
    return mapping


def _extract_guardrails(ws) -> dict[str, str]:
    guardrails: dict[str, str] = {}
    in_guardrails = False
    for row in ws.iter_rows(values_only=True):
        first_raw = row[0] if len(row) > 0 else None
        second_raw = row[1] if len(row) > 1 else None
        first_norm = ingest.normalize_space_only(first_raw)
        if first_norm == "guardrails":
            in_guardrails = True
            continue
        if first_norm == "to map":
            break
        if not in_guardrails:
            continue
        key = _normalize_key(first_raw)
        value = _as_text(second_raw)
        if key and value:
            guardrails[key] = value
    return guardrails


def load_status_sheet_config(raw_root: Path, *, repo_root: Path | None = None) -> dict[str, list[dict[str, str]]]:
    config_path = ingest.resolve_dpr_config_path(raw_root, repo_root=repo_root)
    if config_path is None:
        return {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception:
        return {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}
        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}
        headers = [ingest.normalize_space_only(v) for v in header_row]
        if "project code" not in headers or "status sheet names" not in headers:
            return {}
        project_idx = headers.index("project code")
        status_idx = headers.index("status sheet names")
        line_idx = headers.index("status line names") if "status line names" in headers else None

        mapping: dict[str, list[dict[str, str]]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project_val = row[project_idx] if project_idx < len(row) else None
            if project_val in (None, ""):
                continue
            project_key = ingest.normalize_project_code_key(project_val)
            raw_status = row[status_idx] if status_idx < len(row) else None
            if raw_status in (None, ""):
                mapping[project_key] = []
                continue

            raw_line_names = row[line_idx] if line_idx is not None and line_idx < len(row) else None
            entries = parse_sheet_line_entries(
                raw_status,
                raw_line_names,
                "status",
                infer_from_sheet_name=False,
            )
            deduped_entries: list[dict[str, str]] = []
            seen_sheet_keys: set[str] = set()
            for entry in entries:
                key = ingest.normalize_space_only(entry.get("sheet_name"))
                if not key or key in seen_sheet_keys:
                    continue
                seen_sheet_keys.add(key)
                deduped_entries.append(entry)
            mapping[project_key] = deduped_entries
        return mapping
    finally:
        wb.close()


def load_status_template_mapping_catalog(
    raw_root: Path,
    *,
    repo_root: Path | None = None,
    include_unchecked: bool = False,
) -> tuple[dict[str, list[dict[str, object]]], dict[str, str]]:
    config_path = ingest.resolve_dpr_config_path(raw_root, repo_root=repo_root)
    if config_path is None:
        return {}, {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception:
        return {}, {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}, {}
        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}, {}
        headers = [ingest.normalize_space_only(value) for value in header_row]
        project_idx = headers.index("project code") if "project code" in headers else None
        check_idx = None
        for candidate in ("status template check", "status"):
            if candidate in headers:
                check_idx = headers.index(candidate)
                break
        if project_idx is None or (check_idx is None and not include_unchecked):
            return {}, {}

        catalog: dict[str, list[dict[str, object]]] = {}
        errors: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project_val = row[project_idx] if project_idx < len(row) else None
            if project_val in (None, ""):
                continue
            check_val = row[check_idx] if check_idx is not None and check_idx < len(row) else None
            check_enabled = ingest.normalize_space_only(check_val) == "yes"
            if not check_enabled and not include_unchecked:
                continue

            project_key = ingest.normalize_project_code_key(project_val)
            template_sheets = _resolve_project_template_sheets(wb, project_val, "Status")
            if not template_sheets:
                if check_enabled:
                    errors[project_key] = (
                        f"Status Template Check is Yes but no mapping tab matching project '{str(project_val).strip()}' was found."
                    )
                continue

            options: list[dict[str, object]] = []
            for sheet_name in template_sheets:
                ws_template = wb[sheet_name]
                col_map = _extract_template_column_map(ws_template)
                guardrails = _extract_guardrails(ws_template)
                if not col_map and not guardrails:
                    continue
                options.append(
                    {
                        "column_map": col_map,
                        "template_sheet": sheet_name,
                        "guardrails": guardrails,
                        "numeric_tokens": _extract_numeric_tokens(sheet_name),
                    }
                )
            if not options:
                if check_enabled:
                    errors[project_key] = (
                        f"Status template tab(s) for project '{str(project_val).strip()}' have no usable Guardrails/To Map rows."
                    )
                continue
            catalog[project_key] = options
        return catalog, errors
    finally:
        wb.close()


def select_status_template_for_sheet(
    template_options: list[dict[str, object]] | None,
    *,
    configured_sheet_name: str = "",
    resolved_sheet_name: str = "",
    line_name: str = "",
) -> dict[str, object] | None:
    if not template_options:
        return None
    hints = [configured_sheet_name or "", resolved_sheet_name or "", line_name or ""]
    hint_keys = [ingest.normalize_space_only(value) for value in hints if ingest.normalize_space_only(value)]
    hint_numbers: set[str] = set()
    for value in hints:
        hint_numbers.update(_extract_numeric_tokens(value))

    best: dict[str, object] | None = None
    best_score = float("-inf")
    for idx, option in enumerate(template_options):
        sheet_name = str(option.get("template_sheet", "")).strip()
        sheet_key = ingest.normalize_space_only(sheet_name)
        sheet_numbers = set(option.get("numeric_tokens", set()) or set())
        col_map = option.get("column_map", {}) or {}
        score = 0.0
        for hint in hint_keys:
            if sheet_key == hint:
                score += 1000.0
            elif hint and (hint in sheet_key or sheet_key in hint):
                score += 200.0
        if hint_numbers:
            score += float(len(sheet_numbers.intersection(hint_numbers))) * 120.0
        score += float(len(col_map))
        score -= idx * 1e-4
        if score > best_score:
            best_score = score
            best = option
    return best


def _build_exact_sheet_selector(sheet_name: str):
    expected_space = ingest.normalize_space_only(sheet_name)
    expected_compact = ingest.normalize_sheet_key(sheet_name)

    def _selector(sheet_names: list[str]) -> str | None:
        by_space: dict[str, str] = {}
        by_compact: dict[str, str] = {}
        for existing in sheet_names:
            space_key = ingest.normalize_space_only(existing)
            compact_key = ingest.normalize_sheet_key(existing)
            if space_key and space_key not in by_space:
                by_space[space_key] = existing
            if compact_key and compact_key not in by_compact:
                by_compact[compact_key] = existing
        hit = by_space.get(expected_space)
        if hit:
            return hit
        hit = by_compact.get(expected_compact)
        if hit:
            return hit
        for existing in sheet_names:
            compact = ingest.normalize_sheet_key(existing)
            if expected_compact and (expected_compact in compact or compact in expected_compact):
                return existing
        return None

    return _selector


@dataclass(frozen=True)
class StatusParseResult:
    data: pd.DataFrame
    parse_status: str
    parse_reason: str
    sections_detected: int
    headers_detected: int
    rows_emitted: int
    template_changes: list[str]


def _parse_status_sheet_dataframe(
    df_raw: pd.DataFrame,
    *,
    guardrails: dict[str, str],
    template_map: dict[int, str],
    project_code: str,
    project_display: str,
    project_scope_key: str,
    line_name: str,
    line_name_source: str,
    source_file: str,
    source_sheet: str,
    configured_sheet: str,
    template_sheet: str,
) -> StatusParseResult:
    if df_raw is None or df_raw.empty:
        return StatusParseResult(
            data=pd.DataFrame(columns=RAWDATA_COLUMNS),
            parse_status="EMPTY_SHEET",
            parse_reason="Sheet is empty.",
            sections_detected=0,
            headers_detected=0,
            rows_emitted=0,
            template_changes=[],
        )

    activity_allowlist = _split_tokens(guardrails.get("activity_allowlist", "")) or list(DEFAULT_ACTIVITY_ALLOWLIST)
    activity_exclude = _split_tokens(guardrails.get("activity_exclude", ""))
    required_tokens = [_normalize_text(token) for token in _split_tokens(guardrails.get("required_tokens", ""))]
    required_tokens = [token for token in required_tokens if token] or list(DEFAULT_REQUIRED_HEADER_TOKENS)
    stop_tokens = [_normalize_text(token) for token in _split_tokens(guardrails.get("stop_tokens", ""))]
    header_rows_options = _parse_header_rows_guardrail(guardrails.get("header_rows", ""))
    section_ranges = _extract_section_ranges(
        df_raw,
        start_row=0,
        end_row=len(df_raw.index) - 1,
        section_contains=guardrails.get("section_start_contains", ""),
        section_regex=guardrails.get("section_split_regex", ""),
        section_label_regex=guardrails.get("section_label_regex", ""),
    )

    all_rows: list[pd.DataFrame] = []
    headers_detected = 0
    sections_detected = 0
    template_changes_all: list[str] = []
    for section_label, section_start, section_end in section_ranges:
        anchor_row = _find_anchor_row(
            df_raw,
            section_start,
            section_end,
            anchor_text=guardrails.get("block_anchor", ""),
            anchor_regex=guardrails.get("block_anchor_regex", ""),
        )
        header_row, header_span, header_labels = _detect_header(
            df_raw,
            start_row=anchor_row,
            end_row=section_end,
            required_tokens=required_tokens,
            header_rows_options=header_rows_options,
        )
        if header_row is None or header_span is None:
            continue

        headers_detected += 1
        sections_detected += 1
        width = len(header_labels)
        data_rows: list[tuple[int, list[object]]] = []
        blank_run = 0
        started = False
        for row_idx in range(header_row + header_span, section_end + 1):
            row_values = [df_raw.iat[row_idx, col_idx] if col_idx < len(df_raw.columns) else None for col_idx in range(width)]
            row_text_norm = _normalize_text(" ".join(_as_text(v) for v in row_values[: min(25, width)]))
            if stop_tokens and any(token in row_text_norm for token in stop_tokens):
                if started:
                    break
                continue
            if _row_is_blank(row_values):
                if started:
                    blank_run += 1
                    if blank_run >= 2:
                        break
                continue
            started = True
            blank_run = 0
            data_rows.append((row_idx + 1, row_values))

        if not data_rows:
            continue

        block_df = pd.DataFrame([values for _, values in data_rows], columns=header_labels)
        block_df["__source_row_number"] = [row_number for row_number, _ in data_rows]
        block_df, template_changes = _apply_template_column_mapping(block_df, template_map)
        template_changes_all.extend(template_changes)
        prepared = _build_status_rows(
            block_df,
            project_code=project_code,
            project_display=project_display,
            project_scope_key=project_scope_key,
            line_name=line_name,
            line_name_source=line_name_source,
            section_label=section_label,
            source_file=source_file,
            source_sheet=source_sheet,
            configured_sheet=configured_sheet,
            template_sheet=template_sheet,
            header_row_number=header_row + 1,
            activity_allowlist=activity_allowlist,
            activity_exclude=activity_exclude,
        )
        if not prepared.empty:
            all_rows.append(prepared)

    if not all_rows:
        reason = "No matching rows after guardrails/activity filters."
        status = "NO_MATCHED_ROWS" if headers_detected > 0 else "HEADER_NOT_FOUND"
        return StatusParseResult(
            data=pd.DataFrame(columns=RAWDATA_COLUMNS),
            parse_status=status,
            parse_reason=reason,
            sections_detected=sections_detected,
            headers_detected=headers_detected,
            rows_emitted=0,
            template_changes=template_changes_all,
        )

    combined = pd.concat(all_rows, ignore_index=True)
    return StatusParseResult(
        data=combined.reindex(columns=RAWDATA_COLUMNS),
        parse_status="OK",
        parse_reason="",
        sections_detected=sections_detected,
        headers_detected=headers_detected,
        rows_emitted=int(len(combined.index)),
        template_changes=template_changes_all,
    )


def _status_candidates(input_dir: Optional[Path], files: Optional[list[Path]]) -> list[Path]:
    if files:
        return [
            path
            for path in files
            if path.suffix.lower() in (".xlsx", ".xlsm", ".xls")
            and path.exists()
            and not path.name.startswith("~$")
        ]
    if input_dir and input_dir.exists():
        return sorted(
            [
                path
                for path in input_dir.rglob("*.xls*")
                if path.is_file() and not path.name.startswith("~$")
            ]
        )
    return []


def compile_progress_status_to_workbook(
    input_dir: Optional[Path],
    files: Optional[list[Path]],
    output_path: Path,
    *,
    repo_root: Path | None = None,
) -> Path | None:
    candidates = _status_candidates(input_dir, files)
    if not candidates:
        print("[pipeline] ProgressStatus: no candidate files found; skipping.")
        return None

    if input_dir is not None:
        raw_root = input_dir
    elif files:
        raw_root = files[0].parent
    else:
        raw_root = Path(".")

    status_sheet_config = load_status_sheet_config(raw_root, repo_root=repo_root)
    status_template_catalog, status_template_errors = load_status_template_mapping_catalog(
        raw_root,
        repo_root=repo_root,
    )
    status_template_all_catalog, _ = load_status_template_mapping_catalog(
        raw_root,
        repo_root=repo_root,
        include_unchecked=True,
    )

    has_status_config = bool(status_sheet_config)
    workbook_sheet_cache: dict[str, list[str]] = {}
    workbooks_by_project: dict[str, list[Path]] = {}
    for workbook in candidates:
        identity = parse_project_identity_from_filename(workbook.name)
        project_code = str(identity.get("project_code", "")).strip() or workbook.stem
        project_key = ingest.normalize_project_code_key(project_code)
        workbooks_by_project.setdefault(project_key, []).append(workbook)
        sheet_names, _ = ingest.list_excel_sheet_names(workbook)
        workbook_sheet_cache[str(workbook.resolve())] = list(sheet_names or [])

    raw_frames: list[pd.DataFrame] = []
    diagnostics_rows: list[dict[str, object]] = []
    issue_rows: list[dict[str, object]] = []
    coverage_rows: list[dict[str, object]] = []

    skipped_not_in_config = 0
    skipped_blank_config = 0

    for project_key, project_workbooks in sorted(workbooks_by_project.items()):
        configured_entries = status_sheet_config.get(project_key)
        if has_status_config and configured_entries is None:
            skipped_not_in_config += len(project_workbooks)
            continue
        if configured_entries is not None and not configured_entries:
            skipped_blank_config += len(project_workbooks)
            identity = parse_project_identity_from_filename(project_workbooks[0].name)
            project_code = str(identity.get("project_code", "")).strip() or project_workbooks[0].stem
            coverage_rows.append(
                {
                    "project_code": project_code,
                    "project_display": project_code,
                    "status": "SKIPPED_NO_STATUS_CONFIG",
                    "reason_code": "SKIPPED_NO_STATUS_CONFIG",
                    "reason": "Project has blank Status Sheet Names in DPR_Config.",
                    "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                    "configured_sheet": "",
                    "resolved_sheet": "",
                    "rows": 0,
                    "available_sheets": "",
                }
            )
            continue

        sheet_requests = configured_entries if configured_entries else [{"sheet_name": "", "line_name": "", "line_name_source": ""}]
        template_error = status_template_errors.get(project_key, "")
        for request in sheet_requests:
            configured_sheet = str(request.get("sheet_name", "")).strip()
            configured_line_name = normalize_line_name(request.get("line_name", ""))
            configured_line_source = str(request.get("line_name_source", "")).strip()
            if configured_sheet:
                selected_workbook, resolved_sheet_guess = _pick_best_workbook_for_sheet(
                    project_workbooks,
                    workbook_sheet_cache,
                    configured_sheet,
                )
            else:
                selected_workbook = project_workbooks[0] if project_workbooks else None
                resolved_sheet_guess = ""

            if selected_workbook is None:
                identity = parse_project_identity_from_filename(project_workbooks[0].name)
                project_code = str(identity.get("project_code", "")).strip() or project_workbooks[0].stem
                project_display = build_project_display(project_code, configured_line_name, project_code) or project_code
                coverage_rows.append(
                    {
                        "project_code": project_code,
                        "project_display": project_display,
                        "status": "NO_TARGET_SHEET",
                        "reason_code": "NO_TARGET_SHEET",
                        "reason": "Configured status sheet not found in any workbook for project.",
                        "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                        "configured_sheet": configured_sheet,
                        "resolved_sheet": "",
                        "rows": 0,
                        "available_sheets": "",
                    }
                )
                issue_rows.append(
                    {
                        "Workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                        "Project": project_code,
                        "Sheet": "",
                        "ConfiguredSheet": configured_sheet,
                        "LineName": configured_line_name,
                        "LineNameSource": configured_line_source,
                        "Issue": "NO_TARGET_SHEET",
                        "Reason": "Configured status sheet not found in project workbooks.",
                    }
                )
                continue

            selected_identity = parse_project_identity_from_filename(selected_workbook.name)
            project_code = str(selected_identity.get("project_code", "")).strip() or selected_workbook.stem
            line_name = configured_line_name or normalize_line_name(selected_identity.get("line_name", ""))
            line_source = configured_line_source or ("config" if configured_line_name else "filename")
            project_display = build_project_display(project_code, line_name, project_code) or project_code
            project_scope_key = build_project_scope_key(project_code, line_name, project_display)
            available_sheet_text = "; ".join(workbook_sheet_cache.get(str(selected_workbook.resolve()), []))

            if template_error:
                coverage_rows.append(
                    {
                        "project_code": project_code,
                        "project_display": project_display,
                        "status": "TEMPLATE_CONFIG_ERROR",
                        "reason_code": "TEMPLATE_CONFIG_ERROR",
                        "reason": template_error,
                        "workbook": selected_workbook.name,
                        "configured_sheet": configured_sheet,
                        "resolved_sheet": resolved_sheet_guess,
                        "rows": 0,
                        "available_sheets": available_sheet_text,
                    }
                )
                issue_rows.append(
                    {
                        "Workbook": selected_workbook.name,
                        "Project": project_code,
                        "Sheet": resolved_sheet_guess,
                        "ConfiguredSheet": configured_sheet,
                        "LineName": line_name,
                        "LineNameSource": line_source,
                        "Issue": "TEMPLATE_CONFIG_ERROR",
                        "Reason": template_error,
                    }
                )
                continue

            selector = _build_exact_sheet_selector(configured_sheet) if configured_sheet else (
                lambda names: names[0] if names else None
            )
            try:
                df_raw, resolved_sheet, fallback_note = load_sheet_with_csv_fallback(
                    selected_workbook,
                    selector,
                    read_excel_kwargs={"header": None},
                    read_csv_kwargs={"header": None},
                )
            except Exception as exc:
                issue_rows.append(
                    {
                        "Workbook": selected_workbook.name,
                        "Project": project_code,
                        "Sheet": resolved_sheet_guess,
                        "ConfiguredSheet": configured_sheet,
                        "LineName": line_name,
                        "LineNameSource": line_source,
                        "Issue": "READ_FAIL",
                        "Reason": str(exc),
                    }
                )
                coverage_rows.append(
                    {
                        "project_code": project_code,
                        "project_display": project_display,
                        "status": "READ_FAIL",
                        "reason_code": "READ_FAIL",
                        "reason": str(exc),
                        "workbook": selected_workbook.name,
                        "configured_sheet": configured_sheet,
                        "resolved_sheet": resolved_sheet_guess,
                        "rows": 0,
                        "available_sheets": available_sheet_text,
                    }
                )
                continue

            if df_raw is None or resolved_sheet is None:
                issue_rows.append(
                    {
                        "Workbook": selected_workbook.name,
                        "Project": project_code,
                        "Sheet": resolved_sheet_guess,
                        "ConfiguredSheet": configured_sheet,
                        "LineName": line_name,
                        "LineNameSource": line_source,
                        "Issue": "NO_TARGET_SHEET",
                        "Reason": "Configured status sheet not found in selected workbook.",
                    }
                )
                coverage_rows.append(
                    {
                        "project_code": project_code,
                        "project_display": project_display,
                        "status": "NO_TARGET_SHEET",
                        "reason_code": "NO_TARGET_SHEET",
                        "reason": "Configured status sheet not found in selected workbook.",
                        "workbook": selected_workbook.name,
                        "configured_sheet": configured_sheet,
                        "resolved_sheet": "",
                        "rows": 0,
                        "available_sheets": available_sheet_text,
                    }
                )
                continue

            selected_template = select_status_template_for_sheet(
                status_template_catalog.get(project_key, []),
                configured_sheet_name=configured_sheet,
                resolved_sheet_name=resolved_sheet,
                line_name=line_name,
            )
            if selected_template is None:
                selected_template = select_status_template_for_sheet(
                    status_template_all_catalog.get(project_key, []),
                    configured_sheet_name=configured_sheet,
                    resolved_sheet_name=resolved_sheet,
                    line_name=line_name,
                )
            template_map = dict(selected_template.get("column_map", {}) if selected_template else {})
            guardrails = dict(selected_template.get("guardrails", {}) if selected_template else {})
            template_sheet = str(selected_template.get("template_sheet", "") if selected_template else "")

            parse_result = _parse_status_sheet_dataframe(
                df_raw,
                guardrails=guardrails,
                template_map=template_map,
                project_code=project_code,
                project_display=project_display,
                project_scope_key=project_scope_key,
                line_name=line_name,
                line_name_source=line_source,
                source_file=selected_workbook.name,
                source_sheet=resolved_sheet,
                configured_sheet=configured_sheet,
                template_sheet=template_sheet,
            )
            if not parse_result.data.empty:
                raw_frames.append(parse_result.data)

            diagnostics_rows.append(
                {
                    "Workbook": selected_workbook.name,
                    "Project": project_code,
                    "Sheet": resolved_sheet,
                    "ConfiguredSheet": configured_sheet,
                    "LineName": line_name,
                    "LineNameSource": line_source,
                    "TemplateSheet": template_sheet,
                    "TemplateApplied": bool(template_map),
                    "TemplateChanges": "; ".join(parse_result.template_changes),
                    "FallbackNote": fallback_note or "",
                    "SectionsDetected": int(parse_result.sections_detected),
                    "HeadersDetected": int(parse_result.headers_detected),
                    "Rows": int(parse_result.rows_emitted),
                    "Status": parse_result.parse_status,
                    "Reason": parse_result.parse_reason,
                }
            )
            if parse_result.parse_status != "OK":
                issue_rows.append(
                    {
                        "Workbook": selected_workbook.name,
                        "Project": project_code,
                        "Sheet": resolved_sheet,
                        "ConfiguredSheet": configured_sheet,
                        "LineName": line_name,
                        "LineNameSource": line_source,
                        "Issue": parse_result.parse_status,
                        "Reason": parse_result.parse_reason,
                    }
                )
            coverage_rows.append(
                {
                    "project_code": project_code,
                    "project_display": project_display,
                    "status": parse_result.parse_status,
                    "reason_code": parse_result.parse_status,
                    "reason": parse_result.parse_reason,
                    "workbook": selected_workbook.name,
                    "configured_sheet": configured_sheet,
                    "resolved_sheet": resolved_sheet,
                    "rows": int(parse_result.rows_emitted),
                    "available_sheets": available_sheet_text,
                }
            )

    if skipped_blank_config:
        print(f"[pipeline] ProgressStatus: skipped {skipped_blank_config} workbook(s) per DPR_Config (blank status sheet config).")
    if skipped_not_in_config:
        print(f"[pipeline] ProgressStatus: skipped {skipped_not_in_config} workbook(s) not listed in DPR_Config.")

    raw_df = pd.concat(raw_frames, ignore_index=True) if raw_frames else pd.DataFrame(columns=RAWDATA_COLUMNS)
    diagnostics_df = pd.DataFrame(diagnostics_rows, columns=DIAGNOSTICS_COLUMNS)
    issues_df = pd.DataFrame(issue_rows, columns=ISSUES_COLUMNS)
    coverage_df = pd.DataFrame(coverage_rows, columns=COVERAGE_COLUMNS)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output_path.with_suffix(f"{output_path.suffix}.tmp")

    try:
        with pd.ExcelWriter(temp_output, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name="RawData", index=False)
            diagnostics_df.to_excel(writer, sheet_name="Diagnostics", index=False)
            issues_df.to_excel(writer, sheet_name="Issues", index=False)
            coverage_df.to_excel(writer, sheet_name="Coverage", index=False)
        temp_output.replace(output_path)
    finally:
        if temp_output.exists():
            try:
                temp_output.unlink()
            except Exception:
                pass

    print(
        f"[pipeline] ProgressStatus: wrote workbook {output_path} "
        f"(rows={len(raw_df.index)}, diagnostics={len(diagnostics_df.index)}, issues={len(issues_df.index)})."
    )
    return output_path

# =========================
# Stretch readiness ingestion
# =========================

STRETCH_RAWDATA_COLUMNS = [
    "project_code",
    "project_display",
    "project_scope_key",
    "line_name",
    "line_name_source",
    "section_label",
    "source_file",
    "source_sheet",
    "configured_sheet",
    "template_sheet",
    "report_date",
    "header_row_number",
    "source_row_number",
    "stretch_identifier",
    "from_ap",
    "to_ap",
    "length_m_raw",
    "length_km",
    "readiness_raw",
    "final_check_raw",
    "tack_welding_raw",
    "balance_towers",
    "readiness_state",
    "remarks",
    "readiness_source",
    "source_tag",
    "location_nos_raw",
    "location_parse_status",
    "location_parse_issue",
    "required_location_count",
    "matched_location_count",
    "unmatched_location_count",
    "required_locations",
    "matched_locations",
    "unmatched_locations",
]

STRETCH_SUMMARY_COLUMNS = [
    "project_code",
    "project_display",
    "project_scope_key",
    "line_name",
    "line_name_source",
    "report_date",
    "source_files",
    "source_sheets",
    "total_count",
    "ready_count",
    "partial_count",
    "not_ready_count",
    "unknown_count",
    "balance_count",
    "total_km",
    "ready_km",
    "balance_km",
    "readiness_pct",
    "basis",
]

STRETCH_MANPOWER_COLUMNS = [
    "project_code",
    "project_display",
    "project_scope_key",
    "line_name",
    "line_name_source",
    "source_file",
    "source_sheet",
    "configured_sheet",
    "header_row_number",
    "manpower_fields",
    "readiness_fields",
    "readiness_column_present",
    "signal_type",
    "non_empty_count",
    "sample_values",
    "expected_manpower",
    "expected_match",
    "status",
    "reason",
]

STRETCH_DIAGNOSTICS_COLUMNS = [
    "Workbook",
    "Project",
    "Category",
    "Sheet",
    "ConfiguredSheet",
    "LineName",
    "LineNameSource",
    "TemplateSheet",
    "TemplateApplied",
    "TemplateChanges",
    "FallbackNote",
    "SectionsDetected",
    "HeadersDetected",
    "Rows",
    "Status",
    "Reason",
]

STRETCH_ISSUES_COLUMNS = [
    "Workbook",
    "Project",
    "Category",
    "Sheet",
    "ConfiguredSheet",
    "LineName",
    "LineNameSource",
    "Issue",
    "Reason",
]

STRETCH_COVERAGE_COLUMNS = [
    "project_code",
    "project_display",
    "category",
    "status",
    "reason_code",
    "reason",
    "workbook",
    "configured_sheet",
    "resolved_sheet",
    "rows",
    "available_sheets",
]

STRETCH_READY_TOKENS_DEFAULT = ("ready", "done", "completed", "complete", "c", "yes")
STRETCH_NOT_READY_TOKENS_DEFAULT = ("not ready", "pending", "wip", "balance", "row", "hold", "no", "blocked")

STRETCH_SOURCE_DERIVED = "DERIVED_ENDPOINT_TIGHTENING"
STRETCH_SOURCE_LEGACY = "LEGACY_STRETCH_SHEET"

_LOCATION_TOKEN_SPLIT_RE = re.compile(r"\s*,\s*")
_LOCATION_FULL_TOKEN_RE = re.compile(r"^\d+[A-Z]*/\d+[A-Z]*$", flags=re.IGNORECASE)
_LOCATION_SHORTHAND_RE = re.compile(r"^[A-Z0-9]+$", flags=re.IGNORECASE)


@dataclass(frozen=True)
class StretchProjectConfig:
    stretch_entries: list[dict[str, str]]
    daily_entries: list[dict[str, str]]
    manpower_expected: str


@dataclass(frozen=True)
class StretchParseResult:
    data: pd.DataFrame
    parse_status: str
    parse_reason: str
    sections_detected: int
    headers_detected: int
    rows_emitted: int
    template_changes: list[str]


def _extract_report_date_from_filename(file_name: str) -> str:
    match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", str(file_name or ""))
    return match.group(1) if match else ""


def _col_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _scope_key_for_match(scope: object, project_code: object, line_name: object) -> str:
    scope_text = _as_text(scope)
    if scope_text:
        return ingest.normalize_project_code_key(scope_text)
    return ingest.normalize_project_code_key(f"{_as_text(project_code)}::{normalize_line_name(line_name)}")


def _normalize_location_token(value: object) -> str:
    text = _as_text(value).replace("\u00a0", " ").strip()
    if not text:
        return ""
    text = re.sub(r"^\s*AP[\s\-_./]*", "", text, flags=re.IGNORECASE)
    text = text.upper()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"(\d)\.0\b", r"\1", text)
    return text


def _is_valid_date_value(value: object) -> bool:
    text = _as_text(value)
    if not text:
        return False
    normalized = text.strip().lower()
    if normalized in {"yes", "y", "true", "ok", "done", "completed", "complete", "c"}:
        return True
    if normalized in {"no", "n", "false", "pending", "wip", "balance", "row", "hold", "blocked"}:
        return False
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
            if pd.isna(parsed):
                parsed = pd.to_datetime(text, errors="coerce")
    except Exception:
        return False
    return pd.notna(parsed)


def _report_date_with_fallback(report_date: object, source_file: object) -> str:
    text = _as_text(report_date)
    if text:
        return text
    return _extract_report_date_from_filename(_as_text(source_file))


def _report_timestamp_with_fallback(report_date: object, source_file: object) -> pd.Timestamp:
    text = _report_date_with_fallback(report_date, source_file)
    if not text:
        return pd.NaT
    try:
        return pd.to_datetime(text, errors="coerce").normalize()
    except Exception:
        return pd.NaT


def _pick_column(work: pd.DataFrame, candidates: Iterable[str]) -> str | None:
    if work is None or work.empty:
        return None
    lookup = {_col_key(col): col for col in work.columns}
    for candidate in candidates:
        key = _col_key(candidate)
        if key in lookup:
            return lookup[key]
    return None


def _pick_location_nos_column(work: pd.DataFrame) -> str | None:
    exact = _pick_column(
        work,
        (
            "location nos",
            "location_no_s",
            "locations nos",
            "location numbers",
            "location list",
            "locationnos",
        ),
    )
    if exact:
        return exact
    for col in work.columns:
        norm = _col_key(col)
        if "location" in norm and ("nos" in norm or "numbers" in norm):
            return col
    return None


def _normalize_required_location_sequence(values: Iterable[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        token = _normalize_location_token(value)
        if not token or token in seen:
            continue
        seen.add(token)
        out.append(token)
    return out


def _derive_stretch_identifier_from_stringing_row(row: pd.Series, from_ap: str, to_ap: str) -> tuple[str, str]:
    for key in ("stretch_identifier", "section", "section_name", "section_id", "section label", "stretch"):
        text = _as_text(row.get(key))
        if text:
            return text, text
    if from_ap and to_ap:
        label = f"{from_ap} - {to_ap}"
        return label, label
    section_label = _as_text(row.get("section_label"))
    if section_label:
        return section_label, section_label
    if from_ap or to_ap:
        label = from_ap or to_ap
        return label, label
    return "", ""


def _parse_location_nos_extras(location_nos_raw: object) -> tuple[list[str], str, str]:
    raw_text = _as_text(location_nos_raw)
    if not raw_text:
        return [], "EMPTY", ""

    tokens = [token.strip() for token in _LOCATION_TOKEN_SPLIT_RE.split(raw_text) if token.strip()]
    if not tokens:
        return [], "EMPTY", ""

    extras: list[str] = []
    current_anchor_prefix = ""
    saw_full_anchor = False
    unparsed_tokens: list[str] = []

    for token in tokens:
        normalized = _normalize_location_token(token)
        if not normalized:
            continue
        if _LOCATION_FULL_TOKEN_RE.fullmatch(normalized):
            saw_full_anchor = True
            current_anchor_prefix = normalized.split("/", 1)[0]
            extras.append(normalized)
            continue
        if _LOCATION_SHORTHAND_RE.fullmatch(normalized):
            if saw_full_anchor and current_anchor_prefix:
                extras.append(f"{current_anchor_prefix}/{normalized}")
            else:
                unparsed_tokens.append(normalized)
            continue
        unparsed_tokens.append(normalized)

    extras = _normalize_required_location_sequence(extras)
    if unparsed_tokens and not saw_full_anchor:
        return [], "SHORTHAND_NO_ANCHOR", "Location Nos has shorthand tokens without an explicit anchor token."
    if unparsed_tokens:
        return extras, "PARTIAL_PARSE", f"Unparsed tokens: {', '.join(unparsed_tokens[:5])}"
    return extras, "OK", ""


def _build_required_locations(from_ap: object, to_ap: object, location_nos_raw: object) -> tuple[list[str], str, str]:
    endpoints = _normalize_required_location_sequence((_as_text(from_ap), _as_text(to_ap)))
    extras, parse_status, parse_issue = _parse_location_nos_extras(location_nos_raw)
    required = _normalize_required_location_sequence([*endpoints, *extras])
    if parse_status == "SHORTHAND_NO_ANCHOR":
        required = endpoints
    if not endpoints and not required:
        return [], parse_status, parse_issue
    return required, parse_status, parse_issue


def _build_tightening_completion_maps(erection_raw: pd.DataFrame) -> tuple[dict[tuple[str, str], bool], dict[tuple[str, str], bool]]:
    if erection_raw is None or erection_raw.empty:
        return {}, {}

    work = erection_raw.copy()
    location_col = _pick_column(work, ("location no", "location no.", "location_no", "location number", "location"))
    tightening_col = _pick_column(work, ("tower tightening", "tower_tightening", "tower tightening date", "tightening date"))
    if not location_col or not tightening_col:
        return {}, {}

    for col in ("project_scope_key", "project_code", "line_name", "source_file", "report_date", "source_row_number"):
        if col not in work.columns:
            work[col] = ""
    work["__loc_norm"] = work[location_col].map(_normalize_location_token)
    work = work[work["__loc_norm"].astype(bool)].copy()
    if work.empty:
        return {}, {}

    work["__scope_norm"] = [
        _scope_key_for_match(scope, code, line)
        for scope, code, line in zip(work["project_scope_key"], work["project_code"], work["line_name"])
    ]
    work["__project_norm"] = work["project_code"].map(ingest.normalize_project_code_key)
    work["__complete"] = work[tightening_col].map(_is_valid_date_value)
    work["__report_ts"] = [
        _report_timestamp_with_fallback(report_date, source_file)
        for report_date, source_file in zip(work["report_date"], work["source_file"])
    ]
    work["__source_row"] = pd.to_numeric(work.get("source_row_number"), errors="coerce").fillna(0).astype(int)
    work["_seq"] = range(len(work.index))
    work = work.sort_values(["__scope_norm", "__project_norm", "__loc_norm", "__report_ts", "__source_row", "_seq"])
    work = work.drop_duplicates(subset=["__scope_norm", "__project_norm", "__loc_norm"], keep="last")

    by_scope: dict[tuple[str, str], bool] = {}
    by_project: dict[tuple[str, str], bool] = {}
    for _, row in work.iterrows():
        loc_norm = _as_text(row.get("__loc_norm"))
        if not loc_norm:
            continue
        complete = bool(row.get("__complete", False))
        scope_norm = _as_text(row.get("__scope_norm"))
        if scope_norm:
            by_scope[(scope_norm, loc_norm)] = complete
        project_norm = _as_text(row.get("__project_norm"))
        if project_norm:
            by_project[(project_norm, loc_norm)] = complete
    return by_scope, by_project


def _build_stretch_section_key_series(frame: pd.DataFrame) -> pd.Series:
    if frame is None or frame.empty:
        return pd.Series([], dtype="string")
    scope_series = [
        _scope_key_for_match(scope, code, line)
        for scope, code, line in zip(
            frame.get("project_scope_key", pd.Series("", index=frame.index)),
            frame.get("project_code", pd.Series("", index=frame.index)),
            frame.get("line_name", pd.Series("", index=frame.index)),
        )
    ]
    from_norm = frame.get("from_ap", pd.Series("", index=frame.index)).map(_normalize_location_token)
    to_norm = frame.get("to_ap", pd.Series("", index=frame.index)).map(_normalize_location_token)
    identifier_norm = frame.get("stretch_identifier", pd.Series("", index=frame.index)).fillna("").astype(str).map(_normalize_text)
    section_norm = frame.get("section_label", pd.Series("", index=frame.index)).fillna("").astype(str).map(_normalize_text)
    keys: list[str] = []
    for scope, from_loc, to_loc, ident, section in zip(
        scope_series,
        from_norm.tolist(),
        to_norm.tolist(),
        identifier_norm.tolist(),
        section_norm.tolist(),
    ):
        if from_loc or to_loc:
            keys.append(f"{scope}|{from_loc}|{to_loc}")
            continue
        label = ident or section
        keys.append(f"{scope}|{label}" if label else "")
    return pd.Series(keys, index=frame.index, dtype="string")


def _merge_stretch_sources_prefer_derived(legacy_df: pd.DataFrame, derived_df: pd.DataFrame) -> pd.DataFrame:
    if legacy_df is None or legacy_df.empty:
        return derived_df.reindex(columns=STRETCH_RAWDATA_COLUMNS) if derived_df is not None else pd.DataFrame(columns=STRETCH_RAWDATA_COLUMNS)
    if derived_df is None or derived_df.empty:
        return legacy_df.reindex(columns=STRETCH_RAWDATA_COLUMNS)

    legacy = legacy_df.copy()
    derived = derived_df.copy()
    legacy["__section_key"] = _build_stretch_section_key_series(legacy)
    derived["__section_key"] = _build_stretch_section_key_series(derived)
    derived_keys = {
        key
        for key in derived["__section_key"].dropna().astype(str).str.strip().tolist()
        if key
    }
    if derived_keys:
        legacy = legacy[~legacy["__section_key"].astype(str).isin(derived_keys)].copy()
    merged = pd.concat([legacy.drop(columns=["__section_key"], errors="ignore"), derived.drop(columns=["__section_key"], errors="ignore")], ignore_index=True)
    return merged.reindex(columns=STRETCH_RAWDATA_COLUMNS)


def _dedupe_stretch_latest_sections(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df is None or raw_df.empty:
        return pd.DataFrame(columns=STRETCH_RAWDATA_COLUMNS)
    work = raw_df.copy()
    work["__section_key"] = _build_stretch_section_key_series(work)
    blank_key_mask = ~work["__section_key"].fillna("").astype(str).str.strip().astype(bool)
    if blank_key_mask.any():
        work.loc[blank_key_mask, "__section_key"] = [
            f"__row_{idx}" for idx in work.index[blank_key_mask].tolist()
        ]

    work["__report_ts"] = [
        _report_timestamp_with_fallback(report_date, source_file)
        for report_date, source_file in zip(work.get("report_date", pd.Series("", index=work.index)), work.get("source_file", pd.Series("", index=work.index)))
    ]
    work["__source_rank"] = work.get("readiness_source", pd.Series("", index=work.index)).fillna("").astype(str).map(
        lambda v: 2 if v == STRETCH_SOURCE_DERIVED else 1
    )
    work["__source_row"] = pd.to_numeric(work.get("source_row_number"), errors="coerce").fillna(0).astype(int)
    work["_seq"] = range(len(work.index))
    work = work.sort_values(["__section_key", "__report_ts", "__source_rank", "__source_row", "_seq"])
    latest = work.drop_duplicates(subset=["__section_key"], keep="last")
    return latest.reindex(columns=STRETCH_RAWDATA_COLUMNS).reset_index(drop=True)


def _load_artifact_frame(
    root: Path,
    *,
    parquet_name: str,
    workbook_name: str,
    sheet_name: str,
) -> pd.DataFrame:
    parquet_path = root / parquet_name
    if parquet_path.exists():
        try:
            frame = pd.read_parquet(parquet_path)
            if isinstance(frame, pd.DataFrame):
                return frame
        except Exception:
            pass
    workbook_path = root / workbook_name
    if workbook_path.exists():
        try:
            with pd.ExcelFile(workbook_path) as xl:
                target_sheet = sheet_name
                if target_sheet not in xl.sheet_names:
                    target_sheet = next((name for name in xl.sheet_names if ingest.normalize_sheet_key(name) == ingest.normalize_sheet_key(sheet_name)), "")
                if target_sheet:
                    return xl.parse(target_sheet)
        except Exception:
            pass
    return pd.DataFrame()


def _build_derived_stretch_rows(
    *,
    parquets_root: Path,
    allowed_project_keys: set[str],
) -> pd.DataFrame:
    stringing_root = parquets_root / "Stringing"
    erection_root = parquets_root / "Erection"
    stringing_compiled = _load_artifact_frame(
        stringing_root,
        parquet_name="StringingCompiled.parquet",
        workbook_name="StringingCompiled_Output.xlsx",
        sheet_name="Stringing Compiled",
    )
    erection_raw = _load_artifact_frame(
        erection_root,
        parquet_name="RawData.parquet",
        workbook_name="ErectionCompiled_Output.xlsx",
        sheet_name="RawData",
    )
    if stringing_compiled.empty or erection_raw.empty:
        return pd.DataFrame(columns=STRETCH_RAWDATA_COLUMNS)

    compiled, _ = normalize_stringing_columns(stringing_compiled)
    compiled, _ = add_length_units(compiled)
    for column in ("project_code", "project_name", "project_display", "project", "line_name", "project_scope_key", "from_ap", "to_ap", "source_file", "source_sheet", "section_readiness", "length_m", "length_km"):
        if column not in compiled.columns:
            compiled[column] = ""
    location_nos_col = _pick_location_nos_column(compiled)
    if location_nos_col is None:
        compiled["__location_nos_raw"] = ""
    else:
        compiled["__location_nos_raw"] = compiled[location_nos_col]

    by_scope, by_project = _build_tightening_completion_maps(erection_raw)
    if not by_scope and not by_project:
        return pd.DataFrame(columns=STRETCH_RAWDATA_COLUMNS)

    rows: list[dict[str, object]] = []
    for idx, row in compiled.iterrows():
        project_code = _as_text(row.get("project_code"))
        if not project_code:
            project_code = _as_text(row.get("project")) or _as_text(row.get("project_name"))
        project_key = ingest.normalize_project_code_key(project_code)
        if allowed_project_keys and project_key not in allowed_project_keys:
            continue
        line_name = normalize_line_name(row.get("line_name", ""))
        project_display = _as_text(row.get("project_display")) or _as_text(row.get("project_name")) or project_code
        if not project_display:
            project_display = build_project_display(project_code, line_name, project_code) or project_code
        project_scope_key = _as_text(row.get("project_scope_key")) or build_project_scope_key(project_code, line_name, project_display)
        scope_norm = _scope_key_for_match(project_scope_key, project_code, line_name)
        from_ap = _as_text(row.get("from_ap") or row.get("from"))
        to_ap = _as_text(row.get("to_ap") or row.get("to"))
        location_nos_raw = row.get("__location_nos_raw")
        required_locations, parse_status, parse_issue = _build_required_locations(from_ap, to_ap, location_nos_raw)

        matched_locations: list[str] = []
        unmatched_locations: list[str] = []
        for location in required_locations:
            complete = by_scope.get((scope_norm, location))
            if complete is None and project_key:
                complete = by_project.get((project_key, location))
            if complete is True:
                matched_locations.append(location)
            else:
                unmatched_locations.append(location)

        if not required_locations:
            readiness_state = "UNKNOWN"
        else:
            readiness_state = "READY" if not unmatched_locations else "NOT_READY"

        stretch_identifier, section_label = _derive_stretch_identifier_from_stringing_row(row, from_ap, to_ap)
        report_date = _report_date_with_fallback(row.get("report_date"), row.get("source_file"))
        length_km = pd.to_numeric(pd.Series([row.get("length_km")]), errors="coerce").iloc[0]
        length_m_raw = row.get("length_m")
        source_row_number = int(pd.to_numeric(pd.Series([row.get("source_row_number", idx + 1)]), errors="coerce").fillna(idx + 1).iloc[0])
        rows.append(
            {
                "project_code": project_code,
                "project_display": project_display,
                "project_scope_key": project_scope_key,
                "line_name": line_name,
                "line_name_source": "stringing_compiled",
                "section_label": section_label,
                "source_file": _as_text(row.get("source_file")),
                "source_sheet": _as_text(row.get("source_sheet")),
                "configured_sheet": "",
                "template_sheet": "",
                "report_date": report_date,
                "header_row_number": int(pd.to_numeric(pd.Series([row.get("header_row_number")]), errors="coerce").fillna(0).iloc[0]),
                "source_row_number": source_row_number,
                "stretch_identifier": stretch_identifier,
                "from_ap": from_ap,
                "to_ap": to_ap,
                "length_m_raw": length_m_raw,
                "length_km": float(length_km) if pd.notna(length_km) else None,
                "readiness_raw": _as_text(row.get("section_readiness")),
                "final_check_raw": "",
                "tack_welding_raw": "",
                "balance_towers": None,
                "readiness_state": readiness_state,
                "remarks": parse_issue,
                "readiness_source": STRETCH_SOURCE_DERIVED,
                "source_tag": STRETCH_SOURCE_DERIVED,
                "location_nos_raw": _as_text(location_nos_raw),
                "location_parse_status": parse_status,
                "location_parse_issue": parse_issue,
                "required_location_count": int(len(required_locations)),
                "matched_location_count": int(len(matched_locations)),
                "unmatched_location_count": int(len(unmatched_locations)),
                "required_locations": ", ".join(required_locations),
                "matched_locations": ", ".join(matched_locations),
                "unmatched_locations": ", ".join(unmatched_locations),
            }
        )

    if not rows:
        return pd.DataFrame(columns=STRETCH_RAWDATA_COLUMNS)
    return pd.DataFrame(rows, columns=STRETCH_RAWDATA_COLUMNS)


def _length_km_from_row(row: pd.Series, *, length_unit: str) -> float | None:
    direct_km = _coerce_numeric(row.get("length_km"))
    if direct_km is not None:
        return direct_km
    raw_length = None
    for key in ("length_m", "section_length_m", "span_m", "length"):
        raw_length = _coerce_numeric(row.get(key))
        if raw_length is not None:
            break
    if raw_length is None:
        return None
    unit = _normalize_text(length_unit)
    if unit == "km":
        return raw_length
    if unit == "m":
        return raw_length / 1000.0
    return raw_length / 1000.0 if raw_length > 50 else raw_length


def _is_positive_ready_token(value: object, ready_tokens: tuple[str, ...]) -> bool:
    text = _normalize_text(value)
    if not text:
        return False
    for token in ready_tokens:
        token_norm = _normalize_text(token)
        if not token_norm:
            continue
        if token_norm == "c":
            if text == "c":
                return True
            continue
        if token_norm in text:
            return True
    return False


def _is_negative_ready_token(value: object, not_ready_tokens: tuple[str, ...]) -> bool:
    text = _normalize_text(value)
    if not text:
        return False
    for token in not_ready_tokens:
        token_norm = _normalize_text(token)
        if token_norm and token_norm in text:
            return True
    return False


def _is_date_like_completion(value: object) -> bool:
    """Treat populated completion-style date fields as READY signal."""
    text = _as_text(value)
    if not text:
        return False
    # Fast path for common spreadsheet date strings.
    if re.search(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", text) or re.search(r"\b\d{4}-\d{2}-\d{2}\b", text):
        return True
    try:
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    except Exception:
        return False
    return pd.notna(parsed)


def _normalize_stretch_readiness_state(
    *,
    rule: str,
    readiness_raw: object,
    final_check_raw: object,
    tack_welding_raw: object,
    balance_towers: object,
    ready_tokens: tuple[str, ...],
    not_ready_tokens: tuple[str, ...],
) -> str:
    rule_norm = _normalize_text(rule)
    readiness = _as_text(readiness_raw)
    final_check = _as_text(final_check_raw)
    tack = _as_text(tack_welding_raw)
    balance = _coerce_numeric(balance_towers)

    if rule_norm == "both_required":
        final_ready = _is_positive_ready_token(final_check, ready_tokens) or _is_date_like_completion(final_check_raw)
        tack_ready = _is_positive_ready_token(tack, ready_tokens) or _is_date_like_completion(tack_welding_raw)
        if final_ready and tack_ready:
            return "READY"
        if final_ready or tack_ready:
            return "PARTIAL"
        if _is_negative_ready_token(final_check, not_ready_tokens) or _is_negative_ready_token(tack, not_ready_tokens):
            return "NOT_READY"
        return "UNKNOWN"

    if _is_positive_ready_token(readiness, ready_tokens):
        return "READY"
    if _is_negative_ready_token(readiness, not_ready_tokens):
        return "NOT_READY"
    if balance is not None:
        return "READY" if balance <= 0 else "NOT_READY"
    final_ready = _is_positive_ready_token(final_check, ready_tokens) or _is_date_like_completion(final_check_raw)
    tack_ready = _is_positive_ready_token(tack, ready_tokens) or _is_date_like_completion(tack_welding_raw)
    if final_ready and tack_ready:
        return "READY"
    if final_ready or tack_ready:
        return "PARTIAL"
    return "UNKNOWN"


def _pick_stretch_identifier(row: pd.Series) -> str:
    for key in ("stretch_identifier", "section", "section_name", "loc_no", "location_no", "sec"):
        text = _as_text(row.get(key))
        if text:
            return text
    from_ap = _as_text(row.get("from_ap") or row.get("from"))
    to_ap = _as_text(row.get("to_ap") or row.get("to"))
    if from_ap and to_ap:
        return f"{from_ap} - {to_ap}"
    return from_ap or to_ap


def _resolve_named_template_sheet_stretch(wb, expected_name: str) -> Optional[str]:
    expected_key = ingest.normalize_space_only(expected_name)
    for name in wb.sheetnames:
        if ingest.normalize_space_only(name) == expected_key:
            return name
    return None


def _resolve_project_template_sheets_stretch(wb, project_name: object, discipline: str) -> list[str]:
    project_text = str(project_name or "").strip()
    if not project_text:
        return []
    resolved: list[str] = []
    seen: set[str] = set()
    for expected in (f"{project_text} {discipline}", f"{project_text} {discipline} Template Check"):
        hit = _resolve_named_template_sheet_stretch(wb, expected)
        if hit and hit not in seen:
            seen.add(hit)
            resolved.append(hit)
    project_key = ingest.normalize_space_only(project_text)
    discipline_key = ingest.normalize_space_only(discipline)
    for name in wb.sheetnames:
        key = ingest.normalize_space_only(name)
        if key and key.startswith(project_key) and key.endswith(discipline_key) and name not in seen:
            seen.add(name)
            resolved.append(name)
    return resolved


def _extract_template_column_map_stretch(ws) -> dict[int, str]:
    to_map_row = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if ingest.normalize_space_only(cell) == "to map":
                to_map_row = row_idx
                break
        if to_map_row is not None:
            break
    if to_map_row is None:
        return {}
    labels_row = to_map_row + 1
    row_values = next(ws.iter_rows(min_row=labels_row, max_row=labels_row, values_only=True), ())
    mapping: dict[int, str] = {}
    for col_idx, value in enumerate(row_values):
        label = str(value).strip() if value is not None else ""
        if label:
            mapping[col_idx] = label
    return mapping


def _extract_guardrails_stretch(ws) -> dict[str, str]:
    guardrails: dict[str, str] = {}
    in_guardrails = False
    for row in ws.iter_rows(values_only=True):
        first_raw = row[0] if len(row) > 0 else None
        second_raw = row[1] if len(row) > 1 else None
        first_norm = ingest.normalize_space_only(first_raw)
        if first_norm == "guardrails":
            in_guardrails = True
            continue
        if first_norm == "to map":
            break
        if not in_guardrails:
            continue
        key = _normalize_key(first_raw)
        value = _as_text(second_raw)
        if key and value:
            guardrails[key] = value
    return guardrails

def load_stretch_sheet_config(raw_root: Path, *, repo_root: Path | None = None) -> dict[str, StretchProjectConfig]:
    config_path = ingest.resolve_dpr_config_path(raw_root, repo_root=repo_root)
    if config_path is None:
        return {}
    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception:
        return {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}
        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}
        headers = [ingest.normalize_space_only(v) for v in header_row]
        required = [
            "project code",
            "stretch readiness sheet names",
            "stretch daily stringing sheet names",
            "stretch line names",
            "stretch manpower expected",
        ]
        if any(col not in headers for col in required):
            return {}

        project_idx = headers.index("project code")
        stretch_idx = headers.index("stretch readiness sheet names")
        daily_idx = headers.index("stretch daily stringing sheet names")
        line_idx = headers.index("stretch line names")
        manpower_idx = headers.index("stretch manpower expected")

        mapping: dict[str, StretchProjectConfig] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project_val = row[project_idx] if project_idx < len(row) else None
            if project_val in (None, ""):
                continue
            project_key = ingest.normalize_project_code_key(project_val)
            raw_stretch = row[stretch_idx] if stretch_idx < len(row) else None
            raw_daily = row[daily_idx] if daily_idx < len(row) else None
            raw_line_names = row[line_idx] if line_idx < len(row) else None
            manpower_expected = _normalize_text(row[manpower_idx] if manpower_idx < len(row) else "")
            if manpower_expected not in {"yes", "no", "unknown"}:
                manpower_expected = "unknown"

            stretch_entries = parse_sheet_line_entries(raw_stretch, raw_line_names, "stretch", infer_from_sheet_name=False)
            daily_entries = parse_sheet_line_entries(raw_daily, raw_line_names, "stringing", infer_from_sheet_name=False)

            def _dedupe(entries: list[dict[str, str]]) -> list[dict[str, str]]:
                out: list[dict[str, str]] = []
                seen: set[str] = set()
                for entry in entries:
                    key = ingest.normalize_space_only(entry.get("sheet_name"))
                    if not key or key in seen:
                        continue
                    seen.add(key)
                    out.append(entry)
                return out

            mapping[project_key] = StretchProjectConfig(
                stretch_entries=_dedupe(stretch_entries),
                daily_entries=_dedupe(daily_entries),
                manpower_expected=manpower_expected,
            )
            if (
                not mapping[project_key].stretch_entries
                and not mapping[project_key].daily_entries
                and mapping[project_key].manpower_expected == "unknown"
            ):
                mapping.pop(project_key, None)
        return mapping
    finally:
        wb.close()


def load_stretch_template_mapping_catalog(
    raw_root: Path,
    *,
    repo_root: Path | None = None,
    include_unchecked: bool = False,
) -> tuple[dict[str, list[dict[str, object]]], dict[str, str]]:
    config_path = ingest.resolve_dpr_config_path(raw_root, repo_root=repo_root)
    if config_path is None:
        return {}, {}
    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception:
        return {}, {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}, {}
        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}, {}
        headers = [ingest.normalize_space_only(value) for value in header_row]
        project_idx = headers.index("project code") if "project code" in headers else None
        check_idx = headers.index("stretch template check") if "stretch template check" in headers else None
        if project_idx is None or (check_idx is None and not include_unchecked):
            return {}, {}

        catalog: dict[str, list[dict[str, object]]] = {}
        errors: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project_val = row[project_idx] if project_idx < len(row) else None
            if project_val in (None, ""):
                continue
            check_val = row[check_idx] if check_idx is not None and check_idx < len(row) else None
            check_enabled = ingest.normalize_space_only(check_val) == "yes"
            if not check_enabled and not include_unchecked:
                continue

            project_key = ingest.normalize_project_code_key(project_val)
            template_sheets = _resolve_project_template_sheets_stretch(wb, project_val, "Stretch")
            if not template_sheets:
                if check_enabled:
                    errors[project_key] = (
                        f"Stretch Template Check is Yes but no mapping tab matching project '{str(project_val).strip()}' was found."
                    )
                continue

            options: list[dict[str, object]] = []
            for sheet_name in template_sheets:
                ws_template = wb[sheet_name]
                col_map = _extract_template_column_map_stretch(ws_template)
                guardrails = _extract_guardrails_stretch(ws_template)
                if not col_map and not guardrails:
                    continue
                options.append(
                    {
                        "column_map": col_map,
                        "template_sheet": sheet_name,
                        "guardrails": guardrails,
                        "numeric_tokens": _extract_numeric_tokens(sheet_name),
                    }
                )
            if not options:
                if check_enabled:
                    errors[project_key] = (
                        f"Stretch template tab(s) for project '{str(project_val).strip()}' have no usable Guardrails/To Map rows."
                    )
                continue
            catalog[project_key] = options
        return catalog, errors
    finally:
        wb.close()


def select_stretch_template_for_sheet(
    template_options: list[dict[str, object]] | None,
    *,
    configured_sheet_name: str = "",
    resolved_sheet_name: str = "",
    line_name: str = "",
) -> dict[str, object] | None:
    if not template_options:
        return None
    hints = [configured_sheet_name or "", resolved_sheet_name or "", line_name or ""]
    hint_keys = [ingest.normalize_space_only(value) for value in hints if ingest.normalize_space_only(value)]
    hint_numbers: set[str] = set()
    for value in hints:
        hint_numbers.update(_extract_numeric_tokens(value))

    best: dict[str, object] | None = None
    best_score = float("-inf")
    for idx, option in enumerate(template_options):
        sheet_name = str(option.get("template_sheet", "")).strip()
        sheet_key = ingest.normalize_space_only(sheet_name)
        sheet_numbers = set(option.get("numeric_tokens", set()) or set())
        col_map = option.get("column_map", {}) or {}
        score = 0.0
        for hint in hint_keys:
            if sheet_key == hint:
                score += 1000.0
            elif hint and (hint in sheet_key or sheet_key in hint):
                score += 200.0
        if hint_numbers:
            score += float(len(sheet_numbers.intersection(hint_numbers))) * 120.0
        score += float(len(col_map))
        score -= idx * 1e-4
        if score > best_score:
            best_score = score
            best = option
    return best


def _sheet_tokens_stretch(value: object) -> set[str]:
    text = str(value or "").strip().lower()
    if not text:
        return set()
    return {token for token in re.findall(r"[a-z0-9]+", text) if token}


def _sheet_match_score_stretch(configured_sheet: str, workbook_sheet: str) -> tuple[int, int, int]:
    configured = str(configured_sheet or "").strip()
    workbook = str(workbook_sheet or "").strip()
    if not configured or not workbook:
        return (0, 0, 0)
    if ingest.normalize_space_only(configured) == ingest.normalize_space_only(workbook):
        return (1, 0, 0)
    if ingest.normalize_sheet_key(configured) == ingest.normalize_sheet_key(workbook):
        return (0, 1, 0)
    c_tokens = _sheet_tokens_stretch(configured)
    w_tokens = _sheet_tokens_stretch(workbook)
    return (0, 0, len(c_tokens.intersection(w_tokens)) if c_tokens and w_tokens else 0)


def _pick_best_workbook_for_sheet_stretch(
    workbooks: list[Path],
    workbook_sheets: dict[str, list[str]],
    configured_sheet: str,
    configured_line_name: str = "",
) -> tuple[Path | None, str]:
    scored: list[tuple[tuple[int, int, int, int, float], Path, str]] = []
    line_hint = ingest.normalize_space_only(configured_line_name)
    for workbook in workbooks:
        names = workbook_sheets.get(str(workbook.resolve()), [])
        best_score = (0, 0, 0)
        best_name = ""
        for name in names:
            score = _sheet_match_score_stretch(configured_sheet, name)
            if score > best_score:
                best_score = score
                best_name = name
        identity = parse_project_identity_from_filename(workbook.name)
        wb_line = ingest.normalize_space_only(identity.get("line_name", ""))
        line_bonus = 1 if line_hint and wb_line and line_hint == wb_line else 0
        mtime = workbook.stat().st_mtime if workbook.exists() else 0.0
        scored.append(((best_score[0], best_score[1], best_score[2], line_bonus, mtime), workbook, best_name))

    if not scored:
        return None, ""
    scored.sort(key=lambda item: item[0], reverse=True)
    top = scored[0]
    if top[0][:3] == (0, 0, 0):
        return None, ""
    return top[1], top[2]


def _pick_matching_workbooks_for_sheet_stretch(
    workbooks: list[Path],
    workbook_sheets: dict[str, list[str]],
    configured_sheet: str,
    configured_line_name: str = "",
) -> list[tuple[Path, str]]:
    scored: list[tuple[tuple[int, int, int, int, float], Path, str, str, str]] = []
    line_hint = ingest.normalize_space_only(configured_line_name)
    for workbook in workbooks:
        names = workbook_sheets.get(str(workbook.resolve()), [])
        if configured_sheet:
            best_score = (0, 0, 0)
            best_name = ""
            for name in names:
                score = _sheet_match_score_stretch(configured_sheet, name)
                if score > best_score:
                    best_score = score
                    best_name = name
            if best_score == (0, 0, 0):
                continue
        else:
            best_score = (1, 0, 0)
            best_name = names[0] if names else ""
        identity = parse_project_identity_from_filename(workbook.name)
        wb_line = ingest.normalize_space_only(identity.get("line_name", ""))
        line_bonus = 1 if line_hint and wb_line and line_hint == wb_line else 0
        mtime = workbook.stat().st_mtime if workbook.exists() else 0.0
        report_date = _extract_report_date_from_filename(workbook.name)
        scored.append(((best_score[0], best_score[1], best_score[2], line_bonus, mtime), workbook, best_name, report_date, wb_line))

    if not scored:
        return []

    dedup: dict[tuple[str, str, str], tuple[tuple[int, int, int, int, float], Path, str, str, str]] = {}
    for entry in scored:
        _, workbook, best_name, report_date, wb_line = entry
        dedupe_key = (
            report_date or f"__{workbook.resolve()}",
            wb_line,
            ingest.normalize_space_only(best_name),
        )
        prev = dedup.get(dedupe_key)
        if prev is None or entry[0] > prev[0]:
            dedup[dedupe_key] = entry

    picked = list(dedup.values())
    picked.sort(
        key=lambda item: (
            item[3] if item[3] else "9999-12-31",
            str(item[1]).lower(),
            item[2].lower(),
        )
    )
    return [(item[1], item[2]) for item in picked]


def _build_exact_sheet_selector_stretch(sheet_name: str):
    expected_space = ingest.normalize_space_only(sheet_name)
    expected_compact = ingest.normalize_sheet_key(sheet_name)

    def _selector(sheet_names: list[str]) -> str | None:
        by_space: dict[str, str] = {}
        by_compact: dict[str, str] = {}
        for existing in sheet_names:
            space_key = ingest.normalize_space_only(existing)
            compact_key = ingest.normalize_sheet_key(existing)
            if space_key and space_key not in by_space:
                by_space[space_key] = existing
            if compact_key and compact_key not in by_compact:
                by_compact[compact_key] = existing
        hit = by_space.get(expected_space)
        if hit:
            return hit
        hit = by_compact.get(expected_compact)
        if hit:
            return hit
        for existing in sheet_names:
            compact = ingest.normalize_sheet_key(existing)
            if expected_compact and (expected_compact in compact or compact in expected_compact):
                return existing
        return None

    return _selector

def _build_stretch_rows_from_block(
    df: pd.DataFrame,
    *,
    project_code: str,
    project_display: str,
    project_scope_key: str,
    line_name: str,
    line_name_source: str,
    section_label: str,
    source_file: str,
    source_sheet: str,
    configured_sheet: str,
    template_sheet: str,
    report_date: str,
    header_row_number: int,
    guardrails: dict[str, str],
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=STRETCH_RAWDATA_COLUMNS)

    ready_tokens = tuple(_split_tokens(guardrails.get("ready_tokens", "")) or STRETCH_READY_TOKENS_DEFAULT)
    not_ready_tokens = tuple(_split_tokens(guardrails.get("not_ready_tokens", "")) or STRETCH_NOT_READY_TOKENS_DEFAULT)
    row_exclude_regex = guardrails.get("row_exclude_regex", "")
    row_exclude_obj = None
    if row_exclude_regex:
        try:
            row_exclude_obj = re.compile(row_exclude_regex, flags=re.IGNORECASE)
        except re.error:
            row_exclude_obj = None

    rows: list[dict[str, object]] = []
    for _, row in df.iterrows():
        identifier = _pick_stretch_identifier(row)
        readiness_raw = _as_text(
            row.get("readiness_raw")
            or row.get("section_readiness")
            or row.get("stretch_readiness")
            or row.get("status_readyness")
            or row.get("status_readiness")
            or row.get("erection")
        )
        final_check_raw = _as_text(row.get("final_check_raw") or row.get("final_checking") or row.get("final_check"))
        tack_welding_raw = _as_text(row.get("tack_welding_raw") or row.get("tack_welding"))
        balance_towers = _coerce_numeric(row.get("balance_towers") or row.get("erection_balance") or row.get("balance"))
        remarks = _as_text(row.get("remarks") or row.get("remark") or row.get("notes"))
        from_ap = _as_text(row.get("from_ap") or row.get("from"))
        to_ap = _as_text(row.get("to_ap") or row.get("to"))
        length_m_raw = row.get("length_m") or row.get("section_length_m") or row.get("span_m") or row.get("length")
        length_km = _length_km_from_row(row, length_unit=guardrails.get("length_unit", "auto"))

        if not identifier and not readiness_raw and not final_check_raw and not tack_welding_raw and balance_towers is None:
            continue
        if row_exclude_obj is not None and identifier and row_exclude_obj.search(identifier):
            continue

        readiness_state = _normalize_stretch_readiness_state(
            rule=guardrails.get("readiness_rule", ""),
            readiness_raw=readiness_raw,
            final_check_raw=final_check_raw,
            tack_welding_raw=tack_welding_raw,
            balance_towers=balance_towers,
            ready_tokens=ready_tokens,
            not_ready_tokens=not_ready_tokens,
        )
        rows.append(
            {
                "project_code": project_code,
                "project_display": project_display,
                "project_scope_key": project_scope_key,
                "line_name": line_name,
                "line_name_source": line_name_source,
                "section_label": section_label,
                "source_file": source_file,
                "source_sheet": source_sheet,
                "configured_sheet": configured_sheet,
                "template_sheet": template_sheet,
                "report_date": report_date,
                "header_row_number": int(header_row_number),
                "source_row_number": int(row.get("__source_row_number", 0) or 0),
                "stretch_identifier": identifier,
                "from_ap": from_ap,
                "to_ap": to_ap,
                "length_m_raw": length_m_raw,
                "length_km": length_km,
                "readiness_raw": readiness_raw,
                "final_check_raw": final_check_raw,
                "tack_welding_raw": tack_welding_raw,
                "balance_towers": balance_towers,
                "readiness_state": readiness_state,
                "remarks": remarks,
                "readiness_source": STRETCH_SOURCE_LEGACY,
                "source_tag": STRETCH_SOURCE_LEGACY,
                "location_nos_raw": "",
                "location_parse_status": "",
                "location_parse_issue": "",
                "required_location_count": None,
                "matched_location_count": None,
                "unmatched_location_count": None,
                "required_locations": "",
                "matched_locations": "",
                "unmatched_locations": "",
            }
        )
    return pd.DataFrame(rows, columns=STRETCH_RAWDATA_COLUMNS)


def _parse_stretch_sheet_dataframe(
    df_raw: pd.DataFrame,
    *,
    guardrails: dict[str, str],
    template_map: dict[int, str],
    project_code: str,
    project_display: str,
    project_scope_key: str,
    line_name: str,
    line_name_source: str,
    source_file: str,
    source_sheet: str,
    configured_sheet: str,
    template_sheet: str,
    report_date: str,
) -> StretchParseResult:
    if df_raw is None or df_raw.empty:
        return StretchParseResult(pd.DataFrame(columns=STRETCH_RAWDATA_COLUMNS), "EMPTY_SHEET", "Sheet is empty.", 0, 0, 0, [])

    required_tokens = [_normalize_text(token) for token in _split_tokens(guardrails.get("required_tokens", ""))]
    required_tokens = [token for token in required_tokens if token] or ["section"]
    stop_tokens = [_normalize_text(token) for token in _split_tokens(guardrails.get("stop_tokens", ""))]
    header_rows_options = _parse_header_rows_guardrail(guardrails.get("header_rows", ""))
    section_ranges = _extract_section_ranges(
        df_raw,
        start_row=0,
        end_row=len(df_raw.index) - 1,
        section_contains=guardrails.get("section_start_contains", ""),
        section_regex=guardrails.get("section_split_regex", ""),
        section_label_regex=guardrails.get("section_label_regex", ""),
    )

    all_rows: list[pd.DataFrame] = []
    headers_detected = 0
    sections_detected = 0
    template_changes_all: list[str] = []

    for section_label, section_start, section_end in section_ranges:
        anchor_row = _find_anchor_row(
            df_raw,
            section_start,
            section_end,
            anchor_text=guardrails.get("block_anchor", ""),
            anchor_regex=guardrails.get("block_anchor_regex", ""),
        )
        header_row, header_span, header_labels = _detect_header(
            df_raw,
            start_row=anchor_row,
            end_row=section_end,
            required_tokens=required_tokens,
            header_rows_options=header_rows_options,
        )
        if header_row is None or header_span is None:
            continue

        headers_detected += 1
        sections_detected += 1
        width = len(header_labels)
        data_rows: list[tuple[int, list[object]]] = []
        blank_run = 0
        started = False
        for row_idx in range(header_row + header_span, section_end + 1):
            row_values = [df_raw.iat[row_idx, col_idx] if col_idx < len(df_raw.columns) else None for col_idx in range(width)]
            row_text_norm = _normalize_text(" ".join(_as_text(v) for v in row_values[: min(30, width)]))
            if stop_tokens and any(token in row_text_norm for token in stop_tokens):
                if started:
                    break
                continue
            if _row_is_blank(row_values):
                if started:
                    blank_run += 1
                    if blank_run >= 2:
                        break
                continue
            started = True
            blank_run = 0
            data_rows.append((row_idx + 1, row_values))

        if not data_rows:
            continue

        block_df = pd.DataFrame([values for _, values in data_rows], columns=header_labels)
        block_df["__source_row_number"] = [row_number for row_number, _ in data_rows]
        block_df, template_changes = _apply_template_column_mapping(block_df, template_map)
        template_changes_all.extend(template_changes)
        prepared = _build_stretch_rows_from_block(
            block_df,
            project_code=project_code,
            project_display=project_display,
            project_scope_key=project_scope_key,
            line_name=line_name,
            line_name_source=line_name_source,
            section_label=section_label,
            source_file=source_file,
            source_sheet=source_sheet,
            configured_sheet=configured_sheet,
            template_sheet=template_sheet,
            report_date=report_date,
            header_row_number=header_row + 1,
            guardrails=guardrails,
        )
        if not prepared.empty:
            all_rows.append(prepared)

    if not all_rows:
        status = "NO_MATCHED_ROWS" if headers_detected > 0 else "HEADER_NOT_FOUND"
        return StretchParseResult(
            pd.DataFrame(columns=STRETCH_RAWDATA_COLUMNS),
            status,
            "No matching rows after guardrails/data filters.",
            sections_detected,
            headers_detected,
            0,
            template_changes_all,
        )

    combined = pd.concat(all_rows, ignore_index=True)
    return StretchParseResult(
        combined.reindex(columns=STRETCH_RAWDATA_COLUMNS),
        "OK",
        "",
        sections_detected,
        headers_detected,
        int(len(combined.index)),
        template_changes_all,
    )


def _detect_manpower_signal_from_frame(frame: pd.DataFrame) -> dict[str, object]:
    if frame is None or frame.empty:
        return {
            "signal_type": "ABSENT",
            "manpower_fields": [],
            "readiness_fields": [],
            "readiness_present": False,
            "non_empty_count": 0,
            "sample_values": [],
        }

    manpower_fields: list[str] = []
    readiness_fields: list[str] = []
    non_empty_total = 0
    sample_values: list[str] = []
    for column in frame.columns:
        col_text = _normalize_text(column)
        col_compact = ingest.normalize_sheet_key(column)
        if "readiness" in col_text or "readyness" in col_text:
            readiness_fields.append(str(column))
        is_manpower = (
            "manpower" in col_text
            or "gang strength" in col_text
            or col_text == "mp"
            or col_compact == "mp"
        )
        if not is_manpower:
            continue
        manpower_fields.append(str(column))
        series = frame[column].astype(str).str.strip()
        series = series[~series.str.lower().isin({"", "nan", "none", "null"})]
        non_empty_total += int(series.shape[0])
        for value in series.head(3).tolist():
            if value not in sample_values:
                sample_values.append(value)

    if not manpower_fields:
        signal_type = "ABSENT"
    elif non_empty_total > 0:
        signal_type = "PRESENT_WITH_VALUES"
    else:
        signal_type = "HEADER_ONLY"

    return {
        "signal_type": signal_type,
        "manpower_fields": manpower_fields,
        "readiness_fields": readiness_fields,
        "readiness_present": bool(readiness_fields),
        "non_empty_count": non_empty_total,
        "sample_values": sample_values[:5],
    }


def _detect_manpower_signal_from_raw_table(df_raw: pd.DataFrame) -> dict[str, object]:
    if df_raw is None or df_raw.empty:
        return {
            "signal_type": "ABSENT",
            "manpower_fields": [],
            "readiness_fields": [],
            "readiness_present": False,
            "non_empty_count": 0,
            "sample_values": [],
        }

    manpower_fields: list[str] = []
    readiness_fields: list[str] = []
    non_empty_total = 0
    sample_values: list[str] = []
    max_rows = min(len(df_raw.index), 20)
    max_cols = min(len(df_raw.columns), 40)

    for r in range(max_rows):
        for c in range(max_cols):
            cell = df_raw.iat[r, c]
            text = _normalize_text(cell)
            if not text:
                continue
            is_readiness = ("readiness" in text) or ("readyness" in text)
            is_manpower = text in {"mp", "manpower", "gang strength"} or ("gang strength" in text)
            if is_readiness:
                readiness_fields.append(f"R{r+1}C{c+1}:{_as_text(cell)}")
            if not is_manpower:
                continue
            manpower_fields.append(f"R{r+1}C{c+1}:{_as_text(cell)}")
            if r + 1 < len(df_raw.index):
                series = df_raw.iloc[r + 1 :, c].astype(str).str.strip()
                series = series[~series.str.lower().isin({"", "nan", "none", "null"})]
                non_empty_total += int(series.shape[0])
                for value in series.head(3).tolist():
                    if value not in sample_values:
                        sample_values.append(value)

    if not manpower_fields:
        signal_type = "ABSENT"
    elif non_empty_total > 0:
        signal_type = "PRESENT_WITH_VALUES"
    else:
        signal_type = "HEADER_ONLY"
    return {
        "signal_type": signal_type,
        "manpower_fields": manpower_fields,
        "readiness_fields": readiness_fields,
        "readiness_present": bool(readiness_fields),
        "non_empty_count": non_empty_total,
        "sample_values": sample_values[:5],
    }


def _expected_match_stretch(expected: str, signal_type: str) -> bool:
    expected_norm = _normalize_text(expected)
    signal_norm = _normalize_text(signal_type)
    if expected_norm == "yes":
        return signal_norm in {"present_with_values", "header_only"}
    if expected_norm == "no":
        return signal_norm in {"absent", "header_only", "no_sheet_config"}
    return True


def _build_stretch_summary(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df is None or raw_df.empty:
        return pd.DataFrame(columns=STRETCH_SUMMARY_COLUMNS)

    work = _dedupe_stretch_latest_sections(raw_df)
    if work.empty:
        return pd.DataFrame(columns=STRETCH_SUMMARY_COLUMNS)

    readiness = work.get("readiness_state", pd.Series("", index=work.index)).fillna("").astype(str).str.upper().str.strip()
    readiness_allowed = {"READY", "PARTIAL", "NOT_READY", "UNKNOWN"}
    work["readiness_state"] = readiness.where(readiness.isin(readiness_allowed), "UNKNOWN")
    work["length_km"] = pd.to_numeric(work.get("length_km"), errors="coerce")
    rows: list[dict[str, object]] = []
    group_cols = ["project_code", "project_display", "project_scope_key", "line_name", "line_name_source"]
    for keys, group in work.groupby(group_cols, dropna=False):
        project_code, project_display, project_scope_key, line_name, line_source = keys
        counts = group["readiness_state"].fillna("UNKNOWN").astype(str).str.upper().value_counts()
        total_count = int(len(group.index))
        ready_count = int(counts.get("READY", 0))
        partial_count = int(counts.get("PARTIAL", 0))
        not_ready_count = int(counts.get("NOT_READY", 0))
        unknown_count = int(counts.get("UNKNOWN", 0))
        total_km = float(group["length_km"].dropna().sum()) if "length_km" in group.columns else 0.0
        ready_km = float(group.loc[group["readiness_state"] == "READY", "length_km"].dropna().sum())
        report_timestamps = [
            _report_timestamp_with_fallback(report_date, source_file)
            for report_date, source_file in zip(
                group.get("report_date", pd.Series("", index=group.index)),
                group.get("source_file", pd.Series("", index=group.index)),
            )
        ]
        report_dates = [ts for ts in report_timestamps if pd.notna(ts)]
        rows.append(
            {
                "project_code": project_code,
                "project_display": project_display,
                "project_scope_key": project_scope_key,
                "line_name": line_name,
                "line_name_source": line_source,
                "report_date": max(report_dates).strftime("%Y-%m-%d") if report_dates else "",
                "source_files": int(group["source_file"].astype(str).nunique()),
                "source_sheets": int(group["source_sheet"].astype(str).nunique()),
                "total_count": total_count,
                "ready_count": ready_count,
                "partial_count": partial_count,
                "not_ready_count": not_ready_count,
                "unknown_count": unknown_count,
                "balance_count": max(total_count - ready_count, 0),
                "total_km": round(total_km, 6) if total_km else 0.0,
                "ready_km": round(ready_km, 6) if ready_km else 0.0,
                "balance_km": round(max(total_km - ready_km, 0.0), 6) if total_km else 0.0,
                "readiness_pct": round((ready_count / total_count) * 100.0, 2) if total_count > 0 else None,
                "basis": "count_and_km",
            }
        )
    return pd.DataFrame(rows, columns=STRETCH_SUMMARY_COLUMNS)

def _stretch_candidates(input_dir: Optional[Path], files: Optional[list[Path]]) -> list[Path]:
    if files:
        return [
            path
            for path in files
            if path.suffix.lower() in (".xlsx", ".xlsm", ".xls") and path.exists() and not path.name.startswith("~$")
        ]
    if input_dir and input_dir.exists():
        return sorted([path for path in input_dir.rglob("*.xls*") if path.is_file() and not path.name.startswith("~$")])
    return []


def compile_stretch_readiness_to_workbook(
    input_dir: Optional[Path],
    files: Optional[list[Path]],
    output_path: Path,
    *,
    repo_root: Path | None = None,
    completed_project_keys: set[str] | None = None,
) -> Path | None:
    output_path = Path(output_path)
    candidates = _stretch_candidates(input_dir, files)
    if completed_project_keys:
        kept: list[Path] = []
        skipped_files = 0
        skipped_projects: set[str] = set()
        for workbook in candidates:
            identity = parse_project_identity_from_filename(workbook.name)
            project_code = str(identity.get("project_code", "")).strip() or workbook.stem
            if project_code and is_completed_project(project_code, completed_project_keys):
                skipped_files += 1
                skipped_projects.add(project_code.upper())
                continue
            kept.append(workbook)
        if skipped_files:
            print(
                f"[pipeline] StretchReadiness: skipped_completed_files={skipped_files}, "
                f"skipped_completed_projects={len(skipped_projects)}"
            )
        candidates = kept
    if not candidates:
        print("[pipeline] StretchReadiness: no candidate files found; skipping.")
        return None

    raw_root = input_dir if input_dir is not None else files[0].parent if files else Path(".")
    stretch_config = load_stretch_sheet_config(raw_root, repo_root=repo_root)
    stretch_template_catalog, stretch_template_errors = load_stretch_template_mapping_catalog(raw_root, repo_root=repo_root)
    stretch_template_all_catalog, _ = load_stretch_template_mapping_catalog(raw_root, repo_root=repo_root, include_unchecked=True)
    has_stretch_config = bool(stretch_config)

    workbook_sheet_cache: dict[str, list[str]] = {}
    workbooks_by_project: dict[str, list[Path]] = {}
    for workbook in candidates:
        identity = parse_project_identity_from_filename(workbook.name)
        project_code = str(identity.get("project_code", "")).strip() or workbook.stem
        project_key = ingest.normalize_project_code_key(project_code)
        workbooks_by_project.setdefault(project_key, []).append(workbook)
        sheet_names, _ = ingest.list_excel_sheet_names(workbook)
        workbook_sheet_cache[str(workbook.resolve())] = list(sheet_names or [])

    raw_frames: list[pd.DataFrame] = []
    manpower_rows: list[dict[str, object]] = []
    diagnostics_rows: list[dict[str, object]] = []
    issue_rows: list[dict[str, object]] = []
    coverage_rows: list[dict[str, object]] = []
    skipped_not_in_config = 0
    processed_project_keys: set[str] = set()

    for project_key, project_workbooks in sorted(workbooks_by_project.items()):
        cfg = stretch_config.get(project_key)
        if has_stretch_config and cfg is None:
            skipped_not_in_config += len(project_workbooks)
            continue
        if cfg is None:
            continue
        processed_project_keys.add(project_key)

        identity_base = parse_project_identity_from_filename(project_workbooks[0].name)
        project_code_base = str(identity_base.get("project_code", "")).strip() or project_workbooks[0].stem
        template_error = stretch_template_errors.get(project_key, "")
        workbook_lines: dict[str, dict[str, object]] = {}
        for workbook in project_workbooks:
            wb_identity = parse_project_identity_from_filename(workbook.name)
            wb_line = normalize_line_name(wb_identity.get("line_name", ""))
            wb_line_key = ingest.normalize_space_only(wb_line)
            entry = workbook_lines.setdefault(
                wb_line_key,
                {
                    "line_name": wb_line,
                    "workbooks": [],
                },
            )
            entry["workbooks"].append(workbook)
        configured_stretch_line_keys: set[str] = set()
        has_wildcard_stretch_line = False

        if not cfg.stretch_entries:
            coverage_rows.append(
                {
                    "project_code": project_code_base,
                    "project_display": project_code_base,
                    "category": "stretch",
                    "status": "MISSING_SOURCE",
                    "reason_code": "MISSING_SOURCE",
                    "reason": "No dedicated stretch readiness sheet configured.",
                    "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                    "configured_sheet": "",
                    "resolved_sheet": "",
                    "rows": 0,
                    "available_sheets": "",
                }
            )
        else:
            for request in cfg.stretch_entries:
                configured_sheet = str(request.get("sheet_name", "")).strip()
                configured_line_name = normalize_line_name(request.get("line_name", ""))
                configured_line_source = str(request.get("line_name_source", "")).strip()
                configured_line_key = ingest.normalize_space_only(configured_line_name)
                if configured_line_key:
                    configured_stretch_line_keys.add(configured_line_key)
                else:
                    has_wildcard_stretch_line = True
                selected_workbooks = _pick_matching_workbooks_for_sheet_stretch(
                    project_workbooks,
                    workbook_sheet_cache,
                    configured_sheet,
                    configured_line_name,
                )

                if not selected_workbooks:
                    project_display = build_project_display(project_code_base, configured_line_name, project_code_base) or project_code_base
                    coverage_rows.append(
                        {
                            "project_code": project_code_base,
                            "project_display": project_display,
                            "category": "stretch",
                            "status": "MISSING_SOURCE",
                            "reason_code": "MISSING_SOURCE",
                            "reason": "Configured stretch readiness sheet not found in any workbook for project.",
                            "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                            "configured_sheet": configured_sheet,
                            "resolved_sheet": "",
                            "rows": 0,
                            "available_sheets": "",
                        }
                    )
                    issue_rows.append(
                        {
                            "Workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                            "Project": project_code_base,
                            "Category": "stretch",
                            "Sheet": "",
                            "ConfiguredSheet": configured_sheet,
                            "LineName": configured_line_name,
                            "LineNameSource": configured_line_source,
                            "Issue": "MISSING_SOURCE",
                            "Reason": "Configured stretch readiness sheet not found in project workbooks.",
                        }
                    )
                    continue

                for selected_workbook, resolved_sheet_guess in selected_workbooks:
                    selected_identity = parse_project_identity_from_filename(selected_workbook.name)
                    project_code = str(selected_identity.get("project_code", "")).strip() or selected_workbook.stem
                    line_name = configured_line_name or normalize_line_name(selected_identity.get("line_name", ""))
                    line_source = configured_line_source or ("config" if configured_line_name else "filename")
                    project_display = build_project_display(project_code, line_name, project_code) or project_code
                    project_scope_key = build_project_scope_key(project_code, line_name, project_display)
                    available_sheet_text = "; ".join(workbook_sheet_cache.get(str(selected_workbook.resolve()), []))

                    if template_error:
                        coverage_rows.append(
                            {
                                "project_code": project_code,
                                "project_display": project_display,
                                "category": "stretch",
                                "status": "TEMPLATE_CONFIG_ERROR",
                                "reason_code": "TEMPLATE_CONFIG_ERROR",
                                "reason": template_error,
                                "workbook": selected_workbook.name,
                                "configured_sheet": configured_sheet,
                                "resolved_sheet": resolved_sheet_guess,
                                "rows": 0,
                                "available_sheets": available_sheet_text,
                            }
                        )
                        issue_rows.append(
                            {
                                "Workbook": selected_workbook.name,
                                "Project": project_code,
                                "Category": "stretch",
                                "Sheet": resolved_sheet_guess,
                                "ConfiguredSheet": configured_sheet,
                                "LineName": line_name,
                                "LineNameSource": line_source,
                                "Issue": "TEMPLATE_CONFIG_ERROR",
                                "Reason": template_error,
                            }
                        )
                        continue

                    selector = _build_exact_sheet_selector_stretch(configured_sheet) if configured_sheet else (lambda names: names[0] if names else None)
                    try:
                        df_raw, resolved_sheet, fallback_note = load_sheet_with_csv_fallback(
                            selected_workbook,
                            selector,
                            read_excel_kwargs={"header": None},
                            read_csv_kwargs={"header": None},
                        )
                    except Exception as exc:
                        issue_rows.append(
                            {
                                "Workbook": selected_workbook.name,
                                "Project": project_code,
                                "Category": "stretch",
                                "Sheet": resolved_sheet_guess,
                                "ConfiguredSheet": configured_sheet,
                                "LineName": line_name,
                                "LineNameSource": line_source,
                                "Issue": "READ_FAIL",
                                "Reason": str(exc),
                            }
                        )
                        coverage_rows.append(
                            {
                                "project_code": project_code,
                                "project_display": project_display,
                                "category": "stretch",
                                "status": "READ_FAIL",
                                "reason_code": "READ_FAIL",
                                "reason": str(exc),
                                "workbook": selected_workbook.name,
                                "configured_sheet": configured_sheet,
                                "resolved_sheet": resolved_sheet_guess,
                                "rows": 0,
                                "available_sheets": available_sheet_text,
                            }
                        )
                        continue

                    if df_raw is None or resolved_sheet is None:
                        issue_rows.append(
                            {
                                "Workbook": selected_workbook.name,
                                "Project": project_code,
                                "Category": "stretch",
                                "Sheet": resolved_sheet_guess,
                                "ConfiguredSheet": configured_sheet,
                                "LineName": line_name,
                                "LineNameSource": line_source,
                                "Issue": "MISSING_SOURCE",
                                "Reason": "Configured stretch readiness sheet not found in selected workbook.",
                            }
                        )
                        coverage_rows.append(
                            {
                                "project_code": project_code,
                                "project_display": project_display,
                                "category": "stretch",
                                "status": "MISSING_SOURCE",
                                "reason_code": "MISSING_SOURCE",
                                "reason": "Configured stretch readiness sheet not found in selected workbook.",
                                "workbook": selected_workbook.name,
                                "configured_sheet": configured_sheet,
                                "resolved_sheet": "",
                                "rows": 0,
                                "available_sheets": available_sheet_text,
                            }
                        )
                        continue

                    selected_template = select_stretch_template_for_sheet(
                        stretch_template_catalog.get(project_key, []),
                        configured_sheet_name=configured_sheet,
                        resolved_sheet_name=resolved_sheet,
                        line_name=line_name,
                    )
                    if selected_template is None:
                        selected_template = select_stretch_template_for_sheet(
                            stretch_template_all_catalog.get(project_key, []),
                            configured_sheet_name=configured_sheet,
                            resolved_sheet_name=resolved_sheet,
                            line_name=line_name,
                        )
                    template_map = dict(selected_template.get("column_map", {}) if selected_template else {})
                    guardrails = dict(selected_template.get("guardrails", {}) if selected_template else {})
                    template_sheet = str(selected_template.get("template_sheet", "") if selected_template else "")

                    parse_result = _parse_stretch_sheet_dataframe(
                        df_raw,
                        guardrails=guardrails,
                        template_map=template_map,
                        project_code=project_code,
                        project_display=project_display,
                        project_scope_key=project_scope_key,
                        line_name=line_name,
                        line_name_source=line_source,
                        source_file=selected_workbook.name,
                        source_sheet=resolved_sheet,
                        configured_sheet=configured_sheet,
                        template_sheet=template_sheet,
                        report_date=_extract_report_date_from_filename(selected_workbook.name),
                    )
                    if not parse_result.data.empty:
                        raw_frames.append(parse_result.data)

                    diagnostics_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Category": "stretch",
                            "Sheet": resolved_sheet,
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "TemplateSheet": template_sheet,
                            "TemplateApplied": bool(template_map),
                            "TemplateChanges": "; ".join(parse_result.template_changes),
                            "FallbackNote": fallback_note or "",
                            "SectionsDetected": int(parse_result.sections_detected),
                            "HeadersDetected": int(parse_result.headers_detected),
                            "Rows": int(parse_result.rows_emitted),
                            "Status": parse_result.parse_status,
                            "Reason": parse_result.parse_reason,
                        }
                    )
                    if parse_result.parse_status != "OK":
                        issue_rows.append(
                            {
                                "Workbook": selected_workbook.name,
                                "Project": project_code,
                                "Category": "stretch",
                                "Sheet": resolved_sheet,
                                "ConfiguredSheet": configured_sheet,
                                "LineName": line_name,
                                "LineNameSource": line_source,
                                "Issue": parse_result.parse_status,
                                "Reason": parse_result.parse_reason,
                            }
                        )
                    coverage_rows.append(
                        {
                            "project_code": project_code,
                            "project_display": project_display,
                            "category": "stretch",
                            "status": parse_result.parse_status,
                            "reason_code": parse_result.parse_status,
                            "reason": parse_result.parse_reason,
                            "workbook": selected_workbook.name,
                            "configured_sheet": configured_sheet,
                            "resolved_sheet": resolved_sheet,
                            "rows": int(parse_result.rows_emitted),
                            "available_sheets": available_sheet_text,
                        }
                    )

            if not has_wildcard_stretch_line:
                for line_key, line_payload in workbook_lines.items():
                    if line_key in configured_stretch_line_keys:
                        continue
                    line_name = normalize_line_name(line_payload.get("line_name", ""))
                    project_display = build_project_display(project_code_base, line_name, project_code_base) or project_code_base
                    line_workbooks = line_payload.get("workbooks", []) or []
                    workbook_names = "; ".join(sorted({w.name for w in line_workbooks}))
                    coverage_rows.append(
                        {
                            "project_code": project_code_base,
                            "project_display": project_display,
                            "category": "stretch",
                            "status": "MISSING_SOURCE",
                            "reason_code": "MISSING_SOURCE",
                            "reason": "No dedicated stretch readiness sheet configured for this project line variant.",
                            "workbook": workbook_names,
                            "configured_sheet": "",
                            "resolved_sheet": "",
                            "rows": 0,
                            "available_sheets": "",
                        }
                    )
                    issue_rows.append(
                        {
                            "Workbook": workbook_names,
                            "Project": project_code_base,
                            "Category": "stretch",
                            "Sheet": "",
                            "ConfiguredSheet": "",
                            "LineName": line_name,
                            "LineNameSource": "filename",
                            "Issue": "MISSING_SOURCE",
                            "Reason": "No dedicated stretch readiness sheet configured for this project line variant.",
                        }
                    )

        if not cfg.daily_entries:
            manpower_rows.append(
                {
                    "project_code": project_code_base,
                    "project_display": project_code_base,
                    "project_scope_key": build_project_scope_key(project_code_base, "", project_code_base),
                    "line_name": "",
                    "line_name_source": "",
                    "source_file": "",
                    "source_sheet": "",
                    "configured_sheet": "",
                    "header_row_number": None,
                    "manpower_fields": "",
                    "readiness_fields": "",
                    "readiness_column_present": False,
                    "signal_type": "NO_SHEET_CONFIG",
                    "non_empty_count": 0,
                    "sample_values": "",
                    "expected_manpower": cfg.manpower_expected,
                    "expected_match": _expected_match_stretch(cfg.manpower_expected, "NO_SHEET_CONFIG"),
                    "status": "NO_SHEET_CONFIG",
                    "reason": "No daily stringing sheet configured for manpower audit.",
                }
            )
            coverage_rows.append(
                {
                    "project_code": project_code_base,
                    "project_display": project_code_base,
                    "category": "manpower",
                    "status": "NO_SHEET_CONFIG",
                    "reason_code": "NO_SHEET_CONFIG",
                    "reason": "No daily stringing sheet configured for manpower audit.",
                    "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                    "configured_sheet": "",
                    "resolved_sheet": "",
                    "rows": 0,
                    "available_sheets": "",
                }
            )
        else:
            for request in cfg.daily_entries:
                configured_sheet = str(request.get("sheet_name", "")).strip()
                configured_line_name = normalize_line_name(request.get("line_name", ""))
                configured_line_source = str(request.get("line_name_source", "")).strip()
                selected_workbook, _ = _pick_best_workbook_for_sheet_stretch(
                    project_workbooks,
                    workbook_sheet_cache,
                    configured_sheet,
                    configured_line_name,
                )
                project_display = build_project_display(project_code_base, configured_line_name, project_code_base) or project_code_base
                project_scope_key = build_project_scope_key(project_code_base, configured_line_name, project_display)
                if selected_workbook is None:
                    manpower_rows.append(
                        {
                            "project_code": project_code_base,
                            "project_display": project_display,
                            "project_scope_key": project_scope_key,
                            "line_name": configured_line_name,
                            "line_name_source": configured_line_source,
                            "source_file": "",
                            "source_sheet": "",
                            "configured_sheet": configured_sheet,
                            "header_row_number": None,
                            "manpower_fields": "",
                            "readiness_fields": "",
                            "readiness_column_present": False,
                            "signal_type": "SHEET_MISSING",
                            "non_empty_count": 0,
                            "sample_values": "",
                            "expected_manpower": cfg.manpower_expected,
                            "expected_match": _expected_match_stretch(cfg.manpower_expected, "SHEET_MISSING"),
                            "status": "SHEET_MISSING",
                            "reason": "Configured daily stringing sheet not found.",
                        }
                    )
                    coverage_rows.append(
                        {
                            "project_code": project_code_base,
                            "project_display": project_display,
                            "category": "manpower",
                            "status": "SHEET_MISSING",
                            "reason_code": "SHEET_MISSING",
                            "reason": "Configured daily stringing sheet not found.",
                            "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                            "configured_sheet": configured_sheet,
                            "resolved_sheet": "",
                            "rows": 0,
                            "available_sheets": "",
                        }
                    )
                    continue

                selected_identity = parse_project_identity_from_filename(selected_workbook.name)
                project_code = str(selected_identity.get("project_code", "")).strip() or selected_workbook.stem
                line_name = configured_line_name or normalize_line_name(selected_identity.get("line_name", ""))
                line_source = configured_line_source or ("config" if configured_line_name else "filename")
                project_display = build_project_display(project_code, line_name, project_code) or project_code
                project_scope_key = build_project_scope_key(project_code, line_name, project_display)
                try:
                    loaded = ingest.load_stringing_sheet_frame(selected_workbook, configured_sheet_name=configured_sheet)
                    frame = loaded.frame if loaded.frame is not None else pd.DataFrame()
                    signal = _detect_manpower_signal_from_frame(frame)
                    if str(signal.get("signal_type", "")).upper() == "ABSENT":
                        selector = _build_exact_sheet_selector_stretch(configured_sheet)
                        try:
                            raw_df, _, _ = load_sheet_with_csv_fallback(
                                selected_workbook,
                                selector,
                                read_excel_kwargs={"header": None},
                                read_csv_kwargs={"header": None},
                            )
                        except Exception:
                            raw_df = None
                        raw_signal = _detect_manpower_signal_from_raw_table(raw_df) if raw_df is not None else {}
                        if str(raw_signal.get("signal_type", "")).upper() != "ABSENT":
                            signal = raw_signal
                    status = str(signal.get("signal_type", "ABSENT"))
                    manpower_rows.append(
                        {
                            "project_code": project_code,
                            "project_display": project_display,
                            "project_scope_key": project_scope_key,
                            "line_name": line_name,
                            "line_name_source": line_source,
                            "source_file": selected_workbook.name,
                            "source_sheet": loaded.resolved_sheet or "",
                            "configured_sheet": configured_sheet,
                            "header_row_number": loaded.header_row,
                            "manpower_fields": "; ".join(signal.get("manpower_fields", [])),
                            "readiness_fields": "; ".join(signal.get("readiness_fields", [])),
                            "readiness_column_present": bool(signal.get("readiness_present", False)),
                            "signal_type": status,
                            "non_empty_count": int(signal.get("non_empty_count", 0) or 0),
                            "sample_values": "; ".join(str(v) for v in signal.get("sample_values", [])),
                            "expected_manpower": cfg.manpower_expected,
                            "expected_match": _expected_match_stretch(cfg.manpower_expected, status),
                            "status": status,
                            "reason": "",
                        }
                    )
                    diagnostics_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Category": "manpower",
                            "Sheet": loaded.resolved_sheet or "",
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "TemplateSheet": "",
                            "TemplateApplied": False,
                            "TemplateChanges": "",
                            "FallbackNote": loaded.fallback_note or "",
                            "SectionsDetected": 0,
                            "HeadersDetected": 1 if loaded.header_row is not None else 0,
                            "Rows": int(len(frame.index)),
                            "Status": status,
                            "Reason": "",
                        }
                    )
                    coverage_rows.append(
                        {
                            "project_code": project_code,
                            "project_display": project_display,
                            "category": "manpower",
                            "status": status,
                            "reason_code": status,
                            "reason": "",
                            "workbook": selected_workbook.name,
                            "configured_sheet": configured_sheet,
                            "resolved_sheet": loaded.resolved_sheet or "",
                            "rows": int(len(frame.index)),
                            "available_sheets": "; ".join(workbook_sheet_cache.get(str(selected_workbook.resolve()), [])),
                        }
                    )
                except Exception as exc:
                    manpower_rows.append(
                        {
                            "project_code": project_code,
                            "project_display": project_display,
                            "project_scope_key": project_scope_key,
                            "line_name": line_name,
                            "line_name_source": line_source,
                            "source_file": selected_workbook.name,
                            "source_sheet": "",
                            "configured_sheet": configured_sheet,
                            "header_row_number": None,
                            "manpower_fields": "",
                            "readiness_fields": "",
                            "readiness_column_present": False,
                            "signal_type": "READ_FAIL",
                            "non_empty_count": 0,
                            "sample_values": "",
                            "expected_manpower": cfg.manpower_expected,
                            "expected_match": _expected_match_stretch(cfg.manpower_expected, "READ_FAIL"),
                            "status": "READ_FAIL",
                            "reason": str(exc),
                        }
                    )
                    diagnostics_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Category": "manpower",
                            "Sheet": "",
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "TemplateSheet": "",
                            "TemplateApplied": False,
                            "TemplateChanges": "",
                            "FallbackNote": "",
                            "SectionsDetected": 0,
                            "HeadersDetected": 0,
                            "Rows": 0,
                            "Status": "READ_FAIL",
                            "Reason": str(exc),
                        }
                    )
                    issue_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Category": "manpower",
                            "Sheet": "",
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "Issue": "READ_FAIL",
                            "Reason": str(exc),
                        }
                    )
                    coverage_rows.append(
                        {
                            "project_code": project_code,
                            "project_display": project_display,
                            "category": "manpower",
                            "status": "READ_FAIL",
                            "reason_code": "READ_FAIL",
                            "reason": str(exc),
                            "workbook": selected_workbook.name,
                            "configured_sheet": configured_sheet,
                            "resolved_sheet": "",
                            "rows": 0,
                            "available_sheets": "; ".join(workbook_sheet_cache.get(str(selected_workbook.resolve()), [])),
                        }
                    )

    if skipped_not_in_config:
        print(f"[pipeline] StretchReadiness: skipped {skipped_not_in_config} workbook(s) not listed in DPR_Config.")

    raw_df = pd.concat(raw_frames, ignore_index=True) if raw_frames else pd.DataFrame(columns=STRETCH_RAWDATA_COLUMNS)
    if output_path.parent.name.lower() == "stretchreadiness" and output_path.parent.parent.name.lower() == "parquets":
        parquets_root = output_path.parent.parent
    elif repo_root is not None:
        parquets_root = Path(repo_root) / "Parquets"
    else:
        parquets_root = output_path.parent.parent

    derived_df = _build_derived_stretch_rows(
        parquets_root=parquets_root,
        allowed_project_keys=processed_project_keys,
    )
    if not derived_df.empty:
        raw_df = _merge_stretch_sources_prefer_derived(raw_df, derived_df)
        diagnostics_rows.append(
            {
                "Workbook": "StringingCompiled.parquet; RawData.parquet",
                "Project": "",
                "Category": "stretch_derived",
                "Sheet": "Stringing Compiled",
                "ConfiguredSheet": "",
                "LineName": "",
                "LineNameSource": "derived",
                "TemplateSheet": "",
                "TemplateApplied": False,
                "TemplateChanges": "",
                "FallbackNote": "",
                "SectionsDetected": int(derived_df.shape[0]),
                "HeadersDetected": 0,
                "Rows": int(derived_df.shape[0]),
                "Status": "OK",
                "Reason": "Derived readiness computed using endpoints + Location Nos against erection Tower Tightening dates.",
            }
        )
    else:
        diagnostics_rows.append(
            {
                "Workbook": "StringingCompiled.parquet; RawData.parquet",
                "Project": "",
                "Category": "stretch_derived",
                "Sheet": "",
                "ConfiguredSheet": "",
                "LineName": "",
                "LineNameSource": "derived",
                "TemplateSheet": "",
                "TemplateApplied": False,
                "TemplateChanges": "",
                "FallbackNote": "",
                "SectionsDetected": 0,
                "HeadersDetected": 0,
                "Rows": 0,
                "Status": "NO_DERIVED_INPUT",
                "Reason": "Derived readiness inputs unavailable; legacy stretch readiness (if present) retained.",
            }
        )

    raw_df = raw_df.reindex(columns=STRETCH_RAWDATA_COLUMNS)
    summary_df = _build_stretch_summary(raw_df)
    manpower_df = pd.DataFrame(manpower_rows, columns=STRETCH_MANPOWER_COLUMNS)
    diagnostics_df = pd.DataFrame(diagnostics_rows, columns=STRETCH_DIAGNOSTICS_COLUMNS)
    issues_df = pd.DataFrame(issue_rows, columns=STRETCH_ISSUES_COLUMNS)
    coverage_df = pd.DataFrame(coverage_rows, columns=STRETCH_COVERAGE_COLUMNS)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output_path.with_suffix(f"{output_path.suffix}.tmp")

    try:
        with pd.ExcelWriter(temp_output, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name="RawData", index=False)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            manpower_df.to_excel(writer, sheet_name="ManpowerAudit", index=False)
            diagnostics_df.to_excel(writer, sheet_name="Diagnostics", index=False)
            issues_df.to_excel(writer, sheet_name="Issues", index=False)
            coverage_df.to_excel(writer, sheet_name="Coverage", index=False)
        temp_output.replace(output_path)
    finally:
        if temp_output.exists():
            try:
                temp_output.unlink()
            except Exception:
                pass

    print(
        f"[pipeline] StretchReadiness: wrote workbook {output_path} "
        f"(rows={len(raw_df.index)}, summary={len(summary_df.index)}, manpower={len(manpower_df.index)}, "
        f"diagnostics={len(diagnostics_df.index)}, issues={len(issues_df.index)})."
    )
    return output_path
