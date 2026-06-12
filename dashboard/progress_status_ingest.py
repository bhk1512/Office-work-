"""Config-driven DPR progress/status ingestion."""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional
import re
import subprocess

import pandas as pd
from openpyxl import load_workbook

from erection_compiled_to_daily_new import load_sheet_with_csv_fallback
from . import stringing_ingest as ingest
from .project_identity import (
    build_project_display,
    build_project_scope_key,
    normalize_line_name,
    parse_project_identity_from_filename,
    parse_sheet_line_entries,
)
from .completed_projects import is_completed_project


DEFAULT_ACTIVITY_ALLOWLIST = (
    "route alignment",
    "detail survey",
    "detailed survey",
    "check survey",
    "soil investigation",
    "excavation",
    "foundation",
    "earthing",
    "tower erection",
    "erection",
    "tower tightening",
    "tack welding",
    "tackwelding",
    "nh crossing",
    "power line crossing",
    "stringing",
    "paying out",
    "final sag",
    "opgw",
)

DEFAULT_REQUIRED_HEADER_TOKENS = ("activity", "progress", "cumulative")

RAWDATA_COLUMNS = [
    "project_code",
    "project_display",
    "project_scope_key",
    "line_name",
    "line_name_source",
    "section_label",
    "report_date",
    "source_file",
    "source_sheet",
    "configured_sheet",
    "template_sheet",
    "stringing_resolution_policy",
    "header_row_number",
    "source_row_number",
    "activity_raw",
    "activity_norm",
    "quantity_loa",
    "quantity_estimated_or_total",
    "quantity_primary",
    "cumulative_last_month",
    "plan_for_month",
    "progress_for_month",
    "today_progress",
    "cumulative_progress",
    "balance_progress",
    "gangs_working",
    "remarks",
]

DIAGNOSTICS_COLUMNS = [
    "Workbook",
    "Project",
    "Sheet",
    "ConfiguredSheet",
    "LineName",
    "LineNameSource",
    "TemplateSheet",
    "TemplateApplied",
    "TemplateChanges",
    "FallbackNote",
    "SectionsDetected",
    "HeadersDetected",
    "Rows",
    "FilenameDate",
    "InternalDate",
    "DateQuality",
    "Status",
    "Reason",
]

ISSUES_COLUMNS = [
    "Workbook",
    "Project",
    "Sheet",
    "ConfiguredSheet",
    "LineName",
    "LineNameSource",
    "Issue",
    "Reason",
]

COVERAGE_COLUMNS = [
    "project_code",
    "project_display",
    "status",
    "reason_code",
    "reason",
    "workbook",
    "configured_sheet",
    "resolved_sheet",
    "rows",
    "filename_date",
    "internal_date",
    "date_quality",
    "available_sheets",
]


def _normalize_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"[.,;:()\\[\\]{}]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _split_tokens(value: object) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    tokens = [part.strip() for part in re.split(r"[;,|]+", text)]
    return [token for token in tokens if token]


def _coerce_numeric(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        cleaned = cleaned.replace(",", "")
        value = cleaned
    try:
        num = pd.to_numeric([value], errors="coerce")[0]
    except Exception:
        return None
    if pd.isna(num):
        return None
    return float(num)


def _as_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def _make_unique_headers(labels: list[str]) -> list[str]:
    unique: list[str] = []
    seen: dict[str, int] = {}
    for idx, label in enumerate(labels, start=1):
        base = label.strip() if label else f"unnamed_col_{idx}"
        key = ingest.normalize_space_only(base) or base.lower()
        count = seen.get(key, 0) + 1
        seen[key] = count
        unique.append(base if count == 1 else f"{base}__{count}")
    return unique


def _extract_numeric_tokens(value: object) -> set[str]:
    text = str(value or "")
    return {token for token in re.findall(r"\d{2,4}", text)}


def _normalize_activity(raw: object) -> str:
    text = _normalize_text(raw)
    if not text:
        return ""
    if "route alignment" in text:
        return "route_alignment"
    if "detailed survey" in text:
        return "detailed_survey"
    if "check survey" in text:
        return "check_survey"
    if "soil investigation" in text:
        return "soil_investigation"
    if "excavation" in text:
        return "excavation"
    if "foundation" in text:
        return "foundation"
    if "earthing" in text:
        return "earthing"
    if "tower erection" in text or text == "erection":
        return "tower_erection"
    if "tack welding" in text or "tackwelding" in text:
        return "tack_welding"
    if "paying out" in text:
        return "paying_out"
    if "final sag" in text:
        return "final_sag"
    if "opgw" in text:
        return "opgw_stringing"
    if "stringing" in text:
        return "stringing"
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _header_tokens_joined(labels: list[str]) -> str:
    return " ".join(_normalize_text(label) for label in labels if _normalize_text(label))


def _combine_header_labels(df: pd.DataFrame, start_row: int, header_rows: int, width: int) -> list[str]:
    labels: list[str] = []
    for col_idx in range(width):
        parts: list[str] = []
        for offset in range(header_rows):
            row_idx = start_row + offset
            if row_idx >= len(df.index):
                continue
            value = _as_text(df.iat[row_idx, col_idx] if col_idx < len(df.columns) else "")
            if value and value not in parts:
                parts.append(value)
        labels.append(" ".join(parts).strip())
    return _make_unique_headers(labels)


def _row_text(df: pd.DataFrame, row_idx: int, max_cols: int = 30) -> str:
    if row_idx < 0 or row_idx >= len(df.index):
        return ""
    values: list[str] = []
    width = min(len(df.columns), max_cols)
    for col_idx in range(width):
        values.append(_as_text(df.iat[row_idx, col_idx]))
    return " | ".join(values)


def _row_primary_text(df: pd.DataFrame, row_idx: int, max_cols: int = 8) -> str:
    if row_idx < 0 or row_idx >= len(df.index):
        return ""
    width = min(len(df.columns), max_cols)
    for col_idx in range(width):
        value = _as_text(df.iat[row_idx, col_idx])
        if value:
            return value
    return ""


def _looks_like_header_marker(text_norm: str, required_tokens: list[str]) -> bool:
    if not text_norm:
        return False
    if "activity" in text_norm and ("progress" in text_norm or "cumm" in text_norm or "cumulative" in text_norm):
        return True
    return bool(required_tokens) and all(token in text_norm for token in required_tokens)


def _infer_block_section_label(
    df: pd.DataFrame,
    *,
    header_row: int,
    section_start: int,
    base_label: str,
    required_tokens: list[str],
) -> str:
    explicit = re.sub(r"\s+", " ", str(base_label or "").strip())
    if explicit:
        return explicit

    for row_idx in range(header_row - 1, max(section_start, header_row - 8) - 1, -1):
        candidate = _row_primary_text(df, row_idx)
        candidate_norm = _normalize_text(candidate)
        if not candidate_norm:
            continue
        if _looks_like_header_marker(candidate_norm, required_tokens):
            continue
        if re.fullmatch(r"[0-9.\-_/() ]+", candidate_norm):
            continue
        return re.sub(r"\s+", " ", candidate).strip()

    return explicit


def _row_is_blank(values: list[object]) -> bool:
    for value in values:
        if _as_text(value):
            return False
    return True


def _parse_header_rows_guardrail(value: object) -> list[int]:
    tokens = _split_tokens(value)
    if not tokens:
        token_text = str(value or "").strip()
        if token_text:
            tokens = [token_text]
    parsed: list[int] = []
    for token in tokens:
        try:
            intval = int(str(token).strip())
        except Exception:
            continue
        if intval > 0:
            parsed.append(intval)
    deduped = sorted(set(parsed))
    return deduped if deduped else [1, 2, 3]


def _find_anchor_row(
    df: pd.DataFrame,
    start_row: int,
    end_row: int,
    *,
    anchor_text: str = "",
    anchor_regex: str = "",
) -> int:
    if anchor_regex:
        try:
            pattern = re.compile(anchor_regex, flags=re.IGNORECASE)
        except re.error:
            pattern = None
        if pattern is not None:
            for row_idx in range(start_row, end_row + 1):
                text = _row_text(df, row_idx, max_cols=25)
                if pattern.search(text):
                    return row_idx

    anchor = _normalize_text(anchor_text)
    if anchor:
        for row_idx in range(start_row, end_row + 1):
            text = _normalize_text(_row_text(df, row_idx, max_cols=25))
            if anchor in text:
                return row_idx
    return start_row


def _detect_header(
    df: pd.DataFrame,
    *,
    start_row: int,
    end_row: int,
    required_tokens: list[str],
    header_rows_options: list[int],
    max_scan_rows: int = 45,
) -> tuple[int | None, int | None, list[str]]:
    if df.empty:
        return None, None, []
    width = min(len(df.columns), 60)
    scan_end = min(end_row, start_row + max_scan_rows)
    for row_idx in range(start_row, scan_end + 1):
        for span in header_rows_options:
            if row_idx + span - 1 > end_row:
                continue
            labels = _combine_header_labels(df, row_idx, span, width)
            joined = _header_tokens_joined(labels)
            if not joined:
                continue
            if all(token in joined for token in required_tokens):
                return row_idx, span, labels
    return None, None, []


def _extract_section_ranges(
    df: pd.DataFrame,
    *,
    start_row: int,
    end_row: int,
    section_contains: str = "",
    section_regex: str = "",
    section_label_regex: str = "",
) -> list[tuple[str, int, int]]:
    contains = _normalize_text(section_contains)
    regex_obj = None
    label_regex_obj = None
    if section_regex:
        try:
            regex_obj = re.compile(section_regex, flags=re.IGNORECASE)
        except re.error:
            regex_obj = None
    if section_label_regex:
        try:
            label_regex_obj = re.compile(section_label_regex, flags=re.IGNORECASE)
        except re.error:
            label_regex_obj = None

    if not contains and regex_obj is None:
        return [("", start_row, end_row)]

    hits: list[tuple[str, int]] = []
    for row_idx in range(start_row, end_row + 1):
        text = _row_text(df, row_idx, max_cols=30)
        text_norm = _normalize_text(text)
        matched = False
        if contains and contains in text_norm:
            matched = True
        if not matched and regex_obj is not None and regex_obj.search(text):
            matched = True
        if not matched:
            continue
        label = text.strip()
        if label_regex_obj is not None:
            label_match = label_regex_obj.search(text)
            if label_match:
                label = label_match.group(1).strip() if label_match.groups() else label_match.group(0).strip()
        label = re.sub(r"\s+", " ", label).strip()
        hits.append((label, row_idx))

    if not hits:
        return [("", start_row, end_row)]

    ranges: list[tuple[str, int, int]] = []
    for idx, (label, section_row) in enumerate(hits):
        section_end = hits[idx + 1][1] - 1 if idx + 1 < len(hits) else end_row
        section_start = max(start_row, section_row)
        if section_start <= section_end:
            ranges.append((label, section_start, section_end))
    return ranges if ranges else [("", start_row, end_row)]


def _apply_template_column_mapping(df: pd.DataFrame, template_map: dict[int, str]) -> tuple[pd.DataFrame, list[str]]:
    if df is None or df.empty or not template_map:
        return df, []
    remapped = df.copy()
    columns = list(remapped.columns)
    changes: list[str] = []
    for idx, mapped_name in sorted(template_map.items()):
        if idx >= len(columns):
            continue
        current = str(columns[idx]).strip()
        target = str(mapped_name).strip()
        if not target:
            continue
        columns[idx] = target
        if ingest.normalize_space_only(current) != ingest.normalize_space_only(target):
            changes.append(f"C{idx + 1}:{current}->{target}")
    remapped.columns = columns
    return remapped, changes


def _resolve_activity_column(df: pd.DataFrame) -> str:
    for column in df.columns:
        key = _normalize_text(column)
        if "activity" in key:
            return str(column)
    return str(df.columns[0]) if len(df.columns) else "activity_raw"


def _first_non_null(*values: object) -> float | None:
    for value in values:
        num = _coerce_numeric(value)
        if num is not None:
            return num
    return None


def _build_status_rows(
    df: pd.DataFrame,
    *,
    project_code: str,
    project_display: str,
    project_scope_key: str,
    line_name: str,
    line_name_source: str,
    section_label: str,
    report_date: str,
    source_file: str,
    source_sheet: str,
    configured_sheet: str,
    template_sheet: str,
    stringing_resolution_policy: str,
    header_row_number: int,
    activity_allowlist: list[str],
    activity_exclude: list[str],
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=RAWDATA_COLUMNS)

    rows: list[dict[str, object]] = []
    activity_col = "activity_raw" if "activity_raw" in df.columns else _resolve_activity_column(df)
    allowlist_norm = [_normalize_text(token) for token in activity_allowlist if _normalize_text(token)]
    exclude_norm = [_normalize_text(token) for token in activity_exclude if _normalize_text(token)]

    for _, row in df.iterrows():
        raw_activity = _as_text(row.get(activity_col, ""))
        activity_norm_text = _normalize_text(raw_activity)
        if not raw_activity:
            continue
        if allowlist_norm and not any(token in activity_norm_text for token in allowlist_norm):
            continue
        if exclude_norm and any(token in activity_norm_text for token in exclude_norm):
            continue

        quantity_loa = _coerce_numeric(row.get("quantity_loa"))
        quantity_estimated = _coerce_numeric(row.get("quantity_estimated_or_total"))
        cumulative_last_month = _first_non_null(
            row.get("cumulative_last_month"),
            row.get("cumulative_progress_last_month"),
        )
        plan_for_month = _coerce_numeric(row.get("plan_for_month"))
        progress_for_month = _coerce_numeric(row.get("progress_for_month"))
        today_progress = _coerce_numeric(row.get("today_progress"))
        cumulative_progress = _coerce_numeric(row.get("cumulative_progress"))
        balance_progress = _coerce_numeric(row.get("balance_progress"))

        rows.append(
            {
                "project_code": project_code,
                "project_display": project_display,
                "project_scope_key": project_scope_key,
                "line_name": line_name,
                "line_name_source": line_name_source,
                "section_label": section_label,
                "report_date": report_date,
                "source_file": source_file,
                "source_sheet": source_sheet,
                "configured_sheet": configured_sheet,
                "template_sheet": template_sheet,
                "stringing_resolution_policy": stringing_resolution_policy,
                "header_row_number": int(header_row_number),
                "source_row_number": int(row.get("__source_row_number", 0) or 0),
                "activity_raw": raw_activity,
                "activity_norm": _normalize_activity(raw_activity),
                "quantity_loa": quantity_loa,
                "quantity_estimated_or_total": quantity_estimated,
                "quantity_primary": quantity_estimated if quantity_estimated is not None else quantity_loa,
                "cumulative_last_month": cumulative_last_month,
                "plan_for_month": plan_for_month,
                "progress_for_month": progress_for_month,
                "today_progress": today_progress,
                "cumulative_progress": cumulative_progress,
                "balance_progress": balance_progress,
                "gangs_working": _as_text(row.get("gangs_working")),
                "remarks": _as_text(row.get("remarks")),
            }
        )
    return pd.DataFrame(rows, columns=RAWDATA_COLUMNS)


def _build_vertical_progress_summary_rows(
    df_raw: pd.DataFrame,
    *,
    project_code: str,
    project_display: str,
    project_scope_key: str,
    line_name: str,
    line_name_source: str,
    report_date: str,
    source_file: str,
    source_sheet: str,
    configured_sheet: str,
    template_sheet: str,
    stringing_resolution_policy: str,
    activity_allowlist: list[str],
    activity_exclude: list[str],
) -> pd.DataFrame:
    """Parse DPR sheets where each activity is a separate vertical summary block."""
    rows: list[dict[str, object]] = []
    allowlist_norm = [_normalize_text(token) for token in activity_allowlist if _normalize_text(token)]
    exclude_norm = [_normalize_text(token) for token in activity_exclude if _normalize_text(token)]

    for row_idx in range(max(len(df_raw.index) - 2, 0)):
        heading = _row_text(df_raw, row_idx, max_cols=12)
        match = re.search(r"(.+?)\s+progress\s+summary", heading, flags=re.IGNORECASE)
        if not match:
            continue

        raw_activity = re.sub(r"\s+", " ", match.group(1)).strip(" |:-")
        activity_text = _normalize_text(raw_activity)
        if not raw_activity:
            continue
        if allowlist_norm and not any(token in activity_text for token in allowlist_norm):
            continue
        if exclude_norm and any(token in activity_text for token in exclude_norm):
            continue

        header_idx = row_idx + 1
        value_idx = row_idx + 2
        headers = [
            _normalize_text(df_raw.iat[header_idx, col_idx])
            for col_idx in range(len(df_raw.columns))
        ]
        values = [
            df_raw.iat[value_idx, col_idx]
            for col_idx in range(len(df_raw.columns))
        ]

        def value_for(*required_tokens: str) -> float | None:
            for col_idx, header in enumerate(headers):
                if header and all(token in header for token in required_tokens):
                    return _coerce_numeric(values[col_idx])
            return None

        quantity = value_for("total", "qty")
        previous = value_for("completed", "prev", "month")
        current_month = value_for("current", "month")
        cumulative = value_for("cumulative")
        balance = value_for("bal")
        if all(value is None for value in (quantity, previous, current_month, cumulative, balance)):
            continue

        rows.append(
            {
                "project_code": project_code,
                "project_display": project_display,
                "project_scope_key": project_scope_key,
                "line_name": line_name,
                "line_name_source": line_name_source,
                "section_label": raw_activity,
                "report_date": report_date,
                "source_file": source_file,
                "source_sheet": source_sheet,
                "configured_sheet": configured_sheet,
                "template_sheet": template_sheet,
                "stringing_resolution_policy": stringing_resolution_policy,
                "header_row_number": header_idx + 1,
                "source_row_number": value_idx + 1,
                "activity_raw": raw_activity,
                "activity_norm": _normalize_activity(raw_activity),
                "quantity_loa": None,
                "quantity_estimated_or_total": quantity,
                "quantity_primary": quantity,
                "cumulative_last_month": previous,
                "plan_for_month": None,
                "progress_for_month": current_month,
                "today_progress": None,
                "cumulative_progress": cumulative,
                "balance_progress": balance,
                "gangs_working": "",
                "remarks": "",
            }
        )

    return pd.DataFrame(rows, columns=RAWDATA_COLUMNS)


def _sheet_tokens(value: object) -> set[str]:
    text = str(value or "").strip().lower()
    if not text:
        return set()
    return {token for token in re.findall(r"[a-z0-9]+", text) if token}


def _extract_report_date_from_filename(file_name: str) -> str:
    match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", str(file_name or ""))
    return match.group(1) if match else ""


def _extract_date_token(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.normalize().strftime("%Y-%m-%d")
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return ""
    iso_match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", text)
    if iso_match:
        return iso_match.group(1)
    dm_match = re.search(r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b", text)
    if dm_match:
        parsed = pd.to_datetime(dm_match.group(1), errors="coerce", dayfirst=True)
        if pd.notna(parsed):
            return pd.Timestamp(parsed).normalize().strftime("%Y-%m-%d")
    return ""


def _extract_internal_report_date(df_raw: pd.DataFrame) -> str:
    if df_raw is None or df_raw.empty:
        return ""
    max_rows = min(len(df_raw.index), 20)
    max_cols = min(len(df_raw.columns), 40)
    disallowed = ("contractual", "revised", "completion", "start")

    for row_idx in range(max_rows):
        for col_idx in range(max_cols):
            raw_value = df_raw.iat[row_idx, col_idx]
            label = _normalize_text(raw_value)
            if "date" not in label:
                continue
            if any(token in label for token in disallowed):
                continue
            if not re.search(r"\bdate\b", label):
                continue

            inline = _extract_date_token(raw_value)
            if inline:
                return inline

            for offset in range(1, 6):
                probe_idx = col_idx + offset
                if probe_idx >= max_cols:
                    break
                probe = df_raw.iat[row_idx, probe_idx]
                candidate = _extract_date_token(probe)
                if candidate:
                    return candidate
    return ""


def _date_quality_flag(filename_date: str, internal_date: str) -> str:
    if not filename_date:
        return "UNREADABLE"
    if not internal_date:
        return "MISSING_INTERNAL"
    if filename_date == internal_date:
        return "MATCH"
    return "MISMATCH"


def _sheet_match_score(configured_sheet: str, workbook_sheet: str) -> tuple[int, int, int]:
    configured = str(configured_sheet or "").strip()
    workbook = str(workbook_sheet or "").strip()
    if not configured or not workbook:
        return (0, 0, 0)
    if ingest.normalize_space_only(configured) == ingest.normalize_space_only(workbook):
        return (1, 0, 0)
    if ingest.normalize_sheet_key(configured) == ingest.normalize_sheet_key(workbook):
        return (0, 1, 0)
    c_tokens = _sheet_tokens(configured)
    w_tokens = _sheet_tokens(workbook)
    overlap = len(c_tokens.intersection(w_tokens)) if c_tokens and w_tokens else 0
    return (0, 0, overlap)


def _pick_best_workbook_for_sheet(
    workbooks: list[Path],
    workbook_sheets: dict[str, list[str]],
    configured_sheet: str,
) -> tuple[Path | None, str]:
    scored: list[tuple[tuple[int, int, int, float], Path, str]] = []
    for workbook in workbooks:
        names = workbook_sheets.get(str(workbook.resolve()), [])
        best_score = (0, 0, 0)
        best_name = ""
        for name in names:
            score = _sheet_match_score(configured_sheet, name)
            if score > best_score:
                best_score = score
                best_name = name
        mtime = workbook.stat().st_mtime if workbook.exists() else 0.0
        scored.append(((best_score[0], best_score[1], best_score[2], mtime), workbook, best_name))

    if not scored:
        return None, ""
    scored.sort(key=lambda item: item[0], reverse=True)
    top = scored[0]
    if top[0][:3] == (0, 0, 0):
        return None, ""
    return top[1], top[2]


def _pick_matching_workbooks_for_sheet(
    workbooks: list[Path],
    workbook_sheets: dict[str, list[str]],
    configured_sheet: str,
    configured_line_name: str = "",
    file_identifier: str = "",
) -> list[tuple[Path, str]]:
    scored: list[tuple[tuple[int, int, int, int, float], Path, str, str]] = []
    line_hint = ingest.normalize_space_only(configured_line_name)
    file_id_hint = str(file_identifier or "").strip()
    for workbook in workbooks:
        if file_id_hint and file_id_hint.lower() not in workbook.name.lower():
            continue
        names = workbook_sheets.get(str(workbook.resolve()), [])
        if configured_sheet:
            best_score = (0, 0, 0)
            best_name = ""
            for name in names:
                score = _sheet_match_score(configured_sheet, name)
                if score > best_score:
                    best_score = score
                    best_name = name
            if best_score == (0, 0, 0):
                continue
        else:
            best_score = (1, 0, 0)
            best_name = names[0] if names else ""
        identity = parse_project_identity_from_filename(workbook.name)
        wb_line = ingest.normalize_space_only(identity.get("line_name", ""))
        line_bonus = 1 if line_hint and wb_line and line_hint == wb_line else 0
        mtime = workbook.stat().st_mtime if workbook.exists() else 0.0
        report_date = _extract_report_date_from_filename(workbook.name)
        scored.append(((best_score[0], best_score[1], best_score[2], line_bonus, mtime), workbook, best_name, report_date))

    if not scored:
        return []

    # Choose a single best workbook per configured sheet to avoid cross-file
    # false NO_TARGET_SHEET noise (e.g. TB 507 MAIN/765 variants).
    best = max(
        scored,
        key=lambda item: (
            item[0][0],  # exact/contains overlap score part 1
            item[0][1],  # overlap part 2
            item[0][2],  # token overlap part 3
            item[0][3],  # line hint bonus
            item[0][4],  # mtime
        ),
    )
    return [(best[1], best[2])]


def _resolve_named_template_sheet(wb, expected_name: str) -> Optional[str]:
    expected_key = ingest.normalize_space_only(expected_name)
    for name in wb.sheetnames:
        if ingest.normalize_space_only(name) == expected_key:
            return name
    return None


def _resolve_project_template_sheets(wb, project_name: object, discipline: str) -> list[str]:
    project_text = str(project_name or "").strip()
    if not project_text:
        return []

    resolved: list[str] = []
    seen: set[str] = set()
    for expected in (
        f"{project_text} {discipline}",
        f"{project_text} {discipline} Template Check",
    ):
        hit = _resolve_named_template_sheet(wb, expected)
        if hit and hit not in seen:
            seen.add(hit)
            resolved.append(hit)

    project_key = ingest.normalize_space_only(project_text)
    discipline_key = ingest.normalize_space_only(discipline)
    for name in wb.sheetnames:
        key = ingest.normalize_space_only(name)
        if key and key.startswith(project_key) and key.endswith(discipline_key) and name not in seen:
            seen.add(name)
            resolved.append(name)
    return resolved


def _extract_template_column_map(ws) -> dict[int, str]:
    to_map_row = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if ingest.normalize_space_only(cell) == "to map":
                to_map_row = row_idx
                break
        if to_map_row is not None:
            break
    if to_map_row is None:
        return {}

    labels_row = to_map_row + 1
    row_values = next(ws.iter_rows(min_row=labels_row, max_row=labels_row, values_only=True), ())
    mapping: dict[int, str] = {}
    for col_idx, value in enumerate(row_values):
        label = str(value).strip() if value is not None else ""
        if label:
            mapping[col_idx] = label
    return mapping


def _extract_guardrails(ws) -> dict[str, str]:
    guardrails: dict[str, str] = {}
    in_guardrails = False
    for row in ws.iter_rows(values_only=True):
        first_raw = row[0] if len(row) > 0 else None
        second_raw = row[1] if len(row) > 1 else None
        first_norm = ingest.normalize_space_only(first_raw)
        if first_norm == "guardrails":
            in_guardrails = True
            continue
        if first_norm == "to map":
            break
        if not in_guardrails:
            continue
        key = _normalize_key(first_raw)
        value = _as_text(second_raw)
        if key and value:
            guardrails[key] = value
    return guardrails


def load_status_sheet_config(raw_root: Path, *, repo_root: Path | None = None) -> dict[str, list[dict[str, str]]]:
    config_path = ingest.resolve_dpr_config_path(raw_root, repo_root=repo_root)
    if config_path is None:
        return {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception:
        return {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}
        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}
        headers = [ingest.normalize_space_only(v) for v in header_row]
        if "project code" not in headers or "status sheet names" not in headers:
            return {}
        project_idx = headers.index("project code")
        status_idx = headers.index("status sheet names")
        line_idx = headers.index("status line names") if "status line names" in headers else None
        file_id_idx = headers.index("status file identifier") if "status file identifier" in headers else None

        mapping: dict[str, list[dict[str, str]]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project_val = row[project_idx] if project_idx < len(row) else None
            if project_val in (None, ""):
                continue
            project_key = ingest.normalize_project_code_key(project_val)
            raw_status = row[status_idx] if status_idx < len(row) else None
            if raw_status in (None, ""):
                mapping[project_key] = []
                continue

            raw_line_names = row[line_idx] if line_idx is not None and line_idx < len(row) else None
            raw_file_ids = row[file_id_idx] if file_id_idx is not None and file_id_idx < len(row) else None
            entries = parse_sheet_line_entries(
                raw_status,
                raw_line_names,
                "status",
                infer_from_sheet_name=False,
            )
            if raw_file_ids:
                file_id_parts = _split_tokens(raw_file_ids)
                for i, entry in enumerate(entries):
                    entry["file_identifier"] = file_id_parts[i] if i < len(file_id_parts) else ""
            deduped_entries: list[dict[str, str]] = []
            seen_sheet_keys: set[str] = set()
            for entry in entries:
                key = ingest.normalize_space_only(entry.get("sheet_name"))
                if not key or key in seen_sheet_keys:
                    continue
                seen_sheet_keys.add(key)
                deduped_entries.append(entry)
            mapping[project_key] = deduped_entries
        return mapping
    finally:
        wb.close()


def load_status_template_mapping_catalog(
    raw_root: Path,
    *,
    repo_root: Path | None = None,
    include_unchecked: bool = False,
) -> tuple[dict[str, list[dict[str, object]]], dict[str, str]]:
    config_path = ingest.resolve_dpr_config_path(raw_root, repo_root=repo_root)
    if config_path is None:
        return {}, {}

    try:
        wb = load_workbook(config_path, data_only=True, read_only=True)
    except Exception:
        return {}, {}

    try:
        if "Sheet Names Check" not in wb.sheetnames:
            return {}, {}
        ws = wb["Sheet Names Check"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {}, {}
        headers = [ingest.normalize_space_only(value) for value in header_row]
        project_idx = headers.index("project code") if "project code" in headers else None
        check_idx = None
        for candidate in ("status template check", "status"):
            if candidate in headers:
                check_idx = headers.index(candidate)
                break
        if project_idx is None or (check_idx is None and not include_unchecked):
            return {}, {}

        catalog: dict[str, list[dict[str, object]]] = {}
        errors: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            project_val = row[project_idx] if project_idx < len(row) else None
            if project_val in (None, ""):
                continue
            check_val = row[check_idx] if check_idx is not None and check_idx < len(row) else None
            check_enabled = ingest.normalize_space_only(check_val) == "yes"
            if not check_enabled and not include_unchecked:
                continue

            project_key = ingest.normalize_project_code_key(project_val)
            template_sheets = _resolve_project_template_sheets(wb, project_val, "Status")
            if not template_sheets:
                if check_enabled:
                    errors[project_key] = (
                        f"Status Template Check is Yes but no mapping tab matching project '{str(project_val).strip()}' was found."
                    )
                continue

            options: list[dict[str, object]] = []
            for sheet_name in template_sheets:
                ws_template = wb[sheet_name]
                col_map = _extract_template_column_map(ws_template)
                guardrails = _extract_guardrails(ws_template)
                if not col_map and not guardrails:
                    continue
                options.append(
                    {
                        "column_map": col_map,
                        "template_sheet": sheet_name,
                        "guardrails": guardrails,
                        "numeric_tokens": _extract_numeric_tokens(sheet_name),
                    }
                )
            if not options:
                if check_enabled:
                    errors[project_key] = (
                        f"Status template tab(s) for project '{str(project_val).strip()}' have no usable Guardrails/To Map rows."
                    )
                continue
            catalog[project_key] = options
        return catalog, errors
    finally:
        wb.close()


def select_status_template_for_sheet(
    template_options: list[dict[str, object]] | None,
    *,
    configured_sheet_name: str = "",
    resolved_sheet_name: str = "",
    line_name: str = "",
) -> dict[str, object] | None:
    if not template_options:
        return None
    hints = [configured_sheet_name or "", resolved_sheet_name or "", line_name or ""]
    hint_keys = [ingest.normalize_space_only(value) for value in hints if ingest.normalize_space_only(value)]
    hint_numbers: set[str] = set()
    for value in hints:
        hint_numbers.update(_extract_numeric_tokens(value))

    best: dict[str, object] | None = None
    best_score = float("-inf")
    for idx, option in enumerate(template_options):
        sheet_name = str(option.get("template_sheet", "")).strip()
        sheet_key = ingest.normalize_space_only(sheet_name)
        sheet_numbers = set(option.get("numeric_tokens", set()) or set())
        col_map = option.get("column_map", {}) or {}
        score = 0.0
        for hint in hint_keys:
            if sheet_key == hint:
                score += 1000.0
            elif hint and (hint in sheet_key or sheet_key in hint):
                score += 200.0
        if hint_numbers:
            score += float(len(sheet_numbers.intersection(hint_numbers))) * 120.0
        score += float(len(col_map))
        score -= idx * 1e-4
        if score > best_score:
            best_score = score
            best = option
    return best


def _build_exact_sheet_selector(sheet_name: str):
    expected_space = ingest.normalize_space_only(sheet_name)
    expected_compact = ingest.normalize_sheet_key(sheet_name)

    def _selector(sheet_names: list[str]) -> str | None:
        by_space: dict[str, str] = {}
        by_compact: dict[str, str] = {}
        for existing in sheet_names:
            space_key = ingest.normalize_space_only(existing)
            compact_key = ingest.normalize_sheet_key(existing)
            if space_key and space_key not in by_space:
                by_space[space_key] = existing
            if compact_key and compact_key not in by_compact:
                by_compact[compact_key] = existing
        hit = by_space.get(expected_space)
        if hit:
            return hit
        hit = by_compact.get(expected_compact)
        if hit:
            return hit
        for existing in sheet_names:
            compact = ingest.normalize_sheet_key(existing)
            if expected_compact and (expected_compact in compact or compact in expected_compact):
                return existing
        return None

    return _selector


@dataclass(frozen=True)
class StatusParseResult:
    data: pd.DataFrame
    parse_status: str
    parse_reason: str
    sections_detected: int
    headers_detected: int
    rows_emitted: int
    template_changes: list[str]


def _parse_status_sheet_dataframe(
    df_raw: pd.DataFrame,
    *,
    guardrails: dict[str, str],
    template_map: dict[int, str],
    project_code: str,
    project_display: str,
    project_scope_key: str,
    line_name: str,
    line_name_source: str,
    report_date: str,
    source_file: str,
    source_sheet: str,
    configured_sheet: str,
    template_sheet: str,
    stringing_resolution_policy: str,
) -> StatusParseResult:
    if df_raw is None or df_raw.empty:
        return StatusParseResult(
            data=pd.DataFrame(columns=RAWDATA_COLUMNS),
            parse_status="EMPTY_SHEET",
            parse_reason="Sheet is empty.",
            sections_detected=0,
            headers_detected=0,
            rows_emitted=0,
            template_changes=[],
        )

    activity_allowlist = _split_tokens(guardrails.get("activity_allowlist", "")) or list(DEFAULT_ACTIVITY_ALLOWLIST)
    activity_exclude = _split_tokens(guardrails.get("activity_exclude", ""))
    required_tokens = [_normalize_text(token) for token in _split_tokens(guardrails.get("required_tokens", ""))]
    required_tokens = [token for token in required_tokens if token] or list(DEFAULT_REQUIRED_HEADER_TOKENS)
    stop_tokens = [_normalize_text(token) for token in _split_tokens(guardrails.get("stop_tokens", ""))]
    header_rows_options = _parse_header_rows_guardrail(guardrails.get("header_rows", ""))
    section_ranges = _extract_section_ranges(
        df_raw,
        start_row=0,
        end_row=len(df_raw.index) - 1,
        section_contains=guardrails.get("section_start_contains", ""),
        section_regex=guardrails.get("section_split_regex", ""),
        section_label_regex=guardrails.get("section_label_regex", ""),
    )

    all_rows: list[pd.DataFrame] = []
    headers_detected = 0
    sections_detected = 0
    template_changes_all: list[str] = []
    for section_label, section_start, section_end in section_ranges:
        cursor = section_start
        while cursor <= section_end:
            anchor_row = _find_anchor_row(
                df_raw,
                cursor,
                section_end,
                anchor_text=guardrails.get("block_anchor", ""),
                anchor_regex=guardrails.get("block_anchor_regex", ""),
            )
            header_row, header_span, header_labels = _detect_header(
                df_raw,
                start_row=anchor_row,
                end_row=section_end,
                required_tokens=required_tokens,
                header_rows_options=header_rows_options,
            )
            if header_row is None or header_span is None:
                break

            headers_detected += 1
            width = len(header_labels)
            data_rows: list[tuple[int, list[object]]] = []
            blank_run = 0
            started = False
            hard_stop = False
            last_row_seen = header_row + header_span - 1
            for row_idx in range(header_row + header_span, section_end + 1):
                row_values = [df_raw.iat[row_idx, col_idx] if col_idx < len(df_raw.columns) else None for col_idx in range(width)]
                row_text_norm = _normalize_text(" ".join(_as_text(v) for v in row_values[: min(25, width)]))
                if stop_tokens and any(token in row_text_norm for token in stop_tokens):
                    if started:
                        hard_stop = True
                        break
                    continue
                if _row_is_blank(row_values):
                    if started:
                        blank_run += 1
                        if blank_run >= 2:
                            break
                    continue
                started = True
                blank_run = 0
                data_rows.append((row_idx + 1, row_values))
                last_row_seen = row_idx

            if data_rows:
                sections_detected += 1
                block_df = pd.DataFrame([values for _, values in data_rows], columns=header_labels)
                block_df["__source_row_number"] = [row_number for row_number, _ in data_rows]
                block_df, template_changes = _apply_template_column_mapping(block_df, template_map)
                template_changes_all.extend(template_changes)
                block_section_label = _infer_block_section_label(
                    df_raw,
                    header_row=header_row,
                    section_start=cursor,
                    base_label=section_label,
                    required_tokens=required_tokens,
                )
                prepared = _build_status_rows(
                    block_df,
                    project_code=project_code,
                    project_display=project_display,
                    project_scope_key=project_scope_key,
                    line_name=line_name,
                    line_name_source=line_name_source,
                    section_label=block_section_label,
                    report_date=report_date,
                    source_file=source_file,
                    source_sheet=source_sheet,
                    configured_sheet=configured_sheet,
                    template_sheet=template_sheet,
                    stringing_resolution_policy=stringing_resolution_policy,
                    header_row_number=header_row + 1,
                    activity_allowlist=activity_allowlist,
                    activity_exclude=activity_exclude,
                )
                if not prepared.empty:
                    all_rows.append(prepared)

            cursor = max(last_row_seen + 1, header_row + header_span + 1)
            if hard_stop:
                break

    if not all_rows:
        vertical_rows = _build_vertical_progress_summary_rows(
            df_raw,
            project_code=project_code,
            project_display=project_display,
            project_scope_key=project_scope_key,
            line_name=line_name,
            line_name_source=line_name_source,
            report_date=report_date,
            source_file=source_file,
            source_sheet=source_sheet,
            configured_sheet=configured_sheet,
            template_sheet=template_sheet,
            stringing_resolution_policy=stringing_resolution_policy,
            activity_allowlist=activity_allowlist,
            activity_exclude=activity_exclude,
        )
        if not vertical_rows.empty:
            return StatusParseResult(
                data=vertical_rows.reindex(columns=RAWDATA_COLUMNS),
                parse_status="OK",
                parse_reason="",
                sections_detected=int(len(vertical_rows.index)),
                headers_detected=int(len(vertical_rows.index)),
                rows_emitted=int(len(vertical_rows.index)),
                template_changes=template_changes_all,
            )

        reason = "No matching rows after guardrails/activity filters."
        status = "NO_MATCHED_ROWS" if headers_detected > 0 else "HEADER_NOT_FOUND"
        return StatusParseResult(
            data=pd.DataFrame(columns=RAWDATA_COLUMNS),
            parse_status=status,
            parse_reason=reason,
            sections_detected=sections_detected,
            headers_detected=headers_detected,
            rows_emitted=0,
            template_changes=template_changes_all,
        )

    combined = pd.concat(all_rows, ignore_index=True)
    return StatusParseResult(
        data=combined.reindex(columns=RAWDATA_COLUMNS),
        parse_status="OK",
        parse_reason="",
        sections_detected=sections_detected,
        headers_detected=headers_detected,
        rows_emitted=int(len(combined.index)),
        template_changes=template_changes_all,
    )


def _merge_status_history(current: pd.DataFrame, output_path: Path) -> pd.DataFrame:
    previous_frames: list[pd.DataFrame] = []
    if output_path.exists():
        try:
            previous_frames.append(pd.read_excel(output_path, sheet_name="RawData"))
        except Exception:
            pass

    try:
        repo_root_text = subprocess.check_output(
            ["git", "rev-parse", "--show-toplevel"],
            cwd=output_path.parent,
            text=True,
            stderr=subprocess.DEVNULL,
        ).strip()
        repo_root = Path(repo_root_text)
        relative_path = output_path.resolve().relative_to(repo_root.resolve()).as_posix()
        committed_bytes = subprocess.check_output(
            ["git", "show", f"HEAD:{relative_path}"],
            cwd=repo_root,
            stderr=subprocess.DEVNULL,
        )
        previous_frames.append(pd.read_excel(io.BytesIO(committed_bytes), sheet_name="RawData"))
    except Exception:
        pass

    previous_frames = [frame for frame in previous_frames if not frame.empty]
    if not previous_frames:
        return current.reindex(columns=RAWDATA_COLUMNS)
    previous = pd.concat(previous_frames, ignore_index=True)

    for column in RAWDATA_COLUMNS:
        if column not in previous.columns:
            previous[column] = pd.NA
        if column not in current.columns:
            current[column] = pd.NA

    previous = previous.reindex(columns=RAWDATA_COLUMNS).copy()
    current = current.reindex(columns=RAWDATA_COLUMNS).copy()
    previous["_current_snapshot"] = 0
    current["_current_snapshot"] = 1
    combined = pd.concat([previous, current], ignore_index=True)
    combined["report_date"] = pd.to_datetime(combined["report_date"], errors="coerce").dt.normalize()

    key_columns = [
        "project_scope_key",
        "report_date",
        "source_sheet",
        "configured_sheet",
        "section_label",
        "activity_norm",
        "source_row_number",
    ]
    combined = combined.sort_values("_current_snapshot").drop_duplicates(key_columns, keep="last")
    return combined.drop(columns="_current_snapshot").reindex(columns=RAWDATA_COLUMNS).reset_index(drop=True)


def _status_candidates(input_dir: Optional[Path], files: Optional[list[Path]]) -> list[Path]:
    if files:
        return [
            path
            for path in files
            if path.suffix.lower() in (".xlsx", ".xlsm", ".xls")
            and path.exists()
            and not path.name.startswith("~$")
        ]
    if input_dir and input_dir.exists():
        return sorted(
            [
                path
                for path in input_dir.rglob("*.xls*")
                if path.is_file() and not path.name.startswith("~$")
            ]
        )
    return []


def compile_progress_status_to_workbook(
    input_dir: Optional[Path],
    files: Optional[list[Path]],
    output_path: Path,
    *,
    repo_root: Path | None = None,
    completed_project_keys: set[str] | None = None,
) -> Path | None:
    candidates = _status_candidates(input_dir, files)
    if completed_project_keys:
        kept: list[Path] = []
        skipped_files = 0
        skipped_projects: set[str] = set()
        for workbook in candidates:
            identity = parse_project_identity_from_filename(workbook.name)
            project_code = str(identity.get("project_code", "")).strip() or workbook.stem
            if project_code and is_completed_project(project_code, completed_project_keys):
                skipped_files += 1
                skipped_projects.add(project_code.upper())
                continue
            kept.append(workbook)
        if skipped_files:
            print(
                f"[pipeline] ProgressStatus: skipped_completed_files={skipped_files}, "
                f"skipped_completed_projects={len(skipped_projects)}"
            )
        candidates = kept

    if not candidates:
        print("[pipeline] ProgressStatus: no candidate files found; skipping.")
        return None

    if input_dir is not None:
        raw_root = input_dir
    elif files:
        raw_root = files[0].parent
    else:
        raw_root = Path(".")

    status_sheet_config = load_status_sheet_config(raw_root, repo_root=repo_root)
    status_template_catalog, status_template_errors = load_status_template_mapping_catalog(
        raw_root,
        repo_root=repo_root,
    )
    status_template_all_catalog, _ = load_status_template_mapping_catalog(
        raw_root,
        repo_root=repo_root,
        include_unchecked=True,
    )

    has_status_config = bool(status_sheet_config)
    workbook_sheet_cache: dict[str, list[str]] = {}
    workbooks_by_project: dict[str, list[Path]] = {}
    for workbook in candidates:
        identity = parse_project_identity_from_filename(workbook.name)
        project_code = str(identity.get("project_code", "")).strip() or workbook.stem
        project_key = ingest.normalize_project_code_key(project_code)
        workbooks_by_project.setdefault(project_key, []).append(workbook)
        sheet_names, _ = ingest.list_excel_sheet_names(workbook)
        workbook_sheet_cache[str(workbook.resolve())] = list(sheet_names or [])

    raw_frames: list[pd.DataFrame] = []
    diagnostics_rows: list[dict[str, object]] = []
    issue_rows: list[dict[str, object]] = []
    coverage_rows: list[dict[str, object]] = []

    skipped_not_in_config = 0
    skipped_blank_config = 0

    for project_key, project_workbooks in sorted(workbooks_by_project.items()):
        configured_entries = status_sheet_config.get(project_key)
        if has_status_config and configured_entries is None:
            skipped_not_in_config += len(project_workbooks)
            continue
        if configured_entries is not None and not configured_entries:
            skipped_blank_config += len(project_workbooks)
            identity = parse_project_identity_from_filename(project_workbooks[0].name)
            project_code = str(identity.get("project_code", "")).strip() or project_workbooks[0].stem
            coverage_rows.append(
                {
                    "project_code": project_code,
                    "project_display": project_code,
                    "status": "SKIPPED_NO_STATUS_CONFIG",
                    "reason_code": "SKIPPED_NO_STATUS_CONFIG",
                    "reason": "Project has blank Status Sheet Names in DPR_Config.",
                    "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                    "configured_sheet": "",
                    "resolved_sheet": "",
                    "rows": 0,
                    "filename_date": "",
                    "internal_date": "",
                    "date_quality": "UNREADABLE",
                    "available_sheets": "",
                }
            )
            continue

        sheet_requests = configured_entries if configured_entries else [{"sheet_name": "", "line_name": "", "line_name_source": ""}]
        template_error = status_template_errors.get(project_key, "")
        for request in sheet_requests:
            configured_sheet = str(request.get("sheet_name", "")).strip()
            configured_line_name = normalize_line_name(request.get("line_name", ""))
            configured_line_source = str(request.get("line_name_source", "")).strip()
            configured_file_identifier = str(request.get("file_identifier", "")).strip()
            selected_workbooks = _pick_matching_workbooks_for_sheet(
                project_workbooks,
                workbook_sheet_cache,
                configured_sheet,
                configured_line_name,
                configured_file_identifier,
            )

            if not selected_workbooks:
                identity = parse_project_identity_from_filename(project_workbooks[0].name)
                project_code = str(identity.get("project_code", "")).strip() or project_workbooks[0].stem
                project_display = build_project_display(project_code, configured_line_name, project_code) or project_code
                coverage_rows.append(
                    {
                        "project_code": project_code,
                        "project_display": project_display,
                        "status": "NO_TARGET_SHEET",
                        "reason_code": "NO_TARGET_SHEET",
                        "reason": "Configured status sheet not found in any workbook for project.",
                        "workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                        "configured_sheet": configured_sheet,
                        "resolved_sheet": "",
                        "rows": 0,
                        "filename_date": "",
                        "internal_date": "",
                        "date_quality": "UNREADABLE",
                        "available_sheets": "",
                    }
                )
                issue_rows.append(
                    {
                        "Workbook": "; ".join(sorted({w.name for w in project_workbooks})),
                        "Project": project_code,
                        "Sheet": "",
                        "ConfiguredSheet": configured_sheet,
                        "LineName": configured_line_name,
                        "LineNameSource": configured_line_source,
                        "Issue": "NO_TARGET_SHEET",
                            "Reason": "Configured status sheet not found in project workbooks.",
                        }
                    )
                continue

            for selected_workbook, resolved_sheet_guess in selected_workbooks:
                selected_identity = parse_project_identity_from_filename(selected_workbook.name)
                project_code = str(selected_identity.get("project_code", "")).strip() or selected_workbook.stem
                line_name = configured_line_name or normalize_line_name(selected_identity.get("line_name", ""))
                line_source = configured_line_source or ("config" if configured_line_name else "filename")
                project_display = build_project_display(project_code, line_name, project_code) or project_code
                project_scope_key = build_project_scope_key(project_code, line_name, project_display)
                available_sheet_text = "; ".join(workbook_sheet_cache.get(str(selected_workbook.resolve()), []))
                filename_date = _extract_report_date_from_filename(selected_workbook.name)

                if template_error:
                    coverage_rows.append(
                        {
                            "project_code": project_code,
                            "project_display": project_display,
                            "status": "TEMPLATE_CONFIG_ERROR",
                            "reason_code": "TEMPLATE_CONFIG_ERROR",
                            "reason": template_error,
                            "workbook": selected_workbook.name,
                            "configured_sheet": configured_sheet,
                            "resolved_sheet": resolved_sheet_guess,
                            "rows": 0,
                            "filename_date": filename_date,
                            "internal_date": "",
                            "date_quality": "UNREADABLE",
                            "available_sheets": available_sheet_text,
                        }
                    )
                    issue_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Sheet": resolved_sheet_guess,
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "Issue": "TEMPLATE_CONFIG_ERROR",
                            "Reason": template_error,
                        }
                    )
                    continue

                selector_name = resolved_sheet_guess or configured_sheet
                selector = _build_exact_sheet_selector(selector_name) if selector_name else (
                    lambda names: names[0] if names else None
                )
                try:
                    df_raw, resolved_sheet, fallback_note = load_sheet_with_csv_fallback(
                        selected_workbook,
                        selector,
                        read_excel_kwargs={"header": None},
                        read_csv_kwargs={"header": None},
                    )
                except Exception as exc:
                    issue_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Sheet": resolved_sheet_guess,
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "Issue": "READ_FAIL",
                            "Reason": str(exc),
                        }
                    )
                    coverage_rows.append(
                        {
                            "project_code": project_code,
                            "project_display": project_display,
                            "status": "READ_FAIL",
                            "reason_code": "READ_FAIL",
                            "reason": str(exc),
                            "workbook": selected_workbook.name,
                            "configured_sheet": configured_sheet,
                            "resolved_sheet": resolved_sheet_guess,
                            "rows": 0,
                            "filename_date": filename_date,
                            "internal_date": "",
                            "date_quality": "UNREADABLE",
                            "available_sheets": available_sheet_text,
                        }
                    )
                    continue

                if df_raw is None or resolved_sheet is None:
                    issue_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Sheet": resolved_sheet_guess,
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "Issue": "NO_TARGET_SHEET",
                            "Reason": "Configured status sheet not found in selected workbook.",
                        }
                    )
                    coverage_rows.append(
                        {
                            "project_code": project_code,
                            "project_display": project_display,
                            "status": "NO_TARGET_SHEET",
                            "reason_code": "NO_TARGET_SHEET",
                            "reason": "Configured status sheet not found in selected workbook.",
                            "workbook": selected_workbook.name,
                            "configured_sheet": configured_sheet,
                            "resolved_sheet": "",
                            "rows": 0,
                            "filename_date": filename_date,
                            "internal_date": "",
                            "date_quality": "UNREADABLE",
                            "available_sheets": available_sheet_text,
                        }
                    )
                    continue

                internal_date = _extract_internal_report_date(df_raw)
                date_quality = _date_quality_flag(filename_date, internal_date)

                selected_template = select_status_template_for_sheet(
                    status_template_catalog.get(project_key, []),
                    configured_sheet_name=configured_sheet,
                    resolved_sheet_name=resolved_sheet,
                    line_name=line_name,
                )
                if selected_template is None:
                    selected_template = select_status_template_for_sheet(
                        status_template_all_catalog.get(project_key, []),
                        configured_sheet_name=configured_sheet,
                        resolved_sheet_name=resolved_sheet,
                        line_name=line_name,
                    )
                template_map = dict(selected_template.get("column_map", {}) if selected_template else {})
                guardrails = dict(selected_template.get("guardrails", {}) if selected_template else {})
                template_sheet = str(selected_template.get("template_sheet", "") if selected_template else "")
                stringing_resolution_policy = _normalize_key(guardrails.get("stringing_resolution", ""))

                parse_result = _parse_status_sheet_dataframe(
                    df_raw,
                    guardrails=guardrails,
                    template_map=template_map,
                    project_code=project_code,
                    project_display=project_display,
                    project_scope_key=project_scope_key,
                    line_name=line_name,
                    line_name_source=line_source,
                    report_date=filename_date,
                    source_file=selected_workbook.name,
                    source_sheet=resolved_sheet,
                    configured_sheet=configured_sheet,
                    template_sheet=template_sheet,
                    stringing_resolution_policy=stringing_resolution_policy,
                )
                if not parse_result.data.empty:
                    raw_frames.append(parse_result.data)

                diagnostics_rows.append(
                    {
                        "Workbook": selected_workbook.name,
                        "Project": project_code,
                        "Sheet": resolved_sheet,
                        "ConfiguredSheet": configured_sheet,
                        "LineName": line_name,
                        "LineNameSource": line_source,
                        "TemplateSheet": template_sheet,
                        "TemplateApplied": bool(template_map),
                        "TemplateChanges": "; ".join(parse_result.template_changes),
                        "FallbackNote": fallback_note or "",
                        "SectionsDetected": int(parse_result.sections_detected),
                        "HeadersDetected": int(parse_result.headers_detected),
                        "Rows": int(parse_result.rows_emitted),
                        "FilenameDate": filename_date,
                        "InternalDate": internal_date,
                        "DateQuality": date_quality,
                        "Status": parse_result.parse_status,
                        "Reason": parse_result.parse_reason,
                    }
                )
                if parse_result.parse_status != "OK":
                    issue_rows.append(
                        {
                            "Workbook": selected_workbook.name,
                            "Project": project_code,
                            "Sheet": resolved_sheet,
                            "ConfiguredSheet": configured_sheet,
                            "LineName": line_name,
                            "LineNameSource": line_source,
                            "Issue": parse_result.parse_status,
                            "Reason": parse_result.parse_reason,
                        }
                    )
                coverage_rows.append(
                    {
                        "project_code": project_code,
                        "project_display": project_display,
                        "status": parse_result.parse_status,
                        "reason_code": parse_result.parse_status,
                        "reason": parse_result.parse_reason,
                        "workbook": selected_workbook.name,
                        "configured_sheet": configured_sheet,
                        "resolved_sheet": resolved_sheet,
                        "rows": int(parse_result.rows_emitted),
                        "filename_date": filename_date,
                        "internal_date": internal_date,
                        "date_quality": date_quality,
                        "available_sheets": available_sheet_text,
                    }
                )

    if skipped_blank_config:
        print(f"[pipeline] ProgressStatus: skipped {skipped_blank_config} workbook(s) per DPR_Config (blank status sheet config).")
    if skipped_not_in_config:
        print(f"[pipeline] ProgressStatus: skipped {skipped_not_in_config} workbook(s) not listed in DPR_Config.")

    raw_df = pd.concat(raw_frames, ignore_index=True) if raw_frames else pd.DataFrame(columns=RAWDATA_COLUMNS)
    raw_df = _merge_status_history(raw_df, Path(output_path))

    _CORE_ACTIVITY_NORMS = {"foundation", "tower_erection", "stringing"}
    if not raw_df.empty:
        core_mask = raw_df["activity_norm"].isin(_CORE_ACTIVITY_NORMS)
        for proj_code, grp in raw_df[core_mask].groupby("project_code"):
            if grp["quantity_primary"].isna().all():
                diagnostics_rows.append({
                    "Workbook": grp["source_file"].iloc[0] if "source_file" in grp.columns else "",
                    "Project": proj_code,
                    "Sheet": grp["source_sheet"].iloc[0] if "source_sheet" in grp.columns else "",
                    "ConfiguredSheet": grp["configured_sheet"].iloc[0] if "configured_sheet" in grp.columns else "",
                    "LineName": "",
                    "LineNameSource": "",
                    "TemplateSheet": grp["template_sheet"].iloc[0] if "template_sheet" in grp.columns else "",
                    "TemplateApplied": False,
                    "TemplateChanges": "",
                    "FallbackNote": "",
                    "SectionsDetected": 0,
                    "HeadersDetected": 0,
                    "Rows": len(grp),
                    "FilenameDate": None,
                    "InternalDate": None,
                    "DateQuality": "",
                    "Status": "WARN",
                    "Reason": (
                        f"quantity_primary_all_null: {len(grp)} core activity row(s) have no quantity "
                        "— template column mapping may be wrong"
                    ),
                })

    diagnostics_df = pd.DataFrame(diagnostics_rows, columns=DIAGNOSTICS_COLUMNS)
    issues_df = pd.DataFrame(issue_rows, columns=ISSUES_COLUMNS)
    coverage_df = pd.DataFrame(coverage_rows, columns=COVERAGE_COLUMNS)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output_path.with_suffix(f"{output_path.suffix}.tmp")

    try:
        with pd.ExcelWriter(temp_output, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name="RawData", index=False)
            diagnostics_df.to_excel(writer, sheet_name="Diagnostics", index=False)
            issues_df.to_excel(writer, sheet_name="Issues", index=False)
            coverage_df.to_excel(writer, sheet_name="Coverage", index=False)
        temp_output.replace(output_path)
    finally:
        if temp_output.exists():
            try:
                temp_output.unlink()
            except Exception:
                pass

    print(
        f"[pipeline] ProgressStatus: wrote workbook {output_path} "
        f"(rows={len(raw_df.index)}, diagnostics={len(diagnostics_df.index)}, issues={len(issues_df.index)})."
    )
    return output_path
