from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from dashboard import stretch_readiness_ingest as stretch_ingest


class StretchReadinessIngestTests(unittest.TestCase):
    def _build_derived_rows(
        self,
        *,
        stringing_rows: list[dict[str, object]],
        erection_rows: list[dict[str, object]],
        allowed_keys: set[str] | None = None,
    ) -> pd.DataFrame:
        with tempfile.TemporaryDirectory() as temp_dir:
            parquets_root = Path(temp_dir) / "Parquets"
            stringing_root = parquets_root / "Stringing"
            erection_root = parquets_root / "Erection"
            stringing_root.mkdir(parents=True, exist_ok=True)
            erection_root.mkdir(parents=True, exist_ok=True)
            pd.DataFrame(stringing_rows).to_parquet(stringing_root / "StringingCompiled.parquet", index=False)
            pd.DataFrame(erection_rows).to_parquet(erection_root / "RawData.parquet", index=False)
            return stretch_ingest._build_derived_stretch_rows(  # type: ignore[attr-defined]
                parquets_root=parquets_root,
                allowed_project_keys=allowed_keys or {"ta501"},
            )

    def test_extract_guardrails(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        ws["A1"] = "Guardrails"
        ws["A2"] = "required_tokens"
        ws["B2"] = "loc; final checking"
        ws["A3"] = "header_rows"
        ws["B3"] = "1"
        ws["A4"] = "To Map"
        ws["A5"] = "stretch_identifier"

        parsed = stretch_ingest._extract_guardrails_stretch(ws)  # type: ignore[attr-defined]
        self.assertEqual(parsed.get("required_tokens"), "loc; final checking")
        self.assertEqual(parsed.get("header_rows"), "1")

    def test_parse_tb507_both_required_rule(self) -> None:
        df = pd.DataFrame(
            [
                ["Client", "", "", ""],
                ["SL No", "Loc No", "Final Checking", "Tack Welding"],
                [1, "65/0", "DONE", "DONE"],
                [2, "65/1", "DONE", ""],
                [3, "66/0", "ROW", ""],
            ]
        )
        guardrails = {
            "required_tokens": "loc; final checking; tack welding",
            "header_rows": "1",
            "readiness_rule": "both_required",
        }
        template_map = {
            1: "stretch_identifier",
            2: "final_check_raw",
            3: "tack_welding_raw",
        }

        result = stretch_ingest._parse_stretch_sheet_dataframe(  # type: ignore[attr-defined]
            df,
            guardrails=guardrails,
            template_map=template_map,
            project_code="TB 507",
            project_display="TB 507 - MAIN",
            project_scope_key="tb507main",
            line_name="MAIN",
            line_name_source="config",
            source_file="TB 507 [MAIN] - DPR - 2026-04-24.xlsx",
            source_sheet="Final Check In And Tack Welding",
            configured_sheet="Final Check In And Tack Welding",
            template_sheet="TB 507 Stretch",
            report_date="2026-04-24",
        )
        self.assertEqual(result.parse_status, "OK")
        states = set(result.data["readiness_state"].astype(str).tolist())
        self.assertIn("READY", states)
        self.assertIn("PARTIAL", states)
        self.assertIn("NOT_READY", states)

    def test_detect_manpower_signal(self) -> None:
        frame = pd.DataFrame(
            {
                "From AP": ["1/0", "2/0"],
                "Gang Strength": [12, None],
                "Section Readiness": ["Ready", "Ready"],
            }
        )
        signal = stretch_ingest._detect_manpower_signal_from_frame(frame)  # type: ignore[attr-defined]
        self.assertEqual(signal.get("signal_type"), "PRESENT_WITH_VALUES")
        self.assertTrue(signal.get("readiness_present"))

    def test_required_locations_include_endpoints_when_location_nos_empty(self) -> None:
        derived = self._build_derived_rows(
            stringing_rows=[
                {
                    "project_code": "TA 501",
                    "project_display": "TA 501 - MAIN",
                    "line_name": "MAIN",
                    "project_scope_key": "ta501main",
                    "from_ap": "65/0",
                    "to_ap": "65/3",
                    "location nos": "",
                    "length_m": 1200,
                    "source_file": "TA 501 - DPR - 2026-05-05.xlsx",
                }
            ],
            erection_rows=[
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/0", "Tower Tightening": "2026-05-02"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/3", "Tower Tightening": "2026-05-03"},
            ],
        )
        self.assertEqual(len(derived.index), 1)
        self.assertEqual(str(derived.loc[0, "readiness_state"]), "READY")
        self.assertEqual(int(derived.loc[0, "required_location_count"]), 2)

    def test_required_locations_add_inbetween_and_keep_endpoints(self) -> None:
        derived = self._build_derived_rows(
            stringing_rows=[
                {
                    "project_code": "TA 501",
                    "project_display": "TA 501 - MAIN",
                    "line_name": "MAIN",
                    "project_scope_key": "ta501main",
                    "from_ap": "65/0",
                    "to_ap": "65/3",
                    "location nos": "65/1,65/2",
                    "length_m": 1200,
                    "source_file": "TA 501 - DPR - 2026-05-05.xlsx",
                }
            ],
            erection_rows=[
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/0", "Tower Tightening": "2026-05-01"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/1", "Tower Tightening": "2026-05-01"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/2", "Tower Tightening": "2026-05-01"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/3", "Tower Tightening": "2026-05-01"},
            ],
        )
        self.assertEqual(str(derived.loc[0, "readiness_state"]), "READY")
        self.assertEqual(int(derived.loc[0, "required_location_count"]), 4)
        required = str(derived.loc[0, "required_locations"])
        self.assertIn("65/0", required)
        self.assertIn("65/3", required)
        self.assertIn("65/1", required)
        self.assertIn("65/2", required)

    def test_shorthand_only_location_nos_falls_back_to_endpoints_only(self) -> None:
        derived = self._build_derived_rows(
            stringing_rows=[
                {
                    "project_code": "TA 501",
                    "project_display": "TA 501 - MAIN",
                    "line_name": "MAIN",
                    "project_scope_key": "ta501main",
                    "from_ap": "65/0",
                    "to_ap": "65/3",
                    "location nos": "1,2,3",
                    "length_m": 900,
                    "source_file": "TA 501 - DPR - 2026-05-05.xlsx",
                }
            ],
            erection_rows=[
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/0", "Tower Tightening": "2026-05-01"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/3", "Tower Tightening": "2026-05-01"},
            ],
        )
        self.assertEqual(str(derived.loc[0, "location_parse_status"]), "SHORTHAND_NO_ANCHOR")
        self.assertEqual(int(derived.loc[0, "required_location_count"]), 2)
        self.assertEqual(str(derived.loc[0, "readiness_state"]), "READY")

    def test_mixed_explicit_and_shorthand_tokens_expand_with_anchor(self) -> None:
        derived = self._build_derived_rows(
            stringing_rows=[
                {
                    "project_code": "TA 501",
                    "project_display": "TA 501 - MAIN",
                    "line_name": "MAIN",
                    "project_scope_key": "ta501main",
                    "from_ap": "65/0",
                    "to_ap": "65/4",
                    "location nos": "65/1,2,3",
                    "length_m": 1300,
                    "source_file": "TA 501 - DPR - 2026-05-05.xlsx",
                }
            ],
            erection_rows=[
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/0", "Tower Tightening": "2026-05-01"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/1", "Tower Tightening": "2026-05-01"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/2", "Tower Tightening": "2026-05-01"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/3", "Tower Tightening": "2026-05-01"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/4", "Tower Tightening": "2026-05-01"},
            ],
        )
        self.assertEqual(str(derived.loc[0, "location_parse_status"]), "OK")
        self.assertEqual(str(derived.loc[0, "readiness_state"]), "READY")
        required = str(derived.loc[0, "required_locations"])
        self.assertIn("65/1", required)
        self.assertIn("65/2", required)
        self.assertIn("65/3", required)

    def test_unmatched_required_location_marks_section_not_ready(self) -> None:
        derived = self._build_derived_rows(
            stringing_rows=[
                {
                    "project_code": "TA 501",
                    "project_display": "TA 501 - MAIN",
                    "line_name": "MAIN",
                    "project_scope_key": "ta501main",
                    "from_ap": "65/0",
                    "to_ap": "65/3",
                    "location nos": "65/1,65/2",
                    "length_m": 1200,
                    "source_file": "TA 501 - DPR - 2026-05-05.xlsx",
                }
            ],
            erection_rows=[
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/0", "Tower Tightening": "2026-05-01"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/1", "Tower Tightening": "2026-05-01"},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/2", "Tower Tightening": ""},
                {"project_code": "TA 501", "line_name": "MAIN", "project_scope_key": "ta501main", "Location No.": "65/3", "Tower Tightening": "2026-05-01"},
            ],
        )
        self.assertEqual(str(derived.loc[0, "readiness_state"]), "NOT_READY")
        self.assertGreater(int(derived.loc[0, "unmatched_location_count"]), 0)

    def test_merge_prefers_derived_and_summary_dedupes_latest_per_section(self) -> None:
        legacy = pd.DataFrame(
            [
                {
                    "project_code": "TA 501",
                    "project_display": "TA 501 - MAIN",
                    "project_scope_key": "ta501main",
                    "line_name": "MAIN",
                    "line_name_source": "legacy",
                    "section_label": "S1",
                    "source_file": "TA 501 - DPR - 2026-05-01.xlsx",
                    "source_sheet": "Stretch",
                    "configured_sheet": "Stretch",
                    "template_sheet": "",
                    "report_date": "2026-05-01",
                    "header_row_number": 1,
                    "source_row_number": 1,
                    "stretch_identifier": "65/0 - 65/3",
                    "from_ap": "65/0",
                    "to_ap": "65/3",
                    "length_m_raw": 1200,
                    "length_km": 1.2,
                    "readiness_raw": "Ready",
                    "final_check_raw": "",
                    "tack_welding_raw": "",
                    "balance_towers": 0,
                    "readiness_state": "READY",
                    "remarks": "",
                    "readiness_source": stretch_ingest.STRETCH_SOURCE_LEGACY,  # type: ignore[attr-defined]
                    "source_tag": stretch_ingest.STRETCH_SOURCE_LEGACY,  # type: ignore[attr-defined]
                    "location_nos_raw": "",
                    "location_parse_status": "",
                    "location_parse_issue": "",
                    "required_location_count": None,
                    "matched_location_count": None,
                    "unmatched_location_count": None,
                    "required_locations": "",
                    "matched_locations": "",
                    "unmatched_locations": "",
                }
            ]
        ).reindex(columns=stretch_ingest.STRETCH_RAWDATA_COLUMNS)  # type: ignore[attr-defined]
        derived = legacy.copy()
        derived.loc[0, "report_date"] = "2026-05-02"
        derived.loc[0, "readiness_state"] = "NOT_READY"
        derived.loc[0, "readiness_source"] = stretch_ingest.STRETCH_SOURCE_DERIVED  # type: ignore[attr-defined]
        derived.loc[0, "source_tag"] = stretch_ingest.STRETCH_SOURCE_DERIVED  # type: ignore[attr-defined]

        merged = stretch_ingest._merge_stretch_sources_prefer_derived(legacy, derived)  # type: ignore[attr-defined]
        self.assertEqual(len(merged.index), 1)
        self.assertEqual(str(merged.loc[0, "readiness_source"]), stretch_ingest.STRETCH_SOURCE_DERIVED)  # type: ignore[attr-defined]

        extra = merged.copy()
        extra.loc[0, "stretch_identifier"] = "66/0 - 66/1"
        extra.loc[0, "from_ap"] = "66/0"
        extra.loc[0, "to_ap"] = "66/1"
        extra.loc[0, "readiness_state"] = "READY"
        extra.loc[0, "length_km"] = 0.8
        raw = pd.concat([merged, extra], ignore_index=True)
        summary = stretch_ingest._build_stretch_summary(raw)  # type: ignore[attr-defined]
        self.assertEqual(int(summary.loc[0, "total_count"]), 2)
        self.assertEqual(int(summary.loc[0, "ready_count"]), 1)
        self.assertEqual(int(summary.loc[0, "not_ready_count"]), 1)


if __name__ == "__main__":
    unittest.main()
