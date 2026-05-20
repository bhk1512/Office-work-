from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

import export_dpr_weekly_analysis
from dashboard.workbook import export_erection_productivity_summary, export_weekly_dpr_analysis
from export_stringing_summary import (
    _build_week_windows,
    _build_weekly_comprehensive_table,
    _build_weekly_gang_table,
    _build_weekly_rollup_tables,
    _monthly_avg_column,
    _monthly_km_column,
)


class CalendarWeekExportTests(unittest.TestCase):
    def _write_compiled_daily(self, path: Path) -> None:
        frame = pd.DataFrame(
            [
                {
                    "project_code": "TA419",
                    "project_name": "TA419",
                    "line_name": "",
                    "work_date": "2026-05-12",
                    "start_date": "2026-05-10",
                    "completion_date": "2026-05-12",
                    "tower_weight": 40.0,
                    "productivity": 4.0,
                    "location_no": "1/0",
                    "tower_type": "DA",
                    "gang_name": "Alpha",
                    "status": "Done",
                },
                {
                    "project_code": "TA419",
                    "project_name": "TA419",
                    "line_name": "",
                    "work_date": "2026-05-17",
                    "start_date": "2026-05-15",
                    "completion_date": "2026-05-17",
                    "tower_weight": 42.0,
                    "productivity": 4.2,
                    "location_no": "1/1",
                    "tower_type": "DA",
                    "gang_name": "Beta",
                    "status": "Done",
                },
                {
                    "project_code": "TA419",
                    "project_name": "TA419",
                    "line_name": "",
                    "work_date": "2026-05-18",
                    "start_date": "2026-05-16",
                    "completion_date": "2026-05-18",
                    "tower_weight": 44.0,
                    "productivity": 4.4,
                    "location_no": "1/2",
                    "tower_type": "DA",
                    "gang_name": "Gamma",
                    "status": "Done",
                },
            ]
        )
        frame.to_excel(path, index=False)

    def test_dpr_calendar_mode_start_date_builds_sunday_windows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            compiled_path = root / "daily.xlsx"
            output_path = root / "weekly.xlsx"
            self._write_compiled_daily(compiled_path)

            export_weekly_dpr_analysis(
                project_code="TA419",
                output_path=output_path,
                daily_path=compiled_path,
                week_mode="calendar_sun_sat",
                week_start_date="2026-05-13",
            )

            summary = pd.read_excel(output_path, sheet_name="WeeklySummary")
            self.assertEqual(summary.loc[0, "Week"], "2026-05-10 to 2026-05-16")
            self.assertEqual(summary.loc[1, "Week"], "2026-05-17 to 2026-05-18")
            self.assertEqual(int(summary.loc[0, "Erections Completed"]), 1)
            self.assertEqual(int(summary.loc[1, "Erections Completed"]), 2)
            gang_weekly = pd.read_excel(output_path, sheet_name="GangWeekly")
            self.assertIn("2026-05-10 to 2026-05-16 MT", gang_weekly.columns)
            self.assertIn("2026-05-17 to 2026-05-18 MT", gang_weekly.columns)

    def test_dpr_calendar_mode_previous_week_uses_as_of_date(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            compiled_path = root / "daily.xlsx"
            output_path = root / "weekly_prev.xlsx"
            self._write_compiled_daily(compiled_path)

            export_weekly_dpr_analysis(
                project_code="TA419",
                output_path=output_path,
                daily_path=compiled_path,
                week_mode="calendar_sun_sat",
                previous_week=True,
                as_of_date="2026-05-19",
            )

            summary = pd.read_excel(output_path, sheet_name="WeeklySummary")
            self.assertEqual(len(summary), 1)
            self.assertEqual(summary.loc[0, "Week"], "2026-05-10 to 2026-05-16")
            self.assertEqual(int(summary.loc[0, "Erections Completed"]), 1)

    def test_erection_calendar_mode_uses_date_range_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            output_path = root / "erection.xlsx"
            daily = pd.DataFrame(
                [
                    {
                        "date": "2026-05-12",
                        "completion_date": "2026-05-12",
                        "start_date": "2026-05-10",
                        "location_no": "1/0",
                        "tower_type": "DA",
                        "tower_weight": 40.0,
                        "daily_prod_mt": 4.0,
                        "gang_name": "Alpha",
                        "status": "Done",
                        "project_name": "TA419",
                        "project_code": "TA419",
                        "line_name": "",
                    },
                    {
                        "date": "2026-05-17",
                        "completion_date": "2026-05-17",
                        "start_date": "2026-05-15",
                        "location_no": "1/1",
                        "tower_type": "DA",
                        "tower_weight": 42.0,
                        "daily_prod_mt": 4.2,
                        "gang_name": "Beta",
                        "status": "Done",
                        "project_name": "TA419",
                        "project_code": "TA419",
                        "line_name": "",
                    },
                ]
            )
            project_info = pd.DataFrame(
                [{"project_code": "TA419", "project_name": "TA419", "pch": "PCH 1", "region": "North"}]
            )

            export_erection_productivity_summary(
                output_path=output_path,
                daily_df=daily,
                project_info=project_info,
                week_mode="calendar_sun_sat",
                week_start_date="2026-05-13",
                week_end_date="2026-05-18",
            )

            wb = load_workbook(output_path, data_only=True)
            ws = wb["Erection Summary Weekly"]
            header_row = [ws.cell(row=2, column=col).value for col in range(1, 14)]
            self.assertIn("2026-05-10 to 2026-05-16", header_row)
            self.assertIn("2026-05-17 to 2026-05-18", header_row)
            self.assertIn("2026-05-10 to 2026-05-16 MT", header_row)
            self.assertIn("2026-05-17 to 2026-05-18 MT", header_row)

    def test_stringing_weekly_table_uses_month_equivalent_productivity(self) -> None:
        week_windows, week_labels = _build_week_windows(pd.Timestamp("2026-05-10"), pd.Timestamp("2026-05-18"))
        daily = pd.DataFrame(
            [
                {"date": "2026-05-10", "daily_km": 1.0, "gang_name": "A", "project_display": "TA419"},
                {"date": "2026-05-12", "daily_km": 2.0, "gang_name": "B", "project_display": "TA419"},
                {"date": "2026-05-18", "daily_km": 3.0, "gang_name": "A", "project_display": "TA419"},
            ]
        )
        compiled = pd.DataFrame(
            [
                {"fs_complete_date": "2026-05-12", "length_km": 1.2, "project_display": "TA419"},
                {"fs_complete_date": "2026-05-18", "length_km": 0.8, "project_display": "TA419"},
            ]
        )
        out = _build_weekly_comprehensive_table(
            daily,
            compiled,
            week_windows=week_windows,
            week_labels=week_labels,
        )
        self.assertEqual(out.loc[0, "Week"], "2026-05-10 to 2026-05-16")
        self.assertEqual(float(out.loc[0, "Avg Productivity (KM/month)"]), 45.0)
        self.assertEqual(float(out.loc[1, "Avg Productivity (KM/month)"]), 90.0)

    def test_stringing_weekly_gang_table_includes_weekly_km_and_productivity(self) -> None:
        week_windows, week_labels = _build_week_windows(pd.Timestamp("2026-05-10"), pd.Timestamp("2026-05-18"))
        daily = pd.DataFrame(
            [
                {"date": "2026-05-10", "daily_km": 1.0, "gang_name": "A", "project_display": "TA419"},
                {"date": "2026-05-12", "daily_km": 2.0, "gang_name": "A", "project_display": "TA419"},
                {"date": "2026-05-17", "daily_km": 3.0, "gang_name": "A", "project_display": "TA419"},
                {"date": "2026-05-11", "daily_km": 2.0, "gang_name": "B", "project_display": "TA419"},
            ]
        )
        out = _build_weekly_gang_table(
            daily,
            week_windows=week_windows,
            week_labels=week_labels,
        )
        self.assertIn("2026-05-10 to 2026-05-16 KM", out.columns)
        self.assertIn("2026-05-17 to 2026-05-18 KM", out.columns)
        gang_a = out.loc[out["Gang Name"] == "A"].iloc[0]
        self.assertEqual(float(gang_a["2026-05-10 to 2026-05-16 KM"]), 3.0)
        self.assertEqual(float(gang_a["2026-05-17 to 2026-05-18 KM"]), 3.0)
        self.assertEqual(float(gang_a["2026-05-10 to 2026-05-16"]), 45.0)
        self.assertEqual(float(gang_a["2026-05-17 to 2026-05-18"]), 90.0)

    def test_stringing_weekly_rollup_includes_project_wise_metrics(self) -> None:
        week_windows, week_labels = _build_week_windows(pd.Timestamp("2026-05-10"), pd.Timestamp("2026-05-18"))
        daily = pd.DataFrame(
            [
                {
                    "date": "2026-05-10",
                    "daily_km": 1.0,
                    "gang_name": "A",
                    "pch_display": "PCH 1",
                    "project_rollup_display": "TA419",
                },
                {
                    "date": "2026-05-12",
                    "daily_km": 2.0,
                    "gang_name": "A",
                    "pch_display": "PCH 1",
                    "project_rollup_display": "TA419",
                },
                {
                    "date": "2026-05-17",
                    "daily_km": 3.0,
                    "gang_name": "B",
                    "pch_display": "PCH 1",
                    "project_rollup_display": "TA419",
                },
            ]
        )
        compiled = pd.DataFrame(
            [
                {
                    "fs_complete_date": "2026-05-12",
                    "length_km": 1.2,
                    "pch_display": "PCH 1",
                    "project_rollup_display": "TA419",
                },
                {
                    "fs_complete_date": "2026-05-18",
                    "length_km": 0.8,
                    "pch_display": "PCH 1",
                    "project_rollup_display": "TA419",
                },
            ]
        )

        _, _, project = _build_weekly_rollup_tables(
            daily,
            compiled,
            week_windows=week_windows,
            week_labels=week_labels,
        )
        self.assertIn("PCH", project.columns)
        self.assertIn("Project", project.columns)
        self.assertIn(_monthly_avg_column("2026-05-10 to 2026-05-16"), project.columns)
        self.assertIn(_monthly_km_column("2026-05-17 to 2026-05-18"), project.columns)
        row = project.loc[(project["PCH"] == "PCH 1") & (project["Project"] == "TA419")].iloc[0]
        self.assertEqual(float(row[_monthly_avg_column("2026-05-10 to 2026-05-16")]), 45.0)
        self.assertEqual(float(row[_monthly_km_column("2026-05-10 to 2026-05-16")]), 1.2)

    def test_cli_rejects_end_date_without_start_date(self) -> None:
        argv = [
            "export_dpr_weekly_analysis.py",
            "TA419",
            "--end-date",
            "2026-05-18",
        ]
        with patch("sys.argv", argv):
            with self.assertRaises(SystemExit):
                export_dpr_weekly_analysis.main()

    def test_cli_previous_week_promotes_legacy_mode_to_calendar_mode(self) -> None:
        captured: dict[str, object] = {}

        def _fake_export(*_args, **kwargs):
            captured.update(kwargs)

        argv = [
            "export_dpr_weekly_analysis.py",
            "TA419",
            "--previous-week",
            "--dpr-only",
            "--dpr-path",
            "C:/tmp/nonexistent.xlsx",
        ]
        with patch("sys.argv", argv):
            with patch.object(export_dpr_weekly_analysis, "configure_logging"):
                with patch.object(export_dpr_weekly_analysis, "export_weekly_dpr_analysis", side_effect=_fake_export):
                    export_dpr_weekly_analysis.main()
        self.assertEqual(captured.get("week_mode"), "calendar_sun_sat")


if __name__ == "__main__":
    unittest.main()
