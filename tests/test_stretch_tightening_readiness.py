from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from dashboard.stretch_readiness_ingest import _is_valid_date_value
from dashboard.stretch_tightening_readiness import (
    build_stretch_tightening_readiness_tables,
    write_stretch_tightening_readiness_workbook,
)


class StretchTighteningReadinessTests(unittest.TestCase):
    def test_tightening_completion_accepts_done_like_tokens_and_rejects_blockers(self) -> None:
        for value in ("2026-05-11", "11/05/2026", "Yes", "Y", "True", "OK", "Done", "Completed", "Complete", "C"):
            with self.subTest(value=value):
                self.assertTrue(_is_valid_date_value(value))

        for value in ("", None, "No", "N", "False", "Pending", "WIP", "Balance", "ROW", "Hold", "Blocked"):
            with self.subTest(value=value):
                self.assertFalse(_is_valid_date_value(value))

    def test_span_ready_requires_erection_tightening_and_unstrung_status(self) -> None:
        erection_raw = pd.DataFrame(
            [
                {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "1/0", "Complete Date": "2026-05-01", "Tower Tightening Raw": "Done"},
                {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "1/1", "Complete Date": "2026-05-02", "Tower Tightening Raw": "OK"},
                {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "2/0", "Complete Date": "2026-05-03", "Tower Tightening Raw": "Completed"},
                {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "2/1", "Complete Date": "2026-05-04", "Tower Tightening Raw": "Yes"},
                {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "3/0", "Complete Date": "2026-05-05", "Tower Tightening Raw": "No"},
                {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "3/1", "Complete Date": "2026-05-06", "Tower Tightening Raw": ""},
            ]
        )
        stringing_compiled_raw = pd.DataFrame(
            [
                {"project_code": "TA 501", "project_scope_key": "ta501", "from_ap": "1/0", "to_ap": "1/1", "location nos": "", "length_m": 1000, "status": "Balance"},
                {"project_code": "TA 501", "project_scope_key": "ta501", "from_ap": "2/0", "to_ap": "2/1", "location nos": "", "length_m": 1500, "status": "Completed"},
                {"project_code": "TA 501", "project_scope_key": "ta501", "from_ap": "3/0", "to_ap": "3/1", "location nos": "", "length_m": 2000, "status": "Balance"},
            ]
        )

        tables = build_stretch_tightening_readiness_tables(
            erection_raw=erection_raw,
            stringing_compiled_raw=stringing_compiled_raw,
        )

        spans = tables["Span Readiness"]
        self.assertEqual(int(spans["ready_to_string"].sum()), 1)
        self.assertEqual(int(spans["already_strung"].sum()), 1)
        self.assertIn("TIGHTENING_PENDING", set(spans["readiness_reason"].astype(str)))

        project = tables["Project Summary"].iloc[0]
        self.assertEqual(int(project["usable_stringing_spans"]), 3)
        self.assertEqual(int(project["ready_to_string_spans"]), 1)
        self.assertAlmostEqual(float(project["already_strung_km"]), 1.5)
        self.assertAlmostEqual(float(project["ready_to_string_km"]), 1.0)
        self.assertAlmostEqual(float(project["tightening_not_started_km"]), 2.0)
        self.assertTrue(pd.isna(project["tightening_partial_km"]))

        buckets = tables["Span Buckets"]
        bucket = buckets.iloc[0]
        self.assertEqual(int(bucket["Total Spans"]), 3)
        self.assertEqual(int(bucket["Already Strung spans"]), 1)
        self.assertAlmostEqual(float(bucket["Already Strung km"]), 1.5)
        self.assertEqual(int(bucket["Stringing Ready"]), 1)
        self.assertAlmostEqual(float(bucket["Stringing Ready km"]), 1.0)
        self.assertEqual(int(bucket["Tightening not started"]), 1)
        self.assertAlmostEqual(float(bucket["Tightening not started km"]), 2.0)
        self.assertEqual(int(bucket["Tightening partial"]), 0)
        self.assertTrue(pd.isna(bucket["Tightening partial km"]))
        self.assertEqual(int(bucket["Erection Gap"]), 0)

        executive = tables["Executive Summary"]
        exec_map = dict(zip(executive["Metric"], executive["Value"]))
        self.assertAlmostEqual(float(exec_map["Already strung km"]), 1.5)
        self.assertAlmostEqual(float(exec_map["Stringing ready km"]), 1.0)
        self.assertAlmostEqual(float(exec_map["Tightening not started km"]), 2.0)
        self.assertTrue(pd.isna(exec_map["Tightening partial km"]))

    def test_location_nos_must_all_be_tightened(self) -> None:
        erection_raw = pd.DataFrame(
            [
                {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "65/0", "Complete Date": "2026-05-01", "Tower Tightening Raw": "Done"},
                {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "65/1", "Complete Date": "2026-05-02", "Tower Tightening Raw": "Done"},
                {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "65/2", "Complete Date": "2026-05-03", "Tower Tightening Raw": "Pending"},
                {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "65/3", "Complete Date": "2026-05-04", "Tower Tightening Raw": "Done"},
            ]
        )
        stringing_compiled_raw = pd.DataFrame(
            [
                {
                    "project_code": "TA 501",
                    "project_scope_key": "ta501",
                    "from_ap": "65/0",
                    "to_ap": "65/3",
                    "location nos": "65/1,65/2",
                    "length_m": 1200,
                    "status": "Balance",
                }
            ]
        )

        tables = build_stretch_tightening_readiness_tables(
            erection_raw=erection_raw,
            stringing_compiled_raw=stringing_compiled_raw,
        )
        span = tables["Span Readiness"].iloc[0]
        self.assertFalse(bool(span["ready_to_string"]))
        self.assertEqual(str(span["missing_tightening_locations"]), "65/2")
        buckets = tables["Span Buckets"]
        self.assertEqual(int(buckets.iloc[0]["Tightening partial"]), 1)

    def test_ta413_and_ta416_tightening_is_assumed_done_when_erected(self) -> None:
        tables = build_stretch_tightening_readiness_tables(
            erection_raw=pd.DataFrame(
                [
                    {"Project Code": "TA 413", "Project Scope Key": "ta413", "Location No.": "1/0", "Complete Date": "2026-05-01", "Tower Tightening Raw": ""},
                    {"Project Code": "TA 413", "Project Scope Key": "ta413", "Location No.": "1/1", "Complete Date": "2026-05-02", "Tower Tightening Raw": ""},
                ]
            ),
            stringing_compiled_raw=pd.DataFrame(
                [{"project_code": "TA 413", "project_scope_key": "ta413", "from_ap": "1/0", "to_ap": "1/1", "length_m": 1000, "status": "Balance"}]
            ),
        )
        buckets = tables["Span Buckets"]
        self.assertEqual(int(buckets.iloc[0]["Stringing Ready"]), 1)
        self.assertEqual(int(tables["Project Summary"].iloc[0]["tightened_towers"]), 2)

    def test_project_variants_are_split_by_scope(self) -> None:
        tables = build_stretch_tightening_readiness_tables(
            erection_raw=pd.DataFrame(
                [
                    {"Project Code": "TB 507", "Project Display": "TB 507", "Project Scope Key": "tb507", "Line Name": "", "Location No.": "1/0", "Complete Date": "2026-05-01", "Tower Tightening Raw": "Done"},
                    {"Project Code": "TB 507", "Project Display": "TB 507", "Project Scope Key": "tb507", "Line Name": "", "Location No.": "1/1", "Complete Date": "2026-05-02", "Tower Tightening Raw": "Done"},
                    {"Project Code": "TB 507", "Project Display": "TB 507 - MAIN", "Project Scope Key": "tb507main", "Line Name": "MAIN", "Location No.": "2/0", "Complete Date": "2026-05-01", "Tower Tightening Raw": ""},
                    {"Project Code": "TB 507", "Project Display": "TB 507 - MAIN", "Project Scope Key": "tb507main", "Line Name": "MAIN", "Location No.": "2/1", "Complete Date": "2026-05-02", "Tower Tightening Raw": ""},
                ]
            ),
            stringing_compiled_raw=pd.DataFrame(
                [
                    {"project_code": "TB 507", "project_display": "TB 507", "project_scope_key": "tb507", "line_name": "", "from_ap": "1/0", "to_ap": "1/1", "length_m": 1000, "status": "Balance"},
                    {"project_code": "TB 507", "project_display": "TB 507 - MAIN", "project_scope_key": "tb507main", "line_name": "MAIN", "from_ap": "2/0", "to_ap": "2/1", "length_m": 1000, "status": "Balance"},
                ]
            ),
        )
        projects = set(tables["Span Buckets"]["Project"].astype(str))
        self.assertEqual(projects, {"TB 507", "TB 507 - MAIN"})

    def test_km_aggregation_excludes_unknown_lengths_and_blanks_when_no_known(self) -> None:
        tables = build_stretch_tightening_readiness_tables(
            erection_raw=pd.DataFrame(
                [
                    {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "1/0", "Complete Date": "2026-05-01", "Tower Tightening Raw": "Done"},
                    {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "1/1", "Complete Date": "2026-05-02", "Tower Tightening Raw": "Done"},
                    {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "2/0", "Complete Date": "2026-05-03", "Tower Tightening Raw": "Done"},
                    {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "2/1", "Complete Date": "2026-05-04", "Tower Tightening Raw": "Done"},
                    {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "3/0", "Complete Date": "2026-05-05", "Tower Tightening Raw": "Done"},
                    {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "3/1", "Complete Date": "2026-05-06", "Tower Tightening Raw": "Pending"},
                    {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "4/0", "Complete Date": "2026-05-07", "Tower Tightening Raw": "Done"},
                    {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "4/1", "Complete Date": "2026-05-08", "Tower Tightening Raw": "Pending"},
                ]
            ),
            stringing_compiled_raw=pd.DataFrame(
                [
                    {"project_code": "TA 501", "project_scope_key": "ta501", "from_ap": "1/0", "to_ap": "1/1", "length_m": 1200, "status": "Completed"},
                    {"project_code": "TA 501", "project_scope_key": "ta501", "from_ap": "2/0", "to_ap": "2/1", "length_m": "", "status": "Completed"},
                    {"project_code": "TA 501", "project_scope_key": "ta501", "from_ap": "3/0", "to_ap": "3/1", "length_m": 1800, "status": "Balance"},
                    {"project_code": "TA 501", "project_scope_key": "ta501", "from_ap": "4/0", "to_ap": "4/1", "length_m": "", "status": "Balance"},
                ]
            ),
        )
        bucket = tables["Span Buckets"].iloc[0]
        self.assertEqual(int(bucket["Already Strung spans"]), 2)
        self.assertAlmostEqual(float(bucket["Already Strung km"]), 1.2)
        self.assertEqual(int(bucket["Tightening partial"]), 2)
        self.assertAlmostEqual(float(bucket["Tightening partial km"]), 1.8)
        self.assertTrue(pd.isna(bucket["Tightening not started km"]))

        project = tables["Project Summary"].iloc[0]
        self.assertAlmostEqual(float(project["already_strung_km"]), 1.2)
        self.assertAlmostEqual(float(project["tightening_partial_km"]), 1.8)
        self.assertTrue(pd.isna(project["tightening_not_started_km"]))

    def test_workbook_export_contains_expected_sheets(self) -> None:
        tables = build_stretch_tightening_readiness_tables(
            erection_raw=pd.DataFrame(
                [
                    {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "1/0", "Complete Date": "2026-05-01", "Tower Tightening Raw": "Done"},
                    {"Project Code": "TA 501", "Project Scope Key": "ta501", "Location No.": "1/1", "Complete Date": "2026-05-02", "Tower Tightening Raw": "Done"},
                ]
            ),
            stringing_compiled_raw=pd.DataFrame(
                [{"project_code": "TA 501", "project_scope_key": "ta501", "from_ap": "1/0", "to_ap": "1/1", "length_m": 1000}]
            ),
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "stretch.xlsx"
            write_stretch_tightening_readiness_workbook(output, tables)
            wb = load_workbook(output, read_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    [
                        "Executive Summary",
                        "Project Summary",
                        "Span Buckets",
                        "Span Readiness",
                        "Tower Gap",
                        "Coverage",
                        "Assumptions",
                    ],
                )
                span_bucket_headers = [cell for cell in next(wb["Span Buckets"].iter_rows(min_row=2, max_row=2, values_only=True))]
                self.assertIn("Already Strung km", span_bucket_headers)
                self.assertIn("Stringing Ready km", span_bucket_headers)
                self.assertIn("Tightening not started km", span_bucket_headers)
                self.assertIn("Tightening partial km", span_bucket_headers)
                project_headers = [cell for cell in next(wb["Project Summary"].iter_rows(min_row=2, max_row=2, values_only=True))]
                self.assertIn("already_strung_km", project_headers)
                self.assertIn("tightening_not_started_km", project_headers)
                self.assertIn("tightening_partial_km", project_headers)
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
