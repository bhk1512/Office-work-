from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from dashboard import progress_status_ingest as status_ingest


class ProgressStatusIngestTests(unittest.TestCase):
    def test_compile_excludes_completed_projects(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            files = [
                root / "TA 419 - DPR - 2026-04-29.xlsx",
                root / "TA 505 - DPR - 2026-04-29.xlsx",
            ]
            for path in files:
                df = pd.DataFrame(
                    [
                        ["Activity", "Cumulative Progress", "Plan for Month", "Progress for Month", "Quantity"],
                        ["Foundation", 47, 10, 7, 100],
                    ]
                )
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name="Status", header=False, index=False)

            output_path = root / "ProgressStatus_Output.xlsx"
            compiled = status_ingest.compile_progress_status_to_workbook(
                root,
                None,
                output_path,
                completed_project_keys={"ta419"},
            )  # type: ignore[attr-defined]
            self.assertIsNotNone(compiled)
            raw_df = pd.read_excel(output_path, sheet_name="RawData")
            self.assertTrue(raw_df["project_code"].astype(str).str.contains("TA 505", case=False, na=False).any())
            self.assertFalse(raw_df["project_code"].astype(str).str.contains("TA 419", case=False, na=False).any())

    def test_compile_includes_report_date_and_backfills_snapshots(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            files = [
                root / "TA 419 - DPR - 2026-03-29.xlsx",
                root / "TA 419 - DPR - 2026-04-29.xlsx",
            ]
            for path, progress_value in zip(files, (3, 7)):
                df = pd.DataFrame(
                    [
                        ["Activity", "Cumulative Progress", "Plan for Month", "Progress for Month", "Quantity"],
                        ["Foundation", 40 + progress_value, 10, progress_value, 100],
                    ]
                )
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name="Status", header=False, index=False)

            output_path = root / "ProgressStatus_Output.xlsx"
            compiled = status_ingest.compile_progress_status_to_workbook(root, None, output_path)  # type: ignore[attr-defined]
            self.assertIsNotNone(compiled)
            self.assertTrue(output_path.exists())

            raw_df = pd.read_excel(output_path, sheet_name="RawData")
            self.assertIn("report_date", raw_df.columns)
            report_dates = pd.to_datetime(raw_df["report_date"], errors="coerce").dropna()
            self.assertGreaterEqual(len(report_dates), 1)
            self.assertGreaterEqual(report_dates.dt.to_period("M").nunique(), 1)

    def test_filename_date_remains_authoritative_with_internal_date_diagnostics(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            path = root / "TA 419 - DPR - 2026-04-29.xlsx"
            df = pd.DataFrame(
                [
                    ["Date:", "2026-04-15", "", "", ""],
                    ["Activity", "Cumulative Progress", "Plan for Month", "Progress for Month", "Quantity"],
                    ["Foundation", 47, 10, 7, 100],
                ]
            )
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Status", header=False, index=False)

            output_path = root / "ProgressStatus_Output.xlsx"
            compiled = status_ingest.compile_progress_status_to_workbook(root, None, output_path)  # type: ignore[attr-defined]
            self.assertIsNotNone(compiled)

            raw_df = pd.read_excel(output_path, sheet_name="RawData")
            self.assertTrue((raw_df["report_date"].astype(str) == "2026-04-29").any())

            diagnostics_df = pd.read_excel(output_path, sheet_name="Diagnostics")
            self.assertIn("DateQuality", diagnostics_df.columns)
            self.assertIn("InternalDate", diagnostics_df.columns)
            self.assertIn("FilenameDate", diagnostics_df.columns)
            self.assertTrue((diagnostics_df["DateQuality"].astype(str) == "MISMATCH").any())

            coverage_df = pd.read_excel(output_path, sheet_name="Coverage")
            self.assertIn("date_quality", coverage_df.columns)
            self.assertTrue((coverage_df["date_quality"].astype(str) == "MISMATCH").any())

    def test_extract_guardrails(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        ws["A1"] = "Guardrails"
        ws["A2"] = "block_anchor"
        ws["B2"] = "Progress Status"
        ws["A3"] = "header_rows"
        ws["B3"] = "2"
        ws["A4"] = "To Map"
        ws["A5"] = "activity_raw"

        parsed = status_ingest._extract_guardrails(ws)  # type: ignore[attr-defined]
        self.assertEqual(parsed.get("block_anchor"), "Progress Status")
        self.assertEqual(parsed.get("header_rows"), "2")

    def test_detect_header_single_two_three_row(self) -> None:
        single = pd.DataFrame(
            [
                ["Sr", "Activity", "Cumulative Progress", "Balance Progress"],
                ["1", "Foundation", "120", "10"],
            ]
        )
        hrow, hspan, _ = status_ingest._detect_header(  # type: ignore[attr-defined]
            single,
            start_row=0,
            end_row=len(single.index) - 1,
            required_tokens=["activity", "cumulative", "balance"],
            header_rows_options=[1],
        )
        self.assertEqual((hrow, hspan), (0, 1))

        two_row = pd.DataFrame(
            [
                ["", "Activity", "Quantity", "", "Cumulative till Last Month", "", "Plan for the Month", "Progress for the Month", "Cumulative Progress", "Balance Progress"],
                ["", "", "LOA", "Estimated", "L2", "Progress", "L2", "Till Date", "", ""],
                ["", "Foundation", "322", "330", "322", "228", "40", "4", "232", "98"],
            ]
        )
        hrow, hspan, _ = status_ingest._detect_header(  # type: ignore[attr-defined]
            two_row,
            start_row=0,
            end_row=len(two_row.index) - 1,
            required_tokens=["activity", "cumulative", "balance"],
            header_rows_options=[2],
        )
        self.assertEqual((hrow, hspan), (0, 2))

        three_row = pd.DataFrame(
            [
                ["", "PARTICULARS", "UOM", "Total", "Comp", "Balance", "April'26", "", "WIP", "Cumulative"],
                ["", "", "", "Qty", "Till", "as on", "Plan", "Actual", "", ""],
                ["", "", "", "", "Mar'26", "1 Apr'26", "", "", "", ""],
                ["", "Foundation", "Nos", "190", "137", "53", "6", "2", "5", "137"],
            ]
        )
        hrow, hspan, _ = status_ingest._detect_header(  # type: ignore[attr-defined]
            three_row,
            start_row=0,
            end_row=len(three_row.index) - 1,
            required_tokens=["particulars", "comp", "balance", "cumulative"],
            header_rows_options=[3],
        )
        self.assertEqual((hrow, hspan), (0, 3))

    def test_parse_with_section_split(self) -> None:
        df = pd.DataFrame(
            [
                ["Progress of 220kV Line", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                ["Sr. No.", "Activity", "UoM", "Total Quantity", "As per L2", "", "Cumm. Progress up to last Month", "", "Progress of Current Month (April'26)", "", "", "Cumulative till current month", "", "", "Resources Status", "", "", "Critical Issues"],
                ["", "", "", "", "Start Date", "End Date", "L2", "Actual", "Plan", "Today", "Progress in Month", "Target (L2)", "Completed", "Balance", "Targeted Gang", "Available Gang", "Shortfall", ""],
                ["1", "Foundation", "Nos.", "199", "", "", "199", "195", "4", "1", "2", "199", "197", "2", "1", "1", "0", ""],
                ["5", "Tower Erection", "Nos.", "199", "", "", "199", "195", "4", "0", "0", "199", "195", "4", "0", "0", "0", ""],
                ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                ["Progress of 132kV Line", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
                ["Sr. No.", "Activity", "UoM", "Total Quantity", "As per L2", "", "Cumm. Progress up to last Month", "", "Progress of Current Month (April'26)", "", "", "Cumulative till current month", "", "", "Resources Status", "", "", "Critical Issues"],
                ["", "", "", "", "Start Date", "End Date", "L2", "Actual", "Plan", "Today", "Progress in Month", "Target (L2)", "Completed", "Balance", "Targeted Gang", "Available Gang", "Shortfall", ""],
                ["1", "Foundation", "Nos.", "39", "", "", "39", "31", "5", "1", "3", "39", "34", "5", "2", "2", "0", ""],
                ["7", "Stringing", "Kms", "9", "", "", "9", "5", "1.4", "0", "0.3", "9", "5.3", "3.7", "0", "0", "0", ""],
            ]
        )

        guardrails = {
            "section_start_contains": "Progress of",
            "section_label_regex": r"Progress of\\s*(.*?)\\s*$",
            "header_rows": "2",
            "required_tokens": "activity; cumm; progress; balance",
            "activity_allowlist": "foundation; tower erection; stringing",
        }
        template_map = {
            1: "activity_raw",
            3: "quantity_estimated_or_total",
            7: "cumulative_last_month",
            8: "plan_for_month",
            9: "today_progress",
            10: "progress_for_month",
            12: "cumulative_progress",
            13: "balance_progress",
            15: "gangs_working",
            17: "remarks",
        }
        result = status_ingest._parse_status_sheet_dataframe(  # type: ignore[attr-defined]
            df,
            guardrails=guardrails,
            template_map=template_map,
            project_code="TB 501",
            project_display="TB 501",
            project_scope_key="tb501",
            line_name="",
            line_name_source="",
            report_date="2026-04-29",
            source_file="TB 501 - DPR.xlsx",
            source_sheet="Summary",
            configured_sheet="Summary",
            template_sheet="TB 501 Status",
        )
        self.assertEqual(result.parse_status, "OK")
        self.assertGreaterEqual(result.rows_emitted, 4)
        self.assertGreaterEqual(result.sections_detected, 2)


if __name__ == "__main__":
    unittest.main()
