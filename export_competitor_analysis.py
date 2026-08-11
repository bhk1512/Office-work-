#!/usr/bin/env python3
"""Build competitor-vs-internal productivity scenario workbook."""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
import calendar
import math
import re

import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from dashboard.config import IDLE_MAX_GAP_DAYS, IDLE_OFF_SYSTEM_GAP_DAYS
from dashboard.idle_utils import compute_intervals_for_dates, split_interval_by_month


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_COMPETITOR_INPUT = Path(r"C:\Users\kaushikb\OneDrive - RPG Enterprises\Competitor Analysis.xlsx")
WORKING_COMPETITOR_INPUT = BASE_DIR / "Productivity Summaries" / "Competitor Analysis - working copy.xlsx"
DEFAULT_OUTPUT = BASE_DIR / "Productivity Summaries" / "Competitor_Performance_Comparison_May_June_2026.xlsx"

MONTHS = {
    "May'26": pd.Timestamp("2026-05-01"),
    "Jun'26": pd.Timestamp("2026-06-01"),
}
ACTIVITIES = {
    "FDN": {"name": "Foundation", "unit": "Foundations"},
    "ERE": {"name": "Erection", "unit": "MT"},
    "STR": {"name": "Stringing", "unit": "KM"},
}
SCENARIOS = [
    ("Idle -20%", 0.80),
    ("Idle -10%", 0.90),
    ("Idle Base", 1.00),
    ("Idle +10%", 1.10),
    ("Idle +20%", 1.20),
]


def norm_code(value: object) -> str:
    text = "" if pd.isna(value) else str(value).strip().upper()
    return re.sub(r"[^A-Z0-9]", "", text)


def display_code(value: object) -> str:
    compact = norm_code(value)
    match = re.match(r"^([A-Z]+)(\d+.*)$", compact)
    return f"{match.group(1)} {match.group(2)}" if match else str(value).strip()


def split_codes(value: object) -> list[str]:
    parts = re.split(r"[/,;&]+", "" if pd.isna(value) else str(value))
    return [norm_code(part) for part in parts if norm_code(part)]


def clean_number(value: object) -> float | None:
    number = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return None if pd.isna(number) else float(number)


def month_end(month_start: pd.Timestamp) -> pd.Timestamp:
    return month_start + pd.offsets.MonthEnd(1)


def days_in_month(month_start: pd.Timestamp) -> int:
    return calendar.monthrange(int(month_start.year), int(month_start.month))[1]


def date_range_days(start: pd.Timestamp, end: pd.Timestamp) -> list[pd.Timestamp]:
    if pd.isna(start) or pd.isna(end) or end < start:
        return []
    return list(pd.date_range(start.normalize(), end.normalize(), freq="D"))


def read_competitor(path: Path) -> pd.DataFrame:
    source = path if path.exists() else WORKING_COMPETITOR_INPUT
    try:
        frame = pd.read_excel(source, header=[0, 1])
    except PermissionError:
        if source != WORKING_COMPETITOR_INPUT and WORKING_COMPETITOR_INPUT.exists():
            source = WORKING_COMPETITOR_INPUT
            frame = pd.read_excel(source, header=[0, 1])
        else:
            raise
    rows: list[dict[str, object]] = []
    for _, row in frame.iterrows():
        sl_no = row.get(("Sl.No.", "Unnamed: 0_level_1"))
        if pd.isna(sl_no):
            continue
        project = row.get(("Project", "Unnamed: 1_level_1"))
        agency = row.get(("Agency", "Unnamed: 2_level_1"))
        code_raw = row.get(("Internal Project Code", "Unnamed: 9_level_1"))
        for month_label, month_start in MONTHS.items():
            for activity_code in ACTIVITIES:
                qty = clean_number(row.get((month_label, activity_code)))
                rows.append(
                    {
                        "Sl.No.": int(sl_no) if float(sl_no).is_integer() else sl_no,
                        "Competitor Project": project,
                        "Agency": agency,
                        "Internal Project Code Raw": code_raw,
                        "Internal Codes Parsed": "/".join(split_codes(code_raw)),
                        "Month": month_start.strftime("%Y-%m"),
                        "Month Start": month_start,
                        "Activity Code": activity_code,
                        "Activity": ACTIVITIES[activity_code]["name"],
                        "Competitor Quantity": 0.0 if qty is None else qty,
                    }
                )
    return pd.DataFrame(rows)


def load_project_presence() -> set[str]:
    codes: set[str] = set()
    candidates = [
        BASE_DIR / "Parquets" / "Erection" / "ProjectDetails.parquet",
        BASE_DIR / "Parquets" / "Foundation" / "FoundationCompletions.parquet",
        BASE_DIR / "Parquets" / "Erection" / "RawData.parquet",
        BASE_DIR / "Parquets" / "Stringing" / "StringingDaily.parquet",
    ]
    for path in candidates:
        if not path.exists():
            continue
        df = pd.read_parquet(path)
        for col in ("project_code", "Project Code", "project"):
            if col in df.columns:
                codes.update(df[col].dropna().map(norm_code).loc[lambda s: s.astype(bool)].unique())
    return codes


def idle_by_code_month(work_dates: pd.DataFrame, activity: str) -> pd.DataFrame:
    columns = [
        "Activity",
        "Internal Code",
        "Month",
        "Work Date Count",
        "Idle Days Capped",
        "Idle Window Count",
    ]
    if work_dates.empty:
        return pd.DataFrame(columns=columns)

    rows: list[dict[str, object]] = []
    for code, group in work_dates.groupby("Internal Code", dropna=False):
        dates = (
            pd.to_datetime(group["Work Date"], errors="coerce")
            .dropna()
            .dt.normalize()
            .drop_duplicates()
            .sort_values()
            .tolist()
        )
        work_counts = (
            pd.Series(dates)
            .dt.to_period("M")
            .dt.to_timestamp()
            .value_counts()
            .to_dict()
            if dates
            else {}
        )
        idle_days: defaultdict[pd.Timestamp, float] = defaultdict(float)
        idle_windows: defaultdict[pd.Timestamp, int] = defaultdict(int)
        intervals = compute_intervals_for_dates([d.date() for d in dates], skip_off_system=True)
        for interval in intervals:
            if interval.get("skipped"):
                continue
            for part in split_interval_by_month(
                interval["interval_start"],
                interval["interval_end"],
                int(interval["capped_gap_days"]),
                int(interval["raw_gap_days"]),
            ):
                month = pd.Timestamp(year=int(part["year"]), month=int(part["month"]), day=1)
                idle_days[month] += float(part["allocated_capped_days"])
                idle_windows[month] += 1

        for month in sorted(set(work_counts) | set(idle_days)):
            rows.append(
                {
                    "Activity": activity,
                    "Internal Code": code,
                    "Month": month.strftime("%Y-%m"),
                    "Work Date Count": int(work_counts.get(month, 0)),
                    "Idle Days Capped": round(float(idle_days.get(month, 0.0)), 3),
                    "Idle Window Count": int(idle_windows.get(month, 0)),
                }
            )
    return pd.DataFrame(rows, columns=columns)


def build_foundation() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    path = BASE_DIR / "Parquets" / "Foundation" / "FoundationCompletions.parquet"
    df = pd.read_parquet(path)
    df["Internal Code"] = df["project_code"].map(norm_code)
    df["event_date"] = pd.to_datetime(df["event_date"], errors="coerce").dt.normalize()
    df["start_date"] = pd.to_datetime(df["start_date"], errors="coerce").dt.normalize()
    df["Month"] = df["event_date"].dt.to_period("M").dt.to_timestamp()
    df["source_type_norm"] = df["source_type"].fillna("").astype(str).str.strip().str.lower()

    detail = df[df["source_type_norm"].eq("detail") & df["event_date"].notna()].copy()
    detail_summary = (
        detail.groupby(["Internal Code", "Month"], dropna=False)
        .agg(
            Detail_Count=("event_date", "size"),
            Valid_Duration_Count=("start_date", lambda s: int(s.notna().sum())),
            Start_Date_Missing_Count=("start_date", lambda s: int(s.isna().sum())),
        )
        .reset_index()
    )

    snapshots = df[df["source_type_norm"].eq("snapshot_fallback")].copy()
    snapshot_rows: list[dict[str, object]] = []
    for code, group in snapshots.groupby("Internal Code", dropna=False):
        group = group.sort_values("event_date")
        for _, month in MONTHS.items():
            start = month
            end = month_end(month)
            before = group[group["event_date"].lt(start)]
            upto = group[group["event_date"].le(end)]
            if upto.empty:
                continue
            end_val = pd.to_numeric(upto["cumulative_foundation"], errors="coerce").dropna()
            if end_val.empty:
                continue
            prev_val = pd.to_numeric(before["cumulative_foundation"], errors="coerce").dropna()
            delta = pd.NA if prev_val.empty else float(end_val.iloc[-1] - prev_val.iloc[-1])
            snapshot_rows.append(
                {
                    "Internal Code": code,
                    "Month": month,
                    "Snapshot_Delta": delta,
                    "Snapshot_Count": int(len(group[(group["event_date"].ge(start)) & (group["event_date"].le(end))].index)),
                }
            )
    snapshot_summary = pd.DataFrame(snapshot_rows)

    summary = pd.merge(detail_summary, snapshot_summary, on=["Internal Code", "Month"], how="outer")
    if summary.empty:
        summary = pd.DataFrame(columns=["Internal Code", "Month"])
    for col in ("Detail_Count", "Valid_Duration_Count", "Start_Date_Missing_Count", "Snapshot_Count"):
        if col not in summary.columns:
            summary[col] = 0
        summary[col] = pd.to_numeric(summary[col], errors="coerce").fillna(0).astype(int)
    if "Snapshot_Delta" not in summary.columns:
        summary["Snapshot_Delta"] = pd.NA
    summary["Month"] = pd.to_datetime(summary["Month"], errors="coerce")
    summary["Internal Quantity"] = summary["Detail_Count"].astype(float)
    use_snapshot = summary["Detail_Count"].eq(0) & pd.to_numeric(summary["Snapshot_Delta"], errors="coerce").notna()
    summary.loc[use_snapshot, "Internal Quantity"] = pd.to_numeric(summary.loc[use_snapshot, "Snapshot_Delta"], errors="coerce")
    summary["Internal Unit"] = "Foundations"
    summary["Foundation Source Used"] = "detail rows"
    summary.loc[use_snapshot, "Foundation Source Used"] = "snapshot cumulative delta"
    summary["Data Quality Note"] = ""
    summary.loc[summary["Start_Date_Missing_Count"].gt(0), "Data Quality Note"] = (
        "Some foundation rows have missing start dates; idle only uses valid start/completion rows."
    )
    summary.loc[use_snapshot, "Data Quality Note"] = (
        "Foundation progress came from snapshot cumulative delta; exact start/completion idle unavailable."
    )

    work_rows: list[dict[str, object]] = []
    valid = detail[detail["start_date"].notna() & detail["event_date"].notna() & detail["event_date"].ge(detail["start_date"])].copy()
    for _, row in valid.iterrows():
        for day in date_range_days(row["start_date"], row["event_date"]):
            work_rows.append({"Internal Code": row["Internal Code"], "Work Date": day})
    work_dates = pd.DataFrame(work_rows, columns=["Internal Code", "Work Date"])
    idle = idle_by_code_month(work_dates, "Foundation")
    return summary, idle, work_dates


def build_erection() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    raw = pd.read_parquet(BASE_DIR / "Parquets" / "Erection" / "RawData.parquet")
    raw["Internal Code"] = raw["Project Code"].map(norm_code)
    raw["Complete Date"] = pd.to_datetime(raw["Complete Date"], errors="coerce").dt.normalize()
    raw["Month"] = raw["Complete Date"].dt.to_period("M").dt.to_timestamp()
    raw["Tower Weight"] = pd.to_numeric(raw["Tower Weight"], errors="coerce")
    completed = raw[raw["Complete Date"].notna()].copy()
    summary = (
        completed.groupby(["Internal Code", "Month"], dropna=False)
        .agg(
            Internal_Tower_Count=("Complete Date", "size"),
            Internal_Quantity=("Tower Weight", "sum"),
            Avg_Tower_Weight_MT=("Tower Weight", "mean"),
            Tower_Weight_Rows=("Tower Weight", lambda s: int(s.notna().sum())),
        )
        .reset_index()
    )
    summary["Internal Unit"] = "MT"
    summary["Data Quality Note"] = ""
    missing_weight = summary["Tower_Weight_Rows"].lt(summary["Internal_Tower_Count"])
    summary.loc[missing_weight, "Data Quality Note"] = "Some completed tower rows have missing tower weight."

    daily = pd.read_parquet(BASE_DIR / "Parquets" / "Erection" / "ProdDailyExpandedSingles.parquet")
    daily["Internal Code"] = daily["Project Code"].map(norm_code)
    daily["Work Date"] = pd.to_datetime(daily["Work Date"], errors="coerce").dt.normalize()
    work_dates = daily[daily["Work Date"].notna()][["Internal Code", "Work Date"]].drop_duplicates()
    idle = idle_by_code_month(work_dates, "Erection")
    return summary, idle, work_dates


def build_stringing() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = pd.read_parquet(BASE_DIR / "Parquets" / "Stringing" / "StringingDaily.parquet")
    df["Internal Code"] = df["project"].map(norm_code)
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
    df["Month"] = df["date"].dt.to_period("M").dt.to_timestamp()
    df["daily_km"] = pd.to_numeric(df["daily_km"], errors="coerce")
    positive = df[df["date"].notna() & df["daily_km"].notna()].copy()
    summary = (
        positive.groupby(["Internal Code", "Month"], dropna=False)
        .agg(
            Internal_Quantity=("daily_km", "sum"),
            Internal_Daily_Row_Count=("daily_km", "size"),
            Internal_Work_Date_Count=("date", lambda s: int(s.dropna().nunique())),
        )
        .reset_index()
    )
    summary["Internal Unit"] = "KM"
    summary["Data Quality Note"] = ""
    work_dates = positive[["Internal Code", "date"]].rename(columns={"date": "Work Date"}).drop_duplicates()
    idle = idle_by_code_month(work_dates, "Stringing")
    return summary, idle, work_dates


@dataclass
class Sources:
    project_presence: set[str]
    summaries: dict[str, pd.DataFrame]
    idles: dict[str, pd.DataFrame]


def lookup_code_metric(sources: Sources, activity: str, code: str, month: pd.Timestamp) -> dict[str, object]:
    month_text = month.strftime("%Y-%m")
    summary = sources.summaries[activity]
    idle = sources.idles[activity]
    row = summary[
        summary["Internal Code"].eq(code)
        & pd.to_datetime(summary["Month"], errors="coerce").dt.to_period("M").astype(str).eq(month_text)
    ]
    idle_row = idle[idle["Internal Code"].eq(code) & idle["Month"].eq(month_text)]

    exists = code in sources.project_presence
    out: dict[str, object] = {
        "Activity": activity,
        "Internal Code": code,
        "Month": month_text,
        "Internal Project Exists": bool(exists),
        "Internal Quantity": 0.0 if exists else pd.NA,
        "Internal Unit": ACTIVITIES[activity]["unit"],
        "Internal Tower Count": pd.NA,
        "Avg Tower Weight MT": pd.NA,
        "Idle Days Capped": pd.NA,
        "Idle Window Count": pd.NA,
        "Work Date Count": pd.NA,
        "Data Quality Note": "",
    }

    if not row.empty:
        r = row.iloc[0]
        out["Internal Quantity"] = float(r.get("Internal Quantity", r.get("Internal_Quantity", 0.0)) or 0.0)
        out["Internal Unit"] = r.get("Internal Unit", ACTIVITIES[activity]["unit"])
        out["Internal Tower Count"] = r.get("Internal_Tower_Count", pd.NA)
        out["Avg Tower Weight MT"] = r.get("Avg_Tower_Weight_MT", pd.NA)
        out["Data Quality Note"] = r.get("Data Quality Note", "")
        if activity == "FDN":
            out["Foundation Source Used"] = r.get("Foundation Source Used", "")
            out["Foundation Detail Count"] = r.get("Detail_Count", pd.NA)
            out["Foundation Snapshot Delta"] = r.get("Snapshot_Delta", pd.NA)
            out["Foundation Valid Duration Count"] = r.get("Valid_Duration_Count", pd.NA)
    elif exists:
        out["Data Quality Note"] = "Project exists internally, but this activity has no completed/progress rows in this month."
    else:
        out["Data Quality Note"] = "Internal project code was not found in extracted foundation/erection/stringing/project-details data."

    if not idle_row.empty:
        ir = idle_row.iloc[0]
        out["Idle Days Capped"] = float(ir.get("Idle Days Capped", 0.0) or 0.0)
        out["Idle Window Count"] = int(ir.get("Idle Window Count", 0) or 0)
        out["Work Date Count"] = int(ir.get("Work Date Count", 0) or 0)

    return out


def build_analysis(comp: pd.DataFrame, sources: Sources) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    code_rows: list[dict[str, object]] = []
    baseline_rows: list[dict[str, object]] = []
    scenario_rows: list[dict[str, object]] = []

    for _, comp_row in comp.iterrows():
        activity = str(comp_row["Activity Code"])
        month = pd.Timestamp(comp_row["Month Start"])
        codes = split_codes(comp_row["Internal Project Code Raw"])
        month_days = days_in_month(month)
        code_metrics = [lookup_code_metric(sources, activity, code, month) for code in codes]
        for metric in code_metrics:
            code_rows.append(
                {
                    **comp_row.drop(labels=["Month Start"]).to_dict(),
                    **metric,
                }
            )

        found_metrics = [m for m in code_metrics if bool(m.get("Internal Project Exists"))]
        qty_values = [float(m["Internal Quantity"]) for m in found_metrics if pd.notna(m.get("Internal Quantity"))]
        internal_qty = sum(qty_values) if qty_values else pd.NA
        idle_values = [float(m["Idle Days Capped"]) for m in found_metrics if pd.notna(m.get("Idle Days Capped"))]
        base_idle = sum(idle_values) / len(idle_values) if idle_values else pd.NA
        internal_work_dates = sum(int(m["Work Date Count"]) for m in found_metrics if pd.notna(m.get("Work Date Count")))

        avg_tower_weight = pd.NA
        if activity == "ERE":
            weights = []
            for m in found_metrics:
                tower_count = pd.to_numeric(pd.Series([m.get("Internal Tower Count")]), errors="coerce").iloc[0]
                avg_weight = pd.to_numeric(pd.Series([m.get("Avg Tower Weight MT")]), errors="coerce").iloc[0]
                if pd.notna(tower_count) and pd.notna(avg_weight) and tower_count > 0:
                    weights.append((float(tower_count), float(avg_weight)))
            if weights:
                avg_tower_weight = sum(count * weight for count, weight in weights) / sum(count for count, _ in weights)

        competitor_qty = float(comp_row["Competitor Quantity"])
        competitor_output = competitor_qty
        output_unit = ACTIVITIES[activity]["unit"]
        if activity == "ERE":
            competitor_output = competitor_qty * float(avg_tower_weight) if pd.notna(avg_tower_weight) else pd.NA
            output_unit = "MT"

        missing_codes = [code for code in codes if code not in sources.project_presence]
        notes = []
        if missing_codes:
            notes.append("Missing internal code(s): " + ", ".join(missing_codes))
        if activity == "ERE" and pd.isna(avg_tower_weight) and competitor_qty:
            notes.append("No same-month internal tower weight available; competitor equivalent MT not calculated.")
        if pd.isna(base_idle):
            notes.append("No same-month internal work-date idle baseline available.")
        if not found_metrics:
            notes.append("No internal comparator data found.")

        internal_productive_days = month_days - float(base_idle) if pd.notna(base_idle) else pd.NA
        internal_idle_adj_prod = (
            float(internal_qty) / internal_productive_days
            if pd.notna(internal_qty) and pd.notna(internal_productive_days) and internal_productive_days > 0
            else pd.NA
        )
        internal_calendar_prod = float(internal_qty) / month_days if pd.notna(internal_qty) else pd.NA

        baseline = {
            **comp_row.drop(labels=["Month Start"]).to_dict(),
            "Month Days": month_days,
            "Internal Codes Count": len(codes),
            "Internal Codes Found": len(found_metrics),
            "Internal Quantity": internal_qty,
            "Internal Unit": output_unit,
            "Internal Work Date Count": internal_work_dates if found_metrics else pd.NA,
            "Internal Avg Idle Days Capped": base_idle,
            "Internal Productive Days Proxy": internal_productive_days,
            "Internal Calendar Productivity": internal_calendar_prod,
            "Internal Idle Adjusted Productivity": internal_idle_adj_prod,
            "Same-Month Avg Tower Weight MT": avg_tower_weight,
            "Competitor Equivalent Output": competitor_output,
            "Output Unit": output_unit,
            "Baseline Notes": " | ".join(notes),
        }
        baseline_rows.append(baseline)

        for label, multiplier in SCENARIOS:
            scenario_idle = float(base_idle) * multiplier if pd.notna(base_idle) else pd.NA
            productive_days = month_days - scenario_idle if pd.notna(scenario_idle) else pd.NA
            comp_idle_prod = (
                float(competitor_output) / productive_days
                if pd.notna(competitor_output) and pd.notna(productive_days) and productive_days > 0
                else pd.NA
            )
            diff = (
                comp_idle_prod - internal_idle_adj_prod
                if pd.notna(comp_idle_prod) and pd.notna(internal_idle_adj_prod)
                else pd.NA
            )
            diff_pct = (
                diff / internal_idle_adj_prod
                if pd.notna(diff) and pd.notna(internal_idle_adj_prod) and internal_idle_adj_prod != 0
                else pd.NA
            )
            scenario_rows.append(
                {
                    **baseline,
                    "Scenario": label,
                    "Idle Multiplier": multiplier,
                    "Competitor Assumed Idle Days": scenario_idle,
                    "Competitor Productive Days Proxy": productive_days,
                    "Competitor Calendar Productivity": (
                        float(competitor_output) / month_days if pd.notna(competitor_output) else pd.NA
                    ),
                    "Competitor Idle Adjusted Productivity": comp_idle_prod,
                    "Competitor vs Internal Difference": diff,
                    "Competitor vs Internal Difference %": diff_pct,
                }
            )

    code_detail = pd.DataFrame(code_rows)
    baseline_detail = pd.DataFrame(baseline_rows)
    scenario_detail = pd.DataFrame(scenario_rows)
    base_case = scenario_detail[scenario_detail["Scenario"].eq("Idle Base")].copy()
    return code_detail, baseline_detail, scenario_detail, base_case


def build_summary(base_case: pd.DataFrame) -> pd.DataFrame:
    if base_case.empty:
        return pd.DataFrame()
    rows = []
    valid = base_case[pd.to_numeric(base_case["Competitor vs Internal Difference %"], errors="coerce").notna()].copy()
    for keys, group in valid.groupby(["Agency", "Activity"], dropna=False):
        agency, activity = keys
        diff_pct = pd.to_numeric(group["Competitor vs Internal Difference %"], errors="coerce")
        rows.append(
            {
                "Agency": agency,
                "Activity": activity,
                "Comparable Rows": int(len(group.index)),
                "Avg Competitor vs Internal %": float(diff_pct.mean()),
                "Median Competitor vs Internal %": float(diff_pct.median()),
                "Rows Competitor Above Internal": int((diff_pct > 0).sum()),
                "Rows Competitor Below Internal": int((diff_pct < 0).sum()),
            }
        )
    return pd.DataFrame(rows).sort_values(["Activity", "Agency"]).reset_index(drop=True)


def methodology() -> pd.DataFrame:
    rows = [
        ("Scope", "Competitor rows from Competitor Analysis.xlsx; May 2026 and June 2026 only; activities FDN, ERE, STR."),
        ("Internal source - Foundation", "Parquets/Foundation/FoundationCompletions.parquet. Detail rows are counted by completion month; snapshot cumulative deltas are used only when no detail rows exist for that code/month."),
        ("Internal source - Erection", "Parquets/Erection/RawData.parquet for completed tower count/MT/average tower weight by completion month; ProdDailyExpandedSingles.parquet for work-date idle baseline."),
        ("Internal source - Stringing", "Parquets/Stringing/StringingDaily.parquet; daily_km is summed by date month."),
        ("Tower weight", "Competitor erection tower count is converted to MT using the weighted same-month average tower weight from the listed internal project code(s). If no same-month internal tower weight exists, equivalent MT is blank."),
        ("Idle baseline", f"Idle uses exact internal work dates. Gaps between work dates are capped at {IDLE_MAX_GAP_DAYS} days and gaps over {IDLE_OFF_SYSTEM_GAP_DAYS} days are treated as off-system/excluded, matching the dashboard idle utilities."),
        ("Slash-separated internal codes", "Codes such as TA505/TA509 are treated as the pooled comparator named by the sheet. Code-level details are retained in Internal_Code_Detail."),
        ("Scenario cases", "Five cases vary the internal idle baseline applied to competitor productivity: -20%, -10%, base, +10%, +20% idle days."),
        ("Productivity", "Calendar productivity = output / calendar days in the month. Idle-adjusted productivity = output / (calendar days - assumed capped idle days)."),
        ("No invented values", "Missing internal project codes, missing same-month tower weights, and missing idle baselines are left blank and flagged in notes."),
    ]
    return pd.DataFrame(rows, columns=["Item", "Definition"])


def write_workbook(
    output: Path,
    comp: pd.DataFrame,
    code_detail: pd.DataFrame,
    baseline: pd.DataFrame,
    scenarios: pd.DataFrame,
    base_case: pd.DataFrame,
    agency_summary: pd.DataFrame,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    sheet_frames = {
        "Methodology": methodology(),
        "Competitor_Input_Long": comp.drop(columns=["Month Start"]),
        "Base_Case_Comparison": base_case,
        "Scenario_Detail": scenarios,
        "Internal_Baseline": baseline,
        "Internal_Code_Detail": code_detail,
        "Agency_Summary": agency_summary,
    }
    with pd.ExcelWriter(output, engine="openpyxl", datetime_format="yyyy-mm-dd", date_format="yyyy-mm-dd") as writer:
        methodology().to_excel(writer, sheet_name="Methodology", index=False)
        comp.drop(columns=["Month Start"]).to_excel(writer, sheet_name="Competitor_Input_Long", index=False)
        base_case.to_excel(writer, sheet_name="Base_Case_Comparison", index=False)
        scenarios.to_excel(writer, sheet_name="Scenario_Detail", index=False)
        baseline.to_excel(writer, sheet_name="Internal_Baseline", index=False)
        code_detail.to_excel(writer, sheet_name="Internal_Code_Detail", index=False)
        agency_summary.to_excel(writer, sheet_name="Agency_Summary", index=False)

        thin = Side(style="thin", color="A6A6A6")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for sheet_name, worksheet in writer.sheets.items():
            df = sheet_frames[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for col_idx, col_name in enumerate(df.columns, start=1):
                width = min(
                    max(12, int(df[col_name].astype(str).str.len().quantile(0.95)) + 2 if not df.empty else len(col_name) + 2),
                    55,
                )
                worksheet.column_dimensions[get_column_letter(col_idx)].width = width
                number_format = None
                if "%" in col_name:
                    number_format = "0.0%"
                elif any(token in col_name for token in ["Quantity", "Productivity", "Weight", "Days", "Output", "Difference", "Idle"]):
                    number_format = "0.00"
                elif "Count" in col_name or col_name == "Sl.No.":
                    number_format = "0"
                if number_format:
                    for row in range(2, worksheet.max_row + 1):
                        worksheet.cell(row=row, column=col_idx).number_format = number_format


def main() -> int:
    input_path = DEFAULT_COMPETITOR_INPUT
    if not input_path.exists() and WORKING_COMPETITOR_INPUT.exists():
        input_path = WORKING_COMPETITOR_INPUT

    comp = read_competitor(input_path)
    fdn_summary, fdn_idle, _ = build_foundation()
    ere_summary, ere_idle, _ = build_erection()
    str_summary, str_idle, _ = build_stringing()
    sources = Sources(
        project_presence=load_project_presence(),
        summaries={"FDN": fdn_summary, "ERE": ere_summary, "STR": str_summary},
        idles={"FDN": fdn_idle, "ERE": ere_idle, "STR": str_idle},
    )
    code_detail, baseline, scenarios, base_case = build_analysis(comp, sources)
    agency_summary = build_summary(base_case)
    write_workbook(DEFAULT_OUTPUT, comp, code_detail, baseline, scenarios, base_case, agency_summary)
    print(DEFAULT_OUTPUT)
    print(f"Competitor activity rows: {len(comp)}")
    print(f"Scenario rows: {len(scenarios)}")
    print(f"Base-case comparable rows: {pd.to_numeric(base_case['Competitor vs Internal Difference %'], errors='coerce').notna().sum()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
