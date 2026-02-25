#!/usr/bin/env python3
"""Standalone CLI to export the stringing productivity summary workbook."""
from __future__ import annotations

import argparse
import logging
import re
from dataclasses import replace
from pathlib import Path

import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from dashboard.config import AppConfig, configure_logging
from dashboard.plan_utils import compact_project_key, normalize_location
from dashboard.state import AppDataStore
from dashboard.stringing import (
    expand_stringing_to_daily,
    expand_stringing_to_daily_payout,
    expand_stringing_to_daily_fs,
    add_length_units,
    normalize_stringing_columns,
    _to_datetime_normalize,
)

LOGGER = logging.getLogger("stringing_export")

BASE_DIR = Path(__file__).resolve().parent
PRODUCTIVITY_ROOT = BASE_DIR / "Productivity Summaries"
DEFAULT_OUTPUT = PRODUCTIVITY_ROOT / "Stringing_Productivity_Summary.xlsx"
_MONTH_PRODUCTIVITY_FACTOR = 30.0
MONTHLY_AVG_LABEL = "Avg Productivity (KM/month)"
MONTHLY_GANGS_LABEL = "No. of Gangs"
MONTHLY_KM_LABEL = "KM Strung"


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate the stringing productivity summary Excel workbook."
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Target Excel file path (default: %(default)s).",
    )
    parser.add_argument(
        "--data-path",
        type=Path,
        help="Optional override for the erection dataset root (Parquets/Erection).",
    )
    parser.add_argument(
        "--stringing-data-path",
        type=Path,
        help="Optional override for the stringing dataset root (Parquets/Stringing).",
    )
    parser.add_argument(
        "--start-date",
        type=str,
        help="Optional YYYY-MM-DD start date to filter daily rows.",
    )
    parser.add_argument(
        "--end-date",
        type=str,
        help="Optional YYYY-MM-DD end date to filter daily rows.",
    )
    return parser.parse_args()


def _resolve_output_path(candidate: Path) -> Path:
    resolved = Path(candidate).expanduser()
    if not resolved.is_absolute():
        resolved = PRODUCTIVITY_ROOT / resolved
    resolved = resolved.resolve()
    resolved.parent.mkdir(parents=True, exist_ok=True)
    return resolved


def _parse_date(value: str | None) -> pd.Timestamp | None:
    if not value:
        return None
    try:
        return pd.Timestamp(value).normalize()
    except Exception as exc:
        raise SystemExit(f"Invalid date '{value}': {exc}") from exc


def _filter_by_date(df: pd.DataFrame, column: str, start: pd.Timestamp | None, end: pd.Timestamp | None) -> pd.DataFrame:
    if df is None or df.empty or column not in df.columns:
        return pd.DataFrame()
    working = df.copy()
    working[column] = pd.to_datetime(working[column], errors="coerce").dt.normalize()
    working = working.dropna(subset=[column])
    if start is not None:
        working = working[working[column] >= start]
    if end is not None:
        working = working[working[column] <= end]
    return working


def _parse_date_series(series: pd.Series) -> pd.Series:
    if series is None or series.empty:
        return pd.Series([], dtype="datetime64[ns]")
    parsed = series.map(_to_datetime_normalize)
    parsed = pd.to_datetime(parsed, errors="coerce")
    return parsed.dt.normalize()


_LOCATION_RE = re.compile(r"^\s*(\d+)([A-Za-z]+)?(?:\s*/\s*(\d+))?\s*$")
_AP_PREFIX_RE = re.compile(r"^\s*ap[\s\-_/]*", flags=re.IGNORECASE)
_GANTRY_RE = re.compile(r"\b(gantry|gty)\b", flags=re.IGNORECASE)


def _letter_rank(value: str) -> int:
    rank = 0
    for ch in value.upper():
        if "A" <= ch <= "Z":
            rank = (rank * 26) + (ord(ch) - ord("A") + 1)
    return rank


def _strip_ap_prefix(value: object) -> str:
    text = normalize_location(value)
    if not text:
        return ""
    return _AP_PREFIX_RE.sub("", text).strip()


def _is_gantry_label(value: object) -> bool:
    text = normalize_location(value)
    if not text:
        return False
    cleaned = re.sub(r"[^a-z]+", " ", text.lower()).strip()
    return bool(_GANTRY_RE.search(cleaned))


def _location_order_key(value: object) -> int | None:
    text = _strip_ap_prefix(value)
    if not text:
        return None
    match = _LOCATION_RE.match(text)
    if not match:
        return None
    main = int(match.group(1))
    letters = match.group(2) or ""
    sub = int(match.group(3) or 0)
    letter_rank = _letter_rank(letters) if letters else 0
    return (main * 1_000_000) + (letter_rank * 1_000) + sub


def _ensure_project_fields(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    working = df.copy()
    if "project_name" not in working.columns:
        working["project_name"] = working.get("project", "")
    if "project" not in working.columns:
        working["project"] = working.get("project_name", "")
    working["project_name"] = working["project_name"].fillna("").astype(str).str.strip()
    if "project_key_norm" not in working.columns:
        working["project_key_norm"] = working["project_name"].map(compact_project_key)
    working["project_key_norm"] = working["project_key_norm"].fillna("").astype(str)
    return working


def _add_span_key(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    working = _ensure_project_fields(df)
    for col in ("from_ap", "to_ap"):
        if col not in working.columns:
            working[col] = ""
    working["from_ap_norm"] = working["from_ap"].map(_strip_ap_prefix)
    working["to_ap_norm"] = working["to_ap"].map(_strip_ap_prefix)
    span_base = (
        working["project_key_norm"].fillna("")
        + "|"
        + working["from_ap_norm"].fillna("")
        + "|"
        + working["to_ap_norm"].fillna("")
    )
    fallback = working.get("row_id", pd.Series("", index=working.index))
    working["span_key"] = span_base.where(
        (working["from_ap_norm"].astype(bool) | working["to_ap_norm"].astype(bool)),
        fallback.astype(str),
    )
    return working


def _build_stage_summary(df: pd.DataFrame, stage_label: str) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(
            [
                {
                    "stage": stage_label,
                    "avg_productivity_km_month": 0.0,
                    "total_km": 0.0,
                    "active_days": 0,
                    "spans": 0,
                    "gangs": 0,
                    "projects": 0,
                    "start_date": pd.NaT,
                    "end_date": pd.NaT,
                }
            ]
        )
    working = _add_span_key(df)
    working["daily_km"] = pd.to_numeric(working.get("daily_km"), errors="coerce").fillna(0.0)
    dates = pd.to_datetime(working.get("date"), errors="coerce").dropna()
    avg_prod = round(float(working["daily_km"].mean() * _MONTH_PRODUCTIVITY_FACTOR), 4) if len(working) else 0.0
    return pd.DataFrame(
        [
            {
                "stage": stage_label,
                "avg_productivity_km_month": avg_prod,
                "total_km": round(float(working["daily_km"].sum()), 4),
                "active_days": int(dates.nunique()) if not dates.empty else 0,
                "spans": int(working["span_key"].nunique()),
                "gangs": int(working.get("gang_name", pd.Series([], dtype=object)).nunique()),
                "projects": int(working["project_key_norm"].nunique()),
                "start_date": dates.min() if not dates.empty else pd.NaT,
                "end_date": dates.max() if not dates.empty else pd.NaT,
            }
        ]
    )


def _build_gang_productivity(df: pd.DataFrame, stage_label: str) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(
            columns=[
                "stage",
                "gang_name",
                "avg_productivity_km_month",
                "total_km",
                "active_days",
                "spans",
                "projects",
            ]
        )
    working = _add_span_key(df)
    working["daily_km"] = pd.to_numeric(working.get("daily_km"), errors="coerce").fillna(0.0)
    working["date"] = pd.to_datetime(working.get("date"), errors="coerce")
    working["gang_name"] = working.get("gang_name", "").fillna("").astype(str).str.strip()
    working = working[working["gang_name"].astype(bool)]
    if working.empty:
        return pd.DataFrame(
            columns=[
                "stage",
                "gang_name",
                "avg_productivity_km_month",
                "total_km",
                "active_days",
                "spans",
                "projects",
            ]
        )
    grouped = (
        working.groupby("gang_name", dropna=False)
        .agg(
            avg_productivity_km_month=("daily_km", "mean"),
            total_km=("daily_km", "sum"),
            active_days=("date", lambda s: s.dropna().nunique()),
            spans=("span_key", "nunique"),
            projects=("project_key_norm", "nunique"),
        )
        .reset_index()
    )
    grouped["avg_productivity_km_month"] = (
        grouped["avg_productivity_km_month"].fillna(0.0) * _MONTH_PRODUCTIVITY_FACTOR
    ).round(4)
    grouped["total_km"] = grouped["total_km"].fillna(0.0).round(4)
    grouped["active_days"] = grouped["active_days"].fillna(0).astype(int)
    grouped["spans"] = grouped["spans"].fillna(0).astype(int)
    grouped["projects"] = grouped["projects"].fillna(0).astype(int)
    grouped.insert(0, "stage", stage_label)
    return grouped.sort_values(["stage", "gang_name"]).reset_index(drop=True)


def _build_project_meta_lookup(project_info: pd.DataFrame | None) -> dict[str, dict[str, str]]:
    if project_info is None or project_info.empty:
        return {}
    info = project_info.copy()
    project_code = info.get("project_code", pd.Series("", index=info.index)).fillna("").astype(str).str.strip()
    project_name = info.get("project_name", info.get("Project Name", pd.Series("", index=info.index)))
    project_name = project_name.fillna("").astype(str).str.strip()
    key_name = info.get("key_name", pd.Series("", index=info.index)).fillna("").astype(str).str.strip()
    pch = info.get("pch", pd.Series("", index=info.index)).fillna("").astype(str).str.strip()

    lookup: dict[str, dict[str, str]] = {}
    for code, name, key, pch_value in zip(project_code, project_name, key_name, pch):
        display = code or name or key
        entry = {
            "project_code": code,
            "project_display": display,
            "pch": pch_value,
        }
        for candidate in (code, name, key):
            compact = compact_project_key(candidate)
            if compact and compact not in lookup:
                lookup[compact] = entry
    return lookup


def _prepare_stringing_scope(daily_df: pd.DataFrame, project_info: pd.DataFrame | None) -> pd.DataFrame:
    if daily_df is None or daily_df.empty:
        return pd.DataFrame()
    working = _add_span_key(daily_df)
    working["date"] = pd.to_datetime(working.get("date"), errors="coerce").dt.normalize()
    working = working.dropna(subset=["date"])
    working["daily_km"] = pd.to_numeric(working.get("daily_km"), errors="coerce").fillna(0.0)
    working["gang_name"] = working.get("gang_name", "").fillna("").astype(str).str.strip()
    working["project_name"] = working.get("project_name", working.get("project", "")).fillna("").astype(str).str.strip()
    working["project_key_norm"] = working.get("project_key_norm", "").fillna("").astype(str)

    lookup = _build_project_meta_lookup(project_info)
    if lookup:
        def _lookup_meta(value: object) -> dict[str, str] | None:
            key = compact_project_key(value)
            return lookup.get(key)

        meta_from_key = working["project_key_norm"].map(_lookup_meta)
        meta_from_name = working["project_name"].map(_lookup_meta)
        meta_series = meta_from_key.where(meta_from_key.notna(), meta_from_name)
        working["project_display"] = meta_series.map(
            lambda rec: rec.get("project_display", "") if isinstance(rec, dict) else ""
        )
        working["project_display"] = working["project_display"].where(
            working["project_display"].astype(bool), working["project_name"]
        )
        working["pch_display"] = meta_series.map(
            lambda rec: rec.get("pch", "") if isinstance(rec, dict) else ""
        )
    else:
        working["project_display"] = working["project_name"]
        working["pch_display"] = ""

    working["pch_display"] = working["pch_display"].fillna("").astype(str).str.strip()
    working["pch_display"] = working["pch_display"].where(working["pch_display"].astype(bool), "Unassigned")
    return working


def _prepare_stringing_compiled_scope(compiled_df: pd.DataFrame, project_info: pd.DataFrame | None) -> pd.DataFrame:
    if compiled_df is None or compiled_df.empty:
        return pd.DataFrame()
    normalized, _ = normalize_stringing_columns(compiled_df)
    normalized = _ensure_project_fields(normalized)
    normalized, _ = add_length_units(normalized)
    normalized["fs_complete_date"] = _parse_date_series(normalized.get("fs_complete_date"))
    normalized["month"] = normalized["fs_complete_date"].dt.to_period("M").dt.to_timestamp()
    normalized["gang_name"] = normalized.get("gang_name", "").fillna("").astype(str).str.strip()
    normalized["project_name"] = normalized.get("project_name", normalized.get("project", "")).fillna("").astype(str).str.strip()
    normalized["project_key_norm"] = normalized.get("project_key_norm", "").fillna("").astype(str)

    lookup = _build_project_meta_lookup(project_info)
    if lookup:
        def _lookup_meta(value: object) -> dict[str, str] | None:
            key = compact_project_key(value)
            return lookup.get(key)

        meta_from_key = normalized["project_key_norm"].map(_lookup_meta)
        meta_from_name = normalized["project_name"].map(_lookup_meta)
        meta_series = meta_from_key.where(meta_from_key.notna(), meta_from_name)
        normalized["project_display"] = meta_series.map(
            lambda rec: rec.get("project_display", "") if isinstance(rec, dict) else ""
        )
        normalized["project_display"] = normalized["project_display"].where(
            normalized["project_display"].astype(bool), normalized["project_name"]
        )
        normalized["pch_display"] = meta_series.map(
            lambda rec: rec.get("pch", "") if isinstance(rec, dict) else ""
        )
    else:
        normalized["project_display"] = normalized["project_name"]
        normalized["pch_display"] = ""

    normalized["pch_display"] = normalized["pch_display"].fillna("").astype(str).str.strip()
    normalized["pch_display"] = normalized["pch_display"].where(normalized["pch_display"].astype(bool), "Unassigned")
    return normalized


def _summarize_scope(
    scope: pd.DataFrame,
    *,
    group_columns: list[str] | None = None,
    include_project_count: bool = True,
) -> pd.DataFrame:
    if scope is None or scope.empty:
        return pd.DataFrame()
    group_columns = group_columns or []
    if group_columns:
        grouped = scope.groupby(group_columns, dropna=False)
    else:
        scope = scope.copy()
        scope["_all"] = "All"
        grouped = scope.groupby("_all", dropna=False)
    summary = grouped.agg(
        avg_km_month=("daily_km", lambda s: float(s.mean() * _MONTH_PRODUCTIVITY_FACTOR)),
        total_km=("daily_km", "sum"),
        spans=("span_key", "nunique"),
        gangs=("gang_name", "nunique"),
        projects=("project_key_norm", "nunique"),
    ).reset_index()
    summary["avg_km_month"] = summary["avg_km_month"].fillna(0.0).round(4)
    summary["total_km"] = summary["total_km"].fillna(0.0).round(4)
    summary["spans"] = summary["spans"].fillna(0).astype(int)
    summary["gangs"] = summary["gangs"].fillna(0).astype(int)
    summary["projects"] = summary["projects"].fillna(0).astype(int)
    if not include_project_count and "projects" in summary.columns:
        summary = summary.drop(columns=["projects"])
    if not group_columns:
        summary = summary.drop(columns=["_all"], errors="ignore")
    return summary


def _build_scope_label(scope: pd.DataFrame, start: pd.Timestamp | None, end: pd.Timestamp | None) -> str:
    if scope is None or scope.empty:
        return "Scope: (empty)"
    start_label = start.strftime("%Y-%m-%d") if start is not None else ""
    end_label = end.strftime("%Y-%m-%d") if end is not None else ""
    project_count = int(scope["project_key_norm"].nunique()) if "project_key_norm" in scope.columns else 0
    gang_count = int(scope["gang_name"].nunique()) if "gang_name" in scope.columns else 0
    span_count = int(scope["span_key"].nunique()) if "span_key" in scope.columns else 0
    return (
        f"Scope: {start_label} to {end_label} | Projects: {project_count} | "
        f"Spans: {span_count} | Gangs: {gang_count}"
    )


def _append_totals_row(df: pd.DataFrame, *, avg_col: str, total_col: str, span_col: str) -> pd.DataFrame:
    if df.empty:
        return df
    totals = {column: "" for column in df.columns}
    totals[avg_col] = "Total"
    totals[total_col] = round(float(df[total_col].sum()), 4) if total_col in df.columns else 0.0
    totals[span_col] = int(df[span_col].sum()) if span_col in df.columns else 0
    return pd.concat([df, pd.DataFrame([totals])], ignore_index=True)


def _build_month_windows_from_series(
    series: pd.Series,
    start: pd.Timestamp | None,
    end: pd.Timestamp | None,
) -> tuple[list[pd.Timestamp], list[str]]:
    if series is None or series.empty:
        return [], []
    series = pd.to_datetime(series, errors="coerce").dropna()
    if series.empty:
        return [], []
    start_ts = start or series.min()
    end_ts = end or series.max()
    if pd.isna(start_ts) or pd.isna(end_ts):
        return [], []
    start_month = pd.Timestamp(start_ts).to_period("M").to_timestamp()
    end_month = pd.Timestamp(end_ts).to_period("M").to_timestamp()
    months = pd.period_range(start_month, end_month, freq="M").to_timestamp().tolist()
    labels = [month.strftime("%b'%y") for month in months]
    return months, labels


def _build_review_table_three(
    daily_scope: pd.DataFrame,
    compiled_scope: pd.DataFrame,
    months: list[pd.Timestamp],
    labels: list[str],
) -> pd.DataFrame:
    rows = {
        "Stringing productivity": "Overall KM/month (Avg)",
    }
    tse_row = {
        "Stringing productivity": "TSE KM/month (Avg)",
    }
    tse_total_row = {
        "Stringing productivity": "TSE - KM Total Strung",
    }
    gang_row = {
        "Stringing productivity": "No of Gangs engaged (Avg)",
    }
    tse_gang_row = {
        "Stringing productivity": "No of Gangs TSE Engaged",
    }
    total_row = {
        "Stringing productivity": "KM Total Strung",
    }
    if not months:
        return pd.DataFrame(
            [rows, tse_row, tse_total_row, gang_row, tse_gang_row, total_row],
            columns=["Stringing productivity", *labels],
        )
    daily_empty = daily_scope is None or daily_scope.empty
    if daily_empty:
        daily = pd.DataFrame()
    else:
        daily = daily_scope.copy()
        method_series = daily.get("method", daily.get("method_norm", pd.Series("", index=daily.index)))
        method_norm = method_series.map(_normalize_method)
        daily["method_group"] = method_norm.map(_method_group)
    if compiled_scope is None or compiled_scope.empty:
        compiled = pd.DataFrame()
    else:
        compiled = compiled_scope.copy()
        method_series = compiled.get("method", compiled.get("method_norm", pd.Series("", index=compiled.index)))
        method_norm = method_series.map(_normalize_method)
        compiled["method_group"] = method_norm.map(_method_group)
    for month_ts, label in zip(months, labels):
        month_start = pd.Timestamp(month_ts).normalize()
        month_end = (month_start + pd.offsets.MonthEnd(1)).normalize()
        if daily_empty:
            rows[label] = 0.0
            tse_row[label] = 0.0
            tse_total_row[label] = 0.0
            gang_row[label] = 0
            tse_gang_row[label] = 0
        else:
            month_scope = daily[(daily["date"] >= month_start) & (daily["date"] <= month_end)]
            if month_scope.empty:
                rows[label] = 0.0
                tse_row[label] = 0.0
                tse_total_row[label] = 0.0
                gang_row[label] = 0
                tse_gang_row[label] = 0
            else:
                avg_prod = float(month_scope["daily_km"].mean() * _MONTH_PRODUCTIVITY_FACTOR)
                rows[label] = round(avg_prod, 2)
                gang_row[label] = int(month_scope["gang_name"].nunique())
                tse_scope = month_scope[month_scope["method_group"] == "TSE"]
                if tse_scope.empty:
                    tse_row[label] = 0.0
                    tse_gang_row[label] = 0
                else:
                    tse_avg = float(tse_scope["daily_km"].mean() * _MONTH_PRODUCTIVITY_FACTOR)
                    tse_row[label] = round(tse_avg, 2)
                    tse_gang_row[label] = int(tse_scope["gang_name"].nunique())

        km_total = 0.0
        tse_km_total = 0.0
        if not compiled.empty:
            if "month" in compiled.columns:
                month_compiled = compiled[compiled["month"] == month_start]
            else:
                month_compiled = compiled[
                    (compiled["fs_complete_date"] >= month_start)
                    & (compiled["fs_complete_date"] <= month_end)
                ]
            km_total = float(pd.to_numeric(month_compiled["length_km"], errors="coerce").sum())
            month_compiled_tse = month_compiled[month_compiled["method_group"] == "TSE"]
            tse_km_total = float(
                pd.to_numeric(month_compiled_tse["length_km"], errors="coerce").sum()
            )
        total_row[label] = round(km_total, 2)
        tse_total_row[label] = round(tse_km_total, 2)
    return pd.DataFrame(
        [rows, tse_row, tse_total_row, gang_row, tse_gang_row, total_row],
        columns=["Stringing productivity", *labels],
    )


def _build_rollup_tables(
    scope: pd.DataFrame,
    scope_tse: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    overall_rollup = _summarize_scope(scope)
    overall_rollup = overall_rollup.rename(
        columns={
            "avg_km_month": "Overall Avg Productivity (KM/month)",
            "total_km": "Total KM",
            "spans": "Spans",
            "gangs": "Gangs",
            "projects": "Projects",
        }
    )
    if scope_tse is None or scope_tse.empty:
        overall_rollup["TSE Avg Productivity (KM/month)"] = 0.0
    else:
        tse_overall = _summarize_scope(scope_tse)
        overall_rollup["TSE Avg Productivity (KM/month)"] = round(
            float(tse_overall["avg_km_month"].iloc[0]) if not tse_overall.empty else 0.0,
            4,
        )
    overall_rollup = overall_rollup[
        [
            "Overall Avg Productivity (KM/month)",
            "TSE Avg Productivity (KM/month)",
            "Total KM",
            "Spans",
            "Gangs",
            "Projects",
        ]
    ]

    pch_rollup = _summarize_scope(scope, group_columns=["pch_display"])
    pch_rollup = pch_rollup.rename(
        columns={
            "pch_display": "PCH",
            "avg_km_month": "Overall Avg Productivity (KM/month)",
            "total_km": "Total KM",
            "spans": "Spans",
            "gangs": "Gangs",
            "projects": "Projects",
        }
    )
    if scope_tse is not None and not scope_tse.empty:
        tse_pch = _summarize_scope(scope_tse, group_columns=["pch_display"]).rename(
            columns={
                "pch_display": "PCH",
                "avg_km_month": "TSE Avg Productivity (KM/month)",
            }
        )
        pch_rollup = pch_rollup.merge(
            tse_pch[["PCH", "TSE Avg Productivity (KM/month)"]],
            on="PCH",
            how="left",
        )
    pch_rollup["TSE Avg Productivity (KM/month)"] = (
        pch_rollup.get("TSE Avg Productivity (KM/month)").fillna(0.0)
    )
    pch_rollup = pch_rollup[
        [
            "PCH",
            "Overall Avg Productivity (KM/month)",
            "TSE Avg Productivity (KM/month)",
            "Total KM",
            "Spans",
            "Gangs",
            "Projects",
        ]
    ]

    project_rollup = _summarize_scope(
        scope,
        group_columns=["pch_display", "project_display"],
        include_project_count=False,
    )
    project_rollup = project_rollup.rename(
        columns={
            "pch_display": "PCH",
            "project_display": "Project",
            "avg_km_month": "Overall Avg Productivity (KM/month)",
            "total_km": "Total KM",
            "spans": "Spans",
            "gangs": "Gangs",
        }
    )
    if scope_tse is not None and not scope_tse.empty:
        tse_project = _summarize_scope(
            scope_tse,
            group_columns=["pch_display", "project_display"],
            include_project_count=False,
        ).rename(
            columns={
                "pch_display": "PCH",
                "project_display": "Project",
                "avg_km_month": "TSE Avg Productivity (KM/month)",
            }
        )
        project_rollup = project_rollup.merge(
            tse_project[["PCH", "Project", "TSE Avg Productivity (KM/month)"]],
            on=["PCH", "Project"],
            how="left",
        )
    project_rollup["TSE Avg Productivity (KM/month)"] = (
        project_rollup.get("TSE Avg Productivity (KM/month)").fillna(0.0)
    )
    project_rollup = project_rollup[
        [
            "PCH",
            "Project",
            "Overall Avg Productivity (KM/month)",
            "TSE Avg Productivity (KM/month)",
            "Total KM",
            "Spans",
            "Gangs",
        ]
    ]

    return overall_rollup, pch_rollup, project_rollup


def _build_monthly_km_table(
    scope: pd.DataFrame,
    months: list[pd.Timestamp],
    labels: list[str],
    *,
    label_col: str = "Stringing productivity",
    date_column: str = "month",
    metric_column: str = "length_km",
) -> pd.DataFrame:
    rows = {
        label_col: "KM/month (Total)",
    }
    gang_row = {
        label_col: "No of Gangs",
    }
    if scope is None or scope.empty or not months:
        return pd.DataFrame([rows, gang_row], columns=[label_col, *labels])
    for month_ts, label in zip(months, labels):
        month_start = pd.Timestamp(month_ts).normalize()
        month_end = (month_start + pd.offsets.MonthEnd(1)).normalize()
        if date_column == "month":
            month_scope = scope[scope[date_column] == month_start]
        else:
            month_scope = scope[(scope[date_column] >= month_start) & (scope[date_column] <= month_end)]
        if month_scope.empty:
            rows[label] = 0.0
            gang_row[label] = 0
            continue
        km_total = float(pd.to_numeric(month_scope[metric_column], errors="coerce").sum())
        rows[label] = round(km_total, 2)
        gang_row[label] = int(month_scope["gang_name"].nunique())
    return pd.DataFrame([rows, gang_row], columns=[label_col, *labels])


def _monthly_avg_column(month_label: str) -> str:
    return f"{month_label}__avg"


def _monthly_gangs_column(month_label: str) -> str:
    return f"{month_label}__gangs"


def _monthly_km_column(month_label: str) -> str:
    return f"{month_label}__km"


def _monthly_column_layout(
    month_labels: list[str],
    *,
    group_labels: list[str],
) -> tuple[list[str], list[str], list[str]]:
    ordered_columns = [*group_labels]
    top_headers: list[str] = []
    bottom_headers: list[str] = []

    for label in group_labels:
        top_headers.append("")
        bottom_headers.append(label)

    for month_label in month_labels:
        ordered_columns.extend(
            [
                _monthly_avg_column(month_label),
                _monthly_gangs_column(month_label),
                _monthly_km_column(month_label),
            ]
        )
        top_headers.extend([month_label, "", ""])
        bottom_headers.extend([MONTHLY_AVG_LABEL, MONTHLY_GANGS_LABEL, MONTHLY_KM_LABEL])

    ordered_columns.extend(
        [
            _monthly_avg_column("Overall"),
            _monthly_gangs_column("Overall"),
            _monthly_km_column("Overall"),
        ]
    )
    top_headers.extend(["Overall", "", ""])
    bottom_headers.extend([MONTHLY_AVG_LABEL, MONTHLY_GANGS_LABEL, MONTHLY_KM_LABEL])
    return ordered_columns, top_headers, bottom_headers


def _build_monthly_group_summary_table(
    daily_scope: pd.DataFrame,
    compiled_scope: pd.DataFrame,
    *,
    month_keys: list[pd.Timestamp],
    month_labels: list[str],
    group_columns: list[str],
    group_labels: list[str],
) -> pd.DataFrame:
    ordered_columns, _, _ = _monthly_column_layout(month_labels, group_labels=group_labels)

    def _avg_key(label: str) -> str:
        return _monthly_avg_column(label)

    def _gang_key(label: str) -> str:
        return _monthly_gangs_column(label)

    def _km_key(label: str) -> str:
        return _monthly_km_column(label)

    groups = list(group_columns)
    overall_avg_key = _avg_key("Overall")
    overall_gang_key = _gang_key("Overall")
    overall_km_key = _km_key("Overall")
    numeric_columns = [overall_avg_key, overall_gang_key, overall_km_key]

    daily = pd.DataFrame()
    if daily_scope is not None and not daily_scope.empty:
        daily = daily_scope.copy()
        daily["date"] = pd.to_datetime(daily.get("date"), errors="coerce").dt.normalize()
        daily = daily.dropna(subset=["date"])
        daily["month_key"] = daily["date"].dt.to_period("M").dt.to_timestamp()
        daily["daily_km"] = pd.to_numeric(daily.get("daily_km"), errors="coerce")
        daily["gang_name"] = daily.get("gang_name", "").fillna("").astype(str).str.strip()
        for column in groups:
            if column not in daily.columns:
                daily[column] = ""
            daily[column] = daily[column].fillna("").astype(str).str.strip()

    compiled = pd.DataFrame()
    if compiled_scope is not None and not compiled_scope.empty:
        compiled = compiled_scope.copy()
        if "month" in compiled.columns:
            compiled["month_key"] = pd.to_datetime(compiled.get("month"), errors="coerce").dt.to_period("M").dt.to_timestamp()
        else:
            compiled["month_key"] = (
                pd.to_datetime(compiled.get("fs_complete_date"), errors="coerce").dt.to_period("M").dt.to_timestamp()
            )
        compiled["length_km"] = pd.to_numeric(compiled.get("length_km"), errors="coerce").fillna(0.0)
        for column in groups:
            if column not in compiled.columns:
                compiled[column] = ""
            compiled[column] = compiled[column].fillna("").astype(str).str.strip()

    if not groups:
        row: dict[str, object] = {}
        for month_key, month_label in zip(month_keys, month_labels):
            month_daily = daily[daily["month_key"] == month_key] if not daily.empty else pd.DataFrame()
            month_avg = month_daily["daily_km"].dropna().mean() if not month_daily.empty else 0.0
            month_gangs = int(month_daily[month_daily["gang_name"].astype(bool)]["gang_name"].nunique()) if not month_daily.empty else 0
            month_compiled = compiled[compiled["month_key"] == month_key] if not compiled.empty else pd.DataFrame()
            month_km = float(month_compiled["length_km"].sum()) if not month_compiled.empty else 0.0

            row[_avg_key(month_label)] = round(float(month_avg) * _MONTH_PRODUCTIVITY_FACTOR if not pd.isna(month_avg) else 0.0, 2)
            row[_gang_key(month_label)] = month_gangs
            row[_km_key(month_label)] = round(month_km, 2)
            numeric_columns.extend([_avg_key(month_label), _gang_key(month_label), _km_key(month_label)])

        overall_avg = daily["daily_km"].dropna().mean() if not daily.empty else 0.0
        overall_gangs = int(daily[daily["gang_name"].astype(bool)]["gang_name"].nunique()) if not daily.empty else 0
        overall_km = float(compiled["length_km"].sum()) if not compiled.empty else 0.0
        row[overall_avg_key] = round(float(overall_avg) * _MONTH_PRODUCTIVITY_FACTOR if not pd.isna(overall_avg) else 0.0, 2)
        row[overall_gang_key] = overall_gangs
        row[overall_km_key] = round(overall_km, 2)

        result = pd.DataFrame([row], columns=ordered_columns).fillna(0.0)
        for key in set(numeric_columns):
            if key not in result.columns:
                continue
            if key.endswith("__gangs"):
                result[key] = pd.to_numeric(result[key], errors="coerce").fillna(0).astype(int)
            else:
                result[key] = pd.to_numeric(result[key], errors="coerce").fillna(0.0).round(2)
        return result

    key_frames: list[pd.DataFrame] = []
    if not daily.empty:
        key_frames.append(daily[groups].drop_duplicates())
    if not compiled.empty:
        key_frames.append(compiled[groups].drop_duplicates())
    if not key_frames:
        return pd.DataFrame(columns=ordered_columns)
    summary = pd.concat(key_frames, ignore_index=True).drop_duplicates().reset_index(drop=True)

    if not daily.empty:
        overall_avg = (
            daily.groupby(groups, dropna=False)["daily_km"]
            .mean()
            .mul(_MONTH_PRODUCTIVITY_FACTOR)
            .reset_index(name=overall_avg_key)
        )
        summary = summary.merge(overall_avg, on=groups, how="left")

        gang_source = daily[daily["gang_name"].astype(bool)]
        if not gang_source.empty:
            overall_gangs = (
                gang_source.groupby(groups, dropna=False)["gang_name"]
                .nunique()
                .reset_index(name=overall_gang_key)
            )
            summary = summary.merge(overall_gangs, on=groups, how="left")
        else:
            summary[overall_gang_key] = 0

        for month_key, month_label in zip(month_keys, month_labels):
            avg_key = _avg_key(month_label)
            month_avg = (
                daily[daily["month_key"] == month_key]
                .groupby(groups, dropna=False)["daily_km"]
                .mean()
                .mul(_MONTH_PRODUCTIVITY_FACTOR)
                .reset_index(name=avg_key)
            )
            summary = summary.merge(month_avg, on=groups, how="left")
            numeric_columns.append(avg_key)

            gang_key = _gang_key(month_label)
            month_gang_source = gang_source[gang_source["month_key"] == month_key] if not gang_source.empty else pd.DataFrame()
            if not month_gang_source.empty:
                month_gangs = (
                    month_gang_source.groupby(groups, dropna=False)["gang_name"]
                    .nunique()
                    .reset_index(name=gang_key)
                )
                summary = summary.merge(month_gangs, on=groups, how="left")
            else:
                summary[gang_key] = 0
            numeric_columns.append(gang_key)
    else:
        summary[overall_avg_key] = 0.0
        summary[overall_gang_key] = 0
        for month_label in month_labels:
            summary[_avg_key(month_label)] = 0.0
            summary[_gang_key(month_label)] = 0
            numeric_columns.extend([_avg_key(month_label), _gang_key(month_label)])

    if not compiled.empty:
        overall_km = (
            compiled.groupby(groups, dropna=False)["length_km"]
            .sum()
            .reset_index(name=overall_km_key)
        )
        summary = summary.merge(overall_km, on=groups, how="left")
        for month_key, month_label in zip(month_keys, month_labels):
            km_key = _km_key(month_label)
            month_km = (
                compiled[compiled["month_key"] == month_key]
                .groupby(groups, dropna=False)["length_km"]
                .sum()
                .reset_index(name=km_key)
            )
            summary = summary.merge(month_km, on=groups, how="left")
            numeric_columns.append(km_key)
    else:
        summary[overall_km_key] = 0.0
        for month_label in month_labels:
            summary[_km_key(month_label)] = 0.0
            numeric_columns.append(_km_key(month_label))

    summary = summary.rename(columns={src: dst for src, dst in zip(group_columns, group_labels)})
    for key in set(numeric_columns):
        if key not in summary.columns:
            summary[key] = 0.0
        if key.endswith("__gangs"):
            summary[key] = pd.to_numeric(summary[key], errors="coerce").fillna(0).astype(int)
        else:
            summary[key] = pd.to_numeric(summary[key], errors="coerce").fillna(0.0).round(2)

    summary = summary.reindex(columns=ordered_columns)
    if group_labels == ["PCH"] and "PCH" in summary.columns:
        summary = summary.sort_values("PCH", key=lambda series: series.astype(str).str.lower()).reset_index(drop=True)
    elif group_labels == ["PCH", "Project"] and {"PCH", "Project"}.issubset(summary.columns):
        summary = summary.sort_values(
            ["PCH", "Project"],
            key=lambda series: series.astype(str).str.lower(),
        ).reset_index(drop=True)
    return summary


def _build_monthly_rollup_tables(
    daily_scope: pd.DataFrame,
    compiled_scope: pd.DataFrame,
    *,
    months: list[pd.Timestamp],
    month_labels: list[str],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    month_keys = [
        pd.Timestamp(month).normalize().to_period("M").to_timestamp()
        for month in months
    ]
    overall = _build_monthly_group_summary_table(
        daily_scope,
        compiled_scope,
        month_keys=month_keys,
        month_labels=month_labels,
        group_columns=[],
        group_labels=[],
    )
    pch = _build_monthly_group_summary_table(
        daily_scope,
        compiled_scope,
        month_keys=month_keys,
        month_labels=month_labels,
        group_columns=["pch_display"],
        group_labels=["PCH"],
    )
    project = _build_monthly_group_summary_table(
        daily_scope,
        compiled_scope,
        month_keys=month_keys,
        month_labels=month_labels,
        group_columns=["pch_display", "project_display"],
        group_labels=["PCH", "Project"],
    )
    return overall, pch, project


def _write_scoped_table(
    writer: pd.ExcelWriter,
    sheet_name: str,
    table: pd.DataFrame,
    scope_label: str,
    *,
    startrow: int,
) -> int:
    header_df = pd.DataFrame([[scope_label]])
    header_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
    table_start = startrow + len(header_df) + 1
    table.to_excel(writer, sheet_name=sheet_name, index=False, startrow=table_start)
    return table_start + len(table) + 1


def _write_labeled_table(
    writer: pd.ExcelWriter,
    sheet_name: str,
    label: str,
    table: pd.DataFrame,
    *,
    startrow: int,
) -> int:
    label_df = pd.DataFrame([[label]])
    label_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
    table_start = startrow + len(label_df) + 1
    if table is None or table.empty:
        return table_start
    table.to_excel(writer, sheet_name=sheet_name, index=False, startrow=table_start)
    return table_start + len(table) + 1


def _write_monthly_scoped_table(
    writer: pd.ExcelWriter,
    sheet_name: str,
    table: pd.DataFrame,
    scope_label: str,
    *,
    month_headers: list[str],
    group_headers: list[str],
    startrow: int,
) -> tuple[int, dict[str, int]]:
    scope_df = pd.DataFrame([[scope_label]])
    scope_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
    header_row = startrow + len(scope_df)
    ordered_columns, top_headers, bottom_headers = _monthly_column_layout(
        month_headers,
        group_labels=group_headers,
    )
    aligned = table.reindex(columns=ordered_columns)
    pd.DataFrame([top_headers, bottom_headers], columns=ordered_columns).to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        header=False,
        startrow=header_row,
    )
    data_row = header_row + 2
    aligned.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        header=False,
        startrow=data_row,
    )
    next_row = data_row + len(aligned) + 1
    block = {
        "scope_row": startrow,
        "header_top_row": header_row,
        "header_bottom_row": header_row + 1,
        "data_start_row": data_row,
        "data_end_row": data_row + max(len(aligned) - 1, 0),
        "group_count": len(group_headers),
        "month_count": len(month_headers),
        "column_count": len(ordered_columns),
    }
    return next_row, block


def _style_monthly_summary_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    blocks: list[dict[str, int]],
) -> None:
    worksheet = writer.sheets.get(sheet_name)
    if worksheet is None or not blocks:
        return

    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    scope_font = Font(bold=True, color="334155")
    header_font = Font(bold=True, color="1F2937")
    data_font = Font(color="111827")
    scope_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    top_header_fill = PatternFill(fill_type="solid", fgColor="EAF1F8")
    bottom_header_fill = PatternFill(fill_type="solid", fgColor="F5F8FC")
    zebra_fill = PatternFill(fill_type="solid", fgColor="FAFCFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    max_columns = max(block["column_count"] for block in blocks)
    for col_idx in range(1, max_columns + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            worksheet.column_dimensions[col_letter].width = 30
        elif col_idx == 2:
            worksheet.column_dimensions[col_letter].width = 20
        else:
            worksheet.column_dimensions[col_letter].width = 14

    for block in blocks:
        first_col = 1
        last_col = block["column_count"]
        scope_row = block["scope_row"] + 1
        top_row = block["header_top_row"] + 1
        bottom_row = block["header_bottom_row"] + 1
        data_start = block["data_start_row"] + 1
        data_end = block["data_end_row"] + 1
        group_count = block["group_count"]

        scope_cell = worksheet.cell(row=scope_row, column=first_col)
        scope_cell.font = scope_font
        scope_cell.fill = scope_fill
        scope_cell.alignment = left

        for col in range(first_col, last_col + 1):
            top_cell = worksheet.cell(row=top_row, column=col)
            top_cell.font = header_font
            top_cell.fill = top_header_fill
            top_cell.alignment = center
            top_cell.border = border

            bottom_cell = worksheet.cell(row=bottom_row, column=col)
            bottom_cell.font = header_font
            bottom_cell.fill = bottom_header_fill
            bottom_cell.alignment = center
            bottom_cell.border = border

        metric_span = 3
        month_block_count = block["month_count"] + 1
        month_start_col = first_col + group_count
        for idx in range(month_block_count):
            span_start = month_start_col + (idx * metric_span)
            span_end = span_start + metric_span - 1
            worksheet.merge_cells(
                start_row=top_row,
                start_column=span_start,
                end_row=top_row,
                end_column=span_end,
            )

        if group_count > 0:
            for offset in range(group_count):
                col = first_col + offset
                label_value = worksheet.cell(row=bottom_row, column=col).value
                worksheet.merge_cells(
                    start_row=top_row,
                    start_column=col,
                    end_row=bottom_row,
                    end_column=col,
                )
                merged_cell = worksheet.cell(row=top_row, column=col)
                merged_cell.value = label_value
                merged_cell.font = header_font
                merged_cell.fill = top_header_fill
                merged_cell.alignment = center
                merged_cell.border = border

        if data_end < data_start:
            continue
        for row in range(data_start, data_end + 1):
            apply_zebra = (row - data_start) % 2 == 1
            for col in range(first_col, last_col + 1):
                cell = worksheet.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.font = data_font
                cell.border = border
                is_group_col = col < (first_col + group_count)
                cell.alignment = left if is_group_col else center
                if apply_zebra:
                    cell.fill = zebra_fill
                if not is_group_col:
                    metric_idx = (col - (first_col + group_count)) % metric_span
                    if metric_idx == 1:
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0.00"

        worksheet.row_dimensions[top_row].height = 22
        worksheet.row_dimensions[bottom_row].height = 22


def _style_clean_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    worksheet = writer.sheets.get(sheet_name)
    if worksheet is None:
        return

    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    label_font = Font(bold=True, color="334155")
    header_font = Font(bold=True, color="1F2937")
    data_font = Font(color="111827")
    label_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    header_fill = PatternFill(fill_type="solid", fgColor="F3F7FB")
    zebra_fill = PatternFill(fill_type="solid", fgColor="FAFCFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 0
    if max_row == 0 or max_col == 0:
        return

    column_widths: dict[int, int] = {}
    data_row_band = 0

    for row_idx in range(1, max_row + 1):
        row_values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
        non_empty_pairs = [
            (idx + 1, value)
            for idx, value in enumerate(row_values)
            if value is not None and str(value).strip() != ""
        ]
        if not non_empty_pairs:
            continue

        non_empty_values = [value for _, value in non_empty_pairs]
        string_count = sum(1 for value in non_empty_values if isinstance(value, str))
        numeric_count = sum(
            1 for value in non_empty_values if isinstance(value, (int, float)) and not isinstance(value, bool)
        )
        is_label_row = len(non_empty_pairs) == 1 and string_count == 1
        is_header_row = (
            (not is_label_row)
            and len(non_empty_pairs) >= 2
            and string_count >= max(2, int(len(non_empty_pairs) * 0.6))
            and numeric_count <= max(2, int(len(non_empty_pairs) * 0.4))
        )
        is_data_row = not is_label_row and not is_header_row
        if is_data_row:
            data_row_band += 1

        for col_idx, value in non_empty_pairs:
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue

            cell_text = str(value)
            width_hint = min(48, max(8, len(cell_text) + 2))
            prev_width = column_widths.get(col_idx, 0)
            if width_hint > prev_width:
                column_widths[col_idx] = width_hint

            cell.border = border
            if is_label_row:
                cell.font = label_font
                cell.fill = label_fill
                cell.alignment = left
                continue
            if is_header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                continue

            cell.font = data_font
            if data_row_band % 2 == 0:
                cell.fill = zebra_fill
            cell.alignment = left if col_idx <= 2 else center
            if isinstance(value, bool):
                continue
            if cell.number_format in {"General", "0", ""}:
                if isinstance(value, int):
                    cell.number_format = "#,##0"
                elif isinstance(value, float):
                    cell.number_format = "#,##0.00"

    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = width


def _normalize_method(value: object) -> str:
    text = str(value or "").strip().lower()
    if text in {"", "nan", "none", "null"}:
        return ""
    return text


def _method_group(value: str) -> str:
    if "tse" in value:
        return "TSE"
    if "manual" in value:
        return "Manual"
    return "Other"


def _build_method_summary(
    scope: pd.DataFrame,
    issues: list[dict[str, object]],
) -> pd.DataFrame:
    if scope is None or scope.empty:
        return pd.DataFrame()
    work = scope.copy()
    method_series = work.get("method", work.get("method_norm", pd.Series("", index=work.index)))
    method_norm = method_series.map(_normalize_method)
    work["method_group"] = method_norm.map(_method_group)
    missing_mask = method_norm.eq("")
    if missing_mask.any():
        for _, row in work.loc[missing_mask].iterrows():
            issues.append(
                {
                    "issue_type": "METHOD_MISSING",
                    "project": row.get("project_name", ""),
                    "from_ap": row.get("from_ap", ""),
                    "to_ap": row.get("to_ap", ""),
                    "date": row.get("date", pd.NaT),
                    "row_id": row.get("row_id", ""),
                    "details": "Method column missing or blank",
                }
            )
    summary = _summarize_scope(work, group_columns=["method_group"])
    if summary.empty:
        return summary
    summary = summary.rename(
        columns={
            "method_group": "Method",
            "avg_km_month": "Avg Productivity (KM/month)",
            "total_km": "Total KM",
            "spans": "Spans",
            "gangs": "Gangs",
            "projects": "Projects",
        }
    )
    return summary


def _filter_tse_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    work = df.copy()
    method_series = work.get("method", work.get("method_norm", pd.Series("", index=work.index)))
    method_norm = method_series.map(_normalize_method)
    work["method_group"] = method_norm.map(_method_group)
    return work[work["method_group"] == "TSE"].copy()


def _build_pareto_table(
    daily_scope: pd.DataFrame,
    compiled_scope: pd.DataFrame,
    *,
    method_filter: str | None = None,
) -> pd.DataFrame:
    buckets = [
        (">=6", 6.0, None),
        ("4-6", 4.0, 6.0),
        ("3-4", 3.0, 4.0),
        ("2-3", 2.0, 3.0),
        ("1-2", 1.0, 2.0),
        ("<1", None, 1.0),
    ]

    if daily_scope is None:
        daily_scope = pd.DataFrame()
    if compiled_scope is None:
        compiled_scope = pd.DataFrame()

    daily = daily_scope.copy()
    compiled = compiled_scope.copy()

    for work in (daily, compiled):
        if work.empty:
            continue
        work["gang_name"] = work.get("gang_name", "").fillna("").astype(str).str.strip()
        method_series = work.get("method", work.get("method_norm", pd.Series("", index=work.index)))
        method_norm = method_series.map(_normalize_method)
        work["method_group"] = method_norm.map(_method_group)
        work.drop(work[work["gang_name"] == ""].index, inplace=True)

    if method_filter:
        if not daily.empty:
            daily = daily[daily["method_group"] == method_filter]
        if not compiled.empty:
            compiled = compiled[compiled["method_group"] == method_filter]

    avg_per_gang = (
        pd.to_numeric(daily.get("daily_km", pd.Series([], dtype=float)), errors="coerce")
        .groupby(daily.get("gang_name", pd.Series([], dtype=object)))
        .mean()
        .mul(_MONTH_PRODUCTIVITY_FACTOR)
    )
    length_per_gang = (
        pd.to_numeric(compiled.get("length_km", pd.Series([], dtype=float)), errors="coerce")
        .groupby(compiled.get("gang_name", pd.Series([], dtype=object)))
        .sum()
    )

    all_gangs = sorted(set(avg_per_gang.index) | set(length_per_gang.index))
    if not all_gangs:
        return pd.DataFrame(
            [
                {
                    "Bucket": label,
                    "KM Delivered": 0.0,
                    "Number of Gangs": 0,
                    "KM % Delivered": "",
                    "Number of Gangs %": "",
                }
                for label, _, _ in buckets
            ]
        )

    avg_per_gang = avg_per_gang.reindex(all_gangs).fillna(0.0)
    length_per_gang = length_per_gang.reindex(all_gangs).fillna(0.0)

    def _bucket_for(value: float) -> str:
        for label, lower, upper in buckets:
            if lower is None and value < upper:
                return label
            if upper is None and value >= lower:
                return label
            if lower is not None and upper is not None and lower <= value < upper:
                return label
        return "<1"

    bucket_series = avg_per_gang.map(_bucket_for)
    total_km = float(length_per_gang.sum())
    total_gangs = len(all_gangs)

    rows: list[dict[str, object]] = []
    for label, _, _ in buckets:
        gangs_in_bucket = bucket_series[bucket_series == label].index
        km_sum = float(length_per_gang.loc[gangs_in_bucket].sum()) if len(gangs_in_bucket) else 0.0
        gangs_count = int(len(gangs_in_bucket))
        rows.append(
            {
                "Bucket": label,
                "KM Delivered": round(km_sum, 2),
                "Number of Gangs": gangs_count,
                "KM % Delivered": "",
                "Number of Gangs %": "",
            }
        )
    return pd.DataFrame(rows)


def _apply_pareto_formulas(
    writer: pd.ExcelWriter,
    sheet_name: str,
    *,
    startrow: int,
    table: pd.DataFrame,
) -> None:
    if table is None or table.empty:
        return
    workbook = writer.book
    if sheet_name not in workbook.sheetnames:
        return
    ws = workbook[sheet_name]
    first_row = startrow + 2
    last_row = startrow + 1 + len(table)
    if last_row < first_row:
        return
    sum_km = f"$B${first_row}:$B${last_row}"
    sum_gangs = f"$C${first_row}:$C${last_row}"
    for row in range(first_row, last_row + 1):
        km_cell = f"B{row}"
        gangs_cell = f"C{row}"
        km_pct_cell = f"D{row}"
        gangs_pct_cell = f"E{row}"
        ws[km_pct_cell].value = f"=IF(SUM({sum_km})=0,0,{km_cell}/SUM({sum_km}))"
        ws[gangs_pct_cell].value = f"=IF(SUM({sum_gangs})=0,0,{gangs_cell}/SUM({sum_gangs}))"
        ws[km_pct_cell].number_format = "0.00%"
        ws[gangs_pct_cell].number_format = "0.00%"
def _normalize_stringing_compiled(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    normalized, _ = normalize_stringing_columns(df)
    normalized = _ensure_project_fields(normalized)
    for col in ("from_ap", "to_ap", "gang_name", "source_file"):
        if col not in normalized.columns:
            normalized[col] = ""
    normalized["from_ap"] = normalized["from_ap"].fillna("").astype(str).str.strip()
    normalized["to_ap"] = normalized["to_ap"].fillna("").astype(str).str.strip()
    normalized["gang_name"] = normalized["gang_name"].fillna("").astype(str).str.strip()
    normalized["source_file"] = normalized["source_file"].fillna("").astype(str).str.strip()
    return normalized


def _build_gap_summary(df: pd.DataFrame, gap_column: str) -> pd.DataFrame:
    if df is None or df.empty or gap_column not in df.columns:
        return pd.DataFrame(
            [
                {
                    "count": 0,
                    "avg_gap_days": 0.0,
                    "median_gap_days": 0.0,
                    "min_gap_days": 0,
                    "max_gap_days": 0,
                }
            ]
        )
    series = pd.to_numeric(df[gap_column], errors="coerce").dropna()
    if series.empty:
        return pd.DataFrame(
            [
                {
                    "count": 0,
                    "avg_gap_days": 0.0,
                    "median_gap_days": 0.0,
                    "min_gap_days": 0,
                    "max_gap_days": 0,
                }
            ]
        )
    stats_series = series.clip(lower=0)
    return pd.DataFrame(
        [
            {
                "count": int(stats_series.size),
                "avg_gap_days": round(float(stats_series.mean()), 2),
                "median_gap_days": round(float(stats_series.median()), 2),
                "min_gap_days": int(stats_series.min()),
                "max_gap_days": int(stats_series.max()),
            }
        ]
    )


def _build_po_fs_gap_table(
    compiled: pd.DataFrame,
    *,
    start_date: pd.Timestamp | None,
    end_date: pd.Timestamp | None,
    issues: list[dict[str, object]],
) -> pd.DataFrame:
    work = _normalize_stringing_compiled(compiled)
    if work.empty:
        return pd.DataFrame()
    work["po_start_date"] = _parse_date_series(work.get("po_start_date"))
    work["po_completion_date"] = _parse_date_series(work.get("po_completion_date"))
    work["fs_starting_date"] = _parse_date_series(work.get("fs_starting_date"))
    work["fs_complete_date"] = _parse_date_series(work.get("fs_complete_date"))

    scoped = work.copy()
    if start_date is not None:
        scoped = scoped[scoped["po_completion_date"] >= start_date]
    if end_date is not None:
        scoped = scoped[scoped["po_completion_date"] <= end_date]
    if scoped.empty:
        return pd.DataFrame()

    gap = (scoped["fs_starting_date"] - scoped["po_completion_date"]).dt.days
    negative_mask = gap < 0
    if negative_mask.any():
        for _, row in scoped.loc[negative_mask].iterrows():
            issues.append(
                {
                    "issue_type": "PO_FS_NEGATIVE_GAP",
                    "project": row.get("project_name", ""),
                    "from_ap": row.get("from_ap", ""),
                    "to_ap": row.get("to_ap", ""),
                    "po_start_date": row.get("po_start_date", pd.NaT),
                    "po_completion_date": row.get("po_completion_date", pd.NaT),
                    "fs_starting_date": row.get("fs_starting_date", pd.NaT),
                    "fs_complete_date": row.get("fs_complete_date", pd.NaT),
                    "details": "FS start is before PO completion",
                }
            )
    scoped["gap_days"] = gap
    scoped["negative_gap_flag"] = negative_mask.fillna(False)

    return scoped[
        [
            "project_name",
            "from_ap",
            "to_ap",
            "gang_name",
            "po_completion_date",
            "fs_starting_date",
            "gap_days",
            "negative_gap_flag",
            "source_file",
        ]
    ].copy()


def _build_erection_location_map(erection_daily: pd.DataFrame) -> dict[str, pd.DataFrame]:
    if erection_daily is None or erection_daily.empty:
        return {}
    working = erection_daily.copy()
    working["completion_date"] = pd.to_datetime(working.get("completion_date"), errors="coerce").dt.normalize()
    working["location_no"] = working.get("location_no", "").fillna("").astype(str).str.strip()
    working["project_key_norm"] = working.get("project_key_norm", "").fillna("").astype(str)
    working = working.dropna(subset=["completion_date"])
    working = working[working["location_no"].astype(bool) & working["project_key_norm"].astype(bool)]
    if working.empty:
        return {}
    working["location_no_norm"] = working["location_no"].map(normalize_location)
    working["loc_order"] = working["location_no_norm"].map(_location_order_key)
    working = working.dropna(subset=["loc_order"])
    if working.empty:
        return {}
    working = (
        working.sort_values("completion_date")
        .drop_duplicates(subset=["project_key_norm", "location_no_norm", "loc_order"], keep="last")
    )
    project_map: dict[str, pd.DataFrame] = {}
    for project_key, group in working.groupby("project_key_norm"):
        project_map[str(project_key)] = group[
            ["loc_order", "completion_date", "location_no_norm", "location_no"]
        ].copy()
    return project_map


def _build_erection_po_gap_table(
    compiled: pd.DataFrame,
    erection_daily: pd.DataFrame,
    *,
    start_date: pd.Timestamp | None,
    end_date: pd.Timestamp | None,
    issues: list[dict[str, object]],
) -> pd.DataFrame:
    work = _normalize_stringing_compiled(compiled)
    if work.empty:
        return pd.DataFrame()
    work["po_start_date"] = _parse_date_series(work.get("po_start_date"))
    if start_date is not None:
        work = work[work["po_start_date"] >= start_date]
    if end_date is not None:
        work = work[work["po_start_date"] <= end_date]
    if work.empty:
        return pd.DataFrame()

    work["from_ap_norm"] = work["from_ap"].map(_strip_ap_prefix)
    work["to_ap_norm"] = work["to_ap"].map(_strip_ap_prefix)
    work["from_order"] = work["from_ap_norm"].map(_location_order_key)
    work["to_order"] = work["to_ap_norm"].map(_location_order_key)
    work["from_is_gantry"] = work["from_ap"].map(_is_gantry_label)
    work["to_is_gantry"] = work["to_ap"].map(_is_gantry_label)

    erection_map = _build_erection_location_map(erection_daily)

    rows: list[dict[str, object]] = []
    for _, row in work.iterrows():
        project_key = row.get("project_key_norm", "")
        po_start = row.get("po_start_date")
        from_order = row.get("from_order")
        to_order = row.get("to_order")
        from_ap = row.get("from_ap", "")
        to_ap = row.get("to_ap", "")
        from_is_gantry = bool(row.get("from_is_gantry", False))
        to_is_gantry = bool(row.get("to_is_gantry", False))
        base = {
            "project_name": row.get("project_name", ""),
            "from_ap": from_ap,
            "to_ap": to_ap,
            "gang_name": row.get("gang_name", ""),
            "po_start_date": po_start,
            "source_file": row.get("source_file", ""),
        }

        if pd.isna(po_start):
            issues.append(
                {
                    "issue_type": "PO_START_MISSING",
                    "project": row.get("project_name", ""),
                    "from_ap": from_ap,
                    "to_ap": to_ap,
                    "po_start_date": po_start,
                    "po_completion_date": row.get("po_completion_date", pd.NaT),
                    "fs_starting_date": row.get("fs_starting_date", pd.NaT),
                    "fs_complete_date": row.get("fs_complete_date", pd.NaT),
                    "details": "Missing PO start date",
                }
            )
            rows.append({**base, "last_erection_completion_date": pd.NaT, "gap_days": pd.NA, "erection_locations": 0})
            continue

        if from_is_gantry and to_is_gantry:
            issues.append(
                {
                    "issue_type": "GANTRY_BOTH_SIDES",
                    "project": row.get("project_name", ""),
                    "from_ap": from_ap,
                    "to_ap": to_ap,
                    "po_start_date": po_start,
                    "po_completion_date": row.get("po_completion_date", pd.NaT),
                    "fs_starting_date": row.get("fs_starting_date", pd.NaT),
                    "fs_complete_date": row.get("fs_complete_date", pd.NaT),
                    "details": "Both From_AP and To_AP marked as gantry/end",
                }
            )
            rows.append({**base, "last_erection_completion_date": pd.NaT, "gap_days": pd.NA, "erection_locations": 0})
            continue

        if (not from_is_gantry and pd.isna(from_order)) or (not to_is_gantry and pd.isna(to_order)):
            issues.append(
                {
                    "issue_type": "LOCATION_PARSE_FAILED",
                    "project": row.get("project_name", ""),
                    "from_ap": from_ap,
                    "to_ap": to_ap,
                    "po_start_date": po_start,
                    "po_completion_date": row.get("po_completion_date", pd.NaT),
                    "fs_starting_date": row.get("fs_starting_date", pd.NaT),
                    "fs_complete_date": row.get("fs_complete_date", pd.NaT),
                    "details": "Unable to parse span location ordering",
                }
            )
            rows.append({**base, "last_erection_completion_date": pd.NaT, "gap_days": pd.NA, "erection_locations": 0})
            continue

        project_df = erection_map.get(str(project_key))
        if project_df is None or project_df.empty:
            issues.append(
                {
                    "issue_type": "ERECTION_PROJECT_NOT_FOUND",
                    "project": row.get("project_name", ""),
                    "from_ap": from_ap,
                    "to_ap": to_ap,
                    "po_start_date": po_start,
                    "po_completion_date": row.get("po_completion_date", pd.NaT),
                    "fs_starting_date": row.get("fs_starting_date", pd.NaT),
                    "fs_complete_date": row.get("fs_complete_date", pd.NaT),
                    "details": "No erection data for project",
                }
            )
            rows.append({**base, "last_erection_completion_date": pd.NaT, "gap_days": pd.NA, "erection_locations": 0})
            continue

        if to_is_gantry and not from_is_gantry:
            lo = int(from_order)
            hi = int(project_df["loc_order"].max())
        elif from_is_gantry and not to_is_gantry:
            lo = int(project_df["loc_order"].min())
            hi = int(to_order)
        else:
            lo = min(int(from_order), int(to_order))
            hi = max(int(from_order), int(to_order))
        span_df = project_df[(project_df["loc_order"] >= lo) & (project_df["loc_order"] <= hi)]
        if span_df.empty:
            issues.append(
                {
                    "issue_type": "ERECTION_SPAN_EMPTY",
                    "project": row.get("project_name", ""),
                    "from_ap": from_ap,
                    "to_ap": to_ap,
                    "po_start_date": po_start,
                    "po_completion_date": row.get("po_completion_date", pd.NaT),
                    "fs_starting_date": row.get("fs_starting_date", pd.NaT),
                    "fs_complete_date": row.get("fs_complete_date", pd.NaT),
                    "details": "No erection locations found in span range",
                }
            )
            rows.append({**base, "last_erection_completion_date": pd.NaT, "gap_days": pd.NA, "erection_locations": 0})
            continue

        last_completion = span_df["completion_date"].max()
        gap_days = (po_start - last_completion).days if pd.notna(last_completion) else pd.NA
        if gap_days is not pd.NA and gap_days < 0:
            issues.append(
                {
                    "issue_type": "ERECTION_PO_NEGATIVE_GAP",
                    "project": row.get("project_name", ""),
                    "from_ap": from_ap,
                    "to_ap": to_ap,
                    "po_start_date": po_start,
                    "po_completion_date": row.get("po_completion_date", pd.NaT),
                    "fs_starting_date": row.get("fs_starting_date", pd.NaT),
                    "fs_complete_date": row.get("fs_complete_date", pd.NaT),
                    "details": "PO start is before last erection completion in span",
                }
            )

        rows.append(
            {
                **base,
                "last_erection_completion_date": last_completion,
                "gap_days": int(gap_days) if gap_days is not pd.NA else pd.NA,
                "erection_locations": int(span_df["location_no_norm"].nunique()),
            }
        )

    return pd.DataFrame(rows)


def main() -> int:
    args = _parse_args()
    configure_logging()

    config = AppConfig()
    if args.data_path:
        config = replace(config, data_path=Path(args.data_path).expanduser())
    if args.stringing_data_path:
        config = replace(config, stringing_data_path=Path(args.stringing_data_path).expanduser())
    config.validate()

    store = AppDataStore(config)
    LOGGER.info("Bootstrapping datasets from %s", config.data_path)
    store.bootstrap(config)

    start_date = _parse_date(args.start_date)
    end_date = _parse_date(args.end_date)
    if start_date and end_date and start_date > end_date:
        raise SystemExit("--start-date must be before or equal to --end-date.")

    stringing_daily = store.get_stringing()
    stringing_compiled = store.get_stringing_compiled()
    erection_daily = store.get_daily()

    if stringing_compiled is None or stringing_compiled.empty:
        LOGGER.warning("Stringing compiled dataset not available; some analyses will be empty.")

    # Overall stringing daily (PO start -> FS complete)
    if stringing_compiled is not None and not stringing_compiled.empty:
        overall_daily = expand_stringing_to_daily(stringing_compiled)
    else:
        overall_daily = stringing_daily
    overall_daily = _filter_by_date(overall_daily, "date", start_date, end_date)
    overall_daily_tse = _filter_tse_rows(overall_daily)

    # PO-only daily (PO start -> PO completion)
    po_daily = expand_stringing_to_daily_payout(stringing_compiled) if not stringing_compiled.empty else pd.DataFrame()
    po_daily = _filter_by_date(po_daily, "date", start_date, end_date)
    po_daily_tse = _filter_tse_rows(po_daily)

    # FS-only daily (FS start -> FS complete)
    fs_daily = expand_stringing_to_daily_fs(stringing_compiled) if not stringing_compiled.empty else pd.DataFrame()
    fs_daily = _filter_by_date(fs_daily, "date", start_date, end_date)
    fs_daily_tse = _filter_tse_rows(fs_daily)

    summary_rows = pd.concat(
        [
            _build_stage_summary(overall_daily, "Overall"),
            _build_stage_summary(po_daily, "P/O"),
            _build_stage_summary(fs_daily, "F/S"),
        ],
        ignore_index=True,
    )
    summary_rows_tse = pd.concat(
        [
            _build_stage_summary(overall_daily_tse, "Overall"),
            _build_stage_summary(po_daily_tse, "P/O"),
            _build_stage_summary(fs_daily_tse, "F/S"),
        ],
        ignore_index=True,
    )

    gang_rows = pd.concat(
        [
            _build_gang_productivity(overall_daily, "Overall"),
            _build_gang_productivity(po_daily, "P/O"),
            _build_gang_productivity(fs_daily, "F/S"),
        ],
        ignore_index=True,
    )
    gang_rows_tse = pd.concat(
        [
            _build_gang_productivity(overall_daily_tse, "Overall"),
            _build_gang_productivity(po_daily_tse, "P/O"),
            _build_gang_productivity(fs_daily_tse, "F/S"),
        ],
        ignore_index=True,
    )

    issues: list[dict[str, object]] = []

    po_fs_gap = _build_po_fs_gap_table(
        stringing_compiled,
        start_date=start_date,
        end_date=end_date,
        issues=issues,
    )
    po_fs_gap_summary = _build_gap_summary(po_fs_gap, "gap_days")
    po_fs_gap_tse = _build_po_fs_gap_table(
        _filter_tse_rows(stringing_compiled),
        start_date=start_date,
        end_date=end_date,
        issues=[],
    )
    if po_fs_gap_tse.empty and not po_fs_gap.empty:
        po_fs_gap_tse = po_fs_gap.head(0)
    po_fs_gap_summary_tse = _build_gap_summary(po_fs_gap_tse, "gap_days")

    erection_po_gap = _build_erection_po_gap_table(
        stringing_compiled,
        erection_daily,
        start_date=start_date,
        end_date=end_date,
        issues=issues,
    )
    erection_po_gap_summary = _build_gap_summary(erection_po_gap, "gap_days")
    erection_po_gap_tse = _build_erection_po_gap_table(
        _filter_tse_rows(stringing_compiled),
        erection_daily,
        start_date=start_date,
        end_date=end_date,
        issues=[],
    )
    if erection_po_gap_tse.empty and not erection_po_gap.empty:
        erection_po_gap_tse = erection_po_gap.head(0)
    erection_po_gap_summary_tse = _build_gap_summary(erection_po_gap_tse, "gap_days")

    # Stringing rollups (overall / PCH / project) and review format
    project_info = store.get_project_info()
    scope = _prepare_stringing_scope(overall_daily, project_info)
    method_series = scope.get("method", scope.get("method_norm", pd.Series("", index=scope.index)))
    method_norm = method_series.map(_normalize_method)
    scope["method_group"] = method_norm.map(_method_group)
    scope_tse = scope[scope["method_group"] == "TSE"].copy()
    compiled_scope = _prepare_stringing_compiled_scope(stringing_compiled, project_info)
    method_summary = _build_method_summary(scope, issues)
    months, month_labels = _build_month_windows_from_series(scope["date"], start_date, end_date)
    monthly_labels = [month.strftime("%b-%y") for month in months]
    scope_label = _build_scope_label(scope, start_date, end_date)

    if not compiled_scope.empty:
        method_series = compiled_scope.get("method", compiled_scope.get("method_norm", pd.Series("", index=compiled_scope.index)))
        method_norm = method_series.map(_normalize_method)
        compiled_scope = compiled_scope.copy()
        compiled_scope["method_group"] = method_norm.map(_method_group)
        compiled_scope_tse = compiled_scope[compiled_scope["method_group"] == "TSE"].copy()
    else:
        compiled_scope_tse = pd.DataFrame()

    overall_rollup, pch_rollup, project_rollup = _build_rollup_tables(scope, scope_tse)
    tse_overall_rollup, tse_pch_rollup, tse_project_rollup = _build_rollup_tables(scope_tse, scope_tse)
    monthly_overall, monthly_pch, monthly_project = _build_monthly_rollup_tables(
        scope,
        compiled_scope,
        months=months,
        month_labels=monthly_labels,
    )
    monthly_tse_overall, monthly_tse_pch, monthly_tse_project = _build_monthly_rollup_tables(
        scope_tse,
        compiled_scope_tse,
        months=months,
        month_labels=monthly_labels,
    )

    review_table_raw = project_rollup.reindex(
        columns=[
            "PCH",
            "Project",
            "Overall Avg Productivity (KM/month)",
            "TSE Avg Productivity (KM/month)",
            "Total KM",
            "Spans",
        ]
    )
    review_table_one_groups: list[pd.DataFrame] = []
    if not review_table_raw.empty:
        overall_table = _append_totals_row(
            review_table_raw.copy(),
            avg_col="Overall Avg Productivity (KM/month)",
            total_col="Total KM",
            span_col="Spans",
        )
        if not overall_table.empty and "PCH" in overall_table.columns:
            overall_table.loc[overall_table.index[-1], "PCH"] = "Overall"
        review_table_one_groups.append(overall_table)

    if review_table_raw.empty or "PCH" not in review_table_raw.columns:
        if review_table_raw.empty:
            review_table_one_groups.append(
                _append_totals_row(
                    review_table_raw,
                    avg_col="Overall Avg Productivity (KM/month)",
                    total_col="Total KM",
                    span_col="Spans",
                )
            )
    else:
        unique_pch = sorted(review_table_raw["PCH"].fillna("").astype(str).unique())
        for pch_name in unique_pch:
            if not pch_name:
                continue
            pch_table = review_table_raw[review_table_raw["PCH"] == pch_name].copy()
            pch_with_total = _append_totals_row(
                pch_table,
                avg_col="Overall Avg Productivity (KM/month)",
                total_col="Total KM",
                span_col="Spans",
            )
            if not pch_with_total.empty and "PCH" in pch_with_total.columns:
                pch_with_total.loc[pch_with_total.index[-1], "PCH"] = pch_name
            review_table_one_groups.append(pch_with_total)

    review_table_two_groups: list[tuple[pd.DataFrame, bool]] = []
    if not scope.empty and month_labels and "pch_display" in scope.columns:
        months_desc = list(reversed(months))
        labels_desc = list(reversed(month_labels))
        review_table_two_groups.append(
            (_build_review_table_three(scope, compiled_scope, months_desc, labels_desc), True)
        )
        for pch_name in sorted(scope["pch_display"].fillna("").astype(str).unique()):
            if not pch_name:
                continue
            pch_scope = scope[scope["pch_display"] == pch_name]
            header_row = {"Stringing productivity": f"PCH: {pch_name}"}
            for label in labels_desc:
                header_row[label] = ""
            review_table_two_groups.append(
                (pd.DataFrame([header_row], columns=["Stringing productivity", *labels_desc]), False)
            )
            if "project_display" in pch_scope.columns and "project_key_norm" in pch_scope.columns:
                project_pairs = (
                    pch_scope[["project_display", "project_key_norm"]]
                    .dropna(subset=["project_key_norm"])
                    .drop_duplicates()
                )
                for _, proj_row in project_pairs.iterrows():
                    project_name = str(proj_row.get("project_display", "")).strip()
                    project_key = str(proj_row.get("project_key_norm", "")).strip()
                    if not project_key:
                        continue
                    proj_scope = pch_scope[pch_scope["project_key_norm"] == project_key]
                    header_row = {"Stringing productivity": f"Project: {project_name}"}
                    for label in labels_desc:
                        header_row[label] = ""
                    review_table_two_groups.append(
                        (pd.DataFrame([header_row], columns=["Stringing productivity", *labels_desc]), False)
                    )
                    compiled_proj_scope = (
                        compiled_scope[compiled_scope["project_key_norm"] == project_key]
                        if not compiled_scope.empty
                        else pd.DataFrame()
                    )
                    review_table_two_groups.append(
                        (
                            _build_review_table_three(
                                proj_scope, compiled_proj_scope, months_desc, labels_desc
                            ),
                            True,
                        )
                    )

    issues_df = pd.DataFrame(issues)
    if not issues_df.empty:
        for col in ("po_start_date", "po_completion_date", "fs_starting_date", "fs_complete_date"):
            if col not in issues_df.columns:
                issues_df[col] = pd.NaT

    readme_rows = [
        {
            "Assumption": "Productivity units",
            "Details": "Avg productivity is reported per month as mean(daily_km) * 30. Total_km is the sum of daily_km.",
        },
        {
            "Assumption": "Overall / P/O / F/S daily rows",
            "Details": "Overall uses PO start to FS complete. P/O uses PO start to PO completion with PO length as primary metric (fallback to span length). F/S uses FS start to FS complete with span length.",
        },
        {
            "Assumption": "PO-FS gap",
            "Details": "gap_days = fs_starting_date - po_completion_date. Negative gaps are flagged in Data_Issues and retained in PO_FS_Gap; summary stats clamp negatives to 0. Missing dates are excluded from stats.",
        },
        {
            "Assumption": "Erection-PO gap",
            "Details": "gap_days = po_start_date - last erection completion date within the span (inclusive). Negative gaps are flagged and clamped to 0. Missing dates or unmatched spans are logged in Data_Issues.",
        },
        {
            "Assumption": "Location parsing",
            "Details": "From/To AP values strip a leading 'AP' (case-insensitive) with optional separators. Parsed as main number + optional letters + optional '/sub'. Order key = main*1,000,000 + letter_rank*1,000 + sub.",
        },
        {
            "Assumption": "Gantry spans",
            "Details": "From/To AP values containing 'gantry' or 'gty' are treated as line ends. If To_AP is gantry, the span runs from From_AP to the max location for that project; if From_AP is gantry, the span runs from the min location to To_AP. Both ends marked gantry are flagged as data issues.",
        },
        {
            "Assumption": "Length units",
            "Details": "Span lengths are auto-detected per source_file (fallback per project): if p90 <= 50 and max <= 100, values are treated as KM; otherwise meters. KM values are converted to meters for length_m and used directly for length_km.",
        },
        {
            "Assumption": "Span matching",
            "Details": "Span ranges are inclusive of From and To locations. If From/To order is reversed, the range is normalized.",
        },
        {
            "Assumption": "Project matching",
            "Details": "Projects are matched using a compact key (lowercase, alphanumeric only).",
        },
        {
            "Assumption": "Date filters",
            "Details": "PO-FS gaps are filtered by PO completion date; Erection-PO gaps are filtered by PO start date; daily outputs are filtered by the daily date column.",
        },
        {
            "Assumption": "Date parsing",
            "Details": "Stringing dates support mixed formats. Values like DD.MM.YY or DD/MM/YYYY are parsed day-first; Excel serials are supported.",
        },
        {
            "Assumption": "Method split",
            "Details": "Method is read from the stringing daily data (method/method_norm). Rows without method are logged in Data_Issues. Summary groups into TSE, Manual, and Other.",
        },
        {
            "Assumption": "Pareto analysis",
            "Details": "Pareto buckets use gang-level avg_km_month (mean daily_km * 30). KM % Delivered is based on sum(length_km) from compiled rows for gangs in each bucket; counts and percentages reflect unique gangs per bucket.",
        },
        {
            "Assumption": "Monthly totals",
            "Details": "Review_Format monthly tables use sum(length_km) grouped by F/S completion month and count unique gangs per month.",
        },
    ]
    readme_df = pd.DataFrame(readme_rows)

    output_path = _resolve_output_path(args.output)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        current_row = 0
        current_row = _write_labeled_table(
            writer, "Summary", "TSE", summary_rows_tse, startrow=current_row
        )
        current_row += 1
        _write_labeled_table(writer, "Summary", "Overall", summary_rows, startrow=current_row)

        current_row = 0
        current_row = _write_labeled_table(
            writer, "Gang_Productivity", "TSE", gang_rows_tse, startrow=current_row
        )
        current_row += 1
        _write_labeled_table(
            writer, "Gang_Productivity", "Overall", gang_rows, startrow=current_row
        )

        current_row = 0
        current_row = _write_labeled_table(
            writer, "PO_FS_Gap", "TSE", po_fs_gap_tse, startrow=current_row
        )
        current_row += 1
        _write_labeled_table(
            writer, "PO_FS_Gap", "Overall", po_fs_gap, startrow=current_row
        )

        current_row = 0
        current_row = _write_labeled_table(
            writer, "PO_FS_Gap_Summary", "TSE", po_fs_gap_summary_tse, startrow=current_row
        )
        current_row += 1
        _write_labeled_table(
            writer, "PO_FS_Gap_Summary", "Overall", po_fs_gap_summary, startrow=current_row
        )

        current_row = 0
        current_row = _write_labeled_table(
            writer, "Erection_PO_Gap", "TSE", erection_po_gap_tse, startrow=current_row
        )
        current_row += 1
        _write_labeled_table(
            writer, "Erection_PO_Gap", "Overall", erection_po_gap, startrow=current_row
        )

        current_row = 0
        current_row = _write_labeled_table(
            writer, "Erection_PO_Gap_Summary", "TSE", erection_po_gap_summary_tse, startrow=current_row
        )
        current_row += 1
        _write_labeled_table(
            writer,
            "Erection_PO_Gap_Summary",
            "Overall",
            erection_po_gap_summary,
            startrow=current_row,
        )
        if not method_summary.empty:
            current_row = 0
            method_summary.to_excel(writer, sheet_name="Method_Summary", index=False, startrow=current_row)
            current_row += len(method_summary) + 2

            pareto_overall = _build_pareto_table(scope, compiled_scope)
            header = pd.DataFrame(
                [{
                    "Bucket": "Pareto Analysis - Overall",
                    "KM Delivered": "",
                    "Number of Gangs": "",
                    "KM % Delivered": "",
                    "Number of Gangs %": "",
                }],
                columns=["Bucket", "KM Delivered", "Number of Gangs", "KM % Delivered", "Number of Gangs %"],
            )
            header.to_excel(writer, sheet_name="Method_Summary", index=False, header=False, startrow=current_row)
            current_row += 1
            pareto_overall.to_excel(writer, sheet_name="Method_Summary", index=False, startrow=current_row)
            _apply_pareto_formulas(
                writer, "Method_Summary", startrow=current_row, table=pareto_overall
            )
            current_row += len(pareto_overall) + 2

            pareto_tse = _build_pareto_table(scope, compiled_scope, method_filter="TSE")
            header = pd.DataFrame(
                [{
                    "Bucket": "Pareto Analysis - TSE",
                    "KM Delivered": "",
                    "Number of Gangs": "",
                    "KM % Delivered": "",
                    "Number of Gangs %": "",
                }],
                columns=["Bucket", "KM Delivered", "Number of Gangs", "KM % Delivered", "Number of Gangs %"],
            )
            header.to_excel(writer, sheet_name="Method_Summary", index=False, header=False, startrow=current_row)
            current_row += 1
            pareto_tse.to_excel(writer, sheet_name="Method_Summary", index=False, startrow=current_row)
            _apply_pareto_formulas(
                writer, "Method_Summary", startrow=current_row, table=pareto_tse
            )
        stringing_summary_overall_sheet = "Stringing Summary Overall"
        stringing_summary_monthly_sheet = "Stringing Summary Monthly"
        if not overall_rollup.empty or not pch_rollup.empty or not project_rollup.empty:
            summary_row = 0
            summary_row = _write_labeled_table(
                writer, stringing_summary_overall_sheet, "TSE", pd.DataFrame(), startrow=summary_row
            )
            scope_label_tse = _build_scope_label(scope_tse, start_date, end_date).replace("Scope:", "Scope (TSE):")
            summary_row = _write_scoped_table(
                writer, stringing_summary_overall_sheet, tse_overall_rollup, scope_label_tse, startrow=summary_row
            )
            summary_row += 2
            summary_row = _write_scoped_table(
                writer, stringing_summary_overall_sheet, tse_pch_rollup, scope_label_tse, startrow=summary_row
            )
            summary_row += 2
            summary_row = _write_scoped_table(
                writer, stringing_summary_overall_sheet, tse_project_rollup, scope_label_tse, startrow=summary_row
            )
            summary_row += 2
            summary_row = _write_labeled_table(
                writer, stringing_summary_overall_sheet, "Overall", pd.DataFrame(), startrow=summary_row
            )
            summary_row = _write_scoped_table(
                writer, stringing_summary_overall_sheet, overall_rollup, scope_label, startrow=summary_row
            )
            summary_row += 2
            summary_row = _write_scoped_table(
                writer, stringing_summary_overall_sheet, pch_rollup, scope_label, startrow=summary_row
            )
            summary_row += 2
            _write_scoped_table(
                writer, stringing_summary_overall_sheet, project_rollup, scope_label, startrow=summary_row
            )

        monthly_row = 0
        monthly_blocks: list[dict[str, int]] = []
        monthly_row = _write_labeled_table(
            writer, stringing_summary_monthly_sheet, "TSE", pd.DataFrame(), startrow=monthly_row
        )
        scope_label_tse = _build_scope_label(scope_tse, start_date, end_date).replace("Scope:", "Scope (TSE):")
        monthly_row, monthly_block = _write_monthly_scoped_table(
            writer,
            stringing_summary_monthly_sheet,
            monthly_tse_overall,
            scope_label_tse,
            month_headers=monthly_labels,
            group_headers=[],
            startrow=monthly_row,
        )
        monthly_blocks.append(monthly_block)
        monthly_row += 2
        monthly_row, monthly_block = _write_monthly_scoped_table(
            writer,
            stringing_summary_monthly_sheet,
            monthly_tse_pch,
            scope_label_tse,
            month_headers=monthly_labels,
            group_headers=["PCH"],
            startrow=monthly_row,
        )
        monthly_blocks.append(monthly_block)
        monthly_row += 2
        monthly_row, monthly_block = _write_monthly_scoped_table(
            writer,
            stringing_summary_monthly_sheet,
            monthly_tse_project,
            scope_label_tse,
            month_headers=monthly_labels,
            group_headers=["PCH", "Project"],
            startrow=monthly_row,
        )
        monthly_blocks.append(monthly_block)
        monthly_row += 2
        monthly_row = _write_labeled_table(
            writer, stringing_summary_monthly_sheet, "Overall", pd.DataFrame(), startrow=monthly_row
        )
        monthly_row, monthly_block = _write_monthly_scoped_table(
            writer,
            stringing_summary_monthly_sheet,
            monthly_overall,
            scope_label,
            month_headers=monthly_labels,
            group_headers=[],
            startrow=monthly_row,
        )
        monthly_blocks.append(monthly_block)
        monthly_row += 2
        monthly_row, monthly_block = _write_monthly_scoped_table(
            writer,
            stringing_summary_monthly_sheet,
            monthly_pch,
            scope_label,
            month_headers=monthly_labels,
            group_headers=["PCH"],
            startrow=monthly_row,
        )
        monthly_blocks.append(monthly_block)
        monthly_row += 2
        _, monthly_block = _write_monthly_scoped_table(
            writer,
            stringing_summary_monthly_sheet,
            monthly_project,
            scope_label,
            month_headers=monthly_labels,
            group_headers=["PCH", "Project"],
            startrow=monthly_row,
        )
        monthly_blocks.append(monthly_block)
        _style_monthly_summary_sheet(writer, stringing_summary_monthly_sheet, monthly_blocks)
        if review_table_one_groups or review_table_two_groups:
            current_row = 0
            for table in review_table_one_groups:
                if table.empty:
                    continue
                table.to_excel(writer, sheet_name="Review_Format", index=False, startrow=current_row)
                current_row += len(table) + 2
            current_row += 1
            for table, include_header in review_table_two_groups:
                if table.empty:
                    continue
                table.to_excel(
                    writer,
                    sheet_name="Review_Format",
                    index=False,
                    header=include_header,
                    startrow=current_row,
                )
                current_row += len(table) + (1 if include_header else 0) + 1
        readme_df.to_excel(writer, sheet_name="README", index=False)
        if not issues_df.empty:
            issues_df.to_excel(writer, sheet_name="Data_Issues", index=False)
        for styled_sheet in writer.book.sheetnames:
            if styled_sheet == stringing_summary_monthly_sheet:
                continue
            _style_clean_sheet(writer, styled_sheet)

    LOGGER.info("Stringing summary exported to %s", output_path)
    print(f"[export] Wrote '{output_path}'")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
