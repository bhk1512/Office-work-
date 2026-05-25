"""Export a CMD presentation DPR audit workbook and a conservative PPT fill copy.

This is a one-time/analysis exporter for the CMD review deck. It uses the prior
scope workbook as the table baseline, the current Raw Data/DPRs folder as the
evidence base, and existing normalized parquet outputs for high-confidence
execution/status fields.
"""

from __future__ import annotations

import argparse
import math
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from erection_compiled_to_daily_new import scrub_defined_names_from_workbook


REPO_ROOT = Path(__file__).resolve().parent
DEFAULT_RUN_DATE = "2026-05-22"
DEFAULT_SCOPE_WORKBOOK = REPO_ROOT / "Productivity Summaries" / "CMD_Presentation_DPR_Automation_Scope_2026-05-21.xlsx"
DEFAULT_DPR_DIR = REPO_ROOT / "Raw Data" / "DPRs"
DEFAULT_PARQUET_ROOT = REPO_ROOT / "Parquets"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "Productivity Summaries"
DEFAULT_PPT_TEMPLATE = Path.home() / "Downloads" / "CMD Presentation R2 18.05.2026.pptx"

STATUS_PROJECTS = {
    "TA 413",
    "TA 414",
    "TA 416",
    "TA 418",
    "TA 419",
    "TA 421",
    "TA 504",
    "TA 505",
    "TA 506",
    "TA 509",
    "TA 510",
    "TA 512",
    "TA 513",
    "TA 515",
    "TA 601",
    "TA 602",
}

EXTRA_DECK_WITH_RAW = {"TA 310", "TA 325"}
MATCHING_SCOPE_PROJECTS = STATUS_PROJECTS | EXTRA_DECK_WITH_RAW

NON_FILLABLE_FAMILIES = {
    "commercial",
    "pending_amendments",
    "running_project_master",
    "project_summary",
    "priority_projects",
}

PPT_CELL_COL_RE = re.compile(r"^col_(\d+)$")


@dataclass(frozen=True)
class TableBlock:
    slide: int
    slide_title: str
    slide_type: str
    table_no: int
    project_codes: tuple[str, ...]
    family: str
    rows: tuple[tuple[str, ...], ...]
    supply_start_row: int | None = None


@dataclass(frozen=True)
class SourceRef:
    source_file: str = ""
    source_sheet: str = ""
    source_row: str = ""
    source_cell: str = ""
    source_field: str = ""
    source_note: str = ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).replace("\u00a0", " ")
    text = text.replace("\r", " ").replace("\n", " ")
    text = text.replace("–", "-").replace("—", "-")
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).lower()
    text = text.replace("qty.", "qty").replace("nos.", "nos")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def compact_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_text(value).lower())


def normalize_project_code(value: Any) -> str:
    """Normalize deck/file project references into a stable display code."""
    text = normalize_text(value).upper()
    if not text:
        return ""

    match = re.search(r"\b(TA|TB|SA)\s*-?\s*(\d{3})\b", text)
    if match:
        return f"{match.group(1)} {match.group(2)}"

    match = re.search(r"\bA\s*-?\s*(\d{3}\s*/\s*\d{3})\b", text)
    if match:
        return "A-" + re.sub(r"\s+", "", match.group(1))

    match = re.search(r"\bA\s*-?\s*(\d{3})\b", text)
    if match:
        return "A" + match.group(1)

    return text


def project_sheet_name(project_code: str) -> str:
    base = re.sub(r"[^A-Za-z0-9]+", "", project_code) or "Project"
    return base[:31]


def extract_project_codes(value: Any) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    codes: list[str] = []
    patterns = (
        r"\b(?:TA|TB|SA)\s*-?\s*\d{3}\b",
        r"\bA\s*-?\s*\d{3}\s*/\s*\d{3}\b",
        r"\bA\s*-?\s*\d{3}\b",
    )
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            code = normalize_project_code(match.group(0))
            if code and code not in codes:
                codes.append(code)
    return codes


def split_slide_project_codes(value: Any) -> tuple[str, ...]:
    codes: list[str] = []
    for part in re.split(r",|;", normalize_text(value)):
        code = normalize_project_code(part)
        if code and code not in codes:
            codes.append(code)
    return tuple(codes)


def maybe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = normalize_text(value)
        if not cleaned or cleaned in {"-", "--"}:
            return None
        if cleaned.lower() in {"comp", "completed", "nil"}:
            return None
        cleaned = cleaned.replace(",", "").replace("%", "")
        value = cleaned
    try:
        result = pd.to_numeric([value], errors="coerce")[0]
    except Exception:
        return None
    if pd.isna(result):
        return None
    return float(result)


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return normalize_text(value) == ""


def clean_scalar(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, float):
        if math.isfinite(value) and abs(value - round(value)) < 1e-9:
            return str(int(round(value)))
        return f"{value:.3f}".rstrip("0").rstrip(".")
    return normalize_text(value)


def compare_values(presentation_value: Any, generated_value: Any) -> str:
    if is_blank(generated_value):
        return "not_generated"
    if is_blank(presentation_value):
        return "missing_in_presentation"

    left_num = maybe_float(presentation_value)
    right_num = maybe_float(generated_value)
    if left_num is not None and right_num is not None:
        return "match" if abs(left_num - right_num) <= 0.01 else "differs"

    left = compact_key(presentation_value)
    right = compact_key(generated_value)
    return "match" if left == right else "differs"


def first_date_key(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    match = re.search(r"(\d{1,2})[-/ .](\d{1,2}|[A-Za-z]{3,9})[-/ .](\d{2,4})", text)
    if not match:
        return ""
    parsed = pd.to_datetime(match.group(0), errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return ""
    return parsed.strftime("%Y-%m-%d")


def should_keep_ppt_update(
    *,
    row_family: str,
    availability: str,
    presentation_value: Any,
    generated_value: Any,
    source_note: str,
    requested_update: bool,
) -> bool:
    if not requested_update or availability not in {"direct_from_dpr", "derived_from_dpr"} or is_blank(generated_value):
        return False
    if row_family == "contract_header":
        old_date = first_date_key(presentation_value)
        new_date = first_date_key(generated_value)
        if old_date and new_date:
            return old_date == new_date
        return is_blank(presentation_value)
    if row_family != "supply_status":
        return True
    if "summary row" not in normalize_text(source_note).lower():
        return False
    old_num = maybe_float(presentation_value)
    new_num = maybe_float(generated_value)
    if old_num is not None and new_num is not None:
        if abs(old_num) < 1e-9:
            return abs(new_num) < 1e-9
        ratio = abs(new_num / old_num)
        return 0.5 <= ratio <= 2.0
    return True


def should_blank_unfilled_ppt_cell(row: pd.Series) -> bool:
    if bool(row.get("ppt_update", False)):
        return False
    availability = normalize_text(row.get("availability", ""))
    if availability in {"deck_structure", "not_applicable"}:
        return False
    row_family = normalize_text(row.get("row_family", ""))
    table_family = normalize_text(row.get("table_family", ""))
    row_no = int(row.get("row_no", 0) or 0)
    col_no = int(row.get("col_no", 0) or 0)
    field = normalize_text(row.get("field", ""))
    presentation_value = row.get("presentation_value", "")
    if is_blank(presentation_value):
        return False

    if row_family == "contract_header":
        return availability in {"parser_needed", "external_required", "no_matching_dpr"}

    if row_family == "activity_status":
        if field:
            return True
        return availability in {"no_matching_dpr", "out_of_first_pass_scope"} and row_no > 2 and col_no >= 3

    if row_family == "supply_status":
        if field:
            return True
        return availability in {"no_matching_dpr", "out_of_first_pass_scope"} and row_no > 2 and col_no >= 3

    if row_family == "front_availability":
        return row_no >= 4 and col_no >= 3

    if row_family in {"monthly_plan", "monthly_supply_plan"}:
        return row_no >= 3 and col_no >= 4

    if table_family == "project_summary":
        return row_no > 1 and col_no >= 3

    if table_family == "running_project_master":
        return row_no > 1 and col_no >= 6

    if table_family in {"priority_projects", "pending_amendments"}:
        return row_no > 1 and col_no >= 4

    if table_family == "commercial":
        return row_no > 1 and col_no >= 5

    if availability in {"parser_needed", "planning_not_captured", "history_unavailable", "external_required", "no_matching_dpr"}:
        return bool(field)
    return False


def classify_supply_item(value: Any) -> str:
    text = normalize_key(value)
    if not text:
        return ""
    if "earthing" in text and "earthwire" not in text and "earth_wire" not in text:
        return ""
    if "tower_accessor" in text or "accessor" in text and "tower" in text:
        return ""
    if "stub" in text:
        return "stub"
    if "tower" in text and "erection" not in text and "bolt" not in text and "nut" not in text and "plate" not in text:
        return "tower"
    if "conductor" in text:
        return "conductor"
    if "earthwire" in text or "earth_wire" in text or "earth_wire" in text:
        return "earthwire"
    if "opgw" in text:
        return "opgw"
    if "160" in text and ("insulator" in text or "clr" in text or "kn" in text):
        return "insulator_160"
    if "210" in text and ("insulator" in text or "clr" in text or "kn" in text):
        return "insulator_210"
    if "insulator" in text or "clr" in text:
        return "insulator"
    if "hardware" in text or "fitting" in text or "fittings" in text:
        return "hardware"
    return ""


def activity_norm_for_label(value: Any) -> str:
    text = normalize_key(value)
    if not text:
        return ""
    if "opgw" in text:
        return "opgw_stringing"
    if "check_survey" in text:
        return "check_survey"
    if "detailed_survey" in text or text == "detail_survey":
        return "detailed_survey"
    if "route_alignment" in text:
        return "route_alignment"
    if "soil" in text and "invest" in text:
        return "soil_investigation"
    if "foundation" in text:
        return "foundation"
    if "tower_erection" in text or text == "erection":
        return "tower_erection"
    if "tack" in text and "weld" in text:
        return "tack_welding"
    if "earthing" in text:
        return "earthing"
    if "stringing" in text:
        return "stringing"
    if "paying" in text:
        return "paying_out"
    if "final_sag" in text:
        return "final_sag"
    return ""


def classify_table_family(slide_type: Any, slide_title: Any, rows: Iterable[Iterable[Any]]) -> str:
    row_lists = [[normalize_text(cell) for cell in row] for row in rows]
    table_text = " ".join(" ".join(row) for row in row_lists)
    table_key = normalize_key(table_text)
    title_key = normalize_key(slide_title)
    slide_type_key = normalize_key(slide_type)

    if slide_type_key == "issues_support" or "issues_support_required" in title_key:
        return "issues_support"
    if slide_type_key == "commercial" or "pending_claim" in title_key or "financial_detail" in title_key:
        return "commercial"
    if slide_type_key == "running_project_master" or "project_value" in table_key and "ueob" in table_key:
        return "running_project_master"
    if "pending_amendment" in title_key:
        return "pending_amendments"
    if "priority_projects" in title_key or "priority_projects" in table_key:
        return "priority_projects"
    if slide_type_key == "project_summary" or "summary_of_projects_under_execution" in title_key:
        return "project_summary"
    if "loa_date" in table_key and "contractual_completion" in table_key:
        return "contract_header"
    if "front_availability_status" in table_key:
        return "front_availability"
    if "no_of_month_from_noa" in table_key or "completion_plan" in table_key and "front_required" in table_key:
        return "monthly_plan"
    if "supply_status" in table_key:
        return "activity_supply_status"
    if "activity" in table_key and ("as_per_loa" in table_key or "completed" in table_key):
        return "activity_status"
    return slide_type_key or "unknown"


def classify_supply_field(header_value: Any) -> str:
    text = normalize_key(header_value)
    if not text:
        return ""
    if "balance_to_receive" in text or "balance_to_received" in text:
        return "balance"
    if "total_received" in text or "actual_as_on_date" in text or "cum_completed" in text or "cumulative_completed" in text:
        return "actual_supplied"
    if text in {"received", "received_qty", "receipt_qty"}:
        return "actual_supplied_secondary"
    if "actual" in text or "supplied" in text or "received" in text or text in {"rec", "receipt"}:
        return "actual_supplied"
    if "balance" in text or text == "bal":
        return "balance"
    if "as_per_l2" in text or text == "l2" or "l2_sch" in text or "released_for_supply" in text:
        return "l2_qty"
    if "final_est" in text or "latest_estimate" in text or "estimated_qty" in text or "estimate" in text or "revised" in text or "line_qty" in text:
        return "revised_qty"
    if "loa" in text or "total" in text or "scope" in text:
        return "quantity_loa"
    if "plan" in text:
        return "plan"
    return ""


def classify_activity_field(header_value: Any) -> str:
    text = normalize_key(header_value)
    if not text:
        return ""
    if "as_per_loa" in text:
        return "quantity_loa"
    if "revised" in text:
        return "revised_qty"
    if "as_per_l2" in text or text == "l2":
        return "l2_qty"
    if "completed" in text or "cumulative" in text or "cu_progress" in text:
        return "completed"
    if "balance" in text or "balnce" in text:
        return "balance"
    if "front_availability" in text or text == "front":
        return "front_availability"
    if "gangs_available" in text or "gang_available" in text or "gangs_working" in text:
        return "gangs_available"
    if "remarks" in text or "remark" in text:
        return "remarks"
    if "plan" in text and "may" in text:
        return "plan_current_month"
    if text in {"plan", "plan_may_26", "planmay_26"}:
        return "plan_current_month"
    if "act" in text and "may" in text:
        return "actual_current_month"
    return ""


def prefer_raw_dpr(existing: Path | None, candidate: Path) -> Path:
    if existing is None:
        return candidate
    existing_main = "[MAIN]" in existing.name.upper()
    candidate_main = "[MAIN]" in candidate.name.upper()
    if candidate_main and not existing_main:
        return candidate
    return max(existing, candidate, key=lambda path: path.stat().st_mtime)


def discover_raw_dprs(dpr_dir: Path) -> dict[str, Path]:
    mapping: dict[str, Path] = {}
    for path in sorted(dpr_dir.glob("*DPR*.*")):
        if path.name.startswith("~$"):
            continue
        code = normalize_project_code(path.name)
        if not code:
            continue
        mapping[code] = prefer_raw_dpr(mapping.get(code), path)
    return mapping


def dpr_date_from_name(path: Path | None) -> str:
    if path is None:
        return ""
    match = re.search(r"(20\d{2}-\d{2}-\d{2})", path.name)
    return match.group(1) if match else ""


def safe_read_parquet(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_parquet(path)
    except Exception:
        return pd.DataFrame()


def read_scope(scope_workbook: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    slides = pd.read_excel(scope_workbook, sheet_name="PPT Slides", dtype=str).fillna("")
    table_raw = pd.read_excel(scope_workbook, sheet_name="PPT Table Raw", dtype=str).fillna("")
    try:
        summary = pd.read_excel(scope_workbook, sheet_name="Presentation Summary Rows", dtype=str).fillna("")
    except Exception:
        summary = pd.DataFrame()
    for frame in (slides, table_raw):
        for column in ("slide", "table_no", "row_no"):
            if column in frame.columns:
                frame[column] = pd.to_numeric(frame[column], errors="coerce").fillna(0).astype(int)
    return slides, table_raw, summary


def ppt_value_columns(table_raw: pd.DataFrame) -> list[str]:
    columns: list[tuple[int, str]] = []
    for column in table_raw.columns:
        match = PPT_CELL_COL_RE.match(str(column))
        if match:
            columns.append((int(match.group(1)), column))
    return [column for _, column in sorted(columns)]


def build_table_blocks(slides: pd.DataFrame, table_raw: pd.DataFrame) -> list[TableBlock]:
    slide_meta: dict[int, dict[str, str]] = {}
    for _, row in slides.iterrows():
        slide_no = int(row.get("slide", 0) or 0)
        slide_meta[slide_no] = {
            "title": normalize_text(row.get("title", "")),
            "type": normalize_text(row.get("slide_type", "")),
            "project_codes": normalize_text(row.get("project_codes", "")),
        }

    value_columns = ppt_value_columns(table_raw)
    blocks: list[TableBlock] = []
    for (slide, table_no), group in table_raw.groupby(["slide", "table_no"], sort=True):
        ordered = group.sort_values("row_no")
        rows: list[tuple[str, ...]] = []
        for _, raw_row in ordered.iterrows():
            values = tuple(clean_scalar(raw_row.get(column, "")) for column in value_columns)
            rows.append(values)
        meta = slide_meta.get(int(slide), {})
        project_codes = split_slide_project_codes(meta.get("project_codes", ""))
        family = classify_table_family(meta.get("type", ""), meta.get("title", ""), rows)
        supply_start: int | None = None
        for idx, row_values in enumerate(rows, start=1):
            if "supply_status" in normalize_key(" ".join(row_values)):
                supply_start = idx
                break
        blocks.append(
            TableBlock(
                slide=int(slide),
                slide_title=meta.get("title", ""),
                slide_type=meta.get("type", ""),
                table_no=int(table_no),
                project_codes=project_codes,
                family=family,
                rows=tuple(rows),
                supply_start_row=supply_start,
            )
        )
    return blocks


def list_workbook_sheets(path: Path) -> tuple[list[str], str]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            return list(workbook.sheetnames), "ok"
        finally:
            workbook.close()
    except Exception as primary_error:
        scrubbed: Path | None = None
        try:
            scrubbed = scrub_defined_names_from_workbook(path)
            workbook = load_workbook(scrubbed, read_only=True, data_only=True, keep_links=False)
            try:
                return list(workbook.sheetnames), f"repaired_defined_names ({type(primary_error).__name__})"
            finally:
                workbook.close()
        except Exception as repair_error:
            return [], f"open_failed ({type(repair_error).__name__}: {repair_error})"
        finally:
            if scrubbed is not None:
                shutil.rmtree(scrubbed.parent, ignore_errors=True)


def load_openpyxl_workbook(path: Path):
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        return workbook, None, "ok"
    except Exception:
        scrubbed = scrub_defined_names_from_workbook(path)
        workbook = load_workbook(scrubbed, read_only=True, data_only=True, keep_links=False)
        return workbook, scrubbed, "repaired_defined_names"


def candidate_value_to_right(row: tuple[Any, ...], col_index: int) -> tuple[Any, int | None]:
    for next_col in range(col_index + 1, min(len(row), col_index + 5)):
        if not is_blank(row[next_col]):
            return row[next_col], next_col
    return "", None


def classify_project_detail_label(label: Any) -> str:
    text = normalize_key(label)
    if not text:
        return ""
    if ("loa_end" in text or "contractual_completion" in text or "completion_date" in text) and "planned" not in text:
        return "contractual_completion"
    if ("loa_start" in text or "noa_start" in text or "noa_date" in text or "loa_date" in text) and "end" not in text:
        return "loa_date"
    if ("loa" in text or "contract" in text or "noa" in text) and ("value" in text or "amount" in text):
        return "loa_value"
    if "contractual" in text and "completion" in text:
        return "contractual_completion"
    if ("planned" in text or "target" in text or "expected" in text) and "completion" in text:
        return "planned_completion"
    return ""


def format_date_like(value: Any) -> str:
    if is_blank(value):
        return ""
    try:
        timestamp = pd.to_datetime(value, errors="coerce")
    except Exception:
        timestamp = pd.NaT
    if pd.notna(timestamp):
        return timestamp.strftime("%d/%m/%Y")
    text = normalize_text(value)
    match = re.search(r"(\d{1,2})[-/ .]([A-Za-z]{3,9}|\d{1,2})[-/ .](\d{2,4})", text)
    if match:
        parsed = pd.to_datetime(match.group(0), errors="coerce", dayfirst=True)
        if pd.notna(parsed):
            return parsed.strftime("%d/%m/%Y")
    return text


def project_detail_record(
    *,
    project_code: str,
    field: str,
    value: Any,
    workbook_path: Path,
    sheet_name: str,
    source_row: int,
    source_col: int,
    source_label: Any,
    open_status: str,
) -> dict[str, Any]:
    formatted = format_date_like(value) if field in {"loa_date", "contractual_completion", "planned_completion"} else clean_scalar(value)
    return {
        "project_code": project_code,
        "field": field,
        "value": formatted,
        "source_file": workbook_path.name,
        "source_sheet": sheet_name,
        "source_row": source_row,
        "source_cell": f"'{sheet_name}'!{get_column_letter(source_col)}{source_row}",
        "source_label": clean_scalar(source_label),
        "open_status": open_status,
    }


def extract_project_details(workbook_path: Path, project_code: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    workbook = None
    scrubbed = None
    try:
        workbook, scrubbed, open_status = load_openpyxl_workbook(workbook_path)
        candidate_sheets = [
            sheet
            for sheet in workbook.sheetnames
            if any(token in normalize_key(sheet) for token in ("project_details", "summary", "dpr"))
        ]
        for sheet_name in candidate_sheets[:8]:
            worksheet = workbook[sheet_name]
            max_row = min(worksheet.max_row or 1, 80)
            max_col = min(worksheet.max_column or 1, 20)
            structured_found = False
            for header_row in worksheet.iter_rows(min_row=1, max_row=min(max_row, 8), max_col=max_col):
                header_values = [cell.value for cell in header_row]
                fields_by_col = {
                    cell.column: classify_project_detail_label(cell.value)
                    for cell in header_row
                    if classify_project_detail_label(cell.value)
                }
                if not fields_by_col:
                    continue
                next_row_number = header_row[0].row + 1
                if next_row_number > max_row:
                    continue
                next_values = list(worksheet.iter_rows(min_row=next_row_number, max_row=next_row_number, max_col=max_col))[0]
                for value_cell in next_values:
                    field = fields_by_col.get(value_cell.column)
                    if not field or is_blank(value_cell.value):
                        continue
                    label = header_values[value_cell.column - 1] if value_cell.column - 1 < len(header_values) else field
                    rows.append(
                        project_detail_record(
                            project_code=project_code,
                            field=field,
                            value=value_cell.value,
                            workbook_path=workbook_path,
                            sheet_name=sheet_name,
                            source_row=value_cell.row,
                            source_col=value_cell.column,
                            source_label=label,
                            open_status=open_status,
                        )
                    )
                    structured_found = True
                if structured_found:
                    break

            if structured_found:
                continue

            for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                values = tuple(cell.value for cell in row)
                for cell in row:
                    field = classify_project_detail_label(cell.value)
                    if not field:
                        continue
                    value, value_col = candidate_value_to_right(values, cell.column - 1)
                    if is_blank(value):
                        continue
                    rows.append(
                        project_detail_record(
                            project_code=project_code,
                            field=field,
                            value=value,
                            workbook_path=workbook_path,
                            sheet_name=sheet_name,
                            source_row=cell.row,
                            source_col=(value_col + 1) if value_col is not None else cell.column,
                            source_label=cell.value,
                            open_status=open_status,
                        )
                    )
    except Exception as exc:
        rows.append(
            {
                "project_code": project_code,
                "field": "project_details",
                "value": "",
                "source_file": workbook_path.name,
                "source_sheet": "",
                "source_row": "",
                "source_cell": "",
                "source_label": "",
                "open_status": f"open_failed ({type(exc).__name__}: {exc})",
            }
        )
    finally:
        if workbook is not None:
            workbook.close()
        if scrubbed is not None:
            shutil.rmtree(scrubbed.parent, ignore_errors=True)
    return rows


def detect_supply_header(values: list[Any]) -> tuple[dict[int, str], list[int]] | None:
    fields: dict[int, str] = {}
    description_cols: list[int] = []
    for idx, value in enumerate(values):
        key = normalize_key(value)
        if "description" in key or key in {"item", "particulars", "material", "tower_type"}:
            description_cols.append(idx)
        field = classify_supply_field(value)
        if field:
            fields[idx] = field
    useful_fields = {field for field in fields.values() if field not in {"plan"}}
    if description_cols and len(useful_fields) >= 2:
        return fields, description_cols
    return None


def supply_item_score(description: str, category: str) -> int:
    key = normalize_key(description)
    if not key:
        return 0
    exact_by_category = {
        "stub": {"stub", "stubs"},
        "tower": {"tower", "towers"},
        "conductor": {"conductor"},
        "earthwire": {"earthwire", "earth_wire"},
        "insulator": {"insulator", "insulators"},
        "hardware": {"hardware", "hardware_fittings", "hw_fittings"},
    }
    if category in exact_by_category and key in exact_by_category[category]:
        return 100
    if category == "insulator_160" and "160" in key:
        return 100
    if category == "insulator_210" and "210" in key:
        return 100
    if category == "conductor" and "accessor" in key:
        return 0
    if category == "earthwire" and "accessor" in key:
        return 0
    return 20


def numeric_or_completion(value: Any) -> bool:
    return maybe_float(value) is not None or compact_key(value) in {"comp", "completed", "nil"}


def aggregate_supply_candidates(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for candidate in candidates:
        grouped[(candidate["item_category"], candidate["field"])].append(candidate)

    rows: list[dict[str, Any]] = []
    for (category, field), group in grouped.items():
        exact = [row for row in group if int(row.get("score", 0)) >= 80]
        chosen_group = exact or group
        numeric_values = [maybe_float(row.get("value")) for row in chosen_group]
        numeric_values = [value for value in numeric_values if value is not None]
        if exact:
            selected = exact[0]
            value = selected["value"]
            note = "summary row"
        elif numeric_values:
            selected = chosen_group[0]
            value = sum(numeric_values)
            note = f"aggregated {len(numeric_values)} detail row(s)"
        else:
            selected = chosen_group[0]
            value = selected["value"]
            note = "detail row"
        out = dict(selected)
        out["value"] = clean_scalar(value)
        out["header_context"] = f"{selected.get('header_context', '')} | {note}".strip(" |")
        rows.append(out)
    return rows


def extract_supply_signals(workbook_path: Path, project_code: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    workbook = None
    scrubbed = None
    try:
        workbook, scrubbed, open_status = load_openpyxl_workbook(workbook_path)
        supply_sheets = [
            sheet
            for sheet in workbook.sheetnames
            if any(token in normalize_key(sheet) for token in ("supply", "store_stock", "material_status"))
        ]
        for sheet_name in supply_sheets:
            worksheet = workbook[sheet_name]
            max_row = min(worksheet.max_row or 1, 400)
            max_col = min(worksheet.max_column or 1, 80)
            scanned_rows: list[list[Any]] = []
            scanned_cells: list[list[Any]] = []
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                scanned_cells.append(list(row))
                scanned_rows.append([cell.value for cell in row])

            active_header: tuple[dict[int, str], list[int], int] | None = None
            sheet_candidates: list[dict[str, Any]] = []
            for row_idx, values in enumerate(scanned_rows):
                header = detect_supply_header(values)
                if header is not None:
                    active_header = (header[0], header[1], row_idx)
                    continue
                if active_header is None:
                    continue

                fields, description_cols, header_idx = active_header
                description_parts = [clean_scalar(values[idx]) for idx in description_cols if idx < len(values) and not is_blank(values[idx])]
                description = " ".join(description_parts)
                item_category = classify_supply_item(description)
                if not item_category:
                    continue
                score = supply_item_score(description, item_category)
                if score <= 0:
                    continue

                row_field_values: dict[str, tuple[int, Any]] = {}
                secondary_actual: tuple[int, Any] | None = None
                for col_idx, field in fields.items():
                    if col_idx >= len(values):
                        continue
                    value = values[col_idx]
                    if not numeric_or_completion(value):
                        continue
                    if field == "actual_supplied_secondary":
                        secondary_actual = (col_idx, value)
                        continue
                    row_field_values.setdefault(field, (col_idx, value))
                if "actual_supplied" not in row_field_values and secondary_actual is not None:
                    row_field_values["actual_supplied"] = secondary_actual

                row_number = row_idx + 1
                row_text = " ".join(clean_scalar(value) for value in values)
                for field, (col_idx, value) in row_field_values.items():
                    cell = scanned_cells[row_idx][col_idx]
                    sheet_candidates.append(
                        {
                            "project_code": project_code,
                            "item_category": item_category,
                            "item_raw": description,
                            "field": field,
                            "value": clean_scalar(value),
                            "source_file": workbook_path.name,
                            "source_sheet": sheet_name,
                            "source_row": row_number,
                            "source_cell": f"'{sheet_name}'!{cell.coordinate}",
                            "header_context": clean_scalar(scanned_rows[header_idx][col_idx]),
                            "row_context": row_text[:500],
                            "open_status": open_status,
                            "score": score,
                        }
                    )

                if not row_field_values:
                    first_col = description_cols[0] if description_cols else 0
                    first_cell = scanned_cells[row_idx][first_col]
                    sheet_candidates.append(
                        {
                            "project_code": project_code,
                            "item_category": item_category,
                            "item_raw": description,
                            "field": "item_row",
                            "value": "",
                            "source_file": workbook_path.name,
                            "source_sheet": sheet_name,
                            "source_row": row_number,
                            "source_cell": f"'{sheet_name}'!{first_cell.coordinate}",
                            "header_context": "",
                            "row_context": row_text[:500],
                            "open_status": open_status,
                            "score": score,
                        }
                    )
            rows.extend(aggregate_supply_candidates(sheet_candidates))
    except Exception as exc:
        rows.append(
            {
                "project_code": project_code,
                "item_category": "",
                "item_raw": "",
                "field": "supply_scan_error",
                "value": "",
                "source_file": workbook_path.name,
                "source_sheet": "",
                "source_row": "",
                "source_cell": "",
                "header_context": "",
                "row_context": "",
                "open_status": f"open_failed ({type(exc).__name__}: {exc})",
            }
        )
    finally:
        if workbook is not None:
            workbook.close()
        if scrubbed is not None:
            shutil.rmtree(scrubbed.parent, ignore_errors=True)
    return rows


def load_status_data(parquet_root: Path) -> pd.DataFrame:
    status = safe_read_parquet(parquet_root / "ProgressStatus" / "RawData.parquet")
    if status.empty:
        return status
    status = status.copy()
    status["project_code"] = status["project_code"].map(normalize_project_code)
    return status


def status_row_score(row: pd.Series, desired_activity: str) -> tuple[int, int, int]:
    raw = normalize_key(row.get("activity_raw", ""))
    if desired_activity == "foundation":
        bad = int("classification" in raw or "wip" in raw or "excavation" in raw)
        good = int(raw in {"foundation", "foundation_nos", "foundation_loc", "foundation_no"} or "foundation_nos" in raw)
        return (bad * -10 + good * 10, int(not is_blank(row.get("cumulative_progress"))), int(not is_blank(row.get("balance_progress"))))
    if desired_activity == "tower_erection":
        good = int("tower" in raw or raw == "erection")
        bad = int("loc" in raw and row.get("activity_norm") != "tower_erection")
        return (bad * -10 + good * 10, int(not is_blank(row.get("cumulative_progress"))), int(not is_blank(row.get("balance_progress"))))
    return (int(not is_blank(row.get("cumulative_progress"))), int(not is_blank(row.get("quantity_primary"))), 0)


def select_status_row(status: pd.DataFrame, project_code: str, activity_norm: str) -> pd.Series | None:
    if status.empty or not activity_norm:
        return None
    project_rows = status[status["project_code"].eq(project_code)]
    if project_rows.empty:
        return None

    if activity_norm == "tower_erection":
        candidates = project_rows[project_rows["activity_norm"].astype(str).isin(["tower_erection", "erection"])]
    else:
        candidates = project_rows[project_rows["activity_norm"].astype(str).eq(activity_norm)]
    if candidates.empty:
        return None

    ordered = sorted(
        (row for _, row in candidates.iterrows()),
        key=lambda row: status_row_score(row, activity_norm),
        reverse=True,
    )
    return ordered[0] if ordered else None


def build_supply_lookup(supply_rows: pd.DataFrame) -> dict[tuple[str, str, str], pd.Series]:
    lookup: dict[tuple[str, str, str], pd.Series] = {}
    if supply_rows.empty:
        return lookup
    for _, row in supply_rows.iterrows():
        field = normalize_text(row.get("field", ""))
        if field in {"", "item_row", "supply_scan_error"}:
            continue
        key = (normalize_project_code(row.get("project_code", "")), normalize_text(row.get("item_category", "")), field)
        if key not in lookup:
            lookup[key] = row
            continue
        current_score = maybe_float(lookup[key].get("score", 0)) or 0
        next_score = maybe_float(row.get("score", 0)) or 0
        current_value = maybe_float(lookup[key].get("value", ""))
        next_value = maybe_float(row.get("value", ""))
        if next_score > current_score or (next_score == current_score and next_value is not None and current_value is None):
            lookup[key] = row
    return lookup


def build_detail_lookup(detail_rows: pd.DataFrame) -> dict[tuple[str, str], pd.Series]:
    lookup: dict[tuple[str, str], pd.Series] = {}
    if detail_rows.empty:
        return lookup
    priority = {"loa_date": 3, "contractual_completion": 3, "planned_completion": 2, "loa_value": 1}
    for _, row in detail_rows.iterrows():
        project = normalize_project_code(row.get("project_code", ""))
        field = normalize_text(row.get("field", ""))
        value = row.get("value", "")
        if not project or not field or is_blank(value):
            continue
        key = (project, field)
        if key not in lookup or priority.get(field, 0) >= priority.get(normalize_text(lookup[key].get("field", "")), 0):
            lookup[key] = row
    return lookup


def supply_lookup_value(
    supply_lookup: dict[tuple[str, str, str], pd.Series],
    project_code: str,
    item: str,
    field: str,
) -> tuple[str, SourceRef] | None:
    row = supply_lookup.get((project_code, item, field))
    if row is not None:
        return clean_scalar(row.get("value", "")), supply_source_ref(row, field)
    if item == "insulator":
        parts: list[tuple[float, SourceRef]] = []
        for specific in ("insulator_160", "insulator_210"):
            specific_row = supply_lookup.get((project_code, specific, field))
            if specific_row is None:
                continue
            numeric = maybe_float(specific_row.get("value"))
            if numeric is not None:
                parts.append((numeric, supply_source_ref(specific_row, field)))
        if parts:
            value = sum(value for value, _ in parts)
            source = parts[0][1]
            source = SourceRef(
                source_file=source.source_file,
                source_sheet=source.source_sheet,
                source_row=source.source_row,
                source_cell=source.source_cell,
                source_field=field,
                source_note=f"aggregated {len(parts)} insulator subtype(s)",
            )
            return clean_scalar(value), source
    return None


def table_header_for_cell(block: TableBlock, row_no: int, col_no: int, row_family: str) -> str:
    rows = block.rows
    idx = col_no - 1
    if row_family == "supply_status" and block.supply_start_row:
        header_rows = range(block.supply_start_row, min(len(rows), block.supply_start_row + 2))
    elif block.family == "monthly_plan":
        header_rows = range(0, min(2, len(rows)))
    else:
        header_rows = range(0, min(2, len(rows)))
    parts: list[str] = []
    for header_row_idx in header_rows:
        if idx < len(rows[header_row_idx]) and rows[header_row_idx][idx]:
            parts.append(rows[header_row_idx][idx])
    return " ".join(parts)


def row_label_for(block: TableBlock, row_no: int) -> str:
    if row_no < 1 or row_no > len(block.rows):
        return ""
    row = block.rows[row_no - 1]
    for value in row[:3]:
        if normalize_text(value):
            return normalize_text(value)
    return ""


def row_family_for(block: TableBlock, row_no: int) -> str:
    row = block.rows[row_no - 1]
    row_key = normalize_key(" ".join(row))
    if block.family == "activity_supply_status":
        if block.supply_start_row and row_no > block.supply_start_row:
            return "supply_status"
        if row_key == "supply_status":
            return "supply_separator"
        return "activity_status"
    if block.family == "monthly_plan" and classify_supply_item(" ".join(row)):
        return "monthly_supply_plan"
    return block.family


def project_for_cell(block: TableBlock, row: tuple[str, ...]) -> str:
    row_codes = extract_project_codes(" ".join(row))
    if row_codes:
        return row_codes[0]
    if len(block.project_codes) == 1:
        return block.project_codes[0]
    return ""


def status_source_ref(row: pd.Series, field: str) -> SourceRef:
    source_row = clean_scalar(row.get("source_row_number", ""))
    source_sheet = clean_scalar(row.get("source_sheet", ""))
    return SourceRef(
        source_file=clean_scalar(row.get("source_file", "")),
        source_sheet=source_sheet,
        source_row=source_row,
        source_cell=f"'{source_sheet}'!{source_row}:{source_row}" if source_sheet and source_row else "",
        source_field=field,
        source_note="normalized ProgressStatus row",
    )


def supply_source_ref(row: pd.Series, field: str) -> SourceRef:
    return SourceRef(
        source_file=clean_scalar(row.get("source_file", "")),
        source_sheet=clean_scalar(row.get("source_sheet", "")),
        source_row=clean_scalar(row.get("source_row", "")),
        source_cell=clean_scalar(row.get("source_cell", "")),
        source_field=field,
        source_note=clean_scalar(row.get("header_context", "")) or "supply sheet scan",
    )


def detail_source_ref(row: pd.Series, field: str) -> SourceRef:
    return SourceRef(
        source_file=clean_scalar(row.get("source_file", "")),
        source_sheet=clean_scalar(row.get("source_sheet", "")),
        source_row=clean_scalar(row.get("source_row", "")),
        source_cell=clean_scalar(row.get("source_cell", "")),
        source_field=field,
        source_note=clean_scalar(row.get("source_label", "")) or "Project Details",
    )


def format_contract_header_cell(field: str, value: str) -> str:
    if field == "loa_date":
        return f"LOA DATE: {value}"
    if field == "contractual_completion":
        return f"CONTRACTUAL COMPLETION: {value}"
    if field == "loa_value":
        return f"LOA Value in Cr.: {value}"
    if field == "planned_completion":
        return f"PLANNED COMPLETION: {value}"
    return value


def direct_value_from_status(row: pd.Series, field: str) -> tuple[str, str, SourceRef]:
    if field == "quantity_loa":
        value = row.get("quantity_loa", "")
        if is_blank(value):
            return "", "parser_needed", SourceRef(source_note="LOA quantity not normalized for this activity")
        return clean_scalar(value), "direct_from_dpr", status_source_ref(row, field)
    if field == "revised_qty":
        value = row.get("quantity_estimated_or_total", "")
        if is_blank(value):
            value = row.get("quantity_primary", "")
        if is_blank(value):
            return "", "parser_needed", SourceRef(source_note="revised/total quantity not normalized for this activity")
        return clean_scalar(value), "direct_from_dpr", status_source_ref(row, field)
    if field == "completed":
        value = row.get("cumulative_progress", "")
        if is_blank(value):
            return "", "parser_needed", SourceRef(source_note="cumulative progress not normalized for this activity")
        return clean_scalar(value), "direct_from_dpr", status_source_ref(row, field)
    if field == "balance":
        value = row.get("balance_progress", "")
        if not is_blank(value):
            return clean_scalar(value), "direct_from_dpr", status_source_ref(row, field)
        total = maybe_float(row.get("quantity_primary", ""))
        completed = maybe_float(row.get("cumulative_progress", ""))
        if total is not None and completed is not None:
            return clean_scalar(total - completed), "derived_from_dpr", status_source_ref(row, "quantity_primary - cumulative_progress")
        return "", "parser_needed", SourceRef(source_note="balance cannot be derived from normalized status row")
    if field == "gangs_available":
        value = row.get("gangs_working", "")
        if is_blank(value):
            return "", "parser_needed", SourceRef(source_note="gang availability not populated in normalized status row")
        return clean_scalar(value), "direct_from_dpr", status_source_ref(row, field)
    if field == "plan_current_month":
        value = row.get("plan_for_month", "")
        if is_blank(value):
            return "", "planning_not_captured", SourceRef(source_note="current-month plan not populated in normalized status row")
        return clean_scalar(value), "direct_from_dpr", status_source_ref(row, field)
    if field == "actual_current_month":
        value = row.get("progress_for_month", "")
        if is_blank(value):
            return "", "parser_needed", SourceRef(source_note="month actual not populated in normalized status row")
        return clean_scalar(value), "direct_from_dpr", status_source_ref(row, field)
    if field == "remarks":
        value = row.get("remarks", "")
        if is_blank(value):
            return "", "parser_needed", SourceRef(source_note="remarks blank in normalized status row")
        return clean_scalar(value), "direct_from_dpr", status_source_ref(row, field)
    if field == "front_availability":
        return "", "parser_needed", SourceRef(source_note="front availability is visible in DPRs but not normalized in this pass")
    if field == "l2_qty":
        return "", "planning_not_captured", SourceRef(source_note="L2/planning baseline not standardized in current DPR automation")
    return "", "not_applicable", SourceRef()


def generated_for_cell(
    block: TableBlock,
    row_no: int,
    col_no: int,
    project_code: str,
    status: pd.DataFrame,
    supply_lookup: dict[tuple[str, str, str], pd.Series],
    detail_lookup: dict[tuple[str, str], pd.Series],
) -> tuple[str, str, str, SourceRef, bool]:
    row_family = row_family_for(block, row_no)
    row = block.rows[row_no - 1]
    header = table_header_for_cell(block, row_no, col_no, row_family)
    label = row_label_for(block, row_no)

    if not project_code:
        if row_family in NON_FILLABLE_FAMILIES:
            return "", "external_required", "", SourceRef(source_note="common deck table; no row-level matching DPR project"), False
        return "", "not_applicable", "", SourceRef(), False

    if row_family == "activity_status":
        activity = activity_norm_for_label(label)
        field = classify_activity_field(header)
        if not activity or not field:
            return "", "not_applicable", field, SourceRef(), False
        if field == "front_availability":
            return "", "parser_needed", field, SourceRef(source_note="front availability requires a dedicated DPR parser"), False
        status_row = select_status_row(status, project_code, activity)
        if status_row is None:
            return "", "parser_needed", field, SourceRef(source_note=f"no normalized status row for {project_code} {activity}"), False
        value, availability, source = direct_value_from_status(status_row, field)
        ppt_update = availability in {"direct_from_dpr", "derived_from_dpr"} and field not in {"remarks"}
        return value, availability, field, source, ppt_update

    if row_family == "supply_status":
        item = classify_supply_item(label)
        field = classify_supply_field(header)
        if not item or not field:
            return "", "parser_needed", field, SourceRef(source_note="supply row/category exists but field is not mapped"), False
        supply_value = supply_lookup_value(supply_lookup, project_code, item, field)
        if supply_value is None:
            return "", "parser_needed", field, SourceRef(source_note=f"supply parser did not find {item}/{field}"), False
        value, source = supply_value
        return value, "direct_from_dpr", field, source, True

    if row_family in {"front_availability"}:
        return "", "history_unavailable", "", SourceRef(source_note="latest DPR snapshot cannot reconstruct Feb/Mar/Apr history"), False
    if row_family in {"monthly_plan", "monthly_supply_plan"}:
        return "", "planning_not_captured", "", SourceRef(source_note="monthly planning matrix is not standardized in current DPR automation"), False
    if row_family in {"commercial", "running_project_master", "pending_amendments"}:
        return "", "external_required", "", SourceRef(source_note="requires commercial/amendment/running master tracker outside DPR execution sheets"), False
    if row_family == "project_summary":
        return "", "external_required", "", SourceRef(source_note="expected completion/value fields are not reliable from DPR-only data"), False
    if row_family == "priority_projects":
        return "", "external_required", "", SourceRef(source_note="priority/estimated completion table is a review tracker, not DPR execution data"), False
    if row_family == "contract_header":
        current_cell = block.rows[row_no - 1][col_no - 1] if col_no - 1 < len(block.rows[row_no - 1]) else ""
        field = classify_project_detail_label(current_cell)
        if field in {"loa_date", "contractual_completion", "loa_value", "planned_completion"}:
            detail_row = detail_lookup.get((project_code, field))
            if detail_row is not None and not is_blank(detail_row.get("value", "")):
                value = format_contract_header_cell(field, clean_scalar(detail_row.get("value", "")))
                return value, "direct_from_dpr", field, detail_source_ref(detail_row, field), True
            if field in {"planned_completion", "loa_value"}:
                return "", "external_required", field, SourceRef(source_note=f"{field} not available in standardized DPR project details"), False
            return "", "parser_needed", field, SourceRef(source_note=f"{field} not found in DPR project details"), False
        return "", "not_applicable", "", SourceRef(), False
    return "", "not_applicable", "", SourceRef(), False


def build_gap_matrix(
    blocks: list[TableBlock],
    raw_dprs: dict[str, Path],
    status: pd.DataFrame,
    supply_lookup: dict[tuple[str, str, str], pd.Series],
    detail_lookup: dict[tuple[str, str], pd.Series],
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for block in blocks:
        if block.family == "issues_support":
            continue
        for row_no, row in enumerate(block.rows, start=1):
            row_family = row_family_for(block, row_no)
            project_code = project_for_cell(block, row)
            if not project_code and len(block.project_codes) == 1:
                project_code = block.project_codes[0]
            has_matching_dpr = project_code in raw_dprs if project_code else False
            for col_no, presentation_value in enumerate(row, start=1):
                generated_value = ""
                availability = "not_applicable"
                field = ""
                source = SourceRef()
                ppt_update = False
                if project_code and not has_matching_dpr:
                    availability = "no_matching_dpr"
                    source = SourceRef(source_note="no matching RAW DPR file")
                elif (row_no <= 2 and block.family != "contract_header") or row_family == "supply_separator":
                    availability = "deck_structure"
                elif project_code in MATCHING_SCOPE_PROJECTS:
                    generated_value, availability, field, source, ppt_update = generated_for_cell(
                        block, row_no, col_no, project_code, status, supply_lookup, detail_lookup
                    )
                elif row_family in NON_FILLABLE_FAMILIES:
                    availability = "external_required"
                elif project_code:
                    availability = "out_of_first_pass_scope"

                comparison = compare_values(presentation_value, generated_value)
                if comparison == "differs" and availability in {"direct_from_dpr", "derived_from_dpr"}:
                    comparison = "differs_latest_dpr_vs_presentation"
                ppt_update = should_keep_ppt_update(
                    row_family=row_family,
                    availability=availability,
                    presentation_value=presentation_value,
                    generated_value=generated_value,
                    source_note=source.source_note,
                    requested_update=ppt_update,
                )

                records.append(
                    {
                        "slide": block.slide,
                        "slide_title": block.slide_title,
                        "slide_type": block.slide_type,
                        "table_no": block.table_no,
                        "row_no": row_no,
                        "col_no": col_no,
                        "project_code": project_code,
                        "table_family": block.family,
                        "row_family": row_family,
                        "row_label": row_label_for(block, row_no),
                        "field": field,
                        "presentation_value": presentation_value,
                        "generated_value": generated_value,
                        "availability": availability,
                        "comparison": comparison,
                        "source_file": source.source_file,
                        "source_sheet": source.source_sheet,
                        "source_row": source.source_row,
                        "source_cell": source.source_cell,
                        "source_field": source.source_field,
                        "source_note": source.source_note,
                        "ppt_update": bool(ppt_update and generated_value),
                    }
                )
    return pd.DataFrame(records)


def build_project_coverage(blocks: list[TableBlock], raw_dprs: dict[str, Path]) -> pd.DataFrame:
    project_to_slides: dict[str, set[int]] = defaultdict(set)
    project_to_issue_slides: dict[str, set[int]] = defaultdict(set)
    project_to_families: dict[str, set[str]] = defaultdict(set)
    deck_projects: set[str] = set()
    for block in blocks:
        codes = set(block.project_codes)
        for row in block.rows:
            codes.update(extract_project_codes(" ".join(row)))
        for code in codes:
            if not code:
                continue
            deck_projects.add(code)
            if block.family == "issues_support":
                project_to_issue_slides[code].add(block.slide)
            else:
                project_to_slides[code].add(block.slide)
                project_to_families[code].add(block.family)

    records: list[dict[str, Any]] = []
    for code in sorted(deck_projects):
        dpr_path = raw_dprs.get(code)
        sheets, open_status = ([], "no_matching_dpr")
        if dpr_path is not None:
            sheets, open_status = list_workbook_sheets(dpr_path)
        records.append(
            {
                "project_code": code,
                "in_first_pass_scope": code in MATCHING_SCOPE_PROJECTS,
                "matching_raw_dpr": "Yes" if dpr_path else "No",
                "raw_dpr_file": dpr_path.name if dpr_path else "",
                "raw_dpr_date": dpr_date_from_name(dpr_path),
                "open_read_status": open_status,
                "non_issue_slides": ", ".join(map(str, sorted(project_to_slides.get(code, set())))),
                "issue_slides_excluded": ", ".join(map(str, sorted(project_to_issue_slides.get(code, set())))),
                "table_families": ", ".join(sorted(project_to_families.get(code, set()))),
                "sheet_count": len(sheets),
                "available_sheets": "; ".join(sheets),
            }
        )
    return pd.DataFrame(records)


def build_summary(project_coverage: pd.DataFrame, gap_matrix: pd.DataFrame) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    scoped = project_coverage[project_coverage["in_first_pass_scope"].astype(bool)]
    records.append({"metric": "Deck projects", "value": len(project_coverage), "notes": ""})
    records.append({"metric": "First-pass scoped projects", "value": len(scoped), "notes": "CMD deck projects with matching RAW DPRs, plus TA 310/TA 325 references"})
    records.append({"metric": "Scoped projects with RAW DPR", "value": int((scoped["matching_raw_dpr"] == "Yes").sum()), "notes": ""})
    records.append({"metric": "Deck projects without RAW DPR", "value": int((project_coverage["matching_raw_dpr"] != "Yes").sum()), "notes": "Marked no_matching_dpr"})
    for availability, count in gap_matrix["availability"].value_counts().sort_index().items():
        records.append({"metric": f"Cells: {availability}", "value": int(count), "notes": ""})
    for comparison, count in gap_matrix["comparison"].value_counts().sort_index().items():
        records.append({"metric": f"Comparison: {comparison}", "value": int(count), "notes": ""})
    return pd.DataFrame(records)


def build_source_map(gap_matrix: pd.DataFrame, detail_rows: pd.DataFrame, supply_rows: pd.DataFrame) -> pd.DataFrame:
    source_cols = [
        "project_code",
        "table_family",
        "field",
        "generated_value",
        "availability",
        "source_file",
        "source_sheet",
        "source_row",
        "source_cell",
        "source_field",
        "source_note",
        "slide",
        "table_no",
        "row_no",
        "col_no",
    ]
    mapped = gap_matrix[
        gap_matrix["source_file"].astype(str).str.len().gt(0)
        | gap_matrix["source_note"].astype(str).str.len().gt(0)
    ][source_cols].copy()

    extras: list[pd.DataFrame] = []
    if not detail_rows.empty:
        detail = detail_rows.copy()
        detail["table_family"] = "contract_header_source_candidate"
        detail["generated_value"] = detail.get("value", "")
        detail["availability"] = detail["value"].map(lambda value: "direct_from_dpr" if not is_blank(value) else "parser_needed")
        detail["source_field"] = detail.get("field", "")
        detail["source_note"] = detail.get("source_label", "")
        for column in ("slide", "table_no", "row_no", "col_no"):
            detail[column] = ""
        extras.append(detail.reindex(columns=source_cols))
    if not supply_rows.empty:
        supply = supply_rows.copy()
        supply["table_family"] = "supply_source_candidate"
        supply["generated_value"] = supply.get("value", "")
        supply["availability"] = supply["field"].map(lambda field: "parser_needed" if field in {"item_row", "supply_scan_error"} else "direct_from_dpr")
        supply["source_field"] = supply.get("field", "")
        supply["source_note"] = supply.get("header_context", "")
        for column in ("slide", "table_no", "row_no", "col_no"):
            supply[column] = ""
        extras.append(supply.reindex(columns=source_cols))
    if extras:
        mapped = pd.concat([mapped, *extras], ignore_index=True)
    return mapped.drop_duplicates().reset_index(drop=True)


def write_dataframe(writer: pd.ExcelWriter, frame: pd.DataFrame, sheet_name: str) -> None:
    safe_name = sheet_name[:31]
    frame.to_excel(writer, sheet_name=safe_name, index=False)
    worksheet = writer.sheets[safe_name]
    worksheet.freeze_panes = "A2"
    for idx, column in enumerate(frame.columns, start=1):
        max_len = min(60, max([len(str(column)), *[len(clean_scalar(value)) for value in frame[column].head(500)]]) + 2)
        worksheet.column_dimensions[get_column_letter(idx)].width = max(10, max_len)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_project_sheets(writer: pd.ExcelWriter, blocks: list[TableBlock], gap_matrix: pd.DataFrame) -> None:
    workbook = writer.book
    projects = sorted(project for project in MATCHING_SCOPE_PROJECTS if project in set(gap_matrix["project_code"]))
    used_names: set[str] = set(writer.sheets)
    for project in projects:
        base_name = project_sheet_name(project)
        sheet_name = base_name
        suffix = 1
        while sheet_name in used_names:
            suffix += 1
            sheet_name = f"{base_name[:28]}{suffix}"
        used_names.add(sheet_name)
        worksheet = workbook.create_sheet(sheet_name)
        writer.sheets[sheet_name] = worksheet
        worksheet.freeze_panes = "A2"
        row_cursor = 1
        worksheet.cell(row_cursor, 1, f"{project} recreated presentation-shaped tables")
        worksheet.cell(row_cursor, 1).font = Font(bold=True, size=13)
        row_cursor += 2
        for block in blocks:
            if block.family == "issues_support":
                continue
            block_projects = set(block.project_codes)
            for table_row in block.rows:
                block_projects.update(extract_project_codes(" ".join(table_row)))
            if project not in block_projects and not (len(block.project_codes) == 1 and block.project_codes[0] == project):
                continue
            worksheet.cell(row_cursor, 1, f"Slide {block.slide} Table {block.table_no}: {block.slide_title} [{block.family}]")
            worksheet.cell(row_cursor, 1).font = Font(bold=True)
            row_cursor += 1
            for r_idx, table_row in enumerate(block.rows, start=1):
                for c_idx, value in enumerate(table_row, start=1):
                    cell_gap = gap_matrix[
                        (gap_matrix["slide"] == block.slide)
                        & (gap_matrix["table_no"] == block.table_no)
                        & (gap_matrix["row_no"] == r_idx)
                        & (gap_matrix["col_no"] == c_idx)
                        & ((gap_matrix["project_code"] == project) | (gap_matrix["project_code"] == ""))
                    ]
                    generated_value = ""
                    availability = ""
                    if not cell_gap.empty:
                        generated_value = clean_scalar(cell_gap.iloc[0].get("generated_value", ""))
                        availability = clean_scalar(cell_gap.iloc[0].get("availability", ""))
                    out_value = generated_value if generated_value else value
                    target = worksheet.cell(row_cursor, c_idx, out_value)
                    target.alignment = Alignment(wrap_text=True, vertical="top")
                    if generated_value:
                        target.fill = PatternFill("solid", fgColor="E2F0D9")
                    elif availability in {"parser_needed", "planning_not_captured", "history_unavailable", "external_required", "no_matching_dpr"}:
                        target.fill = PatternFill("solid", fgColor="FCE4D6")
                row_cursor += 1
            row_cursor += 2
        for col_idx in range(1, 24):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 16


def write_audit_workbook(
    output_path: Path,
    summary: pd.DataFrame,
    project_coverage: pd.DataFrame,
    gap_matrix: pd.DataFrame,
    source_map: pd.DataFrame,
    cell_audit: pd.DataFrame,
    common_tables: pd.DataFrame,
    blocks: list[TableBlock],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_dataframe(writer, summary, "Summary")
        write_dataframe(writer, project_coverage, "Project_Coverage")
        write_dataframe(writer, gap_matrix, "Gap_Matrix")
        write_dataframe(writer, source_map, "DPR_Source_Map")
        write_dataframe(writer, cell_audit, "Cell_Audit")
        write_dataframe(writer, common_tables, "Common_Tables")
        write_project_sheets(writer, blocks, gap_matrix)


def build_common_tables(gap_matrix: pd.DataFrame) -> pd.DataFrame:
    families = {"project_summary", "priority_projects", "pending_amendments", "running_project_master", "commercial"}
    return gap_matrix[gap_matrix["table_family"].isin(families)].copy()


def apply_ppt_updates(
    ppt_template: Path,
    output_path: Path,
    gap_matrix: pd.DataFrame,
    *,
    blank_unfilled: bool = False,
) -> tuple[bool, str, int, int]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(ppt_template, output_path)

    updates = gap_matrix[gap_matrix["ppt_update"].astype(bool)].copy()
    updates["ppt_action"] = "update"
    blank_rows = pd.DataFrame()
    if blank_unfilled:
        blank_mask = gap_matrix.apply(should_blank_unfilled_ppt_cell, axis=1)
        blank_rows = gap_matrix[blank_mask].copy()
        blank_rows["ppt_action"] = "blank"

    actions = pd.concat([updates, blank_rows], ignore_index=True) if not blank_rows.empty else updates
    if actions.empty:
        return True, "no PPT cell actions; copy created", 0, 0
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        return False, f"pywin32 import failed: {exc}; copy created without table updates", 0, 0

    powerpoint = None
    presentation = None
    update_count = 0
    blank_count = 0
    try:
        pythoncom.CoInitialize()
        powerpoint = win32com.client.DispatchEx("PowerPoint.Application")
        presentation = powerpoint.Presentations.Open(str(output_path), WithWindow=False)
        grouped = actions.groupby(["slide", "table_no"], sort=True)
        for (slide_no, table_no), group in grouped:
            slide = presentation.Slides(int(slide_no))
            table_shapes = []
            for shape in slide.Shapes:
                try:
                    if shape.HasTable:
                        table_shapes.append(shape)
                except Exception:
                    continue
            if int(table_no) > len(table_shapes):
                continue
            table = table_shapes[int(table_no) - 1].Table
            for _, row in group.iterrows():
                r = int(row["row_no"])
                c = int(row["col_no"])
                if r < 1 or c < 1 or r > table.Rows.Count or c > table.Columns.Count:
                    continue
                action = normalize_text(row.get("ppt_action", "update"))
                if action == "blank":
                    value = ""
                else:
                    value = clean_scalar(row.get("generated_value", ""))
                    if not value:
                        continue
                try:
                    table.Cell(r, c).Shape.TextFrame.TextRange.Text = value
                    if action == "blank":
                        blank_count += 1
                    else:
                        update_count += 1
                except Exception:
                    continue
        presentation.Save()
        if blank_unfilled:
            return True, "PPT copy created, DPR cells updated, and unresolved data cells blanked", update_count, blank_count
        return True, "PPT copy created and high-confidence table cells updated", update_count, blank_count
    except Exception as exc:
        return False, f"PPT update failed: {type(exc).__name__}: {exc}; copy exists but may be partially updated", update_count, blank_count
    finally:
        if presentation is not None:
            try:
                presentation.Close()
            except Exception:
                pass
        if powerpoint is not None:
            try:
                powerpoint.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def collect_project_sources(raw_dprs: dict[str, Path], scope_projects: set[str]) -> tuple[pd.DataFrame, pd.DataFrame]:
    detail_rows: list[dict[str, Any]] = []
    supply_rows: list[dict[str, Any]] = []
    for project in sorted(scope_projects):
        path = raw_dprs.get(project)
        if path is None:
            continue
        detail_rows.extend(extract_project_details(path, project))
        supply_rows.extend(extract_supply_signals(path, project))
    return pd.DataFrame(detail_rows), pd.DataFrame(supply_rows)


def export_cmd_dpr_audit(
    *,
    scope_workbook: Path = DEFAULT_SCOPE_WORKBOOK,
    dpr_dir: Path = DEFAULT_DPR_DIR,
    parquet_root: Path = DEFAULT_PARQUET_ROOT,
    ppt_template: Path = DEFAULT_PPT_TEMPLATE,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    run_date: str = DEFAULT_RUN_DATE,
    skip_ppt: bool = False,
    blank_unfilled_ppt: bool = False,
) -> dict[str, Any]:
    slides, table_raw, _presentation_summary = read_scope(scope_workbook)
    blocks = build_table_blocks(slides, table_raw)
    raw_dprs = discover_raw_dprs(dpr_dir)
    status = load_status_data(parquet_root)
    detail_rows, supply_rows = collect_project_sources(raw_dprs, MATCHING_SCOPE_PROJECTS)
    supply_lookup = build_supply_lookup(supply_rows)
    detail_lookup = build_detail_lookup(detail_rows)
    project_coverage = build_project_coverage(blocks, raw_dprs)
    gap_matrix = build_gap_matrix(blocks, raw_dprs, status, supply_lookup, detail_lookup)
    source_map = build_source_map(gap_matrix, detail_rows, supply_rows)
    summary = build_summary(project_coverage, gap_matrix)
    common_tables = build_common_tables(gap_matrix)
    cell_audit = gap_matrix.copy()

    audit_path = output_dir / f"CMD_DPR_Table_Audit_{run_date}.xlsx"
    write_audit_workbook(audit_path, summary, project_coverage, gap_matrix, source_map, cell_audit, common_tables, blocks)

    ppt_path = output_dir / f"CMD Presentation R2 18.05.2026_DPR_Fill_Copy_{run_date}.pptx"
    ppt_status = "skipped"
    ppt_updates = 0
    ppt_blanks = 0
    if not skip_ppt:
        success, message, update_count, blank_count = apply_ppt_updates(
            ppt_template,
            ppt_path,
            gap_matrix,
            blank_unfilled=blank_unfilled_ppt,
        )
        ppt_status = ("ok: " if success else "warning: ") + message
        ppt_updates = update_count
        ppt_blanks = blank_count

    return {
        "audit_path": audit_path,
        "ppt_path": ppt_path if not skip_ppt else None,
        "ppt_status": ppt_status,
        "ppt_updates": ppt_updates,
        "ppt_blanks": ppt_blanks,
        "gap_rows": len(gap_matrix),
        "source_rows": len(source_map),
        "projects": len(project_coverage),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export CMD presentation DPR audit workbook and PPT fill copy.")
    parser.add_argument("--scope-workbook", type=Path, default=DEFAULT_SCOPE_WORKBOOK)
    parser.add_argument("--dpr-dir", type=Path, default=DEFAULT_DPR_DIR)
    parser.add_argument("--parquet-root", type=Path, default=DEFAULT_PARQUET_ROOT)
    parser.add_argument("--ppt-template", type=Path, default=DEFAULT_PPT_TEMPLATE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--run-date", default=DEFAULT_RUN_DATE)
    parser.add_argument("--skip-ppt", action="store_true")
    parser.add_argument("--blank-unfilled-ppt", action="store_true", help="Blank unresolved data cells in the PPT copy after applying DPR-filled cells.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    started = datetime.now()
    result = export_cmd_dpr_audit(
        scope_workbook=args.scope_workbook,
        dpr_dir=args.dpr_dir,
        parquet_root=args.parquet_root,
        ppt_template=args.ppt_template,
        output_dir=args.output_dir,
        run_date=args.run_date,
        skip_ppt=args.skip_ppt,
        blank_unfilled_ppt=args.blank_unfilled_ppt,
    )
    elapsed = datetime.now() - started
    print(f"Audit workbook: {result['audit_path']}")
    if result["ppt_path"] is not None:
        print(f"PPT copy: {result['ppt_path']}")
    print(f"PPT status: {result['ppt_status']} ({result['ppt_updates']} updates, {result['ppt_blanks']} blanks)")
    print(f"Rows: gap={result['gap_rows']} source={result['source_rows']} projects={result['projects']}")
    print(f"Elapsed: {elapsed}")


if __name__ == "__main__":
    main()
