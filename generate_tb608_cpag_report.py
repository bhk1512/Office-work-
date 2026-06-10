from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime, timedelta
from html import escape
from pathlib import Path
from statistics import mean, median
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "Raw Data" / "DPRs" / "TB 608 - DPR - 2026-05-31.xlsm"
OUT_DIR = ROOT / "Productivity Summaries"
OUT_FILE = OUT_DIR / "TB608_CPAG_Comprehensive_Report.html"


def is_blank(value: Any) -> bool:
    return value is None or value == ""


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).replace("\n", " ").strip()


def num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def date_value(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    return None


def pct(value: float | None, total: float | None) -> float | None:
    if value is None or total in (None, 0):
        return None
    return value / total * 100


def fmt_num(value: Any, digits: int = 1) -> str:
    n = num(value)
    if n is None:
        return clean(value)
    if abs(n - round(n)) < 0.00001:
        return f"{int(round(n)):,}"
    return f"{n:,.{digits}f}"


def fmt_pct(value: float | None) -> str:
    if value is None:
        return "-"
    return f"{value:.1f}%"


def status_class(progress: float | None) -> str:
    if progress is None:
        return "neutral"
    if progress >= 75:
        return "good"
    if progress >= 40:
        return "warn"
    return "bad"


def html_table(headers: list[str], rows: Iterable[Iterable[Any]], classes: str = "") -> str:
    head = "".join(f"<th>{escape(h)}</th>" for h in headers)
    body_rows = []
    for row in rows:
        cells = "".join(f"<td>{escape(clean(c))}</td>" for c in row)
        body_rows.append(f"<tr>{cells}</tr>")
    return f"<table class='{classes}'><thead><tr>{head}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"


def bar(value: float | None, label: str | None = None) -> str:
    if value is None:
        return "<div class='bar'><span style='width:0%'></span></div>"
    width = max(0, min(100, value))
    text = label or f"{value:.1f}%"
    return (
        "<div class='bar'>"
        f"<span class='{status_class(value)}' style='width:{width:.2f}%'></span>"
        f"<em>{escape(text)}</em>"
        "</div>"
    )


def sheet_profile(wb_values, wb_formula) -> list[dict[str, Any]]:
    profiles = []
    for name in wb_values.sheetnames:
        ws_v = wb_values[name]
        ws_f = wb_formula[name]
        nonempty = 0
        formulas = 0
        min_row = min_col = 10**9
        max_row = max_col = 0
        for row in ws_v.iter_rows():
            for c in row:
                if not is_blank(c.value):
                    nonempty += 1
                    min_row = min(min_row, c.row)
                    min_col = min(min_col, c.column)
                    max_row = max(max_row, c.row)
                    max_col = max(max_col, c.column)
        for row in ws_f.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas += 1
        bbox = "" if nonempty == 0 else f"{min_row}:{max_row}, {get_column_letter(min_col)}:{get_column_letter(max_col)}"
        profiles.append(
            {
                "sheet": name,
                "rows": ws_v.max_row,
                "cols": ws_v.max_column,
                "nonempty": nonempty,
                "formulas": formulas,
                "merged": len(ws_f.merged_cells.ranges),
                "bbox": bbox,
            }
        )
    return profiles


def parse_dpr_summary(ws) -> tuple[dict[str, str], list[dict[str, Any]]]:
    meta = {
        "client": clean(ws.cell(1, 3).value),
        "contractor": clean(ws.cell(2, 3).value),
        "project": clean(ws.cell(3, 3).value),
        "loop_in": clean(ws.cell(3, 8).value),
        "loop_out": clean(ws.cell(3, 11).value),
        "loa": clean(ws.cell(4, 3).value),
        "status_date": clean(ws.cell(5, 3).value),
    }
    rows = []
    current_category = ""
    for r in range(8, 55):
        desc = clean(ws.cell(r, 2).value)
        unit = clean(ws.cell(r, 3).value)
        sl = clean(ws.cell(r, 1).value)
        if not desc:
            continue
        if not unit:
            current_category = desc
            continue
        rows.append(
            {
                "row": r,
                "category": current_category,
                "sl": sl,
                "desc": desc,
                "unit": unit,
                "loa": num(ws.cell(r, 4).value) or 0,
                "revised": num(ws.cell(r, 5).value) or 0,
                "l2_last": num(ws.cell(r, 6).value) or 0,
                "l2_this": num(ws.cell(r, 7).value) or 0,
                "l2_cum": num(ws.cell(r, 8).value) or 0,
                "actual_last": num(ws.cell(r, 9).value) or 0,
                "actual_this": num(ws.cell(r, 10).value) or 0,
                "actual_cum": num(ws.cell(r, 11).value) or 0,
                "balance": num(ws.cell(r, 12).value) or 0,
            }
        )
    return meta, rows


ROLLUP_RULES = [
    ("Route alignment", lambda d: "route alignment" in d),
    ("Detailed survey", lambda d: "detailed survey" in d),
    ("Check survey", lambda d: "check survey" in d),
    ("Normal foundation", lambda d: "normal foundation" in d),
    ("Pile foundation", lambda d: "pile foundation" in d and "pile cap" not in d),
    ("Pile cap", lambda d: "pile cap" in d),
    ("Earthing", lambda d: "pipe/counterpoise" in d or "earthing" in d),
    ("Tower erection", lambda d: "tower erection" in d),
    ("Tack welding", lambda d: "tack welding" in d),
    ("Rough sag", lambda d: "rough sag" in d),
    ("Final sag", lambda d: "final sag" in d),
    ("OPGW stringing", lambda d: "opgw" in d),
]


def rollup_dpr(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        desc = row["desc"].lower()
        group = next((name for name, fn in ROLLUP_RULES if fn(desc)), None)
        if not group:
            continue
        item = grouped.setdefault(
            group,
            {"activity": group, "unit": row["unit"], "loa": 0, "revised": 0, "l2_cum": 0, "actual_cum": 0, "actual_this": 0, "balance": 0},
        )
        for key in ["loa", "revised", "l2_cum", "actual_cum", "actual_this", "balance"]:
            item[key] += row[key]
    ordered = [grouped[name] for name, _ in ROLLUP_RULES if name in grouped]
    for item in ordered:
        item["progress"] = pct(item["actual_cum"], item["revised"])
        item["l2_gap"] = item["actual_cum"] - item["l2_cum"]
        item["l2_attainment"] = pct(item["actual_cum"], item["l2_cum"])
    return ordered


def parse_master_block(ws, start_row: int, label: str) -> list[dict[str, Any]]:
    rows = []
    for r in range(start_row + 1, start_row + 15):
        activity = clean(ws.cell(r, 2).value)
        if not activity:
            continue
        rows.append(
            {
                "block": label,
                "activity": activity,
                "loa": num(ws.cell(r, 3).value) or 0,
                "l2": num(ws.cell(r, 4).value) or 0,
                "revised": num(ws.cell(r, 5).value) or 0,
                "completed": num(ws.cell(r, 6).value) or 0,
                "wip": num(ws.cell(r, 7).value) or 0,
                "balance": num(ws.cell(r, 8).value) or 0,
            }
        )
    return rows


def parse_wip_fdn(ws, status_date: datetime) -> dict[str, Any]:
    records = []
    for r in range(8, ws.max_row + 1):
        loc = clean(ws.cell(r, 2).value)
        if not loc:
            continue
        start = date_value(ws.cell(r, 7).value)
        end = date_value(ws.cell(r, 8).value)
        status = clean(ws.cell(r, 11).value)
        earthing_status = clean(ws.cell(r, 15).value)
        record = {
            "row": r,
            "loc": loc,
            "tower_type": clean(ws.cell(r, 3).value),
            "classification": clean(ws.cell(r, 4).value),
            "gang": clean(ws.cell(r, 5).value),
            "manpower": num(ws.cell(r, 6).value),
            "start": start,
            "end": end,
            "section": clean(ws.cell(r, 9).value),
            "engineer": clean(ws.cell(r, 10).value),
            "status": status,
            "earthing_start": date_value(ws.cell(r, 12).value),
            "earthing_end": date_value(ws.cell(r, 13).value),
            "earthing_gang": clean(ws.cell(r, 14).value),
            "earthing_status": earthing_status,
            "ctq": clean(ws.cell(r, 16).value),
        }
        records.append(record)

    def bucket(status: str, start: datetime | None) -> str:
        s = status.lower()
        if "completed" in s:
            return "Completed"
        if "wip" in s or start:
            return "WIP / Started"
        return "Not started / blank"

    status_counts = Counter(bucket(r["status"], r["start"]) for r in records)
    earthing_counts = Counter(r["earthing_status"] or "Blank" for r in records)
    class_counts = Counter(r["classification"] or "Blank" for r in records)
    gang_counts = Counter(r["gang"] or "Blank" for r in records if bucket(r["status"], r["start"]) == "Completed")
    durations = [(r["end"] - r["start"]).days + 1 for r in records if r["start"] and r["end"] and "completed" in r["status"].lower()]
    ageing = []
    for r in records:
        if "completed" not in r["status"].lower() and r["start"]:
            ageing.append({**r, "age_days": (status_date - r["start"]).days})
    ageing.sort(key=lambda x: x["age_days"], reverse=True)
    return {
        "records": records,
        "status_counts": status_counts,
        "earthing_counts": earthing_counts,
        "class_counts": class_counts,
        "gang_counts": gang_counts,
        "durations": durations,
        "ageing": ageing,
    }


def parse_wip_erc(ws, status_date: datetime) -> dict[str, Any]:
    records = []
    for r in range(8, ws.max_row + 1):
        loc = clean(ws.cell(r, 2).value)
        if not loc:
            continue
        start = date_value(ws.cell(r, 7).value)
        end = date_value(ws.cell(r, 8).value)
        remarks = clean(ws.cell(r, 11).value)
        records.append(
            {
                "row": r,
                "loc": loc,
                "tower_type": clean(ws.cell(r, 3).value),
                "fdn_type": clean(ws.cell(r, 4).value),
                "gang": clean(ws.cell(r, 5).value),
                "manpower": num(ws.cell(r, 6).value),
                "start": start,
                "end": end,
                "section": clean(ws.cell(r, 9).value),
                "engineer": clean(ws.cell(r, 10).value),
                "remarks": remarks,
                "weight": num(ws.cell(r, 12).value) or 0,
            }
        )
    status_counts = Counter("Completed" if "completed" in r["remarks"].lower() else ("WIP / Started" if r["start"] else "Not started / blank") for r in records)
    gang_counts = Counter(r["gang"] or "Blank" for r in records if "completed" in r["remarks"].lower())
    durations = [(r["end"] - r["start"]).days + 1 for r in records if r["start"] and r["end"] and "completed" in r["remarks"].lower()]
    ageing = []
    for r in records:
        if "completed" not in r["remarks"].lower() and r["start"]:
            ageing.append({**r, "age_days": (status_date - r["start"]).days})
    ageing.sort(key=lambda x: x["age_days"], reverse=True)
    return {"records": records, "status_counts": status_counts, "gang_counts": gang_counts, "durations": durations, "ageing": ageing}


def parse_foundation_activity(ws, status_date: datetime) -> dict[str, Any]:
    locations = []
    pit_rows = []
    anomalies = []
    activity_pairs = [
        ("Pit marking", 11, 12),
        ("Excavation", 13, 14),
        ("Stub setting", 15, 16),
        ("PCC", 17, 18),
        ("Steel binding", 19, 20),
        ("FS concreting", 21, 22),
        ("Pyramid concreting", 23, 24),
        ("Chimney concreting", 25, 26),
        ("Backfilling", 27, 28),
    ]
    earliest_allowed = datetime(2025, 7, 1)
    latest_allowed = status_date + timedelta(days=7)
    for r in range(3, ws.max_row + 1):
        loc = clean(ws.cell(r, 2).value)
        pit = clean(ws.cell(r, 4).value)
        if loc and not pit:
            start = date_value(ws.cell(r, 10).value)
            end = date_value(ws.cell(r, 29).value)
            locations.append(
                {
                    "row": r,
                    "loc": loc,
                    "tower_type": clean(ws.cell(r, 3).value),
                    "classification": clean(ws.cell(r, 5).value),
                    "gang": clean(ws.cell(r, 6).value),
                    "manpower": num(ws.cell(r, 7).value),
                    "section": clean(ws.cell(r, 8).value),
                    "engineer": clean(ws.cell(r, 9).value),
                    "start": start,
                    "end": end,
                    "status": clean(ws.cell(r, 30).value),
                }
            )
        if pit:
            pit_rows.append(r)
            for activity, c_start, c_end in activity_pairs:
                start = date_value(ws.cell(r, c_start).value)
                end = date_value(ws.cell(r, c_end).value)
                for label, dt in [("start", start), ("end", end)]:
                    if dt and (dt < earliest_allowed or dt > latest_allowed):
                        anomalies.append([r, clean(ws.cell(r, 2).value) or "(pit row)", pit, activity, label, dt.strftime("%Y-%m-%d")])
                if start and end and start > end:
                    anomalies.append([r, clean(ws.cell(r, 2).value) or "(pit row)", pit, activity, "start>end", f"{start:%Y-%m-%d} > {end:%Y-%m-%d}"])
    status_counts = Counter(l["status"] or "Blank" for l in locations)
    durations = [(l["end"] - l["start"]).days + 1 for l in locations if l["start"] and l["end"] and "completed" in l["status"].lower()]
    activity_completion = {}
    for activity, _c_start, c_end in activity_pairs:
        count = 0
        for r in pit_rows:
            if date_value(ws.cell(r, c_end).value):
                count += 1
        activity_completion[activity] = count
    return {
        "locations": locations,
        "pit_count": len(pit_rows),
        "status_counts": status_counts,
        "durations": durations,
        "activity_completion": activity_completion,
        "anomalies": anomalies,
    }


def parse_supply(ws) -> dict[str, Any]:
    materials = []
    current_group = "Unclassified"
    for r in range(7, ws.max_row + 1):
        desc = clean(ws.cell(r, 2).value)
        if not desc:
            continue
        received = num(ws.cell(r, 4).value)
        issued = num(ws.cell(r, 5).value)
        balance_qty = num(ws.cell(r, 6).value)
        unit = clean(ws.cell(r, 3).value)
        if received is None and issued is None and balance_qty is None:
            current_group = desc.title()
            continue
        if received is None and issued is None and balance_qty is None:
            continue
        materials.append(
            {
                "row": r,
                "group": current_group,
                "desc": desc,
                "unit": unit,
                "received": received or 0,
                "issued": issued or 0,
                "balance": balance_qty or 0,
                "remarks": clean(ws.cell(r, 7).value),
            }
        )
    by_group_unit = defaultdict(lambda: {"received": 0.0, "issued": 0.0, "balance": 0.0, "items": 0})
    for m in materials:
        key = (m["group"], m["unit"])
        by_group_unit[key]["received"] += m["received"]
        by_group_unit[key]["issued"] += m["issued"]
        by_group_unit[key]["balance"] += m["balance"]
        by_group_unit[key]["items"] += 1
    low = [
        m
        for m in materials
        if m["received"] > 0 and (m["balance"] <= 0 or m["balance"] / m["received"] <= 0.1)
    ]
    low.sort(key=lambda m: (m["balance"] / m["received"] if m["received"] else 1, m["desc"]))
    return {"materials": materials, "by_group_unit": by_group_unit, "low": low}


def parse_line_details(ws) -> dict[str, Any]:
    records = []
    crossing_cols = list(range(13, 26))
    crossing_names = {c: clean(ws.cell(6, c).value) for c in crossing_cols}
    for r in range(7, ws.max_row + 1):
        tower = clean(ws.cell(r, 2).value)
        if not tower:
            continue
        span = num(ws.cell(r, 4).value) or 0
        crossings = []
        for c in crossing_cols:
            value = ws.cell(r, c).value
            if not is_blank(value) and clean(value) not in {"0", "-"}:
                crossings.append(crossing_names[c])
        records.append(
            {
                "row": r,
                "sl": clean(ws.cell(r, 1).value),
                "tower": tower,
                "tower_type": clean(ws.cell(r, 3).value),
                "span": span,
                "foundation": clean(ws.cell(r, 6).value),
                "earthing": clean(ws.cell(r, 7).value),
                "erection": clean(ws.cell(r, 8).value),
                "stringing": clean(ws.cell(r, 9).value),
                "village": clean(ws.cell(r, 10).value),
                "division": clean(ws.cell(r, 11).value),
                "crossing_details": clean(ws.cell(r, 12).value),
                "crossings": crossings,
                "classification": clean(ws.cell(r, 27).value),
            }
        )
    status_counts = {
        "Foundation": Counter(r["foundation"] or "Blank" for r in records),
        "Earthing": Counter(r["earthing"] or "Blank" for r in records),
        "Erection": Counter(r["erection"] or "Blank" for r in records),
        "Stringing": Counter(r["stringing"] or "Blank" for r in records),
    }
    crossing_counts = Counter(c for r in records for c in r["crossings"])
    classification_counts = Counter(r["classification"] or "Blank" for r in records)
    long_spans = sorted([r for r in records if r["span"]], key=lambda r: r["span"], reverse=True)[:12]
    crossing_rows = [r for r in records if r["crossings"] or r["crossing_details"]][:20]
    return {
        "records": records,
        "status_counts": status_counts,
        "crossing_counts": crossing_counts,
        "classification_counts": classification_counts,
        "long_spans": long_spans,
        "crossing_rows": crossing_rows,
    }


def parse_gang_movement(ws) -> dict[str, Any]:
    active = []
    for r in range(3, 40):
        loc = clean(ws.cell(r, 2).value)
        if not loc:
            continue
        active.append(
            [
                clean(ws.cell(r, 1).value),
                loc,
                clean(ws.cell(r, 3).value),
                clean(ws.cell(r, 4).value),
                clean(ws.cell(r, 5).value),
                clean(ws.cell(r, 6).value),
                clean(ws.cell(r, 7).value),
                clean(ws.cell(r, 8).value),
                clean(ws.cell(r, 9).value),
            ]
        )
    monthly = []
    for r in range(3, 25):
        gang = clean(ws.cell(r, 17).value)
        plan = clean(ws.cell(r, 18).value)
        if gang or plan:
            monthly.append([gang, plan])
    return {"active": active, "monthly": monthly}


def top_counter_rows(counter: Counter, n: int = 10) -> list[list[Any]]:
    return [[k, v] for k, v in counter.most_common(n)]


def duration_summary(durations: list[int]) -> str:
    if not durations:
        return "No completed durations available"
    return f"avg {mean(durations):.1f} days, median {median(durations):.1f}, max {max(durations)}"


def main() -> None:
    if not WORKBOOK.exists():
        raise FileNotFoundError(WORKBOOK)
    wb_values = load_workbook(WORKBOOK, data_only=True, read_only=False, keep_vba=False)
    wb_formula = load_workbook(WORKBOOK, data_only=False, read_only=False, keep_vba=True)

    meta, dpr_rows = parse_dpr_summary(wb_values["DPR Sum."])
    status_date = datetime.strptime(meta["status_date"], "%Y-%m-%d")
    dpr_rollup = rollup_dpr(dpr_rows)
    profiles = sheet_profile(wb_values, wb_formula)

    master_blocks = []
    master_blocks.extend(parse_master_block(wb_values["Master Sheet"], 1, "Master Sheet: combined/top"))
    master_blocks.extend(parse_master_block(wb_values["Master Sheet "], 1, "Master Sheet: main line"))
    master_blocks.extend(parse_master_block(wb_values["Master Sheet "], 464, "Master Sheet: loop in"))
    master_blocks.extend(parse_master_block(wb_values["Master Sheet "], 521, "Master Sheet: loop out"))

    wip_fdn = parse_wip_fdn(wb_values["WIP-FDN"], status_date)
    wip_erc = parse_wip_erc(wb_values["WIP-ERC"], status_date)
    foundation_activity = parse_foundation_activity(wb_values["Foundation Activity"], status_date)
    supply = parse_supply(wb_values["Supply"])
    line = parse_line_details(wb_values["Line Details"])
    gang = parse_gang_movement(wb_values["Gang Movement"])

    rollup_by_activity = {r["activity"]: r for r in dpr_rollup}
    foundation_completed = (
        rollup_by_activity.get("Normal foundation", {}).get("actual_cum", 0)
        + rollup_by_activity.get("Pile cap", {}).get("actual_cum", 0)
    )
    foundation_scope = (
        rollup_by_activity.get("Normal foundation", {}).get("revised", 0)
        + rollup_by_activity.get("Pile cap", {}).get("revised", 0)
    )
    key_kpis = [
        ("Foundation readiness", foundation_completed, foundation_scope, pct(foundation_completed, foundation_scope)),
        ("Earthing", rollup_by_activity["Earthing"]["actual_cum"], rollup_by_activity["Earthing"]["revised"], rollup_by_activity["Earthing"]["progress"]),
        ("Tower erection", rollup_by_activity["Tower erection"]["actual_cum"], rollup_by_activity["Tower erection"]["revised"], rollup_by_activity["Tower erection"]["progress"]),
        ("Tack welding", rollup_by_activity["Tack welding"]["actual_cum"], rollup_by_activity["Tack welding"]["revised"], rollup_by_activity["Tack welding"]["progress"]),
        ("Final sag", rollup_by_activity["Final sag"]["actual_cum"], rollup_by_activity["Final sag"]["revised"], rollup_by_activity["Final sag"]["progress"]),
        ("OPGW", rollup_by_activity["OPGW stringing"]["actual_cum"], rollup_by_activity["OPGW stringing"]["revised"], rollup_by_activity["OPGW stringing"]["progress"]),
    ]

    l2_risks = sorted(
        [r for r in dpr_rollup if r["l2_cum"] > 0 and r["actual_cum"] < r["l2_cum"]],
        key=lambda r: r["l2_cum"] - r["actual_cum"],
        reverse=True,
    )
    zero_started = [r for r in dpr_rollup if r["revised"] > 0 and r["actual_cum"] == 0]

    dpr_revised_erection = rollup_by_activity["Tower erection"]["revised"]
    combined_master_erection = next((r for r in master_blocks if r["block"] == "Master Sheet: combined/top" and r["activity"] == "Tower Erection"), None)
    data_flags = []
    if combined_master_erection and abs(combined_master_erection["revised"] - dpr_revised_erection) > 0.1:
        data_flags.append(
            [
                "Scope mismatch",
                f"DPR summary tower erection revised scope is {fmt_num(dpr_revised_erection)}, while the combined/top Master Sheet shows {fmt_num(combined_master_erection['revised'])}.",
            ]
        )
    erc_completed = wip_erc["status_counts"].get("Completed", 0)
    dpr_erection_actual = rollup_by_activity["Tower erection"]["actual_cum"]
    if abs(erc_completed - dpr_erection_actual) > 0.1:
        data_flags.append(["Erection count mismatch", f"WIP-ERC completed records count {erc_completed}, DPR actual cumulative {fmt_num(dpr_erection_actual)}."])
    fdn_completed_records = wip_fdn["status_counts"].get("Completed", 0)
    if abs(fdn_completed_records - foundation_completed) > 2:
        data_flags.append(["Foundation count mismatch", f"WIP-FDN completed records count {fdn_completed_records}, DPR foundation readiness estimate {fmt_num(foundation_completed)}."])
    if foundation_activity["anomalies"]:
        data_flags.append(["Date anomalies", f"{len(foundation_activity['anomalies'])} foundation activity date cells are outside the expected project/status window."])
    data_flags.append(["Coverage gap", "The workbook contains physical progress, WIP, line, gang, and supply data. Cashflow, tender comparison, revenue, margin/cost escalation, receivables, AS7, and customer correspondence are not present in this raw file."])

    kpi_cards = []
    for label, actual, scope, progress in key_kpis:
        kpi_cards.append(
            "<div class='kpi'>"
            f"<h3>{escape(label)}</h3>"
            f"<strong>{fmt_pct(progress)}</strong>"
            f"{bar(progress)}"
            f"<p>{fmt_num(actual)} of {fmt_num(scope)}</p>"
            "</div>"
        )

    dpr_table_rows = []
    for r in dpr_rollup:
        dpr_table_rows.append(
            [
                r["activity"],
                r["unit"],
                fmt_num(r["revised"]),
                fmt_num(r["actual_cum"]),
                fmt_num(r["balance"]),
                fmt_num(r["actual_this"]),
                fmt_pct(r["progress"]),
                f"{fmt_num(r['actual_cum'])} / {fmt_num(r['l2_cum'])} ({fmt_pct(r['l2_attainment'])})",
                fmt_num(r["l2_gap"]),
            ]
        )

    supply_group_rows = [
        [group, unit, vals["items"], fmt_num(vals["received"]), fmt_num(vals["issued"]), fmt_num(vals["balance"])]
        for (group, unit), vals in sorted(supply["by_group_unit"].items())
    ]
    low_supply_rows = [
        [m["group"], m["desc"], m["unit"], fmt_num(m["received"]), fmt_num(m["issued"]), fmt_num(m["balance"])]
        for m in supply["low"][:25]
    ]

    line_status_rows = []
    for stage, counter in line["status_counts"].items():
        for k, v in counter.most_common():
            line_status_rows.append([stage, k, v])

    html = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>TB 608 CPAG Comprehensive Report</title>
<style>
:root {{
  --ink:#172026; --muted:#667085; --line:#d8dee6; --bg:#f7f8fa; --panel:#fff;
  --good:#138a5e; --warn:#b26a00; --bad:#c2410c; --blue:#195d8d; --soft:#eef3f7;
}}
* {{ box-sizing:border-box; }}
body {{ margin:0; font:14px/1.45 "Segoe UI", Arial, sans-serif; color:var(--ink); background:var(--bg); }}
header {{ background:#183547; color:white; padding:28px 36px; }}
header h1 {{ margin:0 0 8px; font-size:28px; letter-spacing:0; }}
header p {{ margin:2px 0; color:#d7e4ec; }}
main {{ padding:24px 36px 40px; max-width:1500px; margin:0 auto; }}
section {{ margin:0 0 26px; }}
h2 {{ margin:0 0 12px; font-size:20px; }}
h3 {{ margin:0 0 8px; font-size:15px; }}
.grid {{ display:grid; grid-template-columns:repeat(6,minmax(150px,1fr)); gap:12px; }}
.two {{ display:grid; grid-template-columns:1fr 1fr; gap:18px; }}
.kpi,.panel {{ background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:14px; }}
.kpi strong {{ display:block; font-size:25px; margin:2px 0 8px; }}
.kpi p,.note {{ color:var(--muted); margin:8px 0 0; }}
.bar {{ position:relative; height:20px; background:#e8edf2; border-radius:999px; overflow:hidden; }}
.bar span {{ display:block; height:100%; background:var(--blue); }}
.bar span.good {{ background:var(--good); }}
.bar span.warn {{ background:var(--warn); }}
.bar span.bad {{ background:var(--bad); }}
.bar em {{ position:absolute; inset:0; text-align:center; font-style:normal; font-size:12px; color:#111; line-height:20px; }}
table {{ width:100%; border-collapse:collapse; background:white; border:1px solid var(--line); border-radius:8px; overflow:hidden; }}
th,td {{ border-bottom:1px solid var(--line); padding:8px 9px; text-align:left; vertical-align:top; }}
th {{ background:#eef3f7; font-size:12px; text-transform:uppercase; color:#41505c; }}
tr:last-child td {{ border-bottom:0; }}
.dense td {{ font-size:12px; padding:6px 7px; }}
.flag {{ border-left:5px solid var(--bad); background:#fff7ed; }}
.warnflag {{ border-left:5px solid var(--warn); background:#fffbeb; }}
.goodtxt {{ color:var(--good); font-weight:600; }}
.badtxt {{ color:var(--bad); font-weight:600; }}
.small {{ font-size:12px; color:var(--muted); }}
@media (max-width:1100px) {{ .grid,.two {{ grid-template-columns:1fr 1fr; }} }}
@media (max-width:720px) {{ main,header {{ padding-left:16px; padding-right:16px; }} .grid,.two {{ grid-template-columns:1fr; }} table {{ font-size:12px; }} }}
</style>
</head>
<body>
<header>
  <h1>TB 608 CPAG Assurance Report</h1>
  <p>{escape(meta['project'])}</p>
  <p>Client: {escape(meta['client'])} | Contractor: {escape(meta['contractor'])} | Status date: {escape(meta['status_date'])}</p>
  <p>Raw file: {escape(str(WORKBOOK.relative_to(ROOT)))}</p>
</header>
<main>
<section>
  <h2>Executive View</h2>
  <div class="grid">{''.join(kpi_cards)}</div>
  <p class="note">Foundation readiness is calculated as normal foundation plus pile cap completion from the DPR summary. Progress percentages use revised scope as the denominator.</p>
</section>

<section class="two">
  <div class="panel flag">
    <h2>Critical Early Warnings</h2>
    <ul>
      <li><span class="badtxt">Downstream works have not opened:</span> tack welding, rough/final sag, and OPGW show zero actual progress against non-zero revised scope.</li>
      <li><span class="badtxt">Tower erection is behind L2:</span> {fmt_num(rollup_by_activity['Tower erection']['actual_cum'])} actual versus {fmt_num(rollup_by_activity['Tower erection']['l2_cum'])} L2 cumulative.</li>
      <li><span class="badtxt">Earthing is constraining erection readiness:</span> {fmt_num(rollup_by_activity['Earthing']['actual_cum'])} actual versus {fmt_num(rollup_by_activity['Earthing']['l2_cum'])} L2 cumulative.</li>
      <li><span class="badtxt">Pile cap closure is weak:</span> {fmt_num(rollup_by_activity['Pile cap']['actual_cum'])} actual versus {fmt_num(rollup_by_activity['Pile cap']['l2_cum'])} L2 cumulative.</li>
    </ul>
  </div>
  <div class="panel warnflag">
    <h2>Data And CPAG Coverage Caveats</h2>
    <ul>
      {''.join(f"<li><b>{escape(flag[0])}:</b> {escape(flag[1])}</li>" for flag in data_flags)}
    </ul>
  </div>
</section>

<section>
  <h2>DPR Summary: Scope, Actuals, L2 And Balance</h2>
  {html_table(["Activity","Unit","Revised","Actual Cum.","Balance","This Month","Progress","Actual / L2 Cum.","L2 Gap"], dpr_table_rows)}
</section>

<section class="two">
  <div>
    <h2>L2 Slippage Ranking</h2>
    {html_table(["Activity","Revised","L2 Cum.","Actual Cum.","Behind L2"], [[r["activity"], fmt_num(r["revised"]), fmt_num(r["l2_cum"]), fmt_num(r["actual_cum"]), fmt_num(r["l2_cum"]-r["actual_cum"])] for r in l2_risks], "dense")}
  </div>
  <div>
    <h2>Zero-Start Activities</h2>
    {html_table(["Activity","Revised Scope","Balance"], [[r["activity"], fmt_num(r["revised"]), fmt_num(r["balance"])] for r in zero_started], "dense")}
  </div>
</section>

<section>
  <h2>Line/Loop Detailed DPR Rows</h2>
  {html_table(["Category","Description","Unit","LOA","Revised","L2 Cum.","Actual Cum.","This Month","Balance"], [[r["category"], r["desc"], r["unit"], fmt_num(r["loa"]), fmt_num(r["revised"]), fmt_num(r["l2_cum"]), fmt_num(r["actual_cum"]), fmt_num(r["actual_this"]), fmt_num(r["balance"])] for r in dpr_rows], "dense")}
</section>

<section class="two">
  <div>
    <h2>WIP Foundation And Earthing</h2>
    <p class="note">Records: {len(wip_fdn['records'])}. Foundation duration: {escape(duration_summary(wip_fdn['durations']))}.</p>
    {html_table(["Foundation Status","Count"], top_counter_rows(wip_fdn["status_counts"]), "dense")}
    <h3>Earthing Status</h3>
    {html_table(["Earthing Status","Count"], top_counter_rows(wip_fdn["earthing_counts"]), "dense")}
  </div>
  <div>
    <h2>WIP Erection</h2>
    <p class="note">Records: {len(wip_erc['records'])}. Erection duration: {escape(duration_summary(wip_erc['durations']))}.</p>
    {html_table(["Erection Status","Count"], top_counter_rows(wip_erc["status_counts"]), "dense")}
    <h3>Completed Erection By Gang</h3>
    {html_table(["Gang","Completed Locations"], top_counter_rows(wip_erc["gang_counts"], 12), "dense")}
  </div>
</section>

<section class="two">
  <div>
    <h2>Oldest Open Foundation Items</h2>
    {html_table(["Location","Tower Type","Gang","Start","Age Days","Status"], [[r["loc"], r["tower_type"], r["gang"], clean(r["start"]), r["age_days"], r["status"]] for r in wip_fdn["ageing"][:15]], "dense")}
  </div>
  <div>
    <h2>Oldest Open Erection Items</h2>
    {html_table(["Location","Tower Type","Gang","Start","Age Days","Remarks"], [[r["loc"], r["tower_type"], r["gang"], clean(r["start"]), r["age_days"], r["remarks"]] for r in wip_erc["ageing"][:15]], "dense")}
  </div>
</section>

<section class="two">
  <div>
    <h2>Foundation Activity Sheet</h2>
    <p class="note">Location rows: {len(foundation_activity['locations'])}; pit rows: {foundation_activity['pit_count']}; location duration: {escape(duration_summary(foundation_activity['durations']))}.</p>
    {html_table(["Location Status","Count"], top_counter_rows(foundation_activity["status_counts"]), "dense")}
    <h3>Pit-Level Activity End Dates Captured</h3>
    {html_table(["Activity","Pit Rows With End Date"], foundation_activity["activity_completion"].items(), "dense")}
  </div>
  <div>
    <h2>Foundation Date Anomalies</h2>
    {html_table(["Excel Row","Location","Pit","Activity","Field","Value"], foundation_activity["anomalies"][:25], "dense")}
  </div>
</section>

<section>
  <h2>Line Details Sheet</h2>
  <div class="two">
    <div>
      <p class="note">Tower/location records: {len(line['records'])}; total span captured: {fmt_num(sum(r['span'] for r in line['records'])/1000)} km.</p>
      {html_table(["Stage","Status","Count"], line_status_rows, "dense")}
    </div>
    <div>
      <h3>Crossing Count By Type</h3>
      {html_table(["Crossing Type","Count"], top_counter_rows(line["crossing_counts"], 20), "dense")}
      <h3>Top Long Spans</h3>
      {html_table(["Tower","Tower Type","Span m","Crossing Details"], [[r["tower"], r["tower_type"], fmt_num(r["span"]), r["crossing_details"]] for r in line["long_spans"]], "dense")}
    </div>
  </div>
</section>

<section>
  <h2>Crossing Locations From Line Details</h2>
  {html_table(["Tower","Tower Type","Span m","Marked Crossings","Crossing Details","Foundation","Earthing","Erection"], [[r["tower"], r["tower_type"], fmt_num(r["span"]), ", ".join(r["crossings"]), r["crossing_details"], r["foundation"], r["earthing"], r["erection"]] for r in line["crossing_rows"]], "dense")}
</section>

<section class="two">
  <div>
    <h2>Supply Stock Statement</h2>
    {html_table(["Group","Unit","Items","Received","Issued","Balance"], supply_group_rows, "dense")}
  </div>
  <div>
    <h2>Low / Zero Balance Supply Items</h2>
    {html_table(["Group","Material","Unit","Received","Issued","Balance"], low_supply_rows, "dense")}
  </div>
</section>

<section class="two">
  <div>
    <h2>Gang Movement</h2>
    <p class="note">The active movement area is mostly formula-driven and currently has many blank derived fields; monthly plan entries are present.</p>
    {html_table(["Gang","Monthly Plan Locations"], gang["monthly"], "dense")}
  </div>
  <div>
    <h2>Active Location Tracker</h2>
    {html_table(["S.No","Loc","Tower Type","FDN Type","Gang","Start","Expected Completion","Progress","Next Location"], gang["active"][:25], "dense")}
  </div>
</section>

<section>
  <h2>Master Sheet Blocks</h2>
  {html_table(["Block","Activity","LOA","L2","Revised","Completed","WIP","Balance"], [[r["block"], r["activity"], fmt_num(r["loa"]), fmt_num(r["l2"]), fmt_num(r["revised"]), fmt_num(r["completed"]), fmt_num(r["wip"]), fmt_num(r["balance"])] for r in master_blocks], "dense")}
</section>

<section>
  <h2>Workbook Sheet Audit</h2>
  {html_table(["Sheet","Rows","Cols","Non-empty Cells","Formula Cells","Merged Ranges","Non-empty Range"], [[p["sheet"], p["rows"], p["cols"], p["nonempty"], p["formulas"], p["merged"], p["bbox"]] for p in profiles], "dense")}
</section>

<section>
  <h2>Support Required Sheet</h2>
  <p class="note">No actionable support items were filled in the raw workbook. The sheet contains only the header row.</p>
</section>
</main>
</body>
</html>
"""

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text(html, encoding="utf-8")
    print(OUT_FILE)


if __name__ == "__main__":
    main()
