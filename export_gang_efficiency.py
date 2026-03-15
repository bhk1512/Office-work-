#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

BASE = Path(__file__).resolve().parent
PROD = BASE / "Productivity Summaries"
SRC_DEFAULT = PROD / "FY24_FY25_Erection_Summary.xlsx"
OUT_DEFAULT = PROD / "Gang_Efficiency_FY2425.xlsx"
THR = [">=1", ">=2", ">=3"]
THR_MARK = {">=1": "Erections >=1", ">=2": "Erections >=2", ">=3": "Erections >=3"}
BUCKETS = ["0-1", "1-2", "2-3", "3-4", "4-5", "5-6", "6-7", "7-8", "8-9"]
BASELINE = [0.5, 1.5, 2.5, 3.5, 4.5, 5.5, 6.5, 7.5, 8.5]
METRICS = [
    ("gang_count", "Gang Count", "sum"),
    ("weight_mt", "Weight Erected (MT) - Daily Prod", "sum"),
    ("deployment_months", "Deployment months (avg)", "avg"),
    ("idle_days_deploy", "Idle days per deployment month (normalized, primary)", "avg"),
    ("idle_days_scope", "Idle days per month (normalized)", "avg"),
    ("avg_max_raw_gap", "Avg max raw gap days", "avg"),
    ("scope_months", "Scope months (avg)", "avg"),
    ("baseline", "Baseline MT/day Assumption (Hardcoded)", "avg"),
]


def n(v) -> str:
    return "" if pd.isna(v) else str(v).strip()


def abs_ref(coord: str) -> str:
    c = "".join(ch for ch in coord if ch.isalpha())
    r = "".join(ch for ch in coord if ch.isdigit())
    return f"${c}${r}"


def parse() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Export Gang Efficiency workbook")
    p.add_argument("--source", type=Path, default=SRC_DEFAULT)
    p.add_argument("--output", type=Path, default=OUT_DEFAULT)
    return p.parse_args()


def extract(source: Path) -> dict[str, dict[str, dict[str, float]]]:
    df = pd.read_excel(source, sheet_name="Idle Days Buckets", usecols="A:M", header=None, engine="openpyxl")
    starts: dict[str, int] = {}
    for i, v in df[0].items():
        t = n(v)
        for k, m in THR_MARK.items():
            if t == m:
                starts[k] = i
    miss = [k for k in THR if k not in starts]
    if miss:
        raise ValueError(f"Missing threshold section(s): {miss}")

    out = {k: {} for k in THR}
    for i, k in enumerate(THR):
        st = starts[k]
        en = starts[THR[i + 1]] if i + 1 < len(THR) else len(df)
        hdr = next((r for r in range(st, en) if n(df.iat[r, 0]) == "Metric"), None)
        if hdr is None:
            raise ValueError(f"Metric header missing in {k}")
        bcols = {}
        for c in range(1, 13):
            lab = n(df.iat[hdr, c])
            if lab in BUCKETS:
                bcols[lab] = c
        if any(b not in bcols for b in BUCKETS):
            raise ValueError(f"Bucket columns missing in {k}")

        for mid, label, _ in METRICS:
            if mid == "baseline":
                out[k][mid] = {b: float(v) for b, v in zip(BUCKETS, BASELINE)}
                continue
            rr = next((r for r in range(hdr + 1, en) if n(df.iat[r, 0]) == label), None)
            if rr is None:
                raise ValueError(f"Row '{label}' missing in {k}")
            out[k][mid] = {b: float(df.iat[rr, bcols[b]]) if pd.notna(df.iat[rr, bcols[b]]) else 0.0 for b in BUCKETS}
    return out


def styles() -> dict[str, object]:
    thin = Side(style="thin", color="FFBFBFBF")
    return {
        "dark": PatternFill("solid", fgColor="FF1F3864"),
        "mid": PatternFill("solid", fgColor="FF2F5496"),
        "light": PatternFill("solid", fgColor="FFD6E4F0"),
        "yellow": PatternFill("solid", fgColor="FFFFF2CC"),
        "green": PatternFill("solid", fgColor="FFE2EFDA"),
        "white": PatternFill("solid", fgColor="FFFFFFFF"),
        "fw": Font(color="FFFFFFFF", bold=True),
        "fb": Font(color="FF000000"),
        "fbb": Font(color="FF000000", bold=True),
        "fblue": Font(color="FF2F5496"),
        "fgreen": Font(color="FF548235"),
        "fcall": Font(color="FF000000", bold=True, size=11),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "b": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def paint(ws, row: int, c1: int, c2: int, fill=None, font=None, align=None, border=None):
    for c in range(c1, c2 + 1):
        cl = ws.cell(row=row, column=c)
        if fill is not None:
            cl.fill = fill
        if font is not None:
            cl.font = font
        if align is not None:
            cl.alignment = align
        if border is not None:
            cl.border = border

def build_sheet1(wb: Workbook, src: dict[str, dict[str, dict[str, float]]], s: dict[str, object]):
    ws = wb.create_sheet("1. Raw Inputs")
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 35
    for c in range(2, 12):
        ws.column_dimensions[get_column_letter(c)].width = 11

    ws["A1"] = "Gang Deployment Efficiency Inputs (Overall Block Only)"
    ws.merge_cells("A1:K1")
    paint(ws, 1, 1, 11, s["dark"], s["fw"], s["left"], s["b"])
    ws["A2"] = "Source: FY24_FY25_Erection_Summary.xlsx -> Idle Days Buckets (Overall A:M)"
    ws.merge_cells("A2:K2")
    paint(ws, 2, 1, 11, s["light"], s["fb"], s["left"], s["b"])

    refs = {m[0]: {} for m in METRICS}
    rp = 4
    for mid, mlabel, agg in METRICS:
        ws.cell(rp, 1, mlabel)
        ws.merge_cells(start_row=rp, start_column=1, end_row=rp, end_column=11)
        paint(ws, rp, 1, 11, s["mid"], s["fw"], s["left"], s["b"])

        hr = rp + 1
        ws.cell(hr, 1, "Threshold")
        for i, b in enumerate(BUCKETS, start=2):
            ws.cell(hr, i, b)
        ws.cell(hr, 11, "TOTAL" if agg == "sum" else "AVG")
        paint(ws, hr, 1, 11, s["light"], s["fbb"], s["center"], s["b"])

        for ti, t in enumerate(THR):
            r = rp + 2 + ti
            refs[mid][t] = {"buckets": {}, "summary": ""}
            ws.cell(r, 1, t)
            paint(ws, r, 1, 1, s["light"], s["fbb"], s["center"], s["b"])
            for i, b in enumerate(BUCKETS, start=2):
                cl = ws.cell(r, i, src[t][mid][b])
                refs[mid][t]["buckets"][b] = cl.coordinate
                cl.fill = s["yellow"]
                cl.font = s["fblue"]
                cl.alignment = s["center"]
                cl.border = s["b"]
            f = f"=SUM(B{r}:J{r})" if agg == "sum" else f"=AVERAGE(B{r}:J{r})"
            cs = ws.cell(r, 11, f)
            refs[mid][t]["summary"] = cs.coordinate
            cs.fill = s["green"]
            cs.font = s["fb"]
            cs.alignment = s["center"]
            cs.border = s["b"]
        rp += 6

    for mid, _, _ in METRICS:
        for t in THR:
            for b in BUCKETS:
                ws[refs[mid][t]["buckets"][b]].number_format = "0" if mid == "gang_count" else "0.00"
            ws[refs[mid][t]["summary"]].number_format = "0" if mid == "gang_count" else "0.00"
    return refs


def build_sheet2(wb: Workbook, refs: dict, s: dict[str, object]):
    ws = wb.create_sheet("2. Fleet Efficiency")
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 24

    ws["A1"] = "Fleet Efficiency (Formula Driven)"
    ws.merge_cells("A1:B1")
    paint(ws, 1, 1, 2, s["dark"], s["fw"], s["left"], s["b"])
    for r, t in ((3, "Section A - Fleet Coverage"), (10, "Section B - Throughput"), (19, "Section C - Loss and Outperformance"), (25, "Summary KPIs")):
        ws.cell(r, 1, t)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        paint(ws, r, 1, 2, s["mid"], s["fw"], s["left"], s["b"])

    def wrow(r: int, label: str, formula: str, cross: bool, fmt: str):
        ws.cell(r, 1, label)
        paint(ws, r, 1, 1, s["light"], s["fbb"], s["left"], s["b"])
        ws.cell(r, 2, formula)
        paint(ws, r, 2, 2, s["green"] if cross else s["white"], s["fgreen"] if cross else s["fb"], s["center"], s["b"])
        ws.cell(r, 2).number_format = fmt

    raw = "'1. Raw Inputs'"
    wrow(4, "Gangs >=1", f"={raw}!{abs_ref(refs['gang_count']['>=1']['summary'])}", True, "0")
    wrow(5, "Gangs >=2", f"={raw}!{abs_ref(refs['gang_count']['>=2']['summary'])}", True, "0")
    wrow(6, "Gangs >=3", f"={raw}!{abs_ref(refs['gang_count']['>=3']['summary'])}", True, "0")
    wrow(7, "Gangs Dropped (>=1 minus >=3)", "=B4-B6", False, "0")
    wrow(8, "% of Fleet Removed", "=IFERROR(B7/B4,0)", False, "0.00%")

    wrow(11, "Total MT >=1", f"={raw}!{abs_ref(refs['weight_mt']['>=1']['summary'])}", True, "0.00")
    wrow(12, "Total MT >=2", f"={raw}!{abs_ref(refs['weight_mt']['>=2']['summary'])}", True, "0.00")
    wrow(13, "Total MT >=3", f"={raw}!{abs_ref(refs['weight_mt']['>=3']['summary'])}", True, "0.00")
    wrow(14, "MT per Gang >=1", "=IFERROR(B11/B4,0)", False, "0.00")
    wrow(15, "MT per Gang >=2", "=IFERROR(B12/B5,0)", False, "0.00")
    wrow(16, "MT per Gang >=3", "=IFERROR(B13/B6,0)", False, "0.00")
    wrow(17, "Efficiency Ratio (MT/gang >=3 vs >=1)", "=IFERROR(B16/B14,0)", False, "0.00")

    wrow(20, "MT Lost (>=1 minus >=3)", "=B11-B13", False, "0.00")
    wrow(21, "% MT Lost", "=IFERROR(B20/B11,0)", False, "0.00%")
    wrow(22, "MT per Dropped Gang", "=IFERROR(B20/B7,0)", False, "0.00")
    wrow(23, "Retained Outperformance", "=IFERROR(B16/B22,0)", False, "0.00")

    wrow(26, "% Fleet Removed", "=B8", False, "0.00%")
    wrow(27, "% MT from Marginal Gangs", "=B21", False, "0.00%")
    wrow(28, "Deployment Inefficiency Ratio", "=IFERROR(B26/B27,0)", False, "0.00")
    wrow(29, "MT/gang retained", "=B16", False, "0.00")
    wrow(30, "MT/gang dropped", "=B22", False, "0.00")
    wrow(31, "Retained Outperformance", "=B23", False, "0.00")

def build_sheet3(wb: Workbook, refs: dict, s: dict[str, object]):
    ws = wb.create_sheet("3. Idle Recovery")
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 52
    for c in range(2, 12):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws["A1"] = "Idle Recovery (Deployment-Month Basis)"
    ws.merge_cells("A1:K1")
    paint(ws, 1, 1, 11, s["dark"], s["fw"], s["left"], s["b"])

    for r, t in ((3, "Section A - Inputs (>=3, pulled from Sheet 1)"), (10, "Section B - Recovery Calculation"), (14, "Section C - Idle Gap Analysis"), (20, "Section D - Coverage Test")):
        ws.cell(r, 1, t)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        paint(ws, r, 1, 11, s["mid"], s["fw"], s["left"], s["b"])

    ws.cell(4, 1, "Metric")
    for i, b in enumerate(BUCKETS, start=2):
        ws.cell(4, i, b)
    ws.cell(4, 11, "TOTAL/AVG")
    paint(ws, 4, 1, 11, s["light"], s["fbb"], s["center"], s["b"])

    inputs = {
        5: ("Gang Count (>=3)", "gang_count", "sum", "0"),
        6: ("Deployment Months (>=3)", "deployment_months", "avg", "0.00"),
        7: ("Idle Days / Deployment Month (>=3)", "idle_days_deploy", "avg", "0.00"),
        8: ("Baseline MT/day", "baseline", "avg", "0.00"),
    }
    raw = "'1. Raw Inputs'"
    for r, (label, mid, agg, fmt) in inputs.items():
        ws.cell(r, 1, label)
        paint(ws, r, 1, 1, s["light"], s["fbb"], s["left"], s["b"])
        for i, b in enumerate(BUCKETS, start=2):
            ws.cell(r, i, f"={raw}!{abs_ref(refs[mid]['>=3']['buckets'][b])}")
            paint(ws, r, i, i, s["green"], s["fgreen"], s["center"], s["b"])
            ws.cell(r, i).number_format = fmt
        ws.cell(r, 11, f"=SUM(B{r}:J{r})" if agg == "sum" else f"=AVERAGE(B{r}:J{r})")
        paint(ws, r, 11, 11, s["white"], s["fb"], s["center"], s["b"])
        ws.cell(r, 11).number_format = fmt

    ws.cell(11, 1, "Total Idle Days")
    ws.cell(12, 1, "Recoverable MT")
    paint(ws, 11, 1, 1, s["light"], s["fbb"], s["left"], s["b"])
    paint(ws, 12, 1, 1, s["light"], s["fbb"], s["left"], s["b"])

    for col in range(2, 11):
        let = get_column_letter(col)
        ws.cell(11, col, f"={let}5*{let}6*{let}7")
        ws.cell(12, col, f"={let}11*{let}8")
        paint(ws, 11, col, col, s["white"], s["fb"], s["center"], s["b"])
        paint(ws, 12, col, col, s["white"], s["fb"], s["center"], s["b"])
        ws.cell(11, col).number_format = "0.00"
        ws.cell(12, col).number_format = "0.00"
    ws.cell(11, 11, "=SUM(B11:J11)")
    ws.cell(12, 11, "=SUM(B12:J12)")
    paint(ws, 11, 11, 11, s["green"], s["fbb"], s["center"], s["b"])
    paint(ws, 12, 11, 11, s["green"], s["fbb"], s["center"], s["b"])
    ws.cell(11, 11).number_format = "0.00"
    ws.cell(12, 11).number_format = "0.00"

    for r, label, f, fmt in [
        (15, "4-5 Bucket Idle Rate (>=3)", "=F7", "0.00"),
        (16, "6-7 Bucket Idle Rate (>=3)", "=H7", "0.00"),
        (17, "Idle Rate Gap (4-5 minus 6-7)", "=MAX(B15-B16,0)", "0.00"),
        (18, "Targeted Recovery MT (4-5 closes to 6-7 benchmark)", "=B17*F5*F6*F8", "0.00"),
    ]:
        ws.cell(r, 1, label)
        paint(ws, r, 1, 1, s["light"], s["fbb"], s["left"], s["b"])
        ws.cell(r, 2, f)
        paint(ws, r, 2, 2, s["white"], s["fb"], s["center"], s["b"])
        ws.cell(r, 2).number_format = fmt

    for r, label, f, cross, fmt in [
        (21, "MT Lost (from Sheet 2)", "='2. Fleet Efficiency'!$B$20", True, "0.00"),
        (22, "Recoverable MT (Section B total)", "=K12", False, "0.00"),
        (23, "Net Coverage (Recoverable - Lost)", "=B22-B21", False, "0.00"),
        (24, "Coverage Ratio", "=IFERROR(B22/B21,0)", False, "0.00%"),
        (25, "Feasibility", '=IF(B24>=1,"FULL SUBSTITUTION POSSIBLE","PARTIAL - gap remains")', False, None),
    ]:
        ws.cell(r, 1, label)
        paint(ws, r, 1, 1, s["light"], s["fbb"], s["left"], s["b"])
        ws.cell(r, 2, f)
        paint(ws, r, 2, 2, s["green"] if cross else s["white"], s["fgreen"] if cross else s["fb"], s["center"], s["b"])
        if fmt:
            ws.cell(r, 2).number_format = fmt


def build_sheet4(wb: Workbook, s: dict[str, object]):
    ws = wb.create_sheet("4. Leadership Summary")
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["B"].width = 28

    ws["A1"] = "Leadership Summary"
    ws.merge_cells("A1:B1")
    paint(ws, 1, 1, 2, s["dark"], s["fw"], s["left"], s["b"])

    ws["A3"] = "KPI"
    ws["B3"] = "Value"
    paint(ws, 3, 1, 2, s["mid"], s["fw"], s["center"], s["b"])

    rows = [
        (4, "Total Gangs Deployed (>=1)", "='2. Fleet Efficiency'!$B$4", "0"),
        (5, "Marginal Gangs (<3 erections)", "='2. Fleet Efficiency'!$B$7", "0"),
        (6, "% of Fleet That Was Marginal", "='2. Fleet Efficiency'!$B$8", "0.00%"),
        (7, "MT from Marginal Gangs", "='2. Fleet Efficiency'!$B$20", "0.00"),
        (8, "% of Total MT from Marginal Gangs", "='2. Fleet Efficiency'!$B$21", "0.00%"),
        (9, "Deployment Inefficiency Ratio", "='2. Fleet Efficiency'!$B$28", "0.00"),
        (10, "MT/gang Retained (>=3)", "='2. Fleet Efficiency'!$B$16", "0.00"),
        (11, "MT/gang Marginal (dropped)", "='2. Fleet Efficiency'!$B$22", "0.00"),
        (12, "Retained Outperformance", "='2. Fleet Efficiency'!$B$23", "0.00"),
        (13, "Total Idle MT Recoverable (productive >=3 gangs, deployment-month basis)", "='3. Idle Recovery'!$K$12", "0.00"),
        (14, "Coverage Ratio (Idle Recovery vs MT Gap)", "='3. Idle Recovery'!$B$24", "0.00%"),
        (15, "Targeted Recovery: MT recovered if 4-5 bucket closes to 6-7 benchmark", "='3. Idle Recovery'!$B$18", "0.00"),
    ]
    for r, label, f, fmt in rows:
        ws.cell(r, 1, label)
        paint(ws, r, 1, 1, s["light"], s["fbb"], s["left"], s["b"])
        ws.cell(r, 2, f)
        paint(ws, r, 2, 2, s["green"], s["fgreen"], s["center"], s["b"])
        ws.cell(r, 2).number_format = fmt

    ws["A18"] = "Slide Bullets (Manual)"
    ws.merge_cells("A18:B18")
    paint(ws, 18, 1, 2, s["mid"], s["fw"], s["left"], s["b"])
    for i in range(1, 6):
        r = 18 + i
        ws.cell(r, 1, f"[BULLET {i}]")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        paint(ws, r, 1, 2, s["white"], s["fb"], s["left"], s["b"])

def build_sheet5(wb: Workbook, s: dict[str, object]):
    ws = wb.create_sheet("5. Idle Narrative")
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 26
    ws.column_dimensions["G"].width = 18

    ws["A1"] = "Idle Narrative (Deployment-Month Basis)"
    ws.merge_cells("A1:G1")
    paint(ws, 1, 1, 7, s["dark"], s["fw"], s["left"], s["b"])
    ws["A2"] = "All calculations are live and linked to Section A controllable share assumption."
    ws.merge_cells("A2:G2")
    paint(ws, 2, 1, 7, s["light"], s["fb"], s["left"], s["b"])

    ws["A4"] = "Section A - Assumptions"
    ws.merge_cells("A4:G4")
    paint(ws, 4, 1, 7, s["mid"], s["fw"], s["left"], s["b"])

    def assum(r: int, label: str, v, fmt: str | None = None):
        ws.cell(r, 1, label)
        paint(ws, r, 1, 1, s["light"], s["fbb"], s["left"], s["b"])
        ws.cell(r, 2, v)
        paint(ws, r, 2, 2, s["yellow"], s["fblue"], s["left"], s["b"])
        if fmt:
            ws.cell(r, 2).number_format = fmt

    assum(5, "Controllable idle share (Others)", 0.35, "0%")
    assum(6, "Uncontrollable idle share (ROW)", "=1-B5", "0%")
    assum(7, "Productivity band - Low (MT/day)", "0-4 (buckets 0-1, 1-2, 2-3, 3-4)")
    assum(8, "Productivity band - Mid (MT/day)", "4-6 (buckets 4-5, 5-6)")
    assum(9, "Productivity band - High (MT/day)", "6-10 (buckets 6-7, 7-8, 8-9)")

    if "controllable_share" in wb.defined_names:
        del wb.defined_names["controllable_share"]
    wb.defined_names.add(DefinedName(name="controllable_share", attr_text="'5. Idle Narrative'!$B$5"))

    ws["A11"] = "Section B - Band-level Idle Summary (Threshold >=3)"
    ws.merge_cells("A11:G11")
    paint(ws, 11, 1, 7, s["mid"], s["fw"], s["left"], s["b"])

    for i, h in enumerate([
        "Band", "Total Gangs", "Weighted Avg Idle Days / Deployment Month", "Weighted Avg Deployment Months",
        "Weighted Avg Baseline MT/day", "Total Idle Days in Scope", "Idle Index vs Low",
    ], start=1):
        ws.cell(12, i, h)
    paint(ws, 12, 1, 7, s["light"], s["fbb"], s["center"], s["b"])

    ws["A13"] = "Low"
    ws["B13"] = "=SUM('1. Raw Inputs'!B8:E8)"
    ws["C13"] = "=IFERROR(SUMPRODUCT('1. Raw Inputs'!B8:E8,'1. Raw Inputs'!B26:E26)/B13,0)"
    ws["D13"] = "=IFERROR(SUMPRODUCT('1. Raw Inputs'!B8:E8,'1. Raw Inputs'!B20:E20)/B13,0)"
    ws["E13"] = "=IFERROR(SUMPRODUCT('1. Raw Inputs'!B8:E8,'1. Raw Inputs'!B50:E50)/B13,0)"
    ws["F13"] = "=SUMPRODUCT('1. Raw Inputs'!B8:E8,'1. Raw Inputs'!B26:E26,'1. Raw Inputs'!B20:E20)"
    ws["G13"] = "=1"

    ws["A14"] = "Mid"
    ws["B14"] = "=SUM('1. Raw Inputs'!F8:G8)"
    ws["C14"] = "=IFERROR(SUMPRODUCT('1. Raw Inputs'!F8:G8,'1. Raw Inputs'!F26:G26)/B14,0)"
    ws["D14"] = "=IFERROR(SUMPRODUCT('1. Raw Inputs'!F8:G8,'1. Raw Inputs'!F20:G20)/B14,0)"
    ws["E14"] = "=IFERROR(SUMPRODUCT('1. Raw Inputs'!F8:G8,'1. Raw Inputs'!F50:G50)/B14,0)"
    ws["F14"] = "=SUMPRODUCT('1. Raw Inputs'!F8:G8,'1. Raw Inputs'!F26:G26,'1. Raw Inputs'!F20:G20)"
    ws["G14"] = "=IFERROR(C14/$C$13,0)"

    ws["A15"] = "High"
    ws["B15"] = "=SUM('1. Raw Inputs'!H8:J8)"
    ws["C15"] = "=IFERROR(SUMPRODUCT('1. Raw Inputs'!H8:J8,'1. Raw Inputs'!H26:J26)/B15,0)"
    ws["D15"] = "=IFERROR(SUMPRODUCT('1. Raw Inputs'!H8:J8,'1. Raw Inputs'!H20:J20)/B15,0)"
    ws["E15"] = "=IFERROR(SUMPRODUCT('1. Raw Inputs'!H8:J8,'1. Raw Inputs'!H50:J50)/B15,0)"
    ws["F15"] = "=SUMPRODUCT('1. Raw Inputs'!H8:J8,'1. Raw Inputs'!H26:J26,'1. Raw Inputs'!H20:J20)"
    ws["G15"] = "=IFERROR(C15/$C$13,0)"

    for r in (13, 14, 15):
        paint(ws, r, 1, 1, s["light"], s["fbb"], s["center"], s["b"])
        paint(ws, r, 2, 6, s["green"], s["fgreen"], s["center"], s["b"])
        paint(ws, r, 7, 7, s["white"], s["fb"], s["center"], s["b"])
        ws.cell(r, 2).number_format = "0"
        for c in (3, 4, 5, 6):
            ws.cell(r, c).number_format = "0.00"
        ws.cell(r, 7).number_format = "0.00x"

    ws["A17"] = "Idle Index = Band idle days / Low band idle days"
    paint(ws, 17, 1, 1, s["light"], s["fbb"], s["left"], s["b"])
    ws["B16"], ws["C16"], ws["D16"] = "Low", "Mid", "High"
    paint(ws, 16, 2, 4, s["light"], s["fbb"], s["center"], s["b"])
    ws["B17"], ws["C17"], ws["D17"] = "=1", "=IFERROR(C14/$C$13,0)", "=IFERROR(C15/$C$13,0)"
    paint(ws, 17, 2, 4, s["white"], s["fb"], s["center"], s["b"])
    ws["B17"].number_format = ws["C17"].number_format = ws["D17"].number_format = "0.00x"

    ws["A19"] = "Section C - Controllable Idle Recovery for High-productivity Gangs"
    ws.merge_cells("A19:G19")
    paint(ws, 19, 1, 7, s["mid"], s["fw"], s["left"], s["b"])
    for r, label, f, fmt in [
        (20, "Total idle days in scope (High band)", "=F15", "0.00"),
        (21, "Controllable share", "=B5", "0%"),
        (22, "Controllable idle days", "=B20*B21", "0.00"),
        (23, "Baseline MT/day (High band weighted avg)", "=E15", "0.00"),
        (24, "Recoverable MT (total, full scope)", "=B22*B23", "0.00"),
        (25, "Weighted avg deployment months (High band)", "=D15", "0.00"),
        (26, "Recoverable MT per month", "=IFERROR(B24/B25,0)", "0.00"),
        (27, "Total High-band gangs", "=B15", "0"),
        (28, "Recoverable MT per gang per month", "=IFERROR(B26/B27,0)", "0.00"),
    ]:
        ws.cell(r, 1, label)
        paint(ws, r, 1, 1, s["light"], s["fbb"], s["left"], s["b"])
        ws.cell(r, 2, f)
        paint(ws, r, 2, 2, s["white"], s["fb"], s["center"], s["b"])
        ws.cell(r, 2).number_format = fmt
    for r in (20, 23, 25, 27):
        paint(ws, r, 2, 2, s["green"], s["fgreen"], s["center"], s["b"])

    ws["A30"] = "Section D - Sensitivity (Controllable Share 20% to 50%)"
    ws.merge_cells("A30:D30")
    paint(ws, 30, 1, 4, s["mid"], s["fw"], s["left"], s["b"])
    ws["A31"], ws["B31"], ws["C31"], ws["D31"] = "Controllable Share", "Controllable Idle Days", "Recoverable MT", "Recoverable MT / Month"
    paint(ws, 31, 1, 4, s["light"], s["fbb"], s["center"], s["b"])
    for i, p in enumerate([0.20, 0.25, 0.30, 0.35, 0.40, 0.45, 0.50], start=32):
        ws.cell(i, 1, p)
        paint(ws, i, 1, 1, s["yellow"], s["fblue"], s["center"], s["b"])
        ws.cell(i, 1).number_format = "0%"
        ws.cell(i, 2, f"=A{i}*$B$20")
        ws.cell(i, 3, f"=B{i}*$B$23")
        ws.cell(i, 4, f"=IFERROR(C{i}/$B$25,0)")
        paint(ws, i, 2, 4, s["white"], s["fb"], s["center"], s["b"])
        ws.cell(i, 2).number_format = ws.cell(i, 3).number_format = ws.cell(i, 4).number_format = "0.00"

    ws["A40"] = "Section E - Narrative Output (Live Formula Callouts)"
    ws.merge_cells("A40:D40")
    paint(ws, 40, 1, 4, s["mid"], s["fw"], s["left"], s["b"])
    cals = {
        41: '="High-productivity gangs (6-10 MT/day) idle "&TEXT($C$15,"0.0")&" days/deployment-month - "&TEXT(IFERROR($C$15/$C$13,0),"0.0")&"x more than low-productivity gangs ("&TEXT($C$13,"0.0")&" days/deployment-month)."',
        43: '="At "&TEXT($B$5,"0%")&" controllable share, high-band controllable idle equals "&TEXT($B$22,"0.0")&" days and "&TEXT($B$24,"0.0")&" MT recoverable over full scope."',
        45: '="Recovery opportunity is "&TEXT($B$26,"0.0")&" MT/month from high-productivity gangs under current assumptions."',
        47: '="Per-gang opportunity is "&TEXT($B$28,"0.00")&" MT/gang/month for high-band crews."',
    }
    for r, f in cals.items():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, f)
        paint(ws, r, 1, 4, s["light"], s["fcall"], s["left"], s["b"])
        ws.row_dimensions[r].height = 34
    for r in (42, 44, 46):
        paint(ws, r, 1, 4, border=s["b"])


def validate(path: Path):
    wb = load_workbook(path, data_only=False)
    need = ["1. Raw Inputs", "2. Fleet Efficiency", "3. Idle Recovery", "4. Leadership Summary", "5. Idle Narrative"]
    for sh in need:
        if sh not in wb.sheetnames:
            raise RuntimeError(f"Missing sheet {sh}")
        if wb[sh].freeze_panes != "B5":
            raise RuntimeError(f"{sh} freeze pane is {wb[sh].freeze_panes}, expected B5")
    if "controllable_share" not in wb.defined_names:
        raise RuntimeError("Named range controllable_share missing")
    if "'5. Idle Narrative'!$B$5" not in str(wb.defined_names["controllable_share"].attr_text):
        raise RuntimeError("Named range controllable_share points to wrong cell")
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.data_type == "e" and cell.value in ("#VALUE!", "#N/A"):
                    raise RuntimeError(f"Error cell {ws.title}!{cell.coordinate}={cell.value}")
                if isinstance(cell.value, str) and cell.value in ("#VALUE!", "#N/A"):
                    raise RuntimeError(f"Literal error text {ws.title}!{cell.coordinate}={cell.value}")


def main() -> int:
    a = parse()
    src = Path(a.source).expanduser().resolve()
    out = Path(a.output).expanduser()
    if not out.is_absolute():
        out = (PROD / out).resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    if not src.exists():
        raise SystemExit(f"Source workbook not found: {src}")

    vals = extract(src)
    s = styles()
    wb = Workbook()
    wb.remove(wb.active)
    refs = build_sheet1(wb, vals, s)
    build_sheet2(wb, refs, s)
    build_sheet3(wb, refs, s)
    build_sheet4(wb, s)
    build_sheet5(wb, s)
    wb.save(out)
    validate(out)
    print(f"[export] Wrote '{out}'")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
