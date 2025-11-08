# microplan_compile.py
# ------------------------------------------------------------
# Build a tidy, analysis-ready "Micro Plan" responsibilities table
# and write it into the compiled workbook.
#
# What it does:
#   - Discovers files whose name contains "Micro Plan" (case-insensitive)
#   - Auto-detects the header row within the first 50 rows
#   - Canonicalizes columns to a stable schema (easy to tweak below)
#   - Aggregates responsibilities (sum of Revenue, Tower Weight) for:
#       Gang / Section Incharge / Supervisor
#   - Writes a single tidy sheet "MicroPlanResponsibilities"
#   - Optionally writes per-project cleaned micro plan sheets
#   - Writes an index sheet for traceability ("MicroPlanIndex")
#
# How to use (CLI):
#   python microplan_compile.py --input-dir "<folder with project files>" \
#                               --output-xlsx "ErectionCompiled_Output.xlsx"
#
# How to use (import):
#   from microplan_compile import compile_microplans_to_workbook
#   compile_microplans_to_workbook(input_dir="...", output_path="...")
#
# Dependencies:
#   pip install pandas openpyxl
# ------------------------------------------------------------

from __future__ import annotations

import argparse
import os
import re
from glob import glob
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

# =========================
# === Config (Editable) ===
# =========================

# --- Micro Plan: schema config (edit here when your header standard changes) ---
MICROPLAN_SCHEMA_V1 = {
    "required": {
        "gang_name":        ["gang name"],
        "section_incharge": ["section incharge"],
        "supervisor":       ["supervisor"],
        "revenue_planned":   ["revenue planned", "revenue"],
        "revenue_realised":  ["revenue realised", "revenue realized", "revenue realization"],
        "location_no":       ["location no.", "loc no", "loc no.", "location no"],

    },
    "optional": {
        "tower_weight": ["tower weight"],
        "tower_type": ["tower type"],
        "manpower":            ["manpower"],
        "power_tools_issued":  ["power tools issued (yes/no)", "power tools issued"],
        "material_feeding":    ["material feeding", "matl feeding"],
        "starting_date":       ["starting date", "start date"],
        "completion_date":     ["completion date", "end date"],
        "tack_welding":        ["tack-welding", "tack welding"],
        "final_checking":      ["final checking", "final check"],
    }
}
CURRENT_MICROPLAN_SCHEMA = MICROPLAN_SCHEMA_V1


# Output sheet names
MICROPLAN_AGG_SHEET_NAME = "MicroPlanResponsibilities"
MICROPLAN_INDEX_SHEET    = "MicroPlanIndex"

# Search pattern for "Micro Plan" files
MICROPLAN_GLOB_PATTERN = "**/*micro*plan*.xls*"

# Header search window
HEADER_SCAN_MAX_ROWS = 50
MICROPLAN_LEFT_WIDTH = 30

# ==========================
# === Helper / Utilities ===
# ==========================

def _norm_text(x: str) -> str:
    if pd.isna(x):
        return ""
    x = str(x).strip().lower()
    x = re.sub(r"[\s\-_./()]+", " ", x)
    x = re.sub(r"[^a-z0-9 %]+", "", x)
    return " ".join(x.split())

_DETECTION_TOKENS_RAW = {
    # core entity/metrics
    "gang name", "section incharge", "supervisor",
    "tower weight", "revenue planned", "revenue realised", "revenue realized",
    "location no", "loc no",
    # helpful supporting headers (boosts scoring)
    # "sl", "loc no.", "tower type", "manpower",
    # "power tools issued (yes/no)", "matl feeding",
    # "starting date", "completion date", "tack-welding", "final checking",
}

_DETECTION_TOKENS = {_norm_text(s) for s in _DETECTION_TOKENS_RAW}

def _pick_erection_sheet(path: str) -> str:
    """
    Return the sheet whose name contains 'erection' (case-insensitive).
    If none found, fall back to the first sheet.
    """
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    
    for ws in wb.worksheets:
        if "erection" in ws.title.strip().lower():
            return ws.title
    # fallback: first visible
    return wb.worksheets[0].title


def _detect_header_row(path: str, sheet: str, max_rows: int = HEADER_SCAN_MAX_ROWS) -> int:
    """
    Detect the header row by scanning the first `max_rows` rows without any
    index-based column limits. If you want to ignore far-right noise, we slice
    *after* reading instead of using `usecols`, which can go out of bounds on
    narrow sheets.
    """
    prev = pd.read_excel(
        path,
        sheet_name=sheet,
        header=None,
        nrows=max_rows,
        engine="openpyxl",
        dtype=object,
    )

    best_row, best_score = None, -1

    # Evaluate each candidate row; only consider the first MICROPLAN_LEFT_WIDTH cells
    width = MICROPLAN_LEFT_WIDTH if isinstance(MICROPLAN_LEFT_WIDTH, int) and MICROPLAN_LEFT_WIDTH > 0 else prev.shape[1]

    for r in range(len(prev)):
        # slice row to left window, coerce to strings safely
        row_vals = prev.iloc[r].tolist()[:width]
        toks = {_norm_text(v) for v in row_vals if v is not None and str(v).strip() != ""}
        score = sum(1 for t in _DETECTION_TOKENS if t in toks)
        if score > best_score:
            best_row, best_score = r, score

    # Require a minimal confidence to avoid false positives
    if best_row is None or best_score < 3:
        raise ValueError(f"Header row not found on sheet '{sheet}' (best_score={best_score}).")
    return best_row







def _alias_to_canonical(colname: str, schema: dict) -> Optional[str]:
    """Return canonical key (gang_name, supervisor, etc.) if colname matches an alias."""
    n = _norm_text(colname)
    for canon, aliases in {**schema["required"], **schema.get("optional", {})}.items():
        if any(n == _norm_text(a) for a in aliases):
            return canon
    return None




def infer_project_name_from_filename(path: str) -> str:
    """
    Extract project code from filenames like:
    'Micro Plan - TA-416 Sep'25.xlsx' → 'TA416'
    """
    stem = os.path.splitext(os.path.basename(path))[0]

    # Strip the "micro plan" words
    s = re.sub(r"(?i)\bmicro\s*plan\b", " ", stem)

    # Remove month/year tokens (Sep'25, Oct 2025, etc.)
    s = re.sub(r"(?i)(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)['\-\s]*\d{2,4}", " ", s)
    s = re.sub(r"\b20\d{2}\b", " ", s)
    s = re.sub(r"\b\d{2}\b", " ", s)

    s = " ".join(s.split())

    # Extract first letter+3 digit code, remove dashes/spaces
    m = re.search(r"(?i)\b([A-Z]{1,3})[\s\-_]*([0-9]{3})\b", s)
    if m:
        return f"{m.group(1).upper()}{m.group(2)}"

    return stem


def infer_plan_month_from_filename(path: str) -> Optional[pd.Timestamp]:
    """
    Try to extract a (month, year) from filename tokens like:
      - "Sep'25", "Sept-25", "Oct 2025", "Jul2024", etc.

    Returns a pandas Timestamp normalized to the first day of that month,
    or None if no (month, year) pair is confidently found.
    """
    name = os.path.splitext(os.path.basename(path))[0]
    s = name.lower()

    # Normalize separators to spaces
    s = re.sub(r"[\-_]+", " ", s)

    # Month name patterns (short and long names)
    month_map = {
        # short
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
        # long
        "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
        "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
    }

    # Look for month tokens optionally followed or preceded by a year
    # Examples matched:
    #   sep'25, sept-25, oct 2025, jul2024, 2025 oct, 25-sep, october 2025, 2025 november
    month_alt = r"jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec|january|february|march|april|august|september|october|november|december|june|july"
    patterns = [
        fr"\b({month_alt})[ '\-]*([0-9]{{2,4}})\b",
        fr"\b([0-9]{{2,4}})[ '\-]*?({month_alt})\b",
    ]

    def _match_month_year(text: str) -> Optional[pd.Timestamp]:
        for pat in patterns:
            m = re.search(pat, text)
            if not m or m.lastindex != 2:
                continue
            a, b = m.group(1), m.group(2)
            if a in month_map:
                mon = month_map[a]
                year_txt = b
            else:
                mon = month_map.get(b)
                year_txt = a
            if not mon:
                continue
            try:
                year = int(year_txt)
                if year < 100:
                    year += 2000
                if 1990 <= year <= 2100:
                    return pd.Timestamp(year=year, month=mon, day=1)
            except Exception:
                continue
        return None

    # 1) Prefer parent directories (e.g., "October 2025" folder)
    try:
        parent = os.path.dirname(path)
        # walk up a few levels to be safe
        for _ in range(4):
            if not parent:
                break
            seg = os.path.basename(parent).lower()
            ts = _match_month_year(seg)
            if ts is not None:
                return ts
            new_parent = os.path.dirname(parent)
            if new_parent == parent:
                break
            parent = new_parent
    except Exception:
        pass

    # 2) Fallback to filename, if needed
    ts = _match_month_year(s)
    if ts is not None:
        return ts

    return None


def normalize_key(s: str) -> str:
    """Stable snake_key → 'TA416' → 'ta416'"""
    return re.sub(r"[^a-z0-9]+", "", _norm_text(s))


def _safe_read_excel_by_header(path: str, sheet_name: str, header_row: int) -> pd.DataFrame:
    """
    Read a sheet using the detected header row. Avoid index-based `usecols`;
    we'll trim by names (and optionally left width) after reading.
    """
    df = pd.read_excel(
        path,
        sheet_name=sheet_name,
        header=header_row,
        dtype=object,
        engine="openpyxl",
    )
    # standardize NA and strip column label whitespace
    df = df.where(df.notna(), None)
    df.columns = [str(c).strip() if c is not None else "" for c in df.columns]
    return df

def read_microplan_file(path: str,
                        schema: dict,
                        header_scan_rows: int = HEADER_SCAN_MAX_ROWS) -> pd.DataFrame:
    sheet = _pick_erection_sheet(path)
    header_row = _detect_header_row(path,sheet, max_rows=header_scan_rows)

    df = _safe_read_excel_by_header(path, sheet, header_row)

     # --- NEW TRUNCATION LOGIC ---
    # Find first completely empty row after header
    mask_empty = df.isna().all(axis=1)
    if mask_empty.any():
        first_empty_idx = mask_empty.idxmax()  # index of first True
        # truncate everything after this row
        df = df.loc[:first_empty_idx-1]

    rename_map = {}
    for col in df.columns:
        canon = _alias_to_canonical(col, schema)
        if canon:
            rename_map[col] = canon
    df = df.rename(columns=rename_map)

    needed = ["gang_name", "section_incharge", "supervisor", "tower_weight", "revenue_planned", "revenue_realised", "location_no"]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise ValueError(f"Required columns missing: {missing}. Columns read: {list(df.columns)}")


    for col in ["gang_name", "section_incharge", "supervisor"]:
        if col in df.columns:
            df[col] = df[col].ffill().astype(str).str.strip()
    
    # location as stable string id (not numeric), trimmed
    df["location_no"] = df["location_no"].astype(object).map(
        lambda x: str(x).strip() if x is not None and str(x).strip() != "" else None
    )

    for c in ("tower_weight", "revenue_planned", "revenue_realised"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    mask = (df["revenue_planned"].fillna(0) == 0) & (df["revenue_realised"].fillna(0) == 0) & (df["tower_weight"].fillna(0) == 0)
    df = df.loc[~mask].reset_index(drop=True)

    # If there is no usable completion_date in the sheet, fall back to month from filename.
    plan_month = infer_plan_month_from_filename(path)
    if plan_month is not None:
        # Ensure column exists, then fill missing values only
        if "completion_date" not in df.columns:
            df["completion_date"] = plan_month
        else:
            # Treat empty strings and NA as missing
            missing_mask = df["completion_date"].isna() | (df["completion_date"].astype(str).str.strip() == "")
            if missing_mask.any():
                df.loc[missing_mask, "completion_date"] = plan_month

    return df.reset_index(drop=True)







def build_responsibilities_atomic(df_clean: pd.DataFrame,
                                  project_name: str,
                                  project_key: str,
                                  plan_month: Optional[pd.Timestamp]) -> pd.DataFrame:
    """
    Return atomic rows: one row per (project, entity_type, entity_name, location_no)
    with the *per-row* revenue/tower_weight values from the microplan (no pre-aggregation).
    """
    # Ensure required columns exist (reader enforces schema earlier)
    for c in ("revenue_planned", "revenue_realised", "tower_weight"):
        if c not in df_clean.columns:
            df_clean[c] = 0.0

    # normalize location as stable string (already done earlier, but keep safe)
    loc = df_clean["location_no"].astype(object).map(
        lambda x: str(x).strip() if x is not None and str(x).strip() != "" else None
    )

    frames = []
    pairs = [
        ("gang_name",        "Gang"),
        ("section_incharge", "Section Incharge"),
        ("supervisor",       "Supervisor"),
    ]

    for col, etype in pairs:
        if col not in df_clean.columns:
            continue
        # atomic slice for this entity type
        cols_to_take = [c for c in [
            col, "location_no",
            "revenue_planned","revenue_realised","tower_weight",
            "tower_type","manpower","power_tools_issued","material_feeding",
            "starting_date","completion_date","tack_welding","final_checking"
        ] if c in df_clean.columns]
        part = df_clean[cols_to_take].copy()
        part.insert(0, "project_key",  project_key)
        part.insert(1, "project_name", project_name)
        part.insert(2, "plan_month",   plan_month)
        part.insert(3, "entity_type",  etype)
        part = part.rename(columns={col: "entity_name"})
        # Drop rows with missing entity_name or missing/blank location
        part["entity_name"] = part["entity_name"].astype(str).str.strip()
        part["location_no"] = loc  # normalized above
        part = part[(part["entity_name"] != "") & part["location_no"].notna()]
        frames.append(part)

    if not frames:
        return pd.DataFrame(columns=[
            "project_key","project_name","plan_month","entity_type","entity_name",
            "location_no","revenue_planned","revenue_realised","tower_weight"
        ])

    out = pd.concat(frames, ignore_index=True)
    # Column order for consistency
    cols = [
        "project_key","project_name","plan_month","entity_type","entity_name","location_no",
        "revenue_planned","revenue_realised","tower_weight",
        "tower_type","manpower","power_tools_issued","material_feeding",
        "starting_date","completion_date","tack_welding","final_checking"
    ]
    return out[cols]




def _safe_write_df(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, index: bool = False) -> None:
    """
    Overwrite a sheet if it exists. Uses writer.book provided by pandas'
    openpyxl engine; no manual assignment to writer.book.
    """
    wb = writer.book  # read-only property provided by pandas
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
        # Also clean pandas' internal cache to avoid stale references
        if hasattr(writer, "sheets") and sheet_name in writer.sheets:
            del writer.sheets[sheet_name]
    df.to_excel(writer, sheet_name=sheet_name, index=index)


def _open_writer_overwriting_sheets(output_path: str) -> pd.ExcelWriter:
    """
    Open an ExcelWriter. If the file exists, open in append mode and allow
    us to overwrite specific sheets later. Do NOT assign writer.book.
    """
    if os.path.exists(output_path):
        return pd.ExcelWriter(
            output_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="overlay",  # we'll delete/replace targeted sheets ourselves
        )
    else:
        return pd.ExcelWriter(output_path, engine="openpyxl", mode="w")


# ======================
# === Main Compile  ===
# ======================

def compile_microplans_to_workbook(
    input_dir: str,
    output_path: str,
    *,
    schema: dict = CURRENT_MICROPLAN_SCHEMA,
    glob_pattern: str = MICROPLAN_GLOB_PATTERN
) -> None:
    """
    - Finds files with 'micro plan' in the name (case-insensitive).
    - For each file: detect header, clean, aggregate responsibilities.
    - Writes one combined sheet: MICROPLAN_AGG_SHEET_NAME
    - Optionally writes per-project cleaned sheet(s).
    - Also writes an index (file → project) for traceability.
    """
    paths = sorted(glob(os.path.join(input_dir, glob_pattern), recursive=True))

    atomic_all = []
    index_rows = []
    
    issues: list[dict] = []

    for p in paths:
        proj_name = infer_project_name_from_filename(p)
        proj_key  = normalize_key(proj_name)
        plan_month = infer_plan_month_from_filename(p)
        
        # if write_raw_per_project else None
        try:
            df_clean = read_microplan_file(p, schema)
            # --- Data Issues: record rows where in-file completion_date's month differs from folder month
            try:
                if "completion_date" in df_clean.columns and plan_month is not None:
                    comp = pd.to_datetime(df_clean["completion_date"], errors="coerce")
                    comp_month = comp.dt.to_period("M").dt.to_timestamp()
                    mismatch = comp_month.notna() & (comp_month != plan_month)
                    if mismatch.any():
                        sub = df_clean.loc[mismatch, [c for c in ["location_no", "gang_name", "section_incharge", "supervisor", "completion_date"] if c in df_clean.columns]].copy()
                        sub["file_path"] = p
                        sub["project_name"] = proj_name
                        sub["project_key"] = proj_key
                        sub["expected_month"] = plan_month
                        issues.extend(sub.to_dict("records"))
            except Exception:
                # Keep pipeline resilient; issues sheet is best-effort
                pass

            # Regardless of in-sheet values, enforce completion_date from folder month for downstream filtering
            if plan_month is not None:
                df_clean["completion_date"] = plan_month
                        # ensure numeric responsibility columns exist
            for c in ("revenue_planned", "tower_weight"):
                if c not in df_clean.columns:
                    df_clean[c] = 0.0

            # build and collect per-project responsibilities
            resp_long = build_responsibilities_atomic(df_clean, proj_name, proj_key, plan_month)
            if not resp_long.empty:
                atomic_all.append(resp_long)         

            # ...
            index_rows.append({
                "file_path": p,
                "project_name": proj_name,
                "project_key": proj_key,
                "rows_cleaned": len(df_clean),
                "status": "ok",
                "error": "",
                "plan_month": plan_month
            })
        except Exception as e:
            index_rows.append({
                "file_path": p,
                "project_name": proj_name,   # keep for debugging
                "project_key": proj_key,     # keep for debugging
                "rows_cleaned": 0,
                "status": "error",
                "error": str(e),
                "plan_month": plan_month
            })


    # Write combined outputs in one go
    # --- Write combined outputs (atomic only) ---
    with _open_writer_overwriting_sheets(output_path) as writer:
        # Atomic responsibilities sheet
        if atomic_all:
            responsibilities = pd.concat(atomic_all, ignore_index=True)
        else:
            responsibilities = pd.DataFrame(columns=[
                "project_key","project_name","plan_month","entity_type","entity_name",
                "location_no","revenue_planned","revenue_realised","tower_weight",
                "tower_type","manpower","power_tools_issued","material_feeding",
                "starting_date","completion_date","tack_welding","final_checking"
            ])

        _safe_write_df(writer, responsibilities, MICROPLAN_AGG_SHEET_NAME, index=False)

        # Index sheet for traceability
        idx_df = pd.DataFrame(index_rows)
        _safe_write_df(writer, idx_df, MICROPLAN_INDEX_SHEET, index=False)

        # Data Issues sheet for Micro Plan
        try:
            issues_df = pd.DataFrame(issues)
            if not issues_df.empty:
                # Order columns if present
                cols = [
                    "file_path", "project_name", "project_key", "location_no",
                    "gang_name", "section_incharge", "supervisor",
                    "completion_date", "expected_month"
                ]
                ordered = [c for c in cols if c in issues_df.columns] + [c for c in issues_df.columns if c not in cols]
                issues_df = issues_df[ordered]
            _safe_write_df(writer, issues_df, "MicroPlanDataIssues", index=False)
        except Exception:
            # Non-fatal if we cannot write issues
            pass




# ======================
# === CLI Entrypoint ===
# ======================

def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Compile 'Micro Plan' responsibilities into the compiled workbook.")
    p.add_argument("--input-dir", required=True, help="Root folder containing project Excel files (where 'Micro Plan' files reside).")
    p.add_argument("--output-xlsx", required=True, help="Path to the compiled workbook (e.g., ErectionCompiled_Output.xlsx).")
    # p.add_argument("--write-raw", action="store_true",
                #    help="Also write per-project cleaned Micro Plan sheets (MicroPlan_<project_key>).")
    return p


def main() -> None:
    args = _build_arg_parser().parse_args()
    compile_microplans_to_workbook(
        input_dir=args.input_dir,
        output_path=args.output_xlsx,
    )


if __name__ == "__main__":
    main()
